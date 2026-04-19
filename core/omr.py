"""
core/omr.py
-----------
Lógica OMR (Optical Mark Recognition) independiente de la interfaz tkinter.
Puede ser invocada desde la API Flask o desde el módulo de escritorio.
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import cv2
    import numpy as np

    _HAS_CV2 = True
except ImportError:
    _HAS_CV2 = False

from . import get_connection, get_db_path


# ---------------------------------------------------------------------------
# Parsing de QR
# ---------------------------------------------------------------------------


def parse_qr_payload(qr_text: str) -> dict[str, str]:
    """Interpreta el texto de un código QR y devuelve los campos del examen.

    Soporta dos formatos:
    - JSON: ``{"examen_id":"...","version":"A",...}``
    - SEA pipe-separated: ``SEA|doc|nombre|grado|curso|area|eval|ID:XXXXXX``
    """
    txt = str(qr_text or "").strip()
    if not txt:
        return {}

    if txt.startswith("{") and txt.endswith("}"):
        try:
            payload = json.loads(txt)
            exam_id_json = (
                str(
                    payload.get("examen_id")
                    or payload.get("id_examen")
                    or payload.get("exam_id")
                    or ""
                )
                .strip()
                .upper()
            )
            return {
                "raw": txt,
                "id_examen": exam_id_json,
                "documento": str(payload.get("documento", "") or "").strip(),
                "nombre": str(payload.get("nombre", "") or "").strip(),
                "grado": str(payload.get("grado", "") or "").strip(),
                "curso": str(payload.get("curso", "") or "").strip(),
                "area": str(payload.get("area", "") or "").strip(),
                "evaluacion": str(payload.get("evaluacion", "") or "").strip(),
                "version": str(payload.get("version", "") or "").strip().upper(),
            }
        except Exception:
            pass

    m = re.search(r"ID:([A-Za-z0-9]+)", txt, re.IGNORECASE)
    exam_id = m.group(1).upper() if m else ""

    data: dict[str, str] = {
        "raw": txt,
        "id_examen": exam_id,
        "documento": "",
        "nombre": "",
        "grado": "",
        "curso": "",
        "area": "",
        "evaluacion": "",
        "version": "",
    }

    parts = txt.split("|")
    if len(parts) >= 8 and str(parts[0]).strip().upper() == "SEA":
        data["documento"] = str(parts[1]).strip()
        data["nombre"] = str(parts[2]).strip()
        data["grado"] = str(parts[3]).strip()
        data["curso"] = str(parts[4]).strip()
        data["area"] = str(parts[5]).strip()
        data["evaluacion"] = str(parts[6]).strip()
        if not data["id_examen"]:
            m2 = re.search(r"ID:([A-Za-z0-9]+)", str(parts[7]), re.IGNORECASE)
            if m2:
                data["id_examen"] = m2.group(1).upper()

    return data


# ---------------------------------------------------------------------------
# Detección OMR
# ---------------------------------------------------------------------------


def detectar_respuestas_omr(
    frame: Any, total_preguntas: int
) -> tuple[dict[int, str], Any]:
    """Detecta respuestas marcadas en una hoja OMR a partir de un frame BGR.

    Returns
    -------
    (respuestas, debug_frame)
        *respuestas*: ``{n_pregunta: "A"|"B"|"C"|"D"}``
        *debug_frame*: imagen con burbujas resaltadas (o el frame original si hay error)
    """
    if not _HAS_CV2:
        return {}, frame
    if frame is None:
        return {}, frame

    # ------------------------------------------------------------------
    # Helpers internos
    # ------------------------------------------------------------------

    def _ordenar_puntos(pts):
        rect = np.zeros((4, 2), dtype="float32")
        s = pts.sum(axis=1)
        rect[0] = pts[np.argmin(s)]
        rect[2] = pts[np.argmax(s)]
        diff = np.diff(pts, axis=1)
        rect[1] = pts[np.argmin(diff)]
        rect[3] = pts[np.argmax(diff)]
        return rect

    def _warp_perspectiva(img, pts):
        rect = _ordenar_puntos(pts)
        (tl, tr, br, bl) = rect
        width_a = np.linalg.norm(br - bl)
        width_b = np.linalg.norm(tr - tl)
        max_w = int(max(width_a, width_b))
        height_a = np.linalg.norm(tr - br)
        height_b = np.linalg.norm(tl - bl)
        max_h = int(max(height_a, height_b))
        if max_w < 50 or max_h < 50:
            return img
        dst = np.array(
            [[0, 0], [max_w - 1, 0], [max_w - 1, max_h - 1], [0, max_h - 1]],
            dtype="float32",
        )
        m = cv2.getPerspectiveTransform(rect, dst)
        return cv2.warpPerspective(img, m, (max_w, max_h))

    def _sheet_column_ranges(total):
        total = max(1, int(total or 1))
        if total <= 30:
            split = max(1, (total + 1) // 2)
            ranges = [(1, split)]
            if split < total:
                ranges.append((split + 1, total))
        else:
            ranges = []
            start = 1
            while start <= total and len(ranges) < 3:
                end = min(total, start + 29)
                ranges.append((start, end))
                start = end + 1
            if start <= total:
                ranges.append((start, total))

        while len(ranges) < 2:
            ranges.append((0, 0))
        if len(ranges) > 4:
            merged = ranges[:3]
            merged.append((ranges[3][0], ranges[-1][1]))
            ranges = merged
        return ranges

    def _find_document_quad(bin_img):
        cnts, _ = cv2.findContours(bin_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if not cnts:
            return None
        frame_area = float(bin_img.shape[0] * bin_img.shape[1])
        for c in sorted(cnts, key=cv2.contourArea, reverse=True)[:20]:
            area = float(cv2.contourArea(c))
            if area < frame_area * 0.18:
                continue
            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.02 * peri, True)
            if len(approx) == 4:
                return approx.reshape(4, 2).astype("float32")
        return None

    def _detect_corner_square(th_img, corner_name):
        h, w = th_img.shape[:2]
        roi_w = max(90, int(w * 0.2))
        roi_h = max(90, int(h * 0.2))
        if corner_name == "TL":
            x0, y0 = 0, 0
        elif corner_name == "TR":
            x0, y0 = w - roi_w, 0
        elif corner_name == "BL":
            x0, y0 = 0, h - roi_h
        else:
            x0, y0 = w - roi_w, h - roi_h

        roi = th_img[y0 : y0 + roi_h, x0 : x0 + roi_w]
        if roi.size == 0:
            return None

        cnts, _ = cv2.findContours(roi, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
        roi_area = float(roi_w * roi_h)
        best = None
        best_score = -1.0

        for c in cnts:
            area = float(cv2.contourArea(c))
            if area < roi_area * 0.004 or area > roi_area * 0.35:
                continue
            x, y, bw, bh = cv2.boundingRect(c)
            if bw <= 0 or bh <= 0:
                continue
            ar = float(bw) / float(bh)
            if ar < 0.65 or ar > 1.35:
                continue
            rect_area = float(bw * bh)
            solidity = area / rect_area if rect_area > 0 else 0.0
            if solidity < 0.70:
                continue
            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.04 * peri, True)
            if len(approx) < 4 or len(approx) > 8:
                continue
            m = cv2.moments(c)
            if m["m00"] == 0:
                continue
            cx = int(m["m10"] / m["m00"])
            cy = int(m["m01"] / m["m00"])
            score = area * solidity
            if score > best_score:
                best_score = score
                best = np.array([x0 + cx, y0 + cy], dtype="float32")

        return best

    # ------------------------------------------------------------------
    # Pipeline principal
    # ------------------------------------------------------------------
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    th_doc = cv2.adaptiveThreshold(
        blur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 41, 9
    )

    page = _find_document_quad(th_doc)
    warped = _warp_perspectiva(frame, page) if page is not None else frame.copy()

    g2 = cv2.cvtColor(warped, cv2.COLOR_BGR2GRAY)
    g2 = cv2.GaussianBlur(g2, (5, 5), 0)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    g2 = clahe.apply(g2)

    th_marks = cv2.adaptiveThreshold(
        g2, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 41, 7
    )

    tl = _detect_corner_square(th_marks, "TL")
    tr = _detect_corner_square(th_marks, "TR")
    br_pt = _detect_corner_square(th_marks, "BR")
    bl = _detect_corner_square(th_marks, "BL")
    if any(p is None for p in (tl, tr, br_pt, bl)):
        return {}, warped

    src_l = np.array([tl, tr, br_pt, bl], dtype="float32")
    tpl_w, tpl_h = 1600, 2200
    dst_l = np.array(
        [[0, 0], [tpl_w - 1, 0], [tpl_w - 1, tpl_h - 1], [0, tpl_h - 1]],
        dtype="float32",
    )
    mat_l = cv2.getPerspectiveTransform(src_l, dst_l)
    aligned_gray = cv2.warpPerspective(g2, mat_l, (tpl_w, tpl_h))
    aligned = cv2.cvtColor(aligned_gray, cv2.COLOR_GRAY2BGR)
    th = cv2.adaptiveThreshold(
        aligned_gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 35, 7
    )

    ranges = _sheet_column_ranges(total_preguntas)
    num_columnas = len(ranges)
    max_rows = max(
        1,
        max(
            ((r[1] - r[0] + 1) for r in ranges if r[0] > 0 and r[1] >= r[0]),
            default=1,
        ),
    )

    zone_left = 0.07 * float(tpl_w)
    zone_right = 0.93 * float(tpl_w)
    zone_top = 0.23 * float(tpl_h)
    zone_bottom = 0.95 * float(tpl_h)

    zone_w = max(1.0, zone_right - zone_left)
    zone_h = max(1.0, zone_bottom - zone_top)
    inner_left = zone_left + (0.03 * zone_w)
    inner_right = zone_right - (0.03 * zone_w)
    usable_w = max(1.0, inner_right - inner_left)
    col_gap = (0.02 * zone_w) if num_columnas > 1 else 0.0
    col_width = (usable_w - ((num_columnas - 1) * col_gap)) / max(1, num_columnas)
    block_header_h = max(12.0, 0.075 * zone_h)
    rows_top = zone_top + block_header_h
    rows_bottom = zone_bottom - (0.03 * zone_h)
    row_gap = max(9.5, (rows_bottom - rows_top) / max(1, max_rows))
    number_w = max(26.0, 0.20 * col_width)
    bubble_track = max(20.0, col_width - number_w - (0.06 * col_width))
    bubble_gap = bubble_track / 3.0
    bubble_r = min(19.0, max(11.0, row_gap * 0.22))

    respuestas: dict[int, str] = {}
    letras = ["A", "B", "C", "D"]

    for col_idx, (r_start, r_end) in enumerate(ranges):
        if r_start <= 0 or r_end < r_start:
            continue
        x_col = inner_left + (col_idx * (col_width + col_gap))
        for row_idx in range(r_end - r_start + 1):
            pregunta = r_start + row_idx
            if pregunta > int(total_preguntas):
                break

            cy = int(rows_top + ((row_idx + 0.5) * row_gap))
            start_x = x_col + number_w + (0.03 * col_width)
            scores: list[tuple[float, str, int, int, int]] = []

            for opt_idx, letra in enumerate(letras):
                cx = int(start_x + (opt_idx * bubble_gap))
                rr = int(max(8, bubble_r * 0.72))
                x1 = max(0, cx - rr)
                y1 = max(0, cy - rr)
                x2 = min(tpl_w, cx + rr + 1)
                y2 = min(tpl_h, cy + rr + 1)
                roi = th[y1:y2, x1:x2]
                if roi.size == 0:
                    scores.append((0.0, letra, cx, cy, rr))
                    continue

                mask = np.zeros(roi.shape, dtype="uint8")
                cv2.circle(mask, (cx - x1, cy - y1), rr, 255, -1)
                area_px = max(1, cv2.countNonZero(mask))
                ink_px = cv2.countNonZero(cv2.bitwise_and(roi, roi, mask=mask))
                fill_ratio = float(ink_px) / float(area_px)

                circ_score = 0.0
                roi_cnts, _ = cv2.findContours(
                    roi, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
                )
                if roi_cnts:
                    best_c = max(roi_cnts, key=cv2.contourArea)
                    area_c = float(cv2.contourArea(best_c))
                    peri_c = float(cv2.arcLength(best_c, True))
                    if peri_c > 0.0:
                        circ = (4.0 * np.pi * area_c) / (peri_c * peri_c)
                        circ_score = max(0.0, min(1.0, circ))

                robust_ratio = fill_ratio + (0.08 * circ_score)
                scores.append((robust_ratio, letra, cx, cy, rr))

            scores.sort(key=lambda t: t[0], reverse=True)
            top = scores[0][0]
            second = scores[1][0] if len(scores) > 1 else 0.0

            if top >= 0.30 and (top - second) >= 0.07:
                respuestas[pregunta] = scores[0][1]
                cv2.circle(
                    aligned, (scores[0][2], scores[0][3]), scores[0][4], (0, 255, 0), 2
                )
            else:
                for _s, _l, cx, cy_b, rr in scores:
                    cv2.circle(aligned, (cx, cy_b), rr, (0, 180, 255), 1)

    return respuestas, aligned


# ---------------------------------------------------------------------------
# Persistencia de calificaciones
# ---------------------------------------------------------------------------


def guardar_calificacion_camara(
    qr_data: dict[str, str],
    exam_code: str,
    respuestas_est: dict[int, str],
    correctas: int,
    total: int,
    nota: float,
    db_path: Path | None = None,
    base_dir: Path | None = None,
) -> None:
    """Guarda la calificación obtenida por cámara en SQLite y en Excel."""
    try:
        with get_connection(db_path) as conn:
            conn.execute(
                """INSERT INTO calificaciones_camara
                       (id_examen, documento, estudiante_nombre, grado, curso, area,
                        evaluacion, total_preguntas, correctas, nota, respuestas_json)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    exam_code,
                    qr_data.get("documento", ""),
                    qr_data.get("nombre", ""),
                    qr_data.get("grado", ""),
                    qr_data.get("curso", ""),
                    qr_data.get("area", ""),
                    qr_data.get("evaluacion", ""),
                    int(total),
                    int(correctas),
                    float(nota),
                    json.dumps(respuestas_est, ensure_ascii=False),
                ),
            )
            conn.commit()
    except Exception:
        pass

    exportar_calificacion_camara_excel(
        qr_data=qr_data,
        exam_code=exam_code,
        respuestas_est=respuestas_est,
        correctas=correctas,
        total=total,
        nota=nota,
        base_dir=base_dir,
    )


def exportar_calificacion_camara_excel(
    qr_data: dict[str, str],
    exam_code: str,
    respuestas_est: dict[int, str],
    correctas: int,
    total: int,
    nota: float,
    base_dir: Path | None = None,
) -> None:
    """Exporta/actualiza el consolidado Excel de lecturas por cámara."""
    try:
        from openpyxl import Workbook, load_workbook

        if base_dir is None:
            base_dir = get_db_path().parent

        reportes_dir = Path(base_dir) / "reportes"
        reportes_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = reportes_dir / "calificaciones_camara.xlsx"

        if xlsx_path.exists():
            wb = load_workbook(str(xlsx_path))
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "calificaciones"
            ws.append(
                [
                    "fecha",
                    "id_examen",
                    "documento",
                    "estudiante",
                    "grado",
                    "curso",
                    "area",
                    "evaluacion",
                    "correctas",
                    "total",
                    "nota",
                    "respuestas_json",
                ]
            )

        ws.append(
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                exam_code,
                qr_data.get("documento", ""),
                qr_data.get("nombre", ""),
                qr_data.get("grado", ""),
                qr_data.get("curso", ""),
                qr_data.get("area", ""),
                qr_data.get("evaluacion", ""),
                int(correctas),
                int(total),
                float(nota),
                json.dumps(respuestas_est, ensure_ascii=False, sort_keys=True),
            ]
        )
        wb.save(str(xlsx_path))
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Consulta de clave de examen
# ---------------------------------------------------------------------------


def obtener_clave_examen(exam_code: str, db_path: Path | None = None) -> dict[int, str]:
    """Devuelve ``{n_pregunta: letra_correcta}`` para un examen dado."""
    try:
        with get_connection(db_path) as conn:
            cur = conn.cursor()
            cur.execute(
                """SELECT numero_pregunta, respuesta_correcta
                     FROM detalle_examen
                    WHERE id_examen = ?
                    ORDER BY numero_pregunta""",
                (exam_code,),
            )
            rows = cur.fetchall()
    except Exception:
        return {}

    clave: dict[int, str] = {}
    for num, resp in rows:
        try:
            n = int(num)
        except Exception:
            continue
        letra = str(resp or "").strip().upper()[:1]
        if letra in ("A", "B", "C", "D"):
            clave[n] = letra
    return clave


def procesar_imagen_omr(
    image_bytes: bytes,
    qr_text: str,
    db_path: Path | None = None,
    base_dir: Path | None = None,
) -> dict:
    """Punto de entrada principal para la API móvil.

    Recibe los bytes de la imagen y el texto del QR, detecta respuestas,
    calcula nota y persiste el resultado.

    Returns
    -------
    dict con claves:
        ok, id_examen, estudiante, nota, correctas, total, respuestas, error
    """
    if not _HAS_CV2:
        return {"ok": False, "error": "opencv_no_disponible"}

    qr_data = parse_qr_payload(qr_text)
    exam_code = qr_data.get("id_examen", "").strip().upper()
    if not exam_code:
        return {"ok": False, "error": "qr_sin_id_examen"}

    clave = obtener_clave_examen(exam_code, db_path)
    if not clave:
        return {"ok": False, "error": f"examen_{exam_code}_sin_clave"}

    try:
        buf = np.frombuffer(image_bytes, dtype=np.uint8)
        frame = cv2.imdecode(buf, cv2.IMREAD_COLOR)
    except Exception as exc:
        return {"ok": False, "error": f"imagen_invalida: {exc}"}

    if frame is None:
        return {"ok": False, "error": "imagen_no_decodificable"}

    respuestas_est, _ = detectar_respuestas_omr(frame, len(clave))

    if not respuestas_est:
        return {"ok": False, "error": "no_se_detectaron_burbujas"}

    correctas = sum(1 for n, resp in clave.items() if respuestas_est.get(n, "") == resp)
    total = len(clave)
    nota = round((correctas / float(total)) * 100.0, 2)

    guardar_calificacion_camara(
        qr_data=qr_data,
        exam_code=exam_code,
        respuestas_est=respuestas_est,
        correctas=correctas,
        total=total,
        nota=nota,
        db_path=db_path,
        base_dir=base_dir,
    )

    return {
        "ok": True,
        "id_examen": exam_code,
        "estudiante": qr_data.get("nombre", ""),
        "documento": qr_data.get("documento", ""),
        "grado": qr_data.get("grado", ""),
        "curso": qr_data.get("curso", ""),
        "area": qr_data.get("area", ""),
        "evaluacion": qr_data.get("evaluacion", ""),
        "nota": nota,
        "correctas": correctas,
        "total": total,
        "respuestas": {str(k): v for k, v in respuestas_est.items()},
    }
