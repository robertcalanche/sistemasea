from core.construir_nombre import construir_nombre

# -*- coding: utf-8 -*-
import os
import sqlite3
import tkinter as tk
from tkinter import ttk, simpledialog, messagebox, filedialog
import shutil
import random
import io
from pathlib import Path
import sys
import re
import json
import unicodedata
from datetime import datetime
from types import SimpleNamespace
from ui_footer import crear_footer
from core import usuarios as core_usuarios


def _runtime_dir():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def _leer_config_sistema(config_path):
    config = {"modo": "local", "ruta_servidor": ""}
    try:
        for raw_line in config_path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            config[key.strip().lower()] = value.strip().strip('"').strip("'")
    except Exception:
        pass
    return config


def _resolver_rutas_sistema():
    runtime_dir = _runtime_dir()
    config_path = runtime_dir / "config_sistema"
    config = _leer_config_sistema(config_path)

    modo = str(config.get("modo", "local")).strip().lower()
    ruta_servidor = str(config.get("ruta_servidor", "")).strip().strip('"').strip("'")

    if modo == "red" and ruta_servidor:
        ruta = Path(ruta_servidor)
        if not ruta.is_absolute():
            # Permite rutas relativas al directorio del ejecutable/script.
            ruta = (runtime_dir / ruta).resolve()
        if ruta.suffix.lower() == ".db":
            base_dir = ruta.parent
            db_path = ruta
        else:
            base_dir = ruta
            db_path = base_dir / "sistema.db"
    else:
        base_dir = runtime_dir
        db_path = base_dir / "sistema.db"

    return base_dir, str(db_path), config_path


# rutas de base de datos unificada
BASE_DIR, DB_PATH, CONFIG_SISTEMA_FILE = _resolver_rutas_sistema()
DB_FILE = Path(DB_PATH)

try:
    from PIL import Image, ImageTk

    _HAS_PIL = True
except Exception:
    _HAS_PIL = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    _HAS_REPORTLAB = True
except Exception:
    _HAS_REPORTLAB = False

try:
    import pandas as pd

    _HAS_PANDAS = True
except Exception:
    from openpyxl import load_workbook, Workbook

    _HAS_PANDAS = False

try:
    cv2 = __import__("cv2")
    np = __import__("numpy")

    _HAS_CV2 = True
except Exception:
    cv2 = None
    np = None
    _HAS_CV2 = False


ANSWER_SHEET_MODE_NONE = "none"
ANSWER_SHEET_MODE_APPEND = "append"
ANSWER_SHEET_MODE_INLINE = "inline"

EXAM_FORMAT_STANDARD = "standard"
EXAM_FORMAT_MATH_ICFES = "math_icfes"
EXAM_FORMAT_LANGUAGE_ICFES = "language_icfes"


def _normalizar_texto_sin_tildes(texto):
    txt = str(texto or "").strip().lower()
    if not txt:
        return ""
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def _es_area_lenguaje(area):
    area_norm = _normalizar_texto_sin_tildes(area)
    if not area_norm:
        return False

    equivalentes = {
        "lenguaje",
        "lengua castellana",
        "castellano",
        "espanol",
        "comunicacion",
        "humanidades lengua castellana",
        "lectura critica",
    }
    return area_norm in equivalentes


def _extraer_expresiones_latex(texto):
    """Extrae expresiones entre $...$ y valida balance básico de delimitadores."""
    txt = str(texto or "")
    if "$" not in txt:
        return []
    if txt.count("$") % 2 != 0:
        raise ValueError("Delimitadores '$' desbalanceados.")
    return [frag.strip() for frag in re.findall(r"\$([^$]+)\$", txt) if frag.strip()]


_SUPERSCRIPT_MAP = {
    "0": "⁰",
    "1": "¹",
    "2": "²",
    "3": "³",
    "4": "⁴",
    "5": "⁵",
    "6": "⁶",
    "7": "⁷",
    "8": "⁸",
    "9": "⁹",
    "+": "⁺",
    "-": "⁻",
    "=": "⁼",
    "(": "⁽",
    ")": "⁾",
}


def _to_superscript(token):
    out = []
    for ch in str(token or ""):
        sup = _SUPERSCRIPT_MAP.get(ch)
        if sup is None:
            return f"^{token}"
        out.append(sup)
    return "".join(out)


def _formula_a_texto_legible(expr):
    txt = str(expr or "").strip()
    if not txt:
        return ""

    if txt.startswith("$") and txt.endswith("$") and len(txt) >= 2:
        txt = txt[1:-1].strip()

    txt = txt.replace("**", "^")
    txt = re.sub(r"\\frac\s*\{([^{}]+)\}\{([^{}]+)\}", r"\1/\2", txt)
    txt = re.sub(r"\\sqrt\s*\{([^{}]+)\}", r"√\1", txt)
    txt = re.sub(r"sqrt\s*\(([^()]+)\)", r"√\1", txt, flags=re.IGNORECASE)
    txt = re.sub(r"sqrt\s*([A-Za-z0-9]+)", r"√\1", txt, flags=re.IGNORECASE)
    txt = txt.replace("\\cdot", "·").replace("\\times", "x")

    txt = re.sub(r"\^\{([^{}]+)\}", lambda m: _to_superscript(m.group(1)), txt)
    txt = re.sub(r"\^([A-Za-z0-9+\-=()]+)", lambda m: _to_superscript(m.group(1)), txt)

    txt = txt.replace("{", "").replace("}", "")
    txt = txt.replace("\\", "")
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def _detectar_formula_matematica(texto):
    txt = str(texto or "")
    if not txt.strip():
        return False
    return bool(re.search(r"\$[^$]+\$", txt))


def _natural_sort_key(valor):
    texto = str(valor or "").strip()
    if not texto:
        return []
    return [
        int(fragmento) if fragmento.isdigit() else fragmento.lower()
        for fragmento in re.split(r"(\d+)", texto)
    ]


def renderizar_formula(expr, dpi=220):
    """Renderiza una fórmula a PNG en memoria y retorna (buffer, dpi) o None."""
    expr_txt = str(expr or "").strip()
    if not expr_txt:
        return None

    if expr_txt.startswith("$") and expr_txt.endswith("$") and len(expr_txt) >= 2:
        expr_txt = expr_txt[1:-1].strip()

    if not expr_txt:
        return None

    try:
        import importlib

        mathtext = importlib.import_module("matplotlib.mathtext")
        buf = io.BytesIO()
        mathtext.math_to_image(f"${expr_txt}$", buf, dpi=dpi, format="png")
        buf.seek(0)
        return buf, float(dpi)
    except Exception:
        return None


def _validar_expresion_latex(expr):
    """Valida de forma liviana una expresión LaTeX sin volver obligatoria ninguna dependencia."""
    expr_txt = str(expr or "").strip()
    if not expr_txt:
        raise ValueError("Expresión LaTeX vacía.")

    if expr_txt.count("{") != expr_txt.count("}"):
        raise ValueError("Llaves '{' y '}' desbalanceadas en la expresión.")

    renderizar_formula(expr_txt, dpi=180)
    return True


def _render_latex_a_buffer(expr, dpi=220):
    """Renderiza LaTeX a PNG y retorna (buffer, dpi) o None si no se puede."""
    return renderizar_formula(expr, dpi=dpi)


def normalizar_modo_hoja_respuestas(modo=None, generar=False):
    if isinstance(modo, bool):
        return ANSWER_SHEET_MODE_APPEND if modo else ANSWER_SHEET_MODE_NONE

    modo_txt = str(modo or "").strip().lower()
    if modo_txt in {
        ANSWER_SHEET_MODE_NONE,
        ANSWER_SHEET_MODE_APPEND,
        ANSWER_SHEET_MODE_INLINE,
    }:
        return modo_txt

    return ANSWER_SHEET_MODE_APPEND if generar else ANSWER_SHEET_MODE_NONE


def normalizar_formato_examen(formato=None):
    formato_txt = str(formato or "").strip().lower()
    if formato_txt in {
        EXAM_FORMAT_STANDARD,
        EXAM_FORMAT_MATH_ICFES,
        EXAM_FORMAT_LANGUAGE_ICFES,
    }:
        return formato_txt

    if formato_txt in {
        "estándar",
        "estandar",
        "estándar (actual)",
        "estandar (actual)",
    }:
        return EXAM_FORMAT_STANDARD

    if formato_txt in {
        "matemáticas (tipo icfes)",
        "matematicas (tipo icfes)",
        "matemáticas tipo icfes",
        "matematicas tipo icfes",
        "icfes",
    }:
        return EXAM_FORMAT_MATH_ICFES

    if formato_txt in {
        "lenguaje",
        "lenguaje (lectura y abierta)",
        "lenguaje (lectura + abierta)",
        "lenguaje icfes",
        "lectura y abierta",
    }:
        return EXAM_FORMAT_LANGUAGE_ICFES

    return EXAM_FORMAT_STANDARD


# ============= FUNCIONES DE SOPORTE (Importadas de Admin.py) =============


def normalizar_grado(grado):
    """Normaliza el grado: strip, lower, elimina 'grado', '°', espacios extras."""
    if grado is None:
        return ""
    result = str(grado).strip().lower()
    result = result.replace("grado", "").strip()
    result = result.replace("°", "").strip()
    result = " ".join(result.split())
    return result


def cargar_grados_desde_preguntas(preguntas_path=None):
    """Devuelve los grados disponibles en SQLite (tabla banco_preguntas)."""
    try:
        return core_preguntas.cargar_grados_banco()
    except Exception:
        return []


def cargar_estudiantes_por_grado(
    grado, curso=None, estudiantes_path=None, db_path=None
):
    """Recupera alumnos por grado (y opcionalmente curso) desde SQLite."""
    # normalizar/convertir grado
    try:
        grado_val = int(float(grado))
    except Exception:
        grado_val = normalizar_grado(grado)
    # normalizar curso si se provee
    curso_val = None
    if curso is not None:
        try:
            curso_val = str(curso).strip()
        except Exception:
            curso_val = None

    try:
        return core_matricula.listar_estudiantes_por_grado(
            grado=grado_val,
            curso=curso_val,
            solo_activos=False,
        )
    except Exception:
        return []


def cargar_cursos_por_grado(grado, estudiantes_path=None, db_path=None):
    """Devuelve la lista de cursos únicos para un grado dado.

    Utiliza la misma fuente que `cargar_estudiantes_por_grado` (DB o Excel).
    Si el campo "curso" no existe se retorna lista vacía.
    """
    cursos = []
    estudiantes = cargar_estudiantes_por_grado(
        grado, estudiantes_path=estudiantes_path, db_path=db_path
    )
    for e in estudiantes:
        c = e.get("curso")
        if c is not None:
            cursos.append(str(c).strip())
    cursos = sorted([c for c in set(cursos) if c])
    return cursos


def cargar_areas_por_grado(grado, preguntas_path=None):
    """Carga areas para el grado desde SQLite."""
    try:
        return core_preguntas.cargar_areas_por_grado(grado)
    except Exception:
        return []


def cargar_evaluaciones_por_grado_y_area(grado, area, preguntas_path=None):
    """Carga evaluaciones para grado y area desde SQLite."""
    try:
        if not grado or not area:
            return []
        return core_preguntas.cargar_evaluaciones_por_grado_y_area(grado, area)
    except Exception:
        return []


def cargar_preguntas_filtradas(grado, area, evaluacion=None, preguntas_path=None):
    """Devuelve DataFrame filtrado por grado+area (+evaluacion) desde SQLite."""
    try:
        if not _HAS_PANDAS:
            return pd.DataFrame()
        # Buscar id_evaluacion único para los filtros dados
        id_eval = core_preguntas.obtener_id_evaluacion(grado, area, evaluacion)
        if id_eval:
            return core_preguntas.cargar_preguntas_filtradas(
                area=area, grado=grado, evaluacion=evaluacion, id_evaluacion=id_eval
            )
        # Fallback: lógica anterior si no se encuentra id_evaluacion
        return core_preguntas.cargar_preguntas_filtradas(
            area=area, grado=grado, evaluacion=evaluacion
        )
    except Exception:
        return pd.DataFrame()


# ============= FUNCIONES DE CONVERSIÓN DE CÓDIGOS =============


def convertir_genero(codigo):
    """Convierte códigos de género del Excel al formato completo.

    Args:
        codigo: "F" → "Femenino", "M" → "Masculino"

    Returns:
        Género convertido o el valor original si no coincide.
    """
    if not codigo:
        return codigo
    codigo = str(codigo).strip().upper()
    mapeo = {
        "F": "Femenino",
        "M": "Masculino",
    }
    return mapeo.get(codigo, codigo)


def convertir_jornada(codigo):
    """Convierte códigos de jornada del Excel al formato completo.

    Args:
        codigo: "M" → "Mañana", "T" → "Tarde", "N" → "Nocturna"

    Returns:
        Jornada convertida o el valor original si no coincide.
    """
    if not codigo:
        return codigo
    codigo = str(codigo).strip().upper()
    mapeo = {
        "M": "Mañana",
        "T": "Tarde",
        "N": "Nocturna",
    }
    return mapeo.get(codigo, codigo)


def convertir_estado_academico(codigo):
    """Convierte códigos de estado académico del Excel al formato completo.

    Args:
        codigo: "MA" → "Matriculado", "GR" → "Graduado", "TR" → "Trasladado", "RE" → "Retirado"

    Returns:
        Estado académico convertido o el valor original si no coincide.
    """
    if not codigo:
        return codigo
    codigo = str(codigo).strip().upper()
    mapeo = {
        "MA": "Matriculado",
        "GR": "Graduado",
        "TR": "Trasladado",
        "RE": "Retirado",
    }
    return mapeo.get(codigo, codigo)


def convertir_tipo_documento(codigo):
    """Convierte códigos de tipo de documento del Excel al nombre completo.

    Args:
        codigo: Código abreviado (RC, TI, CC, etc.)

    Returns:
        Nombre completo del tipo de documento o el código original si no coincide.
    """
    if not codigo:
        return codigo
    codigo = str(codigo).strip().upper()
    mapeo = {
        "RC": "Registro civil de nacimiento",
        "TI": "Tarjeta de identidad",
        "NUIP": "Número único de identificación personal",
        "CC": "Cédula de ciudadanía",
        "CE": "Cédula de extranjería",
        "PPT": "Permiso de protección temporal",
        "PEP": "Permiso especial permanencia",
        "RUMV": "Registro único de migrantes venezolanos",
        "PA": "Pasaporte",
        "PN": "Partida de nacimiento",
        "NIP": "Número de identificación personal",
        "NES": "Número establecido por la secretaría",
        "TMF": "Tarjeta movilidad fronteriza",
        "CCA": "Certificado de cabildo",
        "VISA": "Visa",
    }
    return mapeo.get(codigo, codigo)


from interfaz_banco_preguntas import InterfazBancoPreguntasAvanzada
from core import examenes as core_examenes
from core import examenes_generacion as core_examenes_generacion
from core import examenes_pdf as core_examenes_pdf
from core import matricula as core_matricula
from core import preguntas as core_preguntas
from core import docentes as core_docentes


class ModuloSuperAdmin(InterfazBancoPreguntasAvanzada):
    def _ejecutar_generar_cuadernillo_multi_area(
        self,
        grado,
        evaluaciones_por_area,
        areas,
        preguntas_por_area,
        pdf_path,
        estudiante=None,
    ):
        try:
            from multi_area_exam_pdf import write_multi_area_exam_pdf
        except ImportError:
            messagebox.showerror(
                "Dependencia faltante",
                "No se encontró el módulo para cuadernillo multi-área.\nVerifique la instalación.",
                parent=self.win,
            )
            return
        # Recopilar preguntas por área
        areas_preguntas_dict = {}
        for area in areas:
            n = preguntas_por_area.get(area, 0)
            evaluacion_area = (evaluaciones_por_area or {}).get(area, "")
            # 1. Intentar consulta exacta
            df = self.banco.obtener_preguntas_filtradas(
                grado=grado, area=area, evaluacion=evaluacion_area
            )
            if len(df) < n:
                # 2. Intentar solo por área y grado
                df_area = self.banco.obtener_preguntas_filtradas(grado=grado, area=area)
                # 3. Intentar solo por área (todas las evaluaciones y grados)
                df_area_sola = self.banco.obtener_preguntas_filtradas(area=area)
                if len(df_area) > 0:
                    msg = (
                        f"No hay suficientes preguntas para el área '{area}' y evaluación '{evaluacion_area}'.\n"
                        f"Disponibles para el área en este grado (todas las evaluaciones): {len(df_area)}.\n"
                        "Verifique si la evaluación coincide exactamente con el banco de preguntas."
                    )
                elif len(df_area_sola) > 0:
                    msg = (
                        f"No hay preguntas para el área '{area}' en el grado y evaluación seleccionados.\n"
                        f"Pero existen {len(df_area_sola)} preguntas para el área en otros grados o evaluaciones.\n"
                        "Verifique la asociación de área, grado y evaluación en el banco de preguntas."
                    )
                else:
                    msg = (
                        f"No hay preguntas registradas para el área '{area}'.\n"
                        "Verifique el banco de preguntas."
                    )
                messagebox.showerror(
                    "Preguntas insuficientes",
                    msg,
                    parent=self.win,
                )
                # Cerrar el diálogo de generación si existe
                try:
                    if (
                        hasattr(self, "dialogo_generar_cuadernillo")
                        and self.dialogo_generar_cuadernillo
                    ):
                        self.dialogo_generar_cuadernillo.destroy()
                except Exception:
                    pass
                return
            areas_preguntas_dict[area] = df.sample(
                n=n, random_state=random.randint(1, 99999)
            )
        # Datos de estudiante: asegurar estructura completa para PDF
        if estudiante is None:
            estudiante = {
                "apellido1": "",
                "apellido2": "",
                "nombre1": "",
                "nombre2": "",
                "documento": "",
                "grado": grado,
                "curso": "",
            }
        elif isinstance(estudiante, str):
            estudiante = {
                "apellido1": "",
                "apellido2": "",
                "nombre1": estudiante,
                "nombre2": "",
                "documento": "",
                "grado": grado,
                "curso": "",
            }
        # Si es dict, asegurar todos los campos requeridos
        if isinstance(estudiante, dict):
            estudiante.setdefault("apellido1", "")
            estudiante.setdefault("apellido2", "")
            estudiante.setdefault("nombre1", "")
            estudiante.setdefault("nombre2", "")
            estudiante.setdefault("documento", "")
            estudiante.setdefault("grado", grado)
            estudiante.setdefault("curso", "")
        # Configuración de numeración y versión
        config_numeracion = "por_area"
        version = "A"
        evaluacion_pdf = " / ".join(
            f"{area}: {evaluacion}"
            for area, evaluacion in (evaluaciones_por_area or {}).items()
            if str(evaluacion).strip()
        )
        try:
            write_multi_area_exam_pdf(
                areas_preguntas_dict=areas_preguntas_dict,
                estudiante=estudiante,
                path=pdf_path,
                evaluacion=evaluacion_pdf,
                version=version,
                config_numeracion=config_numeracion,
                preguntas_por_area=preguntas_por_area,
                instrucciones_generales=None,
            )
            messagebox.showinfo(
                "Cuadernillo generado",
                f"El cuadernillo multi-área se generó correctamente en:\n{pdf_path}",
                parent=self.win,
            )
        except Exception as e:
            import traceback

            messagebox.showerror(
                "Error al generar cuadernillo",
                f"Ocurrió un error:\n{e}\n\n{traceback.format_exc()}",
                parent=self.win,
            )

    """
    Módulo SuperAdmin listo para integrar.

    Uso típico:
        msa = ModuloSuperAdmin(root)       # root: instancia Tk o Toplevel
        if msa.authenticate():
            msa.open_interface()

    Mantiene configuración en SQLite (tabla config_sistema) y centraliza
    Planta de Personal en base de datos para facilitar trabajo en red y futura
    migración a sistema web.
    """

    def __init__(self, parent, db_path=None, base_dir=None, usuario_actual=None):
        self.parent = parent
        self.db_path = db_path or DB_PATH
        self.base_dir = str(base_dir or BASE_DIR)
        self.docentes_path = os.path.join(self.base_dir, "docentes.xlsx")
        self.estudiantes_path = os.path.join(self.base_dir, "estudiantes.xlsx")
        self.preguntas_path = os.path.join(self.base_dir, "preguntas.xlsx")
        # Carpeta para imágenes de preguntas
        self.imagenes_dir = os.path.join(self.base_dir, "imagenes_preguntas")
        os.makedirs(self.imagenes_dir, exist_ok=True)
        self._connect_db()
        self._ensure_config()

        # Inicializar usuario_actual para control de roles
        self.usuario_actual = core_usuarios.enriquecer_usuario_rbac(
            usuario_actual
            if usuario_actual is not None
            else {"rol": core_usuarios.ROL_SUPERADMIN}
        )

        # Data placeholders
        self.docentes_df = None
        self.estudiantes_df = None

        # Inicializar Banco de Preguntas
        from banco_preguntas_profesional import BancoPreguntasProfesional

        self.banco = BancoPreguntasProfesional(db_path=self.db_path)
        self._preguntas_sync_configurada = False

    def _tiene_permiso(self, permiso):
        return core_usuarios.tiene_permiso(self.usuario_actual, permiso)

    def _denegar_permiso(self, permiso, mensaje=None):
        messagebox.showerror(
            "Acceso no autorizado",
            mensaje or f"No tiene permiso para ejecutar esta acción: {permiso}",
            parent=getattr(self, "win", None),
        )

    def _requiere_permiso(self, permiso, mensaje=None):
        if self._tiene_permiso(permiso):
            return True
        self._denegar_permiso(permiso, mensaje=mensaje)
        return False

    def _crear_boton_si_permiso(
        self,
        widget_class,
        contenedor,
        permiso,
        layout="pack",
        layout_kwargs=None,
        **kwargs,
    ):
        if not self._tiene_permiso(permiso):
            return None
        boton = widget_class(contenedor, **kwargs)
        layout_kwargs = layout_kwargs or {}
        if layout == "grid":
            boton.grid(**layout_kwargs)
        elif layout == "place":
            boton.place(**layout_kwargs)
        else:
            boton.pack(**layout_kwargs)
        return boton

    def _agregar_pestana_principal_si_permiso(self, frame, texto):
        permiso = core_usuarios.DESKTOP_SUPERADMIN_TAB_PERMISSIONS.get(texto)
        if permiso and not core_usuarios.superadmin_tab_habilitada(
            self.usuario_actual, permiso
        ):
            return False
        self.nb.add(frame, text=texto)
        return True

    # ---------- SQLite config ----------
    def _connect_db(self):
        self.conn = sqlite3.connect(self.db_path)
        self.cur = self.conn.cursor()

    def _ensure_config(self):
        # Crear tabla para configuración de evaluación (si no existe)
        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS config_evaluacion (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        cognitivo INTEGER,
                        examen INTEGER,
                        autoevaluacion INTEGER,
                        cantidad_notas INTEGER,
                        nota_min REAL,
                        nota_max REAL,
                        decimales INTEGER,
                        metodo TEXT,
                        fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )"""
        )
        self.cur.execute("PRAGMA table_info(config_evaluacion)")
        columnas_config_eval = {
            str(row[1]).strip().lower() for row in self.cur.fetchall()
        }
        for columna in (
            "notas_cognitivo",
            "notas_examen",
            "notas_autoevaluacion",
        ):
            if columna not in columnas_config_eval:
                self.cur.execute(
                    f"ALTER TABLE config_evaluacion ADD COLUMN {columna} INTEGER"
                )
        self.conn.commit()
        self._ensure_planillas_calificaciones_schema()
        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS desempenos_plantilla (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                grado TEXT NOT NULL,
                area TEXT NOT NULL,
                periodo TEXT NOT NULL,
                anio_lectivo TEXT NOT NULL,
                nivel TEXT NOT NULL,
                descriptor TEXT,
                actividad_complementaria TEXT,
                dificultad TEXT,
                fecha_actualizacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )"""
        )
        self.cur.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_desempenos_plantilla_unica "
            "ON desempenos_plantilla(grado, area, periodo, anio_lectivo, nivel)"
        )
        self.conn.commit()

    def _coerce_int(self, value, default=0):
        try:
            return int(float(value))
        except Exception:
            return default

    def _coerce_float(self, value, default=0.0):
        try:
            return float(value)
        except Exception:
            return default

    def _obtener_config_evaluacion_actual(self):
        self._ensure_config()
        try:
            self.cur.execute("SELECT * FROM config_evaluacion ORDER BY id DESC LIMIT 1")
            row = self.cur.fetchone()
            if not row:
                return {}
            columnas = [desc[0] for desc in self.cur.description]
            return dict(zip(columnas, row))
        except Exception:
            return {}

    def _obtener_componentes_evaluacion(self):
        config_eval = self._obtener_config_evaluacion_actual()
        cantidad_legacy = self._coerce_int(
            self._get_config_plantel("cantidad_notas")
            or config_eval.get("cantidad_notas")
            or 6,
            6,
        )

        def _leer_porcentaje(clave_plantel, clave_legacy, default):
            valor = self._get_config_plantel(clave_plantel)
            if str(valor or "").strip() != "":
                return self._coerce_int(valor, default)
            valor_legacy = config_eval.get(clave_legacy)
            if valor_legacy not in (None, ""):
                return self._coerce_int(valor_legacy, default)
            return default

        notas_cognitivo = self._get_config_plantel("notas_cognitivo")
        notas_examen = self._get_config_plantel("notas_examen")
        notas_auto = self._get_config_plantel("notas_autoevaluacion")
        usa_nuevo_modelo = any(
            str(valor or "").strip() != ""
            for valor in (notas_cognitivo, notas_examen, notas_auto)
        ) or any(
            config_eval.get(columna) not in (None, "")
            for columna in (
                "notas_cognitivo",
                "notas_examen",
                "notas_autoevaluacion",
            )
        )

        componentes = {
            "cognitivo": {
                "nombre": self._get_config_plantel("nombre_cognitivo") or "Cognitivo",
                "prefijo": "C",
                "porcentaje": _leer_porcentaje("porcentaje_cognitivo", "cognitivo", 70),
                "cantidad": self._coerce_int(
                    notas_cognitivo
                    or config_eval.get("notas_cognitivo")
                    or (cantidad_legacy if not usa_nuevo_modelo else 0),
                    cantidad_legacy if not usa_nuevo_modelo else 0,
                ),
            },
            "examen": {
                "nombre": self._get_config_plantel("nombre_examen") or "Examen",
                "prefijo": "E",
                "porcentaje": _leer_porcentaje("porcentaje_examen", "examen", 20),
                "cantidad": self._coerce_int(
                    notas_examen or config_eval.get("notas_examen") or 0,
                    0,
                ),
            },
            "autoevaluacion": {
                "nombre": self._get_config_plantel("nombre_autoevaluacion")
                or "Autoevaluación",
                "prefijo": "A",
                "porcentaje": _leer_porcentaje(
                    "porcentaje_autoevaluacion", "autoevaluacion", 10
                ),
                "cantidad": self._coerce_int(
                    notas_auto or config_eval.get("notas_autoevaluacion") or 0,
                    0,
                ),
            },
        }
        return componentes

    def _ensure_planillas_calificaciones_schema(self):
        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS planillas_calificaciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                estudiante_id TEXT,
                area_id INTEGER,
                asignatura_id INTEGER,
                periodo TEXT,
                nota REAL,
                observacion TEXT,
                nota_clave TEXT,
                anio_lectivo TEXT,
                fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )"""
        )
        self.cur.execute("PRAGMA table_info(planillas_calificaciones)")
        columnas_planilla = {str(row[1]).strip().lower() for row in self.cur.fetchall()}
        if "nota_clave" not in columnas_planilla:
            self.cur.execute(
                "ALTER TABLE planillas_calificaciones ADD COLUMN nota_clave TEXT DEFAULT 'N1'"
            )
        self.cur.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_planilla_calif_unica ON planillas_calificaciones(estudiante_id, area_id, periodo, anio_lectivo, nota_clave)"
        )
        self.conn.commit()

    def _tabla_existe(self, nombre_tabla):
        try:
            fila = self.cur.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (nombre_tabla,),
            ).fetchone()
            return bool(fila)
        except Exception:
            return False

    def _obtener_anio_lectivo_base(self):
        anio = ""
        try:
            anio = str(self.obtener_anio_lectivo_activo() or "").strip()
        except Exception:
            anio = ""
        if not anio:
            anio = str(self._get_config_plantel("anio_lectivo") or "").strip()
        return anio or str(datetime.now().year)

    def _listar_periodos_configurados(self, incluir_todos=False):
        cantidad = self._coerce_int(
            self._get_config_plantel("cantidad_periodos") or 4,
            4,
        )
        cantidad = max(1, min(10, cantidad))
        periodos = [f"Periodo {indice}" for indice in range(1, cantidad + 1)]
        if incluir_todos:
            return ["Todos"] + periodos
        return periodos

    def _listar_grados_catalogo(self, incluir_todos=False):
        grados = set()
        consultas = [
            "SELECT DISTINCT TRIM(grado) FROM estudiantes WHERE TRIM(COALESCE(grado, '')) <> ''",
            "SELECT DISTINCT TRIM(grado) FROM plan_estudio WHERE TRIM(COALESCE(grado, '')) <> ''",
            "SELECT DISTINCT TRIM(grado) FROM carga_academica WHERE TRIM(COALESCE(grado, '')) <> ''",
            "SELECT DISTINCT TRIM(grado) FROM desempenos_plantilla WHERE TRIM(COALESCE(grado, '')) <> ''",
        ]
        for consulta in consultas:
            try:
                for fila in self.cur.execute(consulta).fetchall():
                    valor = str((fila or [""])[0] or "").strip()
                    if valor:
                        grados.add(valor)
            except Exception:
                continue
        items = sorted(grados, key=lambda valor: (_natural_sort_key(valor), valor))
        if incluir_todos:
            return ["Todos"] + items
        return items

    def _listar_cursos_catalogo(self, grado=None, incluir_todos=False):
        cursos = set()
        try:
            consulta = "SELECT DISTINCT TRIM(curso) FROM estudiantes WHERE TRIM(COALESCE(curso, '')) <> ''"
            params = []
            grado_txt = str(grado or "").strip()
            if grado_txt and grado_txt != "Todos":
                consulta += " AND TRIM(COALESCE(grado, '')) = ?"
                params.append(grado_txt)
            for fila in self.cur.execute(consulta, params).fetchall():
                valor = str((fila or [""])[0] or "").strip()
                if valor:
                    cursos.add(valor)
        except Exception:
            pass
        items = sorted(cursos, key=lambda valor: (_natural_sort_key(valor), valor))
        if incluir_todos:
            return ["Todos"] + items
        return items

    def _listar_areas_catalogo(self, incluir_todos=False):
        areas = set()
        consultas = [
            "SELECT DISTINCT TRIM(nombre) FROM areas WHERE TRIM(COALESCE(nombre, '')) <> ''",
            "SELECT DISTINCT TRIM(area) FROM plan_estudio WHERE TRIM(COALESCE(area, '')) <> ''",
            "SELECT DISTINCT TRIM(area) FROM carga_academica WHERE TRIM(COALESCE(area, '')) <> ''",
            "SELECT DISTINCT TRIM(area) FROM resultados WHERE TRIM(COALESCE(area, '')) <> ''",
            "SELECT DISTINCT TRIM(area) FROM desempenos_plantilla WHERE TRIM(COALESCE(area, '')) <> ''",
        ]
        if self._tabla_existe("autoevaluacion_planilla_sync"):
            consultas.append(
                "SELECT DISTINCT TRIM(area) FROM autoevaluacion_planilla_sync WHERE TRIM(COALESCE(area, '')) <> ''"
            )
        for consulta in consultas:
            try:
                for fila in self.cur.execute(consulta).fetchall():
                    valor = str((fila or [""])[0] or "").strip()
                    if valor:
                        areas.add(valor)
            except Exception:
                continue
        items = sorted(areas, key=lambda valor: valor.lower())
        if incluir_todos:
            return ["Todos"] + items
        return items

    def _construir_nombre_estudiante(self, fila):
        apellido1 = str(fila.get("apellido1") or "").strip()
        apellido2 = str(fila.get("apellido2") or "").strip()
        nombre1 = str(fila.get("nombre1") or "").strip()
        nombre2 = str(fila.get("nombre2") or "").strip()
        nombre_legacy = str(fila.get("nombre") or "").strip()
        partes = [parte for parte in (apellido1, apellido2, nombre1, nombre2) if parte]
        return " ".join(partes) if partes else nombre_legacy

    def _obtener_escala_valoracion(self):
        try:
            filas = self.cur.execute(
                "SELECT desde, hasta, letra, desempeno, recomendacion FROM escala_valoracion ORDER BY desde DESC"
            ).fetchall()
        except Exception:
            filas = []
        resultado = []
        for fila in filas:
            resultado.append(
                {
                    "desde": self._coerce_float(fila[0], 0.0),
                    "hasta": self._coerce_float(fila[1], 0.0),
                    "letra": str(fila[2] or "").strip(),
                    "desempeno": str(fila[3] or "").strip(),
                    "recomendacion": str(fila[4] or "").strip(),
                }
            )
        if resultado:
            return resultado
        return [
            {
                "desde": 4.6,
                "hasta": 5.0,
                "letra": "S",
                "desempeno": "Superior",
                "recomendacion": "Mantener retos de profundización.",
            },
            {
                "desde": 4.0,
                "hasta": 4.5,
                "letra": "A",
                "desempeno": "Alto",
                "recomendacion": "Consolidar y ampliar los logros alcanzados.",
            },
            {
                "desde": 3.0,
                "hasta": 3.9,
                "letra": "B",
                "desempeno": "Básico",
                "recomendacion": "Reforzar los aprendizajes esenciales.",
            },
            {
                "desde": 1.0,
                "hasta": 2.9,
                "letra": "I",
                "desempeno": "Bajo",
                "recomendacion": "Diseñar un plan de apoyo intensivo.",
            },
        ]

    def _clasificar_desempeno_nota(self, nota, escala=None):
        try:
            valor = float(nota)
        except Exception:
            return "Sin nota"
        for fila in escala or self._obtener_escala_valoracion():
            if fila["desde"] <= valor <= fila["hasta"]:
                return fila["desempeno"] or fila["letra"] or "Sin clasificar"
        return "Sin clasificar"

    def _resumir_texto(self, texto, limite=90):
        contenido = " ".join(str(texto or "").strip().split())
        if len(contenido) <= limite:
            return contenido
        return contenido[: max(0, limite - 1)].rstrip() + "…"

    def _guardar_config_evaluacion(self, config):
        """Guarda la configuración de evaluación en la tabla config_evaluacion, asegurando que solo quede un registro."""
        self._ensure_config()
        # Eliminar todos los registros previos para evitar duplicados
        self.cur.execute("DELETE FROM config_evaluacion")
        self.conn.commit()
        # Insertar el nuevo registro
        self.cur.execute(
            """
            INSERT INTO config_evaluacion (
                cognitivo, examen, autoevaluacion, cantidad_notas, nota_min, nota_max,
                decimales, metodo, notas_cognitivo, notas_examen, notas_autoevaluacion
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                config["cognitivo"],
                config["examen"],
                config["autoevaluacion"],
                config["cantidad_notas"],
                config["nota_min"],
                config["nota_max"],
                config["decimales"],
                config["metodo"],
                config.get("notas_cognitivo", 0),
                config.get("notas_examen", 0),
                config.get("notas_autoevaluacion", 0),
            ),
        )
        self.conn.commit()
        # Agregar claves nuevas a config_sistema si no existen (migración segura)
        # No es necesario alterar la tabla, solo asegurar que los nuevos campos se pueden guardar en config_sistema
        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS config_sistema (
                   clave TEXT PRIMARY KEY,
                   valor TEXT
               )"""
        )
        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS config_examenes (
                   id INTEGER PRIMARY KEY AUTOINCREMENT,
                   grado TEXT,
                   area TEXT,
                   evaluacion TEXT,
                   duracion_segundos INTEGER,
                   cantidad_preguntas INTEGER,
                   max_intentos INTEGER DEFAULT 1,
                   permitir_reintentos INTEGER DEFAULT 1,
                   examen_activo INTEGER DEFAULT 0
               )"""
        )
        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS docentes (
                   id INTEGER PRIMARY KEY AUTOINCREMENT,
                   tipo_documento TEXT,
                   documento TEXT UNIQUE,
                   nombre TEXT,
                   sexo TEXT,
                   fecha_nacimiento TEXT,
                   telefono TEXT,
                   correo TEXT,
                   cargo TEXT,
                   jornada TEXT,
                   sede TEXT,
                   estado TEXT,
                   fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
               )"""
        )
        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS carga_academica (
                   id INTEGER PRIMARY KEY AUTOINCREMENT,
                   docente_documento TEXT,
                   area TEXT,
                   grado TEXT,
                   curso TEXT,
                   horas_asignadas INTEGER DEFAULT 0,
                   horas_extras_usadas INTEGER DEFAULT 0,
                   anio_lectivo TEXT,
                   estado TEXT DEFAULT 'Activo',
                   fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
               )"""
        )
        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS docente_horas_config (
                   docente_documento TEXT PRIMARY KEY,
                   horas_normales_max INTEGER DEFAULT 22,
                   horas_extras_max INTEGER DEFAULT 0,
                   fecha_actualizacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
               )"""
        )
        try:
            self.cur.execute(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS idx_carga_academica_unica
                ON carga_academica (docente_documento, area, grado, curso)
                """
            )
        except sqlite3.IntegrityError:
            # Si existen duplicados antiguos, se mantiene validacion por aplicacion.
            pass
        self.cur.execute(
            """
            CREATE TABLE IF NOT EXISTS plan_estudio (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nivel TEXT,
                grado TEXT,
                curso TEXT,
                area TEXT,
                horas INTEGER,
                estado INTEGER DEFAULT 1
            )
            """
        )
        self.cur.execute(
            """
            CREATE TABLE IF NOT EXISTS areas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT UNIQUE,
                estado INTEGER DEFAULT 1
            )
            """
        )

        areas_iniciales = [
            (1, "Dimensión Cognitiva"),
            (2, "Dimensión Comunicativa"),
            (3, "Dimensión Corporal"),
            (4, "Dimensión Espiritual"),
            (5, "Dimensión Estética"),
            (6, "Dimensión Ética"),
            (7, "Matemáticas"),
            (8, "Ciencias Sociales"),
            (9, "Idioma Extranjero"),
            (10, "Cátedra de la Paz"),
            (11, "Tecnología e Informática"),
            (12, "Educación Artística"),
            (13, "Religión"),
            (14, "Español"),
            (15, "Educación Física"),
            (16, "Ciencias Naturales"),
            (17, "Lectura Crítica"),
            (18, "Educación Ética y Valores"),
            (19, "Historia de Colombia"),
            (21, "Música"),
            (23, "Física"),
            (24, "Química"),
            (25, "Ciencias Políticas y Económicas"),
            (26, "Filosofía"),
        ]
        self.cur.executemany(
            "INSERT OR IGNORE INTO areas (id, nombre) VALUES (?, ?)",
            areas_iniciales,
        )

        # Migración: plan_estudio ahora guarda IdArea además de nombre de área.
        self.cur.execute("PRAGMA table_info(plan_estudio)")
        cols_plan = {str(r[1]).lower() for r in self.cur.fetchall()}
        if "idarea" not in cols_plan:
            self.cur.execute("ALTER TABLE plan_estudio ADD COLUMN IdArea INTEGER")

        # Intentar mapear registros antiguos por nombre de área hacia IdArea.
        self.cur.execute(
            """
            UPDATE plan_estudio
            SET IdArea = (
                SELECT a.id
                FROM areas a
                WHERE LOWER(TRIM(COALESCE(a.nombre, ''))) =
                      LOWER(TRIM(COALESCE(plan_estudio.area, '')))
                LIMIT 1
            )
            WHERE (IdArea IS NULL OR TRIM(CAST(IdArea AS TEXT)) = '')
              AND area IS NOT NULL
              AND TRIM(COALESCE(area, '')) <> ''
            """
        )

        try:
            self.cur.execute(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS idx_plan_estudio_unica_idarea
                ON plan_estudio (grado, curso, IdArea)
                """
            )
        except Exception:
            pass
        self.conn.commit()

        core_examenes_generacion.asegurar_tablas_generacion()
        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS calificaciones_camara (
                   id INTEGER PRIMARY KEY AUTOINCREMENT,
                   id_examen TEXT,
                   documento TEXT,
                   estudiante_nombre TEXT,
                   grado TEXT,
                   curso TEXT,
                   area TEXT,
                   evaluacion TEXT,
                   total_preguntas INTEGER,
                   correctas INTEGER,
                   nota REAL,
                   respuestas_json TEXT,
                   fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP
               )"""
        )
        self.conn.commit()

        # asegurar clave maestra inicial
        self.cur.execute(
            "SELECT valor FROM config_sistema WHERE clave=?", ("master_key",)
        )
        row = self.cur.fetchone()
        if not row:
            self.cur.execute(
                "INSERT INTO config_sistema(clave,valor) VALUES(?,?)",
                ("master_key", "admin123"),
            )
            self.conn.commit()

        # asegurar año lectivo activo inicial

    def _set_config_sistema(self, clave, valor):
        self.cur.execute(
            "REPLACE INTO config_sistema(clave,valor) VALUES(?,?)",
            (clave, str(valor or "").strip()),
        )
        self.conn.commit()

    def _get_config_sistema(self, clave, default=""):
        self.cur.execute("SELECT valor FROM config_sistema WHERE clave=?", (clave,))
        row = self.cur.fetchone()
        valor = row[0] if row else default
        valor = str(valor or "").strip()
        return valor if valor else str(default or "").strip()

    def get_master_key(self):
        return self._get_config_sistema("master_key", default="admin123")

    def set_master_key(self, nuevo):
        self._set_config_sistema("master_key", nuevo)

    def get_superadmin_profile(self):
        nombres = self._get_config_sistema("superadmin_nombres", default="")
        apellidos = self._get_config_sistema("superadmin_apellidos", default="")
        documento = self._get_config_sistema("superadmin_documento", default="admin")
        nombre_completo = " ".join(
            parte for parte in (nombres, apellidos) if str(parte or "").strip()
        ).strip()
        return {
            "nombres": str(nombres or "").strip(),
            "apellidos": str(apellidos or "").strip(),
            "documento": str(documento or "").strip() or "admin",
            "nombre": nombre_completo or "Administrador",
        }

    def set_superadmin_profile(self, nombres, apellidos, documento):
        self._set_config_sistema("superadmin_nombres", nombres)
        self._set_config_sistema("superadmin_apellidos", apellidos)
        self._set_config_sistema("superadmin_documento", documento)

    def _actualizar_resumen_superadmin(self):
        # Detectar rol y mostrar encabezado adecuado
        nombre = ""
        documento = ""
        rol = ""
        if hasattr(self, "usuario_actual") and self.usuario_actual:
            # Puede ser dict o SimpleNamespace
            if isinstance(self.usuario_actual, dict):
                nombre = self.usuario_actual.get("nombre", "")
                documento = self.usuario_actual.get("documento", "")
                rol = str(
                    self.usuario_actual.get("rol_canonico")
                    or self.usuario_actual.get("rol")
                    or ""
                ).lower()
            else:
                nombre = getattr(self.usuario_actual, "nombre", "")
                documento = getattr(self.usuario_actual, "documento", "")
                rol = str(
                    getattr(self.usuario_actual, "rol_canonico", None)
                    or getattr(self.usuario_actual, "rol", "")
                    or ""
                ).lower()
        if not nombre:
            perfil = self.get_superadmin_profile()
            nombre = perfil["nombre"]
            documento = perfil.get("documento", "")
        # Texto según rol
        if rol == core_usuarios.ROL_SUPERADMIN:
            texto = f"Sesión administrativa: {nombre}"
        elif rol in {
            core_usuarios.ROL_RECTOR,
            core_usuarios.ROL_SECRETARIA,
            core_usuarios.ROL_COORDINADOR,
        }:
            texto = f"Sesión directiva: {nombre}"
        elif rol == core_usuarios.ROL_DOCENTE:
            texto = f"Sesión docente: {nombre}"
        else:
            texto = f"Sesión: {nombre}"
        if documento:
            texto += f" | Documento: {documento}"
        if hasattr(self, "lbl_superadmin_session"):
            try:
                self.lbl_superadmin_session.config(text=texto)
            except Exception:
                pass
        # Mantener compatibilidad con dict
        if isinstance(getattr(self, "usuario_actual", None), dict):
            self.usuario_actual["documento"] = documento or "admin"
            self.usuario_actual["nombre"] = nombre

    def obtener_anio_lectivo_activo(self):
        anio_default = str(datetime.now().year)
        anio = self._get_config_plantel("anio_lectivo")
        anio = str(anio or "").strip()
        if not re.fullmatch(r"\d{4}", anio):
            return anio_default
        return anio

    def establecer_anio_lectivo_activo(self, anio):
        anio_txt = str(anio or "").strip()
        if not re.fullmatch(r"\d{4}", anio_txt):
            raise ValueError("Año lectivo inválido")
        self._set_config_sistema("anio_lectivo", anio_txt)

    # ---------- Autenticación ----------
    def authenticate(self):
        """Solicita la clave maestra mediante simpledialog. Devuelve True si OK."""
        clave_actual = self.get_master_key()
        respuesta = simpledialog.askstring(
            "Acceso SuperAdmin", "Clave maestra:", show="*", parent=self.parent
        )
        if respuesta is None:
            return False
        if respuesta == clave_actual:
            return True
        messagebox.showerror("Acceso denegado", "Clave incorrecta.", parent=self.parent)
        return False

    # ---------- Helpers Excel (pandas o openpyxl) ----------
    def _read_excel(self, path):
        if not os.path.exists(path):
            return None
        if _HAS_PANDAS:
            try:
                return pd.read_excel(path)
            except ImportError as ie:
                # pandas can't find an engine (usually openpyxl)
                messagebox.showerror(
                    "Dependencia faltante",
                    "Para leer archivos Excel se requiere instalar 'openpyxl'.\n"
                    "Ejecuta: pip install openpyxl",
                    parent=getattr(self, "win", None),
                )
                return None
            except Exception:
                # cualquier otro error al leer el excel se ignora y se devuelve None
                return None
        # fallback openpyxl -> DataFrame-lite (list of dicts)
        try:
            wb = load_workbook(path)
        except ImportError:
            messagebox.showerror(
                "Dependencia faltante",
                "Para trabajar sin pandas se requiere 'openpyxl'.\n"
                "Ejecuta: pip install openpyxl",
                parent=getattr(self, "win", None),
            )
            return None
        ws = wb.active
        rows = list(ws.values)
        if not rows:
            return None
        headers = rows[0]
        data = [dict(zip(headers, r)) for r in rows[1:]]
        import collections

        return collections.OrderedDict({"_rows": data, "_headers": headers})

    def _write_excel(self, path, df_like):
        if _HAS_PANDAS:
            df_like.to_excel(path, index=False)
            return
        # df_like is expected to be OrderedDict from _read_excel fallback or similar
        wb = Workbook()
        ws = wb.active
        headers = (
            df_like.get("_headers")
            if isinstance(df_like, dict)
            else list(df_like[0].keys()) if df_like else []
        )
        if headers:
            ws.append(list(headers))
        rows = df_like.get("_rows") if isinstance(df_like, dict) else df_like
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        wb.save(path)

    def _apply_modern_style(self):
        """Aplica un tema y paleta moderna consistente para toda la UI."""
        try:
            style = ttk.Style(self.win)
        except Exception:
            style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        default_font = ("Segoe UI", 10)
        try:
            self.win.option_add("*Font", default_font)
        except Exception:
            try:
                self.parent.option_add("*Font", default_font)
            except Exception:
                pass

        # Paleta
        bg = "#f6f8fb"
        fg = "#222222"
        accent = "#0d6efd"
        muted = "#6c757d"

        style.configure(".", background=bg, foreground=fg)
        style.configure("TFrame", background=bg)
        style.configure("TLabel", background=bg, foreground=fg, padding=4)
        style.configure("TButton", padding=6)
        style.configure("Accent.TButton", background=accent, foreground="white")
        style.map(
            "Accent.TButton",
            background=[("active", "!disabled", "#0b5ed7")],
            foreground=[("disabled", "#cccccc")],
        )
        style.configure("TEntry", fieldbackground="white", background="white")
        style.configure(
            "Treeview",
            rowheight=26,
            fieldbackground="white",
            background="white",
            foreground=fg,
        )
        style.configure("Treeview.Heading", font=(default_font[0], 10, "bold"))
        # --- Estilo barra tabs horizontal limpia ---
        style.configure("TNotebook", background=bg, borderwidth=0, padding=0)
        style.configure(
            "TNotebook.Tab",
            padding=(12, 8),
            font=(default_font[0], 10, "bold"),
            background=bg,
            foreground="#38506b",
            borderwidth=0,
            focuscolor=bg,
        )
        style.map(
            "TNotebook.Tab",
            background=[("selected", accent), ("active", "#e7f0fa"), ("!selected", bg)],
            foreground=[
                ("selected", "white"),
                ("active", accent),
                ("!selected", "#38506b"),
            ],
        )
        # Eliminar estilos tipo card
        style.configure(
            "Card.TLabelframe", background=bg, borderwidth=0, relief="flat", padding=0
        )
        style.configure(
            "Card.TLabelframe.Label",
            background=bg,
            foreground="#16324f",
            font=(default_font[0], 11, "bold"),
        )
        style.configure(
            "SectionTitle.TLabel",
            background="#f6f8fb",
            foreground="#16324f",
            font=(default_font[0], 15, "bold"),
        )
        style.configure(
            "SectionSub.TLabel",
            background="#f6f8fb",
            foreground="#5f7287",
            font=(default_font[0], 10),
        )
        style.configure(
            "CardHint.TLabel",
            background="#ffffff",
            foreground="#5f7287",
            font=(default_font[0], 9),
        )
        style.configure("Horizontal.TSeparator", background="#e9ecef")

    # ---------- Interface principal ----------

    def open_interface(self, cerrar_sesion_cb=None):
        self.win = tk.Toplevel(self.parent)
        self.win.title("Modulo SuperAdmin")
        self.win.geometry("900x600")

        # aplicar estilo moderno centralizado
        self._apply_modern_style()

        # Barra superior con botón Cerrar sesión y botón Certificados
        topbar = tk.Frame(self.win, bg="#0066cc", height=44)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)
        tk.Label(
            topbar,
            text="⚙️  SuperAdmin — Sistema de Evaluación",
            font=("Segoe UI", 12, "bold"),
            bg="#0066cc",
            fg="white",
        ).pack(side="left", padx=15, pady=10)
        self.lbl_superadmin_session = tk.Label(
            topbar,
            text="",
            font=("Segoe UI", 9),
            bg="#0066cc",
            fg="#dbeafe",
        )
        self.lbl_superadmin_session.pack(side="left", padx=(4, 0), pady=10)
        if cerrar_sesion_cb:
            tk.Button(
                topbar,
                text="🚪 Cerrar sesión",
                font=("Segoe UI", 10, "bold"),
                bg="#cc3300",
                fg="white",
                relief="flat",
                cursor="hand2",
                padx=12,
                pady=4,
                command=cerrar_sesion_cb,
            ).pack(side="right", padx=15, pady=7)

            self._actualizar_resumen_superadmin()

        # (El botón de acceso directo a Certificados ha sido eliminado; solo queda la pestaña en el Notebook)

        # Contenedor con scroll horizontal si overflow
        nb_container = tk.Frame(self.win, bg="#f6f8fb")
        nb_container.pack(fill="x", side="top")
        self.nb = ttk.Notebook(nb_container)
        self.nb.pack(fill="x", expand=False, padx=0, pady=0)
        # Habilitar scroll horizontal si hay muchas pestañas
        self.nb.enable_traversal()
        self.nb.bind(
            "<Configure>",
            lambda e: self.nb.tk.call(self.nb._w, "see", self.nb.index("current")),
        )

        # ==============================
        # NUEVA PESTAÑA: PLANILLAS
        # ==============================
        self.tab_planillas = ttk.Frame(self.nb)

        # Subnotebook para subpestañas de planillas
        self.nb_planillas = ttk.Notebook(self.tab_planillas)
        self.nb_planillas.pack(fill="both", expand=True)

        # Subpestaña: Planillas de Asistencia
        self.tab_planillas_asistencia = ttk.Frame(self.nb_planillas)
        self.nb_planillas.add(
            self.tab_planillas_asistencia, text="Planillas de asistencia"
        )

        asistencia_shell = ttk.Frame(self.tab_planillas_asistencia)
        asistencia_shell.pack(fill="both", expand=True)

        hero_asistencia = tk.Frame(asistencia_shell, bg="#0f5db8", padx=18, pady=10)
        hero_asistencia.pack(fill="x", padx=18, pady=(12, 6))
        tk.Label(
            hero_asistencia,
            text="Planillas de asistencia",
            font=("Segoe UI", 15, "bold"),
            bg="#0f5db8",
            fg="white",
        ).pack(anchor="w")

        asistencia_body = ttk.Frame(asistencia_shell, padding=(18, 0, 18, 14))
        asistencia_body.pack(fill="both", expand=True)
        asistencia_body.grid_columnconfigure(0, weight=1)
        asistencia_body.grid_rowconfigure(1, weight=1)

        # --- FILTROS ---
        frame_filtros_asist = ttk.LabelFrame(
            asistencia_body,
            text="Filtros",
            padding=8,
            style="Card.TLabelframe",
        )
        frame_filtros_asist.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        fila_filtros_asist = ttk.Frame(frame_filtros_asist)
        fila_filtros_asist.pack(fill="x")
        for columna in (1, 3, 5, 7, 9):
            fila_filtros_asist.grid_columnconfigure(columna, weight=1)

        ttk.Label(fila_filtros_asist, text="Sede:").grid(
            row=0, column=0, sticky="w", padx=(0, 4), pady=2
        )
        self.combo_sede_asist = ttk.Combobox(
            fila_filtros_asist, width=10, state="readonly"
        )
        self.combo_sede_asist.grid(row=0, column=1, sticky="ew", padx=(0, 10), pady=2)

        ttk.Label(fila_filtros_asist, text="Jornada:").grid(
            row=0, column=2, sticky="w", padx=(0, 4), pady=2
        )
        self.combo_jornada_asist = ttk.Combobox(
            fila_filtros_asist, width=8, state="readonly"
        )
        self.combo_jornada_asist.grid(
            row=0, column=3, sticky="ew", padx=(0, 10), pady=2
        )

        ttk.Label(fila_filtros_asist, text="Grado:").grid(
            row=0, column=4, sticky="w", padx=(0, 4), pady=2
        )
        self.combo_grado_asist = ttk.Combobox(
            fila_filtros_asist, width=8, state="readonly"
        )
        self.combo_grado_asist.grid(row=0, column=5, sticky="ew", padx=(0, 10), pady=2)

        ttk.Label(fila_filtros_asist, text="Curso:").grid(
            row=0, column=6, sticky="w", padx=(0, 4), pady=2
        )
        self.combo_curso_asist = ttk.Combobox(
            fila_filtros_asist, width=8, state="readonly"
        )
        self.combo_curso_asist.grid(row=0, column=7, sticky="ew", padx=(0, 10), pady=2)

        ttk.Label(fila_filtros_asist, text="Mes:").grid(
            row=0, column=8, sticky="w", padx=(0, 4), pady=2
        )
        self.combo_mes_asist = ttk.Combobox(
            fila_filtros_asist, width=10, state="readonly"
        )
        self.combo_mes_asist["values"] = [
            "Enero",
            "Febrero",
            "Marzo",
            "Abril",
            "Mayo",
            "Junio",
            "Julio",
            "Agosto",
            "Septiembre",
            "Octubre",
            "Noviembre",
            "Diciembre",
        ]
        self.combo_mes_asist.grid(row=0, column=9, sticky="ew", padx=(0, 10), pady=2)
        self.combo_mes_asist.set(
            [
                "Enero",
                "Febrero",
                "Marzo",
                "Abril",
                "Mayo",
                "Junio",
                "Julio",
                "Agosto",
                "Septiembre",
                "Octubre",
                "Noviembre",
                "Diciembre",
            ][datetime.now().month - 1]
        )
        self.btn_guardar_asistencia = self._crear_boton_si_permiso(
            ttk.Button,
            fila_filtros_asist,
            "desktop.superadmin.planillas.asistencia.guardar",
            text="Guardar asistencia",
            layout="grid",
            layout_kwargs={
                "row": 0,
                "column": 10,
                "sticky": "e",
                "padx": (0, 8),
                "pady": 2,
            },
        )
        self.btn_descargar_planilla = self._crear_boton_si_permiso(
            ttk.Button,
            fila_filtros_asist,
            "desktop.superadmin.planillas.asistencia.descargar",
            text="Descargar planilla",
            layout="grid",
            layout_kwargs={
                "row": 0,
                "column": 11,
                "sticky": "e",
                "padx": (0, 8),
                "pady": 2,
            },
        )
        self.lbl_total_asistencia = ttk.Label(
            fila_filtros_asist, text="0 estudiantes cargados", style="CardHint.TLabel"
        )
        self.lbl_total_asistencia.grid(
            row=0, column=12, sticky="e", padx=(4, 0), pady=2
        )

        # --- TABLA DE ASISTENCIA ---
        frame_tabla_asist = ttk.LabelFrame(
            asistencia_body,
            text="Listado de asistencia",
            padding=8,
            style="Card.TLabelframe",
        )
        frame_tabla_asist.grid(row=1, column=0, sticky="nsew")
        frame_tabla_asist.grid_columnconfigure(0, weight=1)
        frame_tabla_asist.grid_rowconfigure(0, weight=1)
        tabla_asist_body = ttk.Frame(frame_tabla_asist)
        tabla_asist_body.grid(row=0, column=0, sticky="nsew")
        tabla_asist_body.grid_columnconfigure(0, weight=1)
        tabla_asist_body.grid_rowconfigure(0, weight=1)
        self.asistencia_columnas_fijas = ("indice", "codigo", "nombre")
        self.asistencia_columnas_dias = []
        self.asistencia_columnas_totales = ("total_r", "total_f")
        self.tree_asistencia = ttk.Treeview(
            tabla_asist_body,
            columns=self.asistencia_columnas_fijas + self.asistencia_columnas_totales,
            show="headings",
            height=18,
        )
        self.tree_asistencia.grid(row=0, column=0, sticky="nsew")
        scrollbar_asistencia = ttk.Scrollbar(
            tabla_asist_body, orient="vertical", command=self.tree_asistencia.yview
        )
        scrollbar_asistencia.grid(row=0, column=1, sticky="ns")
        scrollbar_asistencia_x = ttk.Scrollbar(
            tabla_asist_body, orient="horizontal", command=self.tree_asistencia.xview
        )
        scrollbar_asistencia_x.grid(row=1, column=0, sticky="ew")
        self.tree_asistencia.configure(yscrollcommand=scrollbar_asistencia.set)
        self.tree_asistencia.configure(xscrollcommand=scrollbar_asistencia_x.set)

        # --- ESTADOS POSIBLES ---
        self.estados_asistencia = ["", "F", "R", "FJ", "RJ"]
        self.asistencia_celda_actual = {"item": None, "columna": None}

        def obtener_mes_numero_asistencia():
            meses = {
                "Enero": 1,
                "Febrero": 2,
                "Marzo": 3,
                "Abril": 4,
                "Mayo": 5,
                "Junio": 6,
                "Julio": 7,
                "Agosto": 8,
                "Septiembre": 9,
                "Octubre": 10,
                "Noviembre": 11,
                "Diciembre": 12,
            }
            mes_nombre = self.combo_mes_asist.get().strip()
            return meses.get(mes_nombre, datetime.now().month)

        def obtener_anio_asistencia():
            try:
                return int(str(self.obtener_anio_lectivo_activo() or "").strip())
            except Exception:
                return datetime.now().year

        def obtener_dias_habiles_asistencia():
            import calendar

            anio = obtener_anio_asistencia()
            mes_numero = obtener_mes_numero_asistencia()
            dias_habiles = []
            for dia in range(1, calendar.monthrange(anio, mes_numero)[1] + 1):
                fecha_dia = datetime(anio, mes_numero, dia)
                if fecha_dia.weekday() < 5:
                    dias_habiles.append((dia, fecha_dia.weekday()))
            return dias_habiles

        def mapear_estado_a_marca(estado):
            return {
                "": "",
                "P": "",
                "A": "F",
                "T": "R",
                "J": "FJ",
                "F": "F",
                "R": "R",
                "FJ": "FJ",
                "RJ": "RJ",
            }.get(str(estado or "").strip().upper(), "")

        def mapear_marca_a_estado(marca):
            marca_limpia = str(marca or "").strip().upper()
            return marca_limpia or None

        def configurar_columnas_asistencia():
            nombres_dia = ["Lun", "Mar", "Mie", "Jue", "Vie"]
            self.asistencia_columnas_dias = [
                (dia, f"dia_{dia:02d}", weekday)
                for dia, weekday in obtener_dias_habiles_asistencia()
            ]
            columnas = (
                self.asistencia_columnas_fijas
                + tuple(columna for _, columna, _ in self.asistencia_columnas_dias)
                + self.asistencia_columnas_totales
            )
            self.tree_asistencia.configure(columns=columnas)
            self.tree_asistencia.heading("indice", text="N")
            self.tree_asistencia.column(
                "indice", anchor="center", width=45, stretch=False
            )
            self.tree_asistencia.heading("codigo", text="COD.")
            self.tree_asistencia.column(
                "codigo", anchor="center", width=95, stretch=False
            )
            self.tree_asistencia.heading("nombre", text="APELLIDOS Y NOMBRES")
            self.tree_asistencia.column("nombre", anchor="w", width=280, stretch=True)
            for dia, columna, weekday in self.asistencia_columnas_dias:
                self.tree_asistencia.heading(
                    columna, text=f"{dia} {nombres_dia[weekday]}"
                )
                self.tree_asistencia.column(
                    columna, anchor="center", width=48, stretch=False
                )
            self.tree_asistencia.heading("total_r", text="R")
            self.tree_asistencia.column(
                "total_r", anchor="center", width=42, stretch=False
            )
            self.tree_asistencia.heading("total_f", text="F")
            self.tree_asistencia.column(
                "total_f", anchor="center", width=42, stretch=False
            )

        def actualizar_totales_fila(item):
            total_r = 0
            total_f = 0
            for _, columna, _ in self.asistencia_columnas_dias:
                marca = (
                    str(self.tree_asistencia.set(item, columna) or "").strip().upper()
                )
                if marca in {"R", "RJ"}:
                    total_r += 1
                elif marca in {"F", "FJ"}:
                    total_f += 1
            self.tree_asistencia.set(item, "total_r", total_r if total_r > 0 else "")
            self.tree_asistencia.set(item, "total_f", total_f if total_f > 0 else "")

        def obtener_columnas_editables_asistencia():
            return [columna for _, columna, _ in self.asistencia_columnas_dias]

        def seleccionar_celda_asistencia(item, columna):
            if not item or columna not in obtener_columnas_editables_asistencia():
                return False
            self.asistencia_celda_actual = {"item": item, "columna": columna}
            self.tree_asistencia.focus(item)
            self.tree_asistencia.selection_set(item)
            self.tree_asistencia.see(item)
            try:
                self.tree_asistencia.see(item)
            except Exception:
                pass
            self.tree_asistencia.focus_set()
            return True

        def obtener_siguiente_celda_asistencia(direccion_fila=0, direccion_columna=0):
            item_actual = self.asistencia_celda_actual.get("item")
            columna_actual = self.asistencia_celda_actual.get("columna")
            items = list(self.tree_asistencia.get_children())
            columnas_editables = obtener_columnas_editables_asistencia()
            if not items or not columnas_editables:
                return None, None
            if item_actual not in items:
                item_actual = items[0]
            if columna_actual not in columnas_editables:
                columna_actual = columnas_editables[0]
            indice_fila = items.index(item_actual)
            indice_columna = columnas_editables.index(columna_actual)
            indice_fila = max(0, min(len(items) - 1, indice_fila + direccion_fila))
            indice_columna = max(
                0,
                min(len(columnas_editables) - 1, indice_columna + direccion_columna),
            )
            return items[indice_fila], columnas_editables[indice_columna]

        def aplicar_marca_asistencia(marca, mover_horizontal=0):
            item = self.asistencia_celda_actual.get("item")
            columna = self.asistencia_celda_actual.get("columna")
            if not item or columna not in obtener_columnas_editables_asistencia():
                return "break"
            self.tree_asistencia.set(item, columna, marca)
            actualizar_totales_fila(item)
            if mover_horizontal:
                siguiente_item, siguiente_columna = obtener_siguiente_celda_asistencia(
                    direccion_columna=mover_horizontal
                )
                seleccionar_celda_asistencia(siguiente_item, siguiente_columna)
            return "break"

        def manejar_click_asistencia(event):
            item = self.tree_asistencia.identify_row(event.y)
            col = self.tree_asistencia.identify_column(event.x)
            if not item or not col.startswith("#"):
                return
            try:
                indice_columna = int(col[1:]) - 1
            except Exception:
                return
            columnas_actuales = self.tree_asistencia["columns"]
            if indice_columna < 0 or indice_columna >= len(columnas_actuales):
                return
            seleccionar_celda_asistencia(item, columnas_actuales[indice_columna])

        def mover_celda_asistencia(direccion_fila=0, direccion_columna=0):
            item, columna = obtener_siguiente_celda_asistencia(
                direccion_fila=direccion_fila,
                direccion_columna=direccion_columna,
            )
            seleccionar_celda_asistencia(item, columna)
            return "break"

        def manejar_tecla_asistencia(event):
            tecla = str(event.keysym or "").upper()
            char = str(event.char or "").upper()
            if tecla in {"LEFT", "RIGHT", "UP", "DOWN"}:
                desplazamientos = {
                    "LEFT": (0, -1),
                    "RIGHT": (0, 1),
                    "UP": (-1, 0),
                    "DOWN": (1, 0),
                }
                direccion_fila, direccion_columna = desplazamientos[tecla]
                return mover_celda_asistencia(direccion_fila, direccion_columna)
            if tecla in {"RETURN", "KP_ENTER", "SPACE"}:
                item = self.asistencia_celda_actual.get("item")
                columna = self.asistencia_celda_actual.get("columna")
                if item and columna:
                    bbox = self.tree_asistencia.bbox(item, columna)
                    if bbox:

                        class EventoManual:
                            pass

                        evento_manual = EventoManual()
                        evento_manual.x = bbox[0] + 2
                        evento_manual.y = bbox[1] + 2
                        editar_estado(evento_manual)
                return "break"
            if tecla in {"DELETE", "BACKSPACE"}:
                return aplicar_marca_asistencia("")
            atajos = {
                "F": "F",
                "R": "R",
                "J": "FJ",
                "X": "RJ",
                "1": "F",
                "2": "R",
                "3": "FJ",
                "4": "RJ",
                "0": "",
            }
            marca = atajos.get(char) or atajos.get(tecla)
            if marca is not None:
                return aplicar_marca_asistencia(marca, mover_horizontal=1)

        configurar_columnas_asistencia()

        def actualizar_resumen_asistencia():
            registros = self.tree_asistencia.get_children()
            total = len(registros)
            self.lbl_total_asistencia.config(
                text=(
                    "0 estudiantes cargados"
                    if total == 0
                    else (
                        "1 estudiante cargado"
                        if total == 1
                        else f"{total} estudiantes cargados"
                    )
                )
            )

        # --- FUNCIONES AUXILIARES ---
        def cargar_filtros_asistencia():
            try:
                from core import matricula as core_matricula

                sedes = core_matricula.listar_sedes()
                jornadas = core_matricula.listar_jornadas()
                grados = core_matricula.listar_grados_distintos(solo_activos=True)
                self.combo_sede_asist["values"] = sedes
                self.combo_jornada_asist["values"] = jornadas
                self.combo_grado_asist["values"] = grados
            except Exception:
                self.combo_sede_asist["values"] = []
                self.combo_jornada_asist["values"] = []
                self.combo_grado_asist["values"] = []

        cargar_filtros_asistencia()

        def cargar_cursos_asistencia(event=None):
            grado = self.combo_grado_asist.get().strip()
            if not grado:
                self.combo_curso_asist["values"] = []
                self.combo_curso_asist.set("")
                limpiar_tabla_asistencia()
                return
            try:
                from core import matricula as core_matricula

                cursos = core_matricula.listar_cursos_por_grado(grado)
                self.combo_curso_asist["values"] = cursos
                self.combo_curso_asist.set("")
                limpiar_tabla_asistencia()
            except Exception:
                self.combo_curso_asist["values"] = []
                self.combo_curso_asist.set("")
                limpiar_tabla_asistencia()

        self.combo_grado_asist.bind("<<ComboboxSelected>>", cargar_cursos_asistencia)

        def limpiar_tabla_asistencia():
            self.asistencia_celda_actual = {"item": None, "columna": None}
            for row in self.tree_asistencia.get_children():
                self.tree_asistencia.delete(row)
            actualizar_resumen_asistencia()

        def cargar_estudiantes_asistencia(mostrar_advertencia=False):
            limpiar_tabla_asistencia()
            sede = self.combo_sede_asist.get().strip()
            jornada = self.combo_jornada_asist.get().strip()
            grado = self.combo_grado_asist.get().strip()
            curso = self.combo_curso_asist.get().strip()
            mes_nombre = self.combo_mes_asist.get().strip()
            if not (sede and jornada and grado and curso and mes_nombre):
                if mostrar_advertencia:
                    messagebox.showwarning(
                        "Faltan filtros",
                        "Selecciona sede, jornada, grado, curso y mes para cargar la planilla.",
                    )
                return
            try:
                from core import matricula as core_matricula

                configurar_columnas_asistencia()
                estudiantes = core_matricula.listar_estudiantes(
                    sede, jornada, grado, curso
                )
                asistencia_registrada = {}
                anio_lectivo = obtener_anio_asistencia()
                prefijo_mes = (
                    f"{anio_lectivo:04d}-{obtener_mes_numero_asistencia():02d}"
                )
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute(
                    """
                    SELECT estudiante_id, fecha, estado
                    FROM asistencia
                    WHERE grado = ? AND curso = ? AND jornada = ? AND sede = ?
                      AND substr(fecha, 1, 7) = ?
                    """,
                    (grado, curso, jornada, sede, prefijo_mes),
                )
                for estudiante_id, fecha_db, estado_db in cursor.fetchall():
                    try:
                        dia_mes = int(str(fecha_db).split("-")[-1])
                    except Exception:
                        continue
                    asistencia_registrada.setdefault(str(estudiante_id), {})[
                        dia_mes
                    ] = mapear_estado_a_marca(estado_db)
                conn.close()
                for idx, est in enumerate(estudiantes, 1):
                    nombre_completo = " ".join(
                        parte.strip()
                        for parte in [est[2], est[3], est[4], est[5]]
                        if str(parte or "").strip()
                    )
                    marcas_estudiante = asistencia_registrada.get(str(est[0]), {})
                    valores = [idx, est[1], nombre_completo]
                    for dia, _columna, _weekday in self.asistencia_columnas_dias:
                        valores.append(marcas_estudiante.get(dia, ""))
                    valores.extend(["", ""])
                    item = self.tree_asistencia.insert(
                        "",
                        "end",
                        values=tuple(valores),
                        tags=(str(est[0]),),
                    )
                    actualizar_totales_fila(item)
                if (
                    self.tree_asistencia.get_children()
                    and self.asistencia_columnas_dias
                ):
                    seleccionar_celda_asistencia(
                        self.tree_asistencia.get_children()[0],
                        self.asistencia_columnas_dias[0][1],
                    )
                actualizar_resumen_asistencia()
            except Exception as e:
                messagebox.showerror("Error", f"No se pudieron cargar estudiantes: {e}")

        def intentar_carga_automatica(_event=None):
            cargar_estudiantes_asistencia(mostrar_advertencia=False)

        self.combo_sede_asist.bind("<<ComboboxSelected>>", intentar_carga_automatica)
        self.combo_jornada_asist.bind("<<ComboboxSelected>>", intentar_carga_automatica)
        self.combo_curso_asist.bind("<<ComboboxSelected>>", intentar_carga_automatica)
        self.combo_mes_asist.bind("<<ComboboxSelected>>", intentar_carga_automatica)

        def editar_estado(event):
            item = self.tree_asistencia.identify_row(event.y)
            col = self.tree_asistencia.identify_column(event.x)
            if not item or not col.startswith("#"):
                return
            try:
                indice_columna = int(col[1:]) - 1
            except Exception:
                return
            columnas_actuales = self.tree_asistencia["columns"]
            if indice_columna < 0 or indice_columna >= len(columnas_actuales):
                return
            columna_id = columnas_actuales[indice_columna]
            columnas_editables = {
                columna for _, columna, _ in self.asistencia_columnas_dias
            }
            if columna_id not in columnas_editables:
                return
            x, y, width, height = self.tree_asistencia.bbox(item, columna_id)
            if not width or not height:
                return
            estado_actual = self.tree_asistencia.set(item, columna_id)
            top = tk.Toplevel(self.tab_planillas_asistencia)
            top.overrideredirect(True)
            top.geometry(
                f"{width}x{height}+{self.tree_asistencia.winfo_rootx() + x}+{self.tree_asistencia.winfo_rooty() + y}"
            )
            var_estado = tk.StringVar(value=estado_actual)
            combo = ttk.Combobox(
                top,
                values=self.estados_asistencia,
                textvariable=var_estado,
                state="readonly",
                width=3,
            )
            combo.pack(fill="both", expand=True)
            combo.focus_set()

            def cerrar(event=None):
                self.tree_asistencia.set(
                    item, columna_id, str(var_estado.get() or "").strip().upper()
                )
                actualizar_totales_fila(item)
                top.destroy()

            combo.bind("<<ComboboxSelected>>", cerrar)
            combo.bind("<Return>", cerrar)
            combo.bind("<FocusOut>", lambda e: top.destroy())

        self.tree_asistencia.bind("<Double-1>", editar_estado)
        self.tree_asistencia.bind("<Button-1>", manejar_click_asistencia, add="+")
        self.tree_asistencia.bind("<Key>", manejar_tecla_asistencia)
        actualizar_resumen_asistencia()

        def guardar_asistencia():
            if not self.tree_asistencia.get_children():
                messagebox.showwarning("Sin datos", "No hay estudiantes cargados.")
                return
            mes_seleccionado = self.combo_mes_asist.get().strip()
            if not mes_seleccionado:
                messagebox.showwarning(
                    "Mes requerido",
                    "Selecciona el mes de la planilla antes de guardar.",
                )
                return
            if not messagebox.askyesno(
                "Confirmar",
                f"¿Guardar la planilla de asistencia de {mes_seleccionado}?",
            ):
                return
            sede = self.combo_sede_asist.get().strip()
            jornada = self.combo_jornada_asist.get().strip()
            grado = self.combo_grado_asist.get().strip()
            curso = self.combo_curso_asist.get().strip()
            anio_lectivo = obtener_anio_asistencia()
            mes_numero = obtener_mes_numero_asistencia()
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                for item in self.tree_asistencia.get_children():
                    estudiante_id = self.tree_asistencia.item(item, "tags")[0]
                    for dia, columna, _weekday in self.asistencia_columnas_dias:
                        fecha = f"{anio_lectivo:04d}-{mes_numero:02d}-{dia:02d}"
                        estado = mapear_marca_a_estado(
                            self.tree_asistencia.set(item, columna)
                        )
                        if estado:
                            cursor.execute(
                                """
                                INSERT INTO asistencia (estudiante_id, fecha, estado, grado, curso, jornada, sede)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                                ON CONFLICT(estudiante_id, fecha) DO UPDATE SET
                                    estado=excluded.estado,
                                    grado=excluded.grado,
                                    curso=excluded.curso,
                                    jornada=excluded.jornada,
                                    sede=excluded.sede
                            """,
                                (
                                    estudiante_id,
                                    fecha,
                                    estado,
                                    grado,
                                    curso,
                                    jornada,
                                    sede,
                                ),
                            )
                        else:
                            cursor.execute(
                                "DELETE FROM asistencia WHERE estudiante_id = ? AND fecha = ?",
                                (estudiante_id, fecha),
                            )
                conn.commit()
                conn.close()
                messagebox.showinfo(
                    "Éxito",
                    "Planilla guardada correctamente. Los días sin marca quedaron limpios.",
                )
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar la asistencia: {e}")

        self.btn_guardar_asistencia.config(command=guardar_asistencia)

        def descargar_planilla():
            if not self.tree_asistencia.get_children():
                messagebox.showwarning("Sin datos", "No hay estudiantes cargados.")
                return
            try:
                import calendar
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
                from openpyxl.utils import get_column_letter
                from openpyxl.drawing.image import Image as XLImage
                from tkinter import filedialog

                mes_nombre = (
                    self.combo_mes_asist.get().strip() or datetime.now().strftime("%B")
                )
                meses = {
                    "Enero": 1,
                    "Febrero": 2,
                    "Marzo": 3,
                    "Abril": 4,
                    "Mayo": 5,
                    "Junio": 6,
                    "Julio": 7,
                    "Agosto": 8,
                    "Septiembre": 9,
                    "Octubre": 10,
                    "Noviembre": 11,
                    "Diciembre": 12,
                }
                mes_numero = meses.get(mes_nombre, datetime.now().month)
                try:
                    anio_lectivo = int(
                        str(self.obtener_anio_lectivo_activo() or "").strip()
                    )
                except Exception:
                    anio_lectivo = datetime.now().year

                jornada = self.combo_jornada_asist.get().strip() or "-"
                grado = self.combo_grado_asist.get().strip() or "-"
                curso = self.combo_curso_asist.get().strip() or "-"
                sede = self.combo_sede_asist.get().strip() or "-"
                periodo = "-"
                prefijo_mes = f"{anio_lectivo:04d}-{mes_numero:02d}"

                estudiantes = []
                for item in self.tree_asistencia.get_children():
                    tags = self.tree_asistencia.item(item, "tags")
                    estudiantes.append(
                        {
                            "indice": self.tree_asistencia.set(item, "indice"),
                            "codigo": self.tree_asistencia.set(item, "codigo"),
                            "nombre": self.tree_asistencia.set(item, "nombre"),
                            "estudiante_id": str(tags[0]) if tags else "",
                        }
                    )

                asistencia_registrada = {}
                try:
                    conn = sqlite3.connect(self.db_path)
                    cur = conn.cursor()
                    cur.execute(
                        """
                        SELECT estudiante_id, fecha, estado
                        FROM asistencia
                        WHERE grado = ? AND curso = ? AND jornada = ? AND sede = ?
                          AND substr(fecha, 1, 7) = ?
                        """,
                        (grado, curso, jornada, sede, prefijo_mes),
                    )
                    for estudiante_id, fecha_db, estado_db in cur.fetchall():
                        try:
                            dia_mes = int(str(fecha_db).split("-")[-1])
                        except Exception:
                            continue
                        marca = mapear_estado_a_marca(estado_db)
                        asistencia_registrada.setdefault(str(estudiante_id), {})[
                            dia_mes
                        ] = marca
                    conn.close()
                except Exception:
                    asistencia_registrada = {}

                dias_habiles = []
                for dia in range(
                    1, calendar.monthrange(anio_lectivo, mes_numero)[1] + 1
                ):
                    fecha_dia = datetime(anio_lectivo, mes_numero, dia)
                    if fecha_dia.weekday() < 5:
                        dias_habiles.append((dia, fecha_dia.weekday()))

                sugerido = f"planilla_asistencia_{grado}_{curso}_{mes_nombre.lower()}_{anio_lectivo}.xlsx".replace(
                    " ", "_"
                )
                ruta = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel", "*.xlsx")],
                    initialfile=sugerido,
                )
                if not ruta:
                    return

                wb = Workbook()
                ws = wb.active
                ws.title = "Asistencia"

                thin = Side(style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                center = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                left = Alignment(horizontal="left", vertical="center", wrap_text=True)

                institucion = (
                    self._get_config_plantel("nombre_institucion")
                    or "INSTITUCIÓN EDUCATIVA"
                ).upper()
                resolucion = self._get_config_plantel("resolucion_aprobacion") or ""
                decreto = self._get_config_plantel("decreto_funcionamiento") or ""
                dane = self._get_config_plantel("codigo_dane") or ""
                nit = self._get_config_plantel("nit") or ""
                logo_path = self._get_config_plantel("logo_path") or ""

                total_columnas = 3 + len(dias_habiles) + 2
                ultima_col = get_column_letter(total_columnas)

                ws.merge_cells(f"C1:{ultima_col}1")
                ws["C1"] = institucion
                ws["C1"].font = Font(name="Arial", size=14, bold=True)
                ws["C1"].alignment = center

                ws.merge_cells(f"C2:{ultima_col}2")
                ws["C2"] = (
                    f"Resolución de aprobación: {resolucion}    Decreto de funcionamiento: {decreto}"
                ).strip()
                ws["C2"].font = Font(name="Arial", size=8, bold=False)
                ws["C2"].alignment = center

                ws.merge_cells(f"C3:{ultima_col}3")
                ws["C3"] = f"DANE: {dane}    NIT: {nit}".strip()
                ws["C3"].font = Font(name="Arial", size=9, bold=True)
                ws["C3"].alignment = center

                ws.merge_cells(f"A4:{ultima_col}4")
                ws["A4"].border = Border(bottom=thin)

                if logo_path and os.path.exists(logo_path):
                    try:
                        img = XLImage(logo_path)
                        img.width = 55
                        img.height = 55
                        ws.add_image(img, "A1")
                    except Exception:
                        pass

                info_row = 5
                month_row = 6
                bloques = [
                    (1, 4, f"Jornada: {jornada}"),
                    (5, 8, f"Grado: {grado}"),
                    (9, 11, f"Curso: {curso}"),
                    (12, 14, f"Periodo: {periodo}"),
                    (15, total_columnas, f"Sede: {sede}"),
                ]
                for inicio, fin, valor in bloques:
                    ws.merge_cells(
                        start_row=info_row,
                        start_column=inicio,
                        end_row=info_row,
                        end_column=fin,
                    )
                    celda = ws.cell(info_row, inicio)
                    celda.value = valor
                    celda.font = Font(name="Arial", size=9, bold=True)
                    celda.alignment = left if inicio >= 15 else center
                    for col in range(inicio, fin + 1):
                        ws.cell(info_row, col).border = border

                ws.merge_cells(
                    start_row=month_row,
                    start_column=4,
                    end_row=month_row,
                    end_column=total_columnas,
                )
                ws.cell(month_row, 4).value = (
                    f"Mes: {mes_nombre}    Año Lectivo: {anio_lectivo}"
                )
                ws.cell(month_row, 4).font = Font(name="Arial", size=10, bold=True)
                ws.cell(month_row, 4).alignment = center
                for col in range(4, total_columnas + 1):
                    ws.cell(month_row, col).border = border

                header_top = 8
                header_bottom = 9
                ws.merge_cells(
                    start_row=header_top,
                    start_column=1,
                    end_row=header_bottom,
                    end_column=1,
                )
                ws.merge_cells(
                    start_row=header_top,
                    start_column=2,
                    end_row=header_bottom,
                    end_column=2,
                )
                ws.merge_cells(
                    start_row=header_top,
                    start_column=3,
                    end_row=header_bottom,
                    end_column=3,
                )
                ws.cell(header_top, 1).value = "N"
                ws.cell(header_top, 2).value = "COD."
                ws.cell(header_top, 3).value = "APELLIDOS Y NOMBRES"

                day_col_start = 4
                nombres_dia = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]
                for offset, (dia, weekday) in enumerate(dias_habiles):
                    col = day_col_start + offset
                    ws.cell(header_top, col).value = dia
                    ws.cell(header_bottom, col).value = nombres_dia[weekday]

                total_r_col = day_col_start + len(dias_habiles)
                total_f_col = total_r_col + 1
                ws.merge_cells(
                    start_row=header_top,
                    start_column=total_r_col,
                    end_row=header_top,
                    end_column=total_f_col,
                )
                ws.cell(header_top, total_r_col).value = "TOT"
                ws.cell(header_bottom, total_r_col).value = "R"
                ws.cell(header_bottom, total_f_col).value = "F"

                for row in (header_top, header_bottom):
                    for col in range(1, total_columnas + 1):
                        cell = ws.cell(row, col)
                        cell.font = Font(name="Arial", size=8, bold=True)
                        cell.alignment = center
                        cell.border = border

                data_row = 10
                for idx, estudiante in enumerate(estudiantes, start=1):
                    row = data_row + idx - 1
                    ws.cell(row, 1).value = idx
                    ws.cell(row, 2).value = estudiante["codigo"]
                    ws.cell(row, 3).value = estudiante["nombre"]
                    marks = asistencia_registrada.get(estudiante["estudiante_id"], {})
                    total_r = 0
                    total_f = 0
                    for offset, (dia, _weekday) in enumerate(dias_habiles):
                        col = day_col_start + offset
                        marca = marks.get(dia, "")
                        ws.cell(row, col).value = marca
                        if marca in {"R", "RJ"}:
                            total_r += 1
                        elif marca in {"F", "FJ"}:
                            total_f += 1
                    ws.cell(row, total_r_col).value = total_r if total_r > 0 else ""
                    ws.cell(row, total_f_col).value = total_f if total_f > 0 else ""

                    for col in range(1, total_columnas + 1):
                        cell = ws.cell(row, col)
                        cell.border = border
                        cell.alignment = left if col == 3 else center
                        cell.font = Font(name="Arial", size=8)

                leyenda_row = data_row + len(estudiantes) + 1
                ws.merge_cells(
                    start_row=leyenda_row,
                    start_column=3,
                    end_row=leyenda_row,
                    end_column=total_columnas,
                )
                ws.cell(leyenda_row, 3).value = (
                    "Marcar: F=Falta    R=Retardo    FJ=Falta Justificada    RJ=Retardo Justificado"
                )
                ws.cell(leyenda_row, 3).font = Font(name="Arial", size=8)
                ws.cell(leyenda_row, 3).alignment = center

                ws.column_dimensions["A"].width = 4
                ws.column_dimensions["B"].width = 12
                ws.column_dimensions["C"].width = 38
                for col in range(day_col_start, total_r_col):
                    ws.column_dimensions[get_column_letter(col)].width = 4
                ws.column_dimensions[get_column_letter(total_r_col)].width = 4
                ws.column_dimensions[get_column_letter(total_f_col)].width = 4

                for row in range(1, leyenda_row + 1):
                    ws.row_dimensions[row].height = 18
                ws.row_dimensions[1].height = 20
                ws.row_dimensions[2].height = 16
                ws.row_dimensions[3].height = 16
                ws.row_dimensions[4].height = 8
                ws.row_dimensions[header_bottom].height = 18

                ws.freeze_panes = "D10"
                ws.sheet_view.showGridLines = False
                ws.page_setup.orientation = "landscape"
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.print_area = f"A1:{ultima_col}{leyenda_row}"

                wb.save(ruta)
                messagebox.showinfo("Éxito", f"Planilla exportada a {ruta}")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo exportar la planilla: {e}")

        self.btn_descargar_planilla.config(command=descargar_planilla)

        # --- CREAR TABLA EN BD SI NO EXISTE ---
        def crear_tabla_asistencia():
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute(
                    """
                    CREATE TABLE IF NOT EXISTS asistencia (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        estudiante_id INTEGER,
                        fecha TEXT,
                        estado TEXT,
                        grado TEXT,
                        curso TEXT,
                        jornada TEXT,
                        sede TEXT
                    )
                """
                )
                cursor.execute(
                    "CREATE UNIQUE INDEX IF NOT EXISTS idx_asistencia_estudiante_fecha ON asistencia(estudiante_id, fecha)"
                )
                conn.commit()
                conn.close()
            except Exception:
                pass

        crear_tabla_asistencia()

        # Subpestaña: Planillas de Calificaciones
        self.tab_planillas_calificaciones = ttk.Frame(self.nb_planillas)
        self.nb_planillas.add(
            self.tab_planillas_calificaciones, text="Planillas de calificaciones"
        )

        calificaciones_shell = ttk.Frame(self.tab_planillas_calificaciones)
        calificaciones_shell.pack(fill="both", expand=True)

        hero_calificaciones = tk.Frame(
            calificaciones_shell, bg="#1f5f8f", padx=18, pady=10
        )
        hero_calificaciones.pack(fill="x", padx=18, pady=(12, 6))
        tk.Label(
            hero_calificaciones,
            text="Planillas de calificaciones",
            font=("Segoe UI", 15, "bold"),
            bg="#1f5f8f",
            fg="white",
        ).pack(anchor="w")

        planillas_body = ttk.Frame(calificaciones_shell, padding=(18, 0, 18, 14))
        planillas_body.pack(fill="both", expand=True)
        planillas_body.grid_columnconfigure(0, weight=1)
        planillas_body.grid_rowconfigure(1, weight=1)

        frame_filtros_planilla = ttk.LabelFrame(
            planillas_body,
            text="Filtros",
            padding=8,
            style="Card.TLabelframe",
        )
        frame_filtros_planilla.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        fila_filtros_planilla = ttk.Frame(frame_filtros_planilla)
        fila_filtros_planilla.pack(fill="x")
        for columna in (1, 3, 5, 7):
            fila_filtros_planilla.grid_columnconfigure(columna, weight=1)

        self.planilla_calif_areas = {}
        self.planilla_calif_columnas_fijas = ("codigo", "nombre")
        self.planilla_calif_columnas_notas = []
        self.planilla_calif_columna_promedio = "promedio"
        self.planilla_calif_celda_actual = {"item": None, "columna": None}

        ttk.Label(fila_filtros_planilla, text="Grado:").grid(
            row=0, column=0, sticky="w", padx=(0, 4), pady=2
        )
        self.combo_grado_planilla = ttk.Combobox(
            fila_filtros_planilla, width=8, state="readonly"
        )
        self.combo_grado_planilla.grid(
            row=0, column=1, sticky="ew", padx=(0, 10), pady=2
        )

        ttk.Label(fila_filtros_planilla, text="Curso:").grid(
            row=0, column=2, sticky="w", padx=(0, 4), pady=2
        )
        self.combo_curso_planilla = ttk.Combobox(
            fila_filtros_planilla, width=8, state="readonly"
        )
        self.combo_curso_planilla.grid(
            row=0, column=3, sticky="ew", padx=(0, 10), pady=2
        )

        ttk.Label(fila_filtros_planilla, text="Área:").grid(
            row=0, column=4, sticky="w", padx=(0, 4), pady=2
        )
        self.combo_area_planilla = ttk.Combobox(
            fila_filtros_planilla, width=18, state="readonly"
        )
        self.combo_area_planilla.grid(
            row=0, column=5, sticky="ew", padx=(0, 10), pady=2
        )

        ttk.Label(fila_filtros_planilla, text="Periodo:").grid(
            row=0, column=6, sticky="w", padx=(0, 4), pady=2
        )
        self.combo_periodo_planilla = ttk.Combobox(
            fila_filtros_planilla,
            values=["Periodo 1", "Periodo 2", "Periodo 3", "Periodo 4"],
            width=12,
            state="readonly",
        )
        self.combo_periodo_planilla.grid(
            row=0, column=7, sticky="ew", padx=(0, 10), pady=2
        )
        self.combo_periodo_planilla.set("Periodo 1")

        self.btn_guardar_planilla_calif = self._crear_boton_si_permiso(
            ttk.Button,
            fila_filtros_planilla,
            "desktop.superadmin.planillas.calificaciones.guardar",
            text="Guardar planilla",
            layout="grid",
            layout_kwargs={
                "row": 0,
                "column": 8,
                "sticky": "e",
                "padx": (0, 8),
                "pady": 2,
            },
        )
        self.btn_exportar_planilla_calif = self._crear_boton_si_permiso(
            ttk.Button,
            fila_filtros_planilla,
            "desktop.superadmin.planillas.calificaciones.descargar",
            text="Descargar planilla",
            layout="grid",
            layout_kwargs={
                "row": 0,
                "column": 9,
                "sticky": "e",
                "padx": (0, 8),
                "pady": 2,
            },
        )
        self.lbl_total_planilla = ttk.Label(
            fila_filtros_planilla,
            text="0 estudiantes cargados",
            style="CardHint.TLabel",
        )
        self.lbl_total_planilla.grid(row=0, column=10, sticky="e", pady=2)

        frame_tabla_planilla = ttk.LabelFrame(
            planillas_body,
            text="Planilla de notas",
            padding=8,
            style="Card.TLabelframe",
        )
        frame_tabla_planilla.grid(row=1, column=0, sticky="nsew")
        frame_tabla_planilla.grid_columnconfigure(0, weight=1)
        frame_tabla_planilla.grid_rowconfigure(0, weight=1)
        tabla_planilla_body = ttk.Frame(frame_tabla_planilla)
        tabla_planilla_body.grid(row=0, column=0, sticky="nsew")
        tabla_planilla_body.grid_columnconfigure(0, weight=1)
        tabla_planilla_body.grid_rowconfigure(1, weight=1)

        planilla_header_bg = "#356A9A"
        planilla_header_fg = "#FFFFFF"
        self.canvas_header_planilla = tk.Canvas(
            tabla_planilla_body,
            height=84,
            highlightthickness=0,
            background=planilla_header_bg,
        )
        self.canvas_header_planilla.grid(row=0, column=0, sticky="ew")

        self.tree_planilla = ttk.Treeview(
            tabla_planilla_body,
            columns=self.planilla_calif_columnas_fijas
            + (self.planilla_calif_columna_promedio,),
            show="",
            height=18,
        )
        self.tree_planilla.grid(row=1, column=0, sticky="nsew")
        self.tree_planilla.column("#0", width=0, minwidth=0, stretch=False)
        self.tree_planilla.heading("#0", text="")
        scrollbar_planilla_y = ttk.Scrollbar(
            tabla_planilla_body, orient="vertical", command=self.tree_planilla.yview
        )
        scrollbar_planilla_y.grid(row=1, column=1, sticky="ns")
        scrollbar_planilla_x = ttk.Scrollbar(tabla_planilla_body, orient="horizontal")
        scrollbar_planilla_x.grid(row=2, column=0, sticky="ew")
        self.tree_planilla.configure(yscrollcommand=scrollbar_planilla_y.set)

        def desplazar_horizontal_planilla(*args):
            self.tree_planilla.xview(*args)
            self.canvas_header_planilla.xview(*args)

        def actualizar_scroll_horizontal_planilla(inicio, fin):
            scrollbar_planilla_x.set(inicio, fin)
            try:
                self.canvas_header_planilla.xview_moveto(inicio)
            except Exception:
                pass

        scrollbar_planilla_x.configure(command=desplazar_horizontal_planilla)
        self.tree_planilla.configure(
            xscrollcommand=actualizar_scroll_horizontal_planilla
        )
        self.tree_planilla.bind(
            "<Configure>", lambda event: dibujar_encabezado_planilla_calif(), add="+"
        )

        def obtener_anio_planilla_calif():
            try:
                return int(str(self.obtener_anio_lectivo_activo() or "").strip())
            except Exception:
                return datetime.now().year

        def obtener_config_planilla_calif():
            componentes_raw = self._obtener_componentes_evaluacion()
            try:
                nota_min = float(self._get_config_plantel("nota_min") or 1.0)
            except Exception:
                nota_min = 1.0
            try:
                nota_max = float(self._get_config_plantel("nota_max") or 5.0)
            except Exception:
                nota_max = 5.0
            try:
                decimales = int(float(self._get_config_plantel("decimales") or 1))
            except Exception:
                decimales = 1
            componentes = []
            solo_cognitivo = (
                self._coerce_int(componentes_raw["examen"].get("cantidad"), 0) == 0
                and self._coerce_int(
                    componentes_raw["autoevaluacion"].get("cantidad"), 0
                )
                == 0
            )
            for clave in ("cognitivo", "examen", "autoevaluacion"):
                componente = dict(componentes_raw.get(clave) or {})
                componente["clave"] = clave
                componente["cantidad"] = max(
                    0, self._coerce_int(componente.get("cantidad"), 0)
                )
                componente["porcentaje"] = self._coerce_int(
                    componente.get("porcentaje"), 0
                )
                componentes.append(componente)
            return {
                "componentes": componentes,
                "solo_cognitivo": solo_cognitivo,
                "cantidad_total": sum(c["cantidad"] for c in componentes),
                "nota_min": nota_min,
                "nota_max": nota_max,
                "decimales": max(0, min(3, decimales)),
            }

        def formatear_porcentaje_encabezado_planilla(valor):
            try:
                return f"{float(valor):.1f}%"
            except Exception:
                return "0.0%"

        def obtener_columnas_visibles_planilla():
            columnas_visibles = self.tree_planilla.cget("displaycolumns")
            if columnas_visibles in (None, "", "#all"):
                return tuple(self.tree_planilla["columns"])
            return tuple(columnas_visibles)

        def obtener_posiciones_reales_columnas_planilla():
            columnas = obtener_columnas_visibles_planilla()
            posiciones = {}
            item_referencia = None
            try:
                item_referencia = next(iter(self.tree_planilla.get_children()), None)
            except Exception:
                item_referencia = None

            cursor_x = 0
            for columna in columnas:
                ancho = 80
                try:
                    ancho = int(float(self.tree_planilla.column(columna, "width") or 0))
                except Exception:
                    pass
                ancho = max(40, ancho)

                bbox = None
                if item_referencia:
                    try:
                        bbox = self.tree_planilla.bbox(item_referencia, columna)
                    except Exception:
                        bbox = None

                if bbox and len(bbox) == 4:
                    x, _y, width, _height = bbox
                    width = max(40, int(width or ancho))
                    posiciones[columna] = (int(x), int(x) + width)
                    cursor_x = max(cursor_x, int(x) + width)
                else:
                    posiciones[columna] = (cursor_x, cursor_x + ancho)
                    cursor_x += ancho

            return posiciones, max(cursor_x, 1)

        def calcular_ancho_columna_planilla(componente, cantidad):
            cantidad = max(1, int(cantidad or 1))
            nombre = str(componente.get("nombre") or "").strip()
            porcentaje = formatear_porcentaje_encabezado_planilla(
                componente.get("porcentaje")
            )
            ancho_base = 62
            ancho_texto = max(len(nombre), len(f"[{porcentaje}]")) * 7 + 26
            ancho_total = max(ancho_base * cantidad, min(220, ancho_texto))
            return max(ancho_base, int((ancho_total + cantidad - 1) / cantidad))

        def dibujar_encabezado_planilla_calif(event=None):
            canvas = self.canvas_header_planilla
            if not canvas.winfo_exists():
                return
            canvas.delete("all")
            altura_superior = 22
            altura_componente = 36
            altura_notas = 26
            altura_total = altura_superior + altura_componente + altura_notas
            columnas = obtener_columnas_visibles_planilla()
            posiciones, ancho_total = obtener_posiciones_reales_columnas_planilla()
            canvas.configure(scrollregion=(0, 0, ancho_total, altura_total))
            canvas.create_rectangle(
                0,
                0,
                ancho_total,
                altura_total,
                fill=planilla_header_bg,
                outline=planilla_header_bg,
            )

            for columna, texto in (
                ("codigo", "Cód."),
                ("nombre", "Apellidos/Nombres"),
            ):
                if columna not in posiciones:
                    continue
                x1, x2 = posiciones[columna]
                canvas.create_rectangle(
                    x1, 0, x2, altura_total, outline=planilla_header_fg, width=1
                )
                canvas.create_text(
                    (x1 + x2) / 2,
                    altura_total / 2,
                    text=texto,
                    fill=planilla_header_fg,
                    font=("Segoe UI", 10, "bold"),
                    justify="center",
                    width=max(20, x2 - x1 - 8),
                )

            if self.planilla_calif_columnas_notas:
                x_notas_1 = posiciones[
                    self.planilla_calif_columnas_notas[0]["columna"]
                ][0]
                x_notas_2 = posiciones[
                    self.planilla_calif_columnas_notas[-1]["columna"]
                ][1]
                canvas.create_rectangle(
                    x_notas_1,
                    0,
                    x_notas_2,
                    altura_superior,
                    outline=planilla_header_fg,
                    width=1,
                )
                canvas.create_text(
                    (x_notas_1 + x_notas_2) / 2,
                    altura_superior / 2,
                    text="Desempeño Evaluado",
                    fill=planilla_header_fg,
                    font=("Segoe UI", 10, "bold"),
                )

                cfg = obtener_config_planilla_calif()
                metas_por_componente = {}
                for meta in self.planilla_calif_columnas_notas:
                    metas_por_componente.setdefault(meta["componente"], []).append(meta)

                for componente in cfg["componentes"]:
                    metas = metas_por_componente.get(componente["clave"], [])
                    if not metas:
                        continue
                    x1 = posiciones[metas[0]["columna"]][0]
                    x2 = posiciones[metas[-1]["columna"]][1]
                    canvas.create_rectangle(
                        x1,
                        altura_superior,
                        x2,
                        altura_total,
                        outline=planilla_header_fg,
                        width=1,
                    )
                    canvas.create_text(
                        (x1 + x2) / 2,
                        altura_superior + (altura_componente / 2),
                        text=(
                            f"{componente['nombre']}\n"
                            f"[{formatear_porcentaje_encabezado_planilla(componente['porcentaje'])}]"
                        ),
                        fill=planilla_header_fg,
                        font=("Segoe UI", 9, "bold"),
                        justify="center",
                        width=max(20, x2 - x1 - 8),
                    )

                y_notas_1 = altura_superior + altura_componente
                y_notas_2 = altura_total
                for meta in self.planilla_calif_columnas_notas:
                    x1 = posiciones[meta["columna"]][0]
                    x2 = posiciones[meta["columna"]][1]
                    canvas.create_rectangle(
                        x1,
                        y_notas_1,
                        x2,
                        y_notas_2,
                        fill="#F2F4F7",
                        outline=planilla_header_fg,
                        width=1,
                    )
                    canvas.create_text(
                        (x1 + x2) / 2,
                        (y_notas_1 + y_notas_2) / 2,
                        text=meta["encabezado_corto"],
                        fill="#0F2942",
                        font=("Segoe UI", 10, "bold"),
                    )

            if self.planilla_calif_columna_promedio in posiciones:
                x1, x2 = posiciones[self.planilla_calif_columna_promedio]
                canvas.create_rectangle(
                    x1, 0, x2, altura_total, outline=planilla_header_fg, width=1
                )
                canvas.create_text(
                    (x1 + x2) / 2,
                    altura_total / 2,
                    text="Definitiva",
                    fill=planilla_header_fg,
                    font=("Segoe UI", 10, "bold"),
                    justify="center",
                    width=max(20, x2 - x1 - 8),
                )
                y_notas_1 = altura_superior + altura_componente
                y_notas_2 = altura_total
                canvas.create_rectangle(
                    x1,
                    y_notas_1,
                    x2,
                    y_notas_2,
                    fill="#F2F4F7",
                    outline=planilla_header_fg,
                    width=1,
                )
                canvas.create_text(
                    (x1 + x2) / 2,
                    (y_notas_1 + y_notas_2) / 2,
                    text="DEF",
                    fill="#0F2942",
                    font=("Segoe UI", 10, "bold"),
                )

        def formatear_nota_planilla(valor):
            if valor in (None, ""):
                return ""
            cfg = obtener_config_planilla_calif()
            return f"{float(valor):.{cfg['decimales']}f}"

        def configurar_columnas_planilla_calif():
            cfg = obtener_config_planilla_calif()
            self.planilla_calif_columnas_notas = []
            indice_global_nota = 1
            for componente in cfg["componentes"]:
                ancho_columna = calcular_ancho_columna_planilla(
                    componente, componente["cantidad"]
                )
                for indice in range(1, componente["cantidad"] + 1):
                    nota_clave = f"{componente['prefijo']}{indice}".upper()
                    nombre_componente = str(
                        componente.get("nombre") or componente.get("clave") or "Nota"
                    ).strip()
                    titulo = f"{nombre_componente} {indice}"
                    encabezado_corto = f"N{indice_global_nota}"
                    indice_global_nota += 1
                    aliases = [nota_clave, titulo.upper(), encabezado_corto.upper()]
                    if componente["clave"] == "cognitivo" and cfg["solo_cognitivo"]:
                        aliases.append(f"N{indice}")
                    self.planilla_calif_columnas_notas.append(
                        {
                            "titulo": titulo,
                            "encabezado_corto": encabezado_corto,
                            "columna": f"{componente['clave']}_{indice}",
                            "componente": componente["clave"],
                            "ancho": ancho_columna,
                            "nota_clave": nota_clave,
                            "aliases": aliases,
                        }
                    )
            columnas = (
                self.planilla_calif_columnas_fijas
                + tuple(meta["columna"] for meta in self.planilla_calif_columnas_notas)
                + (self.planilla_calif_columna_promedio,)
            )
            self.tree_planilla.configure(columns=columnas)
            self.tree_planilla.configure(displaycolumns=columnas)
            self.tree_planilla.column("#0", width=0, minwidth=0, stretch=False)
            self.tree_planilla.column(
                "codigo", anchor="center", width=95, stretch=False
            )
            self.tree_planilla.column("nombre", anchor="w", width=280, stretch=True)
            for meta in self.planilla_calif_columnas_notas:
                self.tree_planilla.column(
                    meta["columna"],
                    anchor="center",
                    width=meta.get("ancho", 62),
                    stretch=False,
                )
            self.tree_planilla.column(
                self.planilla_calif_columna_promedio,
                anchor="center",
                width=82,
                stretch=False,
            )
            dibujar_encabezado_planilla_calif()

        def actualizar_resumen_planilla():
            total = len(self.tree_planilla.get_children())
            cfg = obtener_config_planilla_calif()
            detalle_componentes = " | ".join(
                f"{componente['nombre']}:{componente['cantidad']}"
                for componente in cfg["componentes"]
            )
            sufijo = f" | {detalle_componentes}"
            self.lbl_total_planilla.config(
                text=(
                    "0 estudiantes cargados" + sufijo
                    if total == 0
                    else (
                        "1 estudiante cargado" + sufijo
                        if total == 1
                        else f"{total} estudiantes cargados{sufijo}"
                    )
                )
            )

        def obtener_columnas_editables_planilla():
            return [meta["columna"] for meta in self.planilla_calif_columnas_notas]

        def resolver_valor_nota_planilla(meta, notas_estudiante):
            for alias in meta.get("aliases", []):
                clave = str(alias or "").strip().upper()
                if clave and clave in notas_estudiante:
                    return formatear_nota_planilla(notas_estudiante.get(clave, ""))
            return ""

        def calcular_definitiva_planilla_desde_valores(valores_notas):
            cfg = obtener_config_planilla_calif()
            definitivo = 0.0
            hay_notas = False
            for componente in cfg["componentes"]:
                notas_componente = []
                for meta in self.planilla_calif_columnas_notas:
                    if meta["componente"] != componente["clave"]:
                        continue
                    valor = (
                        str(valores_notas.get(meta["columna"], "") or "")
                        .strip()
                        .replace(",", ".")
                    )
                    if not valor:
                        continue
                    try:
                        notas_componente.append(float(valor))
                    except Exception:
                        continue
                if notas_componente:
                    hay_notas = True
                    promedio_componente = sum(notas_componente) / len(notas_componente)
                    definitivo += (
                        promedio_componente * float(componente["porcentaje"]) / 100.0
                    )
            if not hay_notas:
                return ""
            return f"{definitivo:.{cfg['decimales']}f}"

        def construir_fila_planilla_estudiante(estudiante, notas_estudiante):
            codigo = str(
                estudiante.get("codigo") or estudiante.get("documento") or ""
            ).strip()
            nombre = str(estudiante.get("nombre") or "").strip() or construir_nombre(
                estudiante
            )
            valores_notas = {}
            for meta in self.planilla_calif_columnas_notas:
                valores_notas[meta["columna"]] = resolver_valor_nota_planilla(
                    meta, notas_estudiante
                )
            definitiva = calcular_definitiva_planilla_desde_valores(valores_notas)
            valores_por_columna = {
                "codigo": codigo,
                "nombre": nombre,
                self.planilla_calif_columna_promedio: definitiva,
            }
            valores_por_columna.update(valores_notas)
            return tuple(
                valores_por_columna.get(columna, "")
                for columna in obtener_columnas_visibles_planilla()
            )

        def actualizar_promedio_fila_planilla(item):
            valores_notas = {
                meta["columna"]: self.tree_planilla.set(item, meta["columna"])
                for meta in self.planilla_calif_columnas_notas
            }
            self.tree_planilla.set(
                item,
                self.planilla_calif_columna_promedio,
                calcular_definitiva_planilla_desde_valores(valores_notas),
            )

        def seleccionar_celda_planilla(item, columna):
            if not item or columna not in obtener_columnas_editables_planilla():
                return False
            self.planilla_calif_celda_actual = {"item": item, "columna": columna}
            self.tree_planilla.focus(item)
            self.tree_planilla.selection_set(item)
            self.tree_planilla.see(item)
            self.tree_planilla.focus_set()
            return True

        def obtener_siguiente_celda_planilla(direccion_fila=0, direccion_columna=0):
            items = list(self.tree_planilla.get_children())
            columnas = obtener_columnas_editables_planilla()
            if not items or not columnas:
                return None, None
            item_actual = self.planilla_calif_celda_actual.get("item")
            columna_actual = self.planilla_calif_celda_actual.get("columna")
            if item_actual not in items:
                item_actual = items[0]
            if columna_actual not in columnas:
                columna_actual = columnas[0]
            indice_fila = max(
                0, min(len(items) - 1, items.index(item_actual) + direccion_fila)
            )
            indice_columna = max(
                0,
                min(
                    len(columnas) - 1,
                    columnas.index(columna_actual) + direccion_columna,
                ),
            )
            return items[indice_fila], columnas[indice_columna]

        def limpiar_tree_planilla():
            self.planilla_calif_celda_actual = {"item": None, "columna": None}
            for item in self.tree_planilla.get_children():
                self.tree_planilla.delete(item)
            actualizar_resumen_planilla()

        def normalizar_nota_ingresada(valor):
            texto = str(valor or "").strip().replace(",", ".")
            if not texto:
                return ""
            cfg = obtener_config_planilla_calif()
            try:
                nota = float(texto)
            except Exception:
                raise ValueError("La nota debe ser numérica.")
            if nota < cfg["nota_min"] or nota > cfg["nota_max"]:
                raise ValueError(
                    f"La nota debe estar entre {cfg['nota_min']:.{cfg['decimales']}f} y {cfg['nota_max']:.{cfg['decimales']}f}."
                )
            return f"{nota:.{cfg['decimales']}f}"

        def editar_nota_planilla(event=None, item=None, columna=None, valor_inicial=""):
            item = item or self.planilla_calif_celda_actual.get("item")
            columna = columna or self.planilla_calif_celda_actual.get("columna")
            if not item or columna not in obtener_columnas_editables_planilla():
                return "break"
            bbox = self.tree_planilla.bbox(item, columna)
            if not bbox:
                return "break"
            x, y, width, height = bbox
            top = tk.Toplevel(self.tab_planillas_calificaciones)
            top.overrideredirect(True)
            top.geometry(
                f"{width}x{height}+{self.tree_planilla.winfo_rootx() + x}+{self.tree_planilla.winfo_rooty() + y}"
            )
            var_nota = tk.StringVar(
                value=(
                    valor_inicial
                    if valor_inicial not in (None, "")
                    else self.tree_planilla.set(item, columna)
                )
            )
            entry = ttk.Entry(top, textvariable=var_nota, justify="center")
            entry.pack(fill="both", expand=True)
            entry.focus_set()
            entry.select_range(0, "end")

            def cerrar_editor(mover_columna=0, guardar=True):
                if guardar:
                    try:
                        nota_normalizada = normalizar_nota_ingresada(var_nota.get())
                    except ValueError as exc:
                        messagebox.showerror("Nota inválida", str(exc))
                        entry.focus_set()
                        entry.select_range(0, "end")
                        return "break"
                    self.tree_planilla.set(item, columna, nota_normalizada)
                    actualizar_promedio_fila_planilla(item)
                top.destroy()
                if mover_columna:
                    item_sig, col_sig = obtener_siguiente_celda_planilla(
                        direccion_columna=mover_columna
                    )
                    seleccionar_celda_planilla(item_sig, col_sig)
                return "break"

            entry.bind("<Return>", lambda e: cerrar_editor(mover_columna=1))
            entry.bind("<KP_Enter>", lambda e: cerrar_editor(mover_columna=1))
            entry.bind("<Escape>", lambda e: top.destroy())
            entry.bind("<FocusOut>", lambda e: cerrar_editor(mover_columna=0))
            return "break"

        def manejar_click_planilla(event):
            item = self.tree_planilla.identify_row(event.y)
            col = self.tree_planilla.identify_column(event.x)
            if not item or not col.startswith("#"):
                return
            try:
                indice = int(col[1:]) - 1
            except Exception:
                return
            columnas = self.tree_planilla.cget("displaycolumns")
            if columnas in (None, "", "#all"):
                columnas = self.tree_planilla["columns"]
            if indice < 0 or indice >= len(columnas):
                return
            seleccionar_celda_planilla(item, columnas[indice])

        def mover_celda_planilla(direccion_fila=0, direccion_columna=0):
            item, columna = obtener_siguiente_celda_planilla(
                direccion_fila=direccion_fila,
                direccion_columna=direccion_columna,
            )
            seleccionar_celda_planilla(item, columna)
            return "break"

        def manejar_tecla_planilla(event):
            tecla = str(event.keysym or "").upper()
            char = str(event.char or "")
            if tecla in {"LEFT", "RIGHT", "UP", "DOWN"}:
                movimientos = {
                    "LEFT": (0, -1),
                    "RIGHT": (0, 1),
                    "UP": (-1, 0),
                    "DOWN": (1, 0),
                }
                return mover_celda_planilla(*movimientos[tecla])
            if tecla in {"RETURN", "KP_ENTER", "SPACE"}:
                return editar_nota_planilla()
            if tecla in {"DELETE", "BACKSPACE"}:
                item = self.planilla_calif_celda_actual.get("item")
                columna = self.planilla_calif_celda_actual.get("columna")
                if item and columna:
                    self.tree_planilla.set(item, columna, "")
                    actualizar_promedio_fila_planilla(item)
                return "break"
            if char and char in "0123456789.,":
                return editar_nota_planilla(valor_inicial=char)

        def cargar_filtros_planilla_calif():
            try:
                from core import matricula as core_matricula

                grados = core_matricula.listar_grados_distintos(solo_activos=True)
                self.combo_grado_planilla["values"] = [
                    str(g).strip() for g in grados if str(g).strip()
                ]
            except Exception:
                self.combo_grado_planilla["values"] = []

        def cargar_cursos_planilla_calif(event=None):
            grado_sel = self.combo_grado_planilla.get().strip()
            if not grado_sel:
                self.combo_curso_planilla["values"] = []
                self.combo_curso_planilla.set("")
                self.combo_area_planilla["values"] = []
                self.combo_area_planilla.set("")
                limpiar_tree_planilla()
                return
            try:
                cursos = cargar_cursos_por_grado(grado_sel, db_path=DB_FILE)
            except Exception:
                cursos = []
            self.combo_curso_planilla["values"] = cursos
            self.combo_curso_planilla.set("")
            self.combo_area_planilla["values"] = []
            self.combo_area_planilla.set("")
            self.planilla_calif_areas = {}
            limpiar_tree_planilla()

        def obtener_area_info_planilla_calif(grado_sel, curso_sel, area_sel=None):
            area_buscada = str(area_sel or "").strip().casefold()
            try:
                grado_db = self._plan_normalizar_grado(grado_sel)
                curso_db = self._plan_normalizar_curso(curso_sel)
                conn = sqlite3.connect(self.db_path)
                cur = conn.cursor()
                cur.execute(
                    """
                    SELECT DISTINCT COALESCE(a.nombre, p.area) AS area_nombre, p.IdArea
                    FROM plan_estudio p
                    LEFT JOIN areas a ON a.id = p.IdArea
                    WHERE p.grado = ? AND p.curso = ? AND p.estado = 1
                      AND TRIM(COALESCE(a.nombre, p.area, '')) <> ''
                    ORDER BY COALESCE(a.nombre, p.area)
                    """,
                    (grado_db, curso_db),
                )
                rows = cur.fetchall()
                conn.close()
            except Exception:
                rows = []

            info = {}
            for area_nombre, area_id in rows:
                nombre = str(area_nombre or "").strip()
                if not nombre:
                    continue
                if nombre not in info:
                    info[nombre] = {"area": nombre, "IdArea": area_id}

            if not info:
                try:
                    for nombre in self.cargar_areas_plan_estudio(grado_sel, curso_sel):
                        nombre_txt = str(nombre or "").strip()
                        if nombre_txt and nombre_txt not in info:
                            info[nombre_txt] = {"area": nombre_txt, "IdArea": None}
                except Exception:
                    pass

            if area_buscada:
                for nombre, meta in info.items():
                    if nombre.strip().casefold() == area_buscada:
                        return meta
            return info

        def cargar_areas_planilla_calif(event=None):
            grado_sel = self.combo_grado_planilla.get().strip()
            curso_sel = self.combo_curso_planilla.get().strip()
            if not grado_sel or not curso_sel:
                self.combo_area_planilla["values"] = []
                self.combo_area_planilla.set("")
                self.planilla_calif_areas = {}
                limpiar_tree_planilla()
                return
            area_previa = self.combo_area_planilla.get().strip()
            try:
                self.planilla_calif_areas = obtener_area_info_planilla_calif(
                    grado_sel, curso_sel
                )
                areas = list(self.planilla_calif_areas.keys())
                self.combo_area_planilla["values"] = areas
                if area_previa in areas:
                    self.combo_area_planilla.set(area_previa)
                else:
                    self.combo_area_planilla.set(areas[0] if areas else "")
            except Exception:
                self.combo_area_planilla["values"] = []
                self.combo_area_planilla.set("")
                self.planilla_calif_areas = {}
            limpiar_tree_planilla()

        def cargar_estudiantes_planilla_calif(mostrar_advertencia=False):
            limpiar_tree_planilla()
            grado_sel = self.combo_grado_planilla.get().strip()
            curso_sel = self.combo_curso_planilla.get().strip()
            area_sel = self.combo_area_planilla.get().strip()
            periodo_sel = self.combo_periodo_planilla.get().strip()
            if not (grado_sel and curso_sel and area_sel and periodo_sel):
                if mostrar_advertencia:
                    messagebox.showwarning(
                        "Faltan filtros",
                        "Selecciona grado, curso, área y periodo para cargar la planilla.",
                    )
                return
            try:
                configurar_columnas_planilla_calif()
                estudiantes = cargar_estudiantes_por_grado(
                    grado_sel, curso_sel, db_path=DB_FILE
                )
                area_info = obtener_area_info_planilla_calif(
                    grado_sel, curso_sel, area_sel
                )
                area_id = area_info.get("IdArea")
                notas_registradas = {}
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute(
                    """
                    SELECT estudiante_id, nota_clave, nota
                    FROM planillas_calificaciones
                    WHERE area_id = ? AND periodo = ? AND anio_lectivo = ?
                    ORDER BY nota_clave
                    """,
                    (area_id, periodo_sel, str(obtener_anio_planilla_calif())),
                )
                for estudiante_id, nota_clave, nota in cursor.fetchall():
                    clave = str(nota_clave or "N1").strip().upper() or "N1"
                    notas_registradas.setdefault(str(estudiante_id), {})[clave] = nota
                conn.close()
                for indice, estudiante in enumerate(estudiantes, 1):
                    estudiante_id = str(
                        estudiante.get("id") or estudiante.get("documento") or ""
                    ).strip()
                    notas_estudiante = notas_registradas.get(estudiante_id, {})
                    fila = construir_fila_planilla_estudiante(
                        estudiante, notas_estudiante
                    )
                    item = self.tree_planilla.insert(
                        "",
                        "end",
                        values=fila,
                        tags=(estudiante_id,),
                    )
                if (
                    self.tree_planilla.get_children()
                    and self.planilla_calif_columnas_notas
                ):
                    seleccionar_celda_planilla(
                        self.tree_planilla.get_children()[0],
                        self.planilla_calif_columnas_notas[0]["columna"],
                    )
                dibujar_encabezado_planilla_calif()
                actualizar_resumen_planilla()
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo cargar la planilla: {e}")

        def intentar_carga_planilla_calif(_event=None):
            cargar_estudiantes_planilla_calif(mostrar_advertencia=False)

        def al_cambiar_curso_planilla_calif(event=None):
            cargar_areas_planilla_calif(event)
            intentar_carga_planilla_calif(event)

        def guardar_planilla_calificaciones():
            if not self.tree_planilla.get_children():
                messagebox.showwarning("Sin datos", "No hay estudiantes cargados.")
                return
            area_sel = self.combo_area_planilla.get().strip()
            periodo_sel = self.combo_periodo_planilla.get().strip()
            if not area_sel or not periodo_sel:
                messagebox.showwarning(
                    "Filtros requeridos",
                    "Selecciona el área y el periodo antes de guardar.",
                )
                return
            if not messagebox.askyesno(
                "Confirmar",
                f"¿Guardar la planilla de calificaciones de {area_sel} en {periodo_sel}?",
            ):
                return
            area_info = obtener_area_info_planilla_calif(
                self.combo_grado_planilla.get().strip(),
                self.combo_curso_planilla.get().strip(),
                area_sel,
            )
            area_id = area_info.get("IdArea")
            anio_lectivo = str(obtener_anio_planilla_calif())
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                for item in self.tree_planilla.get_children():
                    tags = self.tree_planilla.item(item, "tags")
                    estudiante_id = str(tags[0]) if tags else ""
                    for meta in self.planilla_calif_columnas_notas:
                        nota = str(
                            self.tree_planilla.set(item, meta["columna"]) or ""
                        ).strip()
                        if nota:
                            cursor.execute(
                                """
                                INSERT INTO planillas_calificaciones (
                                    estudiante_id, area_id, asignatura_id, periodo, nota,
                                    observacion, nota_clave, anio_lectivo
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                                ON CONFLICT(estudiante_id, area_id, periodo, anio_lectivo, nota_clave)
                                DO UPDATE SET nota=excluded.nota, observacion=excluded.observacion
                                """,
                                (
                                    estudiante_id,
                                    area_id,
                                    None,
                                    periodo_sel,
                                    float(nota.replace(",", ".")),
                                    "",
                                    meta["nota_clave"],
                                    anio_lectivo,
                                ),
                            )
                        else:
                            cursor.execute(
                                """
                                DELETE FROM planillas_calificaciones
                                WHERE estudiante_id = ? AND area_id = ? AND periodo = ?
                                  AND anio_lectivo = ? AND nota_clave = ?
                                """,
                                (
                                    estudiante_id,
                                    area_id,
                                    periodo_sel,
                                    anio_lectivo,
                                    meta["nota_clave"],
                                ),
                            )
                conn.commit()
                conn.close()
                messagebox.showinfo(
                    "Éxito", "Planilla de calificaciones guardada correctamente."
                )
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar la planilla: {e}")

        def exportar_planilla_calificaciones():
            if not self.tree_planilla.get_children():
                messagebox.showwarning("Sin datos", "No hay estudiantes cargados.")
                return
            try:
                from openpyxl import Workbook
                from openpyxl.drawing.image import Image as XLImage
                from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
                from openpyxl.utils import get_column_letter
                from tkinter import filedialog

                grado = self.combo_grado_planilla.get().strip() or "-"
                curso = self.combo_curso_planilla.get().strip() or "-"
                area = self.combo_area_planilla.get().strip() or "-"
                periodo = self.combo_periodo_planilla.get().strip() or "-"
                anio_lectivo = obtener_anio_planilla_calif()
                institucion = (
                    self._get_config_plantel("nombre_institucion")
                    or "INSTITUCIÓN EDUCATIVA"
                ).upper()
                resolucion = self._get_config_plantel("resolucion_aprobacion") or ""
                decreto = self._get_config_plantel("decreto_funcionamiento") or ""
                dane = self._get_config_plantel("codigo_dane") or ""
                nit = self._get_config_plantel("nit") or ""
                logo_path = self._get_config_plantel("logo_path") or ""
                sugerido = f"planilla_calificaciones_{grado}_{curso}_{area}_{periodo}_{anio_lectivo}.xlsx".replace(
                    " ", "_"
                ).lower()
                ruta = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel", "*.xlsx")],
                    initialfile=sugerido,
                )
                if not ruta:
                    return

                wb = Workbook()
                ws = wb.active
                ws.title = "Calificaciones"
                thin = Side(style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                center = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                left = Alignment(horizontal="left", vertical="center", wrap_text=True)
                header_fill = PatternFill(fill_type="solid", fgColor="356A9A")
                subheader_fill = PatternFill(fill_type="solid", fgColor="F2F4F7")
                header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                subheader_font = Font(name="Arial", size=9, bold=True)

                total_columnas = 2 + len(self.planilla_calif_columnas_notas) + 1
                ultima_col = get_column_letter(total_columnas)
                ws.merge_cells(f"B1:{ultima_col}1")
                ws["B1"] = institucion
                ws["B1"].font = Font(name="Arial", size=14, bold=True)
                ws["B1"].alignment = center
                ws.merge_cells(f"B2:{ultima_col}2")
                ws["B2"] = (
                    f"Resolución de aprobación: {resolucion}    Decreto de funcionamiento: {decreto}"
                ).strip()
                ws["B2"].font = Font(name="Arial", size=8, bold=False)
                ws["B2"].alignment = center
                ws.merge_cells(f"B3:{ultima_col}3")
                ws["B3"] = f"DANE: {dane}    NIT: {nit}".strip()
                ws["B3"].font = Font(name="Arial", size=9, bold=True)
                ws["B3"].alignment = center
                ws.merge_cells(f"A4:{ultima_col}4")
                ws["A4"].border = Border(bottom=thin)
                if logo_path and os.path.exists(logo_path):
                    try:
                        img = XLImage(logo_path)
                        img.width = 55
                        img.height = 55
                        ws.add_image(img, "A1")
                    except Exception:
                        pass

                info_row = 5
                bloques_info = [
                    f"Grado: {grado}",
                    f"Curso: {curso}",
                    f"Área: {area}",
                    f"Periodo: {periodo} | Año Lectivo: {anio_lectivo}",
                ]
                bloques = []
                inicio_actual = 1
                restante = total_columnas
                for indice, valor in enumerate(bloques_info, start=1):
                    pendientes = len(bloques_info) - indice + 1
                    if indice == len(bloques_info):
                        fin_actual = total_columnas
                    else:
                        ancho = max(1, restante // pendientes)
                        fin_actual = min(total_columnas, inicio_actual + ancho - 1)
                    bloques.append((inicio_actual, fin_actual, valor))
                    usado = fin_actual - inicio_actual + 1
                    restante -= usado
                    inicio_actual = fin_actual + 1

                for inicio, fin, valor in bloques:
                    ws.merge_cells(
                        start_row=info_row,
                        start_column=inicio,
                        end_row=info_row,
                        end_column=fin,
                    )
                    celda = ws.cell(info_row, inicio)
                    celda.value = valor
                    celda.font = Font(name="Arial", size=9, bold=True)
                    celda.alignment = left if "Periodo:" in valor else center
                    for col in range(inicio, fin + 1):
                        ws.cell(info_row, col).border = border

                header_row_top = 7
                header_row_mid = 8
                header_row_bottom = 9
                data_row = 10

                ws.merge_cells(
                    start_row=header_row_top,
                    start_column=1,
                    end_row=header_row_bottom,
                    end_column=1,
                )
                ws.cell(header_row_top, 1).value = "Cód."

                ws.merge_cells(
                    start_row=header_row_top,
                    start_column=2,
                    end_row=header_row_bottom,
                    end_column=2,
                )
                ws.cell(header_row_top, 2).value = "Apellidos/Nombres"

                notas_inicio = 3
                notas_fin = notas_inicio + len(self.planilla_calif_columnas_notas) - 1
                def_col = notas_fin + 1

                if self.planilla_calif_columnas_notas:
                    ws.merge_cells(
                        start_row=header_row_top,
                        start_column=notas_inicio,
                        end_row=header_row_top,
                        end_column=notas_fin,
                    )
                    ws.cell(header_row_top, notas_inicio).value = "Desempeño Evaluado"

                    cfg_export = obtener_config_planilla_calif()
                    metas_por_componente = {}
                    for meta in self.planilla_calif_columnas_notas:
                        metas_por_componente.setdefault(meta["componente"], []).append(
                            meta
                        )

                    columna_por_meta = {
                        meta["columna"]: notas_inicio + indice
                        for indice, meta in enumerate(
                            self.planilla_calif_columnas_notas
                        )
                    }

                    for componente in cfg_export["componentes"]:
                        metas = metas_por_componente.get(componente["clave"], [])
                        if not metas:
                            continue
                        inicio_comp = columna_por_meta[metas[0]["columna"]]
                        fin_comp = columna_por_meta[metas[-1]["columna"]]
                        ws.merge_cells(
                            start_row=header_row_mid,
                            start_column=inicio_comp,
                            end_row=header_row_mid,
                            end_column=fin_comp,
                        )
                        ws.cell(header_row_mid, inicio_comp).value = (
                            f"{componente['nombre']}\n"
                            f"[{formatear_porcentaje_encabezado_planilla(componente['porcentaje'])}]"
                        )

                    for indice, meta in enumerate(
                        self.planilla_calif_columnas_notas, start=notas_inicio
                    ):
                        ws.cell(header_row_bottom, indice).value = meta[
                            "encabezado_corto"
                        ]

                ws.merge_cells(
                    start_row=header_row_top,
                    start_column=def_col,
                    end_row=header_row_mid,
                    end_column=def_col,
                )
                ws.cell(header_row_top, def_col).value = "Definitiva"
                ws.cell(header_row_bottom, def_col).value = "DEF"

                for fila in (header_row_top, header_row_mid):
                    for col in range(1, total_columnas + 1):
                        celda = ws.cell(fila, col)
                        celda.fill = header_fill
                        celda.font = header_font
                        celda.alignment = center
                        celda.border = border

                for col in range(1, total_columnas + 1):
                    celda = ws.cell(header_row_bottom, col)
                    if col in (1, 2):
                        celda.fill = header_fill
                        celda.font = header_font
                    else:
                        celda.fill = subheader_fill
                        celda.font = subheader_font
                    celda.alignment = center
                    celda.border = border

                for offset, item in enumerate(
                    self.tree_planilla.get_children(), start=0
                ):
                    row = data_row + offset
                    ws.cell(row, 1).value = self.tree_planilla.set(item, "codigo")
                    ws.cell(row, 2).value = self.tree_planilla.set(item, "nombre")
                    col_excel = 3
                    for meta in self.planilla_calif_columnas_notas:
                        ws.cell(row, col_excel).value = self.tree_planilla.set(
                            item, meta["columna"]
                        )
                        col_excel += 1
                    ws.cell(row, col_excel).value = self.tree_planilla.set(
                        item, self.planilla_calif_columna_promedio
                    )
                    for col in range(1, total_columnas + 1):
                        celda = ws.cell(row, col)
                        celda.border = border
                        celda.alignment = left if col == 2 else center
                        celda.font = Font(name="Arial", size=8)

                ws.column_dimensions["A"].width = 12
                ws.column_dimensions["B"].width = 38
                for col in range(3, total_columnas + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 9
                ws.freeze_panes = "C10"
                ws.sheet_view.showGridLines = False
                ws.page_setup.orientation = "landscape"
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                wb.save(ruta)
                messagebox.showinfo("Éxito", f"Planilla exportada a {ruta}")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo exportar la planilla: {e}")

        cargar_filtros_planilla_calif()
        configurar_columnas_planilla_calif()
        self.combo_grado_planilla.bind(
            "<<ComboboxSelected>>", cargar_cursos_planilla_calif
        )
        self.combo_curso_planilla.bind(
            "<<ComboboxSelected>>", al_cambiar_curso_planilla_calif
        )
        self.combo_area_planilla.bind(
            "<<ComboboxSelected>>", intentar_carga_planilla_calif
        )
        self.combo_periodo_planilla.bind(
            "<<ComboboxSelected>>", intentar_carga_planilla_calif
        )
        self.tree_planilla.bind("<Button-1>", manejar_click_planilla, add="+")
        self.tree_planilla.bind("<Double-1>", editar_nota_planilla)
        self.tree_planilla.bind("<Key>", manejar_tecla_planilla)
        self.btn_guardar_planilla_calif.config(command=guardar_planilla_calificaciones)
        self.btn_exportar_planilla_calif.config(
            command=exportar_planilla_calificaciones
        )
        actualizar_resumen_planilla()

        # Pestañas principales
        self.tab_docentes = ttk.Frame(self.nb)
        self.tab_carga_academica = ttk.Frame(self.nb)
        self.tab_estudiantes = ttk.Frame(self.nb)
        self.tab_preguntas = ttk.Frame(self.nb)
        self.tab_examenes = ttk.Frame(self.nb)
        self.tab_plan_estudio = ttk.Frame(self.nb)
        self.tab_maestro = ttk.Frame(self.nb)
        # --- Nueva pestaña principal: Académico ---
        self.tab_academico = ttk.Frame(self.nb)
        self.nb_academico = ttk.Notebook(self.tab_academico)
        self.nb_academico.pack(fill="both", expand=True)
        # Subpestañas de Académico
        self.tab_desempenos = ttk.Frame(self.nb_academico)
        self.tab_plan_aula = ttk.Frame(self.nb_academico)
        self.tab_planeador_clase = ttk.Frame(self.nb_academico)
        self.nb_academico.add(self.tab_desempenos, text="Desempeños")
        self.nb_academico.add(self.tab_plan_aula, text="Plan de aula")
        self.nb_academico.add(self.tab_planeador_clase, text="Planeador de clase")
        # Crear la pestaña principal de Configuración
        self.tab_configuracion = ttk.Frame(self.nb)
        # Subnotebook para las subpestañas de configuración
        self.nb_configuracion = ttk.Notebook(self.tab_configuracion)
        self.nb_configuracion.pack(fill="both", expand=True)
        # Subpestañas
        self.tab_configuracion_plantel = ttk.Frame(self.nb_configuracion)
        self.tab_contenido_institucional = ttk.Frame(self.nb_configuracion)
        self.tab_calendario_academico = ttk.Frame(self.nb_configuracion)
        # Integrar CalendarioAcadémico en la subpestaña
        try:
            from app import CalendarioAcademico

            CalendarioAcademico(
                self.tab_calendario_academico,
                anio_lectivo=self._get_config_plantel("anio_lectivo"),
            )
        except Exception as e:
            lbl_error = tk.Label(
                self.tab_calendario_academico,
                text=f"No se pudo cargar el Calendario Académico:\n{e}",
                fg="red",
                font=("Segoe UI", 11, "bold"),
                justify="center",
                wraplength=600,
            )
            lbl_error.pack(padx=30, pady=30)
        self.tab_seguridad = ttk.Frame(self.nb_configuracion)
        # Certificados
        try:
            from certificados.modulo_certificados import ModuloCertificados

            self.tab_certificados = ttk.Frame(self.nb)
            # Subnotebook para subpestañas de certificados
            self.nb_certificados = ttk.Notebook(self.tab_certificados)
            self.nb_certificados.pack(fill="both", expand=True)

            # Subpestañas
            self.tab_cert_matricula = ttk.Frame(self.nb_certificados)
            self.tab_cert_calificaciones = ttk.Frame(self.nb_certificados)
            self.tab_acta = ttk.Frame(self.nb_certificados)
            self.tab_diploma = ttk.Frame(self.nb_certificados)

            # Instanciar la interfaz de certificados en cada subpestaña (idéntica)
            ModuloCertificados(self.tab_cert_matricula, self.db_path)
            ModuloCertificados(self.tab_cert_calificaciones, self.db_path)
            ModuloCertificados(self.tab_acta, self.db_path)
            ModuloCertificados(self.tab_diploma, self.db_path)

            self.nb_certificados.add(
                self.tab_cert_matricula, text="Certificado de Matrícula"
            )
            self.nb_certificados.add(
                self.tab_cert_calificaciones, text="Certificado de Calificaciones"
            )
            self.nb_certificados.add(self.tab_acta, text="Acta de grado")
            self.nb_certificados.add(self.tab_diploma, text="Diploma")

            certificados_ok = True
        except Exception as e:
            print(f"[ERROR] No se pudo cargar la pestaña Certificados: {e}")
            self.tab_certificados = None
            certificados_ok = False

        # Agregar pestañas en el orden correcto solicitado
        self.nb_configuracion.add(
            self.tab_configuracion_plantel, text="Configuración Plantel"
        )
        self.nb_configuracion.add(
            self.tab_contenido_institucional, text="Contenido Institucional"
        )
        self.nb_configuracion.add(
            self.tab_calendario_academico, text="Calendario Académico"
        )
        self.nb_configuracion.add(self.tab_seguridad, text="Seguridad")
        # Agregar pestañas principales en el orden solicitado
        self._agregar_pestana_principal_si_permiso(self.tab_academico, "Académico")
        self._agregar_pestana_principal_si_permiso(
            self.tab_configuracion, "Configuración"
        )
        self._agregar_pestana_principal_si_permiso(
            self.tab_docentes, "Planta de Personal"
        )
        self._agregar_pestana_principal_si_permiso(
            self.tab_carga_academica, "Carga Académica"
        )
        self._agregar_pestana_principal_si_permiso(
            self.tab_plan_estudio, "Plan de Estudios"
        )
        self._agregar_pestana_principal_si_permiso(self.tab_estudiantes, "Matrícula")
        self._agregar_pestana_principal_si_permiso(
            self.tab_preguntas, "Banco Preguntas"
        )
        self._agregar_pestana_principal_si_permiso(self.tab_examenes, "Evaluaciones")
        self._agregar_pestana_principal_si_permiso(self.tab_planillas, "Planillas")
        if certificados_ok and self.tab_certificados is not None:
            self._agregar_pestana_principal_si_permiso(
                self.tab_certificados, "Certificados"
            )
        self._agregar_pestana_principal_si_permiso(
            self.tab_maestro, "Panel del Docente"
        )
        # Nueva pestaña: Estadística
        self.tab_estadistica = ttk.Frame(self.nb)
        self._agregar_pestana_principal_si_permiso(self.tab_estadistica, "Estadística")

        # Subpestaña Totalizar en Estadística
        self.nb_estadistica = ttk.Notebook(self.tab_estadistica)
        self.nb_estadistica.pack(fill="both", expand=True)
        self.tab_totalizar = ttk.Frame(self.nb_estadistica)
        self.nb_estadistica.add(self.tab_totalizar, text="Totalizar")
        self._build_totalizar_tab()

    def _build_totalizar_tab(self):
        frame = self.tab_totalizar
        for widget in frame.winfo_children():
            widget.destroy()
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(0, weight=1)
        columnas = ("sede", "jornada", "grado", "curso", "total")
        titulos = {
            "sede": "Sede",
            "jornada": "Jornada",
            "grado": "Grado",
            "curso": "Curso",
            "total": "Total",
        }
        tree = ttk.Treeview(frame, columns=columnas, show="headings")
        for col in columnas:
            tree.heading(col, text=titulos[col])
            tree.column(col, width=120, anchor="center")
        tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar_y.set)
        scrollbar_y.grid(row=0, column=1, sticky="ns")

        # Llenar la tabla con los totales agrupados
        try:
            import sqlite3

            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            cur.execute(
                """
                SELECT sede, jornada, grado, curso, COUNT(*) as total
                FROM estudiantes
                WHERE estado_academico IS NULL OR estado_academico = '' OR estado_academico = 'Activo'
                GROUP BY sede, jornada, grado, curso
                ORDER BY sede, jornada, grado, curso
            """
            )
            rows = cur.fetchall()
            total_estudiantes = 0
            for row in rows:
                sede, jornada, grado, curso, total = row
                total_estudiantes += total
                tree.insert("", "end", values=(sede, jornada, grado, curso, total))
            # Fila resumen total
            if rows:
                # Añadir una fila separadora visual
                tree.insert("", "end", values=("", "", "", "", ""))
                # Añadir la fila de total general
                tree.insert(
                    "",
                    "end",
                    values=("", "", "", "TOTAL ESTUDIANTES", total_estudiantes),
                )

                # Resumen por sede y jornada
                tree.insert("", "end", values=("", "", "", "", ""))
                tree.insert(
                    "",
                    "end",
                    values=(
                        "Total de Estudiantes Matriculados Por Sede y Jornada",
                        "",
                        "",
                        "",
                        "",
                    ),
                )
                tree.insert("", "end", values=("Sede", "Jornada", "", "Total", ""))
                cur.execute(
                    """
                    SELECT sede, jornada, COUNT(*) as total
                    FROM estudiantes
                    WHERE estado_academico IS NULL OR estado_academico = '' OR estado_academico = 'Activo'
                    GROUP BY sede, jornada
                    ORDER BY sede, jornada
                """
                )
                resumen = cur.fetchall()
                for sede, jornada, total in resumen:
                    tree.insert("", "end", values=(sede, jornada, "", total, ""))
                # Resumen por sede
                tree.insert("", "end", values=("", "", "", "", ""))
                tree.insert(
                    "",
                    "end",
                    values=(
                        "Total de Estudiantes Matriculados Por Sede",
                        "",
                        "",
                        "",
                        "",
                    ),
                )
                tree.insert(
                    "",
                    "end",
                    values=("Cod. Sede", "Sede", "Total", "Total Activos", "% (DE/RE)"),
                )
                cur.execute(
                    """
                    SELECT sede, COUNT(*) as total, SUM(CASE WHEN estado_academico IS NULL OR estado_academico = '' OR estado_academico = 'Activo' THEN 1 ELSE 0 END) as activos
                    FROM estudiantes
                    GROUP BY sede
                    ORDER BY sede
                """
                )
                resumen_sede = cur.fetchall()
                total_de = 0
                total_re = 0
                for idx, (sede, total, activos) in enumerate(resumen_sede, 1):
                    cod_sede = f"{idx:02d}"
                    porcentaje = f"{(activos/total*100):.1f}%" if total else ""
                    tree.insert(
                        "", "end", values=(cod_sede, sede, total, activos, porcentaje)
                    )
                    total_de += total
                    total_re += activos
                # Fila total por sede
                if resumen_sede:
                    tree.insert(
                        "",
                        "end",
                        values=(
                            "Total:",
                            "",
                            total_de,
                            total_re,
                            f"{(total_re/total_de*100):.1f}%" if total_de else "",
                        ),
                    )

                # Total institucional
                tree.insert("", "end", values=("", "", "", "", ""))
                tree.insert(
                    "",
                    "end",
                    values=("Total de Estudiantes Matriculados", "", "", "", ""),
                )
                tree.insert(
                    "", "end", values=("Institución Educativa", "Total", "", "", "")
                )
                cur.execute(
                    """
                    SELECT COUNT(*) FROM estudiantes
                """
                )
                total_inst = cur.fetchone()[0]
                tree.insert(
                    "", "end", values=("Toda la Institución", total_inst, "", "", "")
                )
            conn.close()
        except Exception as e:
            tree.insert(
                "", "end", values=("Error al cargar datos", str(e), "", "", "", "")
            )
        # Pestaña Boletines
        self.tab_boletines = ttk.Frame(self.nb_academico)
        self.nb_academico.add(self.tab_boletines, text="Boletines")

        # (La pestaña Desempeños ahora es subpestaña de Académico)

        self._build_docentes_tab()  # Considera renombrar a _build_personal_tab si refactorizas funciones
        self._build_carga_academica_tab()
        self._build_estudiantes_tab()
        self._build_preguntas_tab()
        self._build_examenes_tab()
        self._build_plan_estudio_tab()
        self._build_maestro_tab()
        self._build_boletines_tab()
        self._build_desempenos_tab()
        self._build_seguridad_tab(self.tab_seguridad)
        self._build_configuracion_plantel_tab(self.tab_configuracion_plantel)
        self._build_contenido_institucional_tab(self.tab_contenido_institucional)
        self._configurar_sincronizacion_preguntas()
        crear_footer(self.win)

    def _build_boletines_tab(self):
        frame = self.tab_boletines
        for widget in frame.winfo_children():
            widget.destroy()

        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(2, weight=1)

        encabezado = ttk.Frame(frame, padding=(12, 10, 12, 0))
        encabezado.grid(row=0, column=0, sticky="ew")
        encabezado.grid_columnconfigure(0, weight=1)
        ttk.Label(
            encabezado,
            text="Boletines y alistamiento académico",
            font=("Segoe UI", 14, "bold"),
        ).grid(row=0, column=0, sticky="w")
        ttk.Label(
            encabezado,
            text="Valida cobertura de notas por período, promedio general y disponibilidad de descriptores antes de emitir boletines.",
            style="CardHint.TLabel",
        ).grid(row=1, column=0, sticky="w", pady=(2, 0))

        filtros = ttk.LabelFrame(
            frame, text="Filtros", padding=10, style="Card.TLabelframe"
        )
        filtros.grid(row=1, column=0, sticky="ew", padx=12, pady=(10, 8))
        for columna in (1, 3, 5, 7):
            filtros.grid_columnconfigure(columna, weight=1)

        self.boletines_var_grado = tk.StringVar(value="Todos")
        self.boletines_var_curso = tk.StringVar(value="Todos")
        self.boletines_var_area = tk.StringVar(value="Todos")
        self.boletines_var_periodo = tk.StringVar()
        self.boletines_var_anio = tk.StringVar(value=self._obtener_anio_lectivo_base())

        ttk.Label(filtros, text="Grado:").grid(
            row=0, column=0, sticky="w", padx=(0, 6), pady=2
        )
        self.combo_boletines_grado = ttk.Combobox(
            filtros,
            textvariable=self.boletines_var_grado,
            state="readonly",
            width=18,
            values=self._listar_grados_catalogo(incluir_todos=True),
        )
        self.combo_boletines_grado.grid(row=0, column=1, sticky="ew", pady=2)
        self.combo_boletines_grado.bind(
            "<<ComboboxSelected>>", self._boletines_refrescar_cursos
        )

        ttk.Label(filtros, text="Curso:").grid(
            row=0, column=2, sticky="w", padx=(12, 6), pady=2
        )
        self.combo_boletines_curso = ttk.Combobox(
            filtros,
            textvariable=self.boletines_var_curso,
            state="readonly",
            width=12,
        )
        self.combo_boletines_curso.grid(row=0, column=3, sticky="ew", pady=2)
        self.combo_boletines_curso.bind(
            "<<ComboboxSelected>>", lambda _e: self._boletines_recargar_tablero()
        )

        ttk.Label(filtros, text="Área:").grid(
            row=0, column=4, sticky="w", padx=(12, 6), pady=2
        )
        self.combo_boletines_area = ttk.Combobox(
            filtros,
            textvariable=self.boletines_var_area,
            state="readonly",
            width=22,
            values=self._listar_areas_catalogo(incluir_todos=True),
        )
        self.combo_boletines_area.grid(row=0, column=5, sticky="ew", pady=2)
        self.combo_boletines_area.bind(
            "<<ComboboxSelected>>", lambda _e: self._boletines_recargar_tablero()
        )

        ttk.Label(filtros, text="Período:").grid(
            row=1, column=0, sticky="w", padx=(0, 6), pady=2
        )
        self.combo_boletines_periodo = ttk.Combobox(
            filtros,
            textvariable=self.boletines_var_periodo,
            state="readonly",
            width=14,
            values=self._listar_periodos_configurados(),
        )
        self.combo_boletines_periodo.grid(row=1, column=1, sticky="ew", pady=2)
        self.combo_boletines_periodo.bind(
            "<<ComboboxSelected>>", lambda _e: self._boletines_recargar_tablero()
        )

        ttk.Label(filtros, text="Año:").grid(
            row=1, column=2, sticky="w", padx=(12, 6), pady=2
        )
        self.entry_boletines_anio = ttk.Entry(
            filtros, textvariable=self.boletines_var_anio, width=10
        )
        self.entry_boletines_anio.grid(row=1, column=3, sticky="ew", pady=2)

        self._crear_boton_si_permiso(
            ttk.Button,
            filtros,
            "desktop.superadmin.boletines.actualizar",
            layout="grid",
            layout_kwargs={
                "row": 1,
                "column": 6,
                "sticky": "e",
                "padx": (12, 0),
                "pady": 2,
            },
            text="Actualizar panel",
            command=self._boletines_recargar_tablero,
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            filtros,
            "desktop.superadmin.boletines.exportar_pdf",
            layout="grid",
            layout_kwargs={
                "row": 1,
                "column": 7,
                "sticky": "e",
                "padx": (8, 0),
                "pady": 2,
            },
            text="Generar PDF",
            command=self._boletines_exportar_pdf_referencia,
        )

        tablero = ttk.Frame(frame, padding=(12, 0, 12, 12))
        tablero.grid(row=2, column=0, sticky="nsew")
        tablero.grid_columnconfigure(0, weight=3)
        tablero.grid_columnconfigure(1, weight=2)
        tablero.grid_rowconfigure(2, weight=1)

        estadisticas = ttk.Frame(tablero)
        estadisticas.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        for columna in range(6):
            estadisticas.grid_columnconfigure(columna, weight=1)
        self.boletines_stats_labels = {}
        for indice, titulo in enumerate(
            [
                "Estudiantes",
                "Con promedio",
                "Listos",
                "Pendientes",
                "Sin notas",
                "Promedio grupo",
            ]
        ):
            tarjeta = ttk.LabelFrame(
                estadisticas, text=titulo, padding=8, style="Card.TLabelframe"
            )
            tarjeta.grid(
                row=0, column=indice, sticky="ew", padx=(0 if indice == 0 else 6, 0)
            )
            valor = ttk.Label(tarjeta, text="0", font=("Segoe UI", 13, "bold"))
            valor.pack(anchor="w")
            self.boletines_stats_labels[titulo] = valor

        self.lbl_boletines_estado = ttk.Label(
            tablero,
            text="Seleccione filtros y cargue el panel para revisar el estado académico del período.",
            style="CardHint.TLabel",
        )
        self.lbl_boletines_estado.grid(
            row=1, column=0, columnspan=2, sticky="w", pady=(0, 6)
        )

        panel_estudiantes = ttk.LabelFrame(
            tablero, text="Resumen por estudiante", padding=8, style="Card.TLabelframe"
        )
        panel_estudiantes.grid(row=2, column=0, sticky="nsew", padx=(0, 8))
        panel_estudiantes.grid_columnconfigure(0, weight=1)
        panel_estudiantes.grid_rowconfigure(0, weight=1)

        columnas_boletines = (
            "documento",
            "estudiante",
            "grado",
            "curso",
            "areas",
            "promedio",
            "desempeno",
            "estado",
        )
        self.tree_boletines = ttk.Treeview(
            panel_estudiantes, columns=columnas_boletines, show="headings", height=18
        )
        anchos = {
            "documento": 110,
            "estudiante": 260,
            "grado": 80,
            "curso": 70,
            "areas": 90,
            "promedio": 90,
            "desempeno": 110,
            "estado": 120,
        }
        titulos = {
            "documento": "Documento",
            "estudiante": "Estudiante",
            "grado": "Grado",
            "curso": "Curso",
            "areas": "Áreas",
            "promedio": "Promedio",
            "desempeno": "Desempeño",
            "estado": "Estado",
        }
        for columna in columnas_boletines:
            self.tree_boletines.heading(columna, text=titulos[columna])
            self.tree_boletines.column(
                columna,
                width=anchos[columna],
                anchor="center" if columna != "estudiante" else "w",
            )
        self.tree_boletines.column("estudiante", anchor="w")
        self.tree_boletines.grid(row=0, column=0, sticky="nsew")
        scrollbar_boletines_y = ttk.Scrollbar(
            panel_estudiantes, orient="vertical", command=self.tree_boletines.yview
        )
        scrollbar_boletines_y.grid(row=0, column=1, sticky="ns")
        scrollbar_boletines_x = ttk.Scrollbar(
            panel_estudiantes, orient="horizontal", command=self.tree_boletines.xview
        )
        scrollbar_boletines_x.grid(row=1, column=0, sticky="ew")
        self.tree_boletines.configure(
            yscrollcommand=scrollbar_boletines_y.set,
            xscrollcommand=scrollbar_boletines_x.set,
        )
        self.tree_boletines.bind(
            "<<TreeviewSelect>>", self._boletines_mostrar_detalle_estudiante
        )

        panel_detalle = ttk.LabelFrame(
            tablero, text="Detalle por áreas", padding=8, style="Card.TLabelframe"
        )
        panel_detalle.grid(row=2, column=1, sticky="nsew")
        panel_detalle.grid_columnconfigure(0, weight=1)
        panel_detalle.grid_rowconfigure(1, weight=1)

        self.lbl_boletines_detalle = ttk.Label(
            panel_detalle,
            text="Seleccione un estudiante para revisar promedios por área y descriptor sugerido.",
            style="CardHint.TLabel",
            wraplength=380,
            justify="left",
        )
        self.lbl_boletines_detalle.grid(row=0, column=0, sticky="w", pady=(0, 8))

        columnas_detalle = ("area", "promedio", "desempeno", "registros", "descriptor")
        self.tree_boletines_detalle = ttk.Treeview(
            panel_detalle, columns=columnas_detalle, show="headings", height=18
        )
        titulos_detalle = {
            "area": "Área",
            "promedio": "Promedio",
            "desempeno": "Desempeño",
            "registros": "Registros",
            "descriptor": "Descriptor disponible",
        }
        anchos_detalle = {
            "area": 140,
            "promedio": 80,
            "desempeno": 100,
            "registros": 80,
            "descriptor": 260,
        }
        for columna in columnas_detalle:
            self.tree_boletines_detalle.heading(columna, text=titulos_detalle[columna])
            self.tree_boletines_detalle.column(
                columna,
                width=anchos_detalle[columna],
                anchor="w" if columna in ("area", "descriptor") else "center",
            )
        self.tree_boletines_detalle.grid(row=1, column=0, sticky="nsew")
        scrollbar_detalle_y = ttk.Scrollbar(
            panel_detalle, orient="vertical", command=self.tree_boletines_detalle.yview
        )
        scrollbar_detalle_y.grid(row=1, column=1, sticky="ns")
        scrollbar_detalle_x = ttk.Scrollbar(
            panel_detalle,
            orient="horizontal",
            command=self.tree_boletines_detalle.xview,
        )
        scrollbar_detalle_x.grid(row=2, column=0, sticky="ew")
        self.tree_boletines_detalle.configure(
            yscrollcommand=scrollbar_detalle_y.set,
            xscrollcommand=scrollbar_detalle_x.set,
        )

        self._boletines_detalle_por_documento = {}
        self._boletines_refrescar_cursos()
        periodos = self._listar_periodos_configurados()
        if periodos:
            self.boletines_var_periodo.set(periodos[0])
        self._boletines_recargar_tablero()

    def _boletines_refrescar_cursos(self, event=None):
        grado = (
            self.boletines_var_grado.get()
            if hasattr(self, "boletines_var_grado")
            else "Todos"
        )
        cursos = self._listar_cursos_catalogo(grado=grado, incluir_todos=True)
        if hasattr(self, "combo_boletines_curso"):
            self.combo_boletines_curso.configure(values=cursos)
        curso_actual = (
            self.boletines_var_curso.get()
            if hasattr(self, "boletines_var_curso")
            else "Todos"
        )
        if curso_actual not in cursos:
            self.boletines_var_curso.set("Todos")
        if event is not None:
            self._boletines_recargar_tablero()

    def _boletines_recargar_tablero(self):
        if not self._tiene_permiso("desktop.superadmin.boletines.actualizar"):
            return
        if not hasattr(self, "tree_boletines"):
            return

        for tree in (self.tree_boletines, self.tree_boletines_detalle):
            for item in tree.get_children():
                tree.delete(item)
        self._boletines_detalle_por_documento = {}

        grado = str(self.boletines_var_grado.get() or "Todos").strip()
        curso = str(self.boletines_var_curso.get() or "Todos").strip()
        area_sel = str(self.boletines_var_area.get() or "Todos").strip()
        periodo = str(self.boletines_var_periodo.get() or "").strip()
        anio = (
            str(self.boletines_var_anio.get() or "").strip()
            or self._obtener_anio_lectivo_base()
        )
        if not periodo:
            self.lbl_boletines_estado.config(
                text="Seleccione un período para calcular el alistamiento de boletines."
            )
            return

        consulta = (
            "SELECT documento, nombre, apellido1, apellido2, nombre1, nombre2, grado, curso "
            "FROM estudiantes WHERE COALESCE(estado, 'Activo') = 'Activo'"
        )
        params = []
        if grado and grado != "Todos":
            consulta += " AND TRIM(COALESCE(grado, '')) = ?"
            params.append(grado)
        if curso and curso != "Todos":
            consulta += " AND TRIM(COALESCE(curso, '')) = ?"
            params.append(curso)
        consulta += (
            " ORDER BY grado, curso, apellido1, apellido2, nombre1, nombre2, nombre"
        )

        try:
            estudiantes = [
                {
                    "documento": str(fila[0] or "").strip(),
                    "nombre": fila[1],
                    "apellido1": fila[2],
                    "apellido2": fila[3],
                    "nombre1": fila[4],
                    "nombre2": fila[5],
                    "grado": str(fila[6] or "").strip(),
                    "curso": str(fila[7] or "").strip(),
                }
                for fila in self.cur.execute(consulta, params).fetchall()
            ]
        except Exception as exc:
            self.lbl_boletines_estado.config(
                text=f"No fue posible cargar la matrícula activa: {exc}"
            )
            return

        if not estudiantes:
            self.lbl_boletines_estado.config(
                text="No hay estudiantes activos para los filtros seleccionados."
            )
            return

        escala = self._obtener_escala_valoracion()

        plan_por_grupo = {}
        try:
            for fila in self.cur.execute(
                "SELECT TRIM(COALESCE(grado, '')), TRIM(COALESCE(curso, '')), COUNT(DISTINCT TRIM(COALESCE(area, ''))) "
                "FROM plan_estudio WHERE COALESCE(estado, 1) = 1 GROUP BY TRIM(COALESCE(grado, '')), TRIM(COALESCE(curso, ''))"
            ).fetchall():
                plan_por_grupo[
                    (str(fila[0] or "").strip(), str(fila[1] or "").strip())
                ] = int(fila[2] or 0)
        except Exception:
            plan_por_grupo = {}

        notas_por_documento = self._boletines_obtener_notas_definitivas_por_area(
            periodo,
            anio,
            area_sel=area_sel,
        )

        autoeval_por_documento = {}
        if self._tabla_existe("autoevaluacion_planilla_sync"):
            try:
                consulta_autoeval = "SELECT documento, COUNT(*) FROM autoevaluacion_planilla_sync WHERE TRIM(COALESCE(periodo, '')) = ? "
                params_autoeval = [periodo]
                if area_sel and area_sel != "Todos":
                    consulta_autoeval += "AND TRIM(COALESCE(area, '')) = ? "
                    params_autoeval.append(area_sel)
                consulta_autoeval += "GROUP BY documento"
                for documento, cantidad in self.cur.execute(
                    consulta_autoeval,
                    params_autoeval,
                ).fetchall():
                    autoeval_por_documento[str(documento or "").strip()] = (
                        self._coerce_int(cantidad, 0)
                    )
            except Exception:
                autoeval_por_documento = {}

        plantillas_desempeno = {}
        try:
            consulta_descriptores = (
                "SELECT grado, area, nivel, descriptor FROM desempenos_plantilla "
                "WHERE TRIM(COALESCE(periodo, '')) = ? AND TRIM(COALESCE(anio_lectivo, '')) = ?"
            )
            params_descriptores = [periodo, anio]
            if grado and grado != "Todos":
                consulta_descriptores += " AND TRIM(COALESCE(grado, '')) = ?"
                params_descriptores.append(grado)
            if area_sel and area_sel != "Todos":
                consulta_descriptores += " AND TRIM(COALESCE(area, '')) = ?"
                params_descriptores.append(area_sel)
            for grado_db, area_db, nivel_db, descriptor_db in self.cur.execute(
                consulta_descriptores,
                params_descriptores,
            ).fetchall():
                plantillas_desempeno[
                    (
                        str(grado_db or "").strip(),
                        str(area_db or "").strip(),
                        str(nivel_db or "").strip(),
                    )
                ] = str(descriptor_db or "").strip()
        except Exception:
            plantillas_desempeno = {}

        total_con_promedio = 0
        total_listos = 0
        total_pendientes = 0
        total_sin_notas = 0
        suma_promedios = 0.0

        for estudiante in estudiantes:
            documento = estudiante["documento"]
            nombre = self._construir_nombre_estudiante(estudiante) or documento
            grado_est = estudiante["grado"]
            curso_est = estudiante["curso"]
            areas = sorted(
                notas_por_documento.get(documento, []),
                key=lambda item: item["area"].lower(),
            )
            areas_plan = (
                1
                if area_sel and area_sel != "Todos"
                else (
                    plan_por_grupo.get((grado_est, curso_est))
                    or plan_por_grupo.get((grado_est, ""))
                    or 0
                )
            )
            areas_calificadas = len(areas)
            promedio_general = None
            desempeno_general = "Sin nota"
            if areas:
                promedio_general = sum(item["promedio"] for item in areas) / max(
                    1, len(areas)
                )
                desempeno_general = self._clasificar_desempeno_nota(
                    promedio_general, escala
                )
                total_con_promedio += 1
                suma_promedios += promedio_general

            if areas_calificadas == 0:
                estado = "Sin notas"
                total_sin_notas += 1
            elif areas_plan and areas_calificadas < areas_plan:
                estado = "Pendiente"
                total_pendientes += 1
            else:
                estado = "Listo"
                total_listos += 1

            detalle_areas = []
            for area in areas:
                nivel_area = self._clasificar_desempeno_nota(area["promedio"], escala)
                descriptor = plantillas_desempeno.get(
                    (grado_est, area["area"], nivel_area), ""
                )
                detalle_areas.append(
                    {
                        "area": area["area"],
                        "promedio": area["promedio"],
                        "desempeno": nivel_area,
                        "registros": area["registros"],
                        "descriptor": descriptor,
                    }
                )

            self._boletines_detalle_por_documento[documento] = {
                "nombre": nombre,
                "grado": grado_est,
                "curso": curso_est,
                "estado": estado,
                "promedio_general": promedio_general,
                "areas": detalle_areas,
                "autoevaluaciones": autoeval_por_documento.get(documento, 0),
                "areas_plan": areas_plan,
            }
            self.tree_boletines.insert(
                "",
                "end",
                iid=documento,
                values=(
                    documento,
                    nombre,
                    grado_est,
                    curso_est,
                    f"{areas_calificadas}/{areas_plan or 0}",
                    "" if promedio_general is None else f"{promedio_general:.2f}",
                    desempeno_general,
                    estado,
                ),
            )

        total_estudiantes = len(estudiantes)
        promedio_grupo = (
            (suma_promedios / total_con_promedio) if total_con_promedio else 0.0
        )
        self.boletines_stats_labels["Estudiantes"].config(text=str(total_estudiantes))
        self.boletines_stats_labels["Con promedio"].config(text=str(total_con_promedio))
        self.boletines_stats_labels["Listos"].config(text=str(total_listos))
        self.boletines_stats_labels["Pendientes"].config(text=str(total_pendientes))
        self.boletines_stats_labels["Sin notas"].config(text=str(total_sin_notas))
        self.boletines_stats_labels["Promedio grupo"].config(
            text=f"{promedio_grupo:.2f}"
        )

        plantillas_count = len({(k[0], k[1]) for k in plantillas_desempeno.keys()})
        self.lbl_boletines_estado.config(
            text=(
                f"Período {periodo}, año {anio}"
                + (f", área {area_sel}" if area_sel and area_sel != "Todos" else "")
                + f": {total_listos} estudiantes listos, "
                f"{total_pendientes} con cobertura incompleta, {total_sin_notas} sin registros y "
                f"{plantillas_count} áreas con descriptor pedagógico disponible."
            )
        )

    def _boletines_mostrar_detalle_estudiante(self, event=None):
        if not hasattr(self, "tree_boletines"):
            return
        seleccion = self.tree_boletines.selection()
        for item in self.tree_boletines_detalle.get_children():
            self.tree_boletines_detalle.delete(item)
        if not seleccion:
            self.lbl_boletines_detalle.config(
                text="Seleccione un estudiante para revisar promedios por área y descriptor sugerido."
            )
            return
        documento = seleccion[0]
        detalle = self._boletines_detalle_por_documento.get(documento, {})
        nombre = detalle.get("nombre", documento)
        promedio_general = detalle.get("promedio_general")
        promedio_txt = (
            "Sin promedio" if promedio_general is None else f"{promedio_general:.2f}"
        )
        self.lbl_boletines_detalle.config(
            text=(
                f"{nombre} | Grado {detalle.get('grado', '')} {detalle.get('curso', '')} | "
                f"Estado: {detalle.get('estado', 'Sin definir')} | "
                f"Promedio general: {promedio_txt} | "
                f"Autoevaluaciones sincronizadas: {detalle.get('autoevaluaciones', 0)}"
            )
        )
        for area in detalle.get("areas", []):
            self.tree_boletines_detalle.insert(
                "",
                "end",
                values=(
                    area["area"],
                    f"{area['promedio']:.2f}",
                    area["desempeno"],
                    area["registros"],
                    self._resumir_texto(
                        area.get("descriptor", "Sin plantilla registrada"), 120
                    ),
                ),
            )

    def _boletines_obtener_columnas_nota(self):
        cfg = self._obtener_componentes_evaluacion()
        columnas = []
        indice_global = 1
        for clave in ("cognitivo", "examen", "autoevaluacion"):
            componente = dict(cfg.get(clave) or {})
            cantidad = max(0, self._coerce_int(componente.get("cantidad"), 0))
            nombre = str(componente.get("nombre") or clave.title()).strip()
            prefijo = str(componente.get("prefijo") or clave[:1]).strip().upper()
            for indice in range(1, cantidad + 1):
                columnas.append(
                    {
                        "nota_clave": f"{prefijo}{indice}",
                        "titulo": f"N{indice_global}",
                        "aspecto": nombre,
                    }
                )
                indice_global += 1
        return columnas

    def _boletines_calcular_definitiva_desde_notas(self, notas_por_clave):
        cfg = self._obtener_componentes_evaluacion()
        notas_map = {
            str(clave or "").strip().upper(): self._coerce_float(valor, None)
            for clave, valor in (notas_por_clave or {}).items()
            if str(clave or "").strip()
        }
        definitiva = 0.0
        hay_notas = False
        indice_global = 1

        for clave in ("cognitivo", "examen", "autoevaluacion"):
            componente = dict(cfg.get(clave) or {})
            cantidad = max(0, self._coerce_int(componente.get("cantidad"), 0))
            prefijo = str(componente.get("prefijo") or clave[:1]).strip().upper()
            notas_componente = []
            for indice in range(1, cantidad + 1):
                valor = None
                for alias in (f"{prefijo}{indice}", f"N{indice_global}"):
                    if alias in notas_map and notas_map[alias] is not None:
                        valor = notas_map[alias]
                        break
                indice_global += 1
                if valor is None:
                    continue
                notas_componente.append(valor)
            if notas_componente:
                hay_notas = True
                promedio_componente = sum(notas_componente) / len(notas_componente)
                definitiva += (
                    promedio_componente
                    * float(self._coerce_float(componente.get("porcentaje"), 0.0))
                    / 100.0
                )

        if not hay_notas:
            return None
        return round(definitiva, 4)

    def _boletines_obtener_notas_definitivas_por_area(
        self, periodo, anio, area_sel=None
    ):
        notas_por_documento = {}
        try:
            consulta_notas = (
                "SELECT COALESCE(NULLIF(TRIM(e.documento), ''), CAST(pc.estudiante_id AS TEXT)), "
                "COALESCE(NULLIF(TRIM(a.nombre), ''), 'Área sin nombre'), pc.area_id, pc.nota_clave, pc.nota "
                "FROM planillas_calificaciones pc "
                "LEFT JOIN areas a ON a.id = pc.area_id "
                "LEFT JOIN estudiantes e ON CAST(e.id AS TEXT) = CAST(pc.estudiante_id AS TEXT) "
                "WHERE TRIM(COALESCE(pc.periodo, '')) = ? AND TRIM(COALESCE(pc.anio_lectivo, '')) = ? "
            )
            params_notas = [periodo, anio]
            if area_sel and area_sel != "Todos":
                consulta_notas += "AND TRIM(COALESCE(a.nombre, '')) = ? "
                params_notas.append(area_sel)
            consulta_notas += "ORDER BY pc.estudiante_id, COALESCE(NULLIF(TRIM(a.nombre), ''), 'Área sin nombre'), pc.nota_clave"

            acumulado = {}
            for documento, area, area_id, nota_clave, nota in self.cur.execute(
                consulta_notas, params_notas
            ).fetchall():
                documento_txt = str(documento or "").strip()
                area_txt = str(area or "Área sin nombre").strip() or "Área sin nombre"
                if not documento_txt:
                    continue
                llave = (documento_txt, area_txt, area_id)
                bucket = acumulado.setdefault(llave, {"notas": {}, "registros": 0})
                bucket["notas"][str(nota_clave or "").strip().upper() or "N1"] = nota
                if nota not in (None, ""):
                    bucket["registros"] += 1

            for (documento_txt, area_txt, _area_id), bucket in acumulado.items():
                definitiva = self._boletines_calcular_definitiva_desde_notas(
                    bucket.get("notas")
                )
                if definitiva is None:
                    continue
                notas_por_documento.setdefault(documento_txt, []).append(
                    {
                        "area": area_txt,
                        "promedio": definitiva,
                        "registros": self._coerce_int(bucket.get("registros"), 0),
                    }
                )
        except Exception:
            notas_por_documento = {}
        return notas_por_documento

    def _boletines_resolver_area_id(self, area_nombre):
        try:
            fila = self.cur.execute(
                "SELECT id FROM areas WHERE TRIM(COALESCE(nombre, '')) = ? LIMIT 1",
                (str(area_nombre or "").strip(),),
            ).fetchone()
            if fila:
                return fila[0]
        except Exception:
            pass
        return None

    def _boletines_obtener_dataset_pdf(self):
        grado = str(self.boletines_var_grado.get() or "Todos").strip()
        curso = str(self.boletines_var_curso.get() or "Todos").strip()
        periodo = str(self.boletines_var_periodo.get() or "").strip()
        anio = (
            str(self.boletines_var_anio.get() or "").strip()
            or self._obtener_anio_lectivo_base()
        )
        if not periodo:
            raise ValueError("Seleccione un período.")
        if not grado or grado == "Todos" or not curso or curso == "Todos":
            raise ValueError(
                "Seleccione un grado y un curso específicos para generar el boletín del curso."
            )

        consulta = (
            "SELECT documento, codigo, apellido1, apellido2, nombre1, nombre2, nombre, grado, curso, jornada, sede "
            "FROM estudiantes WHERE COALESCE(estado, 'Activo') = 'Activo'"
        )
        params = []
        consulta += " AND TRIM(COALESCE(grado, '')) = ?"
        params.append(grado)
        consulta += " AND TRIM(COALESCE(curso, '')) = ?"
        params.append(curso)
        consulta += (
            " ORDER BY grado, curso, apellido1, apellido2, nombre1, nombre2, nombre"
        )
        estudiantes = self.cur.execute(consulta, params).fetchall()
        if not estudiantes:
            raise ValueError(
                "No hay estudiantes activos para el grado y curso seleccionados."
            )

        notas_por_documento = self._boletines_obtener_notas_definitivas_por_area(
            periodo,
            anio,
            area_sel=None,
        )

        plantillas_desempeno = {}
        try:
            for grado_db, area_db, nivel_db, descriptor_db in self.cur.execute(
                "SELECT grado, area, nivel, descriptor FROM desempenos_plantilla "
                "WHERE TRIM(COALESCE(periodo, '')) = ? AND TRIM(COALESCE(anio_lectivo, '')) = ? AND TRIM(COALESCE(grado, '')) = ?",
                (periodo, anio, grado),
            ).fetchall():
                plantillas_desempeno[
                    (
                        str(grado_db or "").strip(),
                        str(area_db or "").strip(),
                        str(nivel_db or "").strip(),
                    )
                ] = str(descriptor_db or "").strip()
        except Exception:
            plantillas_desempeno = {}

        institucion = self._get_config_plantel("nombre_institucion") or "Colegio"
        sede = ""
        jornada = ""
        filas = []
        escala = self._obtener_escala_valoracion()
        for indice, fila in enumerate(estudiantes, start=1):
            documento = str(fila[0] or "").strip()
            codigo = str(fila[1] or documento).strip() or documento
            apellido1 = str(fila[2] or "").strip()
            apellido2 = str(fila[3] or "").strip()
            nombre1 = str(fila[4] or "").strip()
            nombre2 = str(fila[5] or "").strip()
            nombre = str(fila[6] or "").strip()
            grado_est = str(fila[7] or "").strip()
            curso_est = str(fila[8] or "").strip()
            jornada = jornada or str(fila[9] or "").strip()
            sede = sede or str(fila[10] or "").strip()
            areas_est = sorted(
                notas_por_documento.get(documento, []),
                key=lambda item: item["area"].lower(),
            )
            notas_validas = [
                item["promedio"]
                for item in areas_est
                if item.get("promedio") is not None
            ]
            promedio = (
                sum(notas_validas) / len(notas_validas) if notas_validas else None
            )
            desempeno = self._clasificar_desempeno_nota(promedio, escala)
            detalle_areas = []
            for area_meta in areas_est:
                nivel_area = self._clasificar_desempeno_nota(
                    area_meta["promedio"], escala
                )
                descriptor = plantillas_desempeno.get(
                    (grado_est, area_meta["area"], nivel_area), ""
                )
                detalle_areas.append(
                    {
                        "area": area_meta["area"],
                        "promedio": area_meta["promedio"],
                        "desempeno": nivel_area,
                        "descriptor": descriptor,
                        "registros": area_meta["registros"],
                    }
                )
            filas.append(
                {
                    "no": indice,
                    "codigo": codigo,
                    "apellido1": apellido1,
                    "apellido2": apellido2,
                    "nombre1": nombre1 or nombre,
                    "nombre2": nombre2,
                    "promedio": "" if promedio is None else f"{promedio:.1f}",
                    "desempeno": desempeno if promedio is not None else "",
                    "grado": grado_est,
                    "curso": curso_est,
                    "documento": documento,
                    "nombre": self._construir_nombre_estudiante(
                        {
                            "nombre": nombre,
                            "apellido1": apellido1,
                            "apellido2": apellido2,
                            "nombre1": nombre1,
                            "nombre2": nombre2,
                        }
                    )
                    or documento,
                    "areas": detalle_areas,
                }
            )

        return {
            "institucion": institucion,
            "resolucion": self._get_config_plantel("resolucion_aprobacion") or "",
            "decreto": self._get_config_plantel("decreto_funcionamiento") or "",
            "dane": self._get_config_plantel("codigo_dane") or "",
            "nit": self._get_config_plantel("nit") or "",
            "logo_path": self._get_config_plantel("logo_path") or "",
            "sede": sede or "-",
            "jornada": jornada or "-",
            "grado": grado,
            "curso": curso,
            "periodo": periodo,
            "anio": anio,
            "filas": filas,
        }

    def _boletines_obtener_dataset_pdf_planilla(self):
        grado = str(self.boletines_var_grado.get() or "Todos").strip()
        curso = str(self.boletines_var_curso.get() or "Todos").strip()
        periodo = str(self.boletines_var_periodo.get() or "").strip()
        anio = (
            str(self.boletines_var_anio.get() or "").strip()
            or self._obtener_anio_lectivo_base()
        )
        if not periodo:
            raise ValueError("Seleccione un período.")
        if not grado or grado == "Todos" or not curso or curso == "Todos":
            raise ValueError(
                "Seleccione un grado y un curso específicos para generar el boletín del curso."
            )

        consulta = (
            "SELECT id, documento, codigo, apellido1, apellido2, nombre1, nombre2, nombre, grado, curso, jornada, sede "
            "FROM estudiantes WHERE COALESCE(estado, 'Activo') = 'Activo' "
            "AND TRIM(COALESCE(grado, '')) = ? AND TRIM(COALESCE(curso, '')) = ? "
            "ORDER BY apellido1, apellido2, nombre1, nombre2, nombre"
        )
        estudiantes_raw = self.cur.execute(consulta, (grado, curso)).fetchall()
        if not estudiantes_raw:
            raise ValueError(
                "No hay estudiantes activos para el grado y curso seleccionados."
            )

        estudiantes = []
        documentos = []
        sede = ""
        jornada = ""
        for fila in estudiantes_raw:
            estudiante = {
                "id": str(fila[0] or "").strip(),
                "documento": str(fila[1] or "").strip(),
                "codigo": str(fila[2] or fila[1] or "").strip(),
                "apellido1": str(fila[3] or "").strip(),
                "apellido2": str(fila[4] or "").strip(),
                "nombre1": str(fila[5] or "").strip(),
                "nombre2": str(fila[6] or "").strip(),
                "nombre": str(fila[7] or "").strip(),
                "grado": str(fila[8] or "").strip(),
                "curso": str(fila[9] or "").strip(),
            }
            sede = sede or str(fila[11] or "").strip()
            jornada = jornada or str(fila[10] or "").strip()
            estudiantes.append(estudiante)
            documentos.append(estudiante["documento"])

        area_sel = str(self.boletines_var_area.get() or "Todos").strip()
        areas = []
        try:
            consulta_areas = (
                "SELECT DISTINCT COALESCE(NULLIF(TRIM(a.nombre), ''), TRIM(COALESCE(p.area, ''))) AS area_nombre "
                "FROM plan_estudio p "
                "LEFT JOIN areas a ON a.id = p.IdArea "
                "WHERE TRIM(COALESCE(p.grado, '')) = ? AND TRIM(COALESCE(p.curso, '')) = ? AND COALESCE(p.estado, 1) = 1 "
                "ORDER BY area_nombre"
            )
            areas = [
                str(fila[0] or "").strip()
                for fila in self.cur.execute(consulta_areas, (grado, curso)).fetchall()
                if str(fila[0] or "").strip()
            ]
        except Exception:
            areas = []

        if area_sel and area_sel != "Todos":
            areas = [area_sel]

        if not areas:
            notas_tmp = self._boletines_obtener_notas_definitivas_por_area(
                periodo,
                anio,
                area_sel=None,
            )
            areas = sorted(
                {
                    str(item.get("area") or "").strip()
                    for lista in notas_tmp.values()
                    for item in lista
                    if str(item.get("area") or "").strip()
                }
            )

        columnas_nota = self._boletines_obtener_columnas_nota()
        escala = self._obtener_escala_valoracion()
        institucion = self._get_config_plantel("nombre_institucion") or "Colegio"

        plantillas_desempeno = {}
        try:
            for grado_db, area_db, nivel_db, descriptor_db in self.cur.execute(
                "SELECT grado, area, nivel, descriptor FROM desempenos_plantilla "
                "WHERE TRIM(COALESCE(periodo, '')) = ? AND TRIM(COALESCE(anio_lectivo, '')) = ? AND TRIM(COALESCE(grado, '')) = ?",
                (periodo, anio, grado),
            ).fetchall():
                plantillas_desempeno[
                    (
                        str(grado_db or "").strip(),
                        str(area_db or "").strip(),
                        str(nivel_db or "").strip(),
                    )
                ] = str(descriptor_db or "").strip()
        except Exception:
            plantillas_desempeno = {}

        notas_crudas = {}
        try:
            consulta_notas = (
                "SELECT COALESCE(NULLIF(TRIM(e.documento), ''), CAST(pc.estudiante_id AS TEXT)), "
                "COALESCE(NULLIF(TRIM(a.nombre), ''), 'Área sin nombre'), pc.nota_clave, pc.nota "
                "FROM planillas_calificaciones pc "
                "LEFT JOIN areas a ON a.id = pc.area_id "
                "LEFT JOIN estudiantes e ON CAST(e.id AS TEXT) = CAST(pc.estudiante_id AS TEXT) "
                "WHERE TRIM(COALESCE(pc.periodo, '')) = ? AND TRIM(COALESCE(pc.anio_lectivo, '')) = ?"
            )
            for documento_db, area_db, nota_clave, nota in self.cur.execute(
                consulta_notas,
                (periodo, anio),
            ).fetchall():
                documento_txt = str(documento_db or "").strip()
                area_txt = str(area_db or "").strip()
                if not documento_txt or not area_txt:
                    continue
                notas_crudas.setdefault(area_txt, {}).setdefault(documento_txt, {})[
                    str(nota_clave or "").strip().upper() or "N1"
                ] = nota
        except Exception:
            notas_crudas = {}

        secciones = []
        for area in areas:
            docente = ""
            try:
                fila_doc = self.cur.execute(
                    "SELECT d.nombre, ca.docente_documento FROM carga_academica ca "
                    "LEFT JOIN docentes d ON d.documento = ca.docente_documento "
                    "WHERE TRIM(COALESCE(ca.grado, '')) = ? AND TRIM(COALESCE(ca.curso, '')) = ? "
                    "AND TRIM(COALESCE(ca.area, '')) = ? LIMIT 1",
                    (grado, curso, area),
                ).fetchone()
                if fila_doc:
                    docente = str(fila_doc[0] or fila_doc[1] or "").strip()
            except Exception:
                docente = ""

            filas_area = []
            for indice, estudiante in enumerate(estudiantes, start=1):
                documento = estudiante["documento"]
                notas_est = notas_crudas.get(area, {}).get(documento, {})
                nota_cols = []
                notas_def = {}
                for meta in columnas_nota:
                    valor = ""
                    for alias in (
                        str(meta.get("nota_clave") or "").strip().upper(),
                        str(meta.get("titulo") or "").strip().upper(),
                    ):
                        if alias and alias in notas_est:
                            valor = notas_est.get(alias)
                            break
                    nota_cols.append(
                        "" if valor in (None, "") else f"{float(valor):.1f}"
                    )
                    if valor not in (None, ""):
                        notas_def[str(meta.get("nota_clave") or "").strip().upper()] = (
                            valor
                        )
                        notas_def[str(meta.get("titulo") or "").strip().upper()] = valor

                definitiva_num = self._boletines_calcular_definitiva_desde_notas(
                    notas_def
                )
                definitiva_txt = (
                    "" if definitiva_num is None else f"{definitiva_num:.1f}"
                )
                desempeno = self._clasificar_desempeno_nota(definitiva_num, escala)
                descriptor = plantillas_desempeno.get(
                    (estudiante["grado"], area, desempeno),
                    "",
                )
                filas_area.append(
                    {
                        "no": indice,
                        "codigo": estudiante["codigo"],
                        "apellido1": estudiante["apellido1"],
                        "apellido2": estudiante["apellido2"],
                        "nombre1": estudiante["nombre1"] or estudiante["nombre"],
                        "nombre2": estudiante["nombre2"],
                        "notas": nota_cols,
                        "promedio": definitiva_txt,
                        "descriptor": descriptor,
                        "desempeno": desempeno if definitiva_num is not None else "",
                    }
                )

            if not any(str(fila.get("promedio") or "").strip() for fila in filas_area):
                continue

            secciones.append(
                {
                    "area": area,
                    "docente": docente,
                    "filas": filas_area,
                }
            )

        if not secciones:
            raise ValueError(
                "No hay notas definitivas registradas en las planillas para el curso y período seleccionados."
            )

        return {
            "institucion": institucion,
            "resolucion": self._get_config_plantel("resolucion_aprobacion") or "",
            "decreto": self._get_config_plantel("decreto_funcionamiento") or "",
            "dane": self._get_config_plantel("codigo_dane") or "",
            "nit": self._get_config_plantel("nit") or "",
            "logo_path": self._get_config_plantel("logo_path") or "",
            "sede": sede or "-",
            "jornada": jornada or "-",
            "grado": grado,
            "curso": curso,
            "periodo": periodo,
            "anio": anio,
            "columnas_nota": columnas_nota,
            "secciones": secciones,
        }

    def _boletines_numero_periodo(self, periodo):
        import re
        import unicodedata

        texto = str(periodo or "").strip().lower()
        if not texto:
            return None
        match = re.search(r"(\d+)", texto)
        if match:
            return self._coerce_int(match.group(1), None)
        normalizado = unicodedata.normalize("NFKD", texto)
        normalizado = "".join(
            caracter for caracter in normalizado if not unicodedata.combining(caracter)
        )
        equivalencias = {
            "primer periodo": 1,
            "periodo primero": 1,
            "primero": 1,
            "segundo periodo": 2,
            "periodo segundo": 2,
            "segundo": 2,
            "tercer periodo": 3,
            "tercero": 3,
            "cuarto periodo": 4,
            "cuarto": 4,
            "quinto periodo": 5,
            "quinto": 5,
            "sexto periodo": 6,
            "sexto": 6,
        }
        return equivalencias.get(normalizado)

    def _boletines_obtener_regla_escala(self, nota, escala=None):
        try:
            valor = float(nota)
        except Exception:
            return None
        for fila in escala or self._obtener_escala_valoracion():
            if fila["desde"] <= valor <= fila["hasta"]:
                return fila
        return None

    def _boletines_resolver_rango_periodo(self, periodo, anio):
        numero_periodo = self._boletines_numero_periodo(periodo)
        if not numero_periodo:
            return None
        try:
            filas = self.cur.execute(
                "SELECT periodo, fecha_inicio, fecha_fin FROM calendario_academico"
            ).fetchall()
        except Exception:
            filas = []
        candidatos = []
        for periodo_db, fecha_inicio, fecha_fin in filas:
            if self._boletines_numero_periodo(periodo_db) != numero_periodo:
                continue
            try:
                inicio_dt = datetime.strptime(
                    str(fecha_inicio or "").strip(), "%Y-%m-%d"
                )
                fin_dt = datetime.strptime(str(fecha_fin or "").strip(), "%Y-%m-%d")
            except Exception:
                continue
            if inicio_dt.year == int(anio) or fin_dt.year == int(anio):
                return (
                    inicio_dt.strftime("%Y-%m-%d"),
                    fin_dt.strftime("%Y-%m-%d"),
                )
            candidatos.append((inicio_dt, fin_dt))
        if not candidatos:
            return None
        inicio_base, fin_base = candidatos[0]
        try:
            inicio = inicio_base.replace(year=int(anio)).strftime("%Y-%m-%d")
            fin = fin_base.replace(year=int(anio)).strftime("%Y-%m-%d")
            return (inicio, fin)
        except Exception:
            return None

    def _boletines_obtener_inasistencias_grupo(
        self,
        estudiantes,
        grado,
        curso,
        jornada,
        sede,
        periodo,
        anio,
    ):
        if not estudiantes:
            return {}
        ids = [
            str(est.get("id") or "").strip()
            for est in estudiantes
            if str(est.get("id") or "").strip()
        ]
        if not ids:
            return {}
        placeholders = ",".join("?" for _ in ids)
        params = [grado, curso, jornada, sede]
        rango = self._boletines_resolver_rango_periodo(periodo, anio)
        consulta = (
            "SELECT CAST(estudiante_id AS TEXT), COUNT(*) FROM asistencia "
            "WHERE TRIM(COALESCE(grado, '')) = ? AND TRIM(COALESCE(curso, '')) = ? "
            "AND TRIM(COALESCE(jornada, '')) = ? AND TRIM(COALESCE(sede, '')) = ? "
            "AND UPPER(TRIM(COALESCE(estado, ''))) IN ('F', 'FJ', 'A', 'J') "
            f"AND CAST(estudiante_id AS TEXT) IN ({placeholders}) "
        )
        params.extend(ids)
        if rango:
            consulta += "AND fecha BETWEEN ? AND ? "
            params.extend(rango)
        else:
            consulta += "AND substr(COALESCE(fecha, ''), 1, 4) = ? "
            params.append(str(anio))
        consulta += "GROUP BY CAST(estudiante_id AS TEXT)"
        try:
            return {
                str(estudiante_id or "").strip(): self._coerce_int(total, 0)
                for estudiante_id, total in self.cur.execute(
                    consulta, params
                ).fetchall()
            }
        except Exception:
            return {}

    def _boletines_obtener_plan_grupo(self, grado, curso):
        try:
            filas = self.cur.execute(
                "SELECT COALESCE(NULLIF(TRIM(a.nombre), ''), TRIM(COALESCE(p.area, ''))), COALESCE(p.horas, 0) "
                "FROM plan_estudio p "
                "LEFT JOIN areas a ON a.id = p.IdArea "
                "WHERE TRIM(COALESCE(p.grado, '')) = ? AND TRIM(COALESCE(p.curso, '')) = ? "
                "AND COALESCE(p.estado, 1) = 1 "
                "ORDER BY COALESCE(NULLIF(TRIM(a.nombre), ''), TRIM(COALESCE(p.area, ''))) ASC",
                (grado, curso),
            ).fetchall()
        except Exception:
            filas = []
        resultado = []
        for area, horas in filas:
            area_txt = str(area or "").strip()
            if not area_txt:
                continue
            resultado.append(
                {
                    "area": area_txt,
                    "horas": self._coerce_int(horas, 0),
                }
            )
        return resultado

    def _boletines_obtener_docentes_grupo(self, grado, curso):
        try:
            filas = self.cur.execute(
                "SELECT TRIM(COALESCE(ca.area, '')), COALESCE(NULLIF(TRIM(d.nombre), ''), TRIM(COALESCE(ca.docente_documento, ''))) "
                "FROM carga_academica ca "
                "LEFT JOIN docentes d ON d.documento = ca.docente_documento "
                "WHERE TRIM(COALESCE(ca.grado, '')) = ? AND TRIM(COALESCE(ca.curso, '')) = ?",
                (grado, curso),
            ).fetchall()
        except Exception:
            filas = []
        return {
            str(area or "").strip(): str(docente or "").strip()
            for area, docente in filas
            if str(area or "").strip()
        }

    def _boletines_obtener_descriptores_periodo(self, periodo, anio):
        try:
            filas = self.cur.execute(
                "SELECT grado, area, nivel, descriptor FROM desempenos_plantilla "
                "WHERE TRIM(COALESCE(periodo, '')) = ? AND TRIM(COALESCE(anio_lectivo, '')) = ?",
                (periodo, anio),
            ).fetchall()
        except Exception:
            filas = []
        return {
            (
                str(grado or "").strip(),
                str(area or "").strip(),
                str(nivel or "").strip(),
            ): str(descriptor or "").strip()
            for grado, area, nivel, descriptor in filas
        }

    def _boletines_calcular_resumen_curso(self, estudiantes_grupo, notas_por_documento):
        promedios = {}
        for estudiante in estudiantes_grupo:
            documento = str(estudiante.get("documento") or "").strip()
            notas_validas = [
                fila["promedio"]
                for fila in notas_por_documento.get(documento, [])
                if fila.get("promedio") is not None
            ]
            promedios[documento] = (
                (sum(notas_validas) / len(notas_validas)) if notas_validas else None
            )
        ordenados = sorted(
            [
                (doc, promedio)
                for doc, promedio in promedios.items()
                if promedio is not None
            ],
            key=lambda item: (-item[1], item[0]),
        )
        puestos = {}
        posicion_actual = 0
        promedio_previo = None
        for indice, (documento, promedio) in enumerate(ordenados, start=1):
            if promedio_previo is None or round(promedio, 4) != round(
                promedio_previo, 4
            ):
                posicion_actual = indice
                promedio_previo = promedio
            puestos[documento] = posicion_actual
        promedios_validos = [
            promedio for promedio in promedios.values() if promedio is not None
        ]
        promedio_curso = (
            (sum(promedios_validos) / len(promedios_validos))
            if promedios_validos
            else None
        )
        return promedios, puestos, promedio_curso

    def _boletines_obtener_dataset_pdf_individual(self, documentos=None):
        periodo = str(self.boletines_var_periodo.get() or "").strip()
        anio = (
            str(self.boletines_var_anio.get() or "").strip()
            or self._obtener_anio_lectivo_base()
        )
        if not periodo:
            raise ValueError("Seleccione un período.")

        documentos_filtrados = [
            str(documento or "").strip()
            for documento in (documentos or [])
            if str(documento or "").strip()
        ]
        consulta = (
            "SELECT id, documento, codigo, apellido1, apellido2, nombre1, nombre2, nombre, grado, curso, jornada, sede "
            "FROM estudiantes WHERE COALESCE(estado, 'Activo') = 'Activo'"
        )
        params = []
        if documentos_filtrados:
            consulta += (
                " AND TRIM(COALESCE(documento, '')) IN ("
                + ",".join("?" for _ in documentos_filtrados)
                + ")"
            )
            params.extend(documentos_filtrados)
        else:
            grado = str(self.boletines_var_grado.get() or "Todos").strip()
            curso = str(self.boletines_var_curso.get() or "Todos").strip()
            if not grado or grado == "Todos" or not curso or curso == "Todos":
                raise ValueError(
                    "Seleccione un grado y un curso específicos o marque un estudiante para generar boletines."
                )
            consulta += (
                " AND TRIM(COALESCE(grado, '')) = ? AND TRIM(COALESCE(curso, '')) = ?"
            )
            params.extend([grado, curso])
        consulta += (
            " ORDER BY grado, curso, apellido1, apellido2, nombre1, nombre2, nombre"
        )
        estudiantes_raw = self.cur.execute(consulta, params).fetchall()
        if not estudiantes_raw:
            raise ValueError("No hay estudiantes activos para generar boletines.")

        estudiantes = []
        for fila in estudiantes_raw:
            estudiantes.append(
                {
                    "id": str(fila[0] or "").strip(),
                    "documento": str(fila[1] or "").strip(),
                    "codigo": str(fila[2] or fila[1] or "").strip(),
                    "apellido1": str(fila[3] or "").strip(),
                    "apellido2": str(fila[4] or "").strip(),
                    "nombre1": str(fila[5] or "").strip(),
                    "nombre2": str(fila[6] or "").strip(),
                    "nombre": str(fila[7] or "").strip(),
                    "grado": str(fila[8] or "").strip(),
                    "curso": str(fila[9] or "").strip(),
                    "jornada": str(fila[10] or "").strip(),
                    "sede": str(fila[11] or "").strip(),
                }
            )

        notas_por_documento = self._boletines_obtener_notas_definitivas_por_area(
            periodo,
            anio,
            area_sel=None,
        )
        escala = self._obtener_escala_valoracion()
        descriptores = self._boletines_obtener_descriptores_periodo(periodo, anio)
        grupos = {}
        for estudiante in estudiantes:
            grupos.setdefault((estudiante["grado"], estudiante["curso"]), []).append(
                estudiante
            )

        boletines = []
        promedio_curso_por_grupo = {}
        puestos_por_grupo = {}
        promedio_estudiante_por_grupo = {}
        planes_por_grupo = {}
        docentes_por_grupo = {}
        inasistencias_por_grupo = {}

        for grupo, estudiantes_seleccionados in grupos.items():
            grado, curso = grupo
            estudiantes_grupo = [
                {
                    "documento": str(fila[0] or "").strip(),
                    "grado": grado,
                    "curso": curso,
                }
                for fila in self.cur.execute(
                    "SELECT documento, grado, curso FROM estudiantes "
                    "WHERE COALESCE(estado, 'Activo') = 'Activo' AND TRIM(COALESCE(grado, '')) = ? AND TRIM(COALESCE(curso, '')) = ?",
                    (grado, curso),
                ).fetchall()
            ]
            (
                promedio_estudiante_por_grupo[grupo],
                puestos_por_grupo[grupo],
                promedio_curso_por_grupo[grupo],
            ) = self._boletines_calcular_resumen_curso(
                estudiantes_grupo,
                notas_por_documento,
            )
            planes_por_grupo[grupo] = self._boletines_obtener_plan_grupo(grado, curso)
            docentes_por_grupo[grupo] = self._boletines_obtener_docentes_grupo(
                grado, curso
            )
            muestra = estudiantes_seleccionados[0]
            inasistencias_por_grupo[grupo] = (
                self._boletines_obtener_inasistencias_grupo(
                    estudiantes_seleccionados,
                    grado,
                    curso,
                    muestra.get("jornada", ""),
                    muestra.get("sede", ""),
                    periodo,
                    anio,
                )
            )

        for estudiante in estudiantes:
            grupo = (estudiante["grado"], estudiante["curso"])
            notas_estudiante = {
                str(item.get("area") or "").strip(): item
                for item in notas_por_documento.get(estudiante["documento"], [])
                if str(item.get("area") or "").strip()
            }
            areas_plan = list(planes_por_grupo.get(grupo) or [])
            if not areas_plan:
                areas_plan = [
                    {"area": area, "horas": 0}
                    for area in sorted(
                        notas_estudiante.keys(), key=lambda valor: valor.lower()
                    )
                ]
            filas_areas = []
            docentes_firma = []
            for area_info in areas_plan:
                area = str(area_info.get("area") or "").strip()
                nota_meta = notas_estudiante.get(area, {})
                nota = nota_meta.get("promedio")
                regla = self._boletines_obtener_regla_escala(nota, escala)
                desempeno = (regla or {}).get(
                    "desempeno"
                ) or self._clasificar_desempeno_nota(nota, escala)
                descriptor = descriptores.get(
                    (estudiante["grado"], area, desempeno), ""
                )
                observacion = (
                    descriptor or str((regla or {}).get("recomendacion") or "").strip()
                )
                docente = str(docentes_por_grupo.get(grupo, {}).get(area) or "").strip()
                if docente and docente not in docentes_firma:
                    docentes_firma.append(docente)
                filas_areas.append(
                    {
                        "area": area,
                        "horas": self._coerce_int(area_info.get("horas"), 0),
                        "inasistencias": self._coerce_int(
                            inasistencias_por_grupo.get(grupo, {}).get(
                                estudiante["id"]
                            ),
                            0,
                        ),
                        "definitiva": nota,
                        "definitiva_txt": "" if nota is None else f"{float(nota):.1f}",
                        "desempeno": "" if nota is None else desempeno,
                        "docente": docente,
                        "observacion": observacion,
                    }
                )
            promedio_general = promedio_estudiante_por_grupo.get(grupo, {}).get(
                estudiante["documento"]
            )
            boletines.append(
                {
                    "estudiante_id": estudiante["id"],
                    "documento": estudiante["documento"],
                    "codigo": estudiante["codigo"],
                    "nombre": self._construir_nombre_estudiante(estudiante)
                    or estudiante["documento"],
                    "grado": estudiante["grado"],
                    "curso": estudiante["curso"],
                    "jornada": estudiante["jornada"] or "-",
                    "sede": estudiante["sede"] or "-",
                    "periodo": periodo,
                    "anio": anio,
                    "areas": filas_areas,
                    "promedio_general": promedio_general,
                    "promedio_general_txt": (
                        "" if promedio_general is None else f"{promedio_general:.2f}"
                    ),
                    "puesto": puestos_por_grupo.get(grupo, {}).get(
                        estudiante["documento"]
                    ),
                    "promedio_curso": promedio_curso_por_grupo.get(grupo),
                    "escala": escala,
                    "firma_docente": (
                        ", ".join(docentes_firma)
                        if docentes_firma
                        else "Docente(s) del curso"
                    ),
                }
            )

        boletines.sort(key=lambda item: (item["grado"], item["curso"], item["nombre"]))
        return {
            "institucion": self._get_config_plantel("nombre_institucion") or "Colegio",
            "resolucion": self._get_config_plantel("resolucion_aprobacion") or "",
            "decreto": self._get_config_plantel("decreto_funcionamiento") or "",
            "dane": self._get_config_plantel("codigo_dane") or "",
            "nit": self._get_config_plantel("nit") or "",
            "logo_path": self._get_config_plantel("logo_path") or "",
            "rector_nombre": self._get_config_plantel("rector_nombre") or "",
            "rector_identificacion": self._get_config_plantel("rector_identificacion")
            or "",
            "rector_cargo": self._get_config_plantel("rector_cargo") or "Rector",
            "boletines": boletines,
        }

    def generar_boletin_estudiante(self, estudiante_id):
        fila = self.cur.execute(
            "SELECT documento FROM estudiantes WHERE CAST(id AS TEXT) = CAST(? AS TEXT) LIMIT 1",
            (str(estudiante_id or "").strip(),),
        ).fetchone()
        if not fila:
            raise ValueError("No se encontró el estudiante solicitado.")
        datos = self._boletines_obtener_dataset_pdf_individual([fila[0]])
        boletines = list(datos.get("boletines") or [])
        if not boletines:
            raise ValueError("No fue posible construir el boletín del estudiante.")
        return boletines[0]

    def _boletines_generar_pdf_individual(self, ruta, datos_globales, boletin):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Image,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        doc = SimpleDocTemplate(
            ruta,
            pagesize=A4,
            leftMargin=1.0 * cm,
            rightMargin=1.0 * cm,
            topMargin=0.8 * cm,
            bottomMargin=0.8 * cm,
        )
        styles = getSampleStyleSheet()
        estilo_texto = ParagraphStyle(
            "BoletinTexto",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8,
            leading=9.5,
        )
        estilo_titulo = ParagraphStyle(
            "BoletinTituloIndividual",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=12,
            alignment=1,
            leading=13,
        )
        estilo_subtitulo = ParagraphStyle(
            "BoletinSubtituloIndividual",
            parent=estilo_texto,
            fontName="Helvetica-Bold",
            alignment=1,
        )
        estilo_tabla = ParagraphStyle(
            "BoletinTablaTexto",
            parent=estilo_texto,
            fontSize=7,
            leading=8.2,
        )

        story = []
        logo_path = str(datos_globales.get("logo_path") or "").strip()
        if logo_path and os.path.exists(logo_path):
            try:
                imagen = Image(logo_path, width=1.8 * cm, height=1.8 * cm)
                cabecera = Table(
                    [
                        [
                            imagen,
                            Paragraph(
                                str(datos_globales.get("institucion") or "").upper(),
                                estilo_titulo,
                            ),
                        ]
                    ],
                    colWidths=[2.2 * cm, 15.6 * cm],
                )
                cabecera.setStyle(
                    TableStyle(
                        [
                            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ("LEFTPADDING", (0, 0), (-1, -1), 0),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                            ("TOPPADDING", (0, 0), (-1, -1), 0),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                        ]
                    )
                )
                story.append(cabecera)
            except Exception:
                story.append(
                    Paragraph(
                        str(datos_globales.get("institucion") or "").upper(),
                        estilo_titulo,
                    )
                )
        else:
            story.append(
                Paragraph(
                    str(datos_globales.get("institucion") or "").upper(),
                    estilo_titulo,
                )
            )
        story.append(
            Paragraph(
                (
                    f"Resolución: {datos_globales.get('resolucion') or '-'}  |  "
                    f"Decreto: {datos_globales.get('decreto') or '-'}  |  "
                    f"DANE: {datos_globales.get('dane') or '-'}  |  "
                    f"NIT: {datos_globales.get('nit') or '-'}"
                ),
                estilo_subtitulo,
            )
        )
        story.append(Spacer(1, 0.18 * cm))
        story.append(Paragraph("Boletín académico del estudiante", estilo_subtitulo))
        story.append(Spacer(1, 0.18 * cm))

        datos_estudiante = Table(
            [
                [
                    Paragraph(f"<b>Estudiante:</b> {boletin['nombre']}", estilo_texto),
                    Paragraph(
                        f"<b>Documento:</b> {boletin['documento']}", estilo_texto
                    ),
                ],
                [
                    Paragraph(
                        f"<b>Grado:</b> {boletin['grado']}  <b>Curso:</b> {boletin['curso']}",
                        estilo_texto,
                    ),
                    Paragraph(
                        f"<b>Período:</b> {boletin['periodo']}  <b>Año:</b> {boletin['anio']}",
                        estilo_texto,
                    ),
                ],
                [
                    Paragraph(f"<b>Sede:</b> {boletin['sede']}", estilo_texto),
                    Paragraph(f"<b>Jornada:</b> {boletin['jornada']}", estilo_texto),
                ],
            ],
            colWidths=[9.3 * cm, 9.0 * cm],
        )
        datos_estudiante.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#9ca3af")),
                    ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(datos_estudiante)
        story.append(Spacer(1, 0.24 * cm))

        tabla_areas = [
            [
                Paragraph("<b>Área / Asignatura</b>", estilo_tabla),
                Paragraph("<b>I.H.</b>", estilo_tabla),
                Paragraph("<b>Inas.</b>", estilo_tabla),
                Paragraph("<b>Def.</b>", estilo_tabla),
                Paragraph("<b>Desempeño</b>", estilo_tabla),
                Paragraph("<b>Docente</b>", estilo_tabla),
                Paragraph("<b>Observación descriptiva</b>", estilo_tabla),
            ]
        ]
        for area in boletin.get("areas") or []:
            tabla_areas.append(
                [
                    Paragraph(str(area.get("area") or ""), estilo_tabla),
                    Paragraph(str(area.get("horas") or ""), estilo_tabla),
                    Paragraph(str(area.get("inasistencias") or 0), estilo_tabla),
                    Paragraph(str(area.get("definitiva_txt") or ""), estilo_tabla),
                    Paragraph(str(area.get("desempeno") or ""), estilo_tabla),
                    Paragraph(str(area.get("docente") or ""), estilo_tabla),
                    Paragraph(str(area.get("observacion") or ""), estilo_tabla),
                ]
            )
        tabla = Table(
            tabla_areas,
            repeatRows=1,
            colWidths=[
                4.1 * cm,
                1.0 * cm,
                1.1 * cm,
                1.0 * cm,
                2.0 * cm,
                3.2 * cm,
                6.0 * cm,
            ],
        )
        tabla.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5eef9")),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#9ca3af")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (1, 1), (4, -1), "CENTER"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(tabla)
        story.append(Spacer(1, 0.22 * cm))

        resumen = Table(
            [
                [
                    Paragraph(
                        f"<b>Promedio general:</b> {boletin.get('promedio_general_txt') or '-'}",
                        estilo_texto,
                    ),
                    Paragraph(
                        f"<b>Puesto en el curso:</b> {boletin.get('puesto') or '-'}",
                        estilo_texto,
                    ),
                    Paragraph(
                        "<b>Promedio del curso:</b> "
                        + (
                            "-"
                            if boletin.get("promedio_curso") is None
                            else f"{float(boletin['promedio_curso']):.2f}"
                        ),
                        estilo_texto,
                    ),
                ]
            ],
            colWidths=[6.2 * cm, 6.0 * cm, 6.1 * cm],
        )
        resumen.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#9ca3af")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(resumen)
        story.append(Spacer(1, 0.24 * cm))

        escala_filas = [
            [
                Paragraph("<b>Desde</b>", estilo_tabla),
                Paragraph("<b>Hasta</b>", estilo_tabla),
                Paragraph("<b>Letra</b>", estilo_tabla),
                Paragraph("<b>Desempeño</b>", estilo_tabla),
                Paragraph("<b>Orientación</b>", estilo_tabla),
            ]
        ]
        for fila in boletin.get("escala") or []:
            escala_filas.append(
                [
                    Paragraph(f"{float(fila.get('desde', 0)):.1f}", estilo_tabla),
                    Paragraph(f"{float(fila.get('hasta', 0)):.1f}", estilo_tabla),
                    Paragraph(str(fila.get("letra") or ""), estilo_tabla),
                    Paragraph(str(fila.get("desempeno") or ""), estilo_tabla),
                    Paragraph(str(fila.get("recomendacion") or ""), estilo_tabla),
                ]
            )
        tabla_escala = Table(
            escala_filas,
            repeatRows=1,
            colWidths=[1.6 * cm, 1.6 * cm, 1.3 * cm, 3.2 * cm, 11.0 * cm],
        )
        tabla_escala.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#9ca3af")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (0, 1), (2, -1), "CENTER"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(Paragraph("Escala de valoración", estilo_subtitulo))
        story.append(Spacer(1, 0.08 * cm))
        story.append(tabla_escala)
        story.append(Spacer(1, 0.55 * cm))

        firma_docente = str(boletin.get("firma_docente") or "Docente").strip()
        firma_rector = str(datos_globales.get("rector_nombre") or "Rectoría").strip()
        cargo_rector = str(datos_globales.get("rector_cargo") or "Rector").strip()
        firmas = Table(
            [
                [
                    Paragraph("_______________________________", estilo_texto),
                    Paragraph("_______________________________", estilo_texto),
                ],
                [
                    Paragraph(f"Docente: {firma_docente}", estilo_texto),
                    Paragraph(f"{cargo_rector}: {firma_rector}", estilo_texto),
                ],
                [
                    Paragraph("Firma docente", estilo_texto),
                    Paragraph("Firma rector", estilo_texto),
                ],
            ],
            colWidths=[9.1 * cm, 9.1 * cm],
        )
        firmas.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(firmas)
        doc.build(story)

    def _boletines_exportar_pdf_referencia(self):
        if not self._requiere_permiso("desktop.superadmin.boletines.exportar_pdf"):
            return
        if not _HAS_REPORTLAB:
            messagebox.showerror(
                "Boletines",
                "No es posible generar el PDF porque reportlab no está disponible.",
                parent=self.win,
            )
            return
        try:
            seleccion = []
            if hasattr(self, "tree_boletines"):
                seleccion = [
                    str(item or "").strip()
                    for item in self.tree_boletines.selection()
                    if str(item or "").strip()
                ]
            datos = self._boletines_obtener_dataset_pdf_individual(seleccion or None)
        except Exception as exc:
            messagebox.showwarning("Boletines", str(exc), parent=self.win)
            return

        boletines = list(datos.get("boletines") or [])
        if not boletines:
            return

        try:
            if len(boletines) == 1:
                boletin = boletines[0]
                archivo_sugerido = f"Boletin_{boletin['documento']}_{boletin['periodo'].replace(' ', '')}.pdf"
                ruta = filedialog.asksaveasfilename(
                    parent=self.win,
                    title="Guardar boletín PDF",
                    defaultextension=".pdf",
                    filetypes=[("PDF", "*.pdf")],
                    initialfile=archivo_sugerido,
                )
                if not ruta:
                    return
                self._boletines_generar_pdf_individual(ruta, datos, boletin)
                messagebox.showinfo(
                    "Boletines",
                    f"PDF generado correctamente en:\n{ruta}",
                    parent=self.win,
                )
                return

            carpeta = filedialog.askdirectory(
                parent=self.win,
                title="Seleccione la carpeta donde se guardarán los boletines",
            )
            if not carpeta:
                return
            generados = 0
            errores = []
            for boletin in boletines:
                nombre_archivo = f"Boletin_{boletin['grado']}_{boletin['curso']}_{boletin['documento']}_{boletin['periodo'].replace(' ', '')}.pdf"
                ruta_pdf = os.path.join(carpeta, nombre_archivo)
                try:
                    self._boletines_generar_pdf_individual(ruta_pdf, datos, boletin)
                    generados += 1
                except Exception as exc_item:
                    errores.append(f"{boletin['documento']}: {exc_item}")
            mensaje = f"Boletines generados: {generados}"
            if errores:
                mensaje += "\n\nErrores:\n" + "\n".join(errores[:10])
            messagebox.showinfo("Boletines", mensaje, parent=self.win)
        except Exception as exc:
            messagebox.showerror(
                "Boletines",
                f"No fue posible generar el PDF del boletín:\n{exc}",
                parent=self.win,
            )

    def _build_desempenos_tab(self):
        frame = self.tab_desempenos
        for widget in frame.winfo_children():
            widget.destroy()

        canvas = tk.Canvas(frame, bg="#f6f8fb", highlightthickness=0)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        scrollbar.pack(side="right", fill="y")
        canvas.configure(yscrollcommand=scrollbar.set)

        cuerpo = ttk.Frame(canvas, padding=(12, 10, 12, 12))
        ventana_canvas = canvas.create_window((0, 0), window=cuerpo, anchor="nw")
        cuerpo.bind(
            "<Configure>",
            lambda event: canvas.configure(scrollregion=canvas.bbox("all")),
        )
        canvas.bind(
            "<Configure>",
            lambda event: canvas.itemconfigure(ventana_canvas, width=event.width),
        )

        cuerpo.grid_columnconfigure(0, weight=1)
        cuerpo.grid_rowconfigure(3, weight=1)

        ttk.Label(
            cuerpo,
            text="Desempeños y observaciones pedagógicas",
            font=("Segoe UI", 14, "bold"),
        ).grid(row=0, column=0, sticky="w")
        ttk.Label(
            cuerpo,
            text="Cree plantillas por grado, área y período para alimentar observaciones consistentes en boletines y seguimiento académico.",
            style="CardHint.TLabel",
        ).grid(row=1, column=0, sticky="w", pady=(2, 10))

        filtros = ttk.LabelFrame(
            cuerpo,
            text="Contexto de la plantilla",
            padding=10,
            style="Card.TLabelframe",
        )
        filtros.grid(row=2, column=0, sticky="ew")
        for columna in (1, 3, 5, 7):
            filtros.grid_columnconfigure(columna, weight=1)

        self.desempenos_var_grado = tk.StringVar()
        self.desempenos_var_area = tk.StringVar()
        self.desempenos_var_periodo = tk.StringVar()
        self.desempenos_var_anio = tk.StringVar(value=self._obtener_anio_lectivo_base())

        ttk.Label(filtros, text="Grado:").grid(
            row=0, column=0, sticky="w", padx=(0, 6), pady=2
        )
        self.combo_desempenos_grado = ttk.Combobox(
            filtros,
            textvariable=self.desempenos_var_grado,
            state="readonly",
            width=18,
            values=self._listar_grados_catalogo(),
        )
        self.combo_desempenos_grado.grid(row=0, column=1, sticky="ew", pady=2)
        self.combo_desempenos_grado.bind(
            "<<ComboboxSelected>>", self._desempenos_actualizar_contexto_estado
        )

        ttk.Label(filtros, text="Área:").grid(
            row=0, column=2, sticky="w", padx=(12, 6), pady=2
        )
        self.combo_desempenos_area = ttk.Combobox(
            filtros,
            textvariable=self.desempenos_var_area,
            state="readonly",
            width=24,
            values=self._listar_areas_catalogo(),
        )
        self.combo_desempenos_area.grid(row=0, column=3, sticky="ew", pady=2)
        self.combo_desempenos_area.bind(
            "<<ComboboxSelected>>", self._desempenos_actualizar_contexto_estado
        )

        ttk.Label(filtros, text="Período:").grid(
            row=0, column=4, sticky="w", padx=(12, 6), pady=2
        )
        self.combo_desempenos_periodo = ttk.Combobox(
            filtros,
            textvariable=self.desempenos_var_periodo,
            state="readonly",
            width=14,
            values=self._listar_periodos_configurados(),
        )
        self.combo_desempenos_periodo.grid(row=0, column=5, sticky="ew", pady=2)
        self.combo_desempenos_periodo.bind(
            "<<ComboboxSelected>>", self._desempenos_actualizar_contexto_estado
        )

        ttk.Label(filtros, text="Año:").grid(
            row=0, column=6, sticky="w", padx=(12, 6), pady=2
        )
        ttk.Entry(filtros, textvariable=self.desempenos_var_anio, width=10).grid(
            row=0, column=7, sticky="ew", pady=2
        )

        acciones = ttk.Frame(filtros)
        acciones.grid(row=1, column=0, columnspan=8, sticky="ew", pady=(10, 0))
        self._crear_boton_si_permiso(
            ttk.Button,
            acciones,
            "desktop.superadmin.desempenos.cargar",
            text="Cargar plantilla",
            command=self._desempenos_cargar_plantilla,
            layout_kwargs={"side": "left"},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            acciones,
            "desktop.superadmin.desempenos.guardar",
            text="Guardar plantilla",
            command=self._desempenos_guardar_plantilla,
            layout_kwargs={"side": "left", "padx": (8, 0)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            acciones,
            "desktop.superadmin.desempenos.sugerir",
            text="Sugerir desde escala",
            command=self._desempenos_aplicar_sugerencias,
            layout_kwargs={"side": "left", "padx": (8, 0)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            acciones,
            "desktop.superadmin.desempenos.guardar",
            text="Agregar desempeño",
            command=self._desempenos_agregar_descripcion,
            layout_kwargs={"side": "left", "padx": (8, 0)},
        )
        ttk.Button(
            acciones, text="Limpiar editor", command=self._desempenos_limpiar_editor
        ).pack(side="left", padx=(8, 0))
        self._crear_boton_si_permiso(
            ttk.Button,
            acciones,
            "desktop.superadmin.desempenos.eliminar",
            text="Eliminar plantilla",
            command=self._desempenos_eliminar_plantilla,
            layout_kwargs={"side": "left", "padx": (8, 0)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            acciones,
            "desktop.superadmin.desempenos.actualizar",
            text="Actualizar catálogos",
            command=self._desempenos_refrescar_catalogos,
            layout_kwargs={"side": "left", "padx": (8, 0)},
        )

        self.lbl_desempenos_estado = ttk.Label(
            filtros,
            text="Sin plantilla cargada. Puede generar una base inicial a partir de la escala institucional.",
            style="CardHint.TLabel",
        )
        self.lbl_desempenos_estado.grid(
            row=2, column=0, columnspan=8, sticky="w", pady=(8, 0)
        )

        editor = ttk.Frame(cuerpo)
        editor.grid(row=3, column=0, sticky="nsew", pady=(10, 0))
        editor.grid_columnconfigure(0, weight=1)

        self.desempenos_editor = editor
        self.desempenos_items = []
        self._desempenos_reiniciar_items_base()

        listado = ttk.LabelFrame(
            cuerpo, text="Plantillas guardadas", padding=10, style="Card.TLabelframe"
        )
        listado.grid(row=4, column=0, sticky="nsew")
        listado.grid_columnconfigure(0, weight=1)
        listado.grid_rowconfigure(0, weight=1)

        columnas = ("grado", "area", "periodo", "anio", "niveles", "actualizacion")
        self.tree_desempenos = ttk.Treeview(
            listado, columns=columnas, show="headings", height=10
        )
        titulos = {
            "grado": "Grado",
            "area": "Área",
            "periodo": "Período",
            "anio": "Año",
            "niveles": "Desempeños",
            "actualizacion": "Actualización",
        }
        anchos = {
            "grado": 90,
            "area": 190,
            "periodo": 90,
            "anio": 70,
            "niveles": 70,
            "actualizacion": 150,
        }
        for columna in columnas:
            self.tree_desempenos.heading(columna, text=titulos[columna])
            self.tree_desempenos.column(
                columna,
                width=anchos[columna],
                anchor="center" if columna != "area" else "w",
            )
        self.tree_desempenos.grid(row=0, column=0, sticky="nsew")
        scrollbar_desempenos = ttk.Scrollbar(
            listado, orient="vertical", command=self.tree_desempenos.yview
        )
        scrollbar_desempenos.grid(row=0, column=1, sticky="ns")
        self.tree_desempenos.configure(yscrollcommand=scrollbar_desempenos.set)
        self.tree_desempenos.bind(
            "<<TreeviewSelect>>", self._desempenos_cargar_desde_listado
        )

        grados = self._listar_grados_catalogo()
        areas = self._listar_areas_catalogo()
        periodos = self._listar_periodos_configurados()
        if grados:
            self.desempenos_var_grado.set(grados[0])
        if areas:
            self.desempenos_var_area.set(areas[0])
        if periodos:
            self.desempenos_var_periodo.set(periodos[0])
        self._desempenos_refrescar_listado()
        self._desempenos_actualizar_contexto_estado()

    def _desempenos_refrescar_catalogos(self):
        if not self._tiene_permiso("desktop.superadmin.desempenos.actualizar"):
            return
        if not hasattr(self, "combo_desempenos_grado"):
            return
        grado_actual = str(self.desempenos_var_grado.get() or "").strip()
        area_actual = str(self.desempenos_var_area.get() or "").strip()
        periodo_actual = str(self.desempenos_var_periodo.get() or "").strip()

        grados = self._listar_grados_catalogo()
        areas = self._listar_areas_catalogo()
        periodos = self._listar_periodos_configurados()

        self.combo_desempenos_grado.configure(values=grados)
        self.combo_desempenos_area.configure(values=areas)
        self.combo_desempenos_periodo.configure(values=periodos)

        if grados:
            self.desempenos_var_grado.set(
                grado_actual if grado_actual in grados else grados[0]
            )
        else:
            self.desempenos_var_grado.set("")

        if areas:
            self.desempenos_var_area.set(
                area_actual if area_actual in areas else areas[0]
            )
        else:
            self.desempenos_var_area.set("")

        if periodos:
            self.desempenos_var_periodo.set(
                periodo_actual if periodo_actual in periodos else periodos[0]
            )
        else:
            self.desempenos_var_periodo.set("")

        self._desempenos_refrescar_listado()
        self._desempenos_actualizar_contexto_estado()

    def _desempenos_actualizar_contexto_estado(self, event=None):
        if not hasattr(self, "lbl_desempenos_estado"):
            return
        contexto = self._desempenos_obtener_contexto(validar=False)
        if not contexto:
            self.lbl_desempenos_estado.config(
                text="Seleccione grado, área, período y año para editar o consultar una plantilla."
            )
            return
        grado, area, periodo, anio = contexto
        if not grado or not area or not periodo:
            self.lbl_desempenos_estado.config(
                text="Seleccione grado, área, período y año para editar o consultar una plantilla."
            )
            return
        try:
            fila = self.cur.execute(
                "SELECT COUNT(*), MAX(fecha_actualizacion) FROM desempenos_plantilla WHERE grado = ? AND area = ? AND periodo = ? AND anio_lectivo = ?",
                (grado, area, periodo, anio),
            ).fetchone()
        except Exception:
            fila = (0, "")
        total = self._coerce_int((fila or [0])[0], 0)
        fecha = str((fila or [0, ""])[1] or "").strip()
        if total > 0:
            self.lbl_desempenos_estado.config(
                text=f"Existe una plantilla guardada para {grado} | {area} | {periodo} | {anio}. Última actualización: {fecha or 'sin fecha registrada'}."
            )
        else:
            self.lbl_desempenos_estado.config(
                text=f"No existe plantilla para {grado} | {area} | {periodo} | {anio}. Puede crearla desde cero o usar 'Sugerir desde escala'."
            )

    def _desempenos_obtener_contexto(self, validar=True):
        grado = (
            str(self.desempenos_var_grado.get() or "").strip()
            if hasattr(self, "desempenos_var_grado")
            else ""
        )
        area = (
            str(self.desempenos_var_area.get() or "").strip()
            if hasattr(self, "desempenos_var_area")
            else ""
        )
        periodo = (
            str(self.desempenos_var_periodo.get() or "").strip()
            if hasattr(self, "desempenos_var_periodo")
            else ""
        )
        anio = (
            str(self.desempenos_var_anio.get() or "").strip()
            if hasattr(self, "desempenos_var_anio")
            else self._obtener_anio_lectivo_base()
        )
        if validar and (not grado or not area or not periodo):
            messagebox.showwarning(
                "Contexto incompleto",
                "Seleccione grado, área y período antes de continuar.",
                parent=self.win,
            )
            return None
        return grado, area, periodo, anio or self._obtener_anio_lectivo_base()

    def _desempenos_limpiar_editor(self):
        if not hasattr(self, "desempenos_items"):
            return
        for item in self.desempenos_items:
            item["descripcion"].delete("1.0", "end")
        if hasattr(self, "lbl_desempenos_estado"):
            self.lbl_desempenos_estado.config(
                text="Editor limpio. Puede cargar una plantilla existente o generar una sugerencia nueva."
            )

    def _desempenos_aplicar_sugerencias(self):
        if not self._requiere_permiso("desktop.superadmin.desempenos.sugerir"):
            return
        contexto = self._desempenos_obtener_contexto(validar=True)
        if not contexto:
            return
        grado, area, periodo, _anio = contexto
        for item in self.desempenos_items:
            nivel = item["nivel"]
            descripcion = item["descripcion"]
            descripcion.delete("1.0", "end")
            descripcion.insert(
                "1.0",
                f"El estudiante evidencia un desempeño {nivel.lower()} en {area} durante {periodo} del grado {grado}.",
            )
        self.lbl_desempenos_estado.config(
            text=f"Se generó una base inicial para {area} en {grado}, {periodo}. Revise y ajuste los textos antes de guardar."
        )

    def _desempenos_cargar_plantilla(self):
        if not self._requiere_permiso("desktop.superadmin.desempenos.cargar"):
            return
        contexto = self._desempenos_obtener_contexto(validar=True)
        if not contexto:
            return
        grado, area, periodo, anio = contexto
        self._desempenos_reiniciar_items_base()
        filas = self.cur.execute(
            "SELECT nivel, descriptor, fecha_actualizacion "
            "FROM desempenos_plantilla WHERE grado = ? AND area = ? AND periodo = ? AND anio_lectivo = ?",
            (grado, area, periodo, anio),
        ).fetchall()
        if not filas:
            self.lbl_desempenos_estado.config(
                text=f"No existe plantilla guardada para {grado} | {area} | {periodo} | {anio}."
            )
            return
        ultima_actualizacion = ""
        niveles_base = {item["nivel"] for item in self.desempenos_items}
        for nivel, descriptor, fecha in filas:
            nivel_txt = str(nivel or "").strip()
            item = self._desempenos_obtener_item(nivel_txt)
            if not item:
                item = self._desempenos_crear_item(
                    nivel_txt, removable=nivel_txt not in niveles_base
                )
            item["descripcion"].insert("1.0", str(descriptor or ""))
            ultima_actualizacion = str(fecha or ultima_actualizacion)
        self.lbl_desempenos_estado.config(
            text=f"Plantilla cargada para {grado} | {area} | {periodo} | {anio}. Última actualización: {ultima_actualizacion or 'sin fecha registrada'}."
        )

    def _desempenos_guardar_plantilla(self):
        if not self._requiere_permiso("desktop.superadmin.desempenos.guardar"):
            return
        contexto = self._desempenos_obtener_contexto(validar=True)
        if not contexto:
            return
        grado, area, periodo, anio = contexto
        filas = []
        marca_tiempo = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for item in self.desempenos_items:
            nivel = item["nivel"]
            descriptor = item["descripcion"].get("1.0", "end").strip()
            if descriptor:
                filas.append(
                    (
                        grado,
                        area,
                        periodo,
                        anio,
                        nivel,
                        descriptor,
                        "",
                        "",
                        marca_tiempo,
                    )
                )
        if not filas:
            messagebox.showwarning(
                "Sin contenido",
                "Debe diligenciar al menos una descripción del desempeño antes de guardar la plantilla.",
                parent=self.win,
            )
            return
        self.cur.execute(
            "DELETE FROM desempenos_plantilla WHERE grado = ? AND area = ? AND periodo = ? AND anio_lectivo = ?",
            (grado, area, periodo, anio),
        )
        self.cur.executemany(
            "INSERT INTO desempenos_plantilla (grado, area, periodo, anio_lectivo, nivel, descriptor, actividad_complementaria, dificultad, fecha_actualizacion) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            filas,
        )
        self.conn.commit()
        self._desempenos_refrescar_catalogos()
        self._desempenos_refrescar_listado()
        self.lbl_desempenos_estado.config(
            text=f"Plantilla guardada para {grado} | {area} | {periodo} | {anio}."
        )
        messagebox.showinfo(
            "Desempeños",
            "La plantilla de desempeños fue guardada correctamente.",
            parent=self.win,
        )

    def _desempenos_eliminar_plantilla(self):
        if not self._requiere_permiso("desktop.superadmin.desempenos.eliminar"):
            return
        contexto = self._desempenos_obtener_contexto(validar=True)
        if not contexto:
            return
        grado, area, periodo, anio = contexto
        if not messagebox.askyesno(
            "Eliminar plantilla",
            f"Se eliminará la plantilla de desempeños para {grado}, {area}, {periodo}, {anio}.\n\n¿Desea continuar?",
            parent=self.win,
        ):
            return
        self.cur.execute(
            "DELETE FROM desempenos_plantilla WHERE grado = ? AND area = ? AND periodo = ? AND anio_lectivo = ?",
            (grado, area, periodo, anio),
        )
        self.conn.commit()
        self._desempenos_reiniciar_items_base()
        self._desempenos_limpiar_editor()
        self._desempenos_refrescar_catalogos()
        self._desempenos_refrescar_listado()
        self.lbl_desempenos_estado.config(
            text=f"Se eliminó la plantilla para {grado} | {area} | {periodo} | {anio}."
        )

    def _desempenos_refrescar_listado(self):
        if not hasattr(self, "tree_desempenos"):
            return
        for item in self.tree_desempenos.get_children():
            self.tree_desempenos.delete(item)
        try:
            filas = self.cur.execute(
                "SELECT grado, area, periodo, anio_lectivo, COUNT(*), MAX(fecha_actualizacion) "
                "FROM desempenos_plantilla "
                "GROUP BY grado, area, periodo, anio_lectivo "
                "ORDER BY anio_lectivo DESC, grado, area, periodo"
            ).fetchall()
        except Exception:
            filas = []
        for fila in filas:
            grado, area, periodo, anio, niveles, actualizacion = fila
            iid = "|".join(
                [str(grado or ""), str(area or ""), str(periodo or ""), str(anio or "")]
            )
            self.tree_desempenos.insert(
                "",
                "end",
                iid=iid,
                values=(grado, area, periodo, anio, niveles, actualizacion),
            )

    def _desempenos_cargar_desde_listado(self, event=None):
        if not self._tiene_permiso("desktop.superadmin.desempenos.cargar"):
            return
        if not hasattr(self, "tree_desempenos"):
            return
        seleccion = self.tree_desempenos.selection()
        if not seleccion:
            return
        grado, area, periodo, anio = self.tree_desempenos.item(seleccion[0], "values")[
            :4
        ]
        self._desempenos_refrescar_catalogos()
        self.desempenos_var_grado.set(str(grado or ""))
        self.desempenos_var_area.set(str(area or ""))
        self.desempenos_var_periodo.set(str(periodo or ""))
        self.desempenos_var_anio.set(str(anio or ""))
        self._desempenos_cargar_plantilla()

    def _desempenos_reiniciar_items_base(self):
        if not hasattr(self, "desempenos_editor"):
            return
        for child in self.desempenos_editor.winfo_children():
            child.destroy()
        self.desempenos_items = []
        for nombre_nivel in self._desempenos_niveles_base():
            self._desempenos_crear_item(nombre_nivel, removable=False)

    def _desempenos_niveles_base(self):
        return ["Superior", "Alto", "Básico", "Bajo"]

    def _desempenos_obtener_item(self, nivel):
        nivel_txt = str(nivel or "").strip()
        for item in getattr(self, "desempenos_items", []):
            if item["nivel"] == nivel_txt:
                return item
        return None

    def _desempenos_crear_item(self, nivel, removable=True):
        if not hasattr(self, "desempenos_editor"):
            return None
        nivel_txt = (
            str(nivel or "").strip() or f"Desempeño {len(self.desempenos_items) + 1}"
        )
        tarjeta = ttk.LabelFrame(
            self.desempenos_editor,
            text=nivel_txt,
            padding=10,
            style="Card.TLabelframe",
        )
        tarjeta.grid(
            row=len(self.desempenos_items),
            column=0,
            sticky="ew",
            pady=(0, 10),
        )
        tarjeta.grid_columnconfigure(0, weight=1)

        ttk.Label(tarjeta, text="Descripción del desempeño:").grid(
            row=0, column=0, sticky="w"
        )
        txt_descripcion = tk.Text(tarjeta, height=5, wrap="word", font=("Segoe UI", 9))
        txt_descripcion.grid(row=1, column=0, sticky="ew", pady=(2, 0))

        item = {
            "nivel": nivel_txt,
            "frame": tarjeta,
            "descripcion": txt_descripcion,
        }
        self.desempenos_items.append(item)

        if removable:
            ttk.Button(
                tarjeta,
                text="Eliminar",
                command=lambda ref=item: self._desempenos_eliminar_item(ref),
            ).grid(row=0, column=1, padx=(8, 0), sticky="e")

        return item

    def _desempenos_eliminar_item(self, item):
        if item not in getattr(self, "desempenos_items", []):
            return
        try:
            item["frame"].destroy()
        except Exception:
            pass
        self.desempenos_items = [
            actual for actual in self.desempenos_items if actual is not item
        ]
        for indice, actual in enumerate(self.desempenos_items):
            actual["frame"].grid_configure(row=indice)

    def _desempenos_agregar_descripcion(self):
        existentes = {
            str(item.get("nivel") or "").strip().casefold()
            for item in getattr(self, "desempenos_items", [])
        }
        agregados = 0
        for nombre in self._desempenos_niveles_base():
            if nombre.casefold() in existentes:
                continue
            self._desempenos_crear_item(nombre, removable=False)
            agregados += 1

        if agregados == 0:
            messagebox.showinfo(
                "Desempeños",
                "La plantilla ya tiene los niveles base: Superior, Alto, Básico y Bajo.",
                parent=self.win,
            )
            return

        if hasattr(self, "lbl_desempenos_estado"):
            self.lbl_desempenos_estado.config(
                text="Se agregaron automáticamente los desempeños base: Superior, Alto, Básico y Bajo."
            )

    def _configurar_sincronizacion_preguntas(self):
        if getattr(self, "_preguntas_sync_configurada", False):
            return

        try:
            self.win.bind(
                "<<PreguntasActualizadas>>",
                self._on_preguntas_actualizadas,
                add="+",
            )
        except Exception:
            pass

        try:
            self.nb.bind(
                "<<NotebookTabChanged>>",
                self._on_superadmin_tab_changed,
                add="+",
            )
        except Exception:
            pass

        self._preguntas_sync_configurada = True

    def _emit_preguntas_actualizadas(self):
        try:
            if hasattr(self, "win") and self.win.winfo_exists():
                self.win.event_generate("<<PreguntasActualizadas>>", when="tail")
        except Exception:
            try:
                self.recargar_datos_generador_examenes()
            except Exception:
                pass

    def _on_preguntas_actualizadas(self, event=None):
        self.recargar_datos_generador_examenes()

    def _on_superadmin_tab_changed(self, event=None):
        try:
            if not hasattr(self, "nb") or not hasattr(self, "tab_examenes"):
                return
            if str(self.nb.select()) != str(self.tab_examenes):
                return
        except Exception:
            return

        self.recargar_datos_generador_examenes()

    # ---------- Docentes ----------
    def _build_docentes_tab(self):
        frame = self.tab_docentes
        toolbar = ttk.Frame(frame)
        toolbar.pack(fill="x", padx=8, pady=(8, 4))

        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.docentes.crear",
            text="Registrar Personal",
            command=self.docente_agregar,
            layout_kwargs={"side": "left"},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.docentes.editar",
            text="Editar Seleccionado",
            command=self.docente_editar,
            layout_kwargs={"side": "left", "padx": (6, 0)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.docentes.carga_academica",
            text="Carga Académica",
            command=self.docente_carga_academica,
            layout_kwargs={"side": "left", "padx": (6, 0)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.docentes.eliminar",
            text="Eliminar Seleccionado",
            command=self.docente_eliminar,
            layout_kwargs={"side": "left", "padx": (6, 0)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.docentes.recargar",
            text="Recargar",
            command=self._load_docentes,
            layout_kwargs={"side": "left", "padx": (6, 0)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.docentes.importar",
            text="Importar Excel",
            command=self.docentes_importar_excel,
            layout_kwargs={"side": "left", "padx": (6, 0)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.docentes.exportar",
            text="Exportar Excel",
            command=self.docentes_exportar_excel,
            layout_kwargs={"side": "left", "padx": (6, 0)},
        )

        cols = ("documento", "nombre", "cargo", "jornada", "estado", "acciones")
        self.tree_doc = ttk.Treeview(frame, columns=cols, show="headings")

        headers = {
            "documento": "Documento",
            "nombre": "Nombre",
            "cargo": "Cargo",
            "jornada": "Jornada",
            "estado": "Estado",
            "acciones": "Acciones",
        }
        widths = {
            "documento": 130,
            "nombre": 230,
            "cargo": 130,
            "jornada": 100,
            "estado": 100,
            "acciones": 260,
        }
        for c in cols:
            self.tree_doc.heading(c, text=headers[c])
            self.tree_doc.column(
                c, width=widths[c], anchor="w", stretch=(c == "nombre")
            )

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree_doc.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.tree_doc.xview)
        self.tree_doc.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree_doc.pack(fill="both", expand=True, padx=8, pady=(0, 2))
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        self.tree_doc.bind("<Button-1>", self._docente_tree_click_accion, add="+")

        self._load_docentes()

    def _load_docentes(self):
        # poblar tree
        for i in self.tree_doc.get_children():
            self.tree_doc.delete(i)

        rows = []
        try:
            data = core_docentes.listar_docentes(limit=10000, offset=0)
            for d in data.get("docentes", []):
                rows.append(
                    (
                        d.get("documento", ""),
                        d.get("nombre", ""),
                        d.get("cargo", ""),
                        d.get("jornada", ""),
                        d.get("estado", ""),
                    )
                )
        except Exception:
            rows = []
        self.docentes_df = rows

        for documento, nombre, cargo, jornada, estado in rows:
            self.tree_doc.insert(
                "",
                "end",
                values=(
                    documento or "",
                    nombre or "",
                    cargo or "",
                    jornada or "",
                    estado or "",
                    "Editar | Carga académica | Eliminar",
                ),
            )

    def _save_docentes(self):
        # La persistencia en personal se realiza directamente con SQLite.
        return

    def _docente_seleccionado(self, item_id=None, mostrar_msg=True):
        item = item_id
        if not item:
            sel = self.tree_doc.selection()
            if sel:
                item = sel[0]
        if not item:
            if mostrar_msg:
                messagebox.showinfo(
                    "Personal", "Seleccione un miembro del personal.", parent=self.win
                )
            return None, None

        vals = self.tree_doc.item(item).get("values", [])
        if not vals:
            if mostrar_msg:
                messagebox.showinfo(
                    "Personal",
                    "Seleccione un miembro del personal válido.",
                    parent=self.win,
                )
            return None, None
        documento = str(vals[0]).strip()
        if not documento:
            if mostrar_msg:
                messagebox.showinfo(
                    "Personal",
                    "No se encontró documento del registro.",
                    parent=self.win,
                )
            return None, None
        return item, documento

    def _obtener_docente_por_documento(self, documento):
        return core_docentes.buscar_docente(documento)

    def _documento_docente_existe(self, documento, excluir_documento=None):
        return core_docentes.documento_docente_existe(documento, excluir_documento)

    def docente_agregar(self):
        if not self._requiere_permiso("desktop.superadmin.docentes.crear"):
            return
        datos = self._dialog_docente()
        if not datos:
            return

        try:
            core_docentes.crear_o_actualizar_docente(
                tipo_documento=datos.get("tipo_documento"),
                documento=datos.get("documento"),
                nombre=datos.get("nombre"),
                sexo=datos.get("sexo", ""),
                fecha_nacimiento=datos.get("fecha_nacimiento", ""),
                telefono=datos.get("telefono", ""),
                correo=datos.get("correo", ""),
                cargo=datos.get("cargo", "Docente"),
                jornada=datos.get("jornada", "Mañana"),
                sede=datos.get("sede", ""),
                estado=datos.get("estado", "Activo"),
            )
        except ValueError as exc:
            messagebox.showerror("Validación", str(exc), parent=self.win)
            return
        except Exception as exc:
            messagebox.showerror(
                "Personal",
                f"No se pudo registrar miembro del personal: {exc}",
                parent=self.win,
            )
            return

        self._load_docentes()
        messagebox.showinfo(
            "Registro",
            "Docente registrado correctamente",
            parent=self.win,
        )

    def docente_editar(self, item_id=None):
        if not self._requiere_permiso("desktop.superadmin.docentes.editar"):
            return
        item, documento = self._docente_seleccionado(item_id=item_id)
        if not documento:
            return

        actual = self._obtener_docente_por_documento(documento)
        if not actual:
            messagebox.showerror(
                "Personal",
                "No se encontró el docente seleccionado.",
                parent=self.win,
            )
            return

        nuevos = self._dialog_docente(actual)
        if not nuevos:
            return

        try:
            actualizado = core_docentes.crear_o_actualizar_docente(
                tipo_documento=nuevos.get("tipo_documento"),
                documento=nuevos.get("documento"),
                nombre=nuevos.get("nombre"),
                sexo=nuevos.get("sexo", ""),
                fecha_nacimiento=nuevos.get("fecha_nacimiento", ""),
                telefono=nuevos.get("telefono", ""),
                correo=nuevos.get("correo", ""),
                cargo=nuevos.get("cargo", "Docente"),
                jornada=nuevos.get("jornada", "Mañana"),
                sede=nuevos.get("sede", ""),
                estado=nuevos.get("estado", "Activo"),
                documento_original=documento,
            )
            if not actualizado:
                messagebox.showerror(
                    "Docentes",
                    "No se encontró el docente seleccionado.",
                    parent=self.win,
                )
                return
        except ValueError as exc:
            messagebox.showerror("Validación", str(exc), parent=self.win)
            return
        except Exception as exc:
            messagebox.showerror(
                "Personal",
                f"No se pudo actualizar miembro del personal: {exc}",
                parent=self.win,
            )
            return

        self._load_docentes()
        messagebox.showinfo(
            "Personal",
            "Docente actualizado correctamente.",
            parent=self.win,
        )

    def docente_eliminar(self, item_id=None):
        if not self._requiere_permiso("desktop.superadmin.docentes.eliminar"):
            return
        item, documento = self._docente_seleccionado(item_id=item_id)
        if not documento:
            return

        vals = self.tree_doc.item(item).get("values", [])
        nombre = vals[1] if len(vals) > 1 else documento
        if not messagebox.askyesno(
            "Confirmar",
            f"Eliminar docente {nombre}?",
            parent=self.win,
        ):
            return

        try:
            core_docentes.eliminar_docente(documento)
        except Exception as exc:
            messagebox.showerror(
                "Personal",
                f"No se pudo eliminar miembro del personal: {exc}",
                parent=self.win,
            )
            return

        self._load_docentes()

    def docente_carga_academica(self, item_id=None):
        if not self._requiere_permiso("desktop.superadmin.docentes.carga_academica"):
            return
        _, documento = self._docente_seleccionado(item_id=item_id)
        if not documento:
            return
        self._carga_refresh_filtros()

        label_docente = self.ca_docente_doc_to_label.get(documento)
        if label_docente:
            self.ca_filter_doc.set(label_docente)
        else:
            self.ca_filter_doc.set("Todos")

        self._load_carga_academica()
        try:
            self.nb.select(self.tab_carga_academica)
        except Exception:
            pass

    # ---------- Carga académica ----------
    def _build_carga_academica_tab(self):
        self._ca_asegurar_esquema_horas()
        frame = self.tab_carga_academica

        filtros = ttk.Frame(frame)
        filtros.pack(fill="x", padx=8, pady=(8, 4))

        ttk.Label(filtros, text="Docente:").pack(side="left")
        self.ca_filter_doc = ttk.Combobox(filtros, state="readonly", width=36)
        self.ca_filter_doc.pack(side="left", padx=(4, 8))

        ttk.Label(filtros, text="Área:").pack(side="left")
        self.ca_filter_area = ttk.Combobox(filtros, state="readonly", width=18)
        self.ca_filter_area.pack(side="left", padx=(4, 8))

        ttk.Label(filtros, text="Grado:").pack(side="left")
        self.ca_filter_grado = ttk.Combobox(filtros, state="readonly", width=12)
        self.ca_filter_grado.pack(side="left", padx=(4, 8))

        ttk.Label(filtros, text="Curso:").pack(side="left")
        self.ca_filter_curso = ttk.Combobox(filtros, state="readonly", width=12)
        self.ca_filter_curso.pack(side="left", padx=(4, 8))

        self.ca_resumen_horas_var = tk.StringVar(
            value=(
                "Docente: N/D | Horas asignadas: 0 | Horas normales permitidas: 22 | "
                "Horas extras permitidas: 0 | Horas disponibles: 22"
            )
        )
        ttk.Label(frame, textvariable=self.ca_resumen_horas_var).pack(
            fill="x", padx=10, pady=(0, 6)
        )

        acciones = ttk.Frame(frame)
        acciones.pack(fill="x", padx=8, pady=(0, 6))

        self._crear_boton_si_permiso(
            ttk.Button,
            acciones,
            "desktop.superadmin.carga_academica.crear",
            text="Agregar",
            command=self.carga_academica_agregar,
            layout_kwargs={"side": "left"},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            acciones,
            "desktop.superadmin.carga_academica.editar",
            text="Editar",
            command=self.carga_academica_editar,
            layout_kwargs={"side": "left", "padx": (6, 0)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            acciones,
            "desktop.superadmin.carga_academica.eliminar",
            text="Eliminar",
            command=self.carga_academica_eliminar,
            layout_kwargs={"side": "left", "padx": (6, 0)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            acciones,
            "desktop.superadmin.carga_academica.actualizar",
            text="Actualizar",
            command=self.carga_academica_actualizar,
            layout_kwargs={"side": "left", "padx": (6, 0)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            acciones,
            "desktop.superadmin.carga_academica.horas",
            text="Configurar Horas Docente",
            command=self.configuracion_configurar_horas_docente,
            layout_kwargs={"side": "left", "padx": (6, 0)},
        )

        cols = (
            "docente",
            "director_grupo",
            "area",
            "grado",
            "curso",
            "horas_asignadas",
            "horas_extras",
            "anio",
            "estado",
            "acciones",
        )
        self.tree_carga = ttk.Treeview(frame, columns=cols, show="headings")

        headers = {
            "docente": "Docente",
            "director_grupo": "Director de grupo",
            "area": "Área",
            "grado": "Grado",
            "curso": "Curso",
            "horas_asignadas": "Horas Asignadas",
            "horas_extras": "Horas Extras",
            "anio": "Año lectivo",
            "estado": "Estado",
            "acciones": "Acciones",
        }
        widths = {
            "docente": 290,
            "director_grupo": 240,
            "area": 160,
            "grado": 90,
            "curso": 90,
            "horas_asignadas": 125,
            "horas_extras": 115,
            "anio": 110,
            "estado": 100,
            "acciones": 180,
        }
        for c in cols:
            self.tree_carga.heading(c, text=headers[c])
            self.tree_carga.column(
                c,
                width=widths[c],
                anchor="w",
                stretch=(c in ("docente", "director_grupo")),
            )

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree_carga.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.tree_carga.xview)
        self.tree_carga.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree_carga.pack(fill="both", expand=True, padx=8, pady=(0, 2))
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        self.ca_filter_doc.bind("<<ComboboxSelected>>", self._carga_on_docente_filtro)
        self.ca_filter_area.bind(
            "<<ComboboxSelected>>", lambda e: self._load_carga_academica()
        )
        self.ca_filter_grado.bind("<<ComboboxSelected>>", self._carga_on_grado_filtro)
        self.ca_filter_curso.bind(
            "<<ComboboxSelected>>", lambda e: self._load_carga_academica()
        )
        self.tree_carga.bind("<Button-1>", self._carga_tree_click_accion, add="+")

        self._carga_refresh_filtros()
        self._load_carga_academica()

    def _ca_todos_docentes(self, solo_activos=False):
        return core_docentes.listar_docentes_selector(solo_activos=solo_activos)

    def _ca_todas_areas(self):
        """Áreas desde Plan de Estudios (tabla areas); sin dependencia del banco de preguntas."""
        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            # Primero: áreas que tienen al menos un registro activo en plan_estudio
            cur.execute(
                """
                SELECT DISTINCT COALESCE(a.nombre, p.area)
                FROM plan_estudio p
                LEFT JOIN areas a ON a.id = p.IdArea
                WHERE p.estado = 1
                  AND TRIM(COALESCE(a.nombre, p.area, '')) <> ''
                ORDER BY COALESCE(a.nombre, p.area)
                """
            )
            areas = [r[0] for r in cur.fetchall() if r and r[0]]
            conn.close()
            if areas:
                return areas
            # Fallback: catálogo completo de áreas activas
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute("SELECT nombre FROM areas WHERE estado = 1 ORDER BY nombre")
            areas = [r[0] for r in cur.fetchall() if r and r[0]]
            conn.close()
            return areas
        except Exception:
            return []

    def _ca_asegurar_esquema_horas(self):
        """Garantiza columnas de horas en carga_academica para bases antiguas."""
        try:
            core_docentes.asegurar_esquema_carga_academica()
        except Exception:
            pass

    def _ca_obtener_limites_docente(self, docente_documento):
        """Devuelve límites configurados para un docente (normales, extras)."""
        try:
            return core_docentes.obtener_limites_docente(docente_documento)
        except Exception:
            return (22, 0, False)

    def _ca_horas_totales_docente(self, docente_documento, excluir_id=None):
        """Suma horas activas asignadas actualmente a un docente."""
        try:
            return core_docentes.horas_totales_docente(docente_documento, excluir_id)
        except Exception:
            return 0

    def _ca_clave_orden_grado(self, grado):
        txt = str(grado or "").strip().upper()
        if not txt:
            return (99, 999, "")

        if txt in {"0", "TRANSICION", "TRANSICIÓN", "TRANSICION PREESCOLAR"}:
            return (0, 0, txt)
        if txt in {"JA", "JARDIN", "JARDÍN"}:
            return (0, 1, txt)
        if txt in {"PREJ", "PREJARDIN", "PREJARDÍN"}:
            return (0, 2, txt)

        m_num = re.fullmatch(r"(\d+)(?:°)?", txt)
        if m_num:
            return (1, int(m_num.group(1)), txt)

        m_clei = re.fullmatch(r"C\s*([1-6])", txt)
        if m_clei:
            return (2, int(m_clei.group(1)), txt)

        return (98, 999, txt)

    def _ca_ordenar_grados(self, grados):
        vistos = set()
        limpios = []
        for g in grados or []:
            gs = str(g or "").strip()
            if not gs:
                continue
            clave_visto = gs.upper()
            if clave_visto in vistos:
                continue
            vistos.add(clave_visto)
            limpios.append(gs)
        limpios.sort(key=self._ca_clave_orden_grado)
        return limpios

    def _ca_plan_areas_horas(self, grado, curso):
        """Devuelve áreas y horas del plan para un grado/curso."""
        try:
            grado_val = self._plan_normalizar_grado(grado)
            curso_val = self._plan_normalizar_curso(curso)
        except Exception:
            return []

        conn = None
        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            try:
                cur.execute(
                    """
                    SELECT
                        COALESCE(a.nombre, p.area) AS area_nombre,
                        CAST(COALESCE(p.horas, 0) AS INTEGER) AS horas_plan
                    FROM plan_estudio p
                    LEFT JOIN areas a ON a.id = p.IdArea
                    WHERE p.grado = ? AND p.curso = ? AND p.estado = 1
                      AND TRIM(COALESCE(a.nombre, p.area, '')) <> ''
                    ORDER BY COALESCE(a.nombre, p.area)
                    """,
                    (grado_val, curso_val),
                )
            except Exception:
                cur.execute(
                    """
                    SELECT
                        TRIM(COALESCE(p.area, '')) AS area_nombre,
                        CAST(COALESCE(p.horas, 0) AS INTEGER) AS horas_plan
                    FROM plan_estudio p
                    WHERE p.grado = ? AND p.curso = ? AND p.estado = 1
                      AND TRIM(COALESCE(p.area, '')) <> ''
                    ORDER BY TRIM(COALESCE(p.area, ''))
                    """,
                    (grado_val, curso_val),
                )

            rows = cur.fetchall()
            agregado = {}
            for area, horas in rows:
                area_txt = str(area or "").strip()
                if not area_txt:
                    continue
                horas_int = int(horas or 0)
                clave = area_txt.lower()
                if clave not in agregado:
                    agregado[clave] = [area_txt, max(0, horas_int)]
                else:
                    agregado[clave][1] = max(agregado[clave][1], max(0, horas_int))
            out = list(agregado.values())
            out.sort(key=lambda x: str(x[0]).lower())
            return [(a, h) for a, h in out]
        except Exception:
            return []
        finally:
            try:
                if conn is not None:
                    conn.close()
            except Exception:
                pass

    def _ca_todos_grados(self):
        """Grados desde Plan de Estudios; fallback a tabla estudiantes. Sin dependencia del banco de preguntas."""
        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute(
                """
                SELECT DISTINCT TRIM(CAST(grado AS TEXT))
                FROM plan_estudio
                WHERE grado IS NOT NULL AND TRIM(CAST(grado AS TEXT)) <> ''
                                    AND estado = 1
                """
            )
            grados = [str(r[0]).strip() for r in cur.fetchall() if r and r[0]]
            conn.close()
            if grados:
                return self._ca_ordenar_grados(grados)
        except Exception:
            pass

        try:
            grados = core_matricula.listar_grados_distintos(solo_activos=False)
            return self._ca_ordenar_grados(
                [str(g).strip() for g in grados if str(g).strip()]
            )
        except Exception:
            return []

    def _ca_cursos_por_grado(self, grado=None):
        g = str(grado).strip() if grado is not None else ""
        if g and g not in ("Todos", "Seleccione"):
            cursos = cargar_cursos_por_grado(
                g,
                estudiantes_path=self.estudiantes_path,
                db_path=self.db_path,
            )
            return [str(c).strip() for c in cursos if str(c).strip()]

        try:
            return core_matricula.listar_cursos_distintos(
                grado=None, solo_activos=False
            )
        except Exception:
            return []

    def _carga_refresh_filtros(self):
        docentes = self._ca_todos_docentes(solo_activos=False)
        self.ca_docente_label_to_doc = {}
        self.ca_docente_doc_to_label = {}

        labels_doc = []
        for documento, nombre in docentes:
            doc_txt = str(documento or "").strip()
            nom_txt = str(nombre or "").strip()
            if not doc_txt:
                continue
            label = f"{nom_txt} ({doc_txt})" if nom_txt else doc_txt
            labels_doc.append(label)
            self.ca_docente_label_to_doc[label] = doc_txt
            self.ca_docente_doc_to_label[doc_txt] = label

        vals_doc = ["Todos"] + labels_doc
        self.ca_filter_doc["values"] = vals_doc
        if self.ca_filter_doc.get() not in vals_doc:
            self.ca_filter_doc.set("Todos")

        areas = self._ca_todas_areas()
        vals_area = ["Todos"] + areas
        self.ca_filter_area["values"] = vals_area
        if self.ca_filter_area.get() not in vals_area:
            self.ca_filter_area.set("Todos")

        grados = self._ca_todos_grados()
        if not grados:
            vals_grado = ["Sin grados disponibles"]
            self.ca_filter_grado["values"] = vals_grado
            self.ca_filter_grado.set("Sin grados disponibles")
            # Opcional: advertencia visual
            import tkinter.messagebox as mb

            mb.showwarning(
                "Advertencia",
                "No hay grados disponibles en la base de datos. Verifique la tabla plan_estudio o estudiantes.",
            )
        else:
            vals_grado = ["Todos"] + grados
            self.ca_filter_grado["values"] = vals_grado
            if self.ca_filter_grado.get() not in vals_grado:
                self.ca_filter_grado.set("Todos")

        self._carga_refresh_cursos_filtro()
        self._ca_refrescar_resumen_docente_filtro()

    def _ca_refrescar_resumen_docente_filtro(self):
        if not hasattr(self, "ca_resumen_horas_var"):
            return

        docente_sel = ""
        try:
            docente_sel = self.ca_filter_doc.get().strip()
        except Exception:
            docente_sel = ""

        if not docente_sel or docente_sel == "Todos":
            self.ca_resumen_horas_var.set(
                "Docente: N/D | Horas asignadas: 0 | Horas normales permitidas: 22 | "
                "Horas extras permitidas: 0 | Horas disponibles: 22"
            )
            return

        docente_documento = self.ca_docente_label_to_doc.get(docente_sel, "")
        if not docente_documento:
            self.ca_resumen_horas_var.set(
                "Docente: N/D | Horas asignadas: 0 | Horas normales permitidas: 22 | "
                "Horas extras permitidas: 0 | Horas disponibles: 22"
            )
            return

        normales_max, extras_max, _cfg = self._ca_obtener_limites_docente(
            docente_documento
        )
        horas_asignadas = self._ca_horas_totales_docente(docente_documento)
        limite_total = max(0, normales_max) + max(0, extras_max)
        horas_disponibles = max(0, limite_total - horas_asignadas)
        horas_extras_utilizadas = max(0, horas_asignadas - max(0, normales_max))

        resumen = (
            f"Docente: {docente_sel} | Horas asignadas: {horas_asignadas} | "
            f"Horas normales permitidas: {normales_max} | "
            f"Horas extras permitidas: {extras_max} | "
            f"Horas disponibles: {horas_disponibles}"
        )
        if horas_extras_utilizadas > 0:
            resumen += f" | Horas extras utilizadas: {horas_extras_utilizadas}"
        self.ca_resumen_horas_var.set(resumen)

    def _carga_on_docente_filtro(self, event=None):
        self._ca_refrescar_resumen_docente_filtro()
        self._load_carga_academica()

    def _carga_refresh_cursos_filtro(self):
        grado = self.ca_filter_grado.get().strip()
        cursos = self._ca_cursos_por_grado(
            grado if grado and grado != "Todos" else None
        )
        vals_curso = ["Todos"] + cursos
        self.ca_filter_curso["values"] = vals_curso
        if self.ca_filter_curso.get() not in vals_curso:
            self.ca_filter_curso.set("Todos")

    def _carga_on_grado_filtro(self, event=None):
        self._carga_refresh_cursos_filtro()
        self._load_carga_academica()

    def _load_carga_academica(self):
        for i in self.tree_carga.get_children():
            self.tree_carga.delete(i)

        docente_sel = self.ca_filter_doc.get().strip()
        docente_doc = None
        if docente_sel and docente_sel != "Todos":
            docente_doc = self.ca_docente_label_to_doc.get(docente_sel)

        area_sel = self.ca_filter_area.get().strip()
        area_sel = area_sel if area_sel and area_sel != "Todos" else None

        grado_sel = self.ca_filter_grado.get().strip()
        grado_sel = grado_sel if grado_sel and grado_sel != "Todos" else None

        curso_sel = self.ca_filter_curso.get().strip()
        curso_sel = curso_sel if curso_sel and curso_sel != "Todos" else None

        rows = core_docentes.listar_carga_academica(
            docente_documento=docente_doc,
            area=area_sel,
            grado=grado_sel,
            curso=curso_sel,
        )

        for row in rows:
            carga_id = row.get("id")
            doc = row.get("docente_documento")
            nombre = row.get("docente_nombre")
            director_doc = row.get("director_grupo_documento")
            director_nombre = row.get("director_grupo_nombre")
            area = row.get("area")
            grado = row.get("grado")
            curso = row.get("curso")
            horas_asignadas = row.get("horas_asignadas")
            horas_extras_usadas = row.get("horas_extras_usadas")
            anio = row.get("anio_lectivo")
            estado = row.get("estado")
            docente_txt = f"{nombre} ({doc})" if nombre else str(doc or "")
            director_txt = (
                f"{director_nombre} ({director_doc})"
                if director_nombre
                else str(director_doc or "")
            )
            self.tree_carga.insert(
                "",
                "end",
                iid=str(carga_id),
                values=(
                    docente_txt,
                    director_txt,
                    area or "",
                    grado or "",
                    curso or "",
                    int(horas_asignadas or 0),
                    int(horas_extras_usadas or 0),
                    anio or "",
                    estado or "",
                    "Editar | Eliminar",
                ),
            )

    def _carga_id_seleccionada(self, mostrar_msg=True):
        sel = self.tree_carga.selection()
        if not sel:
            if mostrar_msg:
                messagebox.showinfo(
                    "Carga Académica",
                    "Seleccione un registro.",
                    parent=self.win,
                )
            return None
        try:
            return int(sel[0])
        except Exception:
            if mostrar_msg:
                messagebox.showerror(
                    "Carga Académica",
                    "No se pudo identificar el registro seleccionado.",
                    parent=self.win,
                )
            return None

    def _obtener_carga_por_id(self, carga_id):
        return core_docentes.obtener_carga_academica(carga_id)

    def _carga_duplicada(self, docente_documento, area, grado, curso, excluir_id=None):
        return core_docentes.carga_academica_duplicada(
            docente_documento, area, grado, curso, excluir_id=excluir_id
        )

    def _ca_obtener_director_grupo_actual(self, grado, curso, anio_lectivo=None):
        filas = core_docentes.obtener_cargas_academicas_grupo(
            grado, curso, anio_lectivo
        )
        for fila in filas:
            director = str(fila.get("director_grupo_documento") or "").strip()
            if director:
                return director
        return ""

    def _dialog_carga_academica_masiva(self):
        d = tk.Toplevel(self.win)
        d.transient(self.win)
        d.grab_set()
        d.title("Agregar Carga Académica")
        d.geometry("1020x680")
        d.minsize(920, 620)

        cont = ttk.Frame(d, padding=12)
        cont.pack(fill="both", expand=True)
        cont.grid_columnconfigure(0, weight=1)
        cont.grid_rowconfigure(4, weight=1)

        docentes_activos = self._ca_todos_docentes(solo_activos=True)
        docente_label_to_doc = {}
        docente_doc_to_label = {}
        labels_doc = []
        for doc, nom in docentes_activos:
            doc_txt = str(doc or "").strip()
            nom_txt = str(nom or "").strip()
            if not doc_txt:
                continue
            label = f"{nom_txt} ({doc_txt})" if nom_txt else doc_txt
            labels_doc.append(label)
            docente_label_to_doc[label] = doc_txt
            docente_doc_to_label[doc_txt] = label

        ttk.Label(cont, text="Grado:").grid(row=0, column=0, sticky="w", padx=4)
        cb_grado = ttk.Combobox(cont, state="readonly", width=22)
        cb_grado.grid(row=0, column=0, sticky="w", padx=(62, 8), pady=(0, 8))

        ttk.Label(cont, text="Curso:").grid(row=0, column=0, sticky="w", padx=(300, 4))
        cb_curso = ttk.Combobox(cont, state="readonly", width=18)
        cb_curso.grid(row=0, column=0, sticky="w", padx=(352, 8), pady=(0, 8))

        ttk.Label(cont, text="Año lectivo:").grid(
            row=0, column=0, sticky="w", padx=(560, 4)
        )
        anio_activo = self.obtener_anio_lectivo_activo()
        anio_var = tk.StringVar(value=anio_activo)
        en_anio = ttk.Entry(cont, width=12, textvariable=anio_var, state="readonly")
        en_anio.grid(row=0, column=0, sticky="w", padx=(640, 8), pady=(0, 8))

        ttk.Label(cont, text="Estado:").grid(row=0, column=0, sticky="w", padx=(760, 4))
        cb_estado = ttk.Combobox(
            cont, values=["Activo", "Inactivo"], state="readonly", width=12
        )
        cb_estado.grid(row=0, column=0, sticky="w", padx=(815, 0), pady=(0, 8))
        cb_estado.set("Activo")

        ttk.Label(cont, text="Director de grupo:").grid(
            row=1, column=0, sticky="w", padx=4, pady=(0, 8)
        )
        cb_director = ttk.Combobox(cont, state="readonly", width=42)
        cb_director["values"] = labels_doc
        cb_director.grid(row=1, column=0, sticky="w", padx=(122, 8), pady=(0, 8))

        info = ttk.Label(
            cont,
            text="Seleccione grado y curso para cargar automáticamente las áreas del Plan de Estudios.",
        )
        info.grid(row=2, column=0, sticky="w", padx=4, pady=(0, 8))

        docente_actual_var = tk.StringVar(value="")
        resumen_docente_var = tk.StringVar(
            value=(
                "Docente: N/D | Horas asignadas: 0 | "
                "Normales permitidas: 22 | Extras permitidas: 0 | Disponibles: 22"
            )
        )
        ttk.Label(cont, textvariable=resumen_docente_var).grid(
            row=3, column=0, sticky="w", padx=4, pady=(0, 8)
        )

        acciones_sel = ttk.Frame(cont)
        acciones_sel.grid(row=4, column=0, sticky="we", pady=(0, 6))
        ttk.Label(acciones_sel, text="Áreas del Plan de Estudios").pack(side="left")

        frame_areas = ttk.LabelFrame(cont, text="Asignación por área")
        frame_areas.grid(row=5, column=0, sticky="nsew")
        frame_areas.grid_columnconfigure(0, weight=1)
        frame_areas.grid_rowconfigure(0, weight=1)

        canvas = tk.Canvas(frame_areas, highlightthickness=0)
        ysb = ttk.Scrollbar(frame_areas, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas)

        inner.bind(
            "<Configure>",
            lambda _e: canvas.configure(scrollregion=canvas.bbox("all")),
        )
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=ysb.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")

        total_var = tk.StringVar(value="Total de horas asignadas al curso: 0")
        ttk.Label(cont, textvariable=total_var, font=("Segoe UI", 10, "bold")).grid(
            row=6, column=0, sticky="w", padx=4, pady=(8, 6)
        )

        rows_state = []
        cargas_existentes = []
        horas_existentes_por_docente = {}

        def _asegurar_label_docente(documento, nombre=""):
            doc_txt = str(documento or "").strip()
            if not doc_txt:
                return ""
            existente = docente_doc_to_label.get(doc_txt)
            if existente:
                return existente
            nom_txt = str(nombre or "").strip()
            label = f"{nom_txt} ({doc_txt})" if nom_txt else doc_txt
            docente_doc_to_label[doc_txt] = label
            docente_label_to_doc[label] = doc_txt
            if label not in labels_doc:
                labels_doc.append(label)
            return label

        def _totales_temporales_por_docente():
            totales = {}
            for row in rows_state:
                if not row["sel"].get():
                    continue
                docente_label = row["cb_doc"].get().strip()
                docente_documento = docente_label_to_doc.get(docente_label, "")
                if not docente_documento:
                    continue
                try:
                    horas = int(str(row["horas_var"].get()).strip())
                except Exception:
                    horas = 0
                if horas > 0:
                    totales[docente_documento] = (
                        totales.get(docente_documento, 0) + horas
                    )
            return totales

        def _actualizar_resumen_docente(docente_label=None):
            if docente_label is not None:
                docente_actual_var.set(str(docente_label or "").strip())

            docente_label_act = docente_actual_var.get().strip()
            docente_documento = docente_label_to_doc.get(docente_label_act, "")
            if not docente_documento:
                resumen_docente_var.set(
                    "Docente: N/D | Horas asignadas: 0 | "
                    "Normales permitidas: 22 | Extras permitidas: 0 | Disponibles: 22"
                )
                return

            normales_max, extras_max, _cfg = self._ca_obtener_limites_docente(
                docente_documento
            )
            horas_actuales = max(
                0,
                self._ca_horas_totales_docente(docente_documento)
                - horas_existentes_por_docente.get(docente_documento, 0),
            )
            horas_temporales = _totales_temporales_por_docente().get(
                docente_documento, 0
            )
            horas_proyectadas = horas_actuales + horas_temporales
            limite_total = normales_max + extras_max
            horas_disponibles = max(0, limite_total - horas_proyectadas)
            extras_utilizadas = max(0, horas_proyectadas - normales_max)

            resumen = (
                f"Docente: {docente_label_act} | Horas asignadas: {horas_proyectadas} | "
                f"Horas normales permitidas: {normales_max} | "
                f"Horas extras permitidas: {extras_max} | "
                f"Horas disponibles: {horas_disponibles}"
            )
            if extras_utilizadas > 0:
                resumen += f" | Horas extras utilizadas: {extras_utilizadas}"
            resumen_docente_var.set(resumen)

        def _actualizar_total_horas(*_):
            total = 0
            for row in rows_state:
                if not row["sel"].get():
                    continue
                try:
                    total += int(str(row["horas_var"].get()).strip())
                except Exception:
                    pass
            total_var.set(f"Total de horas asignadas al curso: {total}")
            _actualizar_resumen_docente()

        def _on_docente_selected(_event, row):
            _actualizar_resumen_docente(row["cb_doc"].get().strip())

        def _toggle_row(row):
            activo = bool(row["sel"].get())
            if activo:
                row["cb_doc"].configure(state="readonly")
                row["sp_horas"].configure(state="normal")
                if not str(row["horas_var"].get()).strip():
                    row["horas_var"].set(str(row["horas_plan"]))
            else:
                row["cb_doc"].set("")
                row["cb_doc"].configure(state="disabled")
                row["sp_horas"].configure(state="disabled")
            _actualizar_total_horas()

        def _seleccionar_todas():
            for row in rows_state:
                row["sel"].set(True)
                _toggle_row(row)

        def _limpiar_seleccion():
            for row in rows_state:
                row["sel"].set(False)
                _toggle_row(row)

        ttk.Button(
            acciones_sel, text="Seleccionar todas", command=_seleccionar_todas
        ).pack(side="left", padx=(12, 6))
        ttk.Button(
            acciones_sel, text="Limpiar selección", command=_limpiar_seleccion
        ).pack(side="left")

        def _render_areas(areas_horas, precargadas=None):
            for w in inner.winfo_children():
                w.destroy()
            rows_state.clear()
            precargadas = precargadas or {}

            if not areas_horas:
                ttk.Label(
                    inner,
                    text="No hay áreas para este grado/curso en el Plan de Estudios.",
                ).grid(row=0, column=0, sticky="w", padx=8, pady=8)
                _actualizar_total_horas()
                return

            headers = [
                ("Sel", 0),
                ("Área", 1),
                ("Horas plan", 2),
                ("Docente", 3),
                ("Horas asignadas", 4),
            ]
            for txt, col in headers:
                ttk.Label(inner, text=txt, font=("Segoe UI", 9, "bold")).grid(
                    row=0, column=col, sticky="w", padx=6, pady=(6, 4)
                )

            for col in range(5):
                inner.grid_columnconfigure(col, weight=1 if col in (1, 3) else 0)

            for i, (area, horas_plan) in enumerate(areas_horas, start=1):
                horas_plan_int = max(0, int(horas_plan or 0))
                preload = precargadas.get(str(area)) or {}
                horas_precargadas = int(
                    preload.get("horas_asignadas") or horas_plan_int
                )
                var_sel = tk.BooleanVar(value=bool(preload))
                var_horas = tk.StringVar(value=str(horas_precargadas))

                chk = ttk.Checkbutton(inner, variable=var_sel)
                chk.grid(row=i, column=0, sticky="w", padx=6, pady=4)

                ttk.Label(inner, text=str(area)).grid(
                    row=i, column=1, sticky="w", padx=6, pady=4
                )
                ttk.Label(inner, text=str(horas_plan_int)).grid(
                    row=i, column=2, sticky="w", padx=6, pady=4
                )

                cb_doc = ttk.Combobox(inner, state="disabled", width=36)
                cb_doc["values"] = labels_doc
                cb_doc.grid(row=i, column=3, sticky="we", padx=6, pady=4)

                sp_to = max(1, horas_plan_int if horas_plan_int > 0 else 80)
                sp_horas = tk.Spinbox(
                    inner,
                    from_=0,
                    to=sp_to,
                    width=8,
                    textvariable=var_horas,
                    state="disabled",
                )
                sp_horas.grid(row=i, column=4, sticky="w", padx=6, pady=4)

                row = {
                    "area": str(area),
                    "horas_plan": horas_plan_int,
                    "sel": var_sel,
                    "cb_doc": cb_doc,
                    "sp_horas": sp_horas,
                    "horas_var": var_horas,
                }
                rows_state.append(row)

                if preload:
                    label_precargado = _asegurar_label_docente(
                        preload.get("docente_documento"), preload.get("docente_nombre")
                    )
                    cb_doc["values"] = labels_doc
                    cb_doc.set(label_precargado)
                    cb_doc.configure(state="readonly")
                    sp_horas.configure(state="normal")

                chk.configure(command=lambda r=row: _toggle_row(r))
                cb_doc.bind(
                    "<<ComboboxSelected>>",
                    lambda e, r=row: _on_docente_selected(e, r),
                )
                var_horas.trace_add("write", _actualizar_total_horas)
                sp_horas.bind("<FocusOut>", _actualizar_total_horas)

            _actualizar_total_horas()

        grados = self._ca_todos_grados()
        cb_grado["values"] = grados

        def _cargar_areas_por_grado_curso(*_):
            nonlocal cargas_existentes, horas_existentes_por_docente
            grado_sel = cb_grado.get().strip()
            curso_sel = cb_curso.get().strip()
            if not grado_sel or not curso_sel:
                cargas_existentes = []
                horas_existentes_por_docente = {}
                _render_areas([])
                return

            cargas_existentes = core_docentes.obtener_cargas_academicas_grupo(
                grado_sel,
                curso_sel,
                anio_var.get().strip() or anio_activo,
            )
            horas_existentes_por_docente = {}
            precargadas_por_area = {}
            for carga in cargas_existentes:
                doc_carga = str(carga.get("docente_documento") or "").strip()
                if doc_carga:
                    horas_existentes_por_docente[doc_carga] = (
                        horas_existentes_por_docente.get(doc_carga, 0)
                        + int(carga.get("horas_asignadas") or 0)
                    )
                area_carga = str(carga.get("area") or "").strip()
                if area_carga and area_carga not in precargadas_por_area:
                    precargadas_por_area[area_carga] = carga

            areas_horas = list(self._ca_plan_areas_horas(grado_sel, curso_sel) or [])
            areas_presentes = {str(area).strip() for area, _ in areas_horas}
            for area_carga, carga in precargadas_por_area.items():
                if area_carga not in areas_presentes:
                    areas_horas.append(
                        (area_carga, int(carga.get("horas_asignadas") or 0))
                    )

            if cargas_existentes:
                estado_existente = str(cargas_existentes[0].get("estado") or "").strip()
                if estado_existente in ("Activo", "Inactivo"):
                    cb_estado.set(estado_existente)
                director_actual = self._ca_obtener_director_grupo_actual(
                    grado_sel,
                    curso_sel,
                    anio_var.get().strip() or anio_activo,
                )
                cb_director.set(docente_doc_to_label.get(director_actual, ""))
            elif not cb_estado.get().strip():
                cb_estado.set("Activo")
                cb_director.set("")

            if not areas_horas:
                info.config(
                    text="No hay áreas configuradas en Plan de Estudios para el grado/curso seleccionado."
                )
            elif cargas_existentes:
                info.config(
                    text=(
                        f"Se precargaron {len(cargas_existentes)} asignaciones existentes para {grado_sel}-{curso_sel}. "
                        f"Puede ajustarlas y guardar para reemplazar la carga del curso."
                    )
                )
            else:
                info.config(
                    text=f"Se cargaron {len(areas_horas)} áreas del Plan de Estudios para {grado_sel}-{curso_sel}."
                )
            _render_areas(areas_horas, precargadas_por_area)

        def _refrescar_cursos(*_):
            grado_sel = cb_grado.get().strip()
            cursos = self._ca_cursos_por_grado(grado_sel)
            cb_curso["values"] = cursos
            if cb_curso.get().strip() not in cursos:
                cb_curso.set(cursos[0] if cursos else "")
            _cargar_areas_por_grado_curso()

        cb_grado.bind("<<ComboboxSelected>>", _refrescar_cursos)
        cb_curso.bind("<<ComboboxSelected>>", _cargar_areas_por_grado_curso)

        if grados:
            cb_grado.set(grados[0])
            _refrescar_cursos()

        def _guardar():
            grado = cb_grado.get().strip()
            curso = cb_curso.get().strip()
            anio = anio_var.get().strip() or anio_activo
            estado = cb_estado.get().strip() or "Activo"

            if not grado:
                messagebox.showerror("Validación", "Seleccione un grado.", parent=d)
                return
            if not curso:
                messagebox.showerror("Validación", "Seleccione un curso.", parent=d)
                return
            if not anio:
                messagebox.showerror(
                    "Validación", "El año lectivo es obligatorio.", parent=d
                )
                return
            director_label = cb_director.get().strip()
            if not director_label:
                messagebox.showerror(
                    "Validación",
                    "Seleccione un director de grupo.",
                    parent=d,
                )
                return

            director_grupo_documento = docente_label_to_doc.get(director_label, "")
            if not director_grupo_documento:
                messagebox.showerror(
                    "Validación",
                    "El director de grupo seleccionado no es válido.",
                    parent=d,
                )
                return

            seleccionadas = [r for r in rows_state if r["sel"].get()]
            if not seleccionadas:
                messagebox.showerror(
                    "Validación",
                    "Seleccione al menos un área para asignar.",
                    parent=d,
                )
                return

            items = []
            areas_sin_docente = []
            total_plan = 0
            total_asignado = 0
            horas_no_coinciden = []
            base_horas_docente = {}
            acumulado_nuevo_docente = {}
            usa_horas_extras = False

            for row in seleccionadas:
                area = row["area"]
                horas_plan = int(row["horas_plan"] or 0)
                total_plan += horas_plan

                docente_label = row["cb_doc"].get().strip()
                if not docente_label:
                    areas_sin_docente.append(area)
                    continue

                docente_documento = docente_label_to_doc.get(docente_label, "")
                if not docente_documento:
                    areas_sin_docente.append(area)
                    continue

                try:
                    horas_asignadas = int(str(row["horas_var"].get()).strip())
                except Exception:
                    messagebox.showerror(
                        "Validación",
                        f"Horas asignadas inválidas en el área '{area}'.",
                        parent=d,
                    )
                    return

                if horas_asignadas > horas_plan:
                    messagebox.showerror(
                        "Validación",
                        "Las horas asignadas superan las horas permitidas según el Plan de Estudios",
                        parent=d,
                    )
                    return

                if (
                    horas_asignadas <= 0
                    or horas_plan <= 0
                    or horas_asignadas != horas_plan
                ):
                    horas_no_coinciden.append(area)

                if docente_documento not in base_horas_docente:
                    base_horas_docente[docente_documento] = max(
                        0,
                        self._ca_horas_totales_docente(docente_documento)
                        - horas_existentes_por_docente.get(docente_documento, 0),
                    )

                normales_max, extras_max, _cfg = self._ca_obtener_limites_docente(
                    docente_documento
                )
                limite_total = max(0, normales_max) + max(0, extras_max)

                horas_antes = base_horas_docente[
                    docente_documento
                ] + acumulado_nuevo_docente.get(docente_documento, 0)
                horas_despues = horas_antes + max(0, horas_asignadas)

                if horas_despues > limite_total:
                    messagebox.showerror(
                        "Carga Académica",
                        "No es posible asignar más horas. El docente supera el límite máximo permitido.",
                        parent=d,
                    )
                    return

                extras_antes = max(0, horas_antes - normales_max)
                extras_despues = max(0, horas_despues - normales_max)
                horas_extras_registro = max(0, extras_despues - extras_antes)
                if horas_extras_registro > 0:
                    usa_horas_extras = True

                acumulado_nuevo_docente[docente_documento] = (
                    acumulado_nuevo_docente.get(docente_documento, 0)
                    + max(0, horas_asignadas)
                )

                total_asignado += max(0, horas_asignadas)
                items.append(
                    {
                        "docente_documento": docente_documento,
                        "area": area,
                        "grado": grado,
                        "curso": curso,
                        "horas_asignadas": horas_asignadas,
                        "horas_extras_usadas": horas_extras_registro,
                        "anio_lectivo": anio,
                        "estado": estado,
                        "director_grupo_documento": director_grupo_documento,
                    }
                )

            if areas_sin_docente:
                messagebox.showerror(
                    "Validación",
                    "Hay áreas seleccionadas sin docente asignado.",
                    parent=d,
                )
                return

            if horas_no_coinciden or total_asignado != total_plan:
                messagebox.showerror(
                    "Validación",
                    "Las horas no coinciden con la carga permitida.",
                    parent=d,
                )
                return

            d.result = {
                "items": items,
                "total_horas": total_asignado,
                "usa_horas_extras": usa_horas_extras,
                "grado": grado,
                "curso": curso,
                "anio_lectivo": anio,
                "estado": estado,
                "director_grupo_documento": director_grupo_documento,
                "tenia_existentes": bool(cargas_existentes),
                "cantidad_existentes": len(cargas_existentes),
            }
            d.destroy()

        btns = ttk.Frame(cont)
        btns.grid(row=7, column=0, sticky="w", pady=(4, 0))
        ttk.Button(btns, text="Guardar", command=_guardar).pack(side="left")
        ttk.Button(btns, text="Cancelar", command=d.destroy).pack(
            side="left", padx=(8, 0)
        )

        d.wait_window()
        return getattr(d, "result", None)

    def _dialog_carga_academica(self, inicial=None):
        d = tk.Toplevel(self.win)
        d.transient(self.win)
        d.grab_set()
        d.title("Agregar Carga Académica" if not inicial else "Editar Carga Académica")
        d.geometry("620x520")
        d.minsize(560, 480)

        frame = ttk.Frame(d, padding=12)
        frame.pack(fill="both", expand=True)

        entries = {}

        ttk.Label(frame, text="Docente:").grid(
            row=0, column=0, sticky="e", padx=6, pady=8
        )
        cb_docente = ttk.Combobox(frame, state="readonly", width=42)
        cb_docente.grid(row=0, column=1, sticky="we", padx=6, pady=8)
        entries["docente"] = cb_docente

        # Grado y Curso ANTES que Área para que el filtro tenga sentido
        ttk.Label(frame, text="Grado:").grid(
            row=1, column=0, sticky="e", padx=6, pady=8
        )
        cb_grado = ttk.Combobox(frame, state="readonly", width=16)
        cb_grado.grid(row=1, column=1, sticky="w", padx=6, pady=8)
        entries["grado"] = cb_grado

        ttk.Label(frame, text="Curso:").grid(
            row=2, column=0, sticky="e", padx=6, pady=8
        )
        cb_curso = ttk.Combobox(frame, state="readonly", width=16)
        cb_curso.grid(row=2, column=1, sticky="w", padx=6, pady=8)
        entries["curso"] = cb_curso

        # Área se carga desde Plan de Estudios filtrando por Grado+Curso
        lbl_area = ttk.Label(frame, text="Área (Plan de Estudios):")
        lbl_area.grid(row=3, column=0, sticky="e", padx=6, pady=8)
        cb_area = ttk.Combobox(frame, state="readonly", width=30)
        cb_area.grid(row=3, column=1, sticky="we", padx=6, pady=8)
        entries["area"] = cb_area

        ttk.Label(frame, text="Año lectivo:").grid(
            row=4, column=0, sticky="e", padx=6, pady=8
        )
        anio_activo = self.obtener_anio_lectivo_activo()
        anio_var = tk.StringVar(value=anio_activo)
        entry_anio = ttk.Entry(frame, width=16, textvariable=anio_var, state="readonly")
        entry_anio.grid(row=4, column=1, sticky="w", padx=6, pady=8)
        entries["anio"] = entry_anio

        ttk.Label(frame, text="Director de grupo:").grid(
            row=5, column=0, sticky="e", padx=6, pady=8
        )
        cb_director = ttk.Combobox(frame, state="readonly", width=42)
        cb_director.grid(row=5, column=1, sticky="we", padx=6, pady=8)
        entries["director_grupo"] = cb_director

        ttk.Label(frame, text="Estado:").grid(
            row=6, column=0, sticky="e", padx=6, pady=8
        )
        cb_estado = ttk.Combobox(
            frame,
            values=["Activo", "Inactivo"],
            state="readonly",
            width=16,
        )
        cb_estado.grid(row=6, column=1, sticky="w", padx=6, pady=8)
        entries["estado"] = cb_estado

        frame.grid_columnconfigure(1, weight=1)

        docentes_activos = self._ca_todos_docentes(solo_activos=True)
        docente_label_to_doc = {}
        docente_doc_to_label = {}
        labels_doc = []
        for doc, nom in docentes_activos:
            doc_txt = str(doc or "").strip()
            nom_txt = str(nom or "").strip()
            if not doc_txt:
                continue
            label = f"{nom_txt} ({doc_txt})" if nom_txt else doc_txt
            labels_doc.append(label)
            docente_label_to_doc[label] = doc_txt
            docente_doc_to_label[doc_txt] = label
        cb_docente["values"] = labels_doc
        cb_director["values"] = labels_doc

        grados = self._ca_todos_grados()
        cb_grado["values"] = grados

        def refrescar_areas(event=None):
            """Recarga el combo de Área desde Plan de Estudios para el grado+curso actual."""
            grado_sel = cb_grado.get().strip()
            curso_sel = cb_curso.get().strip()
            area_prev = cb_area.get().strip()
            if grado_sel and curso_sel:
                areas_plan = self.cargar_areas_plan_estudio(grado_sel, curso_sel)
            else:
                areas_plan = []
            # Si el plan tiene áreas configuradas, usarlas; si no, mostrar todas
            if areas_plan:
                lbl_area.config(text="Área (Plan de Estudios):")
                areas = areas_plan
            else:
                lbl_area.config(text="Área (todas):")
                areas = self._ca_todas_areas()
            cb_area["values"] = areas
            if area_prev in areas:
                cb_area.set(area_prev)
            else:
                cb_area.set(areas[0] if areas else "")

        def refrescar_cursos(event=None):
            grado_sel = cb_grado.get().strip()
            cursos = self._ca_cursos_por_grado(grado_sel)
            cb_curso["values"] = cursos
            if cb_curso.get() not in cursos:
                cb_curso.set(cursos[0] if cursos else "")
            refrescar_areas()

        cb_grado.bind("<<ComboboxSelected>>", refrescar_cursos)
        cb_curso.bind("<<ComboboxSelected>>", refrescar_areas)

        if inicial:
            doc_ini = str(inicial.get("docente_documento") or "").strip()
            cb_docente.set(docente_doc_to_label.get(doc_ini, ""))

            grado_ini = str(inicial.get("grado") or "").strip()
            if grado_ini in grados:
                cb_grado.set(grado_ini)
            elif grados:
                cb_grado.set(grados[0])

            refrescar_cursos()

            curso_ini = str(inicial.get("curso") or "").strip()
            cursos_actuales = list(cb_curso["values"])
            if curso_ini and curso_ini not in cursos_actuales:
                cursos_actuales.append(curso_ini)
                cb_curso["values"] = cursos_actuales
            if curso_ini:
                cb_curso.set(curso_ini)

            # Recargar áreas según grado+curso definitivos antes de setear área
            refrescar_areas()
            area_ini = str(inicial.get("area") or "").strip()
            areas_actuales = list(cb_area["values"])
            if area_ini and area_ini not in areas_actuales:
                # El área guardada no está en el plan actual; agregarla para no perderla
                areas_actuales.append(area_ini)
                cb_area["values"] = areas_actuales
            if area_ini:
                cb_area.set(area_ini)

            anio_ini = str(inicial.get("anio_lectivo") or "").strip()
            anio_var.set(anio_ini or anio_activo)
            director_ini = str(inicial.get("director_grupo_documento") or "").strip()
            cb_director.set(docente_doc_to_label.get(director_ini, ""))
            estado_ini = str(inicial.get("estado") or "Activo").strip()
            cb_estado.set(
                estado_ini if estado_ini in ("Activo", "Inactivo") else "Activo"
            )
        else:
            if labels_doc:
                cb_docente.set(labels_doc[0])
            if grados:
                cb_grado.set(grados[0])
            refrescar_cursos()  # también llama refrescar_areas
            anio_var.set(anio_activo)
            director_actual = self._ca_obtener_director_grupo_actual(
                cb_grado.get().strip(),
                cb_curso.get().strip(),
                anio_activo,
            )
            cb_director.set(docente_doc_to_label.get(director_actual, ""))
            cb_estado.set("Activo")

        def on_guardar():
            docente_label = cb_docente.get().strip()
            area = cb_area.get().strip()
            grado = cb_grado.get().strip()
            curso = cb_curso.get().strip()
            anio = anio_var.get().strip() or anio_activo
            estado = cb_estado.get().strip() or "Activo"

            if not docente_label:
                messagebox.showerror("Validación", "Seleccione un docente.", parent=d)
                return
            if not area:
                messagebox.showerror("Validación", "Seleccione un área.", parent=d)
                return
            if not grado:
                messagebox.showerror("Validación", "Seleccione un grado.", parent=d)
                return
            if not curso:
                messagebox.showerror("Validación", "Seleccione un curso.", parent=d)
                return
            if not anio:
                messagebox.showerror(
                    "Validación", "El año lectivo es obligatorio.", parent=d
                )
                return
            director_label = cb_director.get().strip()
            if not director_label:
                messagebox.showerror(
                    "Validación", "Seleccione un director de grupo.", parent=d
                )
                return

            docente_documento = docente_label_to_doc.get(docente_label, "")
            if not docente_documento:
                messagebox.showerror(
                    "Validación",
                    "El docente seleccionado no es válido.",
                    parent=d,
                )
                return

            director_grupo_documento = docente_label_to_doc.get(director_label, "")
            if not director_grupo_documento:
                messagebox.showerror(
                    "Validación",
                    "El director de grupo seleccionado no es válido.",
                    parent=d,
                )
                return

            d.result = {
                "docente_documento": docente_documento,
                "area": area,
                "grado": grado,
                "curso": curso,
                "anio_lectivo": anio,
                "estado": estado,
                "director_grupo_documento": director_grupo_documento,
            }
            d.destroy()

        frame_btn = ttk.Frame(frame)
        frame_btn.grid(row=7, column=0, columnspan=2, sticky="w", pady=(14, 0))

        ttk.Button(frame_btn, text="Guardar", command=on_guardar).pack(side="left")
        ttk.Button(frame_btn, text="Cancelar", command=d.destroy).pack(
            side="left", padx=(8, 0)
        )

        d.wait_window()
        return getattr(d, "result", None)

    def carga_academica_agregar(self):
        if not self._requiere_permiso("desktop.superadmin.carga_academica.crear"):
            return
        self._ca_asegurar_esquema_horas()
        datos = self._dialog_carga_academica_masiva()
        if not datos:
            return

        items = datos.get("items") if isinstance(datos, dict) else None
        if not items:
            return

        try:
            resultado = core_docentes.reemplazar_cargas_academicas_grupo(
                datos.get("grado"),
                datos.get("curso"),
                datos.get("anio_lectivo"),
                items,
            )
            guardadas = int(resultado.get("guardadas") or 0)
            usa_horas_extras = bool(resultado.get("usa_horas_extras"))
            tenia_existentes = bool(datos.get("tenia_existentes"))
            reemplazadas = int(resultado.get("reemplazadas") or 0)
        except ValueError as exc:
            messagebox.showerror(
                "Carga Académica",
                str(exc),
                parent=self.win,
            )
            return
        except Exception:
            messagebox.showerror(
                "Carga Académica",
                "No fue posible guardar la carga académica del curso.",
                parent=self.win,
            )
            return

        try:
            core_docentes.actualizar_director_grupo(
                datos.get("grado"),
                datos.get("curso"),
                datos.get("anio_lectivo"),
                datos.get("director_grupo_documento"),
            )
        except Exception:
            pass

        self._load_carga_academica()
        if usa_horas_extras:
            messagebox.showinfo(
                "Carga Académica",
                "Las horas asignadas superan la carga normal del docente. Se registrarán como horas extras.",
                parent=self.win,
            )
        mensaje = f"Carga académica registrada correctamente. Registros guardados: {guardadas}."
        if tenia_existentes:
            mensaje = (
                f"Carga académica actualizada correctamente. Registros guardados: {guardadas}. "
                f"Asignaciones reemplazadas: {reemplazadas}."
            )
        messagebox.showinfo(
            "Carga Académica",
            mensaje,
            parent=self.win,
        )

    def carga_academica_editar(self):
        if not self._requiere_permiso("desktop.superadmin.carga_academica.editar"):
            return
        self._ca_asegurar_esquema_horas()
        carga_id = self._carga_id_seleccionada()
        if carga_id is None:
            return

        actual = self._obtener_carga_por_id(carga_id)
        if not actual:
            messagebox.showerror(
                "Carga Académica",
                "No se encontró el registro seleccionado.",
                parent=self.win,
            )
            return

        nuevos = self._dialog_carga_academica(inicial=actual)
        if not nuevos:
            return

        horas_asignadas = max(0, int(actual.get("horas_asignadas") or 0))
        normales_max, extras_max, _cfg = self._ca_obtener_limites_docente(
            nuevos["docente_documento"]
        )
        limite_total = max(0, normales_max) + max(0, extras_max)
        horas_base = self._ca_horas_totales_docente(
            nuevos["docente_documento"], excluir_id=carga_id
        )
        horas_despues = horas_base + horas_asignadas
        if horas_despues > limite_total:
            messagebox.showerror(
                "Carga Académica",
                "No es posible asignar más horas. El docente supera el límite máximo permitido.",
                parent=self.win,
            )
            return

        extras_antes = max(0, horas_base - normales_max)
        extras_despues = max(0, horas_despues - normales_max)
        horas_extras_registro = max(0, extras_despues - extras_antes)

        if self._carga_duplicada(
            nuevos["docente_documento"],
            nuevos["area"],
            nuevos["grado"],
            nuevos["curso"],
            excluir_id=carga_id,
        ):
            messagebox.showerror(
                "Carga Académica",
                "Ya existe una carga académica con ese docente, área, grado y curso.",
                parent=self.win,
            )
            return

        try:
            core_docentes.actualizar_carga_academica(
                carga_id,
                nuevos["docente_documento"],
                nuevos["area"],
                nuevos["grado"],
                nuevos["curso"],
                horas_asignadas,
                horas_extras_registro,
                nuevos["anio_lectivo"],
                nuevos["estado"],
                nuevos.get("director_grupo_documento"),
            )
        except Exception:
            messagebox.showerror(
                "Carga Académica",
                "Ya existe una carga académica con ese docente, área, grado y curso.",
                parent=self.win,
            )
            return

        try:
            core_docentes.actualizar_director_grupo(
                nuevos["grado"],
                nuevos["curso"],
                nuevos["anio_lectivo"],
                nuevos.get("director_grupo_documento"),
            )
        except Exception:
            pass

        self._load_carga_academica()
        if horas_extras_registro > 0:
            messagebox.showinfo(
                "Carga Académica",
                "Las horas asignadas superan la carga normal del docente. Se registrarán como horas extras.",
                parent=self.win,
            )
        messagebox.showinfo(
            "Carga Académica",
            "Carga académica actualizada correctamente.",
            parent=self.win,
        )

    def carga_academica_eliminar(self):
        if not self._requiere_permiso("desktop.superadmin.carga_academica.eliminar"):
            return
        carga_id = self._carga_id_seleccionada()
        if carga_id is None:
            return

        vals = self.tree_carga.item(str(carga_id)).get("values", [])
        texto = vals[0] if vals else f"ID {carga_id}"
        if not messagebox.askyesno(
            "Confirmar",
            f"¿Eliminar carga académica de {texto}?",
            parent=self.win,
        ):
            return

        core_docentes.eliminar_carga_academica(carga_id)
        self._load_carga_academica()

    def carga_academica_actualizar(self):
        self._carga_refresh_filtros()
        self._load_carga_academica()

    def _carga_tree_click_accion(self, event):
        item = self.tree_carga.identify_row(event.y)
        col = self.tree_carga.identify_column(event.x)
        cols = list(self.tree_carga["columns"])
        try:
            idx_acciones = cols.index("acciones") + 1
            col_acciones = f"#{idx_acciones}"
        except Exception:
            col_acciones = "#9"

        if not item or col != col_acciones:
            return

        bbox = self.tree_carga.bbox(item, column=col)
        if not bbox:
            return
        self.tree_carga.selection_set(item)

        x_rel = event.x - bbox[0]
        width = max(int(bbox[2]), 1)
        tramo = width / 2.0
        if x_rel < tramo:
            self.carga_academica_editar()
        else:
            self.carga_academica_eliminar()
        return "break"

    def _normalizar_columna_docente_excel(self, nombre_columna):
        texto = str(nombre_columna or "").strip().lower()
        reemplazos = {
            "á": "a",
            "é": "e",
            "í": "i",
            "ó": "o",
            "ú": "u",
            "ñ": "n",
        }
        for viejo, nuevo in reemplazos.items():
            texto = texto.replace(viejo, nuevo)
        return re.sub(r"[^a-z0-9]+", "", texto)

    def _limpiar_documento_docente(self, valor):
        if valor is None:
            return ""
        texto = str(valor).strip()
        if texto.lower() == "nan":
            return ""
        if re.fullmatch(r"\d+\.0", texto):
            texto = texto[:-2]
        return texto

    def docentes_importar_excel(self):
        path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel para importar docentes",
            filetypes=[("Excel", "*.xlsx;*.xls")],
            parent=self.win,
        )
        if not path:
            return

        incoming = self._read_excel(path)
        if incoming is None:
            messagebox.showerror(
                "Importar Excel",
                "Archivo vacío o inválido.",
                parent=self.win,
            )
            return

        if _HAS_PANDAS:
            columnas_origen = list(incoming.columns)
            registros = incoming.to_dict(orient="records")
        else:
            registros = incoming.get("_rows", [])
            if not registros:
                messagebox.showerror(
                    "Importar Excel",
                    "Archivo vacío o inválido.",
                    parent=self.win,
                )
                return
            columnas_origen = set()
            for fila in registros:
                columnas_origen.update(fila.keys())
            columnas_origen = list(columnas_origen)

        alias = {
            "tipo_documento": [
                "tipodocumento",
                "tipodoc",
                "tipoid",
                "tipoidentificacion",
                "tipodedocumento",
            ],
            "documento": [
                "documento",
                "numerodedocumento",
                "numeroidentificacion",
                "identificacion",
                "cedula",
                "doc",
            ],
            "nombre": ["nombre", "nombrecompleto", "nombres", "docente"],
            "sexo": ["sexo", "genero"],
            "fecha_nacimiento": [
                "fechanacimiento",
                "fechanac",
                "fechana",
                "nacimiento",
            ],
            "telefono": ["telefono", "celular", "movil"],
            "correo": ["correo", "email", "mail", "correoinstitucional"],
            "cargo": ["cargo", "rol"],
            "jornada": ["jornada"],
            "sede": ["sede"],
            "estado": ["estado"],
        }

        columnas_normalizadas = {}
        for col in columnas_origen:
            columnas_normalizadas[self._normalizar_columna_docente_excel(col)] = col

        columna_por_campo = {}
        for campo, posibles in alias.items():
            encontrada = None
            for posible in posibles:
                if posible in columnas_normalizadas:
                    encontrada = columnas_normalizadas[posible]
                    break
            columna_por_campo[campo] = encontrada

        if not columna_por_campo["documento"] or not columna_por_campo["nombre"]:
            messagebox.showwarning(
                "Importar Excel",
                "El archivo debe incluir columnas de Documento y Nombre.",
                parent=self.win,
            )
            return

        def valor_fila(fila, campo):
            col = columna_por_campo.get(campo)
            if not col:
                return ""
            valor = fila.get(col, "")
            if valor is None:
                return ""
            texto = str(valor).strip()
            return "" if texto.lower() == "nan" else texto

        docs_archivo = set()
        insertados = 0
        duplicados = 0
        invalidos = 0

        for fila in registros:
            documento = self._limpiar_documento_docente(valor_fila(fila, "documento"))
            nombre = valor_fila(fila, "nombre")

            if not documento or not nombre:
                invalidos += 1
                continue

            if not documento.isdigit():
                invalidos += 1
                continue

            if documento in docs_archivo or self._documento_docente_existe(documento):
                duplicados += 1
                continue

            docs_archivo.add(documento)

            datos = {
                "tipo_documento": valor_fila(fila, "tipo_documento") or "CC",
                "documento": documento,
                "nombre": nombre,
                "sexo": valor_fila(fila, "sexo"),
                "fecha_nacimiento": valor_fila(fila, "fecha_nacimiento"),
                "telefono": valor_fila(fila, "telefono"),
                "correo": valor_fila(fila, "correo"),
                "cargo": valor_fila(fila, "cargo") or "Docente",
                "jornada": valor_fila(fila, "jornada") or "Mañana",
                "sede": valor_fila(fila, "sede"),
                "estado": valor_fila(fila, "estado") or "Activo",
            }

            try:
                core_docentes.crear_o_actualizar_docente(
                    tipo_documento=datos["tipo_documento"],
                    documento=datos["documento"],
                    nombre=datos["nombre"],
                    sexo=datos["sexo"],
                    fecha_nacimiento=datos["fecha_nacimiento"],
                    telefono=datos["telefono"],
                    correo=datos["correo"],
                    cargo=datos["cargo"],
                    jornada=datos["jornada"],
                    sede=datos["sede"],
                    estado=datos["estado"],
                )
                insertados += 1
            except ValueError:
                duplicados += 1
            except Exception:
                invalidos += 1

        self._load_docentes()

        if insertados == 0 and duplicados == 0 and invalidos > 0:
            messagebox.showwarning(
                "Importar Excel",
                "No se importaron docentes. Verifica formato y datos obligatorios.",
                parent=self.win,
            )
            return

        messagebox.showinfo(
            "Importar Excel",
            "Importación finalizada.\n"
            f"Insertados: {insertados}\n"
            f"Duplicados: {duplicados}\n"
            f"Inválidos: {invalidos}",
            parent=self.win,
        )

    def docentes_exportar_excel(self):
        path = filedialog.asksaveasfilename(
            title="Exportar docentes a Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            parent=self.win,
        )
        if not path:
            return

        columnas = [
            "tipo_documento",
            "documento",
            "nombre",
            "sexo",
            "fecha_nacimiento",
            "telefono",
            "correo",
            "cargo",
            "jornada",
            "sede",
            "estado",
            "fecha_registro",
        ]

        docentes = core_docentes.listar_docentes_exportacion()
        rows = [tuple(doc.get(col, "") for col in columnas) for doc in docentes]

        if not rows:
            messagebox.showinfo(
                "Exportar Excel",
                "No hay docentes registrados para exportar.",
                parent=self.win,
            )
            return

        if _HAS_PANDAS:
            pd.DataFrame(rows, columns=columnas).to_excel(path, index=False)
        else:
            data = [dict(zip(columnas, row)) for row in rows]
            self._write_excel(path, data)

        messagebox.showinfo(
            "Exportar Excel",
            "Exportación completada correctamente.",
            parent=self.win,
        )

    def _docente_tree_click_accion(self, event):
        item = self.tree_doc.identify_row(event.y)
        col = self.tree_doc.identify_column(event.x)
        if not item or col != "#6":
            return

        bbox = self.tree_doc.bbox(item, column=col)
        if not bbox:
            return
        self.tree_doc.selection_set(item)

        x_rel = event.x - bbox[0]
        width = max(int(bbox[2]), 1)
        tramo = width / 3.0

        if x_rel < tramo:
            self.docente_editar(item_id=item)
        elif x_rel < (2 * tramo):
            self.docente_carga_academica(item_id=item)
        else:
            self.docente_eliminar(item_id=item)
        return "break"

    def _dialog_docente(self, inicial=None):
        d = tk.Toplevel(self.win)
        d.transient(self.win)
        d.grab_set()
        d.title("Registro de personal" if not inicial else "Editar personal")
        d.geometry("680x640")
        d.minsize(620, 560)

        frame_main = ttk.Frame(d, padding=12)
        frame_main.pack(fill="both", expand=True)

        frame_personal = ttk.LabelFrame(frame_main, text="INFORMACIÓN PERSONAL")
        frame_personal.pack(fill="x", pady=(0, 10))

        frame_contacto = ttk.LabelFrame(frame_main, text="INFORMACIÓN DE CONTACTO")
        frame_contacto.pack(fill="x", pady=(0, 10))

        frame_laboral = ttk.LabelFrame(frame_main, text="INFORMACIÓN LABORAL")
        frame_laboral.pack(fill="x")

        entries = {}

        def add_field(parent, row, key, label, kind="entry", values=None):
            # Personalización de nombres de campos
            label_custom = label
            if key == "correo":
                label_custom = "Email"
            elif key == "estado_academico":
                label_custom = "Estado"
            elif key == "codigo":
                label_custom = "A.idalumno"
            ttk.Label(parent, text=label_custom + ":").grid(
                row=row,
                column=0,
                sticky="e",
                padx=6,
                pady=6,
            )
            if kind == "combobox":
                w = ttk.Combobox(
                    parent, values=values or [], state="readonly", width=36
                )
            else:
                w = ttk.Entry(parent, width=38)
            w.grid(row=row, column=1, sticky="we", padx=6, pady=6)
            parent.grid_columnconfigure(1, weight=1)
            entries[key] = w

        add_field(
            frame_personal,
            0,
            "tipo_documento",
            "Tipo de documento",
            "combobox",
            ["CC", "CE", "PPT", "PA"],
        )
        add_field(frame_personal, 1, "documento", "Número de documento")
        add_field(frame_personal, 2, "nombre", "Nombre completo")
        add_field(
            frame_personal,
            3,
            "sexo",
            "Sexo",
            "combobox",
            ["Masculino", "Femenino"],
        )
        add_field(frame_personal, 4, "fecha_nacimiento", "Fecha de nacimiento")

        add_field(frame_contacto, 0, "telefono", "Teléfono")
        add_field(frame_contacto, 1, "correo", "Correo institucional")

        add_field(
            frame_laboral,
            0,
            "cargo",
            "Cargo",
            "combobox",
            [
                "Docente",
                "Docente Tutor PTA",
                "Coordinador",
                "Rector",
                "Administrador Web",
                "Orientador Escolar",
                "Secretaria",
            ],
        )
        add_field(
            frame_laboral,
            1,
            "jornada",
            "Jornada",
            "combobox",
            ["Mañana", "Tarde", "Nocturna"],
        )
        add_field(frame_laboral, 2, "sede", "Sede")
        add_field(
            frame_laboral,
            3,
            "estado",
            "Estado",
            "combobox",
            ["Activo", "Inactivo", "Retirado"],
        )

        if not inicial:
            entries["tipo_documento"].set("CC")
            entries["cargo"].set("Docente")
            entries["jornada"].set("Mañana")
            entries["estado"].set("Activo")

        if inicial:
            for k, v in inicial.items():
                if k in entries and v is not None:
                    if isinstance(entries[k], ttk.Combobox):
                        entries[k].set(str(v))
                    else:
                        entries[k].insert(0, str(v))

        def on_guardar():
            datos = {k: entries[k].get().strip() for k in entries}

            if not datos["documento"]:
                messagebox.showerror(
                    "Validación",
                    "El documento es obligatorio.",
                    parent=d,
                )
                return
            if not datos["documento"].isdigit():
                messagebox.showerror(
                    "Validación",
                    "El documento solo debe contener números.",
                    parent=d,
                )
                return
            if not datos["nombre"]:
                messagebox.showerror(
                    "Validación",
                    "El nombre es obligatorio.",
                    parent=d,
                )
                return
            if datos["correo"]:
                patron_correo = r"^[^@\s]+@[^@\s]+\.[^@\s]+$"
                if not re.match(patron_correo, datos["correo"]):
                    messagebox.showerror(
                        "Validación",
                        "Correo inválido.",
                        parent=d,
                    )
                    return

            doc_original = inicial.get("documento") if inicial else None
            if self._documento_docente_existe(
                datos["documento"], excluir_documento=doc_original
            ):
                messagebox.showerror(
                    "Validación",
                    "El documento ya está registrado",
                    parent=d,
                )
                return

            d.result = datos
            d.destroy()

        frame_botones = ttk.Frame(frame_main)
        frame_botones.pack(fill="x", pady=(14, 0))

        ttk.Button(frame_botones, text="Guardar Docente", command=on_guardar).pack(
            side="left"
        )
        ttk.Button(frame_botones, text="Cancelar", command=d.destroy).pack(
            side="left", padx=(8, 0)
        )

        d.wait_window()
        return getattr(d, "result", None)

    # ---------- Estudiantes ----------

    def _build_estudiantes_tab(self):
        frame = self.tab_estudiantes
        frame_busqueda = ttk.Frame(frame)
        frame_busqueda.pack(fill="x")

        # Botones principales
        self._crear_boton_si_permiso(
            ttk.Button,
            frame_busqueda,
            "desktop.superadmin.matricula.importar",
            text="Importar masivo",
            command=self.estudiantes_importar,
            layout_kwargs={"side": "left"},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            frame_busqueda,
            "desktop.superadmin.matricula.crear",
            text="Agregar",
            command=self.estudiante_agregar,
            layout_kwargs={"side": "left"},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            frame_busqueda,
            "desktop.superadmin.matricula.editar",
            text="Editar",
            command=self.estudiante_editar,
            layout_kwargs={"side": "left"},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            frame_busqueda,
            "desktop.superadmin.matricula.cambiar_curso",
            text="Cambiar de curso",
            command=self.estudiante_cambiar_curso,
            layout_kwargs={"side": "left"},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            frame_busqueda,
            "desktop.superadmin.matricula.eliminar",
            text="Eliminar",
            command=self.estudiante_eliminar,
            layout_kwargs={"side": "left"},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            frame_busqueda,
            "desktop.superadmin.matricula.vaciar",
            text="Vaciar matrícula",
            command=self.vaciar_matricula,
            layout_kwargs={"side": "left"},
        )

        # Filtros
        ttk.Label(frame_busqueda, text="Jornada:").pack(side="left", padx=(10, 0))
        self.filter_jornada = ttk.Combobox(
            frame_busqueda,
            values=["Todos", "Mañana", "Tarde", "Nocturna"],
            state="readonly",
            width=10,
        )
        self.filter_jornada.set("Todos")
        self.filter_jornada.pack(side="left")
        self.filter_jornada.bind(
            "<<ComboboxSelected>>", lambda e: self._load_estudiantes()
        )

        ttk.Label(frame_busqueda, text="Grado:").pack(side="left", padx=(10, 0))
        self.filter_grado = ttk.Combobox(
            frame_busqueda, values=["Todos"], state="readonly", width=8
        )
        self.filter_grado.set("Todos")
        self.filter_grado.pack(side="left")
        self.filter_grado.bind(
            "<<ComboboxSelected>>", lambda e: self._load_estudiantes()
        )

        ttk.Label(frame_busqueda, text="Curso:").pack(side="left", padx=(10, 0))
        self.filter_curso = ttk.Combobox(
            frame_busqueda, values=["Todos"], state="readonly", width=8
        )
        self.filter_curso.set("Todos")
        self.filter_curso.pack(side="left")
        self.filter_curso.bind(
            "<<ComboboxSelected>>", lambda e: self._load_estudiantes()
        )

        ttk.Label(frame_busqueda, text="Buscar:").pack(side="left", padx=(10, 0))
        self.filter_nombre = ttk.Entry(frame_busqueda, width=24)
        self.filter_nombre.pack(side="left")
        self.filter_nombre.bind("<KeyRelease>", lambda e: self._load_estudiantes())

        btn_plantilla = ttk.Button(
            frame_busqueda, text=" Plantilla", command=self.generar_plantilla_excel
        )
        btn_plantilla.pack(side="left", padx=(8, 0))

        # Frame para la tabla y scrollbars
        self.frame_tabla = ttk.Frame(frame)
        self.frame_tabla.pack(fill="both", expand=True)

        # Crear la tabla de estudiantes al construir la pestaña
        cols = (
            "sede",
            "jornada",
            "grado",
            "curso",
            "codigo",
            "apellido1",
            "apellido2",
            "nombre1",
            "nombre2",
            "tipodoc",
            "documento",
            "lugar_expedicion",
            "fecha_expedicion",
            "fecha_nacimiento",
            "lugar_nacimiento",
            "telefono",
            "celular",
            "correo",
            "genero",
            "tipo_sangre",
            "estado_academico",
        )
        self.tree_est = ttk.Treeview(self.frame_tabla, columns=cols, show="headings")
        col_widths = {
            "codigo": 80,
            "apellido1": 110,
            "apellido2": 110,
            "nombre1": 110,
            "nombre2": 110,
            "jornada": 80,
            "grado": 70,
            "curso": 70,
            "tipodoc": 120,
            "documento": 100,
            "lugar_expedicion": 120,
            "fecha_expedicion": 100,
            "fechana": 90,
            "lugar_nacimiento": 120,
            "telefono": 90,
            "celular": 90,
            "email": 150,
            "genero": 80,
            "tipo_sangre": 80,
            "estado": 90,
        }
        for c in cols:
            # Personalización de nombres de columnas
            if c == "correo":
                nombre_col = "Email"
            elif c == "estado_academico":
                nombre_col = "Estado"
            elif c == "codigo":
                nombre_col = "A.idalumno"
            elif c == "fecha_nacimiento":
                nombre_col = "Fecha Nacimiento"
            else:
                nombre_col = c.replace("_", " ").title()
            self.tree_est.heading(c, text=nombre_col)
            self.tree_est.column(
                c,
                width=col_widths.get(
                    c.replace("estado_academico", "estado").replace("correo", "email"),
                    100,
                ),
            )
        vsb = ttk.Scrollbar(
            self.frame_tabla, orient="vertical", command=self.tree_est.yview
        )
        hsb = ttk.Scrollbar(
            self.frame_tabla, orient="horizontal", command=self.tree_est.xview
        )
        self.tree_est.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree_est.pack(fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self._load_estudiantes()

    def vaciar_matricula(self):
        """
        Elimina todos los estudiantes de la base de datos y limpia únicamente el Treeview de matrícula,
        sin recargar ni destruir widgets ni afectar otros módulos.
        Incluye confirmación, manejo de errores y mensaje de éxito.
        """
        if not messagebox.askyesno(
            "Confirmar acción",
            "¿Está seguro que desea vaciar la matrícula?\n\nEsta acción eliminará TODOS los estudiantes de la base de datos y de la tabla actual.\n\nEsta acción NO se puede deshacer.",
            parent=self.win,
        ):
            return
        try:
            # Eliminar todos los estudiantes de la base de datos
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute("DELETE FROM estudiantes")
            conn.commit()
            conn.close()
        except Exception as e:
            messagebox.showerror(
                "Error al vaciar matrícula",
                f"Ocurrió un error al eliminar los estudiantes de la base de datos:\n{e}",
                parent=self.win,
            )
            return
        # Limpiar solo el Treeview de estudiantes
        if hasattr(self, "tree_est"):
            for item in self.tree_est.get_children():
                self.tree_est.delete(item)
        messagebox.showinfo(
            "Matrícula vaciada",
            "Todos los estudiantes han sido eliminados correctamente.",
            parent=self.win,
        )

    def generar_plantilla_excel(self):
        import pandas as pd
        from tkinter import filedialog, messagebox

        # Encabezados exactamente igual a los que reconoce el importador (sincronizado)
        columnas = [
            "Sede",
            "Jornada",
            "Grado",
            "Curso",
            "apellido1",
            "apellido2",
            "nombre1",
            "nombre2",
            "tipodoc",
            "dociden",
            "fechana",
            "telefono",
            "celular",
            "email",
            "genero",
            "tipo_sangre",
            "estado",
            "A.idalumno",
            "nom_muni_exp",
            "nom_lugarnaci",
            "fecha_expe",
        ]
        df = pd.DataFrame(columns=columnas)
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Guardar plantilla de estudiantes",
            initialfile="plantilla_estudiantes.xlsx",
            parent=self.win,
        )
        if not file_path:
            return
        try:
            df.to_excel(file_path, index=False)
            messagebox.showinfo(
                "Plantilla generada", f"Archivo guardado en:\n{file_path}"
            )
        except Exception as e:
            messagebox.showerror("Error al guardar", str(e))

    def _load_estudiantes(self):
        # Orden unificado con el formulario de edición (botón editar)
        cols_est = [
            "sede",
            "jornada",
            "grado",
            "curso",
            "codigo",
            "apellido1",
            "apellido2",
            "nombre1",
            "nombre2",
            "tipodoc",
            "documento",
            "lugar_expedicion",
            "fecha_expedicion",
            "fecha_nacimiento",
            "lugar_nacimiento",
            "telefono",
            "celular",
            "correo",
            "genero",
            "tipo_sangre",
            "estado_academico",
        ]

        registros = []
        try:
            # Cargar estudiantes desde core (desacoplado de SQL)
            filas_db = core_matricula.cargar_todos_estudiantes_dataframe()

            for row in filas_db:
                registro = {}
                for idx, col in enumerate(cols_est):
                    valor = row.get(col, "") if isinstance(row, dict) else row[idx]
                    # Normalización específica para algunos campos
                    if col == "jornada":
                        valor = convertir_jornada(str(valor).strip())
                    elif col == "tipodoc":
                        valor = convertir_tipo_documento(str(valor).strip())
                    elif col == "genero":
                        valor = convertir_genero(str(valor).strip())
                    elif col == "estado_academico":
                        valor = str(valor).strip() or "Matriculado"
                        if valor.lower() in ("activo", "ma", "matriculado", ""):
                            valor = "Matriculado"
                    else:
                        valor = str(valor).strip()
                    registro[col] = valor
                registros.append(registro)
        except Exception:
            registros = []

        if _HAS_PANDAS:
            df = pd.DataFrame(registros, columns=cols_est)
        else:
            df = {
                "_headers": tuple(cols_est),
                "_rows": registros,
            }

        self.estudiantes_df = df
        # si la interfaz aún no se construyó (por ejemplo en pruebas que no
        # llaman a ``open_interface``) simplemente almacenamos el dataframe
        # y regresamos sin intentar tocar widgets inexistentes.
        if not hasattr(self, "tree_est"):
            return
        for i in self.tree_est.get_children():
            self.tree_est.delete(i)

        # Si no hay registros, mostrar una fila de ejemplo para que la tabla sea visible
        if not registros:
            ejemplo = {
                "sede": "Ejemplo",
                "jornada": "Mañana",
                "grado": "1",
                "curso": "A",
                "codigo": "0001",
                "apellido1": "Apellido1",
                "apellido2": "Apellido2",
                "nombre1": "Nombre1",
                "nombre2": "Nombre2",
                "tipodoc": "TI",
                "documento": "123456",
                "lugar_expedicion": "Ciudad",
                "fecha_expedicion": "2020-01-01",
                "fecha_nacimiento": "2010-01-01",
                "lugar_nacimiento": "Ciudad",
                "telefono": "1234567",
                "celular": "3001234567",
                "correo": "ejemplo@correo.com",
                "genero": "M",
                "tipo_sangre": "O+",
                "estado_academico": "Matriculado",
            }
            registros.append(ejemplo)

        # actualizar valores de filtros si existen
        if _HAS_PANDAS:
            if hasattr(self, "filter_jornada"):
                jornadas = sorted(
                    {str(x).strip() for x in df["jornada"].dropna().unique()}
                )
                self.filter_jornada["values"] = ["Todos"] + jornadas
                if self.filter_jornada.get() not in self.filter_jornada["values"]:
                    self.filter_jornada.set("Todos")
            if hasattr(self, "filter_grado"):
                grados = sorted({str(x).strip() for x in df["grado"].dropna().unique()})
                self.filter_grado["values"] = ["Todos"] + grados
                if self.filter_grado.get() not in self.filter_grado["values"]:
                    self.filter_grado.set("Todos")
            if hasattr(self, "filter_curso"):
                cursos = sorted({str(x).strip() for x in df["curso"].dropna().unique()})
                self.filter_curso["values"] = ["Todos"] + cursos
                if self.filter_curso.get() not in self.filter_curso["values"]:
                    self.filter_curso.set("Todos")

        # aplicar filtros seleccionados
        jornada_sel = None
        grado_sel = None
        curso_sel = None
        nombre_sel = None
        if hasattr(self, "filter_jornada"):
            jornada_sel = self.filter_jornada.get()
        if hasattr(self, "filter_grado"):
            grado_sel = self.filter_grado.get()
        if hasattr(self, "filter_curso"):
            curso_sel = self.filter_curso.get()
        if hasattr(self, "filter_nombre"):
            nombre_sel = self.filter_nombre.get().strip().lower()

        def pasa_filtro(row):
            if jornada_sel and jornada_sel != "Todos":
                if str(row.get("jornada", "")).strip() != jornada_sel:
                    return False
            if grado_sel and grado_sel != "Todos":
                if str(row.get("grado", "")).strip() != grado_sel:
                    return False
            if curso_sel and curso_sel != "Todos":
                if str(row.get("curso", "")).strip() != curso_sel:
                    return False
            if nombre_sel:
                # Buscar en nombre1, nombre2, apellido1, apellido2 y documento
                campos = [
                    str(row.get("nombre1", "")).lower(),
                    str(row.get("nombre2", "")).lower(),
                    str(row.get("apellido1", "")).lower(),
                    str(row.get("apellido2", "")).lower(),
                    str(row.get("documento", "")).lower(),
                ]
                if not any(nombre_sel in campo for campo in campos):
                    return False
            return True

        cols_display = (
            "sede",
            "jornada",
            "grado",
            "curso",
            "codigo",
            "apellido1",
            "apellido2",
            "nombre1",
            "nombre2",
            "tipodoc",
            "documento",
            "lugar_expedicion",
            "fecha_expedicion",
            "fecha_nacimiento",
            "lugar_nacimiento",
            "telefono",
            "celular",
            "correo",
            "genero",
            "tipo_sangre",
            "estado_academico",
        )

        def clean_val(val):
            if _HAS_PANDAS:
                try:
                    if pd.isna(val):
                        return ""
                except Exception:
                    pass
            if val is None:
                return ""
            s = str(val)
            if s.lower() == "nan":
                return ""
            # eliminar hora de fechas en formato ISO
            if re.match(r"^\d{4}-\d{2}-\d{2} ", s):
                s = s.split(" ")[0]
            return s

        def build_nombre(row):
            get = (
                row.get if hasattr(row, "get") else lambda k, d=None: getattr(row, k, d)
            )
            nombre1 = get("nombre1", "").strip()
            nombre2 = get("nombre2", "").strip()
            apellido1 = get("apellido1", "").strip()
            apellido2 = get("apellido2", "").strip()
            partes = [apellido1, apellido2, nombre1, nombre2]
            return " ".join([p for p in partes if p]).strip()

        if _HAS_PANDAS:
            for _, r in df.iterrows():
                if not pasa_filtro(r):
                    continue
                vals = []
                for c in cols_display:
                    if c == "nombre_estudiante":
                        vals.append(build_nombre(r))
                    else:
                        vals.append(clean_val(r.get(c, "")))
                self.tree_est.insert("", "end", values=tuple(vals))
        else:
            for r in df.get("_rows", []):
                if not pasa_filtro(r):
                    continue
                vals = []
                for c in cols_display:
                    if c == "nombre_estudiante":
                        vals.append(build_nombre(r))
                    else:
                        vals.append(clean_val(r.get(c, "")))
                self.tree_est.insert("", "end", values=tuple(vals))

    def _save_estudiantes(self):
        if _HAS_PANDAS:
            rows = self.estudiantes_df.to_dict(orient="records")
        else:
            rows = list(self.estudiantes_df.get("_rows", []))

        anio_lectivo_activo = self.obtener_anio_lectivo_activo()

        # Normalizar campos para backend
        def build_nombre(row):
            return "{} {} {} {}".format(
                str(row.get("nombre1", "")).strip(),
                str(row.get("nombre2", "")).strip(),
                str(row.get("apellido1", "")).strip(),
                str(row.get("apellido2", "")).strip(),
            ).strip()

        rows_norm = []
        for row in rows:
            row_norm = dict(row)  # copia
            # Unir nombres y apellidos
            row_norm["nombre"] = build_nombre(row)
            # Renombrar campos
            if "fecha_nacimiento" in row:
                row_norm["fechana"] = row["fecha_nacimiento"]
            if "correo" in row:
                row_norm["email"] = row["correo"]
            if "estado_academico" in row:
                row_norm["estado"] = row["estado_academico"]
            rows_norm.append(row_norm)

        # Usar servicio core para sincronizar estudiantes
        core_matricula.sincronizar_estudiantes(
            rows_norm, anio_lectivo=anio_lectivo_activo
        )

    def estudiantes_importar(self):
        if not self._requiere_permiso("desktop.superadmin.matricula.importar"):
            return
        path = filedialog.askopenfilename(
            title="Seleccionar archivo para importar",
            filetypes=[
                ("Archivos Excel o CSV", "*.xlsx;*.xls;*.csv"),
                ("Excel", "*.xlsx;*.xls"),
                ("CSV", "*.csv"),
                ("Todos los archivos", "*.*"),
            ],
        )

        # Definición de columnas requeridas y alias
        required_cols = [
            "sede",
            "jornada",
            "grado",
            "curso",
            "codigo",
            "apellido1",
            "apellido2",
            "nombre1",
            "nombre2",
            "tipodoc",
            "documento",
            "lugar_expedicion",
            "fecha_expedicion",
            "fecha_nacimiento",
            "lugar_nacimiento",
            "telefono",
            "celular",
            "correo",
            "genero",
            "tipo_sangre",
            "estado_academico",
        ]

        def alias_variants(*names):
            variants = set()
            for n in names:
                variants.add(n)
                variants.add(n.lower())
                variants.add(n.capitalize())
                variants.add(n.upper())
            return list(variants)

        col_aliases = {
            "sede": alias_variants("sede", "Sede"),
            "jornada": alias_variants("jornada", "Jornada"),
            "grado": alias_variants("grado", "Grado"),
            "curso": alias_variants("curso", "Curso"),
            "codigo": alias_variants("codigo", "A.idalumno", "aidalumno"),
            "apellido1": alias_variants("apellido1"),
            "apellido2": alias_variants("apellido2"),
            "nombre1": alias_variants("nombre1"),
            "nombre2": alias_variants("nombre2"),
            "tipodoc": alias_variants("tipodoc"),
            "documento": alias_variants("documento", "dociden"),
            "lugar_expedicion": alias_variants(
                "lugar_expedicion", "nom_muni_exp", "nommuniexp"
            ),
            "fecha_expedicion": alias_variants(
                "fecha_expedicion", "fecha_expe", "fechaexpe"
            ),
            "fecha_nacimiento": alias_variants("fecha_nacimiento", "fechana"),
            "lugar_nacimiento": alias_variants(
                "lugar_nacimiento", "nom_lugarnaci", "nomlugarnaci"
            ),
            "telefono": alias_variants("telefono"),
            "celular": alias_variants("celular"),
            "correo": alias_variants("correo", "email"),
            "genero": alias_variants("genero"),
            "tipo_sangre": alias_variants("tipo_sangre"),
            "estado_academico": alias_variants("estado_academico", "estado"),
        }

        def norm(x):
            return (
                str(x)
                .lower()
                .replace(" ", "")
                .replace("_", "")
                .replace(".", "")
                .replace("á", "a")
                .replace("é", "e")
                .replace("í", "i")
                .replace("ó", "o")
                .replace("ú", "u")
                .strip()
            )

        import pandas as pd

        if path.lower().endswith(".csv"):
            try:
                incoming = pd.read_csv(path)
            except Exception as e:
                messagebox.showerror("Error al leer CSV", str(e), parent=self.win)
                return
        else:
            incoming = self._read_excel(path)
        if incoming is None:
            messagebox.showerror(
                "Importar", "Archivo vacío o inválido.", parent=self.win
            )
            return

        incoming.columns = incoming.columns.str.strip()
        cols_presentes = list(incoming.columns)
        norm_presentes = [norm(c) for c in cols_presentes]
        # DEBUG: Mostrar encabezados detectados
        try:
            import tkinter
            from tkinter import messagebox

            messagebox.showinfo(
                "Encabezados detectados",
                "\n".join(
                    [
                        f"{i+1}. '{c}' → '{norm(c)}'"
                        for i, c in enumerate(cols_presentes)
                    ]
                ),
                parent=self.win,
            )
        except Exception:
            print("Encabezados detectados:", cols_presentes)

        # Construir el mapeo de columnas: requerido → presente
        col_map = {}
        for req, aliases in col_aliases.items():
            found = None
            for alias in aliases:
                if norm(alias) in norm_presentes:
                    idx = norm_presentes.index(norm(alias))
                    found = cols_presentes[idx]
                    break
            if found:
                col_map[req] = found
        print("[DEBUG] col_map:", col_map)
        # Verificar faltantes
        faltan = [c for c in required_cols if c not in col_map]
        if faltan:
            messagebox.showwarning(
                "Importar",
                f"El archivo debe tener las siguientes columnas (se aceptan nombres alternativos):\n\nFaltan:\n{', '.join(faltan)}",
                parent=self.win,
            )
            return

        def limpiar(valor):
            if pd.isna(valor):
                return ""
            s = str(valor).strip()
            return "" if s.lower() == "nan" else s

        documentos = set()
        filas_validas = []
        for _, row in incoming.iterrows():
            doc = limpiar(row[col_map["documento"]])
            if not doc or doc in documentos:
                continue
            documentos.add(doc)
            fila = {k: limpiar(row[col_map[k]]) for k in required_cols}
            filas_validas.append(fila)
        if not filas_validas:
            messagebox.showwarning(
                "Importar",
                "No hay registros válidos para importar.",
                parent=self.win,
            )
            return
        df_validado = pd.DataFrame(filas_validas, columns=required_cols)
        self.estudiantes_df = pd.concat(
            [self.estudiantes_df, df_validado], ignore_index=True
        )

        self._save_estudiantes()
        self._load_estudiantes()
        messagebox.showinfo(
            "Importar", "Estudiantes importados correctamente.", parent=self.win
        )

    def estudiante_agregar(self):
        if not self._requiere_permiso("desktop.superadmin.matricula.crear"):
            return
        datos = self._dialog_estudiante()
        if not datos:
            return
        documento = datos.get("documento", "").strip()
        if not documento:
            messagebox.showerror(
                "Error", "El documento es obligatorio.", parent=self.win
            )
            return
        try:
            # Eliminar clave 'nombre' si existe para evitar error de argumento inesperado
            if "nombre" in datos:
                datos.pop("nombre")
            # Unificar nombres y pasar todos los campos requeridos
            core_matricula.crear_o_actualizar_estudiante(
                documento=documento,
                sede=datos.get("sede", "").strip(),
                jornada=datos.get("jornada", "").strip(),
                grado=datos.get("grado", "").strip(),
                curso=datos.get("curso", "").strip(),
                codigo=datos.get("codigo", "").strip(),
                apellido1=datos.get("apellido1", "").strip(),
                apellido2=datos.get("apellido2", "").strip(),
                nombre1=datos.get("nombre1", "").strip(),
                nombre2=datos.get("nombre2", "").strip(),
                tipodoc=datos.get("tipodoc", "").strip(),
                lugar_expedicion=datos.get("lugar_expedicion", "").strip(),
                fecha_expedicion=datos.get("fecha_expedicion", "").strip(),
                fecha_nacimiento=datos.get("fechana", "").strip(),
                lugar_nacimiento=datos.get("lugar_nacimiento", "").strip(),
                telefono=datos.get("telefono", "").strip(),
                celular=datos.get("celular", "").strip(),
                correo=datos.get("email", "").strip(),
                genero=datos.get("genero", "").strip(),
                tipo_sangre=datos.get("tipo_sangre", "").strip(),
                estado_academico=datos.get("estado", "").strip(),
            )
            self._load_estudiantes()
            messagebox.showinfo(
                "Éxito", "Estudiante agregado correctamente.", parent=self.win
            )
        except Exception as e:
            messagebox.showerror(
                "Error", f"No se pudo agregar estudiante: {str(e)}", parent=self.win
            )

    def estudiante_editar(self):
        if not self._requiere_permiso("desktop.superadmin.matricula.editar"):
            return
        sel = self.tree_est.selection()
        if not sel:
            messagebox.showinfo("Editar", "Seleccione un estudiante.", parent=self.win)
            return
        vals = self.tree_est.item(sel[0])["values"]
        # El orden de columnas del Treeview (sin columna legacy de nombre, pero con sede al final)
        cols_tree = [
            "sede",
            "jornada",
            "grado",
            "curso",
            "codigo",
            "apellido1",
            "apellido2",
            "nombre1",
            "nombre2",
            "tipodoc",
            "documento",
            "lugar_expedicion",
            "fecha_expedicion",
            "fecha_nacimiento",
            "lugar_nacimiento",
            "telefono",
            "celular",
            "correo",
            "genero",
            "tipo_sangre",
            "estado_academico",
        ]
        actuales = {
            col: vals[i] if i < len(vals) else "" for i, col in enumerate(cols_tree)
        }
        # El formulario espera los campos con estos nombres:
        # codigo, apellido1, apellido2, nombre1, nombre2, jornada, grado, curso, tipodoc, documento, lugar_expedicion, fecha_expedicion, fechana, lugar_nacimiento, telefono, celular, email, genero, tipo_sangre, estado, sede
        nuevos = self._dialog_estudiante(actuales)
        if not nuevos:
            return

        documento = actuales.get("documento", "").strip()
        if not documento:
            messagebox.showerror(
                "Error", "No se encontró el documento del estudiante.", parent=self.win
            )
            return

        try:
            # Eliminar clave 'nombre' si existe para evitar error de argumento inesperado
            if "nombre" in nuevos:
                nuevos.pop("nombre")
            core_matricula.crear_o_actualizar_estudiante(
                documento=documento,
                sede=nuevos.get("sede", "").strip(),
                jornada=nuevos.get("jornada", "").strip(),
                grado=nuevos.get("grado", "").strip(),
                curso=nuevos.get("curso", "").strip(),
                codigo=nuevos.get("codigo", "").strip(),
                apellido1=nuevos.get("apellido1", "").strip(),
                apellido2=nuevos.get("apellido2", "").strip(),
                nombre1=nuevos.get("nombre1", "").strip(),
                nombre2=nuevos.get("nombre2", "").strip(),
                tipodoc=nuevos.get("tipodoc", "").strip(),
                lugar_expedicion=nuevos.get("lugar_expedicion", "").strip(),
                fecha_expedicion=nuevos.get("fecha_expedicion", "").strip(),
                fecha_nacimiento=nuevos.get("fechana", "").strip(),
                lugar_nacimiento=nuevos.get("lugar_nacimiento", "").strip(),
                telefono=nuevos.get("telefono", "").strip(),
                celular=nuevos.get("celular", "").strip(),
                correo=nuevos.get("email", "").strip(),
                genero=nuevos.get("genero", "").strip(),
                tipo_sangre=nuevos.get("tipo_sangre", "").strip(),
                estado_academico=nuevos.get("estado", "").strip(),
            )
            self._load_estudiantes()
            messagebox.showinfo(
                "Éxito", "Estudiante actualizado correctamente.", parent=self.win
            )
        except Exception as e:
            messagebox.showerror(
                "Error", f"No se pudo actualizar estudiante: {str(e)}", parent=self.win
            )

    def estudiante_cambiar_curso(self):
        """Permite cambiar el grado y/o curso de uno o varios estudiantes."""
        if not self._requiere_permiso("desktop.superadmin.matricula.cambiar_curso"):
            return
        sel = self.tree_est.selection()
        if not sel:
            messagebox.showinfo(
                "Cambiar de curso",
                "Seleccione al menos un estudiante.",
                parent=self.win,
            )
            return

        # Crear diálogo para seleccionar nuevo grado y curso
        d = tk.Toplevel(self.win)
        d.transient(self.win)
        d.grab_set()
        d.title("Cambiar de curso")
        d.geometry("400x250")

        tk.Label(d, text="Nuevo Grado:", font=("Arial", 10)).grid(
            row=0, column=0, sticky="e", padx=10, pady=10
        )
        entry_grado = ttk.Entry(d, width=30)
        entry_grado.grid(row=0, column=1, sticky="ew", padx=10, pady=10)

        tk.Label(d, text="Nuevo Curso:", font=("Arial", 10)).grid(
            row=1, column=0, sticky="e", padx=10, pady=10
        )
        entry_curso = ttk.Entry(d, width=30)
        entry_curso.grid(row=1, column=1, sticky="ew", padx=10, pady=10)

        tk.Label(d, text="Nueva Jornada (opcional):", font=("Arial", 10)).grid(
            row=2, column=0, sticky="e", padx=10, pady=10
        )
        combo_jornada = ttk.Combobox(
            d, values=["Mañana", "Tarde", "Nocturna"], state="readonly", width=28
        )
        combo_jornada.grid(row=2, column=1, sticky="ew", padx=10, pady=10)

        d.columnconfigure(1, weight=1)

        def aplicar_cambios():
            nuevo_grado = entry_grado.get().strip()
            nuevo_curso = entry_curso.get().strip()
            nueva_jornada = combo_jornada.get()

            if not nuevo_grado or not nuevo_curso:
                messagebox.showerror(
                    "Validación", "Grado y curso son obligatorios.", parent=d
                )
                return

            try:
                # Aplicar cambios a los estudiantes seleccionados usando core.matricula
                for item in sel:
                    vals = self.tree_est.item(item)["values"]
                    documento = str(vals[5]).strip()
                    # Remover .0 si existe (conversión numérica excel)
                    documento = re.sub(r"\.0+$", "", documento)
                    core_matricula.cambiar_grado_curso(
                        documento=documento,
                        nuevo_grado=nuevo_grado,
                        nuevo_curso=nuevo_curso,
                        nueva_jornada=nueva_jornada if nueva_jornada else None,
                    )
                self._load_estudiantes()
                messagebox.showinfo(
                    "Éxito",
                    f"Se cambió de curso a {len(sel)} estudiante(s).",
                    parent=self.win,
                )
                d.destroy()
            except Exception as e:
                messagebox.showerror(
                    "Error", f"No se pudieron cambiar de curso: {str(e)}", parent=d
                )

        btn_frame = ttk.Frame(d)
        btn_frame.grid(row=3, column=0, columnspan=2, pady=20)

        ttk.Button(btn_frame, text="Aplicar", command=aplicar_cambios).pack(
            side="left", padx=5
        )
        ttk.Button(btn_frame, text="Cancelar", command=d.destroy).pack(
            side="left", padx=5
        )

    def estudiante_eliminar(self):
        if not self._requiere_permiso("desktop.superadmin.matricula.eliminar"):
            return
        sel = self.tree_est.selection()
        if not sel:
            messagebox.showinfo(
                "Eliminar", "Seleccione un estudiante.", parent=self.win
            )
            return

        vals = self.tree_est.item(sel[0])["values"]
        # El índice de documento depende del orden de columnas en cols_display
        # Buscar el índice de la columna "documento" en self.tree_est["columns"]
        columns = self.tree_est["columns"]
        idx_documento = 10  # documento siempre en la posición 10
        nombre_muestra = vals[7] if len(vals) > 7 else ""
        documento_val = (
            str(vals[idx_documento]).strip() if len(vals) > idx_documento else ""
        )
        # Remover .0 si existe (conversión numérica excel)
        documento_val = re.sub(r"\.0+$", "", documento_val)

        if not messagebox.askyesno(
            "Confirmar", f"Eliminar estudiante {nombre_muestra}?", parent=self.win
        ):
            return

        try:
            core_matricula.eliminar_estudiante(documento=documento_val)
            self._load_estudiantes()
            messagebox.showinfo(
                "Éxito", f"Estudiante eliminado correctamente.", parent=self.win
            )
        except Exception as e:
            messagebox.showerror(
                "Error", f"No se pudo eliminar estudiante: {str(e)}", parent=self.win
            )

    def _dialog_estudiante(self, inicial=None):
        d = tk.Toplevel(self.win)
        d.transient(self.win)
        d.grab_set()
        d.title("Estudiante")
        d.geometry("500x700")

        # Frame con scrollbar
        canvas = tk.Canvas(d, highlightthickness=0)
        scrollbar = ttk.Scrollbar(d, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.grid(row=0, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
        scrollbar.grid(row=0, column=2, sticky="ns")
        d.grid_rowconfigure(0, weight=1)
        d.grid_columnconfigure(0, weight=1)

        # Campos del formulario (nuevo esquema de Excel)
        fields = [
            ("sede", "Sede", "entry", None),
            ("jornada", "Jornada", "combobox", ["Mañana", "Tarde", "Nocturna"]),
            ("grado", "Grado", "entry", None),
            ("curso", "Curso", "entry", None),
            ("codigo", "Código estudiante", "entry", None),
            ("apellido1", "Primer apellido", "entry", None),
            ("apellido2", "Segundo apellido", "entry", None),
            ("nombre1", "Primer nombre", "entry", None),
            ("nombre2", "Segundo nombre", "entry", None),
            (
                "tipodoc",
                "Tipo de documento",
                "combobox",
                [
                    "Registro civil de nacimiento",
                    "Tarjeta de identidad",
                    "Número único de identificación personal",
                    "Cédula de ciudadanía",
                    "Cédula de extranjería",
                    "Permiso de protección temporal",
                    "Permiso especial permanencia",
                    "Registro único de migrantes venezolanos",
                    "Pasaporte",
                    "Partida de nacimiento",
                    "Número de identificación personal",
                    "Número establecido por la secretaría",
                    "Tarjeta movilidad fronteriza",
                    "Certificado de cabildo",
                    "Visa",
                ],
            ),
            ("documento", "Número de documento", "entry", None),
            ("lugar_expedicion", "Lugar de expedición", "entry", None),
            ("fecha_expedicion", "Fecha de expedición", "entry", None),
            ("fechana", "Fecha de nacimiento", "entry", None),
            ("lugar_nacimiento", "Lugar de nacimiento", "entry", None),
            ("telefono", "Teléfono", "entry", None),
            ("celular", "Celular", "entry", None),
            ("email", "Correo electrónico", "entry", None),
            ("genero", "Género", "combobox", ["Masculino", "Femenino"]),
            ("tipo_sangre", "Tipo de sangre", "entry", None),
            (
                "estado",
                "Estado académico",
                "combobox",
                ["Matriculado", "Retirado", "Graduado", "Trasladado"],
            ),
        ]

        entries = {}
        row = 0
        for key, label, kind, values in fields:
            tk.Label(scrollable_frame, text=label + ":", anchor="w").grid(
                row=row, column=0, sticky="w", padx=8, pady=4
            )
            if kind == "entry":
                entry = tk.Entry(scrollable_frame, width=32)
                entry.grid(row=row, column=1, sticky="we", padx=8, pady=4)
                entries[key] = entry
            elif kind == "combobox":
                cb = ttk.Combobox(
                    scrollable_frame, values=values, state="readonly", width=30
                )
                cb.grid(row=row, column=1, sticky="we", padx=8, pady=4)
                entries[key] = cb
            row += 1

        # Si hay datos iniciales (edición), rellenar los campos
        if inicial:
            for k, v in inicial.items():
                if k in entries:
                    if isinstance(entries[k], ttk.Combobox):
                        entries[k].set(v)
                    else:
                        entries[k].insert(0, v)

        def validate_and_save():
            # Validaciones obligatorias
            doc_val = entries["documento"].get().strip()
            if not doc_val:
                messagebox.showerror(
                    "Validación", "El número de documento es obligatorio.", parent=d
                )
                return

            if not entries["grado"].get().strip():
                messagebox.showerror("Validación", "El grado es obligatorio.", parent=d)
                return

            if not entries["curso"].get().strip():
                messagebox.showerror("Validación", "El curso es obligatorio.", parent=d)
                return

            if not entries["jornada"].get():
                messagebox.showerror(
                    "Validación", "La jornada es obligatoria.", parent=d
                )
                return

            if not entries["estado"].get():
                messagebox.showerror(
                    "Validación", "El estado académico es obligatorio.", parent=d
                )
                return

            # Validaciones nuevos campos
            if not entries["codigo"].get().strip():
                messagebox.showerror(
                    "Validación", "El código es obligatorio.", parent=d
                )
                return
            if not entries["apellido1"].get().strip():
                messagebox.showerror(
                    "Validación", "El primer apellido es obligatorio.", parent=d
                )
                return
            if not entries["nombre1"].get().strip():
                messagebox.showerror(
                    "Validación", "El primer nombre es obligatorio.", parent=d
                )
                return

            datos = {}
            for key, _, _, _ in fields:
                widget = entries[key]
                datos[key] = widget.get()

            d.result = datos
            d.destroy()

        button_frame = ttk.Frame(scrollable_frame)
        button_frame.grid(row=row, column=0, columnspan=2, pady=10)
        ttk.Button(button_frame, text="OK", command=validate_and_save).pack(
            side="left", padx=5
        )
        ttk.Button(button_frame, text="Cancelar", command=d.destroy).pack(
            side="left", padx=5
        )

        d.wait_window()
        return getattr(d, "result", None)

        # Frame con scrollbar
        canvas = tk.Canvas(d, highlightthickness=0)
        scrollbar = ttk.Scrollbar(d, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.grid(row=0, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
        scrollbar.grid(row=0, column=2, sticky="ns")
        d.grid_rowconfigure(0, weight=1)
        d.grid_columnconfigure(0, weight=1)

        # Campos del formulario (nuevo esquema de Excel)
        fields = [
            ("sede", "Sede", "entry", None),
            ("jornada", "Jornada", "combobox", ["Mañana", "Tarde", "Nocturna"]),
            ("grado", "Grado", "entry", None),
            ("curso", "Curso", "entry", None),
            ("codigo", "Código estudiante", "entry", None),
            ("apellido1", "Primer apellido", "entry", None),
            ("apellido2", "Segundo apellido", "entry", None),
            ("nombre1", "Primer nombre", "entry", None),
            ("nombre2", "Segundo nombre", "entry", None),
            (
                "tipodoc",
                "Tipo de documento",
                "combobox",
                [
                    "Registro civil de nacimiento",
                    "Tarjeta de identidad",
                    "Número único de identificación personal",
                    "Cédula de ciudadanía",
                    "Cédula de extranjería",
                    "Permiso de protección temporal",
                    "Permiso especial permanencia",
                    "Registro único de migrantes venezolanos",
                    "Pasaporte",
                    "Partida de nacimiento",
                    "Número de identificación personal",
                    "Número establecido por la secretaría",
                    "Tarjeta movilidad fronteriza",
                    "Certificado de cabildo",
                    "Visa",
                ],
            ),
            ("documento", "Número de documento", "entry", None),
            ("lugar_expedicion", "Lugar de expedición", "entry", None),
            ("fecha_expedicion", "Fecha de expedición", "entry", None),
            ("fechana", "Fecha de nacimiento", "entry", None),
            ("lugar_nacimiento", "Lugar de nacimiento", "entry", None),
            ("telefono", "Teléfono", "entry", None),
            ("celular", "Celular", "entry", None),
            ("email", "Correo electrónico", "entry", None),
            ("genero", "Género", "combobox", ["Masculino", "Femenino"]),
            ("tipo_sangre", "Tipo de sangre", "entry", None),
            (
                "estado",
                "Estado académico",
                "combobox",
                ["Matriculado", "Retirado", "Graduado", "Trasladado"],
            ),
        ]

        # ...código original del método continúa aquí...

    # ---------- Banco de Preguntas ----------
    def _build_preguntas_tab(self):
        """Construye la pestaña de Banco de Preguntas (versión profesional)."""
        self._preg_mostrar_filtro_docente = True
        InterfazBancoPreguntasAvanzada._build_preguntas_tab_mejorada(self)
        self._agregar_botones_limpieza_banco()

    def _agregar_botones_limpieza_banco(self):
        """Agrega botones para limpiar por filtros o vaciar el banco completo."""
        if getattr(self, "_botones_limpieza_banco_agregados", False):
            return
        try:
            children = self.tab_preguntas.winfo_children()
            toolbar = children[1] if len(children) > 1 else None
            if toolbar is None:
                return

            ttk.Separator(toolbar, orient="vertical").pack(
                side="left", fill="y", padx=4
            )
            ttk.Button(
                toolbar,
                text="Limpiar Banco (Filtros)",
                command=self.banco_limpiar_filtros,
            ).pack(side="left", padx=2)
            ttk.Button(
                toolbar,
                text="Vaciar Banco Completo",
                command=self.banco_vaciar_completo,
            ).pack(side="left", padx=2)
            ttk.Button(
                toolbar,
                text="Generar con IA",
                command=self.generar_preguntas_con_ia,
            ).pack(side="left", padx=2)

            self._botones_limpieza_banco_agregados = True
        except Exception:
            # Si la UI cambia, no interrumpir carga de la pestaña.
            pass

    def generar_plantilla_excel_banco(self):
        """Genera plantilla Excel segura con catálogos y listas desplegables."""
        path = filedialog.asksaveasfilename(
            title="Guardar plantilla de Banco de Preguntas",
            defaultextension=".xlsx",
            initialfile="plantilla_banco_preguntas.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
            parent=self.win,
        )
        if not path:
            return

        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()

            cur.execute("SELECT COALESCE(MAX(id), 0) FROM banco_preguntas")
            max_id_actual = int(cur.fetchone()[0] or 0)
            conn.close()
        except Exception as exc:
            messagebox.showerror(
                "Plantilla Banco",
                f"No se pudo consultar el banco de preguntas:\n{exc}",
                parent=self.win,
            )
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.worksheet.datavalidation import DataValidation

            wb = Workbook()
            ws_preguntas = wb.active
            ws_preguntas.title = "preguntas"
            ws_catalogos = wb.create_sheet("catalogos")
            ws_instrucciones = wb.create_sheet("instrucciones")

            columnas = [
                "id",
                "grado",
                "area",
                "evaluacion",
                "periodo",
                "id_contexto",
                "contexto",
                "enunciado",
                "tipo_pregunta",
                "opcion_a",
                "opcion_b",
                "opcion_c",
                "opcion_d",
                "correcta",
                "imagen",
            ]
            ws_preguntas.append(columnas)

            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws_preguntas[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # IDs secuenciales con filas en blanco listas para diligenciar.
            total_filas = 50
            for idx in range(total_filas):
                ws_preguntas.append(
                    [
                        max_id_actual + idx + 1,
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "opcion_multiple",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]
                )

            # Hoja de catálogos
            ws_catalogos["A1"] = "tipos_pregunta"
            ws_catalogos["B1"] = "correctas"

            for i, val in enumerate(["opcion_multiple", "abierta"], start=2):
                ws_catalogos.cell(row=i, column=1, value=val)
            for i, val in enumerate(["A", "B", "C", "D"], start=2):
                ws_catalogos.cell(row=i, column=2, value=val)

            last_tipo = 3
            last_correcta = 5

            dv_tipo = DataValidation(
                type="list",
                formula1=f"=catalogos!$A$2:$A${last_tipo}",
                allow_blank=False,
            )
            dv_correcta = DataValidation(
                type="list",
                formula1=f"=catalogos!$B$2:$B${last_correcta}",
                allow_blank=True,
            )

            ws_preguntas.add_data_validation(dv_tipo)
            ws_preguntas.add_data_validation(dv_correcta)

            max_rows_validacion = max(500, total_filas + 200)
            dv_tipo.add(f"I2:I{max_rows_validacion}")
            dv_correcta.add(f"N2:N{max_rows_validacion}")

            ws_preguntas.freeze_panes = "A2"
            ws_preguntas.auto_filter.ref = f"A1:O{total_filas + 1}"
            anchos = {
                "A": 10,
                "B": 12,
                "C": 24,
                "D": 22,
                "E": 12,
                "F": 14,
                "G": 34,
                "H": 48,
                "I": 18,
                "J": 24,
                "K": 24,
                "L": 24,
                "M": 24,
                "N": 12,
                "O": 32,
            }
            for col, ancho in anchos.items():
                ws_preguntas.column_dimensions[col].width = ancho

            instrucciones = [
                ("Campo", "Uso"),
                ("id", "Obligatorio. Debe ser numerico y unico."),
                ("grado", "Opcional."),
                ("area", "Opcional."),
                ("evaluacion", "Obligatorio."),
                ("periodo", "Opcional."),
                (
                    "id_contexto",
                    "Opcional. Sirve para agrupar preguntas de una misma lectura.",
                ),
                ("contexto", "Opcional. Texto base para preguntas asociadas."),
                ("enunciado", "Obligatorio."),
                ("tipo_pregunta", "Use 'opcion_multiple' o 'abierta'."),
                ("opcion_a a opcion_d", "Obligatorias solo para opcion_multiple."),
                (
                    "correcta",
                    "Obligatoria solo para opcion_multiple. Valores A, B, C o D.",
                ),
                ("imagen", "Opcional. Ruta local de la imagen."),
            ]
            for fila_idx, (campo, uso) in enumerate(instrucciones, start=1):
                ws_instrucciones.cell(row=fila_idx, column=1, value=campo)
                ws_instrucciones.cell(row=fila_idx, column=2, value=uso)
            for cell in ws_instrucciones[1]:
                cell.fill = header_fill
                cell.font = header_font
            ws_instrucciones.column_dimensions["A"].width = 22
            ws_instrucciones.column_dimensions["B"].width = 90

            ws_catalogos.sheet_state = "hidden"

            wb.save(path)
        except Exception as exc:
            messagebox.showerror(
                "Plantilla Banco",
                f"No se pudo generar el archivo Excel:\n{exc}",
                parent=self.win,
            )
            return

        messagebox.showinfo(
            "Plantilla Banco",
            f"Plantilla generada con {total_filas} fila(s) listas para diligenciar.",
            parent=self.win,
        )

    def generar_preguntas_con_ia(self):
        """Abre formulario para generar preguntas de selección múltiple con IA."""
        d = tk.Toplevel(self.win)
        d.transient(self.win)
        d.grab_set()
        d.title("Generar preguntas con IA")

        # ── fila 0: grado ──────────────────────────────────────────────────────
        ttk.Label(d, text="grado:").grid(row=0, column=0, sticky="e", padx=6, pady=4)
        cb_grado = ttk.Combobox(
            d, state="readonly", values=self._banco_grados_plan_estudio()
        )
        cb_grado.grid(row=0, column=1, sticky="we", padx=6, pady=4)

        # ── fila 1: área ───────────────────────────────────────────────────────
        ttk.Label(d, text="area:").grid(row=1, column=0, sticky="e", padx=6, pady=4)
        cb_area = ttk.Combobox(d, state="readonly", values=["Seleccione"])
        cb_area.set("Seleccione")
        cb_area.grid(row=1, column=1, sticky="we", padx=6, pady=4)

        # ── fila 2: evaluación ─────────────────────────────────────────────────
        ttk.Label(d, text="evaluacion:").grid(
            row=2, column=0, sticky="e", padx=6, pady=4
        )
        en_evaluacion = ttk.Entry(d)
        en_evaluacion.grid(row=2, column=1, sticky="we", padx=6, pady=4)

        # ── fila 3: cantidad de temas ──────────────────────────────────────────
        ttk.Label(d, text="cantidad de temas:").grid(
            row=3, column=0, sticky="e", padx=6, pady=4
        )
        sp_cant_temas = tk.Spinbox(d, from_=1, to=10, width=6)
        sp_cant_temas.delete(0, "end")
        sp_cant_temas.insert(0, "5")
        sp_cant_temas.grid(row=3, column=1, sticky="w", padx=6, pady=4)

        # ── fila 4: entradas dinámicas de temas ────────────────────────────────
        frame_temas = ttk.LabelFrame(d, text="Temas de la evaluación")
        frame_temas.grid(row=4, column=0, columnspan=2, sticky="we", padx=6, pady=4)
        frame_temas.columnconfigure(1, weight=1)

        tema_entries = []  # lista de ttk.Entry, se reconstruye con cada cambio

        def _rebuild_temas(*_):
            for w in frame_temas.winfo_children():
                w.destroy()
            tema_entries.clear()
            try:
                n = max(1, min(10, int(str(sp_cant_temas.get()).strip())))
            except Exception:
                n = 5
            for j in range(n):
                ttk.Label(frame_temas, text=f"Tema {j + 1}:").grid(
                    row=j, column=0, sticky="e", padx=6, pady=2
                )
                en = ttk.Entry(frame_temas, width=46)
                en.grid(row=j, column=1, sticky="we", padx=6, pady=2)
                tema_entries.append(en)

        sp_cant_temas.bind("<FocusOut>", _rebuild_temas)
        sp_cant_temas.bind("<Return>", _rebuild_temas)
        sp_cant_temas.configure(command=_rebuild_temas)
        _rebuild_temas()  # render inicial con 5 temas

        # ── fila 5: cantidad de preguntas ──────────────────────────────────────
        ttk.Label(d, text="cantidad de preguntas:").grid(
            row=5, column=0, sticky="e", padx=6, pady=4
        )
        sp_cantidad = tk.Spinbox(d, from_=1, to=50, width=8)
        sp_cantidad.delete(0, "end")
        sp_cantidad.insert(0, "20")
        sp_cantidad.grid(row=5, column=1, sticky="w", padx=6, pady=4)

        # ── fila 6: nivel de dificultad ────────────────────────────────────────
        ttk.Label(d, text="nivel de dificultad:").grid(
            row=6, column=0, sticky="e", padx=6, pady=4
        )
        cb_dificultad = ttk.Combobox(
            d, state="readonly", values=["Baja", "Media", "Alta"]
        )
        cb_dificultad.set("Alta")
        cb_dificultad.grid(row=6, column=1, sticky="w", padx=6, pady=4)

        # ── fila 7: cantidad de texto ─────────────────────────────────────────
        ttk.Label(d, text="cantidad de texto:").grid(
            row=7, column=0, sticky="e", padx=6, pady=4
        )
        sp_cantidad_texto = tk.Spinbox(d, from_=1, to=20, width=8)
        sp_cantidad_texto.delete(0, "end")
        sp_cantidad_texto.insert(0, "4")
        sp_cantidad_texto.grid(row=7, column=1, sticky="w", padx=6, pady=4)

        # ── fila 8: mínimo de palabras texto ──────────────────────────────────
        ttk.Label(d, text="mínimo de palabras texto:").grid(
            row=8, column=0, sticky="e", padx=6, pady=4
        )
        sp_min_palabras_texto = tk.Spinbox(d, from_=20, to=500, width=8)
        sp_min_palabras_texto.delete(0, "end")
        sp_min_palabras_texto.insert(0, "150")
        sp_min_palabras_texto.grid(row=8, column=1, sticky="w", padx=6, pady=4)

        # ── fila 9: máximo de palabras texto ──────────────────────────────────
        ttk.Label(d, text="máximo de palabras texto:").grid(
            row=9, column=0, sticky="e", padx=6, pady=4
        )
        sp_max_palabras_texto = tk.Spinbox(d, from_=20, to=800, width=8)
        sp_max_palabras_texto.delete(0, "end")
        sp_max_palabras_texto.insert(0, "250")
        sp_max_palabras_texto.grid(row=9, column=1, sticky="w", padx=6, pady=4)

        d._area_nombre_a_id = {}

        def _cargar_areas_por_grado():
            grado = cb_grado.get().strip()
            if not grado:
                cb_area["values"] = ["Seleccione"]
                cb_area.set("Seleccione")
                d._area_nombre_a_id = {}
                return
            areas = self._banco_areas_por_grado_desde_plan(grado)
            if not areas:
                cb_area["values"] = ["Seleccione"]
                cb_area.set("Seleccione")
                d._area_nombre_a_id = {}
                messagebox.showwarning(
                    "Plan de Estudios",
                    "No existe plan de estudios configurado para este grado.",
                    parent=d,
                )
                return
            nombres = [n for _, n in areas]
            d._area_nombre_a_id = {n: i for i, n in areas}
            cb_area["values"] = ["Seleccione"] + nombres
            cb_area.set("Seleccione")

        cb_grado.bind("<<ComboboxSelected>>", lambda _e: _cargar_areas_por_grado())

        def _on_generar():
            grado = cb_grado.get().strip()
            area = cb_area.get().strip()
            evaluacion = en_evaluacion.get().strip()
            dificultad = cb_dificultad.get().strip()

            # Recolectar temas individuales (solo los que tengan contenido)
            temas_vals = [e.get().strip() for e in tema_entries if e.get().strip()]
            temas = "\n".join(temas_vals)

            try:
                cantidad = int(str(sp_cantidad.get()).strip())
            except Exception:
                cantidad = 0

            try:
                cantidad_texto = int(str(sp_cantidad_texto.get()).strip())
            except Exception:
                cantidad_texto = 0

            try:
                min_palabras_texto = int(str(sp_min_palabras_texto.get()).strip())
            except Exception:
                min_palabras_texto = 0

            try:
                max_palabras_texto = int(str(sp_max_palabras_texto.get()).strip())
            except Exception:
                max_palabras_texto = 0

            if not grado:
                messagebox.showerror("Validación", "Debe seleccionar grado.", parent=d)
                return
            if not area or area == "Seleccione":
                messagebox.showerror("Validación", "Debe seleccionar área.", parent=d)
                return
            if not evaluacion:
                messagebox.showerror("Validación", "Debe indicar evaluación.", parent=d)
                return
            if not temas_vals:
                messagebox.showerror(
                    "Validación", "Debe ingresar al menos un tema.", parent=d
                )
                return
            if cantidad <= 0:
                messagebox.showerror(
                    "Validación", "Cantidad de preguntas inválida.", parent=d
                )
                return
            if cantidad_texto <= 0:
                messagebox.showerror(
                    "Validación", "Cantidad de texto inválida.", parent=d
                )
                return
            if min_palabras_texto <= 0 or max_palabras_texto <= 0:
                messagebox.showerror(
                    "Validación",
                    "Mínimo y máximo de palabras deben ser mayores a 0.",
                    parent=d,
                )
                return
            if max_palabras_texto < min_palabras_texto:
                messagebox.showerror(
                    "Validación",
                    "El máximo de palabras no puede ser menor al mínimo.",
                    parent=d,
                )
                return

            id_area = d._area_nombre_a_id.get(area)
            if id_area is None:
                messagebox.showerror(
                    "Validación", "Área inválida para el grado seleccionado.", parent=d
                )
                return

            preguntas, fuente = self._ia_generar_preguntas(
                grado=grado,
                area=area,
                evaluacion=evaluacion,
                temas=temas,
                cantidad=cantidad,
                dificultad=dificultad,
                cantidad_texto=min(cantidad_texto, cantidad),
                min_palabras_texto=min_palabras_texto,
                max_palabras_texto=max_palabras_texto,
            )
            if not preguntas:
                messagebox.showerror(
                    "IA",
                    "No se pudieron generar preguntas con los datos proporcionados.",
                    parent=d,
                )
                return

            d.destroy()
            self._ia_previsualizar_preguntas(
                preguntas=preguntas,
                meta={
                    "grado": grado,
                    "area": area,
                    "id_area": int(id_area),
                    "evaluacion": evaluacion,
                    "dificultad": dificultad,
                    "cantidad_texto": min(cantidad_texto, cantidad),
                    "min_palabras_texto": min_palabras_texto,
                    "max_palabras_texto": max_palabras_texto,
                    "fuente": fuente,
                },
            )

        btns = ttk.Frame(d)
        btns.grid(row=10, column=0, columnspan=2, sticky="w", padx=6, pady=8)
        ttk.Button(btns, text="Generar", command=_on_generar).pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(btns, text="Cancelar", command=d.destroy).pack(side="left")

        d.columnconfigure(1, weight=1)
        d.wait_window()

    def _ia_generar_preguntas(
        self,
        grado,
        area,
        evaluacion,
        temas,
        cantidad,
        dificultad,
        cantidad_texto=3,
        min_palabras_texto=70,
        max_palabras_texto=120,
    ):
        """Intenta generar preguntas con API compatible OpenAI; si falla usa fallback local."""
        temas_list = [
            t.strip() for t in re.split(r"[,\n;]+", str(temas or "")) if t.strip()
        ]
        if not temas_list:
            temas_list = [f"fundamentos de {area}"]

        cantidad_texto = max(1, int(cantidad_texto or 1))
        min_palabras_texto = max(1, int(min_palabras_texto or 1))
        max_palabras_texto = max(min_palabras_texto, int(max_palabras_texto or 1))
        usar_competencias = False
        competencias = []
        cantidad = int(cantidad)

        temas_fijos = []
        for j in range(5):
            if j < len(temas_list):
                temas_fijos.append(temas_list[j])
            else:
                temas_fijos.append("(sin tema especificado)")

        prompt_sistema = "Eres un experto en evaluación educativa del sistema colombiano y en la elaboración de pruebas tipo ICFES."

        prompt_usuario = (
            "Debes generar una evaluación académica de alta calidad siguiendo estrictamente las siguientes instrucciones.\n\n"
            "DATOS DE LA EVALUACIÓN\n\n"
            f"Grado: {grado}\n"
            f"Área: {area}\n"
            f"Nivel de dificultad: {dificultad}\n\n"
            "Temas de la evaluación:\n\n"
            f"{temas_fijos[0]}\n"
            f"{temas_fijos[1]}\n"
            f"{temas_fijos[2]}\n"
            f"{temas_fijos[3]}\n"
            f"{temas_fijos[4]}\n\n"
            f"Cantidad de textos: {cantidad_texto}\n"
            f"Cantidad total de preguntas: {cantidad}\n\n"
            f"Cada texto debe tener entre {min_palabras_texto} y {max_palabras_texto} palabras.\n\n"
            "INSTRUCCIONES\n\n"
            f"1. Genera exactamente {cantidad_texto} textos informativos, narrativos o explicativos.\n"
            f"2. Cada texto debe tener entre {min_palabras_texto} y {max_palabras_texto} palabras.\n"
            "3. Los textos deben estar relacionados con los temas indicados.\n"
            f"4. Los textos deben ser adecuados para estudiantes de grado {grado}.\n"
            "5. Los textos deben ser claros, coherentes y bien desarrollados.\n"
            f"6. Si un texto tiene menos de {min_palabras_texto} palabras debes ampliarlo hasta cumplir el mínimo.\n\n"
            "PREGUNTAS\n\n"
            f"1. Genera exactamente {cantidad} preguntas.\n"
            "2. Distribuye las preguntas de forma equilibrada entre los textos.\n"
            "3. Todas las preguntas deben basarse directamente en el texto correspondiente.\n"
            "4. Las preguntas deben evaluar habilidades cognitivas como:\n\n"
            "- comprensión literal\n"
            "- interpretación\n"
            "- inferencia\n"
            "- análisis\n\n"
            "5. Cada pregunta debe tener cuatro opciones de respuesta:\n\n"
            "A\n"
            "B\n"
            "C\n"
            "D\n\n"
            "6. Solo una opción debe ser correcta.\n"
            "7. Las opciones incorrectas deben ser plausibles y creíbles.\n"
            "8. Evita respuestas obvias.\n\n"
            "9. Si la pregunta incluye expresiones matemáticas, usa LaTeX delimitado por $...$.\n"
            "10. No uses formato plano como x^2; escribe $x^{2}$, $\\frac{3}{4}$, $\\sqrt{16}$, etc.\n\n"
            "FORMATO DE SALIDA\n\n"
            "TEXTO 1\n"
            "[texto completo]\n\n"
            "Pregunta 1\n"
            "A.\n"
            "B.\n"
            "C.\n"
            "D.\n"
            "Respuesta correcta:\n\n"
            "Pregunta 2\n"
            "A.\n"
            "B.\n"
            "C.\n"
            "D.\n"
            "Respuesta correcta:\n\n"
            "Pregunta 3\n"
            "...\n\n"
            "TEXTO 2\n"
            "...\n\n"
            "Continúa hasta completar todos los textos y preguntas."
        )

        # Intento de generación por API externa (opcional)
        try:
            api_key = os.environ.get("OPENAI_API_KEY", "").strip()
            if api_key:
                import urllib.request

                base_url = os.environ.get(
                    "OPENAI_BASE_URL", "https://api.openai.com/v1"
                ).rstrip("/")
                model = os.environ.get("SEA_IA_MODEL", "gpt-4o-mini")

                prompt = {
                    "model": model,
                    "messages": [
                        {
                            "role": "system",
                            "content": prompt_sistema,
                        },
                        {
                            "role": "user",
                            "content": prompt_usuario,
                        },
                    ],
                    "temperature": 0.7,
                }
                data = json.dumps(prompt).encode("utf-8")
                req = urllib.request.Request(
                    url=f"{base_url}/chat/completions",
                    data=data,
                    method="POST",
                    headers={
                        "Content-Type": "application/json",
                        "Authorization": f"Bearer {api_key}",
                    },
                )
                with urllib.request.urlopen(req, timeout=45) as resp:
                    payload = json.loads(resp.read().decode("utf-8", errors="ignore"))

                content = (
                    payload.get("choices", [{}])[0]
                    .get("message", {})
                    .get("content", "")
                )
                raw = str(content or "").strip()
                raw = re.sub(
                    r"^```json\s*|\s*```$", "", raw, flags=re.IGNORECASE | re.DOTALL
                )
                raw = re.sub(
                    r"^```\s*|\s*```$", "", raw, flags=re.IGNORECASE | re.DOTALL
                )

                parsed = None
                try:
                    parsed = json.loads(raw)
                except Exception:
                    m = re.search(r"\[.*\]", raw, flags=re.DOTALL)
                    if m:
                        parsed = json.loads(m.group(0))

                if isinstance(parsed, dict) and "preguntas" in parsed:
                    parsed = parsed.get("preguntas")

                if isinstance(parsed, list) and parsed:
                    preguntas = []
                    for q in parsed[:cantidad]:
                        p = {
                            "contexto": str(q.get("contexto", "")).strip(),
                            "enunciado": str(q.get("enunciado", "")).strip(),
                            "opcion_a": str(q.get("opcion_a", "")).strip(),
                            "opcion_b": str(q.get("opcion_b", "")).strip(),
                            "opcion_c": str(q.get("opcion_c", "")).strip(),
                            "opcion_d": str(q.get("opcion_d", "")).strip(),
                            "correcta": str(q.get("correcta", "")).strip().upper()[:1],
                            "competencia": str(q.get("competencia", "")).strip(),
                        }
                        if p["correcta"] not in ("A", "B", "C", "D"):
                            p["correcta"] = "A"
                        if usar_competencias and p["competencia"] not in competencias:
                            p["competencia"] = competencias[len(preguntas) % 3]
                        if (
                            p["enunciado"]
                            and p["opcion_a"]
                            and p["opcion_b"]
                            and p["opcion_c"]
                            and p["opcion_d"]
                        ):
                            preguntas.append(p)
                    if preguntas:
                        return preguntas, "IA API"

                # Si el modelo responde en formato textual, lo parseamos.
                preguntas_txt = self._ia_parsear_formato_icfes(
                    raw,
                    cantidad=cantidad,
                    usar_competencias=usar_competencias,
                    competencias=competencias,
                )
                if preguntas_txt:
                    return preguntas_txt, "IA API"
        except Exception:
            pass

        def _ajustar_cantidad_palabras(texto_base, minimo, maximo):
            palabras = str(texto_base or "").split()
            if not palabras:
                palabras = ["Texto", "de", "apoyo", "académico", "sobre", str(area)]

            while len(palabras) < minimo:
                palabras.extend(
                    [
                        "El",
                        "análisis",
                        "de",
                        "la",
                        "situación",
                        "permite",
                        "comprender",
                        "mejor",
                        "el",
                        "tema.",
                    ]
                )

            if len(palabras) > maximo:
                palabras = palabras[:maximo]

            return " ".join(palabras)

        contextos_base = []
        total_textos = min(cantidad_texto, cantidad)
        for j in range(total_textos):
            tema_ctx = temas_list[j % len(temas_list)]
            if j == (total_textos - 1):
                contexto_base = (
                    f"En la comunidad de estudiantes de grado {grado}, se presentó una situación real vinculada con {tema_ctx}. "
                    "Distintas personas propusieron alternativas para resolver el problema y cada opción tenía ventajas y riesgos. "
                    "Para decidir de manera responsable, fue necesario revisar datos, comparar evidencias y anticipar consecuencias. "
                    "La decisión final debía beneficiar al grupo y sostenerse con argumentos claros."
                )
            else:
                idx_ctx = j % 3
                if idx_ctx == 0:
                    contexto_base = (
                        f"{tema_ctx.capitalize()} es un contenido clave del área de {area}. "
                        "Comprenderlo ayuda a interpretar situaciones cotidianas y académicas con mayor precisión. "
                        "Cuando se analiza con ejemplos cercanos, es más fácil identificar relaciones, causas y efectos."
                    )
                elif idx_ctx == 1:
                    contexto_base = (
                        f"Durante una actividad escolar sobre {tema_ctx}, varios estudiantes reunieron información de diferentes fuentes. "
                        "Al comparar los datos, encontraron coincidencias y diferencias que exigían una lectura cuidadosa. "
                        "Ese proceso permitió construir conclusiones mejor fundamentadas."
                    )
                else:
                    contexto_base = (
                        f"En el entorno local, {tema_ctx} aparece en situaciones que afectan decisiones individuales y colectivas. "
                        "Observar lo que ocurre, ordenar la información y evaluarla con criterios claros facilita proponer soluciones viables. "
                        "Por eso, este tema se estudia para fortalecer el razonamiento."
                    )
            contextos_base.append(
                _ajustar_cantidad_palabras(
                    contexto_base, min_palabras_texto, max_palabras_texto
                )
            )

        # Fallback local
        preguntas = []
        for i in range(cantidad):
            tema = temas_list[i % len(temas_list)]
            competencia = (
                competencias[i % len(competencias)] if usar_competencias else ""
            )

            contexto = contextos_base[i % len(contextos_base)]

            if competencia:
                contexto = f"[Competencia: {competencia}] {contexto}"

            nivel = str(dificultad or "").strip().lower()
            area_txt = str(area or "").strip().lower()
            es_matematicas = "matem" in area_txt

            if es_matematicas:
                a = (i % 5) + 2
                b = (i % 7) + 1
                cte = (i % 4) + 1
                x_val = (i % 4) + 1
                correcto_val = (a * (x_val**2)) + (b * x_val) + cte

                enunciado = (
                    "Calcule el valor de la expresión "
                    f"${a}x^{{2}} + {b}x + {cte}$ cuando $x={x_val}$."
                )
                opciones = [
                    f"${correcto_val}$",
                    f"${correcto_val + 1}$",
                    f"${correcto_val - 1}$",
                    f"${correcto_val + 2}$",
                ]
                correct_idx = 0
            elif nivel == "baja":
                enunciado = f"De acuerdo con el contexto, ¿qué idea describe mejor el concepto central de {tema}?"
                opciones = [
                    f"{tema} solo sirve para repetir definiciones sin comprender situaciones.",
                    f"{tema} permite reconocer información relevante para interpretar la situación planteada.",
                    f"{tema} no se relaciona con los propósitos del área de {area}.",
                    f"{tema} impide distinguir datos principales y secundarios en un texto.",
                ]
                correct_idx = 1
            elif nivel == "alta":
                enunciado = (
                    "Con base en el contexto, ¿cuál interpretación relaciona correctamente "
                    "la información y sustenta una conclusión válida?"
                )
                opciones = [
                    "Tomar un dato aislado y generalizarlo sin contrastarlo con el resto del contexto.",
                    "Relacionar evidencias del contexto para inferir una conclusión consistente.",
                    "Descartar la información central y priorizar detalles irrelevantes.",
                    "Aceptar la primera afirmación sin analizar supuestos ni consecuencias.",
                ]
                correct_idx = 1
            else:
                enunciado = f"Según la situación presentada, ¿qué acción aplica de mejor manera lo aprendido sobre {tema}?"
                opciones = [
                    "Repetir la misma respuesta en cualquier caso, sin revisar el contexto.",
                    "Usar la información del contexto para proponer una solución coherente con el problema.",
                    "Ignorar los datos del contexto y decidir únicamente por intuición.",
                    "Elegir la opción más extensa aunque no responda al caso planteado.",
                ]
                correct_idx = 1

            orden = [0, 1, 2, 3]
            rnd = random.Random(f"{grado}|{area}|{tema}|{i}|{dificultad}")
            rnd.shuffle(orden)

            opts = [opciones[j] for j in orden]
            letra = "ABCD"
            correcta = letra[orden.index(correct_idx)]

            preguntas.append(
                {
                    "contexto": contexto,
                    "enunciado": enunciado,
                    "opcion_a": opts[0],
                    "opcion_b": opts[1],
                    "opcion_c": opts[2],
                    "opcion_d": opts[3],
                    "correcta": correcta,
                    "competencia": competencia,
                }
            )

        return preguntas, "IA Local"

    def _ia_parsear_formato_icfes(
        self, texto, cantidad, usar_competencias, competencias
    ):
        """Parsea salida textual (formato etiquetas o formato TEXTO/Pregunta)."""
        txt = str(texto or "").replace("\r\n", "\n").replace("\r", "\n").strip()
        if not txt:
            return []

        bloques = re.split(r"\n\s*\n(?=\s*contexto\s*:)", txt, flags=re.IGNORECASE)
        resultados = []
        keys = [
            "contexto",
            "enunciado",
            "opcion_a",
            "opcion_b",
            "opcion_c",
            "opcion_d",
            "correcta",
        ]
        key_set = set(keys)
        patt = re.compile(
            r"^\s*(contexto|enunciado|opcion_a|opcion_b|opcion_c|opcion_d|correcta)\s*:\s*(.*)$",
            flags=re.IGNORECASE,
        )

        for bloque in bloques:
            lineas = [ln.rstrip() for ln in bloque.split("\n") if ln.strip()]
            if not lineas:
                continue

            data = {k: "" for k in keys}
            actual = None
            for ln in lineas:
                m = patt.match(ln)
                if m:
                    actual = m.group(1).lower()
                    data[actual] = m.group(2).strip()
                elif actual in key_set:
                    if data[actual]:
                        data[actual] += "\n" + ln.strip()
                    else:
                        data[actual] = ln.strip()

            if not all(data.get(k, "").strip() for k in keys[:-1]):
                continue

            correcta = str(data.get("correcta", "")).strip().upper()[:1]
            if correcta not in ("A", "B", "C", "D"):
                continue

            competencia = ""
            if usar_competencias:
                competencia = competencias[len(resultados) % len(competencias)]

            resultados.append(
                {
                    "contexto": data["contexto"].strip(),
                    "enunciado": data["enunciado"].strip(),
                    "opcion_a": data["opcion_a"].strip(),
                    "opcion_b": data["opcion_b"].strip(),
                    "opcion_c": data["opcion_c"].strip(),
                    "opcion_d": data["opcion_d"].strip(),
                    "correcta": correcta,
                    "competencia": competencia,
                }
            )

            if len(resultados) >= int(cantidad):
                break

        if resultados:
            return resultados

        # Formato alterno:
        # TEXTO 1
        # [texto]
        #
        # Pregunta 1
        # [enunciado]
        # A.
        # B.
        # C.
        # D.
        # Respuesta correcta: X
        bloque_textos = re.split(r"(?im)^\s*TEXTO\s+\d+\s*$", txt)
        if len(bloque_textos) <= 1:
            return []

        def _extraer_opcion(qbloque, letra):
            patron = re.compile(
                rf"(?ims)^\s*{letra}\.\s*(.*?)\s*(?=^\s*[ABCD]\.\s|^\s*Respuesta\s+correcta\s*:|\Z)"
            )
            mopt = patron.search(qbloque)
            return mopt.group(1).strip() if mopt else ""

        for bloque_texto in bloque_textos[1:]:
            bt = str(bloque_texto or "").strip()
            if not bt:
                continue

            preguntas_match = list(re.finditer(r"(?im)^\s*Pregunta\s+\d+\s*$", bt))
            if not preguntas_match:
                continue

            contexto = bt[: preguntas_match[0].start()].strip()
            if not contexto:
                continue

            for idx, m_q in enumerate(preguntas_match):
                ini = m_q.end()
                fin = (
                    preguntas_match[idx + 1].start()
                    if idx + 1 < len(preguntas_match)
                    else len(bt)
                )
                qbloque = bt[ini:fin].strip()
                if not qbloque:
                    continue

                m_a = re.search(r"(?im)^\s*A\.\s*", qbloque)
                if not m_a:
                    continue

                enunciado = qbloque[: m_a.start()].strip()
                opcion_a = _extraer_opcion(qbloque, "A")
                opcion_b = _extraer_opcion(qbloque, "B")
                opcion_c = _extraer_opcion(qbloque, "C")
                opcion_d = _extraer_opcion(qbloque, "D")

                m_corr = re.search(
                    r"(?im)^\s*Respuesta\s+correcta\s*:\s*([ABCD])\s*$", qbloque
                )
                if not m_corr:
                    m_corr = re.search(
                        r"(?im)^\s*Respuesta\s+correcta\s*:\s*([ABCD])", qbloque
                    )
                if not m_corr:
                    continue

                correcta = m_corr.group(1).upper().strip()
                if correcta not in ("A", "B", "C", "D"):
                    continue

                if not all([enunciado, opcion_a, opcion_b, opcion_c, opcion_d]):
                    continue

                competencia = ""
                if usar_competencias:
                    competencia = competencias[len(resultados) % len(competencias)]

                resultados.append(
                    {
                        "contexto": contexto,
                        "enunciado": enunciado,
                        "opcion_a": opcion_a,
                        "opcion_b": opcion_b,
                        "opcion_c": opcion_c,
                        "opcion_d": opcion_d,
                        "correcta": correcta,
                        "competencia": competencia,
                    }
                )

                if len(resultados) >= int(cantidad):
                    break

            if len(resultados) >= int(cantidad):
                break

        return resultados

    def _ia_previsualizar_preguntas(self, preguntas, meta):
        """Muestra preguntas generadas antes de guardarlas."""
        d = tk.Toplevel(self.win)
        d.transient(self.win)
        d.grab_set()
        d.title("Preguntas generadas con IA")
        d.geometry("1100x520")

        info = (
            f"Fuente: {meta.get('fuente', '')} | "
            f"Grado: {meta.get('grado', '')} | "
            f"Área: {meta.get('area', '')} | "
            f"Evaluación: {meta.get('evaluacion', '')} | "
            f"Textos: {meta.get('cantidad_texto', '')} | "
            f"Palabras/texto: {meta.get('min_palabras_texto', '')}-{meta.get('max_palabras_texto', '')}"
        )
        ttk.Label(d, text=info).pack(anchor="w", padx=10, pady=(8, 4))

        cols = ("n", "competencia", "enunciado", "a", "b", "c", "d", "correcta")
        tree = ttk.Treeview(d, columns=cols, show="headings")
        tree.heading("n", text="#")
        tree.heading("competencia", text="Competencia")
        tree.heading("enunciado", text="Enunciado")
        tree.heading("a", text="A")
        tree.heading("b", text="B")
        tree.heading("c", text="C")
        tree.heading("d", text="D")
        tree.heading("correcta", text="Correcta")
        tree.column("n", width=40, anchor="center", stretch=False)
        tree.column("competencia", width=130, anchor="center", stretch=False)
        tree.column("enunciado", width=300, anchor="w")
        tree.column("a", width=180, anchor="w")
        tree.column("b", width=180, anchor="w")
        tree.column("c", width=180, anchor="w")
        tree.column("d", width=180, anchor="w")
        tree.column("correcta", width=80, anchor="center", stretch=False)

        ysb = ttk.Scrollbar(d, orient="vertical", command=tree.yview)
        xsb = ttk.Scrollbar(d, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)

        tree.pack(fill="both", expand=True, padx=10, pady=(0, 2))
        ysb.pack(side="right", fill="y")
        xsb.pack(side="bottom", fill="x")

        def _truncate(txt, max_len=80):
            t = str(txt or "")
            return t if len(t) <= max_len else (t[: max_len - 3] + "...")

        def _refresh_tree():
            for iid in tree.get_children():
                tree.delete(iid)
            for i, q in enumerate(preguntas, start=1):
                tree.insert(
                    "",
                    "end",
                    iid=str(i - 1),
                    values=(
                        i,
                        _truncate(q.get("competencia", ""), 25),
                        _truncate(q.get("enunciado", ""), 120),
                        _truncate(q.get("opcion_a", ""), 65),
                        _truncate(q.get("opcion_b", ""), 65),
                        _truncate(q.get("opcion_c", ""), 65),
                        _truncate(q.get("opcion_d", ""), 65),
                        q.get("correcta", ""),
                    ),
                )

        def _editar_seleccionada():
            sel = tree.selection()
            if not sel:
                messagebox.showinfo("IA", "Seleccione una pregunta.", parent=d)
                return
            idx = int(sel[0])
            q = preguntas[idx]

            ed = tk.Toplevel(d)
            ed.transient(d)
            ed.grab_set()
            ed.title(f"Editar pregunta {idx + 1}")

            campos = [
                "contexto",
                "enunciado",
                "opcion_a",
                "opcion_b",
                "opcion_c",
                "opcion_d",
                "correcta",
                "competencia",
            ]
            widgets = {}
            for r, c in enumerate(campos):
                ttk.Label(ed, text=c + ":").grid(
                    row=r, column=0, sticky="ne", padx=6, pady=4
                )
                if c in ("contexto", "enunciado"):
                    w = tk.Text(ed, width=70, height=3)
                    w.insert("1.0", q.get(c, ""))
                else:
                    w = ttk.Entry(ed, width=72)
                    w.insert(0, q.get(c, ""))
                w.grid(row=r, column=1, sticky="we", padx=6, pady=4)
                widgets[c] = w

            def _guardar_edicion():
                nuevo = {}
                for c in campos:
                    w = widgets[c]
                    if isinstance(w, tk.Text):
                        nuevo[c] = w.get("1.0", "end").strip()
                    else:
                        nuevo[c] = w.get().strip()

                for req in (
                    "enunciado",
                    "opcion_a",
                    "opcion_b",
                    "opcion_c",
                    "opcion_d",
                ):
                    if not nuevo.get(req):
                        messagebox.showerror(
                            "Validación", f"Campo requerido: {req}", parent=ed
                        )
                        return

                corr = str(nuevo.get("correcta", "")).upper().strip()
                if corr not in ("A", "B", "C", "D"):
                    messagebox.showerror(
                        "Validación", "La correcta debe ser A, B, C o D.", parent=ed
                    )
                    return

                preguntas[idx].update(nuevo)
                ed.destroy()
                _refresh_tree()

            bf = ttk.Frame(ed)
            bf.grid(row=len(campos), column=0, columnspan=2, sticky="w", padx=6, pady=8)
            ttk.Button(bf, text="Guardar", command=_guardar_edicion).pack(
                side="left", padx=(0, 6)
            )
            ttk.Button(bf, text="Cancelar", command=ed.destroy).pack(side="left")
            ed.columnconfigure(1, weight=1)
            ed.wait_window()

        def _guardar_todas():
            if not preguntas:
                messagebox.showinfo("IA", "No hay preguntas para guardar.", parent=d)
                return

            try:
                guardadas = 0
                for i, q in enumerate(preguntas, start=1):
                    correcta = str(q.get("correcta", "")).upper().strip()
                    if correcta not in ("A", "B", "C", "D"):
                        continue

                    contexto_guardar = str(q.get("contexto", "")).strip()
                    competencia = str(q.get("competencia", "")).strip()
                    if competencia and "[Competencia:" not in contexto_guardar:
                        contexto_guardar = (
                            f"[Competencia: {competencia}] {contexto_guardar}"
                        )

                    core_preguntas.crear_pregunta_banco(
                        evaluacion=str(meta.get("evaluacion", "")).strip(),
                        area=str(meta.get("area", "")).strip(),
                        periodo="",
                        grado=str(meta.get("grado", "")).strip(),
                        curso="",
                        id_contexto=f"IA_{datetime.now().strftime('%Y%m%d%H%M%S')}_{i}",
                        contexto=contexto_guardar,
                        enunciado=str(q.get("enunciado", "")).strip(),
                        opcion_a=str(q.get("opcion_a", "")).strip(),
                        opcion_b=str(q.get("opcion_b", "")).strip(),
                        opcion_c=str(q.get("opcion_c", "")).strip(),
                        opcion_d=str(q.get("opcion_d", "")).strip(),
                        correcta=correcta,
                        imagen="",
                        nombre="IA",
                        id_area=meta.get("id_area"),
                    )
                    guardadas += 1
            except Exception as exc:
                messagebox.showerror(
                    "IA", f"No se pudieron guardar las preguntas: {exc}", parent=d
                )
                return

            self._banco_recargar_treeview()
            messagebox.showinfo(
                "IA", f"Preguntas guardadas: {guardadas}", parent=self.win
            )
            d.destroy()

        _refresh_tree()

        bf = ttk.Frame(d)
        bf.pack(fill="x", padx=10, pady=8)
        ttk.Button(bf, text="Guardar preguntas", command=_guardar_todas).pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(bf, text="Editar pregunta", command=_editar_seleccionada).pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(bf, text="Descartar", command=d.destroy).pack(side="left")

        d.wait_window()

    def _banco_valor_filtro(self, attr_name):
        """Devuelve valor de filtro normalizado o None si equivale a 'todos'."""
        if not hasattr(self, attr_name):
            return None
        try:
            val = str(getattr(self, attr_name).get()).strip()
        except Exception:
            return None
        if not val:
            return None
        if val.lower() in ("(todos)", "todos", "seleccione", "todas"):
            return None
        return val

    def _banco_recargar_treeview(self):
        """Recarga datos del banco y refresca el Treeview de preguntas."""
        try:
            self.banco._cargar_preguntas()
        except Exception:
            pass

        try:
            self._preg_cargar_datos_filtrados()
        except Exception:
            try:
                self._load_preguntas()
            except Exception:
                pass

        self._emit_preguntas_actualizadas()

    def banco_limpiar_filtros(self):
        """Elimina preguntas que coinciden con los filtros seleccionados."""
        if not messagebox.askyesno(
            "Confirmar",
            "¿Desea eliminar todas las preguntas que coinciden con los filtros seleccionados?",
            parent=self.win,
        ):
            return

        grado = self._banco_valor_filtro("combo_grado_preg")
        curso = self._banco_valor_filtro("combo_curso_preg")
        area = self._banco_valor_filtro("combo_area_preg")
        evaluacion = self._banco_valor_filtro("combo_evaluacion_preg")

        eliminadas = 0
        try:
            eliminadas = core_preguntas.eliminar_preguntas_banco(
                grado=grado,
                curso=curso,
                area=area,
                evaluacion=evaluacion,
            )
        except Exception as e:
            messagebox.showerror(
                "Limpiar Banco",
                f"No se pudieron eliminar preguntas: {e}",
                parent=self.win,
            )
            return

        self._banco_recargar_treeview()
        messagebox.showinfo(
            "Limpiar Banco",
            f"Preguntas eliminadas: {eliminadas}",
            parent=self.win,
        )

    def banco_vaciar_completo(self):
        """Elimina todas las preguntas del banco."""
        if not messagebox.askyesno(
            "Confirmar",
            "⚠ Esta acción eliminará TODAS las preguntas del banco de preguntas.\n¿Desea continuar?",
            parent=self.win,
        ):
            return

        eliminadas = 0
        try:
            eliminadas = core_preguntas.vaciar_banco_preguntas()
        except Exception as e:
            messagebox.showerror(
                "Vaciar Banco",
                f"No se pudo vaciar el banco: {e}",
                parent=self.win,
            )
            return

        self._banco_recargar_treeview()
        messagebox.showinfo(
            "Vaciar Banco",
            f"Banco vaciado. Registros eliminados: {eliminadas}",
            parent=self.win,
        )

    def _load_preguntas(self):
        try:
            self.banco._cargar_preguntas()
            df = self.banco.obtener_todas_preguntas()
        except Exception:
            if _HAS_PANDAS:
                df = pd.DataFrame(
                    columns=[
                        "id",
                        "evaluacion",
                        "area",
                        "periodo",
                        "grado",
                        "id_contexto",
                        "contexto",
                        "enunciado",
                        "opcion_a",
                        "opcion_b",
                        "opcion_c",
                        "opcion_d",
                        "correcta",
                        "imagen",
                    ]
                )
            else:
                df = {
                    "_headers": (
                        "id",
                        "evaluacion",
                        "area",
                        "periodo",
                        "grado",
                        "id_contexto",
                        "contexto",
                        "enunciado",
                        "opcion_a",
                        "opcion_b",
                        "opcion_c",
                        "opcion_d",
                        "correcta",
                        "imagen",
                    ),
                    "_rows": [],
                }
        self.preguntas_df = df
        for i in self.tree_preg.get_children():
            self.tree_preg.delete(i)
        if _HAS_PANDAS:
            for _, r in df.iterrows():
                self.tree_preg.insert(
                    "",
                    "end",
                    values=tuple(r.get(c, "") for c in df.columns),
                )
        else:
            for r in df.get("_rows", []):
                self.tree_preg.insert(
                    "",
                    "end",
                    values=(
                        r.get("id", ""),
                        r.get("evaluacion", ""),
                        r.get("area", ""),
                        r.get("periodo", ""),
                        r.get("grado", ""),
                        r.get("id_contexto", ""),
                        r.get("contexto", ""),
                        r.get("enunciado", ""),
                        r.get("opcion_a", ""),
                        r.get("opcion_b", ""),
                        r.get("opcion_c", ""),
                        r.get("opcion_d", ""),
                        r.get("correcta", ""),
                        r.get("imagen", ""),
                    ),
                )

        def aplicar_filtros(self):
            """Aplica filtros al banco de preguntas y recarga `tree_preg`.

            Normaliza selección y columnas, convierte grado a cadena, limpia y
            recarga el Treeview con tuplas en el orden esperado y muestra un
            mensaje si no hay resultados.
            """

            def _norm(val):
                if val is None:
                    return None
                v = str(val).strip()
                if not v or v == "(Todos)":
                    return None
                return v.lower()

            grado_sel = (
                self.combo_grado_preg.get()
                if hasattr(self, "combo_grado_preg")
                else None
            )
            area_sel = (
                self.combo_area_preg.get() if hasattr(self, "combo_area_preg") else None
            )
            eval_sel = (
                self.combo_evaluacion_preg.get()
                if hasattr(self, "combo_evaluacion_preg")
                else None
            )

            grado_val = _norm(grado_sel)
            area_val = _norm(area_sel)
            eval_val = _norm(eval_sel)

            if grado_val is not None:
                grado_val = str(grado_val)

            try:
                df_filtrado = self.banco.obtener_preguntas_filtradas(
                    grado=grado_val, area=area_val, evaluacion=eval_val
                )
            except Exception:
                df_filtrado = None

            if hasattr(df_filtrado, "columns"):
                for col in ("grado", "area", "evaluacion"):
                    if col in df_filtrado.columns:
                        try:
                            df_filtrado[col] = (
                                df_filtrado[col].astype(str).str.strip().str.lower()
                            )
                        except Exception:
                            pass

            for iid in self.tree_preg.get_children():
                self.tree_preg.delete(iid)

            is_empty = False
            if df_filtrado is None:
                is_empty = True
            elif hasattr(df_filtrado, "empty") and df_filtrado.empty:
                is_empty = True
            elif isinstance(df_filtrado, (list, tuple)) and len(df_filtrado) == 0:
                is_empty = True

            if is_empty:
                messagebox.showinfo(
                    "Filtros",
                    "No hay preguntas que coincidan con los criterios seleccionados.",
                    parent=getattr(self, "win", None),
                )
                try:
                    self._preg_actualizar_estadisticas()
                except Exception:
                    pass
                return

            if hasattr(df_filtrado, "iterrows"):
                for _, row in df_filtrado.iterrows():
                    imagen_str = (
                        "✓"
                        if row.get("imagen") and str(row.get("imagen")).strip()
                        else ""
                    )

                    valores = (
                        row.get("id", ""),
                        row.get("evaluacion", ""),
                        row.get("area", ""),
                        row.get("periodo", ""),
                        row.get("grado", ""),
                        row.get("id_contexto", ""),
                        row.get("contexto", ""),
                        row.get("enunciado", ""),
                        row.get("opcion_a", ""),
                        row.get("opcion_b", ""),
                        row.get("opcion_c", ""),
                        row.get("opcion_d", ""),
                        row.get("correcta", ""),
                        imagen_str,
                    )
                    self.tree_preg.insert("", "end", values=valores)
            else:
                for r in df_filtrado:
                    imagen_str = (
                        "✓" if r.get("imagen") and str(r.get("imagen")).strip() else ""
                    )
                    valores = (
                        r.get("id", ""),
                        r.get("evaluacion", ""),
                        r.get("area", ""),
                        r.get("periodo", ""),
                        r.get("grado", ""),
                        r.get("id_contexto", ""),
                        r.get("contexto", ""),
                        r.get("enunciado", ""),
                        r.get("opcion_a", ""),
                        r.get("opcion_b", ""),
                        r.get("opcion_c", ""),
                        r.get("opcion_d", ""),
                        r.get("correcta", ""),
                        imagen_str,
                    )
                    self.tree_preg.insert("", "end", values=valores)

            try:
                self._preg_actualizar_estadisticas()
            except Exception:
                pass

    def _save_preguntas(self):
        if _HAS_PANDAS:
            self.banco.df = self.preguntas_df.copy()
            self.banco.guardar_preguntas()
            self.banco._cargar_preguntas()
        else:
            return

    def _pregunta_id_exists(self, id_val):
        if id_val is None or str(id_val).strip() == "":
            return False
        if _HAS_PANDAS:
            try:
                return str(id_val) in self.preguntas_df["id"].astype(str).values
            except Exception:
                return False
        else:
            return any(
                str(r.get("id", "")) == str(id_val)
                for r in self.preguntas_df.get("_rows", [])
            )

    def preguntas_importar(self):
        path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel para importar preguntas",
            filetypes=[("Excel", "*.xlsx;*.xls")],
            parent=self.win,
        )
        if not path:
            return
        incoming = self._read_excel(path)
        if incoming is None:
            messagebox.showerror(
                "Importar", "Archivo vacío o inválido.", parent=self.win
            )
            return
        if _HAS_PANDAS:
            self.preguntas_df = pd.concat(
                [self.preguntas_df, incoming], ignore_index=True
            )
        else:
            rows = incoming.get("_rows", [])
            self.preguntas_df.setdefault("_rows", []).extend(rows)
        self._save_preguntas()
        self._banco_recargar_treeview()

    def preguntas_exportar(self):
        path = filedialog.asksaveasfilename(
            title="Exportar preguntas a Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            parent=self.win,
        )
        if not path:
            return
        # usar el DataFrame actual para exportar
        if _HAS_PANDAS:
            self.preguntas_df.to_excel(path, index=False)
        else:
            self._write_excel(path, self.preguntas_df)
        messagebox.showinfo("Exportar", "Exportación completada.", parent=self.win)

    def pregunta_agregar(self):
        payload = self._dialog_evaluacion_completa()
        if not payload:
            return

        filas = [dict(datos) for datos in payload if isinstance(datos, dict)]
        if not filas:
            return

        try:
            for datos in filas:
                core_preguntas.crear_pregunta_banco(**datos)
            self._banco_recargar_treeview()
        except Exception as exc:
            messagebox.showerror(
                "Banco de Preguntas",
                f"No se pudieron agregar las preguntas: {exc}",
                parent=self.win,
            )
            return

        if len(filas) > 1:
            messagebox.showinfo(
                "Banco de Preguntas",
                f"Se agregaron {len(filas)} preguntas en una sola evaluacion.",
                parent=self.win,
            )

    def _obtener_siguientes_ids_pregunta(self, cantidad):
        try:
            cantidad = max(0, int(cantidad or 0))
        except Exception:
            cantidad = 0

        if cantidad <= 0:
            return []

        ids = []

        def _append_id(raw):
            txt = str(raw or "").strip()
            if not txt:
                return
            try:
                ids.append(int(float(txt)))
            except Exception:
                return

        try:
            if _HAS_PANDAS and hasattr(self.preguntas_df, "iterrows"):
                for _, row in self.preguntas_df.iterrows():
                    _append_id(row.get("id"))
            else:
                for row in self.preguntas_df.get("_rows", []):
                    _append_id(row.get("id"))
        except Exception:
            pass

        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute("SELECT COALESCE(MAX(id), 0) FROM banco_preguntas")
            row = cur.fetchone()
            conn.close()
            if row:
                _append_id(row[0])
        except Exception:
            pass

        inicio = (max(ids) if ids else 0) + 1
        return [str(inicio + i) for i in range(cantidad)]

    def _validar_contenido_latex(self, datos, parent=None, prefijo="Pregunta"):
        campos_math = [
            "contexto",
            "enunciado",
            "opcion_a",
            "opcion_b",
            "opcion_c",
            "opcion_d",
        ]

        if str(datos.get("tipo_contenido", "normal")).strip().lower() == "latex":
            if not any("$" in str(datos.get(c, "") or "") for c in campos_math):
                messagebox.showerror(
                    "Validacion",
                    f"{prefijo}: selecciono tipo LaTeX pero no se detectaron delimitadores $...$.",
                    parent=parent or self.win,
                )
                return False

        for campo in campos_math:
            texto = str(datos.get(campo, "") or "")
            try:
                expresiones = _extraer_expresiones_latex(texto)
            except ValueError as exc:
                messagebox.showerror(
                    "Validacion LaTeX",
                    f"{prefijo} - campo '{campo}': {exc}",
                    parent=parent or self.win,
                )
                return False

            for expr in expresiones:
                try:
                    _validar_expresion_latex(expr)
                except ValueError as exc:
                    messagebox.showerror(
                        "Validacion LaTeX",
                        f"{prefijo} - campo '{campo}': {exc}",
                        parent=parent or self.win,
                    )
                    return False
        return True

    def _dialog_evaluacion_completa(self):
        d = tk.Toplevel(self.win)
        d.transient(self.win)
        d.grab_set()
        d.title("Agregar evaluacion completa")
        d.geometry("980x820")

        d._area_nombre_a_id = {}
        bloques = []
        var_aplicar_todas = tk.BooleanVar(value=False)

        frame_meta = ttk.LabelFrame(d, text="Datos de la evaluacion", padding=(10, 8))
        frame_meta.pack(fill="x", padx=10, pady=(10, 6))

        ttk.Label(frame_meta, text="grado:").grid(
            row=0, column=0, sticky="e", padx=6, pady=4
        )
        cb_grado = ttk.Combobox(
            frame_meta,
            state="readonly",
            values=self._banco_grados_plan_estudio(),
            width=16,
        )
        cb_grado.grid(row=0, column=1, sticky="we", padx=6, pady=4)

        ttk.Label(frame_meta, text="area:").grid(
            row=0, column=2, sticky="e", padx=6, pady=4
        )
        cb_area = ttk.Combobox(frame_meta, state="readonly", values=["Seleccione"])
        cb_area.set("Seleccione")
        cb_area.grid(row=0, column=3, sticky="we", padx=6, pady=4)

        ttk.Label(frame_meta, text="evaluacion:").grid(
            row=1, column=0, sticky="e", padx=6, pady=4
        )
        en_evaluacion = ttk.Entry(frame_meta)
        en_evaluacion.grid(row=1, column=1, sticky="we", padx=6, pady=4)

        ttk.Label(frame_meta, text="periodo:").grid(
            row=1, column=2, sticky="e", padx=6, pady=4
        )
        en_periodo = ttk.Entry(frame_meta)
        en_periodo.grid(row=1, column=3, sticky="we", padx=6, pady=4)

        frame_meta.columnconfigure(1, weight=1)
        frame_meta.columnconfigure(3, weight=1)

        frame_ctx_comp = ttk.LabelFrame(d, text="Contexto compartido", padding=(10, 8))
        frame_ctx_comp.pack(fill="x", padx=10, pady=(0, 6))

        chk_aplicar_todas = ttk.Checkbutton(
            frame_ctx_comp,
            text="Aplicar contexto a todas las preguntas",
            variable=var_aplicar_todas,
        )
        chk_aplicar_todas.grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 6))

        ttk.Label(frame_ctx_comp, text="id_contexto:").grid(
            row=1, column=0, sticky="e", padx=4, pady=2
        )
        en_ctx_id = ttk.Entry(frame_ctx_comp)
        en_ctx_id.grid(row=1, column=1, sticky="we", padx=4, pady=2)

        ttk.Label(frame_ctx_comp, text="contexto:").grid(
            row=2, column=0, sticky="ne", padx=4, pady=2
        )
        txt_ctx = tk.Text(frame_ctx_comp, height=3, width=80)
        txt_ctx.grid(row=2, column=1, sticky="we", padx=4, pady=2)

        ttk.Label(frame_ctx_comp, text="imagen:").grid(
            row=3, column=0, sticky="e", padx=4, pady=2
        )
        frame_img_ctx = ttk.Frame(frame_ctx_comp)
        frame_img_ctx.grid(row=3, column=1, sticky="we", padx=4, pady=2)
        frame_img_ctx.columnconfigure(0, weight=1)

        en_ctx_imagen = ttk.Entry(frame_img_ctx)
        en_ctx_imagen.grid(row=0, column=0, sticky="we", padx=(0, 4))

        lbl_ctx_preview = ttk.Label(
            frame_img_ctx, text="Sin imagen", relief="sunken", width=22
        )
        lbl_ctx_preview.grid(row=0, column=2)

        frame_ctx_comp.columnconfigure(1, weight=1)

        controles = ttk.Frame(d)
        controles.pack(fill="x", padx=10, pady=(0, 6))

        ttk.Label(controles, text="cantidad de preguntas:").pack(
            side="left", padx=(0, 6)
        )
        sp_cantidad = tk.Spinbox(controles, from_=1, to=300, width=6)
        sp_cantidad.delete(0, "end")
        sp_cantidad.insert(0, "5")
        sp_cantidad.pack(side="left")

        contenedor_scroll = ttk.Frame(d)
        contenedor_scroll.pack(fill="both", expand=True, padx=10, pady=(0, 6))

        canvas = tk.Canvas(contenedor_scroll, highlightthickness=0)
        scroll_y = ttk.Scrollbar(
            contenedor_scroll, orient="vertical", command=canvas.yview
        )
        frame_bloques = ttk.Frame(canvas)
        frame_bloques.bind(
            "<Configure>",
            lambda _e: canvas.configure(scrollregion=canvas.bbox("all")),
        )

        win_frame = canvas.create_window((0, 0), window=frame_bloques, anchor="nw")
        canvas.bind(
            "<Configure>",
            lambda e: canvas.itemconfigure(win_frame, width=e.width),
        )

        canvas.configure(yscrollcommand=scroll_y.set)
        canvas.pack(side="left", fill="both", expand=True)
        scroll_y.pack(side="right", fill="y")

        def _scroll_rueda(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind("<MouseWheel>", _scroll_rueda)
        d.bind("<MouseWheel>", _scroll_rueda)

        def _resolver_ruta_imagen(ruta):
            txt = str(ruta or "").strip()
            if not txt:
                return ""
            if os.path.isabs(txt):
                return txt
            return os.path.join(self.base_dir, txt)

        def _actualizar_preview(preview, ruta_rel):
            ruta_rel = str(ruta_rel or "").strip()
            preview._tkimg = None
            if not ruta_rel:
                preview.configure(image="", text="Sin imagen")
                return
            try:
                ruta_abs = _resolver_ruta_imagen(ruta_rel)
                if _HAS_PIL and os.path.isfile(ruta_abs):
                    img_obj = Image.open(ruta_abs)
                    img_obj.thumbnail((120, 120))
                    tkimg = ImageTk.PhotoImage(img_obj)
                    preview._tkimg = tkimg
                    preview.configure(image=tkimg, text="")
                else:
                    preview.configure(image="", text=os.path.basename(ruta_rel))
            except Exception:
                preview.configure(image="", text=os.path.basename(ruta_rel))

        def _set_entry(entry, valor):
            try:
                if "disabled" in entry.state():
                    entry.state(["!disabled"])
            except Exception:
                pass
            entry.delete(0, "end")
            entry.insert(0, str(valor or ""))

        def _set_text(text_widget, valor):
            estado = str(text_widget.cget("state"))
            if estado == "disabled":
                text_widget.configure(state="normal")
            text_widget.delete("1.0", "end")
            text_widget.insert("1.0", str(valor or ""))
            if estado == "disabled":
                text_widget.configure(state="disabled")

        def _leer_contexto_compartido():
            return {
                "id_contexto": en_ctx_id.get().strip(),
                "contexto": txt_ctx.get("1.0", "end").strip(),
                "imagen": en_ctx_imagen.get().strip(),
            }

        def _actualizar_combo_areas():
            grado = cb_grado.get().strip()
            if not grado:
                d._area_nombre_a_id = {}
                cb_area["values"] = ["Seleccione"]
                cb_area.set("Seleccione")
                return

            areas = self._banco_areas_por_grado_desde_plan(grado)
            if not areas:
                d._area_nombre_a_id = {}
                cb_area["values"] = ["Seleccione"]
                cb_area.set("Seleccione")
                messagebox.showwarning(
                    "Plan de Estudios",
                    "No existe plan de estudios configurado para este grado.",
                    parent=d,
                )
                return

            nombres = [n for _, n in areas]
            d._area_nombre_a_id = {n: i for i, n in areas}
            anterior = cb_area.get().strip().lower()
            cb_area["values"] = ["Seleccione"] + nombres

            seleccionado = None
            for nombre in nombres:
                if str(nombre).strip().lower() == anterior:
                    seleccionado = nombre
                    break

            cb_area.set(seleccionado if seleccionado else "Seleccione")

        cb_grado.bind("<<ComboboxSelected>>", lambda _e: _actualizar_combo_areas())

        def _renumerar_bloques():
            for idx, bloque in enumerate(bloques, start=1):
                bloque["numero"] = idx
                bloque["titulo_var"].set(f"Pregunta {idx}")
                bloque["frame"].grid(row=idx - 1, column=0, sticky="we", pady=4)

        def _toggle_bloque(bloque):
            if bloque.get("colapsado"):
                bloque["cuerpo"].grid()
                bloque["btn_colapsar"].configure(text="Colapsar")
                bloque["colapsado"] = False
            else:
                bloque["cuerpo"].grid_remove()
                bloque["btn_colapsar"].configure(text="Expandir")
                bloque["colapsado"] = True

        def _eliminar_bloque(bloque):
            try:
                bloque["frame"].destroy()
            except Exception:
                pass
            if bloque in bloques:
                bloques.remove(bloque)
            _renumerar_bloques()

        def _on_pick_image(entry, preview):
            p = filedialog.askopenfilename(
                title="Seleccionar imagen",
                filetypes=[("Imagenes", "*.png;*.jpg;*.jpeg;*.gif")],
                parent=d,
            )
            if not p:
                return
            try:
                dst = os.path.join(self.imagenes_dir, os.path.basename(p))
                shutil.copy2(p, dst)
                rel = os.path.relpath(dst, self.base_dir)
                _set_entry(entry, rel)
                _actualizar_preview(preview, rel)
            except Exception as e:
                messagebox.showerror("Imagen", f"Error copiando imagen: {e}", parent=d)

        def _aplicar_contexto_a_bloque(bloque):
            usar_compartido = bool(var_aplicar_todas.get()) or bool(
                bloque["usar_contexto_compartido"].get()
            )
            w = bloque["widgets"]

            if bool(var_aplicar_todas.get()):
                bloque["usar_contexto_compartido"].set(True)
                bloque["chk_contexto_compartido"].state(["disabled"])
            else:
                bloque["chk_contexto_compartido"].state(["!disabled"])

            if usar_compartido:
                if not bloque.get("_tenia_contexto_compartido", False):
                    bloque["_cache_local"] = {
                        "id_contexto": w["id_contexto"].get().strip(),
                        "contexto": w["contexto"].get("1.0", "end").strip(),
                        "imagen": w["imagen"].get().strip(),
                    }

                compartido = _leer_contexto_compartido()
                _set_entry(w["id_contexto"], compartido["id_contexto"])
                _set_text(w["contexto"], compartido["contexto"])
                _set_entry(w["imagen"], compartido["imagen"])
                _actualizar_preview(bloque["preview_imagen"], compartido["imagen"])

                w["id_contexto"].state(["disabled"])
                w["imagen"].state(["disabled"])
                bloque["btn_imagen"].state(["disabled"])
                w["contexto"].configure(state="disabled")
                bloque["_tenia_contexto_compartido"] = True
            else:
                w["id_contexto"].state(["!disabled"])
                w["imagen"].state(["!disabled"])
                bloque["btn_imagen"].state(["!disabled"])
                w["contexto"].configure(state="normal")
                cache_local = bloque.get("_cache_local", {})
                _set_entry(w["id_contexto"], cache_local.get("id_contexto", ""))
                _set_text(w["contexto"], cache_local.get("contexto", ""))
                _set_entry(w["imagen"], cache_local.get("imagen", ""))
                _actualizar_preview(
                    bloque["preview_imagen"], cache_local.get("imagen", "")
                )
                bloque["_tenia_contexto_compartido"] = False

        def _sincronizar_contexto_compartido():
            _actualizar_preview(lbl_ctx_preview, en_ctx_imagen.get().strip())
            for bloque in bloques:
                if bool(var_aplicar_todas.get()) or bool(
                    bloque["usar_contexto_compartido"].get()
                ):
                    _aplicar_contexto_a_bloque(bloque)

        def _toggle_contexto_global():
            for bloque in bloques:
                if bool(var_aplicar_todas.get()):
                    bloque["usar_contexto_compartido"].set(True)
                _aplicar_contexto_a_bloque(bloque)

        chk_aplicar_todas.configure(command=_toggle_contexto_global)

        def _crear_bloque(datos_iniciales=None):
            datos_iniciales = datos_iniciales or {}

            frame = ttk.LabelFrame(frame_bloques)
            frame.grid(column=0, sticky="we", pady=4)
            frame.columnconfigure(0, weight=1)

            encabezado = ttk.Frame(frame)
            encabezado.grid(row=0, column=0, sticky="we", padx=6, pady=(4, 2))
            encabezado.columnconfigure(0, weight=1)

            titulo_var = tk.StringVar(value="Pregunta")
            ttk.Label(encabezado, textvariable=titulo_var).grid(
                row=0, column=0, sticky="w"
            )

            btn_colapsar = ttk.Button(encabezado, text="Colapsar", width=10)
            btn_colapsar.grid(row=0, column=1, sticky="e", padx=(0, 6))

            btn_eliminar = ttk.Button(encabezado, text="Eliminar", width=10)
            btn_eliminar.grid(row=0, column=2, sticky="e")

            cuerpo = ttk.Frame(frame)
            cuerpo.grid(row=1, column=0, sticky="we", padx=6, pady=(0, 6))
            cuerpo.columnconfigure(1, weight=1)

            var_ctx_bloque = tk.BooleanVar(
                value=bool(datos_iniciales.get("usar_contexto_compartido", False))
            )

            chk_ctx_bloque = ttk.Checkbutton(
                cuerpo,
                text="Usar contexto compartido en esta pregunta",
                variable=var_ctx_bloque,
            )
            chk_ctx_bloque.grid(
                row=0, column=0, columnspan=2, sticky="w", padx=4, pady=2
            )

            ttk.Label(cuerpo, text="id_contexto:").grid(
                row=1, column=0, sticky="e", padx=4, pady=2
            )
            en_id_contexto = ttk.Entry(cuerpo)
            en_id_contexto.grid(row=1, column=1, sticky="we", padx=4, pady=2)

            ttk.Label(cuerpo, text="tipo_pregunta:").grid(
                row=2, column=0, sticky="e", padx=4, pady=2
            )
            cb_tipo_pregunta = ttk.Combobox(
                cuerpo,
                state="readonly",
                values=["opcion_multiple", "abierta"],
                width=20,
            )
            cb_tipo_pregunta.set("opcion_multiple")
            cb_tipo_pregunta.grid(row=2, column=1, sticky="w", padx=4, pady=2)

            ttk.Label(cuerpo, text="tipo de contenido:").grid(
                row=3, column=0, sticky="e", padx=4, pady=2
            )
            cb_tipo = ttk.Combobox(cuerpo, state="readonly", values=["normal", "latex"])
            cb_tipo.set("normal")
            cb_tipo.grid(row=3, column=1, sticky="w", padx=4, pady=2)

            ttk.Label(cuerpo, text="contexto:").grid(
                row=4, column=0, sticky="ne", padx=4, pady=2
            )
            txt_contexto = tk.Text(cuerpo, height=3, width=80)
            txt_contexto.grid(row=4, column=1, sticky="we", padx=4, pady=2)

            ttk.Label(cuerpo, text="enunciado:").grid(
                row=5, column=0, sticky="ne", padx=4, pady=2
            )
            txt_enunciado = tk.Text(cuerpo, height=3, width=80)
            txt_enunciado.grid(row=5, column=1, sticky="we", padx=4, pady=2)

            ttk.Label(cuerpo, text="opcion_a:").grid(
                row=6, column=0, sticky="e", padx=4, pady=2
            )
            en_a = ttk.Entry(cuerpo)
            en_a.grid(row=6, column=1, sticky="we", padx=4, pady=2)

            ttk.Label(cuerpo, text="opcion_b:").grid(
                row=7, column=0, sticky="e", padx=4, pady=2
            )
            en_b = ttk.Entry(cuerpo)
            en_b.grid(row=7, column=1, sticky="we", padx=4, pady=2)

            ttk.Label(cuerpo, text="opcion_c:").grid(
                row=8, column=0, sticky="e", padx=4, pady=2
            )
            en_c = ttk.Entry(cuerpo)
            en_c.grid(row=8, column=1, sticky="we", padx=4, pady=2)

            ttk.Label(cuerpo, text="opcion_d:").grid(
                row=9, column=0, sticky="e", padx=4, pady=2
            )
            en_d = ttk.Entry(cuerpo)
            en_d.grid(row=9, column=1, sticky="we", padx=4, pady=2)

            ttk.Label(cuerpo, text="correcta:").grid(
                row=10, column=0, sticky="e", padx=4, pady=2
            )
            cb_correcta = ttk.Combobox(
                cuerpo, state="readonly", values=["A", "B", "C", "D"]
            )
            cb_correcta.grid(row=10, column=1, sticky="w", padx=4, pady=2)

            ttk.Label(cuerpo, text="imagen:").grid(
                row=11, column=0, sticky="e", padx=4, pady=2
            )
            frame_img = ttk.Frame(cuerpo)
            frame_img.grid(row=11, column=1, sticky="we", padx=4, pady=2)
            frame_img.columnconfigure(0, weight=1)

            en_imagen = ttk.Entry(frame_img)
            en_imagen.grid(row=0, column=0, sticky="we", padx=(0, 4))

            btn_imagen = ttk.Button(
                frame_img,
                text="Examinar...",
                width=10,
            )
            btn_imagen.grid(row=0, column=1, padx=(0, 4))

            lbl_preview = ttk.Label(
                frame_img, text="Sin imagen", relief="sunken", width=22
            )
            lbl_preview.grid(row=0, column=2)

            en_id_contexto.insert(0, str(datos_iniciales.get("id_contexto", "")))
            cb_tipo_pregunta.set(
                str(
                    datos_iniciales.get("tipo_pregunta", "opcion_multiple")
                    or "opcion_multiple"
                )
            )
            cb_tipo.set(
                str(datos_iniciales.get("tipo_contenido", "normal") or "normal")
            )
            txt_contexto.insert("1.0", str(datos_iniciales.get("contexto", "")))
            txt_enunciado.insert("1.0", str(datos_iniciales.get("enunciado", "")))
            en_a.insert(0, str(datos_iniciales.get("opcion_a", "")))
            en_b.insert(0, str(datos_iniciales.get("opcion_b", "")))
            en_c.insert(0, str(datos_iniciales.get("opcion_c", "")))
            en_d.insert(0, str(datos_iniciales.get("opcion_d", "")))
            if str(datos_iniciales.get("correcta", "")).strip().upper() in (
                "A",
                "B",
                "C",
                "D",
            ):
                cb_correcta.set(
                    str(datos_iniciales.get("correcta", "")).strip().upper()
                )

            img_inicial = str(datos_iniciales.get("imagen", "") or "").strip()
            if img_inicial:
                en_imagen.insert(0, img_inicial)
            _actualizar_preview(lbl_preview, img_inicial)

            btn_imagen.configure(
                command=lambda e=en_imagen, p=lbl_preview: _on_pick_image(e, p)
            )

            bloque = {
                "frame": frame,
                "cuerpo": cuerpo,
                "titulo_var": titulo_var,
                "btn_colapsar": btn_colapsar,
                "numero": len(bloques) + 1,
                "colapsado": False,
                "usar_contexto_compartido": var_ctx_bloque,
                "chk_contexto_compartido": chk_ctx_bloque,
                "btn_imagen": btn_imagen,
                "preview_imagen": lbl_preview,
                "_cache_local": {
                    "id_contexto": en_id_contexto.get().strip(),
                    "contexto": txt_contexto.get("1.0", "end").strip(),
                    "imagen": en_imagen.get().strip(),
                },
                "_tenia_contexto_compartido": False,
                "widgets": {
                    "id_contexto": en_id_contexto,
                    "tipo_pregunta": cb_tipo_pregunta,
                    "tipo_contenido": cb_tipo,
                    "contexto": txt_contexto,
                    "enunciado": txt_enunciado,
                    "opcion_a": en_a,
                    "opcion_b": en_b,
                    "opcion_c": en_c,
                    "opcion_d": en_d,
                    "correcta": cb_correcta,
                    "imagen": en_imagen,
                },
            }

            for _txt in (txt_contexto, txt_enunciado):
                _txt.bind(
                    "<MouseWheel>",
                    lambda e: (
                        canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"),
                        "break",
                    )[1],
                )

            btn_colapsar.configure(command=lambda b=bloque: _toggle_bloque(b))
            btn_eliminar.configure(command=lambda b=bloque: _eliminar_bloque(b))
            chk_ctx_bloque.configure(
                command=lambda b=bloque: _aplicar_contexto_a_bloque(b)
            )

            en_imagen.bind(
                "<KeyRelease>",
                lambda _e, p=lbl_preview, e=en_imagen: _actualizar_preview(
                    p, e.get().strip()
                ),
            )

            bloques.append(bloque)
            _renumerar_bloques()
            if bool(var_aplicar_todas.get()):
                bloque["usar_contexto_compartido"].set(True)
            _aplicar_contexto_a_bloque(bloque)

        def _leer_bloque(bloque):
            w = bloque["widgets"]
            return {
                "id_contexto": w["id_contexto"].get().strip(),
                "tipo_pregunta": w["tipo_pregunta"].get().strip().lower()
                or "opcion_multiple",
                "tipo_contenido": w["tipo_contenido"].get().strip().lower() or "normal",
                "contexto": w["contexto"].get("1.0", "end").strip(),
                "enunciado": w["enunciado"].get("1.0", "end").strip(),
                "opcion_a": w["opcion_a"].get().strip(),
                "opcion_b": w["opcion_b"].get().strip(),
                "opcion_c": w["opcion_c"].get().strip(),
                "opcion_d": w["opcion_d"].get().strip(),
                "correcta": w["correcta"].get().strip().upper(),
                "imagen": w["imagen"].get().strip(),
                "usa_contexto_compartido": bool(
                    var_aplicar_todas.get() or bloque["usar_contexto_compartido"].get()
                ),
            }

        def _generar_estructura():
            try:
                cantidad = int(str(sp_cantidad.get()).strip())
            except Exception:
                cantidad = 0

            if cantidad <= 0:
                messagebox.showerror(
                    "Validacion",
                    "La cantidad de preguntas debe ser mayor a 0.",
                    parent=d,
                )
                return

            for bloque in list(bloques):
                try:
                    bloque["frame"].destroy()
                except Exception:
                    pass
            bloques.clear()

            for _ in range(cantidad):
                _crear_bloque()

        def _anadir_bloque():
            _crear_bloque()

        ttk.Button(
            controles, text="Generar estructura", command=_generar_estructura
        ).pack(side="left", padx=(8, 4))
        ttk.Button(controles, text="Anadir pregunta", command=_anadir_bloque).pack(
            side="left", padx=(0, 4)
        )

        def _guardar():
            grado = cb_grado.get().strip()
            area = cb_area.get().strip()
            evaluacion = en_evaluacion.get().strip()
            periodo = en_periodo.get().strip()

            if not grado:
                messagebox.showerror("Validacion", "Debe seleccionar grado.", parent=d)
                return
            if not area or area == "Seleccione":
                messagebox.showerror("Validacion", "Debe seleccionar area.", parent=d)
                return
            if not evaluacion:
                messagebox.showerror("Validacion", "Debe indicar evaluacion.", parent=d)
                return
            if not bloques:
                messagebox.showerror(
                    "Validacion",
                    "Debe crear al menos un bloque de pregunta.",
                    parent=d,
                )
                return

            id_area = d._area_nombre_a_id.get(area)
            if id_area is None:
                _actualizar_combo_areas()
                id_area = d._area_nombre_a_id.get(area)
            if id_area is None:
                messagebox.showerror(
                    "Validacion",
                    "El area seleccionada no es valida para el grado.",
                    parent=d,
                )
                return

            ctx_comp = _leer_contexto_compartido()
            hay_bloques_compartidos = bool(var_aplicar_todas.get())
            if not hay_bloques_compartidos:
                hay_bloques_compartidos = any(
                    bool(b["usar_contexto_compartido"].get()) for b in bloques
                )
            if hay_bloques_compartidos:
                if not ctx_comp.get("id_contexto") or not ctx_comp.get("contexto"):
                    messagebox.showerror(
                        "Validacion",
                        "Debe completar id_contexto y contexto compartido para usarlo en preguntas.",
                        parent=d,
                    )
                    return

            firmas_existentes = set()
            try:
                if _HAS_PANDAS and hasattr(self.preguntas_df, "iterrows"):
                    for _, row in self.preguntas_df.iterrows():
                        firma = (
                            str(row.get("grado", "")).strip().lower(),
                            str(row.get("area", "")).strip().lower(),
                            str(row.get("evaluacion", "")).strip().lower(),
                            re.sub(
                                r"\s+",
                                " ",
                                str(row.get("enunciado", "")).strip().lower(),
                            ),
                        )
                        if firma[3]:
                            firmas_existentes.add(firma)
                else:
                    for row in self.preguntas_df.get("_rows", []):
                        firma = (
                            str(row.get("grado", "")).strip().lower(),
                            str(row.get("area", "")).strip().lower(),
                            str(row.get("evaluacion", "")).strip().lower(),
                            re.sub(
                                r"\s+",
                                " ",
                                str(row.get("enunciado", "")).strip().lower(),
                            ),
                        )
                        if firma[3]:
                            firmas_existentes.add(firma)
            except Exception:
                pass

            preguntas = []
            firmas_locales = set()
            requeridos = ["tipo_pregunta", "enunciado"]

            for idx, bloque in enumerate(bloques, start=1):
                datos_bloque = _leer_bloque(bloque)

                for campo in requeridos:
                    if not datos_bloque.get(campo):
                        messagebox.showerror(
                            "Validacion",
                            f"Pregunta {idx}: campo requerido '{campo}'.",
                            parent=d,
                        )
                        return

                tipo_pregunta = (
                    str(datos_bloque.get("tipo_pregunta", "opcion_multiple"))
                    .strip()
                    .lower()
                )
                if tipo_pregunta not in ("opcion_multiple", "abierta"):
                    messagebox.showerror(
                        "Validacion",
                        f"Pregunta {idx}: tipo_pregunta debe ser 'opcion_multiple' o 'abierta'.",
                        parent=d,
                    )
                    return
                datos_bloque["tipo_pregunta"] = tipo_pregunta

                if not datos_bloque.get("id_contexto"):
                    messagebox.showerror(
                        "Validacion",
                        f"Pregunta {idx}: debe indicar id_contexto.",
                        parent=d,
                    )
                    return
                if not datos_bloque.get("contexto"):
                    messagebox.showerror(
                        "Validacion",
                        f"Pregunta {idx}: debe indicar contexto.",
                        parent=d,
                    )
                    return

                if tipo_pregunta == "opcion_multiple":
                    for campo in [
                        "opcion_a",
                        "opcion_b",
                        "opcion_c",
                        "opcion_d",
                        "correcta",
                    ]:
                        if not datos_bloque.get(campo):
                            messagebox.showerror(
                                "Validacion",
                                f"Pregunta {idx}: campo requerido '{campo}' para opcion_multiple.",
                                parent=d,
                            )
                            return
                else:
                    datos_bloque["opcion_a"] = ""
                    datos_bloque["opcion_b"] = ""
                    datos_bloque["opcion_c"] = ""
                    datos_bloque["opcion_d"] = ""
                    datos_bloque["correcta"] = ""

                if tipo_pregunta == "opcion_multiple":
                    correcta = str(datos_bloque.get("correcta", "")).upper().strip()
                    if correcta not in ("A", "B", "C", "D"):
                        messagebox.showerror(
                            "Validacion",
                            f"Pregunta {idx}: la correcta debe ser A, B, C o D.",
                            parent=d,
                        )
                        return
                    datos_bloque["correcta"] = correcta

                if not self._validar_contenido_latex(
                    datos_bloque, parent=d, prefijo=f"Pregunta {idx}"
                ):
                    return

                firma = (
                    str(grado).strip().lower(),
                    str(area).strip().lower(),
                    str(evaluacion).strip().lower(),
                    re.sub(
                        r"\s+",
                        " ",
                        str(datos_bloque.get("enunciado", "")).strip().lower(),
                    ),
                )
                if firma in firmas_locales:
                    messagebox.showerror(
                        "Validacion",
                        f"Pregunta {idx}: enunciado duplicado dentro de la misma evaluacion.",
                        parent=d,
                    )
                    return
                if firma in firmas_existentes:
                    messagebox.showerror(
                        "Validacion",
                        f"Pregunta {idx}: ya existe una pregunta igual en el banco para ese grado, area y evaluacion.",
                        parent=d,
                    )
                    return

                firmas_locales.add(firma)
                preguntas.append(datos_bloque)

            ids = self._obtener_siguientes_ids_pregunta(len(preguntas))
            if len(ids) != len(preguntas):
                messagebox.showerror(
                    "Validacion",
                    "No se pudieron generar IDs para las preguntas.",
                    parent=d,
                )
                return

            payload = []
            for idx, item in enumerate(preguntas):
                payload.append(
                    {
                        "id": ids[idx],
                        "evaluacion": evaluacion,
                        "area": area,
                        "periodo": periodo,
                        "grado": grado,
                        "id_contexto": item.get("id_contexto", ""),
                        "contexto": item.get("contexto", ""),
                        "enunciado": item.get("enunciado", ""),
                        "opcion_a": item.get("opcion_a", ""),
                        "opcion_b": item.get("opcion_b", ""),
                        "opcion_c": item.get("opcion_c", ""),
                        "opcion_d": item.get("opcion_d", ""),
                        "correcta": item.get("correcta", ""),
                        "imagen": item.get("imagen", ""),
                        "tipo_pregunta": item.get("tipo_pregunta", "opcion_multiple"),
                        "tipo_contenido": item.get("tipo_contenido", "normal"),
                        "id_area": int(id_area),
                    }
                )

            d.result = payload
            d.destroy()

        btn_ctx_imagen = ttk.Button(
            frame_img_ctx,
            text="Examinar...",
            command=lambda: (
                _on_pick_image(en_ctx_imagen, lbl_ctx_preview),
                _sincronizar_contexto_compartido(),
            ),
            width=10,
        )
        btn_ctx_imagen.grid(row=0, column=1, padx=(0, 4))
        _actualizar_preview(lbl_ctx_preview, "")

        en_ctx_id.bind("<KeyRelease>", lambda _e: _sincronizar_contexto_compartido())
        txt_ctx.bind("<KeyRelease>", lambda _e: _sincronizar_contexto_compartido())
        en_ctx_imagen.bind(
            "<KeyRelease>",
            lambda _e: _sincronizar_contexto_compartido(),
        )

        frame_btns = ttk.Frame(d)
        frame_btns.pack(fill="x", padx=10, pady=(0, 10))

        ttk.Button(frame_btns, text="OK", command=_guardar).pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(frame_btns, text="Cancelar", command=d.destroy).pack(side="left")

        _generar_estructura()

        d.wait_window()
        return getattr(d, "result", None)

    def pregunta_editar(self):
        def _norm_id(v):
            s = str(v).strip()
            if not s:
                return ""
            try:
                return str(int(float(s)))
            except Exception:
                return s

        sel = self.tree_preg.selection()
        if not sel:
            messagebox.showinfo("Editar", "Seleccione una pregunta.", parent=self.win)
            return
        vals = tuple(self.tree_preg.item(sel[0]).get("values", ()))
        try:
            cols_tree = tuple(self.tree_preg["columns"])
        except Exception:
            cols_tree = ()

        row_map = {}
        if cols_tree:
            for i, col in enumerate(cols_tree):
                row_map[str(col)] = vals[i] if i < len(vals) else ""

        actuales = {
            "id": row_map.get("id", vals[0] if len(vals) > 0 else ""),
            "evaluacion": row_map.get("evaluacion", vals[1] if len(vals) > 1 else ""),
            "area": row_map.get("area", vals[2] if len(vals) > 2 else ""),
            "periodo": row_map.get("periodo", vals[3] if len(vals) > 3 else ""),
            "grado": row_map.get("grado", vals[4] if len(vals) > 4 else ""),
            "id_contexto": row_map.get("id_contexto", vals[5] if len(vals) > 5 else ""),
            "contexto": row_map.get("contexto", vals[6] if len(vals) > 6 else ""),
            "enunciado": row_map.get("enunciado", vals[7] if len(vals) > 7 else ""),
            "opcion_a": row_map.get("opcion_a", vals[8] if len(vals) > 8 else ""),
            "opcion_b": row_map.get("opcion_b", vals[9] if len(vals) > 9 else ""),
            "opcion_c": row_map.get("opcion_c", vals[10] if len(vals) > 10 else ""),
            "opcion_d": row_map.get("opcion_d", vals[11] if len(vals) > 11 else ""),
            "correcta": row_map.get("correcta", vals[12] if len(vals) > 12 else ""),
            "tipo_pregunta": row_map.get("tipo_pregunta", ""),
            "imagen": row_map.get("imagen", vals[13] if len(vals) > 13 else ""),
        }
        if not str(actuales.get("tipo_pregunta", "")).strip():
            tiene_opciones = any(
                str(actuales.get(k, "")).strip()
                for k in ("opcion_a", "opcion_b", "opcion_c", "opcion_d")
            )
            actuales["tipo_pregunta"] = (
                "opcion_multiple" if tiene_opciones else "abierta"
            )

        actuales["id_area"] = self._banco_obtener_area_id_pregunta(actuales.get("id"))
        nuevos = self._dialog_pregunta(actuales)
        if not nuevos:
            return
        id_area_nuevo = nuevos.pop("id_area", None)
        id_anterior = _norm_id(actuales.get("id", ""))
        id_nuevo = _norm_id(nuevos.get("id", ""))
        datos_sql = {k: str(v).strip() for k, v in nuevos.items()}

        try:
            actualizado = core_preguntas.actualizar_pregunta_banco(
                id_anterior,
                id=id_nuevo,
                evaluacion=datos_sql.get("evaluacion", ""),
                area=datos_sql.get("area", ""),
                periodo=datos_sql.get("periodo", ""),
                grado=datos_sql.get("grado", ""),
                id_contexto=datos_sql.get("id_contexto", ""),
                contexto=datos_sql.get("contexto", ""),
                enunciado=datos_sql.get("enunciado", ""),
                opcion_a=datos_sql.get("opcion_a", ""),
                opcion_b=datos_sql.get("opcion_b", ""),
                opcion_c=datos_sql.get("opcion_c", ""),
                opcion_d=datos_sql.get("opcion_d", ""),
                correcta=datos_sql.get("correcta", ""),
                tipo_pregunta=datos_sql.get("tipo_pregunta", "opcion_multiple"),
                imagen=datos_sql.get("imagen", ""),
                id_area=id_area_nuevo,
            )
            if not actualizado:
                messagebox.showerror(
                    "Editar",
                    "No se pudo ubicar la pregunta para actualizar. Intente recargar la lista e intente nuevamente.",
                    parent=self.win,
                )
                return
        except Exception as exc:
            messagebox.showerror(
                "Editar",
                f"No se pudo actualizar la pregunta: {exc}",
                parent=self.win,
            )
            return

        try:
            self.banco._cargar_preguntas()
        except Exception:
            pass

        self._banco_recargar_treeview()

    def pregunta_eliminar(self):
        sel = self.tree_preg.selection()
        if not sel:
            messagebox.showinfo("Eliminar", "Seleccione una pregunta.", parent=self.win)
            return
        vals = self.tree_preg.item(sel[0])["values"]
        id_pregunta = vals[0] if vals else None
        if id_pregunta is None or str(id_pregunta).strip() == "":
            messagebox.showerror(
                "Eliminar",
                "No se pudo identificar la pregunta seleccionada.",
                parent=self.win,
            )
            return
        if not messagebox.askyesno(
            "Confirmar", f"Eliminar pregunta id={id_pregunta}?", parent=self.win
        ):
            return

        try:
            core_preguntas.eliminar_pregunta_banco(id_pregunta)
        except Exception as e:
            messagebox.showerror(
                "Eliminar",
                f"No se pudo eliminar la pregunta: {e}",
                parent=self.win,
            )
            return

        self._banco_recargar_treeview()

    def _banco_asegurar_tabla_area_pregunta(self):
        try:
            core_preguntas.asegurar_tabla_area_pregunta()
        except Exception:
            pass

    def _banco_guardar_area_id_pregunta(self, pregunta_id, id_area):
        try:
            core_preguntas.guardar_area_id_pregunta(pregunta_id, id_area)
        except Exception:
            pass

    def _banco_obtener_area_id_pregunta(self, pregunta_id):
        try:
            return core_preguntas.obtener_area_id_pregunta(pregunta_id)
        except Exception:
            pass
        return None

    def _banco_eliminar_area_id_pregunta(self, pregunta_id):
        try:
            core_preguntas.eliminar_area_id_pregunta(pregunta_id)
        except Exception:
            pass

    def _banco_grados_plan_estudio(self):
        # Orden visual fijo solicitado para el módulo Banco de Preguntas.
        return [
            "0",
            "JA",
            "1",
            "2",
            "3",
            "4",
            "5",
            "6",
            "7",
            "8",
            "9",
            "10",
            "11",
            "C1",
            "C2",
            "C3",
            "C4",
            "C5",
            "C6",
        ]

    def _banco_areas_por_grado_desde_plan(self, grado):
        """Retorna lista de tuplas (id_area, nombre_area) para el grado."""
        grado_norm = str(grado or "").strip()
        if not grado_norm:
            return []
        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute(
                """
                SELECT DISTINCT IdArea
                FROM plan_estudio
                WHERE LOWER(TRIM(CAST(grado AS TEXT))) = LOWER(TRIM(?))
                  AND IdArea IS NOT NULL
                """,
                (grado_norm,),
            )
            ids = [int(r[0]) for r in cur.fetchall() if r and r[0] is not None]
            if not ids:
                conn.close()
                return []

            placeholders = ",".join(["?"] * len(ids))
            cur.execute(
                f"SELECT id, nombre FROM areas WHERE id IN ({placeholders}) ORDER BY nombre",
                tuple(ids),
            )
            rows = [
                (int(r[0]), str(r[1])) for r in cur.fetchall() if r and r[0] is not None
            ]
            conn.close()
            return rows
        except Exception:
            return []

    def _dialog_pregunta(self, inicial=None):
        d = tk.Toplevel(self.win)
        d.transient(self.win)
        d.grab_set()
        d.title("Pregunta")
        labels = [
            "id",
            "evaluacion",
            "area",
            "periodo",
            "grado",
            "id_contexto",
            "contexto",
            "enunciado",
            "opcion_a",
            "opcion_b",
            "opcion_c",
            "opcion_d",
            "correcta",
            "tipo_pregunta",
            "imagen",
        ]
        entries = {}
        combo_grado = None
        combo_area = None
        d._area_nombre_a_id = {}
        tipo_contenido_var = tk.StringVar(value="texto")

        grados_plan = self._banco_grados_plan_estudio()
        for i, lab in enumerate(labels):
            ttk.Label(d, text=lab + ":").grid(row=i, column=0, sticky="e")
            if lab == "contexto" or lab == "enunciado":
                e = tk.Text(d, height=3, width=60)
                e.grid(row=i, column=1, sticky="we")
            elif lab == "grado":
                e = ttk.Combobox(d, state="readonly", values=grados_plan)
                e.grid(row=i, column=1, sticky="we")
                combo_grado = e
            elif lab == "area":
                e = ttk.Combobox(d, state="readonly")
                e.grid(row=i, column=1, sticky="we")
                combo_area = e
            elif lab == "tipo_pregunta":
                e = ttk.Combobox(
                    d,
                    state="readonly",
                    values=["opcion_multiple", "abierta"],
                )
                e.grid(row=i, column=1, sticky="we")
                e.set("opcion_multiple")
            else:
                e = ttk.Entry(d)
                e.grid(row=i, column=1, sticky="we")
            entries[lab] = e

        def _actualizar_combo_areas_desde_grado(mostrar_alerta=True):
            if combo_grado is None or combo_area is None:
                return False
            grado_sel = combo_grado.get().strip()
            if not grado_sel:
                combo_area["values"] = ["Seleccione"]
                combo_area.set("Seleccione")
                d._area_nombre_a_id = {}
                return False

            areas = self._banco_areas_por_grado_desde_plan(grado_sel)
            if not areas:
                combo_area.set("Seleccione")
                combo_area["values"] = ["Seleccione"]
                d._area_nombre_a_id = {}
                if mostrar_alerta:
                    messagebox.showwarning(
                        "Plan de Estudios",
                        "No existe plan de estudios configurado para este grado.",
                        parent=d,
                    )
                return False

            nombres = [n for _, n in areas]
            d._area_nombre_a_id = {n: i for i, n in areas}
            anterior = combo_area.get().strip()
            combo_area["values"] = ["Seleccione"] + nombres

            # Preservar selección original aunque cambie mayúsculas/minúsculas.
            anterior_l = anterior.lower()
            match = None
            for nombre in nombres:
                if str(nombre).strip().lower() == anterior_l:
                    match = nombre
                    break

            if match:
                combo_area.set(match)
            else:
                combo_area.set("Seleccione")
            return True

        if combo_grado is not None:
            combo_grado.bind(
                "<<ComboboxSelected>>",
                lambda _e: _actualizar_combo_areas_desde_grado(mostrar_alerta=True),
            )

        def on_select_image():
            p = filedialog.askopenfilename(
                title="Seleccionar imagen",
                filetypes=[("Imágenes", "*.png;*.jpg;*.jpeg;*.gif")],
                parent=d,
            )
            if not p:
                return
            # copiar imagen a carpeta de imagenes
            try:
                dst = os.path.join(self.imagenes_dir, os.path.basename(p))
                shutil.copy2(p, dst)
                entries["imagen"].delete(0, "end")
                entries["imagen"].insert(0, os.path.relpath(dst, self.base_dir))
                # mostrar vista previa si es posible
                try:
                    if _HAS_PIL:
                        img = Image.open(dst)
                        img.thumbnail((150, 150))
                        tkimg = ImageTk.PhotoImage(img)
                        d.preview_img = tkimg
                        preview.configure(image=tkimg, text="")
                    else:
                        preview.configure(text=os.path.basename(dst))
                except Exception:
                    preview.configure(text=os.path.basename(dst))
            except Exception as e:
                messagebox.showerror("Imagen", f"Error copiando imagen: {e}", parent=d)

        tipo_frame = ttk.LabelFrame(d, text="Tipo de contenido", padding=(6, 4))
        tipo_frame.grid(
            row=len(labels), column=0, columnspan=2, sticky="we", pady=(6, 2)
        )
        ttk.Radiobutton(
            tipo_frame,
            text="Texto normal",
            variable=tipo_contenido_var,
            value="texto",
        ).pack(side="left", padx=(0, 12))
        ttk.Radiobutton(
            tipo_frame,
            text="Expresión matemática (LaTeX)",
            variable=tipo_contenido_var,
            value="latex",
        ).pack(side="left")

        lbl_latex_hint = ttk.Label(
            d,
            text="Use $...$ para fórmulas. Ejemplo: $x^{2} + \\frac{3}{4}$",
            foreground="#555",
        )
        lbl_latex_hint.grid(
            row=len(labels) + 1, column=0, columnspan=2, sticky="w", pady=(0, 4)
        )

        btn_img = ttk.Button(d, text="Seleccionar imagen...", command=on_select_image)
        btn_img.grid(row=len(labels) + 2, column=0)

        preview = ttk.Label(d, text="Sin imagen", relief="sunken", width=40)
        preview.grid(row=len(labels) + 3, column=0, columnspan=2, pady=4)

        if inicial:
            for k, v in inicial.items():
                w = entries.get(k)
                if not w:
                    continue
                if isinstance(w, tk.Text):
                    w.insert("1.0", v)
                elif isinstance(w, ttk.Combobox):
                    w.set(str(v))
                else:
                    w.insert(0, v)
            # mostrar preview de la imagen inicial si existe
            try:
                img_path = inicial.get("imagen")
                if img_path:
                    abs_path = (
                        os.path.join(self.base_dir, img_path)
                        if not os.path.isabs(img_path)
                        else img_path
                    )
                    if os.path.exists(abs_path):
                        if _HAS_PIL:
                            img = Image.open(abs_path)
                            img.thumbnail((150, 150))
                            tkimg = ImageTk.PhotoImage(img)
                            d.preview_img = tkimg
                            preview.configure(image=tkimg, text="")
                        else:
                            preview.configure(text=os.path.basename(abs_path))
            except Exception:
                pass

            # Cargar combo de áreas desde plan para el grado inicial.
            _actualizar_combo_areas_desde_grado(mostrar_alerta=False)
            try:
                campos_math = [
                    "contexto",
                    "enunciado",
                    "opcion_a",
                    "opcion_b",
                    "opcion_c",
                    "opcion_d",
                ]
                if any("$" in str(inicial.get(c, "") or "") for c in campos_math):
                    tipo_contenido_var.set("latex")
            except Exception:
                pass
            id_area_ini = inicial.get("id_area")
            if id_area_ini not in (None, ""):
                try:
                    id_area_ini = int(id_area_ini)
                    for nombre, ida in d._area_nombre_a_id.items():
                        if int(ida) == id_area_ini:
                            if combo_area is not None:
                                combo_area.set(nombre)
                            break
                except Exception:
                    pass
        else:
            if combo_area is not None:
                combo_area["values"] = ["Seleccione"]
                combo_area.set("Seleccione")

        def on_ok():
            datos = {}
            for lab in labels:
                w = entries[lab]
                if isinstance(w, tk.Text):
                    datos[lab] = w.get("1.0", "end").strip()
                else:
                    datos[lab] = w.get().strip()

            tipo_pregunta = (
                str(datos.get("tipo_pregunta", "opcion_multiple")).strip().lower()
            )
            if tipo_pregunta not in ("opcion_multiple", "abierta"):
                messagebox.showerror(
                    "Validación",
                    "tipo_pregunta debe ser 'opcion_multiple' o 'abierta'.",
                    parent=d,
                )
                return
            datos["tipo_pregunta"] = tipo_pregunta

            # Validaciones: campos obligatorios
            required = [
                "id",
                "evaluacion",
                "area",
                "enunciado",
                "tipo_pregunta",
            ]
            for r in required:
                if not datos.get(r):
                    messagebox.showerror(
                        "Validación", f"Campo requerido: {r}", parent=d
                    )
                    return

            if tipo_pregunta == "opcion_multiple":
                for r in ("opcion_a", "opcion_b", "opcion_c", "opcion_d", "correcta"):
                    if not datos.get(r):
                        messagebox.showerror(
                            "Validación", f"Campo requerido: {r}", parent=d
                        )
                        return

                correcta = str(datos.get("correcta", "")).upper().strip()
                if correcta not in ("A", "B", "C", "D"):
                    messagebox.showerror(
                        "Validación",
                        "La correcta debe ser A, B, C o D.",
                        parent=d,
                    )
                    return
                datos["correcta"] = correcta
            else:
                datos["opcion_a"] = ""
                datos["opcion_b"] = ""
                datos["opcion_c"] = ""
                datos["opcion_d"] = ""
                datos["correcta"] = ""

            if str(datos.get("area", "")).strip().lower() == "seleccione":
                messagebox.showerror(
                    "Validación",
                    "Debe seleccionar un área válida.",
                    parent=d,
                )
                return

            # área desde catálogo por grado (nombre visible, IdArea interno)
            area_nombre = datos.get("area", "").strip()
            id_area = d._area_nombre_a_id.get(area_nombre)
            if id_area is None:
                if not _actualizar_combo_areas_desde_grado(mostrar_alerta=True):
                    return
                area_nombre = datos.get("area", "").strip()
                id_area = d._area_nombre_a_id.get(area_nombre)
                if id_area is None:
                    messagebox.showerror(
                        "Validación",
                        "Seleccione un área válida del plan de estudios.",
                        parent=d,
                    )
                    return
            datos["id_area"] = int(id_area)

            campos_math = [
                "contexto",
                "enunciado",
                "opcion_a",
                "opcion_b",
                "opcion_c",
                "opcion_d",
            ]

            if tipo_contenido_var.get() == "latex" and not any(
                "$" in str(datos.get(c, "") or "") for c in campos_math
            ):
                messagebox.showerror(
                    "Validación",
                    "Seleccionó 'Expresión matemática (LaTeX)' pero no se detectaron delimitadores $...$.",
                    parent=d,
                )
                return

            for campo in campos_math:
                texto_campo = str(datos.get(campo, "") or "")
                try:
                    expresiones = _extraer_expresiones_latex(texto_campo)
                except ValueError as exc:
                    messagebox.showerror(
                        "Validación LaTeX",
                        f"Error en campo '{campo}': {exc}",
                        parent=d,
                    )
                    return

                for expr in expresiones:
                    try:
                        _validar_expresion_latex(expr)
                    except ValueError as exc:
                        messagebox.showerror(
                            "Validación LaTeX",
                            f"Error en campo '{campo}': {exc}",
                            parent=d,
                        )
                        return

            # Validación id único
            nuevo_id = datos.get("id")
            if inicial is None:
                if self._pregunta_id_exists(nuevo_id):
                    messagebox.showerror(
                        "Validación", f"El id '{nuevo_id}' ya existe.", parent=d
                    )
                    return
            else:
                if str(nuevo_id) != str(inicial.get("id")) and self._pregunta_id_exists(
                    nuevo_id
                ):
                    messagebox.showerror(
                        "Validación", f"El id '{nuevo_id}' ya existe.", parent=d
                    )
                    return

            d.result = datos
            d.destroy()

        ttk.Button(d, text="OK", command=on_ok).grid(row=len(labels) + 4, column=0)
        ttk.Button(d, text="Cancelar", command=d.destroy).grid(
            row=len(labels) + 4, column=1
        )
        d.wait_window()
        return getattr(d, "result", None)

    # ---------- Acceso Docente Integrado ----------
    def _build_maestro_tab(self):
        frame = self.tab_maestro
        for widget in frame.winfo_children():
            widget.destroy()

        # Mapeo fila->documento para replicar el comportamiento del módulo Docente.
        self._maestro_fila_documento = {}

        config_frame = ttk.LabelFrame(
            frame, text="⚙️  Configuración de Examen", padding=10
        )
        config_frame.pack(fill="x", padx=12, pady=(10, 6))

        frame_identificacion = ttk.Frame(config_frame)
        frame_identificacion.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        frame_configuracion = ttk.Frame(config_frame)
        frame_configuracion.grid(row=1, column=0, sticky="ew")

        config_frame.columnconfigure(0, weight=1)
        for i in range(4):
            frame_identificacion.columnconfigure(i, weight=1)
        for i in range(6):
            frame_configuracion.columnconfigure(i, weight=1)

        ttk.Label(frame_identificacion, text="Grado:").grid(
            row=0, column=0, sticky="w", padx=6, pady=2
        )
        self.combo_grado_cfg = ttk.Combobox(
            frame_identificacion, values=["Seleccione"], state="readonly", width=12
        )
        self.combo_grado_cfg.current(0)
        self.combo_grado_cfg.grid(row=1, column=0, sticky="ew", padx=6)
        self.combo_grado_cfg.bind(
            "<<ComboboxSelected>>", self._maestro_on_grado_selected
        )

        ttk.Label(frame_identificacion, text="Área:").grid(
            row=0, column=1, sticky="w", padx=6, pady=2
        )
        self.combo_area_cfg = ttk.Combobox(
            frame_identificacion, values=["Seleccione"], state="readonly", width=14
        )
        self.combo_area_cfg.current(0)
        self.combo_area_cfg.grid(row=1, column=1, sticky="ew", padx=6)
        self.combo_area_cfg.bind(
            "<<ComboboxSelected>>", self._maestro_on_area_selected_cfg
        )

        ttk.Label(frame_identificacion, text="Evaluación:").grid(
            row=0, column=2, sticky="w", padx=6, pady=2
        )
        self.combo_evaluacion_cfg = ttk.Combobox(
            frame_identificacion, values=["Seleccione"], state="readonly", width=14
        )
        self.combo_evaluacion_cfg.current(0)
        self.combo_evaluacion_cfg.grid(row=1, column=2, sticky="ew", padx=6)
        self.combo_evaluacion_cfg.bind(
            "<<ComboboxSelected>>", self._maestro_cargar_config_examen
        )

        ttk.Label(frame_identificacion, text="Curso:").grid(
            row=0, column=3, sticky="w", padx=6, pady=2
        )
        self.combo_curso_cfg = ttk.Combobox(
            frame_identificacion, values=["Seleccione"], state="readonly", width=12
        )
        self.combo_curso_cfg.current(0)
        self.combo_curso_cfg.grid(row=1, column=3, sticky="ew", padx=6)
        self.combo_curso_cfg.bind(
            "<<ComboboxSelected>>", self._maestro_cargar_config_examen
        )

        # Alias de compatibilidad con código previo.
        self.combo_area = self.combo_area_cfg

        ttk.Label(frame_configuracion, text="Duración (min):").grid(
            row=0, column=0, sticky="w", padx=6, pady=2
        )
        self.entry_duracion = ttk.Entry(frame_configuracion, width=6)
        self.entry_duracion.insert(0, "30")
        self.entry_duracion.grid(row=1, column=0, sticky="w", padx=6)

        ttk.Label(frame_configuracion, text="Preguntas:").grid(
            row=0, column=1, sticky="w", padx=6, pady=2
        )
        self.entry_cantidad = ttk.Entry(frame_configuracion, width=6)
        self.entry_cantidad.insert(0, "10")
        self.entry_cantidad.grid(row=1, column=1, sticky="w", padx=6)

        ttk.Label(frame_configuracion, text="Máx. Intentos:").grid(
            row=0, column=2, sticky="w", padx=6, pady=2
        )
        self.entry_max_intentos = ttk.Entry(frame_configuracion, width=6)
        self.entry_max_intentos.insert(0, "1")
        self.entry_max_intentos.grid(row=1, column=2, sticky="w", padx=6)

        ttk.Label(frame_configuracion, text="Permitir Reintentos:").grid(
            row=0, column=3, sticky="w", padx=6, pady=2
        )
        self.var_permitir_reintentos = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            frame_configuracion, variable=self.var_permitir_reintentos
        ).grid(row=1, column=3, sticky="w", padx=6)

        ttk.Label(frame_configuracion, text="Habilitar Examen:").grid(
            row=0, column=4, sticky="w", padx=6, pady=2
        )
        self.var_examen_activo = tk.BooleanVar(value=True)
        ttk.Checkbutton(frame_configuracion, variable=self.var_examen_activo).grid(
            row=1, column=4, sticky="w", padx=6
        )

        self._crear_boton_si_permiso(
            ttk.Button,
            frame_configuracion,
            "desktop.superadmin.maestro.configuracion.guardar",
            text="💾 Guardar",
            command=self.maestro_guardar_config,
            layout="grid",
            layout_kwargs={
                "row": 1,
                "column": 5,
                "sticky": "e",
                "padx": (6, 2),
            },
        )

        actions_container = tk.Frame(frame, bg="#f5f7fa")
        actions_container.pack(fill="x", padx=12, pady=(6, 0))
        actions_frame = tk.Frame(actions_container, bg="#f5f7fa")
        actions_frame.pack(side="right")

        # Botón Configuración de examen (idéntico a módulo docente)
        self._crear_boton_si_permiso(
            tk.Button,
            actions_frame,
            "desktop.superadmin.maestro.configuracion_examen",
            text="⚙️ Configuración Examen",
            command=(
                self.abrir_configuracion_examen_docente
                if hasattr(self, "abrir_configuracion_examen_docente")
                else lambda: messagebox.showinfo(
                    "No implementado", "Funcionalidad no disponible."
                )
            ),
            font=("Segoe UI", 9, "bold"),
            bg="#0f766e",
            fg="white",
            relief="flat",
            padx=10,
            pady=6,
            cursor="hand2",
            layout_kwargs={"side": "left", "padx": 4},
        )
        self._crear_boton_si_permiso(
            tk.Button,
            actions_frame,
            "desktop.superadmin.maestro.vaciar_calificaciones",
            text="Vaciar calificaciones",
            command=self.maestro_vaciar_calificaciones,
            font=("Segoe UI", 9, "bold"),
            bg="#ff6b6b",
            fg="white",
            relief="flat",
            padx=10,
            pady=6,
            cursor="hand2",
            layout_kwargs={"side": "left", "padx": 4},
        )
        self._crear_boton_si_permiso(
            tk.Button,
            actions_frame,
            "desktop.superadmin.maestro.actualizar",
            text="🔄 Actualizar Datos",
            command=self.maestro_cargar_resultados,
            font=("Segoe UI", 9, "bold"),
            bg="#17A2B8",
            fg="white",
            relief="flat",
            padx=10,
            pady=6,
            cursor="hand2",
            layout_kwargs={"side": "left", "padx": 4},
        )
        self._crear_boton_si_permiso(
            tk.Button,
            actions_frame,
            "desktop.superadmin.maestro.autoevaluacion",
            text="📝 Autoevaluación",
            command=self.maestro_abrir_autoevaluacion,
            font=("Segoe UI", 9, "bold"),
            bg="#00b894",
            fg="white",
            relief="flat",
            padx=10,
            pady=6,
            cursor="hand2",
            layout_kwargs={"side": "left", "padx": 4},
        )
        self._crear_boton_si_permiso(
            tk.Button,
            actions_frame,
            "desktop.superadmin.maestro.exportar_excel",
            text="📥 Exportar Excel",
            command=self.maestro_export_excel,
            font=("Segoe UI", 9, "bold"),
            bg="#0066cc",
            fg="white",
            relief="flat",
            padx=10,
            pady=6,
            cursor="hand2",
            layout_kwargs={"side": "left", "padx": 4},
        )
        self._crear_boton_si_permiso(
            tk.Button,
            actions_frame,
            "desktop.superadmin.maestro.exportar_consolidado",
            text="Consolidado",
            command=self.maestro_export_consolidado,
            font=("Segoe UI", 9, "bold"),
            bg="#17A2B8",
            fg="white",
            relief="flat",
            padx=10,
            pady=6,
            cursor="hand2",
            layout_kwargs={"side": "left", "padx": 4},
        )
        self._crear_boton_si_permiso(
            tk.Button,
            actions_frame,
            "desktop.superadmin.maestro.ver_detalle",
            text="🔎 Ver Detalle",
            command=self.maestro_ver_detalle_selected,
            font=("Segoe UI", 9, "bold"),
            bg="#0078D7",
            fg="white",
            relief="flat",
            padx=10,
            pady=6,
            cursor="hand2",
            layout_kwargs={"side": "left", "padx": 4},
        )
        self._crear_boton_si_permiso(
            tk.Button,
            actions_frame,
            "desktop.superadmin.maestro.autorizar_revision",
            text="⭐ Autorizar Revisión",
            command=self.maestro_autorizar_revision,
            font=("Segoe UI", 9, "bold"),
            bg="#51cf66",
            fg="white",
            relief="flat",
            padx=10,
            pady=6,
            cursor="hand2",
            layout_kwargs={"side": "left", "padx": 4},
        )
        self._crear_boton_si_permiso(
            tk.Button,
            actions_frame,
            "desktop.superadmin.maestro.resetear_nota",
            text="🔄 Resetear Nota",
            command=self.maestro_resetear_examen,
            font=("Segoe UI", 9, "bold"),
            bg="#ff9500",
            fg="white",
            relief="flat",
            padx=10,
            pady=6,
            cursor="hand2",
            layout_kwargs={"side": "left", "padx": 4},
        )

        filtros_frame = tk.LabelFrame(
            frame,
            text="🔍 Consulta de Evaluaciones de Estudiantes",
            font=("Segoe UI", 10, "bold"),
            bg="#f5f7fa",
            fg="#1a1a1a",
            padx=10,
            pady=10,
        )
        filtros_frame.pack(fill="x", padx=12, pady=(8, 0))
        filtros_row = tk.Frame(filtros_frame, bg="#f5f7fa")
        filtros_row.pack(fill="x")

        tk.Label(
            filtros_row,
            text="Grado:",
            font=("Segoe UI", 10),
            bg="#f5f7fa",
            fg="#1a1a1a",
        ).pack(side="left", padx=(0, 5))
        self.combo_filtro_grado = ttk.Combobox(
            filtros_row, values=[], state="readonly", width=8
        )
        self.combo_filtro_grado.pack(side="left", padx=(0, 12))
        self.combo_filtro_grado.bind(
            "<<ComboboxSelected>>", lambda e: self._maestro_on_grado_selected_filtro()
        )

        tk.Label(
            filtros_row,
            text="Curso:",
            font=("Segoe UI", 10),
            bg="#f5f7fa",
            fg="#1a1a1a",
        ).pack(side="left", padx=(0, 5))
        self.combo_filtro_curso = ttk.Combobox(
            filtros_row, values=[], state="readonly", width=12
        )
        self.combo_filtro_curso.pack(side="left", padx=(0, 12))
        self.combo_filtro_curso.bind(
            "<<ComboboxSelected>>", lambda e: self._maestro_on_curso_selected()
        )

        tk.Label(
            filtros_row, text="Área:", font=("Segoe UI", 10), bg="#f5f7fa", fg="#1a1a1a"
        ).pack(side="left", padx=(0, 5))
        self.combo_filtro_area = ttk.Combobox(
            filtros_row, values=[], state="readonly", width=12
        )
        self.combo_filtro_area.pack(side="left", padx=(0, 12))
        self.combo_filtro_area.bind(
            "<<ComboboxSelected>>", lambda e: self._maestro_on_area_selected()
        )

        tk.Label(
            filtros_row,
            text="Evaluación:",
            font=("Segoe UI", 10),
            bg="#f5f7fa",
            fg="#1a1a1a",
        ).pack(side="left", padx=(0, 5))
        self.combo_filtro_evaluacion = ttk.Combobox(
            filtros_row, values=[], state="readonly", width=12
        )
        self.combo_filtro_evaluacion.pack(side="left", padx=(0, 12))
        self.combo_filtro_evaluacion.bind(
            "<<ComboboxSelected>>", lambda e: self.maestro_cargar_resultados()
        )

        self._crear_boton_si_permiso(
            tk.Button,
            filtros_row,
            "desktop.superadmin.maestro.filtros.limpiar",
            text="✖️  Limpiar Filtros",
            command=self._maestro_limpiar_filtros,
            font=("Segoe UI", 9),
            bg="#ff9500",
            fg="white",
            relief="flat",
            padx=10,
            pady=4,
            cursor="hand2",
            layout_kwargs={"side": "left", "padx": 4},
        )
        tk.Label(
            filtros_row,
            text="Apellidos/Nombres/Documento:",
            font=("Segoe UI", 10),
            bg="#f5f7fa",
            fg="#1a1a1a",
        ).pack(side="left", padx=(10, 5))
        self.entry_busqueda_maestro = tk.Entry(
            filtros_row, width=24, font=("Segoe UI", 10)
        )
        self.entry_busqueda_maestro.pack(side="left", padx=(0, 10))
        self.entry_busqueda_maestro.bind(
            "<Return>", lambda e: self._maestro_buscar_estudiante_filtros()
        )
        self._crear_boton_si_permiso(
            tk.Button,
            filtros_row,
            "desktop.superadmin.maestro.filtros.buscar",
            text="🔎 Buscar",
            command=self._maestro_buscar_estudiante_filtros,
            font=("Segoe UI", 9),
            bg="#0066cc",
            fg="white",
            relief="flat",
            padx=10,
            pady=4,
            cursor="hand2",
            layout_kwargs={"side": "left"},
        )

        self._maestro_crear_tabla()
        self._maestro_refresh_grados()
        self._maestro_cargar_config_examen()
        self.maestro_cargar_resultados()

        self._maestro_auto_refresh_job = None

        def _detener_auto_refresh(_event=None):
            job_id = getattr(self, "_maestro_auto_refresh_job", None)
            if job_id:
                try:
                    self.win.after_cancel(job_id)
                except Exception:
                    pass
                self._maestro_auto_refresh_job = None

        def _auto_refresh():
            if not frame.winfo_exists() or not self.win.winfo_exists():
                _detener_auto_refresh()
                return
            try:
                self.maestro_cargar_resultados()
            except Exception:
                pass
            try:
                self._maestro_auto_refresh_job = self.win.after(5000, _auto_refresh)
            except Exception:
                pass

        try:
            frame.bind(
                "<Destroy>",
                lambda e: _detener_auto_refresh() if e.widget is frame else None,
                add="+",
            )
            self.win.bind(
                "<Destroy>",
                lambda e: _detener_auto_refresh() if e.widget is self.win else None,
                add="+",
            )
        except Exception:
            pass

        _auto_refresh()

    def _maestro_refresh_grados(self):
        """Actualiza combos de grado/curso/área en la pestaña Acceso Docente."""
        self._maestro_llenar_combo_grados()
        self._maestro_llenar_combo_areas_por_grado()

    def _maestro_todas_areas(self):
        """Devuelve todas las áreas disponibles desde el banco de preguntas."""
        try:
            return self.banco.obtener_areas_disponibles()
        except Exception:
            return []

    def _maestro_llenar_combo_grados(self):
        """Llena combos de grados/cursos (filtros y configuración) igual que Docente."""
        try:
            grados = core_matricula.listar_grados_distintos(solo_activos=True)
            cursos = core_matricula.listar_cursos_distintos(
                grado=None,
                solo_activos=True,
            )

            if hasattr(self, "combo_filtro_grado"):
                self.combo_filtro_grado["values"] = ["Todos"] + [str(g) for g in grados]
                if (
                    self.combo_filtro_grado.get()
                    not in self.combo_filtro_grado["values"]
                ):
                    self.combo_filtro_grado.current(0)

            if hasattr(self, "combo_filtro_curso"):
                self.combo_filtro_curso["values"] = ["Todos"] + [str(c) for c in cursos]
                if (
                    self.combo_filtro_curso.get()
                    not in self.combo_filtro_curso["values"]
                ):
                    self.combo_filtro_curso.current(0)

            if hasattr(self, "combo_grado_cfg"):
                valores_cfg = ["Seleccione"] + [str(g) for g in grados]
                self.combo_grado_cfg["values"] = valores_cfg
                if self.combo_grado_cfg.get() not in valores_cfg:
                    self.combo_grado_cfg.current(0)

            self._maestro_llenar_combo_cursos_cfg()
        except Exception:
            pass

    def _maestro_llenar_combo_cursos_cfg(self, grado=None):
        """Carga cursos del panel de configuración según grado."""
        try:
            if grado is None and hasattr(self, "combo_grado_cfg"):
                grado = self.combo_grado_cfg.get()

            if not grado or grado == "Seleccione":
                self.combo_curso_cfg["values"] = ["Seleccione"]
                self.combo_curso_cfg.current(0)
                return

            grado_norm = normalizar_grado(grado)
            cursos = core_matricula.listar_cursos_distintos(
                grado=grado_norm,
                solo_activos=True,
            )

            valores = ["Seleccione"] + cursos if cursos else ["Seleccione"]
            self.combo_curso_cfg["values"] = valores
            if self.combo_curso_cfg.get() not in valores:
                self.combo_curso_cfg.current(0)
        except Exception:
            try:
                self.combo_curso_cfg["values"] = ["Seleccione"]
                self.combo_curso_cfg.current(0)
            except Exception:
                pass

    def _maestro_llenar_combo_areas_por_grado(self):
        """Llena áreas para configuración en función del grado seleccionado."""
        try:
            grado_sel = None
            if hasattr(self, "combo_grado_cfg"):
                g = self.combo_grado_cfg.get()
                if g and g not in ("Todos", "Seleccione"):
                    grado_sel = g

            if grado_sel:
                areas = cargar_areas_por_grado(grado_sel)
            else:
                areas = self._maestro_todas_areas()

            valores = ["Seleccione"] + (areas if areas else ["General"])
            self.combo_area_cfg["values"] = valores
            if self.combo_area_cfg.get() not in valores:
                self.combo_area_cfg.current(0)
        except Exception:
            pass

    def _maestro_llenar_combo_evaluaciones_por_grado_area(
        self, grado=None, area=None, destino="filtro"
    ):
        """Llena evaluaciones según grado y área para filtros o configuración."""
        try:
            if not grado or not area:
                if destino == "cfg":
                    self.combo_evaluacion_cfg["values"] = ["Seleccione"]
                    self.combo_evaluacion_cfg.current(0)
                else:
                    self.combo_filtro_evaluacion["values"] = ["Todos"]
                    self.combo_filtro_evaluacion.current(0)
                return

            evaluaciones = cargar_evaluaciones_por_grado_y_area(grado, area)

            if destino == "cfg":
                valores = ["Seleccione"] + (evaluaciones if evaluaciones else [])
                self.combo_evaluacion_cfg["values"] = valores
                self.combo_evaluacion_cfg.current(0)
            else:
                valores = ["Todos"] + (evaluaciones if evaluaciones else [])
                self.combo_filtro_evaluacion["values"] = valores
                self.combo_filtro_evaluacion.current(0)
        except Exception:
            if destino == "cfg":
                try:
                    self.combo_evaluacion_cfg["values"] = ["Seleccione"]
                    self.combo_evaluacion_cfg.current(0)
                except Exception:
                    pass
            else:
                try:
                    self.combo_filtro_evaluacion["values"] = ["Todos"]
                    self.combo_filtro_evaluacion.current(0)
                except Exception:
                    pass

    def _maestro_on_grado_selected_filtro(self):
        """Actualiza curso/área/evaluación cuando cambia el grado del filtro."""
        grado = (
            self.combo_filtro_grado.get()
            if hasattr(self, "combo_filtro_grado")
            else None
        )
        if grado == "Todos":
            grado = None

        try:
            grado_norm = normalizar_grado(grado) if grado else None
            cursos = core_matricula.listar_cursos_distintos(
                grado=grado_norm,
                solo_activos=True,
            )
            self.combo_filtro_curso["values"] = ["Todos"] + cursos
            self.combo_filtro_curso.current(0)
        except Exception:
            pass

        if grado:
            areas = cargar_areas_por_grado(grado)
        else:
            areas = self._maestro_todas_areas()
        self.combo_filtro_area["values"] = ["Todos"] + (areas if areas else ["General"])
        self.combo_filtro_area.current(0)

        self.combo_filtro_evaluacion["values"] = ["Todos"]
        self.combo_filtro_evaluacion.current(0)
        self.maestro_cargar_resultados()

    def _maestro_on_curso_selected(self):
        """Actualiza filtros dependientes cuando cambia el curso."""
        grado = (
            self.combo_filtro_grado.get()
            if hasattr(self, "combo_filtro_grado")
            else None
        )
        if grado == "Todos":
            grado = None

        if grado:
            areas = cargar_areas_por_grado(grado)
        else:
            areas = self._maestro_todas_areas()

        self.combo_filtro_area["values"] = ["Todos"] + (areas if areas else ["General"])
        self.combo_filtro_area.current(0)
        self.combo_filtro_evaluacion["values"] = ["Todos"]
        self.combo_filtro_evaluacion.current(0)
        self.maestro_cargar_resultados()

    def _maestro_on_grado_selected(self, event=None):
        """Callback de configuración al seleccionar grado."""
        self._maestro_llenar_combo_cursos_cfg()
        self._maestro_llenar_combo_areas_por_grado()
        self._maestro_on_area_selected_cfg()
        try:
            self.maestro_cargar_resultados()
        except Exception:
            pass

    def _maestro_on_area_selected(self, event=None):
        """Callback de filtros al seleccionar área."""
        grado = (
            self.combo_filtro_grado.get()
            if hasattr(self, "combo_filtro_grado")
            else None
        )
        if grado == "Todos":
            grado = None
        area = (
            self.combo_filtro_area.get() if hasattr(self, "combo_filtro_area") else None
        )
        if area == "Todos":
            area = None
        self._maestro_llenar_combo_evaluaciones_por_grado_area(
            grado, area, destino="filtro"
        )
        self.maestro_cargar_resultados()

    def _maestro_on_area_selected_cfg(self, event=None):
        """Callback de configuración al seleccionar área."""
        grado = self.combo_grado_cfg.get() if hasattr(self, "combo_grado_cfg") else None
        area = self.combo_area_cfg.get() if hasattr(self, "combo_area_cfg") else None
        if grado == "Seleccione":
            grado = None
        if area == "Seleccione":
            area = None
        self._maestro_llenar_combo_evaluaciones_por_grado_area(
            grado, area, destino="cfg"
        )
        self._maestro_cargar_config_examen()

    def _maestro_limpiar_filtros(self):
        """Limpia filtros y recarga datos."""
        try:
            self.combo_filtro_grado.current(0)
            self.combo_filtro_curso.current(0)
            self.combo_filtro_area.current(0)
            self.combo_filtro_evaluacion.current(0)
            if hasattr(self, "entry_busqueda_maestro"):
                self.entry_busqueda_maestro.delete(0, tk.END)
            self.maestro_cargar_resultados()
        except Exception:
            pass

    def _maestro_buscar_estudiante_filtros(self):
        """Aplica la búsqueda por apellidos y nombres junto con los filtros actuales."""
        self.maestro_cargar_resultados()

    def _maestro_ajustar_columnas(self):
        try:
            if not hasattr(self, "tabla_maestro"):
                return
            ancho_total = self.tabla_maestro.winfo_width()
            if ancho_total <= 50:
                return

            props = [
                0.07,
                0.07,
                0.13,
                0.13,
                0.13,
                0.13,
                0.08,
                0.10,
                0.07,
                0.07,
                0.07,
                0.07,
            ]
            columnas = [
                "Grado",
                "Curso",
                "Apellido1",
                "Apellido2",
                "Nombre1",
                "Nombre2",
                "Area",
                "Evaluacion",
                "Nota",
                "Estado",
                "Fecha",
                "Duracion",
            ]
            for col, p in zip(columnas, props):
                self.tabla_maestro.column(col, width=max(60, int(ancho_total * p)))
        except Exception:
            pass

    def _maestro_crear_tabla(self):
        """Crea la tabla de Acceso Docente con la misma estructura del Docente."""
        self.frame_tabla_maestro = ttk.Frame(self.tab_maestro)
        self.frame_tabla_maestro.pack(fill="both", expand=True, padx=12, pady=(8, 10))

        scrollbar = ttk.Scrollbar(self.frame_tabla_maestro, orient="vertical")
        scrollbar.pack(side="right", fill="y")

        self.tabla_maestro = ttk.Treeview(
            self.frame_tabla_maestro,
            columns=(
                "Grado",
                "Curso",
                "Apellido1",
                "Apellido2",
                "Nombre1",
                "Nombre2",
                "Area",
                "Evaluacion",
                "Nota",
                "Estado",
                "Fecha",
                "Duracion",
            ),
            show="headings",
            yscrollcommand=scrollbar.set,
        )
        scrollbar.config(command=self.tabla_maestro.yview)

        self.tabla_maestro.heading("Grado", text="Grado")
        self.tabla_maestro.heading("Curso", text="Curso")
        self.tabla_maestro.heading("Apellido1", text="Apellido 1")
        self.tabla_maestro.heading("Apellido2", text="Apellido 2")
        self.tabla_maestro.heading("Nombre1", text="Nombre 1")
        self.tabla_maestro.heading("Nombre2", text="Nombre 2")
        self.tabla_maestro.heading("Area", text="Área")
        self.tabla_maestro.heading("Evaluacion", text="Evaluación")
        self.tabla_maestro.heading("Nota", text="Nota")
        self.tabla_maestro.heading("Estado", text="Estado")
        self.tabla_maestro.heading("Fecha", text="Fecha")
        self.tabla_maestro.heading("Duracion", text="Duración")

        for c in (
            "Grado",
            "Curso",
            "Apellido1",
            "Apellido2",
            "Nombre1",
            "Nombre2",
            "Area",
            "Evaluacion",
            "Nota",
            "Estado",
            "Fecha",
            "Duracion",
        ):
            self.tabla_maestro.column(c, stretch=True)

        self.tabla_maestro.pack(fill="both", expand=True)

        # Compatibilidad con métodos existentes.
        self.tabla = self.tabla_maestro

        try:
            self.win.bind("<Configure>", lambda e: self._maestro_ajustar_columnas())
        except Exception:
            pass

    def _maestro_cargar_config_examen(self, event=None):
        """Carga la configuración usando la misma función base que Docente."""
        try:
            from Admin import cargar_config_examen as _cargar_config_examen

            def aplicar_defaults_config():
                self.entry_duracion.delete(0, tk.END)
                self.entry_duracion.insert(0, "30")

                self.entry_cantidad.delete(0, tk.END)
                self.entry_cantidad.insert(0, "10")

                self.entry_max_intentos.delete(0, tk.END)
                self.entry_max_intentos.insert(0, "1")

                self.var_permitir_reintentos.set(True)
                self.var_examen_activo.set(True)

            grado = (
                self.combo_grado_cfg.get() if hasattr(self, "combo_grado_cfg") else None
            )
            area = (
                self.combo_area_cfg.get() if hasattr(self, "combo_area_cfg") else None
            )
            evaluacion = (
                self.combo_evaluacion_cfg.get()
                if hasattr(self, "combo_evaluacion_cfg")
                else None
            )
            curso = (
                self.combo_curso_cfg.get() if hasattr(self, "combo_curso_cfg") else None
            )

            if not area:
                return

            if grado == "Seleccione":
                grado = None
            if area == "Seleccione":
                area = None
            if evaluacion == "Seleccione":
                evaluacion = None
            if curso == "Seleccione":
                curso = None

            if not area:
                return

            if not all([grado, area, evaluacion, curso]):
                aplicar_defaults_config()
                return

            duracion, cantidad, max_intentos, permitir_reintentos, examen_activo = (
                _cargar_config_examen(area, grado, evaluacion, curso)
            )

            self.entry_duracion.delete(0, tk.END)
            self.entry_duracion.insert(0, str(duracion // 60))

            self.entry_cantidad.delete(0, tk.END)
            self.entry_cantidad.insert(0, str(cantidad))

            self.entry_max_intentos.delete(0, tk.END)
            self.entry_max_intentos.insert(0, str(max_intentos))

            self.var_permitir_reintentos.set(bool(permitir_reintentos))
            self.var_examen_activo.set(bool(examen_activo))
        except Exception as e:
            messagebox.showerror(
                "Error", f"No se pudo cargar la configuración: {e}", parent=self.win
            )

    def maestro_guardar_config(self):
        """Guarda configuración usando la misma lógica del módulo Docente."""
        if not self._requiere_permiso(
            "desktop.superadmin.maestro.configuracion.guardar"
        ):
            return
        try:
            from Admin import guardar_config_examen as _guardar_config_examen

            grado = (
                self.combo_grado_cfg.get() if hasattr(self, "combo_grado_cfg") else None
            )
            area = (
                self.combo_area_cfg.get() if hasattr(self, "combo_area_cfg") else None
            )
            evaluacion = (
                self.combo_evaluacion_cfg.get()
                if hasattr(self, "combo_evaluacion_cfg")
                else None
            )
            curso = (
                self.combo_curso_cfg.get() if hasattr(self, "combo_curso_cfg") else None
            )

            if not grado or grado == "Seleccione":
                messagebox.showerror(
                    "Error", "Debe seleccionar un grado.", parent=self.win
                )
                return
            if not area or area == "Seleccione":
                messagebox.showerror(
                    "Error", "Debe seleccionar un área.", parent=self.win
                )
                return
            if not evaluacion or evaluacion == "Seleccione":
                messagebox.showerror(
                    "Error", "Debe seleccionar una evaluación.", parent=self.win
                )
                return
            if not curso or curso == "Seleccione":
                messagebox.showerror(
                    "Error", "Debe seleccionar un curso.", parent=self.win
                )
                return

            duracion_min = int(self.entry_duracion.get())
            cantidad = int(self.entry_cantidad.get())
            max_intentos = int(self.entry_max_intentos.get())
            permitir_reintentos = 1 if self.var_permitir_reintentos.get() else 0
            habilitado = 1 if self.var_examen_activo.get() else 0

            if duracion_min <= 0 or cantidad <= 0 or max_intentos <= 0:
                messagebox.showwarning(
                    "Advertencia",
                    "Duración, cantidad e intentos deben ser mayores que 0.",
                    parent=self.win,
                )
                return

            if _guardar_config_examen(
                grado,
                area,
                evaluacion,
                duracion_min * 60,
                cantidad,
                max_intentos,
                permitir_reintentos,
                habilitado,
                curso=curso,
            ):
                reintentos_txt = "Sí" if permitir_reintentos else "No"
                messagebox.showinfo(
                    "Éxito",
                    f"Configuración guardada:\n"
                    f"- Grado: {grado}\n"
                    f"- Curso: {curso}\n"
                    f"- Área: {area}\n"
                    f"- Evaluación: {evaluacion}\n"
                    f"- Duración: {duracion_min} min\n"
                    f"- Preguntas: {cantidad}\n"
                    f"- Máx. Intentos: {max_intentos}\n"
                    f"- Permitir Reintentos: {reintentos_txt}\n"
                    f"- Habilitado: {'Sí' if habilitado else 'No'}",
                    parent=self.win,
                )
            else:
                messagebox.showerror(
                    "Error", "No se pudo guardar la configuración.", parent=self.win
                )
        except ValueError:
            messagebox.showerror(
                "Error", "Todos los campos deben ser números enteros.", parent=self.win
            )
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar: {e}", parent=self.win)

    def maestro_cargar_resultados(self):
        """Carga la tabla de estudiantes/resultados con el mismo criterio del Docente."""
        if not hasattr(self, "tabla_maestro"):
            return

        for fila in self.tabla_maestro.get_children():
            self.tabla_maestro.delete(fila)
        self._maestro_fila_documento = {}

        grado_sel = (
            self.combo_filtro_grado.get()
            if hasattr(self, "combo_filtro_grado")
            else None
        )
        if grado_sel == "Todos":
            grado_sel = None

        curso_sel = (
            self.combo_filtro_curso.get()
            if hasattr(self, "combo_filtro_curso")
            else None
        )
        if curso_sel == "Todos":
            curso_sel = None

        area_sel = (
            self.combo_filtro_area.get() if hasattr(self, "combo_filtro_area") else None
        )
        if area_sel == "Todos":
            area_sel = None

        evaluacion_sel = (
            self.combo_filtro_evaluacion.get()
            if hasattr(self, "combo_filtro_evaluacion")
            else None
        )
        if evaluacion_sel == "Todos":
            evaluacion_sel = None

        busqueda_estudiante = ""
        if hasattr(self, "entry_busqueda_maestro"):
            busqueda_estudiante = self.entry_busqueda_maestro.get().strip().upper()

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            grado_sel_norm = normalizar_grado(grado_sel) if grado_sel else None
            curso_sel_norm = curso_sel.strip().upper() if curso_sel else None
            query_estudiantes = (
                "SELECT documento, apellido1, apellido2, nombre1, nombre2, grado, curso "
                "FROM estudiantes WHERE estado = 'Activo'"
            )
            params_estudiantes = []
            if grado_sel_norm:
                query_estudiantes += " AND grado = ?"
                params_estudiantes.append(grado_sel_norm)
            if curso_sel_norm:
                query_estudiantes += " AND UPPER(TRIM(curso)) = ?"
                params_estudiantes.append(curso_sel_norm)
            if busqueda_estudiante:
                query_estudiantes += """
                    AND (
                        UPPER(TRIM(COALESCE(documento, ''))) LIKE ?
                        OR UPPER(TRIM(COALESCE(apellido1, ''))) LIKE ?
                        OR UPPER(TRIM(COALESCE(apellido2, ''))) LIKE ?
                        OR UPPER(TRIM(COALESCE(nombre1, ''))) LIKE ?
                        OR UPPER(TRIM(COALESCE(nombre2, ''))) LIKE ?
                        OR UPPER(TRIM(
                            COALESCE(apellido1, '') || ' ' ||
                            COALESCE(apellido2, '') || ' ' ||
                            COALESCE(nombre1, '') || ' ' ||
                            COALESCE(nombre2, '')
                        )) LIKE ?
                    )
                """
                termino_busqueda = f"%{busqueda_estudiante}%"
                params_estudiantes.extend([termino_busqueda] * 6)
            query_estudiantes += (
                " ORDER BY grado, curso, apellido1, apellido2, nombre1, nombre2"
            )
            cursor.execute(query_estudiantes, params_estudiantes)
            estudiantes_filtrados = cursor.fetchall()
        except Exception:
            estudiantes_filtrados = []

        for est in estudiantes_filtrados:
            documento, apellido1, apellido2, nombre1, nombre2, grado, curso = est
            documento = str(documento).strip()
            grado = str(grado).strip()
            curso = str(curso).strip()
            apellido1 = str(apellido1 or "").strip()
            apellido2 = str(apellido2 or "").strip()
            nombre1 = str(nombre1 or "").strip()
            nombre2 = str(nombre2 or "").strip()

            query = (
                "SELECT nota, estado_examen, hora_inicio, hora_fin, evaluacion, area "
                "FROM resultados "
                "WHERE documento = ? AND estado_examen IN ('FINALIZADO','PRESENTADO')"
            )
            params = [documento]
            if area_sel:
                query += " AND area = ?"
                params.append(area_sel)
            if evaluacion_sel:
                query += " AND evaluacion = ?"
                params.append(evaluacion_sel)
            query += " ORDER BY id DESC LIMIT 1"

            cursor.execute(query, params)
            resultado = cursor.fetchone()

            if resultado:
                # Ahora resultado tiene 6 campos: nota, estado_examen, hora_inicio, hora_fin, evaluacion, area
                nota, estado, hora_inicio, hora_fin, evaluacion_db, area_db = resultado
                nota_str = str(nota) if nota is not None else ""
                estado_str = str(estado).upper() if estado else ""
                fecha = hora_fin if hora_fin else hora_inicio
                evaluacion_str = evaluacion_db if evaluacion_db else ""
                area_str = area_db if area_db else ""
                duracion = ""
                try:
                    if hora_inicio and hora_fin:
                        hi = datetime.strptime(hora_inicio, "%Y-%m-%d %H:%M:%S")
                        hf = datetime.strptime(hora_fin, "%Y-%m-%d %H:%M:%S")
                        delta = hf - hi
                        total_seconds = int(delta.total_seconds())
                        horas = total_seconds // 3600
                        minutos = (total_seconds % 3600) // 60
                        segundos = total_seconds % 60
                        duracion = f"{horas:02d}:{minutos:02d}:{segundos:02d}"
                except Exception:
                    duracion = ""
            else:
                nota_str = ""
                estado_str = "EXAMEN NO PRESENTADO"
                fecha = ""
                evaluacion_str = ""
                area_str = ""
                duracion = ""

            item = self.tabla_maestro.insert(
                "",
                "end",
                values=(
                    grado,
                    curso,
                    apellido1,
                    apellido2,
                    nombre1,
                    nombre2,
                    area_str,
                    evaluacion_str,
                    nota_str,
                    estado_str,
                    fecha,
                    duracion,
                ),
            )
            self._maestro_fila_documento[item] = documento

        conn.close()

    def _get_selected_result(self):
        if not hasattr(self, "tabla_maestro"):
            return None
        sel = self.tabla_maestro.selection()
        if not sel:
            messagebox.showwarning(
                "Advertencia", "Seleccione un registro.", parent=self.win
            )
            return None
        return self.tabla_maestro.item(sel[0])["values"]

    def _maestro_get_selected_row(self):
        if not hasattr(self, "tabla_maestro"):
            return None
        sel = self.tabla_maestro.selection()
        if not sel:
            messagebox.showwarning(
                "Advertencia", "Seleccione un registro.", parent=self.win
            )
            return None
        item = sel[0]
        vals = self.tabla_maestro.item(item, "values")
        documento = self._maestro_fila_documento.get(item)
        if not documento:
            messagebox.showerror(
                "Error",
                "No se pudo determinar el documento del registro seleccionado.",
                parent=self.win,
            )
            return None
        return item, vals, documento

    def maestro_resetear_examen(self):
        if not self._requiere_permiso("desktop.superadmin.maestro.resetear_nota"):
            return
        sel = self._maestro_get_selected_row()
        if not sel:
            return

        _, vals, documento = sel
        nombre = " ".join(
            str(vals[i]).strip()
            for i in range(2, 6)
            if len(vals) > i and str(vals[i]).strip()
        )
        area_sel = vals[6] if len(vals) > 6 else None

        if not messagebox.askyesno(
            "Confirmar",
            f"¿Desea resetear la nota de {nombre}?",
            parent=self.win,
        ):
            return

        try:
            from Admin import resetear_examen as _resetear_examen

            ok = _resetear_examen(documento, area_sel)
            if ok:
                messagebox.showinfo(
                    "Éxito",
                    "Intento(s) y respuestas eliminadas correctamente.",
                    parent=self.win,
                )
                self.maestro_cargar_resultados()
                return
        except Exception:
            pass

        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                if area_sel:
                    cur.execute(
                        "DELETE FROM respuestas_estudiantes WHERE documento = ? AND area = ?",
                        (documento, area_sel),
                    )
                    cur.execute(
                        "DELETE FROM resultados WHERE documento = ? AND area = ?",
                        (documento, area_sel),
                    )
                else:
                    cur.execute(
                        "DELETE FROM respuestas_estudiantes WHERE documento = ?",
                        (documento,),
                    )
                    cur.execute(
                        "DELETE FROM resultados WHERE documento = ?",
                        (documento,),
                    )
                conn.commit()

            messagebox.showinfo(
                "Éxito",
                "Intento(s) y respuestas eliminadas correctamente.",
                parent=self.win,
            )
            self.maestro_cargar_resultados()
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"No se pudo eliminar intentos/respuestas: {e}",
                parent=self.win,
            )

    def maestro_abrir_autoevaluacion(self):
        if not self._requiere_permiso("desktop.superadmin.maestro.autoevaluacion"):
            return
        try:
            from autoevaluacion_docente_ui import VentanaAutoevaluacionDocente
        except ImportError:
            messagebox.showerror(
                "Error",
                "No se encontró el módulo de autoevaluación.",
                parent=self.win,
            )
            return

        usuario = self.usuario_actual or {}
        if isinstance(usuario, dict):
            docente_documento = str(usuario.get("documento") or "").strip()
            docente_nombre = str(usuario.get("nombre") or "").strip()
        else:
            docente_documento = str(getattr(usuario, "documento", "") or "").strip()
            docente_nombre = str(getattr(usuario, "nombre", "") or "").strip()

        if not docente_documento:
            messagebox.showwarning(
                "Autoevaluación",
                "No hay un documento disponible para abrir la autoevaluación docente.",
                parent=self.win,
            )
            return

        modulo_proxy = SimpleNamespace(nombre=docente_nombre or "Docente")
        VentanaAutoevaluacionDocente(
            self.win,
            modulo_proxy,
            docente_documento,
            db_path=self.db_path,
        )

    def maestro_vaciar_calificaciones(self):
        if not self._requiere_permiso(
            "desktop.superadmin.maestro.vaciar_calificaciones"
        ):
            return
        if not messagebox.askyesno(
            "Vaciar calificaciones",
            "¿Desea eliminar TODAS las calificaciones registradas en Acceso Docente?\n\n"
            "Esta acción eliminará todos los resultados de examen y las respuestas asociadas.\n\n"
            "Esta acción no se puede deshacer.",
            parent=self.win,
        ):
            return

        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute("SELECT COUNT(*) FROM resultados")
                total_resultados = int(cur.fetchone()[0] or 0)
                cur.execute("SELECT COUNT(*) FROM respuestas_estudiantes")
                total_respuestas = int(cur.fetchone()[0] or 0)

                cur.execute("DELETE FROM respuestas_estudiantes")
                cur.execute("DELETE FROM resultados")
                conn.commit()

            self.maestro_cargar_resultados()
            messagebox.showinfo(
                "Vaciar calificaciones",
                "Calificaciones eliminadas correctamente.\n\n"
                f"Resultados eliminados: {total_resultados}\n"
                f"Respuestas eliminadas: {total_respuestas}",
                parent=self.win,
            )
        except Exception as e:
            messagebox.showerror(
                "Vaciar calificaciones",
                f"No se pudieron eliminar las calificaciones: {e}",
                parent=self.win,
            )

    def _maestro_refresh_filters(self):
        """Compatibilidad: refresca filtros en Acceso Docente."""
        self._maestro_refresh_grados()
        try:
            self._maestro_on_grado_selected_filtro()
        except Exception:
            self.maestro_cargar_resultados()

    def maestro_autorizar_revision(self):
        """Autoriza revisión del estudiante/área seleccionados (mismo flujo de Docente)."""
        if not self._requiere_permiso("desktop.superadmin.maestro.autorizar_revision"):
            return
        sel = self._maestro_get_selected_row()
        if not sel:
            return

        _, vals, documento = sel
        area_sel = vals[6] if len(vals) > 6 else None

        try:
            from Admin import autorizar_revision as _autorizar_revision

            if area_sel:
                _autorizar_revision(documento, area_sel)
                messagebox.showinfo(
                    "Éxito",
                    f"Revisión autorizada para {documento} en {area_sel}.",
                    parent=self.win,
                )
            else:
                _autorizar_revision(documento)
                messagebox.showinfo(
                    "Éxito",
                    f"Revisión autorizada para {documento} (todas las áreas finalizadas).",
                    parent=self.win,
                )
            self.maestro_cargar_resultados()
        except Exception as e:
            messagebox.showerror(
                "Error", f"No se pudo autorizar la revisión: {e}", parent=self.win
            )

    def maestro_ver_detalle_selected(self):
        """Muestra detalle de respuestas del estudiante seleccionado."""
        sel = self._maestro_get_selected_row()
        if not sel:
            return

        _, vals, documento = sel
        area = vals[6] if len(vals) > 6 else None

        try:
            from Admin import (
                obtener_respuestas_estudiante as _obtener_respuestas_estudiante,
                mostrar_detalle_respuestas as _mostrar_detalle_respuestas,
            )

            intento_val = None
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT intento FROM resultados WHERE documento = ? AND area = ? AND estado_examen IN ('FINALIZADO','PRESENTADO') ORDER BY id DESC LIMIT 1",
                    (documento, area),
                )
                fila = cur.fetchone()
                if fila:
                    intento_val = fila[0]

            respuestas = _obtener_respuestas_estudiante(
                documento, area=area, intento=intento_val
            )

            base_dir = getattr(self, "base_dir", None)
            _mostrar_detalle_respuestas(
                self.win, documento, area, intento_val, respuestas, base_dir
            )
        except Exception as e:
            from tkinter import messagebox

            messagebox.showerror(
                "Error", f"No se pudo obtener detalle: {e}", parent=self.win
            )

    def maestro_get_filtered_rows(self, grado=None, area=None):
        sql = "SELECT id, documento, nombre, grado, area, nota, estado_examen, intento, puede_revisar FROM resultados"
        conds = []
        params = []
        if grado:
            conds.append("grado=?")
            params.append(grado)
        if area:
            conds.append("area=?")
            params.append(area)
        if conds:
            sql = sql + " WHERE " + " AND ".join(conds)
        try:
            rows = self.cur.execute(sql, tuple(params)).fetchall()
            return rows
        except Exception:
            return []

    def maestro_export_excel(self):
        """Exporta resultados con los filtros activos (igual que Docente)."""
        try:
            from Admin import (
                exportar_reporte_por_filtros as _exportar_reporte_por_filtros,
            )

            grado_sel = None
            if hasattr(self, "combo_filtro_grado"):
                g = self.combo_filtro_grado.get()
                if g and g != "Todos":
                    grado_sel = g.strip()

            curso_sel = None
            if hasattr(self, "combo_filtro_curso"):
                c = self.combo_filtro_curso.get()
                if c and c.lower() != "todos":
                    curso_sel = c.strip()

            area_sel = None
            if hasattr(self, "combo_filtro_area"):
                a = self.combo_filtro_area.get()
                if a and a.lower() != "todos":
                    area_sel = a.strip()

            evaluacion_sel = None
            if hasattr(self, "combo_filtro_evaluacion"):
                e = self.combo_filtro_evaluacion.get()
                if e and e.lower() != "todos":
                    evaluacion_sel = e.strip()

            _exportar_reporte_por_filtros(
                grado_sel, curso_sel, area_sel, evaluacion_sel
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar: {e}", parent=self.win)

    def maestro_export_consolidado(self):
        """Exporta consolidado por período (mismo comportamiento del Docente)."""
        try:
            from Admin import (
                exportar_consolidado_periodo as _exportar_consolidado_periodo,
            )

            grado_sel = None
            if hasattr(self, "combo_filtro_grado"):
                g = self.combo_filtro_grado.get()
                if g and g != "Todos":
                    grado_sel = g.strip()

            curso_sel = None
            if hasattr(self, "combo_filtro_curso"):
                c = self.combo_filtro_curso.get()
                if c and c.lower() != "todos":
                    curso_sel = c.strip()

            area_sel = None
            if hasattr(self, "combo_filtro_area"):
                a = self.combo_filtro_area.get()
                if a and a.lower() != "todos":
                    area_sel = a.strip()

            if not grado_sel:
                messagebox.showwarning(
                    "Filtro requerido",
                    "Seleccione un Grado para exportar el consolidado.",
                    parent=self.win,
                )
                return

            if not area_sel:
                messagebox.showwarning(
                    "Filtro requerido",
                    "Seleccione un Área para exportar el consolidado.",
                    parent=self.win,
                )
                return

            _exportar_consolidado_periodo(grado_sel, area_sel, curso_sel)
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"No se pudo exportar el consolidado: {e}",
                parent=self.win,
            )

    def maestro_export_pdf(self):
        rows = self.maestro_get_filtered_rows(
            (
                None
                if not hasattr(self, "cb_grado")
                else (None if self.cb_grado.get() == "(Todos)" else self.cb_grado.get())
            ),
            (
                None
                if not hasattr(self, "cb_area")
                else (None if self.cb_area.get() == "(Todos)" else self.cb_area.get())
            ),
        )
        if not rows:
            messagebox.showinfo(
                "Exportar", "No hay datos para exportar.", parent=self.win
            )
            return
        path = filedialog.asksaveasfilename(
            title="Guardar reporte PDF",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            parent=self.win,
        )
        if not path:
            return
        headers = [
            "id",
            "documento",
            "nombre",
            "grado",
            "area",
            "nota",
            "estado_examen",
            "intento",
            "puede_revisar",
        ]
        try:
            if _HAS_REPORTLAB:
                w, h = A4
                c = canvas.Canvas(path, pagesize=A4)
                c.setFont("Helvetica-Bold", 14)
                c.drawString(40, h - 40, "Reporte de Calificaciones")
                c.setFont("Helvetica", 10)
                y = h - 70
                row_h = 14
                # header
                c.setFont("Helvetica-Bold", 9)
                x = 40
                col_w = (w - 80) / len(headers)
                for i, hh in enumerate(headers):
                    c.drawString(x + i * col_w, y, hh)
                y -= row_h
                c.setFont("Helvetica", 9)
                for r in rows:
                    if y < 40:
                        c.showPage()
                        y = h - 40
                        c.setFont("Helvetica-Bold", 9)
                        for i, hh in enumerate(headers):
                            c.drawString(x + i * col_w, y, hh)
                        y -= row_h
                        c.setFont("Helvetica", 9)
                    for i, cell in enumerate(r):
                        txt = str(cell)
                        c.drawString(x + i * col_w, y, txt[: int(col_w / 6)])
                    y -= row_h
                c.save()
                messagebox.showinfo(
                    "Exportar", f"PDF guardado en {path}", parent=self.win
                )
                return
            else:
                if _HAS_PIL:
                    from PIL import ImageDraw

                    pages = []
                    # A4 approx in points 595x842
                    pw, ph = 595, 842
                    draw_font = None
                    for page_start in range(0, len(rows), 40):
                        img = Image.new("RGB", (pw, ph), "white")
                        d = ImageDraw.Draw(img)
                        y = 30
                        d.text((40, y), "Reporte de Calificaciones", fill="black")
                        y += 24
                        for r in rows[page_start : page_start + 40]:
                            line = " | ".join(str(x) for x in r)
                            d.text((40, y), line, fill="black")
                            y += 18
                        pages.append(img.convert("RGB"))
                    pages[0].save(path, save_all=True, append_images=pages[1:])
                    messagebox.showinfo(
                        "Exportar", f"PDF guardado en {path}", parent=self.win
                    )
                    return
        except Exception as e:
            messagebox.showerror(
                "Exportar PDF", f"Error generando PDF: {e}", parent=self.win
            )
            return
        messagebox.showerror(
            "Exportar PDF",
            "No es posible generar PDF: instale reportlab o Pillow.",
            parent=self.win,
        )

    # ---------- Generación de exámenes en PDF ----------
    def _build_examenes_tab(self):
        """Construye la interfaz para generar exámenes en PDF."""

        frame = self.tab_examenes
        # (Eliminado título visual destacado)

        # SECCIÓN 1: Generar exámenes en PDF
        lbl1 = ttk.Label(
            frame,
            text="Sección 1. Generar exámenes en PDF",
            font=("Segoe UI", 13, "bold"),
            foreground="#1a237e",
        )
        lbl1.pack(fill="x", padx=8, pady=(8, 0))
        seccion_generar = ttk.LabelFrame(frame, text="", padding=8)
        seccion_generar.pack(fill="x", padx=8, pady=(0, 2))
        toolbar = ttk.Frame(seccion_generar)
        toolbar.pack(fill="x")
        ttk.Label(toolbar, text="Grado:").pack(side="left", padx=2)
        self.cb_examen_grado = ttk.Combobox(toolbar, state="readonly", width=10)
        self.cb_examen_grado.pack(side="left")
        self.cb_examen_grado.bind("<<ComboboxSelected>>", self._on_examen_grado_changed)
        ttk.Label(toolbar, text="Curso:").pack(side="left", padx=2)
        self.cb_examen_curso = ttk.Combobox(toolbar, state="readonly", width=10)
        self.cb_examen_curso.pack(side="left")
        self.cb_examen_curso.bind("<<ComboboxSelected>>", self._on_examen_curso_changed)
        ttk.Label(toolbar, text="Área:").pack(side="left", padx=2)
        self.cb_examen_area = ttk.Combobox(toolbar, state="readonly", width=15)
        self.cb_examen_area.pack(side="left")
        self.cb_examen_area.bind("<<ComboboxSelected>>", self._on_examen_area_changed)
        ttk.Label(toolbar, text="Evaluación:").pack(side="left", padx=2)
        self.cb_examen_evaluacion = ttk.Combobox(toolbar, state="readonly", width=15)
        self.cb_examen_evaluacion.pack(side="left")
        ttk.Label(toolbar, text="Estudiante:").pack(side="left", padx=2)
        self.cb_examen_estudiante = ttk.Combobox(toolbar, state="readonly", width=25)
        self.cb_examen_estudiante.pack(side="left")
        # Opciones de generación
        opciones = ttk.Frame(seccion_generar)
        opciones.pack(fill="x", pady=(8, 0))
        self.var_examen_tipo = tk.StringVar(value="individual")
        ttk.Radiobutton(
            opciones,
            text="Individual",
            variable=self.var_examen_tipo,
            value="individual",
            command=self._update_examen_mode,
        ).pack(side="left", padx=6)
        ttk.Radiobutton(
            opciones,
            text="Todos",
            variable=self.var_examen_tipo,
            value="todos",
            command=self._update_examen_mode,
        ).pack(side="left")
        ttk.Radiobutton(
            opciones,
            text="Todos (1 PDF)",
            variable=self.var_examen_tipo,
            value="todos_un_pdf",
            command=self._update_examen_mode,
        ).pack(side="left", padx=(6, 0))
        self._crear_boton_si_permiso(
            ttk.Button,
            opciones,
            "desktop.superadmin.evaluaciones.generar_pdf",
            text="Generar PDF",
            command=self.examen_generar,
            layout_kwargs={"side": "left", "padx": 6},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            opciones,
            "desktop.superadmin.evaluaciones.actualizar_preguntas",
            text="Actualizar preguntas",
            command=self.recargar_datos_generador_examenes,
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )

        # SECCIÓN 2: Acciones principales
        lbl2 = ttk.Label(
            frame,
            text="Sección 2. Generar cuadernillos en PDF",
            font=("Segoe UI", 13, "bold"),
            foreground="#1a237e",
        )
        lbl2.pack(fill="x", padx=8, pady=(8, 0))
        seccion_acciones = ttk.LabelFrame(frame, text="", padding=8)
        seccion_acciones.pack(fill="x", padx=8, pady=(0, 2))
        toolbar2 = ttk.Frame(seccion_acciones)
        toolbar2.pack(fill="x")
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar2,
            "desktop.superadmin.evaluaciones.generar_cuadernillo_multi_area",
            text="Generar cuadernillo multi-área",
            command=self._generar_cuadernillo_multi_area,
            style="TButton",
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )

        # SECCIÓN 3: Consulta de examen por ID
        lbl3 = ttk.Label(
            frame,
            text="Sección 3. Consulta de examen por ID",
            font=("Segoe UI", 13, "bold"),
            foreground="#1a237e",
        )
        lbl3.pack(fill="x", padx=8, pady=(8, 0))
        seccion_consulta = ttk.LabelFrame(frame, text="", padding=8)
        seccion_consulta.pack(fill="x", padx=8, pady=(0, 2))
        frm_consulta_top = ttk.Frame(seccion_consulta)
        frm_consulta_top.pack(fill="x")
        ttk.Label(frm_consulta_top, text="ID Examen:").pack(side="left", padx=(0, 4))
        self.entry_consulta_id = ttk.Entry(frm_consulta_top, width=28)
        self.entry_consulta_id.pack(side="left")
        self.entry_consulta_id.bind(
            "<Return>", lambda e: self._consultar_examen_inline()
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            frm_consulta_top,
            "desktop.superadmin.evaluaciones.consultar_examen",
            text="Consultar",
            command=self._consultar_examen_inline,
            layout_kwargs={"side": "left", "padx": (6, 0)},
        )
        self.lbl_consulta_msg = ttk.Label(frm_consulta_top, text="", foreground="red")
        self.lbl_consulta_msg.pack(side="left", padx=(8, 0))
        frm_resultado = ttk.Frame(seccion_consulta)
        frm_resultado.pack(fill="x", pady=(4, 0))
        for col_i, (etiqueta, attr) in enumerate(
            [
                ("Estudiante:", "lbl_c_nombre"),
                ("Área:", "lbl_c_area"),
                ("Fecha:", "lbl_c_fecha"),
            ]
        ):
            ttk.Label(frm_resultado, text=etiqueta).grid(
                row=0, column=col_i * 2, sticky="e", padx=(8 if col_i else 0, 2)
            )
            lv = ttk.Label(frm_resultado, text="—", width=22, anchor="w")
            lv.grid(row=0, column=col_i * 2 + 1, sticky="w")
            setattr(self, attr, lv)
        self.btn_abrir_pdf = ttk.Button(
            frm_resultado,
            text="Abrir PDF",
            state="disabled",
            command=self._abrir_pdf_consultado,
        )
        self.btn_abrir_pdf.grid(row=0, column=6, padx=(16, 0))
        self._ruta_pdf_consultado = None

        # SECCIÓN 4: Acciones secundarias
        lbl4 = ttk.Label(
            frame,
            text="Sección 4. Calificación por cámara",
            font=("Segoe UI", 13, "bold"),
            foreground="#1a237e",
        )
        lbl4.pack(fill="x", padx=8, pady=(8, 0))
        seccion_secundarias = ttk.LabelFrame(frame, text="", padding=8)
        seccion_secundarias.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        toolbar_secundario = ttk.Frame(seccion_secundarias)
        toolbar_secundario.pack(fill="x")
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar_secundario,
            "desktop.superadmin.evaluaciones.calificar_camara",
            text="Calificar con Cámara",
            command=self._abrir_calificacion_camara,
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar_secundario,
            "desktop.superadmin.evaluaciones.ver_resultados_camara",
            text="Ver resultados",
            command=self._abrir_historial_camara,
            layout_kwargs={"side": "left"},
        )
        self.lbl_historial_camara_msg = ttk.Label(
            toolbar_secundario,
            text="",
            foreground="#455a64",
        )
        self.lbl_historial_camara_msg.pack(side="left", padx=(10, 0))

        self._cargar_historial_camara_inline()

    def _generar_cuadernillo_multi_area(self):
        """Rediseño visual del generador de cuadernillo multi-área (ICFES)."""

        def cancelar():
            d.destroy()

        try:
            from multi_area_exam_pdf import write_multi_area_exam_pdf
        except ImportError:
            messagebox.showerror(
                "Dependencia faltante",
                "No se encontró el módulo para cuadernillo multi-área.\nVerifique la instalación.",
                parent=self.win,
            )
            return

        d = tk.Toplevel(self.win)
        d.title("Generar cuadernillo multi-área")
        d.resizable(True, True)
        try:
            d.state("zoomed")
        except Exception:
            pass

        # --- CONTENEDOR PRINCIPAL CON SCROLL ---
        canvas = tk.Canvas(d, borderwidth=0, highlightthickness=0)
        vsb = ttk.Scrollbar(d, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        frame_contenido = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=frame_contenido, anchor="nw")

        def _on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        frame_contenido.bind("<Configure>", _on_frame_configure)

        # Permitir scroll con rueda del mouse
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # --- CARD: DATOS GENERALES ---
        card_datos = ttk.LabelFrame(frame_contenido, text="Datos generales", padding=12)
        card_datos.pack(fill="x", padx=12, pady=(12, 6))
        card_datos.columnconfigure(1, weight=1)
        ttk.Label(card_datos, text="Grado:").grid(row=0, column=0, sticky="w", pady=2)
        grados = self.banco.obtener_grados_disponibles()
        cb_grado = ttk.Combobox(card_datos, values=grados, state="readonly", width=16)
        cb_grado.grid(row=0, column=1, sticky="ew", pady=2)
        ttk.Label(card_datos, text="Curso:").grid(
            row=0, column=2, sticky="w", padx=(12, 0), pady=2
        )
        cb_curso = ttk.Combobox(card_datos, values=[], state="disabled", width=16)
        cb_curso.grid(row=0, column=3, sticky="ew", pady=2)
        lbl_curso_msg = ttk.Label(
            card_datos, text="", foreground="#b71c1c", font=("Segoe UI", 9, "italic")
        )
        lbl_curso_msg.grid(row=1, column=3, sticky="w", pady=(0, 2))

        # --- NUEVA SECCIÓN: MODO DE GENERACIÓN ---
        frame_modo = ttk.LabelFrame(
            frame_contenido, text="Modo de generación", padding=8
        )
        frame_modo.pack(fill="x", padx=12, pady=(0, 6))
        var_modo = tk.StringVar(value="individual")
        rb_ind = ttk.Radiobutton(
            frame_modo, text="Individual", variable=var_modo, value="individual"
        )
        rb_todos = ttk.Radiobutton(
            frame_modo, text="Todos", variable=var_modo, value="todos"
        )
        rb_todos1 = ttk.Radiobutton(
            frame_modo, text="Todos (1 PDF)", variable=var_modo, value="todos1"
        )
        rb_ind.grid(row=0, column=0, padx=4, pady=2, sticky="w")
        rb_todos.grid(row=0, column=1, padx=4, pady=2, sticky="w")
        rb_todos1.grid(row=0, column=2, padx=4, pady=2, sticky="w")

        # --- CAMPO ESTUDIANTE ---
        ttk.Label(frame_modo, text="Estudiante:").grid(
            row=1, column=0, sticky="w", pady=(6, 2)
        )
        cb_estudiante = ttk.Combobox(frame_modo, values=[], state="readonly", width=30)
        cb_estudiante.grid(row=1, column=1, columnspan=2, sticky="ew", pady=(6, 2))

        # Lógica igual a generador de exámenes: cargar_estudiantes_por_grado y mapeo
        estudiantes_map = {}

        def actualizar_estudiantes(*_):
            grado = cb_grado.get()
            curso = cb_curso.get()
            vals = []
            estudiantes_map.clear()
            if not grado:
                cb_estudiante["values"] = []
                cb_estudiante.set("")
                return
            # Si el combo de curso está deshabilitado o vacío, cargar todos los estudiantes del grado
            curso_val = (
                curso if cb_curso.cget("state") == "readonly" and curso else None
            )
            estudiantes = cargar_estudiantes_por_grado(grado, curso=curso_val)
            for e in estudiantes:
                nombre = e.get("nombre") or e.get("Nombre") or ""
                doc = e.get("documento") or e.get("id") or ""
                label = f"{nombre} ({doc})"
                if label in estudiantes_map:
                    label = f"{label}_{len(vals)}"
                estudiantes_map[label] = e
                vals.append(label)
            cb_estudiante["values"] = vals
            if vals:
                cb_estudiante.set(vals[0])
            else:
                cb_estudiante.set("")

        cb_grado.bind("<<ComboboxSelected>>", actualizar_estudiantes)
        cb_curso.bind("<<ComboboxSelected>>", actualizar_estudiantes)
        actualizar_estudiantes()

        def actualizar_estado_estudiante(*_):
            modo = var_modo.get()
            if modo == "individual":
                cb_estudiante.config(state="readonly")
                actualizar_estudiantes()
            else:
                cb_estudiante.set("")
                cb_estudiante.config(state="disabled")

        var_modo.trace_add("write", lambda *_: actualizar_estado_estudiante())
        actualizar_estado_estudiante()

        def actualizar_cursos(*_):
            grado = cb_grado.get()
            cb_curso.set("")
            cb_curso["values"] = []
            cb_curso.config(state="disabled")
            lbl_curso_msg.config(text="")
            # Depuración: mostrar grado seleccionado
            print(f"[DEBUG] Grado seleccionado: {grado}")
            if grado:
                cursos = cargar_cursos_por_grado(grado) or []
                print(f"[DEBUG] Cursos detectados para grado '{grado}': {cursos}")
                cursos = list(dict.fromkeys([str(c).strip() for c in cursos if c]))
                if cursos:
                    cb_curso["values"] = cursos
                    cb_curso.config(state="readonly")
                    cb_curso.current(0)
                    cb_curso.update_idletasks()
                    cb_curso.event_generate("<<ComboboxSelected>>")
                    lbl_curso_msg.config(text=f"Cursos detectados: {', '.join(cursos)}")
                else:
                    cb_curso.set("")
                    cb_curso.config(state="disabled")
                    lbl_curso_msg.config(
                        text="No existen cursos para el grado seleccionado."
                    )
            else:
                cb_curso.set("")
                cb_curso.config(state="disabled")
                lbl_curso_msg.config(text="Seleccione un grado para ver cursos.")

        # Forzar actualización de cursos al abrir el diálogo y al hacer focus en cb_grado
        def forzar_actualizacion_cursos(event=None):
            actualizar_cursos()

        cb_grado.bind("<FocusIn>", forzar_actualizacion_cursos)
        cb_grado.bind("<Button-1>", forzar_actualizacion_cursos)
        cb_grado.bind("<<ComboboxSelected>>", actualizar_cursos)
        actualizar_cursos()

        cb_grado.bind("<<ComboboxSelected>>", actualizar_cursos)
        actualizar_cursos()

        # --- CARD: ÁREAS Y CONFIGURACIÓN ---
        frame_areas_scroll = ttk.LabelFrame(
            frame_contenido, text="Áreas y configuración", padding=0
        )
        frame_areas_scroll.pack(fill="both", expand=True, padx=12, pady=6)

        card_areas = ttk.Frame(frame_areas_scroll, padding=12)
        card_areas.pack(fill="both", expand=True)

        # Encabezados de tabla
        headers = ["", "Área", "Evaluación", "Textos", "Preguntas"]
        for col, h in enumerate(headers):
            ttk.Label(card_areas, text=h, font=("Segoe UI", 10, "bold")).grid(
                row=0, column=col, padx=2, pady=2
            )
        area_vars = []
        eval_vars = []
        textos_vars = []
        preguntas_vars = []
        area_checks = []

        def cargar_areas(*_):
            for w in card_areas.winfo_children():
                if int(w.grid_info().get("row", 0)) > 0 and not isinstance(
                    w, ttk.Label
                ):
                    w.destroy()
            area_vars.clear()
            eval_vars.clear()
            textos_vars.clear()
            preguntas_vars.clear()
            area_checks.clear()
            grado = cb_grado.get()
            curso = cb_curso.get()
            areas = self.banco.obtener_areas_disponibles(grado)
            # --- Label para total de preguntas ---
            total_preguntas_var = tk.StringVar(value="0")
            label_total = ttk.Label(
                card_areas,
                textvariable=total_preguntas_var,
                font=("Segoe UI", 10, "bold"),
                foreground="#1a237e",
            )

            def actualizar_total_preguntas(*_):
                total = 0
                for i, var in enumerate(area_vars):
                    if var.get():
                        try:
                            n = int(preguntas_vars[i].get())
                            total += n
                        except Exception:
                            pass
                total_preguntas_var.set(f"Total preguntas del cuadernillo: {total}")

            # Eliminar label total si existe
            label_total.grid_forget()
            for i, area in enumerate(areas):
                var_chk = tk.BooleanVar(value=False)
                area_vars.append(var_chk)
                chk = ttk.Checkbutton(card_areas, variable=var_chk)
                chk.grid(row=i + 1, column=0, padx=2)
                area_checks.append(chk)
                ttk.Label(card_areas, text=area).grid(
                    row=i + 1, column=1, sticky="w", padx=2
                )
                evals = self.banco.obtener_evaluaciones_disponibles(grado, area) or []
                var_eval = tk.StringVar(value=evals[0] if evals else "")
                cb_eval = ttk.Combobox(
                    card_areas,
                    values=evals,
                    textvariable=var_eval,
                    state="readonly",
                    width=14,
                )
                cb_eval.grid(row=i + 1, column=2, padx=2)
                eval_vars.append(var_eval)

                # --- Cálculo de textos y preguntas disponibles ---
                # Se actualiza dinámicamente al cambiar la evaluación
                def actualizar_disponibles(
                    area_idx=i,
                    area_name=area,
                    var_eval=var_eval,
                    var_textos=None,
                    var_pregs=None,
                ):
                    grado_sel = cb_grado.get()
                    eval_sel = var_eval.get()
                    df = self.banco.obtener_preguntas_filtradas(
                        grado=grado_sel, area=area_name, evaluacion=eval_sel
                    )
                    n_preguntas = len(df)
                    n_textos = (
                        df["contexto"].nunique()
                        if not df.empty and "contexto" in df.columns
                        else 0
                    )
                    if var_textos is not None:
                        var_textos.set(str(n_textos))
                    if var_pregs is not None:
                        var_pregs.set(str(n_preguntas))

                var_textos = tk.StringVar()
                var_pregs = tk.StringVar()
                # Inicializar con los valores actuales
                actualizar_disponibles(i, area, var_eval, var_textos, var_pregs)
                entry_textos = ttk.Entry(
                    card_areas, textvariable=var_textos, width=5, state="disabled"
                )
                entry_textos.grid(row=i + 1, column=3, padx=2)
                textos_vars.append(var_textos)
                entry_pregs = ttk.Entry(
                    card_areas, textvariable=var_pregs, width=5, state="disabled"
                )
                entry_pregs.grid(row=i + 1, column=4, padx=2)
                preguntas_vars.append(var_pregs)

                # Cuando cambia la evaluación, actualizar los disponibles
                def on_eval_change(
                    event=None,
                    area_idx=i,
                    area_name=area,
                    var_eval=var_eval,
                    var_textos=var_textos,
                    var_pregs=var_pregs,
                ):
                    actualizar_disponibles(
                        area_idx, area_name, var_eval, var_textos, var_pregs
                    )
                    actualizar_total_preguntas()

                cb_eval.bind("<<ComboboxSelected>>", on_eval_change)

                def on_preguntas_change(*_, idx=i):
                    actualizar_total_preguntas()

                var_pregs.trace_add("write", on_preguntas_change)
                var_chk.trace_add(
                    "write", lambda *_, idx=i: actualizar_total_preguntas()
                )

                def toggle_entries(var=var_chk, et=entry_textos, ep=entry_pregs):
                    state = "normal" if var.get() else "disabled"
                    et.config(state=state)
                    ep.config(state=state)

                var_chk.trace_add(
                    "write",
                    lambda *_, v=var_chk, et=entry_textos, ep=entry_pregs: toggle_entries(
                        v, et, ep
                    ),
                )

            # Mostrar el label de total debajo de la última fila
            label_total.grid(
                row=len(areas) + 1, column=0, columnspan=5, sticky="w", pady=(8, 0)
            )
            actualizar_total_preguntas()

        cb_grado.bind("<<ComboboxSelected>>", cargar_areas)
        cb_curso.bind("<<ComboboxSelected>>", cargar_areas)
        cargar_areas()

        # --- CARD: CONFIGURACIÓN DEL CUADERNILLO ---
        card_cfg = ttk.LabelFrame(
            frame_contenido, text="Configuración del cuadernillo", padding=12
        )
        card_cfg.pack(fill="x", padx=12, pady=6)
        ttk.Label(card_cfg, text="Cantidad de versiones:").grid(
            row=0, column=0, sticky="w", pady=2
        )
        var_versiones = tk.StringVar(value="1")
        entry_versiones = ttk.Entry(card_cfg, textvariable=var_versiones, width=5)
        entry_versiones.grid(row=0, column=1, sticky="w", pady=2)
        ttk.Label(card_cfg, text="Formato:").grid(
            row=0, column=2, sticky="w", padx=(12, 0), pady=2
        )
        formatos = ["Estándar", "Prueba SABER"]
        cb_formato = ttk.Combobox(card_cfg, values=formatos, state="readonly", width=16)
        cb_formato.current(0)
        cb_formato.grid(row=0, column=3, sticky="w", pady=2)

        # --- CARD: HOJA DE RESPUESTAS ---
        card_hoja = ttk.LabelFrame(
            frame_contenido, text="Generación de hoja de respuestas", padding=12
        )
        card_hoja.pack(fill="x", padx=12, pady=6)
        var_hoja = tk.StringVar(value="ninguna")
        radios = [
            ("No generar hoja de respuestas", "ninguna"),
            ("Generar en página adicional", "adicional"),
            ("Generar en la última página", "final"),
        ]
        for i, (txt, val) in enumerate(radios):
            ttk.Radiobutton(card_hoja, text=txt, variable=var_hoja, value=val).grid(
                row=0, column=i, padx=8, pady=2, sticky="w"
            )

        # --- CARD: ARCHIVO DE SALIDA ---
        card_archivo = ttk.LabelFrame(
            frame_contenido, text="Archivo de salida", padding=12
        )
        card_archivo.pack(fill="x", padx=12, pady=6)
        ttk.Label(card_archivo, text="Ruta/nombre del PDF:").grid(
            row=0, column=0, sticky="w", pady=2
        )
        entry_pdf = ttk.Entry(card_archivo, width=36)
        entry_pdf.grid(row=0, column=1, sticky="ew", pady=2)

        def seleccionar_pdf():
            f = filedialog.asksaveasfilename(
                parent=d,
                title="Guardar cuadernillo como...",
                defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
            )
            if f:
                entry_pdf.delete(0, tk.END)
                entry_pdf.insert(0, f)

        ttk.Button(card_archivo, text="...", width=3, command=seleccionar_pdf).grid(
            row=0, column=2, padx=2
        )

        # --- BOTONES ---
        # Botones dentro del área con scroll
        btns = ttk.Frame(frame_contenido)
        btns.pack(fill="x", padx=12, pady=(8, 12))

        def aceptar():
            import tempfile
            import os
            from tkinter import messagebox

            modo = var_modo.get()
            grado = cb_grado.get()
            curso = cb_curso.get()
            estudiante_label = cb_estudiante.get()
            # Validar grado y curso
            if not grado:
                messagebox.showerror("Faltan datos", "Seleccione el grado.", parent=d)
                return
            if not curso:
                messagebox.showerror("Faltan datos", "Seleccione un curso.", parent=d)
                return
            # Áreas seleccionadas
            areas = []
            preguntas_por_area = {}
            textos_por_area = {}
            evaluaciones_por_area = {}
            for i, var in enumerate(area_vars):
                if var.get():
                    area = self.banco.obtener_areas_disponibles(grado)[i]
                    areas.append(area)
                    try:
                        n_p = int(preguntas_vars[i].get())
                        n_t = int(textos_vars[i].get())
                        if n_p <= 0 or n_t < 0:
                            raise ValueError
                        preguntas_por_area[area] = n_p
                        textos_por_area[area] = n_t
                        evaluaciones_por_area[area] = eval_vars[i].get()
                    except Exception:
                        messagebox.showerror(
                            "Cantidad inválida",
                            f"Ingrese cantidades válidas para el área '{area}'.",
                            parent=d,
                        )
                        return
            if not areas:
                messagebox.showerror(
                    "Faltan datos", "Seleccione al menos un área.", parent=d
                )
                return
            pdf_path = entry_pdf.get().strip()
            if not pdf_path:
                messagebox.showerror(
                    "Falta archivo", "Indique el archivo PDF de salida.", parent=d
                )
                return

            # Validación de modo Individual
            if modo == "individual" and not estudiante_label:
                messagebox.showerror(
                    "Faltan datos", "Seleccione un estudiante.", parent=d
                )
                return

            # Obtener lista de estudiantes según modo
            estudiantes = []
            if modo == "individual":
                estudiante = estudiantes_map.get(estudiante_label)
                if not estudiante:
                    messagebox.showerror(
                        "Faltan datos", "Seleccione un estudiante válido.", parent=d
                    )
                    return
                estudiantes = [estudiante]
            elif modo in ("todos", "todos1"):
                # Siempre recargar la lista de estudiantes según grado y curso seleccionados
                curso_val = (
                    curso if cb_curso.cget("state") == "readonly" and curso else None
                )
                estudiantes = cargar_estudiantes_por_grado(grado, curso=curso_val)
                if not estudiantes:
                    messagebox.showerror(
                        "Sin estudiantes",
                        "No hay estudiantes en el curso seleccionado.",
                        parent=d,
                    )
                    return

            d.destroy()

            # Lógica de generación según modo
            if modo == "individual":
                # Generar solo para el estudiante seleccionado
                est = estudiantes[0]
                nombre_estudiante = f"{est.get('nombre1','')} {est.get('nombre2','')} {est.get('apellido1','')} {est.get('apellido2','')}".strip()
                if not nombre_estudiante:
                    nombre_estudiante = "estudiante"
                nombre_estudiante = nombre_estudiante.replace(" ", "_")
                doc = est.get("documento") or est.get("id") or ""
                nombre_archivo = (
                    os.path.splitext(pdf_path)[0] + f"_{nombre_estudiante}_{doc}.pdf"
                )
                self._ejecutar_generar_cuadernillo_multi_area(
                    grado,
                    evaluaciones_por_area,
                    areas,
                    preguntas_por_area,
                    nombre_archivo,
                    estudiante=est,
                )
            elif modo == "todos":
                # Generar un PDF por cada estudiante
                for est in estudiantes:
                    # Generar nombre del estudiante seguro para el archivo
                    nombre_estudiante = construir_nombre(est).strip()
                    if not nombre_estudiante:
                        nombre_estudiante = "estudiante"
                    nombre_estudiante = nombre_estudiante.replace(" ", "_")
                    doc = est.get("documento") or est.get("id") or ""
                    nombre_archivo = (
                        os.path.splitext(pdf_path)[0]
                        + f"_{nombre_estudiante}_{doc}.pdf"
                    )
                    self._ejecutar_generar_cuadernillo_multi_area(
                        grado,
                        evaluaciones_por_area,
                        areas,
                        preguntas_por_area,
                        nombre_archivo,
                        estudiante=est,
                    )
                messagebox.showinfo(
                    "Cuadernillos generados",
                    f"Se generaron {len(estudiantes)} cuadernillos individuales.",
                    parent=self.win,
                )
            elif modo == "todos1":
                # Generar todos los PDFs y unificarlos
                try:
                    from PyPDF2 import PdfMerger
                except ImportError:
                    try:
                        from pypdf import PdfMerger
                    except ImportError:
                        messagebox.showerror(
                            "Dependencia faltante",
                            "Debe instalar PyPDF2 o pypdf para unificar PDFs.",
                            parent=self.win,
                        )
                        return
                merger = PdfMerger()
                temp_files = []
                try:
                    for est in estudiantes:
                        with tempfile.NamedTemporaryFile(
                            delete=False, suffix=".pdf"
                        ) as tmp:
                            temp_files.append(tmp.name)
                            self._ejecutar_generar_cuadernillo_multi_area(
                                grado,
                                evaluaciones_por_area,
                                areas,
                                preguntas_por_area,
                                tmp.name,
                                estudiante=est,
                            )
                            merger.append(tmp.name)
                    merger.write(pdf_path)
                    merger.close()
                    messagebox.showinfo(
                        "Cuadernillo consolidado",
                        f"Se generó un solo PDF con {len(estudiantes)} cuadernillos.",
                        parent=self.win,
                    )
                finally:
                    for f in temp_files:
                        try:
                            os.remove(f)
                        except Exception:
                            pass

        ttk.Button(btns, text="Cancelar", command=cancelar).pack(
            side="right", padx=(0, 8)
        )
        ttk.Button(btns, text="Generar", command=aceptar).pack(side="right")

    def _consultar_examen_inline(self):
        """Busca un examen por su código ID y muestra el resultado inline en el tab."""
        import os as _os
        import re as _re
        import sys
        import subprocess

        # limpiar estado anterior
        self.lbl_consulta_msg.config(text="", foreground="red")
        self.lbl_c_nombre.config(text="—")
        self.lbl_c_area.config(text="—")
        self.lbl_c_fecha.config(text="—")
        self.btn_abrir_pdf.config(state="disabled")
        self._ruta_pdf_consultado = None

        raw = self.entry_consulta_id.get().strip()
        if not raw:
            self.lbl_consulta_msg.config(text="Ingrese el ID del examen.")
            return

        # soporte QR: SEA|doc|nombre|grado|curso|area|eval|ID:XXXXXX
        m = _re.search(r"ID:([A-Za-z0-9]+)", raw, _re.IGNORECASE)
        codigo = m.group(1).upper() if m else raw.upper()

        try:
            examen = core_examenes_generacion.obtener_examen_generado_por_codigo(codigo)
            row = None
            if examen:
                row = (
                    examen.get("exam_code"),
                    examen.get("grado"),
                    examen.get("curso"),
                    examen.get("area"),
                    examen.get("evaluacion"),
                    examen.get("version"),
                    examen.get("estudiante_nombre"),
                    examen.get("estudiante_documento"),
                    examen.get("ruta_pdf"),
                    examen.get("fecha_generacion"),
                )
        except Exception as exc:
            import traceback

            self.lbl_consulta_msg.config(text=f"Error BD: {exc}")
            messagebox.showerror(
                "Error al consultar", traceback.format_exc(), parent=self.win
            )
            return

        if row is None:
            self.lbl_consulta_msg.config(
                text="El examen no fue encontrado.", foreground="red"
            )
            return

        (
            exam_code,
            grado,
            curso,
            area,
            evaluacion,
            version,
            nombre,
            documento,
            ruta_pdf,
            fecha,
        ) = row

        self.lbl_c_nombre.config(text=f"{nombre or '—'} ({documento or '—'})")
        version_txt = f" - V{version}" if str(version or "").strip() else ""
        self.lbl_c_area.config(text=f"{area or '—'} – {evaluacion or '—'}{version_txt}")
        self.lbl_c_fecha.config(text=str(fecha or "—")[:16])

        self._ruta_pdf_consultado = ruta_pdf
        existe = bool(ruta_pdf and _os.path.exists(ruta_pdf))

        if existe:
            self.lbl_consulta_msg.config(
                text=f"Examen encontrado: {exam_code}  |  {grado} {curso}",
                foreground="green",
            )
            self.btn_abrir_pdf.config(state="normal")
            # abrir automáticamente
            try:
                if sys.platform.startswith("win"):
                    _os.startfile(ruta_pdf)
                elif sys.platform.startswith("darwin"):
                    subprocess.Popen(["open", ruta_pdf])
                else:
                    subprocess.Popen(["xdg-open", ruta_pdf])
            except Exception as exc:
                messagebox.showerror("Error al abrir PDF", str(exc), parent=self.win)
        else:
            self.lbl_consulta_msg.config(
                text="Registro encontrado pero el archivo PDF ya no existe.",
                foreground="#c07000",
            )
            self.btn_abrir_pdf.config(state="disabled")

    def _abrir_pdf_consultado(self):
        """Abre manualmente el PDF de la última consulta exitosa."""
        import os as _os
        import sys
        import subprocess

        ruta = self._ruta_pdf_consultado
        if not ruta or not _os.path.exists(ruta):
            messagebox.showwarning(
                "Archivo no encontrado",
                f"El archivo PDF ya no existe en:\n{ruta}",
                parent=self.win,
            )
            return
        try:
            if sys.platform.startswith("win"):
                _os.startfile(ruta)
            elif sys.platform.startswith("darwin"):
                subprocess.Popen(["open", ruta])
            else:
                subprocess.Popen(["xdg-open", ruta])
        except Exception as exc:
            messagebox.showerror("Error al abrir PDF", str(exc), parent=self.win)

    def _cargar_historial_camara_inline(self):
        try:
            data = core_examenes.listar_calificaciones_camara(limit=100, offset=0)
            rows = data.get("calificaciones", [])
        except Exception as exc:
            if hasattr(self, "lbl_historial_camara_msg"):
                self.lbl_historial_camara_msg.config(
                    text=f"No fue posible cargar el historial: {exc}",
                    foreground="red",
                )
            return

        if hasattr(self, "lbl_historial_camara_msg"):
            if rows:
                self.lbl_historial_camara_msg.config(
                    text=f"Mostrando las últimas {len(rows)} lecturas registradas. Use 'Ver resultados' para abrir la tabla completa.",
                    foreground="#455a64",
                )
            else:
                self.lbl_historial_camara_msg.config(
                    text="Aún no hay lecturas registradas por cámara.",
                    foreground="#455a64",
                )

        if hasattr(self, "_cam_hist_tree") and self._cam_hist_tree:
            self._llenar_historial_camara_tree(self._cam_hist_tree, rows)
        if hasattr(self, "_cam_hist_status_lbl") and self._cam_hist_status_lbl:
            if rows:
                self._cam_hist_status_lbl.config(
                    text=f"Mostrando las últimas {len(rows)} lecturas registradas.",
                    foreground="#455a64",
                )
            else:
                self._cam_hist_status_lbl.config(
                    text="Aún no hay lecturas registradas por cámara.",
                    foreground="#455a64",
                )

    def _abrir_historial_camara(self):
        if (
            hasattr(self, "_cam_hist_win")
            and self._cam_hist_win
            and self._cam_hist_win.winfo_exists()
        ):
            self._cam_hist_win.focus_force()
            self._cargar_historial_camara_inline()
            return

        d = tk.Toplevel(self.win)
        d.title("Resultados de calificación por cámara")
        d.geometry("1180x520")
        d.minsize(980, 420)
        self._cam_hist_win = d

        top = ttk.Frame(d, padding=10)
        top.pack(fill="x")
        ttk.Label(
            top,
            text="Historial de lecturas realizadas por cámara",
            font=("Segoe UI", 12, "bold"),
        ).pack(side="left")
        ttk.Button(
            top,
            text="Actualizar",
            command=self._cargar_historial_camara_inline,
        ).pack(side="right")

        self._cam_hist_status_lbl = ttk.Label(
            d,
            text="",
            foreground="#455a64",
        )
        self._cam_hist_status_lbl.pack(anchor="w", padx=10, pady=(0, 8))

        historial_wrap = ttk.Frame(d, padding=(10, 0, 10, 10))
        historial_wrap.pack(fill="both", expand=True)
        cols_historial = (
            "fecha",
            "id_examen",
            "documento",
            "estudiante",
            "grado",
            "curso",
            "area",
            "evaluacion",
            "resultado",
            "nota",
        )
        tree = ttk.Treeview(
            historial_wrap,
            columns=cols_historial,
            show="headings",
            height=16,
        )
        encabezados = {
            "fecha": "Fecha",
            "id_examen": "ID examen",
            "documento": "Documento",
            "estudiante": "Estudiante",
            "grado": "Grado",
            "curso": "Curso",
            "area": "Área",
            "evaluacion": "Evaluación",
            "resultado": "Resultado",
            "nota": "Nota",
        }
        anchos = {
            "fecha": 145,
            "id_examen": 110,
            "documento": 120,
            "estudiante": 250,
            "grado": 70,
            "curso": 70,
            "area": 150,
            "evaluacion": 150,
            "resultado": 110,
            "nota": 80,
        }
        for col in cols_historial:
            tree.heading(col, text=encabezados[col])
            tree.column(
                col,
                width=anchos[col],
                anchor="center" if col in {"grado", "curso", "nota"} else "w",
                stretch=col in {"estudiante", "area", "evaluacion"},
            )

        hist_y = ttk.Scrollbar(historial_wrap, orient="vertical", command=tree.yview)
        hist_x = ttk.Scrollbar(historial_wrap, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=hist_y.set, xscrollcommand=hist_x.set)
        tree.grid(row=0, column=0, sticky="nsew")
        hist_y.grid(row=0, column=1, sticky="ns")
        hist_x.grid(row=1, column=0, sticky="ew")
        historial_wrap.columnconfigure(0, weight=1)
        historial_wrap.rowconfigure(0, weight=1)

        self._cam_hist_tree = tree

        def _cerrar_historial_camara():
            self._cam_hist_tree = None
            self._cam_hist_status_lbl = None
            self._cam_hist_win = None
            d.destroy()

        d.protocol("WM_DELETE_WINDOW", _cerrar_historial_camara)
        self._cargar_historial_camara_inline()

    def _llenar_historial_camara_tree(self, tree, rows):
        for item in tree.get_children():
            tree.delete(item)

        for row in rows:
            tree.insert(
                "",
                "end",
                values=(
                    row.get("fecha_registro", "") or "",
                    row.get("id_examen", "") or "",
                    row.get("documento", "") or "",
                    row.get("estudiante_nombre", "") or "",
                    row.get("grado", "") or "",
                    row.get("curso", "") or "",
                    row.get("area", "") or "",
                    row.get("evaluacion", "") or "",
                    f"{row.get('correctas', 0) or 0}/{row.get('total_preguntas', 0) or 0}",
                    self._formatear_nota_historial_camara(row.get("nota")),
                ),
            )

    def _formatear_nota_historial_camara(self, nota):
        try:
            return f"{float(nota):.2f}"
        except Exception:
            return ""

    def _abrir_calificacion_camara(self):
        """Abre la ventana de calificación automática por cámara (QR + OMR)."""
        if not _HAS_CV2:
            messagebox.showerror(
                "Dependencia faltante",
                "No se encontró OpenCV. Instale 'opencv-python' y 'numpy' para usar esta función.",
                parent=self.win,
            )
            return

        if hasattr(self, "_cam_win") and self._cam_win and self._cam_win.winfo_exists():
            self._cam_win.focus_force()
            return

        d = tk.Toplevel(self.win)
        d.title("Calificar con Cámara")
        d.geometry("980x720")
        d.minsize(900, 650)
        self._cam_win = d
        self._cam_cap = None
        self._cam_running = False
        self._cam_scan_lock = False
        self._cam_after_id = None
        self._cam_qr_detector = cv2.QRCodeDetector()
        self._cam_last_exam_code = ""
        self._cam_exam_confirm_count = 0

        top = ttk.Frame(d, padding=8)
        top.pack(fill="x")
        ttk.Label(
            top,
            text="Ubique la hoja dentro del recuadro. El sistema captura automáticamente cuando detecta QR + respuestas.",
        ).pack(anchor="w")

        self._cam_status_lbl = ttk.Label(
            top,
            text="Estado: esperando iniciar cámara...",
            foreground="#0b5",
        )
        self._cam_status_lbl.pack(anchor="w", pady=(4, 0))

        preview_wrap = ttk.Frame(d, padding=(8, 0, 8, 0))
        preview_wrap.pack(fill="both", expand=True)
        self._cam_preview_lbl = ttk.Label(preview_wrap)
        self._cam_preview_lbl.pack(fill="both", expand=True)

        resultado = ttk.LabelFrame(d, text="Resultado", padding=8)
        resultado.pack(fill="x", padx=8, pady=6)
        self._cam_res_est = ttk.Label(resultado, text="Estudiante: —")
        self._cam_res_est.grid(row=0, column=0, sticky="w", padx=4)
        self._cam_res_area = ttk.Label(resultado, text="Área: —")
        self._cam_res_area.grid(row=0, column=1, sticky="w", padx=8)
        self._cam_res_id = ttk.Label(resultado, text="ID Examen: —")
        self._cam_res_id.grid(row=0, column=2, sticky="w", padx=8)
        self._cam_res_score = ttk.Label(resultado, text="Preguntas correctas: — / —")
        self._cam_res_score.grid(row=1, column=0, sticky="w", padx=4, pady=(6, 0))
        self._cam_res_nota = ttk.Label(
            resultado, text="Nota final: —", font=("", 11, "bold")
        )
        self._cam_res_nota.grid(row=1, column=1, sticky="w", padx=8, pady=(6, 0))

        btns = ttk.Frame(d, padding=(8, 0, 8, 10))
        btns.pack(fill="x")
        self._cam_btn_start = ttk.Button(
            btns,
            text="Iniciar Cámara",
            command=self._iniciar_calificacion_camara,
        )
        self._cam_btn_start.pack(side="left")
        self._cam_btn_stop = ttk.Button(
            btns,
            text="Detener",
            state="disabled",
            command=self._detener_calificacion_camara,
        )
        self._cam_btn_stop.pack(side="left", padx=(6, 0))
        ttk.Button(btns, text="Cerrar", command=self._cerrar_calificacion_camara).pack(
            side="right"
        )

        d.protocol("WM_DELETE_WINDOW", self._cerrar_calificacion_camara)
        self._iniciar_calificacion_camara()

    def _iniciar_calificacion_camara(self):
        if (
            not hasattr(self, "_cam_win")
            or not self._cam_win
            or not self._cam_win.winfo_exists()
        ):
            return
        if self._cam_running:
            return

        cap = (
            cv2.VideoCapture(0, cv2.CAP_DSHOW)
            if sys.platform.startswith("win")
            else cv2.VideoCapture(0)
        )
        if not cap or not cap.isOpened():
            messagebox.showerror(
                "Cámara",
                "No se pudo acceder a la cámara del equipo.",
                parent=self._cam_win,
            )
            return

        self._cam_cap = cap
        self._cam_running = True
        self._cam_scan_lock = False
        self._cam_last_exam_code = ""
        self._cam_exam_confirm_count = 0
        self._cam_status_lbl.config(text="Estado: cámara activa. Buscando QR y hoja...")
        self._cam_btn_start.config(state="disabled")
        self._cam_btn_stop.config(state="normal")
        self._loop_calificacion_camara()

    def _detener_calificacion_camara(self):
        self._cam_running = False
        if (
            hasattr(self, "_cam_after_id")
            and self._cam_after_id
            and hasattr(self, "_cam_win")
        ):
            try:
                self._cam_win.after_cancel(self._cam_after_id)
            except Exception:
                pass
        self._cam_after_id = None

        cap = getattr(self, "_cam_cap", None)
        if cap is not None:
            try:
                cap.release()
            except Exception:
                pass
        self._cam_cap = None

        if hasattr(self, "_cam_btn_start"):
            try:
                self._cam_btn_start.config(state="normal")
                self._cam_btn_stop.config(state="disabled")
            except Exception:
                pass
        if hasattr(self, "_cam_status_lbl"):
            self._cam_status_lbl.config(text="Estado: cámara detenida.")

    def _cerrar_calificacion_camara(self):
        self._detener_calificacion_camara()
        if hasattr(self, "_cam_win") and self._cam_win and self._cam_win.winfo_exists():
            self._cam_win.destroy()
        self._cam_win = None

    def _loop_calificacion_camara(self):
        if not self._cam_running or self._cam_cap is None:
            return

        ok, frame = self._cam_cap.read()
        if not ok or frame is None:
            self._cam_status_lbl.config(
                text="Estado: no se pudo leer la cámara.", foreground="red"
            )
            self._cam_after_id = self._cam_win.after(
                120, self._loop_calificacion_camara
            )
            return

        h, w = frame.shape[:2]
        gx1, gy1 = int(w * 0.08), int(h * 0.08)
        gx2, gy2 = int(w * 0.92), int(h * 0.92)
        cv2.rectangle(frame, (gx1, gy1), (gx2, gy2), (60, 220, 60), 2)
        cv2.putText(
            frame,
            "Ubique la hoja dentro del recuadro",
            (gx1, max(20, gy1 - 10)),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.65,
            (60, 220, 60),
            2,
            cv2.LINE_AA,
        )

        qr_text, _points, _ = self._cam_qr_detector.detectAndDecode(frame)
        parsed = self._parse_qr_payload(qr_text)
        exam_code = parsed.get("id_examen", "") if parsed else ""
        if exam_code:
            cv2.putText(
                frame,
                f"QR ID: {exam_code}",
                (gx1, gy2 + 24 if gy2 + 24 < h - 6 else h - 8),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.6,
                (0, 200, 255),
                2,
                cv2.LINE_AA,
            )
            if exam_code == self._cam_last_exam_code:
                self._cam_exam_confirm_count += 1
            else:
                self._cam_last_exam_code = exam_code
                self._cam_exam_confirm_count = 1
        else:
            self._cam_last_exam_code = ""
            self._cam_exam_confirm_count = 0

        if exam_code and self._cam_exam_confirm_count >= 6 and not self._cam_scan_lock:
            self._cam_scan_lock = True
            try:
                success, msg, debug_frame = self._procesar_captura_calificacion(
                    frame.copy(), parsed
                )
            finally:
                self._cam_scan_lock = False

            if debug_frame is not None:
                frame = debug_frame

            if success:
                self._cam_status_lbl.config(text=f"Estado: {msg}", foreground="green")
                self._detener_calificacion_camara()
                try:
                    if sys.platform.startswith("win"):
                        import winsound

                        winsound.Beep(1100, 180)
                    else:
                        self._cam_win.bell()
                except Exception:
                    pass
            else:
                self._cam_status_lbl.config(text=f"Estado: {msg}", foreground="#c07000")

        if _HAS_PIL:
            try:
                rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                img = Image.fromarray(rgb)
                max_w = max(640, self._cam_preview_lbl.winfo_width())
                max_h = max(360, self._cam_preview_lbl.winfo_height())
                img.thumbnail((max_w, max_h))
                tkimg = ImageTk.PhotoImage(img)
                self._cam_preview_lbl.configure(image=tkimg)
                self._cam_preview_lbl.image = tkimg
            except Exception:
                pass

        self._cam_after_id = self._cam_win.after(30, self._loop_calificacion_camara)

    def _parse_qr_payload(self, qr_text):
        txt = str(qr_text or "").strip()
        if not txt:
            return {}

        # Soporta QR en formato JSON: {"examen_id":"...","version":"A",...}
        if txt.startswith("{") and txt.endswith("}"):
            try:
                payload = json.loads(txt)
                exam_id_json = (
                    str(
                        payload.get("examen_id")
                        or payload.get("id_examen")
                        or payload.get("exam_id")
                        or ""
                    )
                    .strip()
                    .upper()
                )
                return {
                    "raw": txt,
                    "id_examen": exam_id_json,
                    "documento": str(payload.get("documento", "") or "").strip(),
                    "nombre": str(payload.get("nombre", "") or "").strip(),
                    "grado": str(payload.get("grado", "") or "").strip(),
                    "curso": str(payload.get("curso", "") or "").strip(),
                    "area": str(payload.get("area", "") or "").strip(),
                    "evaluacion": str(payload.get("evaluacion", "") or "").strip(),
                    "version": str(payload.get("version", "") or "").strip().upper(),
                }
            except Exception:
                pass

        m = re.search(r"ID:([A-Za-z0-9]+)", txt, re.IGNORECASE)
        exam_id = m.group(1).upper() if m else ""

        data = {
            "raw": txt,
            "id_examen": exam_id,
            "documento": "",
            "nombre": "",
            "grado": "",
            "curso": "",
            "area": "",
            "evaluacion": "",
            "version": "",
        }

        parts = txt.split("|")
        if len(parts) >= 8 and str(parts[0]).strip().upper() == "SEA":
            data["documento"] = str(parts[1]).strip()
            data["nombre"] = str(parts[2]).strip()
            data["grado"] = str(parts[3]).strip()
            data["curso"] = str(parts[4]).strip()
            data["area"] = str(parts[5]).strip()
            data["evaluacion"] = str(parts[6]).strip()
            if not data["id_examen"]:
                m2 = re.search(r"ID:([A-Za-z0-9]+)", str(parts[7]), re.IGNORECASE)
                if m2:
                    data["id_examen"] = m2.group(1).upper()
        return data

    def _procesar_captura_calificacion(self, frame, qr_data):
        exam_code = str(qr_data.get("id_examen", "") or "").strip().upper()
        if not exam_code:
            return False, "No se detectó ID de examen en el QR.", frame

        try:
            self.cur.execute(
                """SELECT numero_pregunta, respuesta_correcta
                     FROM detalle_examen
                    WHERE id_examen = ?
                    ORDER BY numero_pregunta""",
                (exam_code,),
            )
            rows = self.cur.fetchall()
        except Exception as exc:
            return False, f"Error consultando detalle_examen: {exc}", frame

        if not rows:
            return False, f"No hay clave guardada para el examen ID {exam_code}", frame

        respuestas_correctas = {}
        for num, resp in rows:
            try:
                n = int(num)
            except Exception:
                continue
            letra = str(resp or "").strip().upper()[:1]
            if letra in ("A", "B", "C", "D"):
                respuestas_correctas[n] = letra

        total = len(respuestas_correctas)
        if total == 0:
            return False, "La clave del examen no contiene respuestas válidas.", frame

        respuestas_estudiante, debug_frame = self._detectar_respuestas_omr(frame, total)
        if len(respuestas_estudiante) == 0:
            return False, "No se detectaron burbujas marcadas en la hoja.", debug_frame

        correctas = 0
        for n_preg, resp_corr in respuestas_correctas.items():
            if respuestas_estudiante.get(n_preg, "") == resp_corr:
                correctas += 1

        nota = round((correctas / float(total)) * 100.0, 2)

        nombre = qr_data.get("nombre") or "N/D"
        area = qr_data.get("area") or "N/D"
        self._cam_res_est.config(text=f"Estudiante: {nombre}")
        self._cam_res_area.config(text=f"Área: {area}")
        self._cam_res_id.config(text=f"ID Examen: {exam_code}")
        self._cam_res_score.config(text=f"Preguntas correctas: {correctas} / {total}")
        self._cam_res_nota.config(text=f"Nota final: {nota}")

        self._guardar_calificacion_camara(
            qr_data, exam_code, respuestas_estudiante, correctas, total, nota
        )
        return True, f"Calificado: {correctas}/{total} - Nota {nota}", debug_frame

    def _detectar_respuestas_omr(self, frame, total_preguntas):
        """Detecta respuestas OMR para captura móvil con corrección de perspectiva."""
        if frame is None:
            return {}, frame

        def _ordenar_puntos(pts):
            rect = np.zeros((4, 2), dtype="float32")
            s = pts.sum(axis=1)
            rect[0] = pts[np.argmin(s)]
            rect[2] = pts[np.argmax(s)]
            diff = np.diff(pts, axis=1)
            rect[1] = pts[np.argmin(diff)]
            rect[3] = pts[np.argmax(diff)]
            return rect

        def _warp_perspectiva(img, pts):
            rect = _ordenar_puntos(pts)
            (tl, tr, br, bl) = rect
            width_a = np.linalg.norm(br - bl)
            width_b = np.linalg.norm(tr - tl)
            max_w = int(max(width_a, width_b))
            height_a = np.linalg.norm(tr - br)
            height_b = np.linalg.norm(tl - bl)
            max_h = int(max(height_a, height_b))
            if max_w < 50 or max_h < 50:
                return img
            dst = np.array(
                [[0, 0], [max_w - 1, 0], [max_w - 1, max_h - 1], [0, max_h - 1]],
                dtype="float32",
            )
            m = cv2.getPerspectiveTransform(rect, dst)
            return cv2.warpPerspective(img, m, (max_w, max_h))

        def _sheet_column_ranges(total):
            total = max(1, int(total or 1))
            if total <= 30:
                split = max(1, (total + 1) // 2)
                ranges = [(1, split)]
                if split < total:
                    ranges.append((split + 1, total))
            else:
                ranges = []
                start = 1
                while start <= total and len(ranges) < 3:
                    end = min(total, start + 29)
                    ranges.append((start, end))
                    start = end + 1
                if start <= total:
                    ranges.append((start, total))

            while len(ranges) < 2:
                ranges.append((0, 0))
            if len(ranges) > 4:
                merged = ranges[:3]
                merged.append((ranges[3][0], ranges[-1][1]))
                ranges = merged
            return ranges

        def _find_document_quad(bin_img):
            cnts, _ = cv2.findContours(
                bin_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
            )
            if not cnts:
                return None

            frame_area = float(bin_img.shape[0] * bin_img.shape[1])
            for c in sorted(cnts, key=cv2.contourArea, reverse=True)[:20]:
                area = float(cv2.contourArea(c))
                if area < frame_area * 0.18:
                    continue
                peri = cv2.arcLength(c, True)
                approx = cv2.approxPolyDP(c, 0.02 * peri, True)
                if len(approx) == 4:
                    return approx.reshape(4, 2).astype("float32")
            return None

        def _detect_corner_square(th_img, corner_name):
            h, w = th_img.shape[:2]
            roi_w = max(90, int(w * 0.2))
            roi_h = max(90, int(h * 0.2))

            if corner_name == "TL":
                x0, y0 = 0, 0
            elif corner_name == "TR":
                x0, y0 = w - roi_w, 0
            elif corner_name == "BL":
                x0, y0 = 0, h - roi_h
            else:
                x0, y0 = w - roi_w, h - roi_h

            roi = th_img[y0 : y0 + roi_h, x0 : x0 + roi_w]
            if roi.size == 0:
                return None

            cnts, _ = cv2.findContours(roi, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
            roi_area = float(roi_w * roi_h)
            best = None
            best_score = -1.0

            for c in cnts:
                area = float(cv2.contourArea(c))
                if area < roi_area * 0.004 or area > roi_area * 0.35:
                    continue

                x, y, bw, bh = cv2.boundingRect(c)
                if bw <= 0 or bh <= 0:
                    continue
                ar = float(bw) / float(bh)
                if ar < 0.65 or ar > 1.35:
                    continue

                rect_area = float(bw * bh)
                solidity = area / rect_area if rect_area > 0 else 0.0
                if solidity < 0.70:
                    continue

                peri = cv2.arcLength(c, True)
                approx = cv2.approxPolyDP(c, 0.04 * peri, True)
                if len(approx) < 4 or len(approx) > 8:
                    continue

                m = cv2.moments(c)
                if m["m00"] == 0:
                    continue
                cx = int(m["m10"] / m["m00"])
                cy = int(m["m01"] / m["m00"])

                score = area * solidity
                if score > best_score:
                    best_score = score
                    best = np.array([x0 + cx, y0 + cy], dtype="float32")

            return best

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        blur = cv2.GaussianBlur(gray, (5, 5), 0)
        th_doc = cv2.adaptiveThreshold(
            blur,
            255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY_INV,
            41,
            9,
        )

        page = _find_document_quad(th_doc)
        warped = _warp_perspectiva(frame, page) if page is not None else frame.copy()

        g2 = cv2.cvtColor(warped, cv2.COLOR_BGR2GRAY)
        g2 = cv2.GaussianBlur(g2, (5, 5), 0)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        g2 = clahe.apply(g2)

        th_marks = cv2.adaptiveThreshold(
            g2,
            255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY_INV,
            41,
            7,
        )

        tl = _detect_corner_square(th_marks, "TL")
        tr = _detect_corner_square(th_marks, "TR")
        br = _detect_corner_square(th_marks, "BR")
        bl = _detect_corner_square(th_marks, "BL")
        if any(p is None for p in (tl, tr, br, bl)):
            return {}, warped

        src_l = np.array([tl, tr, br, bl], dtype="float32")
        tpl_w, tpl_h = 1600, 2200
        dst_l = np.array(
            [[0, 0], [tpl_w - 1, 0], [tpl_w - 1, tpl_h - 1], [0, tpl_h - 1]],
            dtype="float32",
        )
        mat_l = cv2.getPerspectiveTransform(src_l, dst_l)
        aligned_gray = cv2.warpPerspective(g2, mat_l, (tpl_w, tpl_h))
        aligned = cv2.cvtColor(aligned_gray, cv2.COLOR_GRAY2BGR)
        th = cv2.adaptiveThreshold(
            aligned_gray,
            255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY_INV,
            35,
            7,
        )

        ranges = _sheet_column_ranges(total_preguntas)
        num_columnas = len(ranges)
        max_rows = max(
            1,
            max(
                ((r[1] - r[0] + 1) for r in ranges if r[0] > 0 and r[1] >= r[0]),
                default=1,
            ),
        )

        # Zona OMR en la hoja aplanada (proporciones ajustadas al diseño SEA móvil)
        zone_left = 0.07 * float(tpl_w)
        zone_right = 0.93 * float(tpl_w)
        zone_top = 0.23 * float(tpl_h)
        zone_bottom = 0.95 * float(tpl_h)

        zone_w = max(1.0, zone_right - zone_left)
        zone_h = max(1.0, zone_bottom - zone_top)
        inner_left = zone_left + (0.03 * zone_w)
        inner_right = zone_right - (0.03 * zone_w)
        usable_w = max(1.0, inner_right - inner_left)
        col_gap = (0.02 * zone_w) if num_columnas > 1 else 0.0
        col_width = (usable_w - ((num_columnas - 1) * col_gap)) / max(1, num_columnas)
        block_header_h = max(12.0, 0.075 * zone_h)
        rows_top = zone_top + block_header_h
        rows_bottom = zone_bottom - (0.03 * zone_h)
        row_gap = max(9.5, (rows_bottom - rows_top) / max(1, max_rows))
        number_w = max(26.0, 0.20 * col_width)
        bubble_track = max(20.0, col_width - number_w - (0.06 * col_width))
        bubble_gap = bubble_track / 3.0
        bubble_r = min(19.0, max(11.0, row_gap * 0.22))

        respuestas = {}
        letras = ["A", "B", "C", "D"]

        for col_idx, (r_start, r_end) in enumerate(ranges):
            if r_start <= 0 or r_end < r_start:
                continue
            x_col = inner_left + (col_idx * (col_width + col_gap))
            for row_idx in range(r_end - r_start + 1):
                pregunta = r_start + row_idx
                if pregunta > int(total_preguntas):
                    break

                cy = int(rows_top + ((row_idx + 0.5) * row_gap))
                start_x = x_col + number_w + (0.03 * col_width)
                scores = []

                for opt_idx, letra in enumerate(letras):
                    cx = int(start_x + (opt_idx * bubble_gap))
                    rr = int(max(8, bubble_r * 0.72))
                    x1 = max(0, cx - rr)
                    y1 = max(0, cy - rr)
                    x2 = min(tpl_w, cx + rr + 1)
                    y2 = min(tpl_h, cy + rr + 1)
                    roi = th[y1:y2, x1:x2]
                    if roi.size == 0:
                        scores.append((0.0, letra, cx, cy, rr))
                        continue

                    mask = np.zeros(roi.shape, dtype="uint8")
                    cv2.circle(mask, (cx - x1, cy - y1), rr, 255, -1)
                    area_px = max(1, cv2.countNonZero(mask))
                    ink_px = cv2.countNonZero(cv2.bitwise_and(roi, roi, mask=mask))
                    fill_ratio = float(ink_px) / float(area_px)

                    # Señal de circularidad local para reducir falsos positivos por sombras.
                    circ_score = 0.0
                    roi_cnts, _ = cv2.findContours(
                        roi, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
                    )
                    if roi_cnts:
                        best_c = max(roi_cnts, key=cv2.contourArea)
                        area_c = float(cv2.contourArea(best_c))
                        peri_c = float(cv2.arcLength(best_c, True))
                        if peri_c > 0.0:
                            circ = (4.0 * np.pi * area_c) / (peri_c * peri_c)
                            circ_score = max(0.0, min(1.0, circ))

                    robust_ratio = fill_ratio + (0.08 * circ_score)
                    scores.append((robust_ratio, letra, cx, cy, rr))

                scores.sort(key=lambda t: t[0], reverse=True)
                top = scores[0][0]
                second = scores[1][0] if len(scores) > 1 else 0.0

                if top >= 0.30 and (top - second) >= 0.07:
                    respuestas[pregunta] = scores[0][1]
                    cv2.circle(
                        aligned,
                        (scores[0][2], scores[0][3]),
                        scores[0][4],
                        (0, 255, 0),
                        2,
                    )
                else:
                    for _score, _letter, cx, cy, rr in scores:
                        cv2.circle(aligned, (cx, cy), rr, (0, 180, 255), 1)

        return respuestas, aligned

    def _guardar_calificacion_camara(
        self, qr_data, exam_code, respuestas_est, correctas, total, nota
    ):
        try:
            self.cur.execute(
                """INSERT INTO calificaciones_camara
                       (id_examen, documento, estudiante_nombre, grado, curso, area,
                        evaluacion, total_preguntas, correctas, nota, respuestas_json)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    exam_code,
                    qr_data.get("documento", ""),
                    qr_data.get("nombre", ""),
                    qr_data.get("grado", ""),
                    qr_data.get("curso", ""),
                    qr_data.get("area", ""),
                    qr_data.get("evaluacion", ""),
                    int(total),
                    int(correctas),
                    float(nota),
                    json.dumps(respuestas_est, ensure_ascii=False),
                ),
            )
            self.conn.commit()
            self._exportar_calificacion_camara_excel(
                qr_data=qr_data,
                exam_code=exam_code,
                respuestas_est=respuestas_est,
                correctas=correctas,
                total=total,
                nota=nota,
            )
            self._cargar_historial_camara_inline()
        except Exception:
            pass

    def _exportar_calificacion_camara_excel(
        self, qr_data, exam_code, respuestas_est, correctas, total, nota
    ):
        """Exporta/actualiza un consolidado Excel de lecturas por cámara."""
        try:
            from openpyxl import Workbook, load_workbook

            reportes_dir = Path(self.base_dir) / "reportes"
            reportes_dir.mkdir(parents=True, exist_ok=True)
            xlsx_path = reportes_dir / "calificaciones_camara.xlsx"

            if xlsx_path.exists():
                wb = load_workbook(str(xlsx_path))
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "calificaciones"
                ws.append(
                    [
                        "fecha",
                        "id_examen",
                        "documento",
                        "estudiante",
                        "grado",
                        "curso",
                        "area",
                        "evaluacion",
                        "correctas",
                        "total",
                        "nota",
                        "respuestas_json",
                    ]
                )

            ws.append(
                [
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    exam_code,
                    qr_data.get("documento", ""),
                    qr_data.get("nombre", ""),
                    qr_data.get("grado", ""),
                    qr_data.get("curso", ""),
                    qr_data.get("area", ""),
                    qr_data.get("evaluacion", ""),
                    int(correctas),
                    int(total),
                    float(nota),
                    json.dumps(respuestas_est, ensure_ascii=False, sort_keys=True),
                ]
            )

            wb.save(str(xlsx_path))
        except Exception:
            pass

    def recargar_datos_generador_examenes(self, event=None):
        """Recarga filtros y listas del generador usando datos actuales de la BD."""
        if not hasattr(self, "cb_examen_grado"):
            return

        grado_actual = str(self.cb_examen_grado.get() or "").strip()
        curso_actual = str(self.cb_examen_curso.get() or "").strip()
        area_actual = str(self.cb_examen_area.get() or "").strip()
        evaluacion_actual = str(self.cb_examen_evaluacion.get() or "").strip()
        estudiante_actual = str(self.cb_examen_estudiante.get() or "").strip()

        grados = [str(g) for g in cargar_grados_desde_preguntas()]
        self.cb_examen_grado["values"] = grados

        if not grados:
            self.cb_examen_grado.set("")
            self.cb_examen_curso["values"] = []
            self.cb_examen_curso.set("")
            self.cb_examen_area["values"] = []
            self.cb_examen_area.set("")
            self.cb_examen_evaluacion["values"] = []
            self.cb_examen_evaluacion.set("")
            self.cb_examen_estudiante["values"] = []
            self.cb_examen_estudiante.set("")
            self._update_examen_mode()
            return

        grado_final = grado_actual if grado_actual in grados else grados[0]
        self.cb_examen_grado.set(grado_final)

        cursos = cargar_cursos_por_grado(
            grado_final,
            estudiantes_path=self.estudiantes_path,
            db_path=self.db_path,
        )
        self.cb_examen_curso["values"] = cursos
        curso_final = (
            curso_actual if curso_actual in cursos else (cursos[0] if cursos else "")
        )
        self.cb_examen_curso.set(curso_final)

        areas = self._examen_areas_canonicas(cargar_areas_por_grado(grado_final))
        self.cb_examen_area["values"] = areas
        area_final = (
            area_actual if area_actual in areas else (areas[0] if areas else "")
        )
        self.cb_examen_area.set(area_final)

        evals = (
            cargar_evaluaciones_por_grado_y_area(grado_final, area_final)
            if area_final
            else []
        )
        self.cb_examen_evaluacion["values"] = evals
        eval_final = (
            evaluacion_actual
            if evaluacion_actual in evals
            else (evals[0] if evals else "")
        )
        self.cb_examen_evaluacion.set(eval_final)

        self._update_student_combo(grado_final, curso_final)
        estudiantes_vals = list(self.cb_examen_estudiante["values"] or [])
        if estudiante_actual and estudiante_actual in estudiantes_vals:
            self.cb_examen_estudiante.set(estudiante_actual)
        elif estudiantes_vals:
            self.cb_examen_estudiante.set(estudiantes_vals[0])
        else:
            self.cb_examen_estudiante.set("")

        self._update_examen_mode()

    def _refresh_examen_grados(self):
        """Compatibilidad: refresca el generador con datos actuales."""
        self.recargar_datos_generador_examenes()

    def _examen_mapa_areas_plan(self):
        """Mapa de áreas canónicas del plan: {nombre_lower: nombre_original}."""
        try:
            areas_plan = self._ca_todas_areas()
        except Exception:
            areas_plan = []

        mapa = {}
        for area in areas_plan or []:
            area_txt = str(area or "").strip()
            if area_txt:
                mapa.setdefault(area_txt.lower(), area_txt)
        return mapa

    def _examen_areas_canonicas(self, areas):
        """Normaliza lista de áreas para mostrar con el nombre del plan de estudio."""
        mapa_plan = self._examen_mapa_areas_plan()
        salida = []
        vistos = set()
        for area in areas or []:
            area_txt = str(area or "").strip()
            if not area_txt:
                continue
            area_canon = mapa_plan.get(area_txt.lower(), area_txt)
            clave = area_canon.lower()
            if clave in vistos:
                continue
            vistos.add(clave)
            salida.append(area_canon)
        return salida

    def _on_examen_grado_changed(self, event=None):
        grado = self.cb_examen_grado.get()
        # cargar cursos disponibles para ese grado (desde matrícula)
        cursos = cargar_cursos_por_grado(
            grado, estudiantes_path=self.estudiantes_path, db_path=self.db_path
        )
        self.cb_examen_curso["values"] = cursos
        self.cb_examen_curso.set(cursos[0] if cursos else "")

        # cargar áreas
        areas = self._examen_areas_canonicas(cargar_areas_por_grado(grado))
        self.cb_examen_area["values"] = areas
        self.cb_examen_area.set(areas[0] if areas else "")
        # actualizar lista de estudiantes (grado+curso)
        self._update_student_combo(grado, self.cb_examen_curso.get())
        # también actualizar evaluaciones si área ya existe
        self._on_examen_area_changed()

    def _on_examen_area_changed(self, event=None):
        grado = self.cb_examen_grado.get()
        area = self.cb_examen_area.get()
        evals = cargar_evaluaciones_por_grado_y_area(grado, area)
        self.cb_examen_evaluacion["values"] = evals
        self.cb_examen_evaluacion.set(evals[0] if evals else "")

    def _on_examen_curso_changed(self, event=None):
        # cuando cambia el curso, simplemente refrescar lista de estudiantes
        grado = self.cb_examen_grado.get()
        curso = self.cb_examen_curso.get()
        self._update_student_combo(grado, curso)

    def _update_student_combo(self, grado, curso=None):
        # grado y curso provienen de los combobox; se normalizan y se utiliza
        # cargar_estudiantes_por_grado para recuperar la lista de alumnos.
        vals = []
        # mapeo texto_combo -> dict completo del estudiante para recuperar
        # el documento real al generar el PDF sin depender del parsing.
        self._examen_estudiantes_map = {}
        if not grado:
            self.cb_examen_estudiante["values"] = []
            self.cb_examen_estudiante.set("")
            return
        estudiantes = cargar_estudiantes_por_grado(
            grado,
            curso=curso,
            estudiantes_path=self.estudiantes_path,
            db_path=self.db_path,
        )
        for e in estudiantes:
            nombre_estudiante = construir_nombre(e)
            doc = e.get("documento") or e.get("id") or ""
            label = f"{nombre_estudiante} ({doc})"
            # ante etiquetas duplicadas añadir sufijo para distinguirlas
            if label in self._examen_estudiantes_map:
                label = f"{label}_{len(vals)}"
            self._examen_estudiantes_map[label] = e
            vals.append(label)
        self.cb_examen_estudiante["values"] = vals
        if vals:
            self.cb_examen_estudiante.set(vals[0])
        else:
            self.cb_examen_estudiante.set("")

    def _update_examen_mode(self):
        modo = self.var_examen_tipo.get()
        state = "normal" if modo == "individual" else "disabled"
        self.cb_examen_estudiante.config(state=state)

    def _generar_etiquetas_version(self, cantidad):
        """Genera etiquetas de versión en secuencia alfabética: A, B, ..., Z, AA..."""
        return core_examenes_generacion.generar_etiquetas_version(cantidad)

    def _examen_listar_docentes_selector(self):
        """Retorna lista de docentes para el selector temporal de generación PDF."""
        try:
            data = core_docentes.listar_docentes(limit=10000, offset=0)
            rows = data.get("docentes", [])
            return [
                (str(r.get("documento", "")).strip(), str(r.get("nombre", "")).strip())
                for r in rows
                if str(r.get("nombre", "")).strip()
            ]
        except Exception:
            return []

    def _dialog_parametros_generar_examen(
        self,
        max_qty,
        max_textos,
        modo_generacion="individual",
    ):
        """Diálogo único para parámetros de generación del examen."""
        docentes = self._examen_listar_docentes_selector()
        if not docentes:
            messagebox.showerror(
                "Generar Exámenes",
                "No hay docentes disponibles para seleccionar.",
                parent=self.win,
            )
            return None

        d = tk.Toplevel(self.win)
        d.title("Parámetros del examen")
        d.transient(self.win)
        d.resizable(False, False)
        d.grab_set()

        frm = ttk.Frame(d, padding=12)
        frm.pack(fill="both", expand=True)
        frm.grid_columnconfigure(1, weight=1)

        txt_max_textos = max(0, int(max_textos or 0))
        txt_default = 0
        if txt_max_textos > 0:
            txt_default = txt_max_textos

        qty_textos_var = tk.IntVar(value=txt_default)
        qty_preg_var = tk.IntVar(value=max(1, int(max_qty or 1)))
        qty_versiones_var = tk.IntVar(value=1)
        version_manual_var = tk.StringVar(value="A")
        modo_hoja_var = tk.StringVar(value=ANSWER_SHEET_MODE_NONE)
        formato_examen_var = tk.StringVar(value="Estándar")
        fecha_var = tk.StringVar(value="")
        docente_var = tk.StringVar(value=docentes[0][1])

        formato_examen_map = {
            "Estándar": EXAM_FORMAT_STANDARD,
            "Matemáticas (Tipo ICFES)": EXAM_FORMAT_MATH_ICFES,
            "Lenguaje (Lectura + Abierta)": EXAM_FORMAT_LANGUAGE_ICFES,
        }

        ttk.Label(frm, text="Cantidad de textos:").grid(
            row=0, column=0, sticky="w", padx=(0, 8), pady=4
        )
        sp_textos = ttk.Spinbox(
            frm,
            from_=0,
            to=txt_max_textos,
            textvariable=qty_textos_var,
            width=12,
        )
        sp_textos.grid(row=0, column=1, sticky="w", pady=4)
        if txt_max_textos <= 0:
            sp_textos.state(["disabled"])

        ttk.Label(frm, text="Cantidad de preguntas:").grid(
            row=1, column=0, sticky="w", padx=(0, 8), pady=4
        )
        sp_preg = ttk.Spinbox(
            frm,
            from_=1,
            to=max(1, int(max_qty)),
            textvariable=qty_preg_var,
            width=12,
        )
        sp_preg.grid(row=1, column=1, sticky="w", pady=4)

        ttk.Label(frm, text="Cantidad de versiones:").grid(
            row=2, column=0, sticky="w", padx=(0, 8), pady=4
        )
        sp_versiones = ttk.Spinbox(
            frm,
            from_=1,
            to=999,
            textvariable=qty_versiones_var,
            width=12,
        )
        sp_versiones.grid(row=2, column=1, sticky="w", pady=4)

        ttk.Label(frm, text="Versión (individual):").grid(
            row=3, column=0, sticky="w", padx=(0, 8), pady=4
        )
        cb_version_manual = ttk.Combobox(
            frm,
            state="readonly",
            values=["A"],
            textvariable=version_manual_var,
            width=12,
        )
        cb_version_manual.grid(row=3, column=1, sticky="w", pady=4)

        def _sync_version_options(*_args):
            try:
                n_ver = int(qty_versiones_var.get())
            except Exception:
                n_ver = 1
            n_ver = max(1, n_ver)
            etiquetas = self._generar_etiquetas_version(n_ver)
            cb_version_manual["values"] = etiquetas
            sel_actual = str(version_manual_var.get() or "").strip().upper()
            if sel_actual not in etiquetas:
                version_manual_var.set(etiquetas[0])

            if str(modo_generacion) == "individual":
                cb_version_manual.state(["!disabled", "readonly"])
            else:
                cb_version_manual.state(["disabled"])

        qty_versiones_var.trace_add("write", _sync_version_options)
        _sync_version_options()

        ttk.Label(frm, text="Formato del examen:").grid(
            row=4, column=0, sticky="w", padx=(0, 8), pady=4
        )
        cb_formato = ttk.Combobox(
            frm,
            state="readonly",
            values=list(formato_examen_map.keys()),
            textvariable=formato_examen_var,
            width=34,
        )
        cb_formato.grid(row=4, column=1, sticky="ew", pady=4)

        hoja_frame = ttk.LabelFrame(frm, text="Generación de hoja de respuestas")
        hoja_frame.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(4, 8))
        hoja_frame.grid_columnconfigure(0, weight=1)

        ttk.Radiobutton(
            hoja_frame,
            text="No generar hoja de respuestas",
            variable=modo_hoja_var,
            value=ANSWER_SHEET_MODE_NONE,
        ).grid(row=0, column=0, sticky="w", padx=8, pady=(6, 2))

        ttk.Radiobutton(
            hoja_frame,
            text="Generar hoja de respuestas en página adicional",
            variable=modo_hoja_var,
            value=ANSWER_SHEET_MODE_APPEND,
        ).grid(row=1, column=0, sticky="w", padx=8, pady=2)

        ttk.Radiobutton(
            hoja_frame,
            text="Generar hoja de respuestas en la última página del examen",
            variable=modo_hoja_var,
            value=ANSWER_SHEET_MODE_INLINE,
        ).grid(row=2, column=0, sticky="w", padx=8, pady=(2, 6))

        ttk.Label(frm, text="Fecha del examen:").grid(
            row=6, column=0, sticky="w", padx=(0, 8), pady=4
        )
        ttk.Entry(frm, textvariable=fecha_var, width=28).grid(
            row=6, column=1, sticky="ew", pady=4
        )

        ttk.Label(frm, text="Docente:").grid(
            row=7, column=0, sticky="w", padx=(0, 8), pady=4
        )
        cb_docente = ttk.Combobox(
            frm,
            state="readonly",
            values=[nombre for _, nombre in docentes],
            textvariable=docente_var,
            width=34,
        )
        cb_docente.grid(row=7, column=1, sticky="ew", pady=4)

        result = {"ok": False}

        def _aceptar():
            try:
                cantidad_preguntas = int(qty_preg_var.get())
            except Exception:
                messagebox.showerror(
                    "Generar Exámenes",
                    "La cantidad de preguntas debe ser un número válido.",
                    parent=d,
                )
                return

            if cantidad_preguntas < 1 or cantidad_preguntas > int(max_qty):
                messagebox.showerror(
                    "Generar Exámenes",
                    f"La cantidad de preguntas debe estar entre 1 y {int(max_qty)}.",
                    parent=d,
                )
                return

            try:
                cantidad_textos = int(qty_textos_var.get())
            except Exception:
                messagebox.showerror(
                    "Generar Exámenes",
                    "La cantidad de textos debe ser un número válido.",
                    parent=d,
                )
                return

            if cantidad_textos < 0 or cantidad_textos > int(txt_max_textos):
                messagebox.showerror(
                    "Generar Exámenes",
                    f"La cantidad de textos debe estar entre 0 y {int(txt_max_textos)}.",
                    parent=d,
                )
                return

            try:
                cantidad_versiones = int(qty_versiones_var.get())
            except Exception:
                messagebox.showerror(
                    "Generar Exámenes",
                    "La cantidad de versiones debe ser un número válido.",
                    parent=d,
                )
                return

            if cantidad_versiones < 1:
                messagebox.showerror(
                    "Generar Exámenes",
                    "La cantidad de versiones no puede ser menor que 1.",
                    parent=d,
                )
                return

            etiquetas_version = self._generar_etiquetas_version(cantidad_versiones)
            version_manual = None
            if str(modo_generacion) == "individual":
                version_manual = str(version_manual_var.get() or "").strip().upper()
                if version_manual not in etiquetas_version:
                    version_manual = etiquetas_version[0]

            docente_nombre = str(docente_var.get() or "").strip()
            if not docente_nombre:
                messagebox.showerror(
                    "Generar Exámenes",
                    "Debe seleccionar un docente antes de generar el PDF.",
                    parent=d,
                )
                return

            docente_id = ""
            for doc_id, nom in docentes:
                if nom == docente_nombre:
                    docente_id = doc_id
                    break

            modo_hoja = normalizar_modo_hoja_respuestas(modo_hoja_var.get())
            formato_examen = normalizar_formato_examen(
                formato_examen_map.get(
                    str(formato_examen_var.get() or "").strip(),
                    EXAM_FORMAT_STANDARD,
                )
            )

            result.update(
                {
                    "ok": True,
                    "cantidad_textos": cantidad_textos,
                    "cantidad_preguntas": cantidad_preguntas,
                    "cantidad_versiones": cantidad_versiones,
                    "version_individual": version_manual,
                    "formato_examen": formato_examen,
                    "modo_hoja_respuestas": modo_hoja,
                    "generar_hoja_respuestas": modo_hoja != ANSWER_SHEET_MODE_NONE,
                    "fecha": str(fecha_var.get() or "").strip(),
                    "docente_id": docente_id,
                    "docente_nombre": docente_nombre,
                }
            )
            d.destroy()

        def _cancelar():
            d.destroy()

        btns = ttk.Frame(frm)
        btns.grid(row=8, column=0, columnspan=2, sticky="e", pady=(10, 0))
        ttk.Button(btns, text="Cancelar", command=_cancelar).pack(side="right")
        ttk.Button(btns, text="Generar", command=_aceptar).pack(
            side="right", padx=(0, 8)
        )

        d.bind("<Return>", lambda _e: _aceptar())
        d.bind("<Escape>", lambda _e: _cancelar())
        cb_docente.focus_set()

        d.wait_window()
        return result if result.get("ok") else None

    def examen_generar(self):
        """Callback del botón para iniciar generación.

        Antes de abrir el diálogo de destino valida que existan alumnos para el
        grado seleccionado y muestra mensajes claros en caso contrario.
        """
        if not self._requiere_permiso("desktop.superadmin.evaluaciones.generar_pdf"):
            return
        grado = self.cb_examen_grado.get()
        curso = self.cb_examen_curso.get()
        area = self.cb_examen_area.get()
        evaluacion = self.cb_examen_evaluacion.get() or None

        if not grado or not curso or not area:
            message = "Debe seleccionar grado, curso y área."
            messagebox.showerror("Error", message, parent=self.win)
            return

        # validar existencia de estudiantes (en todos o en individual cuando ya
        # se haya elegido uno)
        alumnos = cargar_estudiantes_por_grado(
            grado,
            curso=curso,
            estudiantes_path=self.estudiantes_path,
            db_path=self.db_path,
        )
        modo_generacion = self.var_examen_tipo.get()

        # en modos masivos, asegurar alumnos para grado/curso
        if modo_generacion != "individual":
            if not alumnos:
                messagebox.showerror(
                    "Error",
                    "No hay estudiantes registrados para el grado seleccionado.",
                    parent=self.win,
                )
                return

        # validar dependencia de ReportLab
        if not _HAS_REPORTLAB:
            messagebox.showerror(
                "Error",
                "No es posible generar exámenes porque ReportLab no está instalado.",
                parent=self.win,
            )
            return

        # ver que existan preguntas para el filtro seleccionado
        preguntas_df = cargar_preguntas_filtradas(grado, area, evaluacion)
        if preguntas_df.empty:
            messagebox.showerror(
                "Error",
                "No hay preguntas disponibles para los filtros seleccionados.",
                parent=self.win,
            )
            return

        estudiante = None
        dest_file = None

        if modo_generacion == "individual":
            sel = self.cb_examen_estudiante.get()
            if not sel:
                messagebox.showerror(
                    "Error",
                    "Debe seleccionar un estudiante cuando el modo es Individual.",
                    parent=self.win,
                )
                return
            # recuperar dict completo del estudiante desde el mapeo guardado
            # al poblar el combo; esto garantiza que documento, curso, etc.
            # sean exactamente los valores almacenados en la BD.
            mapa = getattr(self, "_examen_estudiantes_map", {})
            est_bd = mapa.get(sel)
            if est_bd:
                estudiante = dict(est_bd)  # copia para no mutar el mapa
            else:
                # fallback: parsear el string del combo
                nombre_raw, doc_raw = sel.rsplit(" (", 1)
                doc_val = doc_raw.strip(")")
                estudiante = {
                    "id": doc_val,
                    "documento": doc_val,
                    "nombre1": nombre_raw.strip(),
                    "nombre2": "",
                    "apellido1": "",
                    "apellido2": "",
                    "grado": grado,
                    "curso": self.cb_examen_curso.get(),
                }
            # Asegurar todos los campos requeridos para PDF
            estudiante.setdefault("apellido1", "")
            estudiante.setdefault("apellido2", "")
            estudiante.setdefault("nombre1", "")
            estudiante.setdefault("nombre2", "")
            estudiante.setdefault("documento", estudiante.get("id", ""))
            estudiante.setdefault("grado", grado)
            estudiante.setdefault("curso", self.cb_examen_curso.get())

            # antes de solicitar destino, pedir cantidad de textos, cantidad de preguntas y fecha
            # consultar cuántas preguntas hay disponibles para los filtros
            df_preg = cargar_preguntas_filtradas(grado, area, evaluacion)
            if df_preg.empty:
                messagebox.showerror(
                    "Generar Exámenes",
                    "No hay preguntas disponibles para los filtros seleccionados.",
                    parent=self.win,
                )
                return
            max_qty = len(df_preg)
            # calcular cantidad de textos disponibles (contextos únicos)
            if "id_contexto" in df_preg.columns:
                unique_ctx = [
                    c
                    for c in df_preg["id_contexto"]
                    if str(c).strip() and str(c) != "nan"
                ]
                # preservar orden sin duplicados
                seen = set()
                ctxts = []
                for c in unique_ctx:
                    if c not in seen:
                        seen.add(c)
                        ctxts.append(c)
                max_textos = len(ctxts)
            else:
                max_textos = 0

            params = self._dialog_parametros_generar_examen(
                max_qty=max_qty,
                max_textos=max_textos,
                modo_generacion=modo_generacion,
            )
            if not params:
                return

            cantidad_textos = params.get("cantidad_textos", 0)
            cantidad_manual = params.get("cantidad_preguntas")
            cantidad_versiones = params.get("cantidad_versiones", 1)
            version_individual = params.get("version_individual")
            formato_examen = normalizar_formato_examen(params.get("formato_examen"))
            modo_hoja_respuestas = normalizar_modo_hoja_respuestas(
                params.get("modo_hoja_respuestas"),
                params.get("generar_hoja_respuestas", False),
            )
            fecha_val = params.get("fecha")
            docente_nombre = params.get("docente_nombre")

            # solicitar nombre y ubicación de archivo
            curso = estudiante.get("curso") or self.cb_examen_curso.get()
            version_suffix = (
                f"_version_{version_individual}"
                if int(cantidad_versiones or 1) > 1 and version_individual
                else ""
            )
            nombre_estudiante = f"{estudiante.get('nombre1','')} {estudiante.get('nombre2','')} {estudiante.get('apellido1','')} {estudiante.get('apellido2','')}".strip()
            if not nombre_estudiante:
                nombre_estudiante = "estudiante"
            sugerido = (
                f"examen_{grado}_{curso}_{area}_{evaluacion or 'general'}_"
                f"{nombre_estudiante}{version_suffix}"
            )
            sugerido = sugerido.replace(" ", "_")
            dest_file = filedialog.asksaveasfilename(
                title="Guardar examen PDF",
                defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
                initialfile=sugerido,
                parent=self.win,
            )
            if not dest_file:
                return

            try:
                # _do_generate_exams soporta un destino específico cuando se pasa
                # ``dest_filename`` y parámetros opcionales.
                self._do_generate_exams(
                    os.path.dirname(dest_file),
                    grado,
                    area,
                    evaluacion,
                    self.cb_examen_curso.get(),
                    estudiante,
                    dest_filename=dest_file,
                    cantidad_manual=cantidad_manual,
                    cantidad_textos=cantidad_textos,
                    formato_examen=formato_examen,
                    fecha=fecha_val,
                    docente_nombre=docente_nombre,
                    modo_hoja_respuestas=modo_hoja_respuestas,
                    cantidad_versiones=cantidad_versiones,
                    version_individual=version_individual,
                )
                messagebox.showinfo(
                    "Generar Exámenes",
                    f"Archivo generado:\n{dest_file}",
                    parent=self.win,
                )
            except Exception as e:
                messagebox.showerror(
                    "Generar Exámenes", f"Error al generar PDF: {e}", parent=self.win
                )
            return

        # modos masivos: "todos" (carpeta) y "todos_un_pdf" (archivo único)
        dest_file = None
        if modo_generacion == "todos":
            path = filedialog.askdirectory(
                title="Seleccionar carpeta destino", parent=self.win
            )
            if not path:
                return
        elif modo_generacion == "todos_un_pdf":
            sugerido = (
                f"Examenes_{grado}_{curso}_{area}_{evaluacion or 'general'}_TODOS.pdf"
            )
            sugerido = sugerido.replace(" ", "_")
            dest_file = filedialog.asksaveasfilename(
                title="Guardar examen único (todos)",
                defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
                initialfile=sugerido,
                parent=self.win,
            )
            if not dest_file:
                return
            path = os.path.dirname(dest_file)
        else:
            messagebox.showerror(
                "Error",
                "Modo de generación no válido.",
                parent=self.win,
            )
            return

        # para modo masivo también pedir cantidad de textos, preguntas, fecha y docente
        df_preg = cargar_preguntas_filtradas(grado, area, evaluacion)
        max_qty = len(df_preg)
        cantidad_manual = None
        cantidad_textos = 0
        cantidad_versiones = 1
        formato_examen = EXAM_FORMAT_STANDARD
        modo_hoja_respuestas = ANSWER_SHEET_MODE_NONE
        fecha_val = None
        docente_nombre = None
        if max_qty > 0:
            # calcular textos disponibles primero
            if "id_contexto" in df_preg.columns:
                ctxs = []
                seen = set()
                for c in df_preg["id_contexto"]:
                    if str(c).strip() and str(c) != "nan" and c not in seen:
                        seen.add(c)
                        ctxs.append(c)
                max_textos = len(ctxs)
            else:
                max_textos = 0
            params = self._dialog_parametros_generar_examen(
                max_qty=max_qty,
                max_textos=max_textos,
                modo_generacion=modo_generacion,
            )
            if not params:
                return
            cantidad_textos = params.get("cantidad_textos", 0)
            cantidad_manual = params.get("cantidad_preguntas")
            cantidad_versiones = params.get("cantidad_versiones", 1)
            formato_examen = normalizar_formato_examen(params.get("formato_examen"))
            modo_hoja_respuestas = normalizar_modo_hoja_respuestas(
                params.get("modo_hoja_respuestas"),
                params.get("generar_hoja_respuestas", False),
            )
            fecha_val = params.get("fecha")
            docente_nombre = params.get("docente_nombre")
        try:
            if modo_generacion == "todos_un_pdf":
                cantidad = self._do_generate_exams(
                    path,
                    grado,
                    area,
                    evaluacion,
                    self.cb_examen_curso.get(),
                    estudiante=None,
                    dest_filename=dest_file,
                    cantidad_manual=cantidad_manual,
                    cantidad_textos=cantidad_textos,
                    formato_examen=formato_examen,
                    fecha=fecha_val,
                    docente_nombre=docente_nombre,
                    modo_hoja_respuestas=modo_hoja_respuestas,
                    un_solo_pdf_todos=True,
                    cantidad_versiones=cantidad_versiones,
                )
                messagebox.showinfo(
                    "Generar Exámenes",
                    f"Archivo generado para {cantidad} estudiante(s):\n{dest_file}",
                    parent=self.win,
                )
                # Preguntar si desea generar el reporte de respuestas correctas y organización por versiones
                if messagebox.askyesno(
                    "Reporte de respuestas",
                    "✔ Exámenes generados correctamente\n¿Desea generar el reporte de respuestas correctas y organización por versiones?",
                    parent=self.win,
                ):
                    self._generar_reporte_respuestas(
                        grado=grado,
                        area=area,
                        evaluacion=evaluacion,
                        curso=self.cb_examen_curso.get(),
                        docente_nombre=docente_nombre,
                        fecha=fecha_val,
                        cantidad_versiones=cantidad_versiones,
                        out_dir=path,
                    )

            else:
                cantidad = self._do_generate_exams(
                    path,
                    grado,
                    area,
                    evaluacion,
                    self.cb_examen_curso.get(),
                    estudiante=None,
                    cantidad_manual=cantidad_manual,
                    cantidad_textos=cantidad_textos,
                    formato_examen=formato_examen,
                    fecha=fecha_val,
                    docente_nombre=docente_nombre,
                    modo_hoja_respuestas=modo_hoja_respuestas,
                    cantidad_versiones=cantidad_versiones,
                )
        except Exception as e:
            messagebox.showerror(
                "Generar Exámenes", f"Error al generar PDF: {e}", parent=self.win
            )

    def _generar_reporte_respuestas(
        self,
        grado,
        area,
        evaluacion,
        curso,
        docente_nombre,
        fecha,
        cantidad_versiones,
        out_dir,
    ):
        """Genera el PDF de reporte de respuestas correctas y organización por versiones."""
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from core import examenes_generacion as core_examenes_generacion
        import datetime

        respuestas_por_version = (
            core_examenes_generacion.obtener_respuestas_correctas_por_version(
                grado, area, evaluacion
            )
        )
        relacion_estudiantes = (
            core_examenes_generacion.obtener_relacion_estudiantes_version(
                grado, area, evaluacion
            )
        )
        if not respuestas_por_version:
            messagebox.showerror(
                "Reporte de respuestas",
                "No se encontraron respuestas correctas para el filtro indicado.",
                parent=self.win,
            )
            return
        if not relacion_estudiantes:
            messagebox.showerror(
                "Reporte de respuestas",
                "No se encontró relación de estudiantes para el filtro indicado.",
                parent=self.win,
            )
            return

        estudiantes_unicos = {}
        for est in relacion_estudiantes:
            documento = str(est.get("documento", "") or "").strip()
            nombre = str(construir_nombre(est) or "").strip()
            version = str(est.get("version", "") or "").strip().upper()
            if not version:
                version = "SIN VERSION"
            if not nombre:
                continue

            clave = documento or f"{version}|{nombre}"
            estudiantes_unicos[clave] = {
                "documento": documento,
                "nombre": nombre,
                "version": version,
            }

        agrupados_por_version = {}
        for est in estudiantes_unicos.values():
            agrupados_por_version.setdefault(est["version"], []).append(est)

        versiones_ordenadas = sorted(agrupados_por_version.keys())
        total_estudiantes = len(estudiantes_unicos)
        total_versiones_reales = len(versiones_ordenadas)

        def _rango_versiones(versiones):
            versiones_limpias = [
                str(v or "").strip().upper() for v in versiones if str(v or "").strip()
            ]
            if not versiones_limpias:
                return "-"
            if len(versiones_limpias) == 1:
                return versiones_limpias[0]
            return f"{versiones_limpias[0]}-{versiones_limpias[-1]}"

        nombre_docente = docente_nombre or "-"
        area_txt = str(area or "-")
        grado_txt = str(grado or "-")
        curso_txt = str(curso or "-")
        fecha_txt = (
            fecha.strftime("%d/%m/%Y")
            if hasattr(fecha, "strftime")
            else str(fecha or datetime.date.today().strftime("%d/%m/%Y"))
        )
        nombre_eval = str(evaluacion or "-")

        nombre_archivo = f"reporte_respuestas_Grado{grado_txt}_{area_txt}.pdf".replace(
            " ", "_"
        )
        ruta_pdf = os.path.join(out_dir, nombre_archivo)

        c = canvas.Canvas(ruta_pdf, pagesize=A4)
        w, h = A4
        margen = 40
        y = h - margen
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredString(
            w / 2,
            y,
            "REPORTE DE RESPUESTAS CORRECTAS Y ORGANIZACIÓN POR VERSIONES",
        )
        y -= 32
        c.setFont("Helvetica", 11)
        c.drawString(margen, y, f"Docente: {nombre_docente}")
        y -= 18
        c.drawString(margen, y, f"Área: {area_txt}")
        y -= 18
        c.drawString(margen, y, f"Grado: {grado_txt}")
        y -= 18
        c.drawString(margen, y, f"Curso: {curso_txt}")
        y -= 18
        c.drawString(margen, y, f"Fecha: {fecha_txt}")
        y -= 18
        c.drawString(margen, y, f"Evaluación: {nombre_eval}")
        y -= 18
        c.drawString(margen, y, f"Total de estudiantes: {total_estudiantes}")
        if total_estudiantes > 1:
            y -= 18
            c.drawString(
                margen,
                y,
                f"Cantidad de versiones: {total_versiones_reales}",
            )
            y -= 18
            c.drawString(
                margen,
                y,
                f"Tipo de generación: Versiones múltiples ({_rango_versiones(versiones_ordenadas)})",
            )
        else:
            y -= 18
            c.drawString(
                margen,
                y,
                f"Versión individual: {_rango_versiones(versiones_ordenadas)}",
            )
        y -= 28

        c.setFont("Helvetica-Bold", 12)
        c.drawString(margen, y, "Instrucciones para el docente:")
        y -= 18
        c.setFont("Helvetica", 10)
        instrucciones = [
            "- Este documento permite identificar rápidamente las respuestas correctas por versión.",
            "- Utilice este reporte para calificación manual o verificación.",
            "- Agrupe los exámenes según versión (A, B, C, ...).",
            "- Verifique que cada estudiante tenga la versión correcta.",
        ]
        for instr in instrucciones:
            c.drawString(margen + 12, y, instr)
            y -= 14
        y -= 10

        c.setFont("Helvetica-Bold", 12)
        c.drawString(margen, y, "Respuestas correctas por versión:")
        y -= 18
        c.setFont("Helvetica", 11)
        for version in sorted(respuestas_por_version.keys()):
            c.setFont("Helvetica-Bold", 11)
            c.drawString(margen + 8, y, f"Versión {version}")
            y -= 16
            c.setFont("Helvetica", 11)
            respuestas = respuestas_por_version[version]
            linea = ""
            for idx, resp in enumerate(respuestas, 1):
                linea += f"{idx}. {resp}   "
                if idx % 10 == 0:
                    c.drawString(margen + 24, y, linea.strip())
                    y -= 14
                    linea = ""
            if linea:
                c.drawString(margen + 24, y, linea.strip())
                y -= 14
            y -= 6
            if y < 80:
                c.showPage()
                y = h - margen
        y -= 10

        c.setFont("Helvetica-Bold", 12)
        c.drawString(margen, y, "Relación de estudiantes por versión:")
        y -= 18
        espacio_por_fila = 14
        margen_interno = margen + 12
        for version in agrupados_por_version:
            agrupados_por_version[version].sort(key=lambda item: item["nombre"])

        def _nueva_pagina_relacion():
            nonlocal y
            c.showPage()
            y = h - margen

        def _asegurar_espacio(lineas_necesarias):
            nonlocal y
            if y - (lineas_necesarias * espacio_por_fila) < 60:
                _nueva_pagina_relacion()

        _asegurar_espacio(2)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(margen_interno, y, "Resumen por versión:")
        y -= 18
        c.setFont("Helvetica", 10)
        for version in sorted(agrupados_por_version.keys()):
            _asegurar_espacio(1)
            total_estudiantes = len(agrupados_por_version[version])
            c.drawString(
                margen_interno + 8, y, f"{version}: {total_estudiantes} estudiantes"
            )
            y -= espacio_por_fila

        y -= 8

        for version in sorted(agrupados_por_version.keys()):
            estudiantes_version = agrupados_por_version[version]
            _asegurar_espacio(3)
            c.setFont("Helvetica-Bold", 11)
            c.drawString(margen_interno, y, f"VERSIÓN {version}")
            y -= 18
            c.setFont("Helvetica", 10)

            for est in estudiantes_version:
                _asegurar_espacio(1)
                c.drawString(margen_interno + 8, y, f"• {est['nombre']}")
                y -= espacio_por_fila

            y -= 10

        c.save()
        messagebox.showinfo(
            "Reporte de respuestas",
            f"Reporte generado correctamente:\n{ruta_pdf}",
            parent=self.win,
        )

    def _merge_pdf_files(self, pdf_paths, output_path):
        """Une varios PDFs en un único archivo de salida y muestra resumen de paridad."""
        rutas = [p for p in (pdf_paths or []) if p and os.path.exists(p)]
        resumen = core_examenes_generacion.unir_pdfs(rutas, output_path)
        if resumen:
            estudiantes = resumen.get("estudiantes", 0)
            total_paginas = resumen.get("total_paginas", 0)
            hojas_blanco = resumen.get("hojas_blanco", 0)
            if hojas_blanco > 0:
                mensaje = (
                    f"Total de estudiantes procesados: {estudiantes}\n"
                    f"Total de páginas generadas: {total_paginas}\n"
                    f"Cantidad de hojas en blanco insertadas: {hojas_blanco}\n\n"
                    "⚠️ Se detectaron exámenes con número impar de páginas. "
                    "Se insertaron hojas en blanco automáticamente para garantizar impresión doble cara correcta."
                )
                messagebox.showwarning(
                    "Control de paridad de páginas", mensaje, parent=self.win
                )
            else:
                mensaje = (
                    f"Total de estudiantes procesados: {estudiantes}\n"
                    f"Total de páginas generadas: {total_paginas}\n"
                    "No fue necesario insertar hojas en blanco."
                )
                messagebox.showinfo(
                    "Control de paridad de páginas", mensaje, parent=self.win
                )

    def _fetch_config_examen(self, grado, area, evaluacion=None):
        """Devuelve la tupla (duracion_segundos, cantidad_preguntas) o None."""
        cfg = core_examenes.cargar_config_examen(
            area=area,
            grado=grado,
            evaluacion=evaluacion,
        )
        if not cfg:
            return None
        return (cfg[0], cfg[1])

    def _do_generate_exams(
        self,
        out_dir,
        grado,
        area,
        evaluacion=None,
        curso=None,
        estudiante=None,
        dest_filename=None,
        cantidad_manual=None,
        cantidad_textos=0,
        formato_examen=EXAM_FORMAT_STANDARD,
        fecha=None,
        docente_nombre=None,
        modo_hoja_respuestas=None,
        un_solo_pdf_todos=False,
        generar_hoja_respuestas=False,
        cantidad_versiones=1,
        version_individual=None,
    ):
        """Genera pdfs para los filtros indicados.

        Las preguntas se filtran SOLO por grado+área+evaluación. El parámetro
        ``grado`` determina el banco de preguntas que se utilizará, sin importar
        el valor de ``curso``.

        El parámetro ``curso`` opcional solo se utiliza para:
        - Filtrar la LISTA DE ESTUDIANTES (grado + curso) cuando se genera
          para todos los estudiantes del grado (``estudiante=None``).
        - En el nombre del archivo generado.

        Es decir: el mismo examen (preguntario) se aplica a todos los cursos
        del grado, pero solo se generan archivos para los estudiantes del
        curso seleccionado.

        Si ``estudiante`` es ``None`` se genera un archivo por cada estudiante
        del grado (y curso si se especificó). De lo contrario se genera
        únicamente para el diccionario recibido.

        El parámetro ``cantidad_textos`` permite limitar cuántos textos
        (grupos de contexto) se incluyen en el examen; si se especifica un
        valor mayor al disponible se produce un error."""
        modo_hoja_respuestas = normalizar_modo_hoja_respuestas(
            modo_hoja_respuestas,
            generar_hoja_respuestas,
        )
        formato_examen = normalizar_formato_examen(formato_examen)
        try:
            cantidad_versiones = int(cantidad_versiones)
        except Exception:
            cantidad_versiones = 1
        cantidad_versiones = max(1, cantidad_versiones)
        etiquetas_version = self._generar_etiquetas_version(cantidad_versiones)
        if "matem" in str(area or "").strip().lower():
            formato_examen = EXAM_FORMAT_MATH_ICFES
        elif _es_area_lenguaje(area):
            if formato_examen in {
                EXAM_FORMAT_STANDARD,
                EXAM_FORMAT_LANGUAGE_ICFES,
            }:
                formato_examen = EXAM_FORMAT_LANGUAGE_ICFES
        elif formato_examen == EXAM_FORMAT_LANGUAGE_ICFES:
            formato_examen = EXAM_FORMAT_STANDARD

        # cargar preguntas con los filtros solicitados
        df = cargar_preguntas_filtradas(grado, area, evaluacion)
        if df.empty:
            raise ValueError("No hay preguntas disponibles para los filtros dados.")

        # determinar cantidad según configuración de examen
        cfg = self._fetch_config_examen(grado, area, evaluacion)
        cantidad = None
        if cfg:
            cantidad = cfg[1]
        # si el usuario proporcionó un valor manual válido usarlo
        if cantidad_manual is not None:
            try:
                manual = int(cantidad_manual)
                if manual > 0 and manual <= len(df):
                    cantidad = manual
            except Exception:
                pass
        if cantidad is None or cantidad <= 0 or cantidad > len(df):
            cantidad = len(df)
        # Si se solicitó `cantidad_textos` la selección y distribución se
        # realizarán por alumno más abajo. No truncamos el DataFrame aquí
        # porque necesitamos conocer la disponibilidad por texto.
        if cantidad < len(df) and not (cantidad_textos and cantidad_textos > 0):
            df = df.head(cantidad)

        # Si se definen múltiples versiones, cada versión mantiene exactamente
        # la misma selección y orden para todos los estudiantes asignados.
        examenes_por_version = {}
        if cantidad_versiones > 1:
            for ver in etiquetas_version:
                seed_ver = f"{grado}|{area}|{evaluacion or ''}|{ver}"
                examenes_por_version[ver] = (
                    core_examenes_generacion.seleccionar_preguntas_por_textos(
                        df,
                        cantidad,
                        cantidad_textos,
                        rnd_seed=seed_ver,
                    )
                )

        # resolver curso a usar: primero preferir el parámetro explícito,
        # luego el campo del estudiante individual si se proporcionó uno.
        curso_final = curso
        if estudiante and curso_final is None:
            curso_final = estudiante.get("curso")
        if estudiante is None:
            # todos los estudiantes del grado (filtrados por curso si se especificó)
            if self.estudiantes_df is None:
                self._load_estudiantes()
            alumnos = []
            if _HAS_PANDAS:
                # la columna grado puede ser numérica tras leer Excel, normalizamos
                df_filtered = self.estudiantes_df[
                    self.estudiantes_df["grado"].astype(str) == str(grado)
                ]
                # filtrar también por curso si se proporcionó
                if curso_final:
                    df_filtered = df_filtered[
                        df_filtered["curso"].astype(str) == str(curso_final)
                    ]
                alumnos = df_filtered.to_dict(orient="records")
            else:
                for r in self.estudiantes_df.get("_rows", []):
                    if str(r.get("grado", "")) == str(grado):
                        # filtrar también por curso si se proporcionó
                        if curso_final and str(r.get("curso", "")) != str(curso_final):
                            continue
                        alumnos.append(r.copy())
            if not alumnos:
                if curso_final:
                    raise ValueError(
                        f"No hay estudiantes registrados para grado {grado} y curso {curso_final}."
                    )
                else:
                    raise ValueError("No hay estudiantes registrados para el grado.")

            archivos_creados = 0
            pdfs_generados = []
            for idx, alum in enumerate(alumnos, start=1):
                # intentar usar un identificador único para el archivo
                ident = alum.get("documento") or alum.get("id") or alum.get("nombre")
                if ident is None or str(ident).strip() == "":
                    # fallback simple: usar índice si no hay dato útil
                    ident = f"alumno{idx}"
                version_asignada = etiquetas_version[(idx - 1) % len(etiquetas_version)]
                # determinar curso para este alumno (usar del alumno,
                # no del curso_final que es solo filtro)
                alum_curso = alum.get("curso") or ""
                if un_solo_pdf_todos:
                    ident_safe = re.sub(r"[^A-Za-z0-9_-]+", "_", str(ident))
                    if cantidad_versiones > 1:
                        fname = (
                            f"tmp_{idx:04d}_{ident_safe}_version_{version_asignada}.pdf"
                        )
                    else:
                        fname = f"tmp_{idx:04d}_{ident_safe}.pdf"
                else:
                    if cantidad_versiones > 1:
                        fname = (
                            f"Examen_{grado}_{alum_curso}_{area}_{evaluacion or 'general'}_"
                            f"{ident}_version_{version_asignada}.pdf"
                        )
                    else:
                        fname = (
                            f"Examen_{grado}_{alum_curso}_{area}_{evaluacion or 'general'}_"
                            f"{ident}.pdf"
                        )
                    fname = fname.replace(" ", "_")
                dest = os.path.join(out_dir, fname)
                # seleccionar preguntas por alumno (aleatorizar por alumno)
                try:
                    if cantidad_versiones > 1:
                        df_sel = examenes_por_version.get(version_asignada)
                        if df_sel is None:
                            raise ValueError(
                                f"No se pudo construir la versión {version_asignada}."
                            )
                    else:
                        seed = ident
                        df_sel = (
                            core_examenes_generacion.seleccionar_preguntas_por_textos(
                                df,
                                cantidad,
                                cantidad_textos,
                                rnd_seed=seed,
                            )
                        )
                except Exception:
                    # propagar la excepción hacia el llamador para informar al usuario
                    raise
                _cod = self._write_exam_pdf(
                    df_sel,
                    alum,
                    dest,
                    cantidad,
                    area,
                    evaluacion,
                    fecha=fecha,
                    cantidad_textos=cantidad_textos,
                    formato_examen=formato_examen,
                    docente_nombre=docente_nombre,
                    modo_hoja_respuestas=modo_hoja_respuestas,
                    version=version_asignada,
                )
                # guardar registro en BD (ignorar errores para no bloquear generación)
                try:
                    core_examenes_generacion.guardar_examen_generado(
                        exam_code=_cod or "",
                        grado=grado,
                        curso=alum.get("curso") or "",
                        area=area,
                        evaluacion=evaluacion or "",
                        version=version_asignada,
                        estudiante_nombre=alum.get("nombre") or "",
                        estudiante_documento=alum.get("documento")
                        or alum.get("id")
                        or "",
                        ruta_pdf=dest,
                    )
                except Exception:
                    pass
                core_examenes_generacion.guardar_detalle_examen(_cod, df_sel)
                if un_solo_pdf_todos:
                    pdfs_generados.append(dest)
                archivos_creados += 1

            if un_solo_pdf_todos:
                destino_unico = str(dest_filename or "").strip()
                if not destino_unico:
                    raise ValueError(
                        "Debe indicar el nombre del archivo destino para generar un único PDF."
                    )
                try:
                    self._merge_pdf_files(pdfs_generados, destino_unico)
                finally:
                    # limpiar temporales; no bloquear por errores de borrado
                    for pdf_tmp in pdfs_generados:
                        try:
                            os.remove(pdf_tmp)
                        except Exception:
                            pass

            return archivos_creados
        else:
            # estudiante individual
            version_asignada = etiquetas_version[0]
            if cantidad_versiones > 1:
                ver_sel = str(version_individual or "").strip().upper()
                if ver_sel in etiquetas_version:
                    version_asignada = ver_sel
            if dest_filename:
                dest = dest_filename
            else:
                curso_usar = curso_final or ""
                if cantidad_versiones > 1:
                    fname = (
                        f"Examen_{grado}_{curso_usar}_{area}_{evaluacion or 'general'}_"
                        f"{estudiante.get('id','')}_version_{version_asignada}.pdf"
                    )
                else:
                    fname = (
                        f"Examen_{grado}_{curso_usar}_{area}_{evaluacion or 'general'}_"
                        f"{estudiante.get('id','')}.pdf"
                    )
                fname = fname.replace(" ", "_")
                dest = os.path.join(out_dir, fname)
            # selección personalizada para el estudiante individual
            if cantidad_versiones > 1:
                df_sel = examenes_por_version.get(version_asignada)
                if df_sel is None:
                    raise ValueError(
                        f"No se pudo construir la versión {version_asignada}."
                    )
            else:
                seed = (
                    estudiante.get("documento")
                    or estudiante.get("id")
                    or estudiante.get("nombre")
                )
                df_sel = core_examenes_generacion.seleccionar_preguntas_por_textos(
                    df, cantidad, cantidad_textos, rnd_seed=seed
                )
            _cod = self._write_exam_pdf(
                df_sel,
                estudiante,
                dest,
                cantidad,
                area,
                evaluacion,
                fecha=fecha,
                cantidad_textos=cantidad_textos,
                formato_examen=formato_examen,
                docente_nombre=docente_nombre,
                modo_hoja_respuestas=modo_hoja_respuestas,
                version=version_asignada,
            )
            # guardar registro individual
            try:
                core_examenes_generacion.guardar_examen_generado(
                    exam_code=_cod or "",
                    grado=grado,
                    curso=estudiante.get("curso") or "",
                    area=area,
                    evaluacion=evaluacion or "",
                    version=version_asignada,
                    estudiante_nombre=estudiante.get("nombre") or "",
                    estudiante_documento=estudiante.get("documento")
                    or estudiante.get("id")
                    or "",
                    ruta_pdf=dest,
                )
            except Exception:
                pass
            core_examenes_generacion.guardar_detalle_examen(_cod, df_sel)
            return 1

    def _write_exam_pdf(
        self,
        preguntas_df,
        estudiante,
        path,
        cantidad,
        area,
        evaluacion,
        fecha=None,
        cantidad_textos=0,
        formato_examen=EXAM_FORMAT_STANDARD,
        docente_nombre=None,
        modo_hoja_respuestas=None,
        generar_hoja_respuestas=False,
        version=None,
    ):
        return core_examenes_pdf.write_exam_pdf(
            self,
            preguntas_df,
            estudiante,
            path,
            cantidad,
            area,
            evaluacion,
            fecha=fecha,
            cantidad_textos=cantidad_textos,
            formato_examen=formato_examen,
            docente_nombre=docente_nombre,
            modo_hoja_respuestas=modo_hoja_respuestas,
            generar_hoja_respuestas=generar_hoja_respuestas,
            version=version,
        )

        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import (
            BaseDocTemplate,
            Frame,
            PageTemplate,
            NextPageTemplate,
            Paragraph,
            Spacer,
            KeepTogether,
            CondPageBreak,
            FrameBreak,
            PageBreak,
            Image as RLImage,
            Flowable,
        )
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfgen.canvas import Canvas
        from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER
        from reportlab.lib.utils import ImageReader
        import importlib
        import os
        import re

        version_label = str(version or "").strip().upper() or "A"

        modo_hoja_respuestas = normalizar_modo_hoja_respuestas(
            modo_hoja_respuestas,
            generar_hoja_respuestas,
        )
        formato_examen = normalizar_formato_examen(formato_examen)
        if "matem" in str(area or "").strip().lower():
            formato_examen = EXAM_FORMAT_MATH_ICFES
        elif _es_area_lenguaje(area):
            if formato_examen in {
                EXAM_FORMAT_STANDARD,
                EXAM_FORMAT_LANGUAGE_ICFES,
            }:
                formato_examen = EXAM_FORMAT_LANGUAGE_ICFES
        elif formato_examen == EXAM_FORMAT_LANGUAGE_ICFES:
            formato_examen = EXAM_FORMAT_STANDARD
        is_math_icfes = formato_examen == EXAM_FORMAT_MATH_ICFES
        is_language_icfes = formato_examen == EXAM_FORMAT_LANGUAGE_ICFES
        generar_hoja_adicional = modo_hoja_respuestas == ANSWER_SHEET_MODE_APPEND
        integrar_hoja_ultima_pagina = modo_hoja_respuestas == ANSWER_SHEET_MODE_INLINE
        two_col_gap = 0.3 * cm
        qr_render_size = 2.95 * cm
        omr_bubble_diameter = 0.85 * cm
        omr_target_h_gap = 2.5 * cm
        omr_target_v_gap = 1.5 * cm
        omr_corner_mark_size = 1.0 * cm
        omr_qr_min_gap = 1.5 * cm
        omr_frame_margin = 0.7 * cm

        total_preguntas_hoja = max(1, int(cantidad or len(preguntas_df) or 1))

        def _sheet_column_ranges(total_preguntas):
            total = max(1, int(total_preguntas or 1))
            if total <= 30:
                split = max(1, (total + 1) // 2)
                ranges = [(1, split)]
                if split < total:
                    ranges.append((split + 1, total))
            else:
                ranges = []
                start = 1
                while start <= total and len(ranges) < 3:
                    end = min(total, start + 29)
                    ranges.append((start, end))
                    start = end + 1
                if start <= total:
                    ranges.append((start, total))

            while len(ranges) < 2:
                ranges.append((0, 0))
            if len(ranges) > 4:
                merged = ranges[:3]
                merged.append((ranges[3][0], ranges[-1][1]))
                ranges = merged
            return ranges

        def _embedded_sheet_metrics(total_preguntas):
            ranges = _sheet_column_ranges(total_preguntas)
            filas = max(
                1,
                max(
                    ((r[1] - r[0] + 1) for r in ranges if r[0] > 0 and r[1] >= r[0]),
                    default=1,
                ),
            )

            safety_gap = 20.0
            title_table_gap = 6.0
            title_font_size = 13
            qr_size = qr_render_size
            block_header_h = 18.0
            row_gap = 13.0
            table_tail = 10.0
            table_height = block_header_h + (filas * row_gap) + table_tail
            bottom_pad = 6.0

            required_from_content = (
                safety_gap
                + title_font_size
                + title_table_gap
                + table_height
                + bottom_pad
            )
            return {
                "safety_gap": safety_gap,
                "title_table_gap": title_table_gap,
                "title_font_size": title_font_size,
                "qr_size": qr_size,
                "table_height": table_height,
                "required_from_content": required_from_content,
            }

        def _get_last_content_y(page_state):
            if isinstance(page_state, dict):
                direct_y = page_state.get("_last_content_y", None)
                if isinstance(direct_y, (int, float)):
                    return float(direct_y)

                anchor_data = page_state.get("_answer_sheet_anchor")
                if isinstance(anchor_data, dict):
                    try:
                        return float(anchor_data.get("y", doc.bottomMargin))
                    except Exception:
                        return float(doc.bottomMargin)
            return float(doc.bottomMargin)

        def _draw_same_qr(target_canvas, x, y, qr_size):
            """
            Dibuja código QR de ALTA RESOLUCIÓN en el canvas PDF.

            El QR se generó con:
            - box_size=10 (alta resolución)
            - error_correction=ERROR_CORRECT_H (máxima tolerancia a errores)
            - border=4 (márgenes de seguridad)
            - formato PNG (sin compresión con pérdida)

            El dibujo utiliza:
            - preserveAspectRatio=True (mantienen proporción exacta)
            - Sin interpolación para evitar distorsión
            - mask='auto' para manejar transparencia correctamente
            """
            if _qr_img_data is not None:
                try:
                    _qr_img_data.seek(0)
                    # Dibujar imagen PNG de QR de alta resolución SIN INTERPOLACIÓN
                    target_canvas.drawImage(
                        ImageReader(_qr_img_data),
                        x,
                        y,
                        width=qr_size,
                        height=qr_size,
                        preserveAspectRatio=True,
                        mask="auto",
                    )
                    return
                except Exception as exc:
                    # Silenciar excepción para pasar al fallback
                    pass

            # Fallback sin dependencia externa: QR vectorial de ReportLab.
            # (Se usa si la generación con qrcode falla)
            if not _qr_content:
                return
            try:
                from reportlab.graphics.barcode import qr as _rl_qr
                from reportlab.graphics.shapes import Drawing
                from reportlab.graphics import renderPDF

                qr_widget = _rl_qr.QrCodeWidget(_qr_content)
                bounds = qr_widget.getBounds()
                bw = max(1.0, float(bounds[2] - bounds[0]))
                bh = max(1.0, float(bounds[3] - bounds[1]))

                drawing = Drawing(
                    qr_size,
                    qr_size,
                    transform=[qr_size / bw, 0, 0, qr_size / bh, 0, 0],
                )
                drawing.add(qr_widget)
                renderPDF.draw(drawing, target_canvas, x, y)
            except Exception as exc:
                # Silenciar excepción silenciosa
                pass

        def _draw_sheet_table(
            target_canvas,
            left_x,
            right_x,
            top_y,
            bottom_y,
            total_preguntas,
            embedded=False,
        ):
            ranges = _sheet_column_ranges(total_preguntas)
            num_columnas = len(ranges)
            filas_por_columna = max(
                1,
                max(
                    ((r[1] - r[0] + 1) for r in ranges if r[0] > 0 and r[1] >= r[0]),
                    default=1,
                ),
            )

            zone_w = max(1.0, right_x - left_x)
            zone_h = max(1.0, top_y - bottom_y)
            inner_left = left_x + (0.03 * zone_w)
            inner_right = right_x - (0.03 * zone_w)
            usable_w = max(1.0, inner_right - inner_left)
            col_gap = (0.02 * zone_w) if num_columnas > 1 else 0.0
            col_width = (usable_w - ((num_columnas - 1) * col_gap)) / max(
                1, num_columnas
            )
            block_header_h = max(12.0, 0.075 * zone_h)
            rows_top = top_y - block_header_h
            rows_bottom = bottom_y + (0.03 * zone_h)
            usable_h = max(1.0, rows_top - rows_bottom)
            row_gap = min(
                omr_target_v_gap,
                usable_h / max(1, filas_por_columna),
            )
            row_gap = max(omr_bubble_diameter + (0.2 * cm), row_gap)
            if (row_gap * filas_por_columna) > usable_h:
                row_gap = usable_h / max(1, filas_por_columna)

            number_w = max(1.25 * cm, 0.20 * col_width)
            left_pad = 0.25 * cm
            right_pad = 0.2 * cm

            bubble_r_target = omr_bubble_diameter / 2.0
            option_font = 10
            number_font = 10 if embedded else 11
            block_font = 8 if embedded else 9

            target_canvas.setFillColor(colors.black)
            target_canvas.setFont(font_family, option_font)

            for col_idx in range(num_columnas):
                x_col = inner_left + (col_idx * (col_width + col_gap))
                r_start, r_end = ranges[col_idx]
                block_label = (
                    f"{r_start}-{r_end}" if r_start > 0 and r_end >= r_start else "--"
                )

                max_right = x_col + col_width - right_pad
                opt_start_x = x_col + number_w + left_pad
                available_track = max(1.0, max_right - opt_start_x)
                bubble_gap = min(omr_target_h_gap, available_track / 3.0)
                bubble_r = min(
                    bubble_r_target,
                    max(3.2, (bubble_gap * 0.44)),
                )

                min_center_gap = (2.0 * bubble_r) + (0.2 * cm)
                if bubble_gap < min_center_gap:
                    bubble_gap = min_center_gap
                final_track = 3.0 * bubble_gap
                if final_track > available_track:
                    bubble_gap = max(1.0, available_track / 3.0)
                    bubble_r = min(bubble_r, bubble_gap * 0.45)

                target_canvas.setFont(f"{font_family}-Bold", block_font)
                target_canvas.drawCentredString(
                    x_col + (col_width / 2.0),
                    top_y - (block_header_h * 0.72),
                    block_label,
                )

                for row_idx in range(filas_por_columna):
                    pregunta = r_start + row_idx
                    if r_start <= 0 or r_end < r_start or pregunta > r_end:
                        continue

                    y_row = rows_top - ((row_idx + 0.5) * row_gap)
                    target_canvas.setFont(f"{font_family}-Bold", number_font)
                    num_y = y_row - 3
                    target_canvas.drawRightString(
                        x_col + number_w - (0.15 * cm), num_y, f"{pregunta}."
                    )

                    target_canvas.setFont(font_family, option_font)
                    target_canvas.setLineWidth(1.25)
                    for opt_idx, opt in enumerate(["A", "B", "C", "D"]):
                        cx = opt_start_x + (opt_idx * bubble_gap)
                        circle_y = y_row
                        label_y = y_row - (bubble_r + 8)
                        target_canvas.circle(cx, circle_y, bubble_r, stroke=1, fill=0)
                        target_canvas.drawCentredString(cx, label_y, opt)

        def _draw_alignment_marks(target_canvas, left_x, right_x, top_y, bottom_y):
            mark = omr_corner_mark_size
            target_canvas.setFillColor(colors.black)
            target_canvas.rect(left_x, top_y - mark, mark, mark, stroke=0, fill=1)
            target_canvas.rect(
                right_x - mark, top_y - mark, mark, mark, stroke=0, fill=1
            )
            target_canvas.rect(left_x, bottom_y, mark, mark, stroke=0, fill=1)
            target_canvas.rect(right_x - mark, bottom_y, mark, mark, stroke=0, fill=1)

        def _draw_document_frame(target_canvas, page_w, page_h):
            frame_w = page_w - (2 * omr_frame_margin)
            frame_h = page_h - (2 * omr_frame_margin)
            target_canvas.setStrokeColor(colors.black)
            target_canvas.setLineWidth(1.2)
            target_canvas.rect(
                omr_frame_margin,
                omr_frame_margin,
                frame_w,
                frame_h,
                stroke=1,
                fill=0,
            )

            _draw_alignment_marks(
                target_canvas,
                omr_frame_margin,
                page_w - omr_frame_margin,
                page_h - omr_frame_margin,
                omr_frame_margin,
            )

        def _draw_embedded_answer_sheet(
            target_canvas,
            page_num,
            total_pages,
            content_end_y,
        ):
            page_w, page_h = target_canvas._pagesize
            page_left = doc.leftMargin
            page_right = page_w - doc.rightMargin
            # Hoja embebida: en formato Matemáticas ICFES usa toda el área útil (1 columna).
            if is_math_icfes:
                left_x = page_left
                right_x = page_right
            else:
                col_width = (
                    page_w - doc.leftMargin - doc.rightMargin - two_col_gap
                ) / 2.0
                left_x = page_left
                right_x = left_x + col_width
            metrics = _embedded_sheet_metrics(total_preguntas_hoja)

            available_bottom = doc.bottomMargin + (0.25 * cm)
            title_y = content_end_y - metrics["safety_gap"]

            qr_size = metrics["qr_size"]
            min_qr_gap_marks = 34.0
            qr_x_lim = right_x - min_qr_gap_marks - qr_size
            qr_x = min(page_right - qr_size, qr_x_lim)
            title_font_size = metrics["title_font_size"]
            qr_y = title_y - qr_size + title_font_size
            pin_text = f"PIN: {exam_id}"

            title_bottom = title_y - title_font_size
            table_top = min(
                title_bottom - metrics["title_table_gap"],
                qr_y - omr_qr_min_gap,
            )
            table_bottom = table_top - metrics["table_height"]
            bottom_y = max(available_bottom, table_bottom - 4)
            row_top = table_top

            target_canvas.saveState()
            _draw_document_frame(target_canvas, page_w, page_h)
            target_canvas.setStrokeColor(colors.HexColor("#555555"))
            target_canvas.setLineWidth(0.8)
            target_canvas.line(left_x, row_top + 3, right_x, row_top + 3)
            target_canvas.setFillColor(colors.black)
            target_canvas.setFont(f"{font_family}-Bold", title_font_size)
            target_canvas.drawString(left_x, title_y, "HOJA DE RESPUESTAS - SEA")
            _draw_same_qr(target_canvas, qr_x, qr_y, qr_size)
            target_canvas.setFont(f"{font_family}-Bold", 9)
            target_canvas.drawString(qr_x, qr_y - 10, pin_text)
            _draw_sheet_table(
                target_canvas,
                left_x,
                right_x,
                row_top,
                bottom_y,
                total_preguntas_hoja,
                embedded=True,
            )
            target_canvas.draw_footer(page_num, total_pages)
            target_canvas.restoreState()

        def _draw_append_answer_sheet_page(target_canvas, total_pages):
            page_num = total_pages
            page_w, page_h = target_canvas._pagesize
            margin_x = 1.5 * cm
            top_y = page_h - (1.3 * cm)

            target_canvas.saveState()
            _draw_document_frame(target_canvas, page_w, page_h)
            target_canvas.setFont(f"{font_family}-Bold", 14)
            target_canvas.drawCentredString(
                page_w / 2, top_y, "HOJA DE RESPUESTAS - SEA"
            )

            target_canvas.setFont(font_family, 10)
            target_canvas.drawCentredString(
                page_w / 2,
                top_y - 12,
                "Sistema SEA - Sistema de Evaluación Automatizada",
            )

            estudiante_nombre = (
                str(construir_nombre(estudiante) or "").strip() or est_nombre_marca
            )
            estudiante_documento = (
                str(estudiante.get("documento", estudiante.get("id", "")) or "").strip()
                or est_doc_marca
            )
            estudiante_grado = str(estudiante.get("grado", "") or "").strip()
            estudiante_curso = str(estudiante.get("curso", "") or "").strip()

            info_y = top_y - 30
            line_gap = 11
            target_canvas.drawString(margin_x, info_y, f"Área: {area}")
            target_canvas.drawString(page_w / 2, info_y, f"Grado: {estudiante_grado}")
            target_canvas.drawString(
                margin_x, info_y - line_gap, f"Curso: {estudiante_curso}"
            )
            target_canvas.drawString(
                page_w / 2, info_y - line_gap, f"Evaluación: {evaluacion or ''}"
            )
            target_canvas.drawString(
                page_w / 2,
                info_y - (2 * line_gap),
                f"Versión: {version_label}",
            )
            target_canvas.drawString(
                margin_x,
                info_y - (3 * line_gap),
                f"Nombre del estudiante: {estudiante_nombre}",
            )
            target_canvas.drawString(
                margin_x,
                info_y - (4 * line_gap),
                f"Documento: {estudiante_documento}",
            )

            # QR en alta resolución con tamaño visual contenido.
            qr_size = qr_render_size
            qr_x = page_w - margin_x - qr_size
            qr_y = info_y - (4 * line_gap) - 2
            _draw_same_qr(target_canvas, qr_x, qr_y, qr_size)
            target_canvas.setFont(f"{font_family}-Bold", 10)
            target_canvas.drawString(qr_x, qr_y - 12, f"PIN: {exam_id}")

            inst_y = info_y - (5.5 * line_gap)
            target_canvas.setFont(f"{font_family}-Bold", 11)
            target_canvas.drawString(margin_x, inst_y, "INSTRUCCIONES")
            target_canvas.setFont(font_family, 9)
            instrucciones = [
                "1. Marque completamente el círculo correspondiente a la respuesta correcta.",
                "2. Utilice lápiz o bolígrafo negro.",
                "3. No marque más de una opción por pregunta.",
                "4. No haga marcas fuera de los círculos.",
            ]
            for idx, texto in enumerate(instrucciones, start=1):
                target_canvas.drawString(
                    margin_x + (0.1 * cm), inst_y - (idx * 11), texto
                )

            table_top = min(inst_y - 66, qr_y - omr_qr_min_gap)
            bottom_y = doc.bottomMargin + (0.35 * cm)
            _draw_sheet_table(
                target_canvas,
                margin_x,
                page_w - margin_x,
                table_top,
                bottom_y,
                total_preguntas_hoja,
                embedded=False,
            )

            target_canvas.draw_footer(page_num, total_pages)
            target_canvas.restoreState()

        # Registrar Arial de forma robusta (ruta local o ruta del sistema).
        try:
            arial_candidates = [
                "arial.ttf",
                os.path.join("C:\\Windows\\Fonts", "arial.ttf"),
            ]
            arial_bold_candidates = [
                "arialbd.ttf",
                os.path.join("C:\\Windows\\Fonts", "arialbd.ttf"),
            ]

            arial_path = next((p for p in arial_candidates if os.path.exists(p)), None)
            arial_bold_path = next(
                (p for p in arial_bold_candidates if os.path.exists(p)), None
            )

            if not arial_path or not arial_bold_path:
                raise FileNotFoundError("No se encontraron archivos de Arial")

            pdfmetrics.registerFont(TTFont("Arial", arial_path))
            pdfmetrics.registerFont(TTFont("Arial-Bold", arial_bold_path))
            font_family = "Arial"
        except Exception:
            font_family = "Helvetica"

        class NumberedCanvas(Canvas):
            def __init__(self, *args, **kwargs):
                Canvas.__init__(self, *args, **kwargs)
                self._saved_page_states = []

            def showPage(self):
                self._saved_page_states.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                append_sheet_page = generar_hoja_adicional
                content_end_y = float(doc.bottomMargin)
                if integrar_hoja_ultima_pagina and self._saved_page_states:
                    content_end_y = _get_last_content_y(self._saved_page_states[-1])
                    available_height = max(
                        0,
                        content_end_y - (doc.bottomMargin + (0.25 * cm)),
                    )
                    required_height = _embedded_sheet_metrics(total_preguntas_hoja)[
                        "required_from_content"
                    ]
                    inline_sheet_fits = available_height >= required_height
                    append_sheet_page = not inline_sheet_fits
                else:
                    inline_sheet_fits = False

                extra_pages = 1 if append_sheet_page else 0
                num_pages = len(self._saved_page_states) + extra_pages
                for page_num, state in enumerate(self._saved_page_states, start=1):
                    self.__dict__.update(state)
                    if (
                        integrar_hoja_ultima_pagina
                        and inline_sheet_fits
                        and page_num == len(self._saved_page_states)
                    ):
                        last_content_y = _get_last_content_y(state)
                        _draw_embedded_answer_sheet(
                            self,
                            page_num,
                            num_pages,
                            last_content_y,
                        )
                    else:
                        self.draw_footer(page_num, num_pages)
                    Canvas.showPage(self)

                if append_sheet_page:
                    _draw_append_answer_sheet_page(self, num_pages)
                    Canvas.showPage(self)

                Canvas.save(self)

            def draw_footer(self, page_num, total_pages):
                self.setFont(font_family, 8)
                docente_impresion = str(docente_nombre or "").strip() or "N/D"
                footer_text = (
                    f"Autor del examen: {docente_impresion} | "
                    f"Generado por SEA Sistema de Evaluación Automatizada | "
                    f"Página {page_num} de {total_pages}"
                )
                self.drawCentredString(self._pagesize[0] / 2, 1.0 * cm, footer_text)

        # --- Generar QR antes del header para reutilizarlo por closure ---
        import io as _io

        version_token = str(version_label or "A").strip().upper() or "A"
        exam_id = core_examenes_generacion.crear_codigo_examen(version_token)

        _qr_img_data = None
        _qr_content = ""
        try:
            _qrcode_lib = importlib.import_module("qrcode")

            _est_doc_qr = estudiante.get("documento", estudiante.get("id", ""))
            _est_nombre_qr = estudiante.get("nombre", "")
            _est_grado_qr = estudiante.get("grado", "")
            _est_curso_qr = estudiante.get("curso", "")
            _qr_content = core_examenes_generacion.construir_contenido_qr(
                documento=_est_doc_qr,
                nombre=_est_nombre_qr,
                grado=_est_grado_qr,
                curso=_est_curso_qr,
                area=area,
                evaluacion=evaluacion or "",
                exam_id=exam_id,
            )
            # Generar QR en ALTA RESOLUCIÓN con parámetros óptimos para PDF
            _qr = _qrcode_lib.QRCode(
                version=None,
                error_correction=_qrcode_lib.constants.ERROR_CORRECT_H,
                box_size=10,  # IMPORTANTE: mínimo 8, preferible 10
                border=4,  # Mínimo 4 para márgenes seguros
            )
            _qr.add_data(_qr_content)
            _qr.make(fit=True)
            # Crear imagen PNG sin interpolación (formato nativo recomendado)
            _qr_pil = _qr.make_image(fill_color="black", back_color="white")
            _qr_buf = _io.BytesIO()
            # Guardar como PNG en memoria (no usar JPG ni otros formatos)
            _qr_pil.save(_qr_buf, format="PNG", optimize=False)
            _qr_buf.seek(0)
            _qr_img_data = _qr_buf
        except Exception:
            _qr_img_data = None
            try:
                _est_doc_qr = estudiante.get("documento", estudiante.get("id", ""))
                _est_nombre_qr = estudiante.get("nombre", "")
                _est_grado_qr = estudiante.get("grado", "")
                _est_curso_qr = estudiante.get("curso", "")
                _qr_content = core_examenes_generacion.construir_contenido_qr(
                    documento=_est_doc_qr,
                    nombre=_est_nombre_qr,
                    grado=_est_grado_qr,
                    curso=_est_curso_qr,
                    area=area,
                    evaluacion=evaluacion or "",
                    exam_id=exam_id,
                )
            except Exception:
                _qr_content = ""

        # Marca de seguridad personalizada del estudiante.
        est_doc_marca = str(
            estudiante.get("documento", estudiante.get("id", "")) or ""
        ).strip()
        est_nombre_marca = str(construir_nombre(estudiante) or "").strip()

        # Intentar obtener nombre/documento actualizados desde BD al momento de generar.
        if est_doc_marca:
            try:
                _est_db = core_matricula.buscar_estudiante(est_doc_marca)
                if _est_db:
                    nombre_estudiante_db = str(_est_db.get("nombre") or "").strip()
                    documento_db = str(_est_db.get("documento") or "").strip()
                    if nombre_estudiante_db:
                        est_nombre_marca = nombre_estudiante_db
                    if documento_db:
                        est_doc_marca = documento_db
            except Exception:
                pass

        if not est_nombre_marca:
            est_nombre_marca = "N/D"
        if not est_doc_marca:
            est_doc_marca = "N/D"

        marca_estudiante = (
            f"Nombre del estudiante: {est_nombre_marca} | Documento: {est_doc_marca}"
        )

        def header_callback(canvas, doc):
            canvas.saveState()
            w, h = A4
            margin = 1.27 * cm
            canvas._last_content_y = h - doc.topMargin

            # datos de institución
            instit = (self._get_config_plantel("nombre_institucion") or "").upper()
            nit = self._get_config_plantel("nit") or ""
            dane = self._get_config_plantel("codigo_dane") or ""
            departamento = self._get_config_plantel("departamento") or ""
            municipio = self._get_config_plantel("municipio") or ""
            corre = self._get_config_plantel("corregimiento_localidad") or ""
            decreto = self._get_config_plantel("decreto_funcionamiento") or ""
            resolucion = self._get_config_plantel("resolucion_aprobacion") or ""
            logo_path = self._get_config_plantel("logo_path")

            # dibujar logo a la izquierda (más alto para no cabalgar sobre texto)
            logo_w = 2.45 * cm
            logo_h = 2.45 * cm
            logo_x = margin - 0.15 * cm
            if logo_path and os.path.exists(logo_path):
                try:
                    canvas.drawImage(
                        logo_path,
                        logo_x,
                        h - margin - logo_h + 12,
                        width=logo_w,
                        height=logo_h,
                        preserveAspectRatio=True,
                        mask="auto",
                    )
                except Exception:
                    pass

            # texto a la derecha del logo
            text_x = logo_x + logo_w + 0.02 * cm
            y1 = h - margin - 0.2 * cm
            titulo_y = y1 + 8
            info_y = y1 - 9
            canvas.setFont(f"{font_family}-Bold", 12)
            canvas.drawCentredString(w / 2, titulo_y, instit)

            canvas.setFont(font_family, 8)
            canvas.drawString(text_x, info_y, f"NIT: {nit}   DANE: {dane}")
            canvas.setFont(font_family, 7)
            canvas.drawString(
                text_x,
                info_y - 8,
                f"Resolución de Funcionamiento: {decreto}",
            )
            canvas.drawString(
                text_x,
                info_y - 16,
                f"Resolución de aprobación: {resolucion}",
            )
            canvas.setFont(f"{font_family}-Bold", 11)
            canvas.drawRightString(w - margin, y1, f"VERSION {version_label}")
            # Fila 4: Corregimiento, Municipio y Departamento
            canvas.setFont(font_family, 8)
            canvas.drawString(
                text_x,
                info_y - 24,
                f"Corregimiento: {corre} Municipio: {municipio} Departamento: {departamento}",
            )
            canvas.line(margin, y1 - 38, w - margin, y1 - 38)

            # líneas del estudiante justo debajo (centradas para evitar que se oculten bajo el logo)
            est_name = estudiante.get("nombre", "")
            est_doc = estudiante.get("documento", estudiante.get("id", ""))
            est_grado = estudiante.get("grado", "")
            est_curso = estudiante.get("curso", "")

            extra_gap_docente = 10 if formato_examen == EXAM_FORMAT_STANDARD else 0
            y2 = y1 - 52 - extra_gap_docente
            canvas.setFont(font_family, 10)
            fecha_txt = str(fecha or "").strip() or "________"
            docente_txt = str(docente_nombre or "").strip() or "________"
            # Fila 1: Docente y Área (valor del área en negrita y mayúsculas)
            _prefix_area = f"Docente: {docente_txt}   Área: "
            _area_upper = area.upper()
            docente_x = margin
            canvas.setFont(font_family, 10)
            canvas.drawString(docente_x, y2, _prefix_area)
            _prefix_area_w = canvas.stringWidth(_prefix_area, font_family, 10)
            canvas.setFont(f"{font_family}-Bold", 10)
            canvas.drawString(docente_x + _prefix_area_w, y2, _area_upper)
            canvas.setFont(font_family, 10)
            # Fila 2: Estudiante (nombre en negrita), Grado, Curso, Fecha, Nota
            _label_est = "Estudiante: "
            _rest_est = f"   Grado: {est_grado}  - {est_curso}   Fecha: {fecha_txt}   Nota:_____"
            student_x = margin
            canvas.setFont(font_family, 10)
            canvas.drawString(student_x, y2 - 11, _label_est)
            _label_w = canvas.stringWidth(_label_est, font_family, 10)
            canvas.setFont(f"{font_family}-Bold", 10)
            canvas.drawString(student_x + _label_w, y2 - 11, est_name)
            _name_w = canvas.stringWidth(est_name, f"{font_family}-Bold", 10)
            canvas.setFont(font_family, 10)
            canvas.drawString(student_x + _label_w + _name_w, y2 - 11, _rest_est)

            # QR en esquina superior derecha del encabezado con tamaño visual contenido.
            qr_size = qr_render_size
            qr_x = w - margin - qr_size
            qr_y = h - margin - qr_size + 12
            _draw_same_qr(canvas, qr_x, qr_y, qr_size)
            canvas.setFont(font_family, 8)
            canvas.drawCentredString(
                qr_x + (qr_size / 2), qr_y - 6, f"ID Examen: {exam_id}"
            )

            # Marca vertical del sistema en margen derecho, fuera del contenido.
            marca_sistema = (
                "Software y Soporte Tecnológico: Robert Calanche Villa | "
                "SEA | Versión 1.0 | 2026"
            )
            canvas.saveState()
            canvas.setFillColor(colors.Color(0.25, 0.25, 0.25))
            canvas.setFont(font_family, 8)
            canvas.translate(w - 0.45 * cm, h / 2)
            canvas.rotate(90)
            canvas.drawCentredString(0, 0, marca_sistema)
            canvas.restoreState()

            # Marca de seguridad del estudiante en margen izquierdo.
            canvas.saveState()
            canvas.setFillColor(colors.HexColor("#555555"))
            canvas.setFont(font_family, 8)
            canvas.translate(0.6 * cm, h / 2)
            canvas.rotate(90)
            canvas.drawCentredString(0, 0, marca_estudiante)
            canvas.restoreState()

            canvas.restoreState()

        # Márgenes formato estrecho para maximizar espacio de contenido
        header_height = 2.5 * cm
        math_margin_top = 110.0
        math_margin_bottom = 60.0
        _side_margin = 1.27 * cm
        _bottom_margin = 1.27 * cm
        _top_margin = 1.27 * cm + header_height

        doc = BaseDocTemplate(
            path,
            pagesize=A4,
            leftMargin=_side_margin,
            rightMargin=_side_margin,
            topMargin=_top_margin,
            bottomMargin=_bottom_margin,
            canvasmaker=NumberedCanvas,
        )

        width, height = A4

        class TrackingFrame(Frame):
            def _track_y(self, canv, added):
                if not added:
                    return
                try:
                    y_actual = float(getattr(self, "_y", 0) or 0)
                    y_prev = getattr(canv, "_last_content_y", None)
                    if y_prev is None:
                        canv._last_content_y = y_actual
                    else:
                        canv._last_content_y = min(float(y_prev), y_actual)
                except Exception:
                    pass

            def _add(self, flowable, canv, trySplit=0):
                added = Frame._add(self, flowable, canv, trySplit=trySplit)
                self._track_y(canv, added)
                return added

            def add(self, flowable, canv, trySplit=0):
                added = Frame.add(self, flowable, canv, trySplit=trySplit)
                self._track_y(canv, added)
                return added

        # Layout del contenido:
        # - Matemáticas ICFES: una columna.
        # - Lenguaje ICFES: texto (una columna) + preguntas (dos columnas).
        # - Estándar: dos columnas.
        gap = 0.3 * cm
        if is_math_icfes:
            col_width = doc.width
            frame_main = TrackingFrame(
                doc.leftMargin,
                doc.bottomMargin,
                doc.width,
                doc.height,
                leftPadding=0,
                rightPadding=0,
                topPadding=0,
                bottomPadding=0,
                id="col1",
            )
            template = PageTemplate(
                id="OneCol", frames=[frame_main], onPage=header_callback
            )
            doc.addPageTemplates([template])
        elif is_language_icfes:
            _lang_col_gap = 20  # pts separación central entre columnas
            _lang_col_w = (doc.width - _lang_col_gap) / 2
            frame_lang_left = TrackingFrame(
                doc.leftMargin,
                doc.bottomMargin,
                _lang_col_w,
                doc.height,
                leftPadding=0,
                rightPadding=_lang_col_gap / 2,
                id="lang_left",
            )
            frame_lang_right = TrackingFrame(
                doc.leftMargin + _lang_col_w + _lang_col_gap / 2,
                doc.bottomMargin,
                _lang_col_w,
                doc.height,
                leftPadding=_lang_col_gap / 2,
                rightPadding=0,
                id="lang_right",
            )

            def _lang_header_callback(canvas, doc_obj):
                header_callback(canvas, doc_obj)
                # Separador vertical entre las dos columnas
                x_sep = doc_obj.leftMargin + _lang_col_w + _lang_col_gap / 2
                canvas.saveState()
                canvas.setStrokeColor(colors.HexColor("#aaaaaa"))
                canvas.setLineWidth(0.5)
                canvas.line(
                    x_sep,
                    doc_obj.bottomMargin,
                    x_sep,
                    doc_obj.bottomMargin + doc_obj.height,
                )
                canvas.restoreState()

            template_lang = PageTemplate(
                id="LangTwoCol",
                frames=[frame_lang_left, frame_lang_right],
                onPage=_lang_header_callback,
            )
            doc.addPageTemplates([template_lang])
        else:
            col_width = (doc.width - gap) / 2
            frame1 = TrackingFrame(
                doc.leftMargin,
                doc.bottomMargin,
                col_width,
                doc.height,
                leftPadding=0,
                rightPadding=gap / 2,
                id="col1",
            )
            frame2 = TrackingFrame(
                doc.leftMargin + col_width + gap / 2,
                doc.bottomMargin,
                col_width,
                doc.height,
                leftPadding=gap / 2,
                rightPadding=0,
                id="col2",
            )
            template = PageTemplate(
                id="TwoCol", frames=[frame1, frame2], onPage=header_callback
            )
            doc.addPageTemplates([template])

        def _track_after_flowable(_flowable):
            try:
                canv = getattr(doc, "canv", None)
                frame = getattr(doc, "frame", None)
                if canv is None or frame is None:
                    return
                y_actual = float(getattr(frame, "_y", 0) or 0)
                y_prev = getattr(canv, "_last_content_y", None)
                if y_prev is None:
                    canv._last_content_y = y_actual
                else:
                    canv._last_content_y = min(float(y_prev), y_actual)
            except Exception:
                pass

        doc.afterFlowable = _track_after_flowable

        styles = getSampleStyleSheet()
        exam_font_size = 12 if is_math_icfes else 11
        exam_leading = 13 if is_math_icfes else 12
        style_instruction = ParagraphStyle(
            "Instruction",
            parent=styles["Normal"],
            fontName=font_family,
            fontSize=exam_font_size,
            leading=exam_leading,
            spaceAfter=0,
            alignment=TA_JUSTIFY,
        )
        style_bold = ParagraphStyle(
            "Bold",
            parent=styles["Normal"],
            fontName=f"{font_family}-Bold",
            fontSize=exam_font_size,
            leading=exam_leading,
            spaceAfter=0,
            alignment=TA_JUSTIFY,
        )
        style_normal = ParagraphStyle(
            "Normal",
            parent=styles["Normal"],
            fontName=font_family,
            fontSize=exam_font_size,
            leading=exam_leading,
            spaceAfter=0,
            alignment=TA_JUSTIFY,
        )
        style_option = ParagraphStyle(
            "Option",
            parent=styles["Normal"],
            fontName=font_family,
            fontSize=exam_font_size,
            leading=max(exam_font_size + 0.5, exam_leading - 1),
            spaceAfter=0,
            leftIndent=0.2 * cm,
            alignment=TA_LEFT,
        )
        style_option_math = ParagraphStyle(
            "OptionMath",
            parent=style_option,
            leftIndent=0.15 * cm,
            spaceAfter=0,
        )
        style_responde = ParagraphStyle(
            "Responde",
            parent=styles["Normal"],
            fontName=font_family,
            fontSize=exam_font_size,
            leading=exam_leading,
            spaceAfter=0,
            alignment=TA_JUSTIFY,
        )
        style_formula_center = ParagraphStyle(
            "FormulaCenter",
            parent=styles["Normal"],
            fontName=font_family,
            fontSize=exam_font_size,
            leading=exam_leading,
            spaceAfter=0,
            alignment=TA_CENTER,
        )

        flowables = []

        class AnswerSheetAnchorFlowable(Flowable):
            def wrap(self, availWidth, availHeight):
                return (0, 0)

            def draw(self):
                return

            def drawOn(self, canv, x, y, _sW=0):
                canv._answer_sheet_anchor = {
                    "page": canv.getPageNumber(),
                    "x": x,
                    "y": y,
                }
                Flowable.drawOn(self, canv, x, y, _sW)

        class IndivisibleBlockFlowable(Flowable):
            def __init__(self, block_flowables):
                super().__init__()
                self.block_flowables = list(block_flowables or [])
                self._wrapped_items = []

            def wrap(self, availWidth, availHeight):
                self._wrapped_items = []
                total_h = 0.0
                max_w = 0.0
                for item in self.block_flowables:
                    w, h = item.wrap(availWidth, availHeight)
                    h = max(0.0, float(h or 0.0))
                    self._wrapped_items.append((item, float(w or 0.0), h))
                    total_h += h
                    max_w = max(max_w, float(w or 0.0))
                self.width = max_w
                self.height = total_h
                return availWidth, total_h

            def split(self, availWidth, availHeight):
                return []

            def draw(self):
                y_cursor = float(getattr(self, "height", 0.0) or 0.0)
                for item, _w, h in self._wrapped_items:
                    y_cursor -= h
                    item.drawOn(self.canv, 0, y_cursor)

        instruccion_lectura = (
            "Lea cada pregunta cuidadosamente y seleccione la respuesta correcta."
            if is_math_icfes
            else "Lee con atención los siguientes textos y responde las preguntas."
        )
        intro_top_spacing = 0.2 * cm

        # --- Depuración: imprimir valores solicitados y forma inicial ---
        try:
            print(f"[DEBUG] cantidad_textos solicitado: {cantidad_textos}")
            print(f"[DEBUG] preguntas_df filas iniciales: {len(preguntas_df)}")
        except Exception:
            pass

        # Agrupar sobre el DataFrame completo para asegurar que la selección
        # de textos (contextos) se realiza antes de truncar por cantidad de
        # preguntas. Esto evita que un .head(cantidad) previo elimine
        # contextos completos.
        preguntas_all = preguntas_df.copy()

        if "id_contexto" in preguntas_all.columns:
            grupos_all = [
                (ctx, grp)
                for ctx, grp in preguntas_all.groupby("id_contexto", sort=False)
            ]
            total_contextos = len(grupos_all)
            try:
                print(f"[DEBUG] contextos encontrados: {total_contextos}")
            except Exception:
                pass
            # aplicar límite de textos si se solicitó
            if cantidad_textos and cantidad_textos > 0:
                grupos_sel = grupos_all[:cantidad_textos]
            else:
                grupos_sel = grupos_all
            try:
                print(
                    f"[DEBUG] contextos seleccionados (post-slice): {len(grupos_sel)}"
                )
            except Exception:
                pass
            # Usar la lista de grupos seleccionados tal cual para garantizar
            # que la cantidad de textos solicitada se preserve. No recomponemos
            # un DataFrame y luego volvemos a agrupar porque ese head() puede
            # eliminar contextos completos.
            grupos = grupos_sel
            total_filas = sum(len(g) for (_, g) in grupos)
            try:
                print(
                    f"[DEBUG] filas totales en contextos seleccionados: {total_filas}"
                )
                print(f"[DEBUG] grupos finales a procesar (sin truncar): {len(grupos)}")
            except Exception:
                pass
        else:
            # sin contexto definido
            preguntas = preguntas_all.copy()
            if cantidad is not None and cantidad > 0 and cantidad < len(preguntas):
                preguntas = preguntas.head(cantidad).copy()
            grupos = [(None, preguntas)]

        pregunta_num = 1
        contextos_usados = {}
        contador_texto = 1
        _math_buffers = []
        _formula_fallback_notificado = False

        def _texto_valido(valor):
            if valor is None:
                return False
            texto = str(valor).strip()
            return bool(texto) and texto.lower() != "nan"

        def _append_group_image(target_block, imagen_ref):
            if not _texto_valido(imagen_ref):
                return
            img_path = (
                imagen_ref
                if os.path.isabs(imagen_ref)
                else os.path.join(self.imagenes_dir, str(imagen_ref))
            )
            if not os.path.exists(img_path):
                return
            try:
                max_w = 6 * cm
                max_h = 4 * cm
                reader = ImageReader(img_path)
                iw, ih = reader.getSize()
                if iw > 0 and ih > 0:
                    scale = min(max_w / iw, max_h / ih, 1)
                    img_w = iw * scale
                    img_h = ih * scale
                else:
                    img_w = max_w
                    img_h = max_h

                target_block.append(RLImage(img_path, width=img_w, height=img_h))
            except Exception:
                pass

        def _append_text_with_math(
            target_block,
            texto,
            style,
            prefijo="",
            centrar_formula=False,
        ):
            nonlocal _formula_fallback_notificado
            txt = str(texto or "")
            if not _texto_valido(txt):
                return

            try:
                expresiones = _extraer_expresiones_latex(txt)
            except ValueError:
                expresiones = []

            def _append_formula_as_text(expr_txt, prefijo_local=""):
                texto_formula = _formula_a_texto_legible(expr_txt) or str(
                    expr_txt or ""
                )
                if prefijo_local:
                    texto_formula = f"{prefijo_local}{texto_formula}"
                estilo = (
                    style_formula_center
                    if (centrar_formula or is_math_icfes)
                    else style
                )
                target_block.append(Paragraph(texto_formula, estilo))

            def _append_formula(expr_txt):
                nonlocal _formula_fallback_notificado
                rendered = renderizar_formula(expr_txt, dpi=220)
                if rendered is None:
                    if not _formula_fallback_notificado:
                        try:
                            print(
                                "No se pudo renderizar la fórmula. Se mostrará en formato texto."
                            )
                        except Exception:
                            pass
                        _formula_fallback_notificado = True
                    return None

                buf, dpi_math = rendered
                _math_buffers.append(buf)
                try:
                    iw, ih = ImageReader(buf).getSize()
                    w_pt = (float(iw) * 72.0) / float(dpi_math)
                    h_pt = (float(ih) * 72.0) / float(dpi_math)
                    max_w = doc.width * (0.94 if is_math_icfes else 0.90)
                    scale = min(1.0, (max_w / w_pt)) if w_pt > 0 else 1.0
                    img = RLImage(buf, width=w_pt * scale, height=h_pt * scale)
                    if centrar_formula and not is_math_icfes:
                        img.hAlign = "CENTER"
                    else:
                        img.hAlign = "LEFT"
                    target_block.append(img)
                    return True
                except Exception:
                    return None

            if not expresiones:
                contenido = f"{prefijo}{txt}" if prefijo else txt
                target_block.append(Paragraph(contenido, style))
                return

            partes = re.split(r"(\$[^$]+\$)", txt)
            prefijo_usado = False

            for parte in partes:
                if not parte:
                    continue

                es_formula = (
                    parte.startswith("$") and parte.endswith("$") and len(parte) >= 2
                )

                if es_formula:
                    expr = parte[1:-1].strip()
                    if not expr:
                        continue
                    appended = _append_formula(expr)
                    if appended:
                        if prefijo and not prefijo_usado:
                            target_block.insert(-1, Paragraph(prefijo.strip(), style))
                            prefijo_usado = True
                    else:
                        prefijo_local = ""
                        if prefijo and not prefijo_usado:
                            prefijo_local = prefijo
                            prefijo_usado = True
                        _append_formula_as_text(expr, prefijo_local=prefijo_local)
                    continue

                chunk = str(parte).strip()
                if not chunk:
                    continue
                if prefijo and not prefijo_usado:
                    chunk = f"{prefijo}{chunk}"
                    prefijo_usado = True
                target_block.append(Paragraph(chunk, style))

            if prefijo and not prefijo_usado:
                target_block.append(Paragraph(prefijo, style))

        def _es_linea_formula_matematica(linea):
            txt_linea = str(linea or "").strip()
            if not txt_linea:
                return False
            return bool(re.fullmatch(r"\$[^$]+\$", txt_linea))

        def _separar_enunciado_formula_matematica(texto):
            txt = str(texto or "").strip()
            if not _texto_valido(txt):
                return "", ""

            try:
                expresiones = _extraer_expresiones_latex(txt)
            except ValueError:
                expresiones = []

            if expresiones:
                enunciado = re.sub(r"\$[^$]+\$", " ", txt)
                enunciado = re.sub(r"\s+", " ", enunciado).strip(" ;:,-")
                formula = " ; ".join(
                    [str(e).strip() for e in expresiones if str(e).strip()]
                )
                return enunciado, formula

            return txt, ""

        def _append_formula_centrada(
            target_block,
            expr_txt,
            espacio_antes=8,
            espacio_despues=6,
        ):
            nonlocal _formula_fallback_notificado

            expr_txt = str(expr_txt or "").strip()
            if not expr_txt:
                return

            formula_block = []
            if espacio_antes and float(espacio_antes) > 0:
                formula_block.append(Spacer(1, espacio_antes))

            rendered = renderizar_formula(expr_txt, dpi=240)
            if rendered is not None:
                buf, dpi_math = rendered
                _math_buffers.append(buf)
                try:
                    iw, ih = ImageReader(buf).getSize()
                    w_pt = (float(iw) * 72.0) / float(dpi_math)
                    h_pt = (float(ih) * 72.0) / float(dpi_math)
                    max_w = doc.width * 0.65
                    scale = min(1.0, (max_w / w_pt)) if w_pt > 0 else 1.0
                    img = RLImage(buf, width=w_pt * scale, height=h_pt * scale)
                    img.hAlign = "CENTER"
                    formula_block.append(img)
                except Exception:
                    rendered = None

            if rendered is None:
                if not _formula_fallback_notificado:
                    try:
                        print(
                            "No se pudo renderizar la fórmula. Se mostrará en formato texto."
                        )
                    except Exception:
                        pass
                    _formula_fallback_notificado = True
                texto_formula = _formula_a_texto_legible(expr_txt) or expr_txt
                formula_block.append(Paragraph(texto_formula, style_formula_center))

            if espacio_despues and float(espacio_despues) > 0:
                formula_block.append(Spacer(1, espacio_despues))

            target_block.extend(formula_block)

        def _build_question_block(row, numero_pregunta):
            bloque = []
            _append_text_with_math(
                bloque,
                row.get("enunciado", ""),
                style_bold,
                prefijo=f"{numero_pregunta}. ",
                centrar_formula=False,
            )
            for i, opt_col in enumerate(
                ["opcion_a", "opcion_b", "opcion_c", "opcion_d"]
            ):
                opt_val = row.get(opt_col, "")
                if (
                    opt_val is not None
                    and str(opt_val).strip()
                    and str(opt_val) != "nan"
                ):
                    letra = chr(65 + i)
                    _append_text_with_math(
                        bloque,
                        str(opt_val),
                        style_option,
                        prefijo=f"{letra}. ",
                        centrar_formula=False,
                    )
            bloque.append(Spacer(1, 0.4 * cm))
            return bloque

        def _build_math_question_block(
            row,
            numero_pregunta,
            incluir_contexto=False,
            contexto_texto="",
            imagen_ref="",
            responde_texto=None,
        ):
            bloque = []
            if responde_texto:
                bloque.append(Paragraph(responde_texto, style_responde))
                bloque.append(Spacer(1, 0.06 * cm))

            bloque.append(Paragraph(f"Pregunta {numero_pregunta}", style_bold))
            bloque.append(Spacer(1, 0.04 * cm))

            if incluir_contexto and _texto_valido(contexto_texto):
                _append_text_with_math(
                    bloque,
                    str(contexto_texto),
                    style_normal,
                    centrar_formula=True,
                )
                _append_group_image(bloque, imagen_ref)

            enunciado_raw = str(row.get("enunciado", "") or "")
            enunciado_txt, formula_txt = _separar_enunciado_formula_matematica(
                enunciado_raw
            )

            if _texto_valido(enunciado_txt):
                _append_text_with_math(
                    bloque,
                    enunciado_txt,
                    style_bold,
                    centrar_formula=False,
                )

            if _texto_valido(formula_txt):
                _append_formula_centrada(bloque, formula_txt)
            elif not _texto_valido(enunciado_txt):
                _append_text_with_math(
                    bloque,
                    enunciado_raw,
                    style_bold,
                    centrar_formula=False,
                )

            opciones_validas = []
            for i, opt_col in enumerate(
                ["opcion_a", "opcion_b", "opcion_c", "opcion_d"]
            ):
                opt_val = row.get(opt_col, "")
                if _texto_valido(opt_val):
                    letra = chr(65 + i)
                    opciones_validas.append((letra, str(opt_val)))

            for idx_opt, (letra, texto_opt) in enumerate(opciones_validas):
                _append_text_with_math(
                    bloque,
                    texto_opt,
                    style_option_math,
                    prefijo=f"{letra}. ",
                    centrar_formula=False,
                )
                if idx_opt < (len(opciones_validas) - 1):
                    bloque.append(Spacer(1, 1))

            bloque.append(Spacer(1, 10))
            return bloque

        def _normalizar_valor_contexto(valor):
            return re.sub(r"\s+", " ", str(valor or "").strip())

        def _build_math_context_block(
            texto_responde,
            contexto_texto="",
            imagen_ref="",
        ):
            bloque = []
            if _texto_valido(texto_responde):
                bloque.append(Paragraph(str(texto_responde), style_responde))
                bloque.append(Spacer(1, 4))

            if _texto_valido(contexto_texto):
                _append_text_with_math(
                    bloque,
                    str(contexto_texto),
                    style_normal,
                    centrar_formula=True,
                )

            if _texto_valido(imagen_ref):
                _append_group_image(bloque, imagen_ref)

            if bloque:
                bloque.append(Spacer(1, 6))
            return bloque

        def _calcular_alto_bloque_pregunta(bloque):
            # Altura total del bloque indivisible: título, enunciado,
            # fórmula (si existe), opciones y márgenes internos.
            return float(_estimate_block_height(bloque, doc.width) or 0.0)

        def _append_math_block_with_paging(target_flowables, bloque, state):
            if not bloque:
                return

            alto_bloque = _calcular_alto_bloque_pregunta(bloque)
            espacio_disponible = max(
                0.0,
                float(state["page_height"])
                - float(state["margin_bottom"])
                - float(state["y_actual"]),
            )

            if alto_bloque > espacio_disponible and state["blocks_on_page"] > 0:
                target_flowables.append(PageBreak())
                state["y_actual"] = float(state["margin_top"])
                state["blocks_on_page"] = 0

            target_flowables.append(IndivisibleBlockFlowable(bloque))
            alto_consumido = min(alto_bloque, float(state["content_height"]))
            state["y_actual"] += alto_consumido
            state["blocks_on_page"] += 1

        def _normalizar_tipo_pregunta(tipo_raw):
            tipo = _normalizar_texto_sin_tildes(tipo_raw)
            tipo = tipo.replace("-", "_").replace(" ", "_")
            tipo = re.sub(r"_+", "_", tipo).strip("_")

            if tipo in {
                "opcion_multiple",
                "seleccion_multiple",
                "multiple",
                "mcq",
                "choice",
            }:
                return "opcion_multiple"

            if tipo in {
                "abierta",
                "pregunta_abierta",
                "respuesta_abierta",
                "texto_libre",
                "desarrollo",
                "open",
                "essay",
            }:
                return "abierta"

            return ""

        def _inferir_tipo_pregunta_lenguaje(row):
            tipo = _normalizar_tipo_pregunta(row.get("tipo_pregunta", ""))
            opciones_ok = all(
                _texto_valido(row.get(col, ""))
                for col in ("opcion_a", "opcion_b", "opcion_c", "opcion_d")
            )

            if not tipo:
                return "opcion_multiple" if opciones_ok else "abierta"

            if tipo == "opcion_multiple" and not opciones_ok:
                return "abierta"

            return tipo

        def _lineas_respuesta_abierta(num_lineas=4):
            lineas = []
            for _ in range(max(3, int(num_lineas))):
                lineas.append(
                    Paragraph(
                        "____________________________________________________________",
                        style_normal,
                    )
                )
                lineas.append(Spacer(1, 0.08 * cm))
            return lineas

        def _build_language_question_block(row, numero_pregunta):
            bloque = []
            tipo_pregunta = _inferir_tipo_pregunta_lenguaje(row)

            bloque.append(Paragraph(f"Pregunta {numero_pregunta}", style_bold))
            bloque.append(Spacer(1, 0.08 * cm))

            _append_text_with_math(
                bloque,
                row.get("enunciado", ""),
                style_normal,
                centrar_formula=False,
            )
            bloque.append(Spacer(1, 0.12 * cm))

            if tipo_pregunta == "opcion_multiple":
                for i, opt_col in enumerate(
                    ["opcion_a", "opcion_b", "opcion_c", "opcion_d"]
                ):
                    letra = chr(65 + i)
                    _append_text_with_math(
                        bloque,
                        row.get(opt_col, ""),
                        style_option,
                        prefijo=f"{letra}. ",
                        centrar_formula=False,
                    )
            else:
                bloque.extend(_lineas_respuesta_abierta(num_lineas=4))

            bloque.append(Spacer(1, 0.28 * cm))
            return bloque

        def _estimate_block_height(block, avail_width):
            total_h = 0.0
            for flowable in block:
                try:
                    if flowable.__class__.__name__ == "KeepTogether":
                        inner = getattr(flowable, "_content", None) or getattr(
                            flowable, "_flowables", []
                        )
                        if inner:
                            total_h += _estimate_block_height(inner, avail_width)
                            continue
                    _w, h = flowable.wrap(avail_width, doc.height)
                    total_h += max(0.0, float(h or 0.0))
                except Exception:
                    continue
            return total_h

        def _compose_balanced_math_flowables(question_blocks, intro_height=0.0):
            if not question_blocks:
                return []
            first_page_capacity = max(
                1.0, float(doc.height) - float(intro_height or 0.0)
            )
            second_page_capacity = float(doc.height)
            heights = [
                float(_estimate_block_height(bloque, doc.width) or 0.0)
                for bloque in question_blocks
            ]

            def _as_flowables(split_index=None):
                result = []
                for idx, bloque in enumerate(question_blocks):
                    if split_index is not None and idx == split_index:
                        result.append(PageBreak())
                    result.append(KeepTogether(bloque))
                return result

            total_height = sum(heights)
            total_blocks = len(question_blocks)

            # Solo forzamos un reparto cuando realmente es factible que todo el
            # contenido quede en dos hojas. En cualquier otro caso dejamos que
            # Platypus use el espacio real disponible y mueva bloques completos.
            if total_blocks >= 2 and total_height <= (
                first_page_capacity + second_page_capacity
            ):
                acumulado = 0.0
                candidatos = []
                for idx, block_height in enumerate(heights[:-1], start=1):
                    acumulado += block_height
                    restante = total_height - acumulado
                    if (
                        acumulado <= first_page_capacity
                        and restante <= second_page_capacity
                    ):
                        diferencia = abs((idx) - (total_blocks - idx))
                        candidatos.append((diferencia, -idx, idx))

                if candidatos:
                    candidatos.sort()
                    return _as_flowables(split_index=candidatos[0][2])

            return _as_flowables()

        if is_math_icfes:
            math_layout_state = {
                "page_height": float(height),
                "margin_top": float(math_margin_top),
                "margin_bottom": float(math_margin_bottom),
                "content_height": max(
                    1.0,
                    float(height) - float(math_margin_top) - float(math_margin_bottom),
                ),
                "y_actual": float(math_margin_top),
                "blocks_on_page": 0,
            }

            math_intro_block = [
                Spacer(1, intro_top_spacing),
                Paragraph(instruccion_lectura, style_instruction),
                Spacer(1, 6),
            ]
            _append_math_block_with_paging(
                flowables, math_intro_block, math_layout_state
            )

            if "id_contexto" in preguntas_all.columns:
                for _context_id, grupo_preg in grupos:
                    if cantidad is not None and pregunta_num > cantidad:
                        break

                    filas_grupo = list(grupo_preg.iterrows())
                    if cantidad is not None and cantidad > 0:
                        restantes = int(cantidad) - pregunta_num + 1
                        if restantes <= 0:
                            break
                        filas_grupo = filas_grupo[:restantes]

                    if not filas_grupo:
                        continue

                    contexto_texto = grupo_preg.iloc[0].get("contexto", "")
                    imagen_ref = grupo_preg.iloc[0].get("imagen", "")

                    contexto_compartido = True
                    contexto_base = _normalizar_valor_contexto(contexto_texto)
                    imagen_base = _normalizar_valor_contexto(imagen_ref)
                    for _idx_row, row_data in filas_grupo:
                        if (
                            _normalizar_valor_contexto(row_data.get("contexto", ""))
                            != contexto_base
                            or _normalizar_valor_contexto(row_data.get("imagen", ""))
                            != imagen_base
                        ):
                            contexto_compartido = False
                            break

                    if contexto_compartido and (
                        _texto_valido(contexto_texto) or _texto_valido(imagen_ref)
                    ):
                        start_num = pregunta_num
                        end_num = pregunta_num + len(filas_grupo) - 1
                        texto_responde = None
                        if len(filas_grupo) >= 2:
                            texto_responde = (
                                "RESPONDA LAS PREGUNTAS "
                                f"{start_num} A {end_num} "
                                "DE ACUERDO CON LA SIGUIENTE INFORMACIÓN"
                            )

                        bloque_contexto = _build_math_context_block(
                            texto_responde=texto_responde,
                            contexto_texto=contexto_texto,
                            imagen_ref=imagen_ref,
                        )
                        _append_math_block_with_paging(
                            flowables,
                            bloque_contexto,
                            math_layout_state,
                        )

                    for idx_fila, (_row_idx, row) in enumerate(filas_grupo):
                        contexto_bloque = row.get("contexto", contexto_texto)
                        imagen_bloque = row.get("imagen", imagen_ref)

                        incluir_contexto = False
                        responde_texto = None
                        if not contexto_compartido:
                            incluir_contexto = _texto_valido(
                                contexto_bloque
                            ) or _texto_valido(imagen_bloque)
                            if (
                                incluir_contexto
                                and idx_fila == 0
                                and len(filas_grupo) >= 2
                            ):
                                start_num = pregunta_num
                                end_num = pregunta_num + len(filas_grupo) - 1
                                responde_texto = (
                                    "RESPONDA LAS PREGUNTAS "
                                    f"{start_num} A {end_num} "
                                    "DE ACUERDO CON LA SIGUIENTE INFORMACIÓN"
                                )

                        bloque = _build_math_question_block(
                            row,
                            pregunta_num,
                            incluir_contexto=incluir_contexto,
                            contexto_texto=contexto_bloque,
                            imagen_ref=imagen_bloque,
                            responde_texto=responde_texto,
                        )
                        _append_math_block_with_paging(
                            flowables, bloque, math_layout_state
                        )
                        pregunta_num += 1
            else:
                for _, row in grupos[0][1].iterrows():
                    if cantidad is not None and pregunta_num > cantidad:
                        break

                    bloque = _build_math_question_block(row, pregunta_num)
                    _append_math_block_with_paging(flowables, bloque, math_layout_state)
                    pregunta_num += 1
        elif is_language_icfes:
            grupos_lenguaje = grupos
            # Ancho efectivo de una columna del template de Lenguaje.
            lang_col_width = max(1.0, float(doc.width - 20.0) / 2.0)

            # Estilos para Lenguaje (11.5 pt / leading 13)
            _FS = 11.5
            _LD = 13
            style_lang_text = ParagraphStyle(
                "LangText",
                parent=style_normal,
                fontSize=_FS,
                leading=_LD,
                spaceAfter=0,
                alignment=TA_JUSTIFY,
            )
            style_lang_q = ParagraphStyle(
                "LangQ",
                parent=style_normal,
                fontSize=_FS,
                leading=_LD,
                spaceAfter=0,
            )
            style_lang_q_bold = ParagraphStyle(
                "LangQBold",
                parent=style_bold,
                fontSize=_FS,
                leading=_LD,
                spaceAfter=0,
            )
            style_lang_opt = ParagraphStyle(
                "LangOpt",
                parent=style_option,
                fontSize=_FS,
                leading=_LD,
                spaceAfter=0,
                leftIndent=0.1 * cm,
            )
            style_lang_instr_group = ParagraphStyle(
                "LangInstrGroup",
                parent=style_lang_text,
                fontName=f"{font_family}-Bold",
                fontSize=8,
                leading=9,
                spaceAfter=0,
            )

            class _HRuleLang(Flowable):
                def wrap(self, aw, ah):
                    return aw, 1

                def draw(self):
                    self.canv.setStrokeColor(colors.HexColor("#888888"))
                    self.canv.setLineWidth(0.5)
                    self.canv.line(0, 0, self._frame._width, 0)

            def _append_lang_guard_page_break(target_flowables, bloque_estimado):
                """Evita iniciar un bloque de texto incompleto al final de columna.

                Si no hay espacio suficiente para el inicio del bloque
                (TEXTO + RESPONDA + contexto), fuerza salto a la siguiente hoja.
                """
                alto_estimado = float(
                    _estimate_block_height(bloque_estimado, lang_col_width) or 0.0
                )
                if alto_estimado <= 0.0:
                    return

                # Limita la guardia para evitar bucles cuando el contexto es muy extenso.
                alto_guardia = min(alto_estimado, max(80.0, float(doc.height) * 0.92))
                target_flowables.append(CondPageBreak(alto_guardia))

            def _build_lang_q_block(row, numero):
                """Bloque indivisible de pregunta: selección múltiple o abierta."""
                b = []
                b.append(Paragraph(f"<b>Pregunta {numero}.</b>", style_lang_q_bold))
                b.append(Spacer(1, 3))
                enunciado = str(row.get("enunciado", "") or "")
                if _texto_valido(enunciado):
                    _append_text_with_math(b, enunciado, style_lang_q)
                tipo_p = _inferir_tipo_pregunta_lenguaje(row)
                if tipo_p == "opcion_multiple":
                    b.append(Spacer(1, 4))
                    for i, col in enumerate(
                        ["opcion_a", "opcion_b", "opcion_c", "opcion_d"]
                    ):
                        val = row.get(col, "")
                        if _texto_valido(val):
                            _append_text_with_math(
                                b,
                                str(val),
                                style_lang_opt,
                                prefijo=f"{chr(65 + i)}. ",
                            )
                            b.append(Spacer(1, 1))
                else:
                    # Pregunta abierta: líneas de respuesta visual
                    b.append(Spacer(1, 8))
                    b.extend(_lineas_respuesta_abierta(num_lineas=4))
                b.append(Spacer(1, 10))
                return b

            # Instrucción inicial del examen
            flowables.append(Spacer(1, intro_top_spacing))
            flowables.append(Paragraph(instruccion_lectura, style_instruction))
            flowables.append(Spacer(1, 8))

            for indice_texto, (_context_id, grupo_preg) in enumerate(grupos_lenguaje):
                if cantidad is not None and pregunta_num > cantidad:
                    break

                filas_grupo = list(grupo_preg.iterrows())
                if cantidad is not None and cantidad > 0:
                    restantes = int(cantidad) - pregunta_num + 1
                    if restantes <= 0:
                        break
                    filas_grupo = filas_grupo[:restantes]

                if not filas_grupo:
                    continue

                fila_cabecera = grupo_preg.iloc[0]
                contexto_texto = fila_cabecera.get("contexto", "")
                imagen_ref = fila_cabecera.get("imagen", "")

                titulo_texto = ""
                for col_titulo in (
                    "titulo_texto",
                    "titulo",
                    "titulo_lectura",
                    "nombre_texto",
                ):
                    if col_titulo in grupo_preg.columns and _texto_valido(
                        fila_cabecera.get(col_titulo, "")
                    ):
                        titulo_texto = str(fila_cabecera.get(col_titulo, "")).strip()
                        break
                if not titulo_texto:
                    titulo_texto = f"TEXTO {indice_texto + 1}"

                start_num = pregunta_num
                end_num = pregunta_num + len(filas_grupo) - 1
                instr_grupo = (
                    f"RESPONDA LAS PREGUNTAS {start_num} A {end_num} "
                    "DE ACUERDO CON EL SIGUIENTE TEXTO"
                )

                titulo_para = Paragraph(f"<b>{titulo_texto}</b>", style_lang_q_bold)
                instr_para = Paragraph(instr_grupo, style_lang_instr_group)

                # Párrafos del texto de lectura (pueden dividirse entre columnas)
                text_paragraphs = []
                _append_group_image(text_paragraphs, imagen_ref)
                if _texto_valido(contexto_texto):
                    partes_ctx = [
                        p.strip() for p in str(contexto_texto).split("\n") if p.strip()
                    ]
                    for parte_ctx in partes_ctx:
                        text_paragraphs.append(Paragraph(parte_ctx, style_lang_text))
                        text_paragraphs.append(Spacer(1, 3))
                else:
                    text_paragraphs.append(
                        Paragraph("Texto de lectura no disponible.", style_lang_text)
                    )

                separador_grupo = []
                if indice_texto > 0:
                    separador_grupo = [Spacer(1, 6), _HRuleLang(), Spacer(1, 6)]

                bloque_estimado_texto = (
                    list(separador_grupo)
                    + [
                        titulo_para,
                        Spacer(1, 4),
                        instr_para,
                        Spacer(1, 6),
                    ]
                    + list(text_paragraphs)
                )
                _append_lang_guard_page_break(flowables, bloque_estimado_texto)

                for item_sep in separador_grupo:
                    flowables.append(item_sep)

                # Título + instrucción anclados al primer párrafo del texto
                # para que nunca queden huérfanos al final de una columna
                encabezado_grupo = [
                    titulo_para,
                    Spacer(1, 4),
                    instr_para,
                    Spacer(1, 6),
                ]
                if text_paragraphs:
                    flowables.append(
                        KeepTogether(encabezado_grupo + [text_paragraphs[0]])
                    )
                    for tp in text_paragraphs[1:]:
                        flowables.append(tp)
                else:
                    flowables.append(KeepTogether(encabezado_grupo))

                # Preguntas del grupo: cada una como bloque indivisible
                for _idx, row in filas_grupo:
                    bloque = _build_lang_q_block(row, pregunta_num)
                    flowables.append(IndivisibleBlockFlowable(bloque))
                    pregunta_num += 1
        else:
            # Para exactamente 4 textos se conserva una distribución fija de 2 hojas.
            # Para más de 4 textos se usa flujo natural en columnas/páginas.
            if "id_contexto" in preguntas_all.columns:
                total_textos = len(grupos)
                usar_distribucion_fija = total_textos == 4

                if usar_distribucion_fija:
                    columnas_por_hoja = {}

                for indice_texto, (context_id, grupo_preg) in enumerate(grupos):
                    encabezado_texto = []
                    if indice_texto == 0:
                        encabezado_texto.append(Spacer(1, intro_top_spacing))
                        encabezado_texto.append(
                            Paragraph(instruccion_lectura, style_instruction)
                        )
                        encabezado_texto.append(Spacer(1, 0.2 * cm))

                    contexto_texto = grupo_preg.iloc[0].get("contexto", "")
                    if _texto_valido(contexto_texto):
                        contexto_clave = str(contexto_texto).strip()
                    else:
                        # fallback por grupo para conservar estabilidad si falta texto
                        contexto_clave = str(context_id).strip()

                    if contexto_clave not in contextos_usados:
                        contextos_usados[contexto_clave] = contador_texto
                        contador_texto += 1

                    context_label = contextos_usados[contexto_clave]
                    encabezado_texto.append(
                        Paragraph(f"TEXTO {context_label}", style_bold)
                    )

                    # imagen bajo el encabezado si existe
                    imagen_ref = grupo_preg.iloc[0].get("imagen", "")
                    _append_group_image(encabezado_texto, imagen_ref)

                    # texto de contexto
                    if _texto_valido(contexto_texto):
                        _append_text_with_math(
                            encabezado_texto,
                            str(contexto_texto),
                            style_normal,
                            centrar_formula=False,
                        )

                    start_num = pregunta_num
                    end_num = pregunta_num + len(grupo_preg) - 1
                    encabezado_texto.append(Spacer(1, 10))
                    encabezado_texto.append(
                        Paragraph(
                            f"RESPONDE LAS PREGUNTAS {start_num} A {end_num} SEGÚN EL TEXTO {context_label}",
                            style_responde,
                        )
                    )

                    preguntas_bloque = []

                    for _, row in grupo_preg.iterrows():
                        preguntas_bloque.append(
                            _build_question_block(row, pregunta_num)
                        )
                        pregunta_num += 1

                    if usar_distribucion_fija:
                        bloque_texto = list(encabezado_texto)
                        for pregunta_bloque in preguntas_bloque:
                            bloque_texto.extend(pregunta_bloque)
                        flowables.append(KeepTogether(bloque_texto))

                        hoja = indice_texto // 2
                        columna = indice_texto % 2
                        columnas_por_hoja.setdefault(hoja, set()).add(columna)

                        try:
                            col_txt = "izquierda" if columna == 0 else "derecha"
                            print(
                                f"[DEBUG] Texto {indice_texto + 1} -> hoja {hoja + 1}, columna {col_txt}"
                            )
                        except Exception:
                            pass

                        if indice_texto < total_textos - 1:
                            if columna == 0:
                                flowables.append(FrameBreak())
                            else:
                                flowables.append(PageBreak())
                    else:
                        # Bloque inicial indivisible: título + contenido + instrucción.
                        # No se agrupa con la primera pregunta para evitar saltos prematuros.
                        bloque_inicial_texto = list(encabezado_texto)
                        flowables.append(KeepTogether(bloque_inicial_texto))

                        # Cada pregunta mantiene su bloque indivisible (pregunta + opciones).
                        for pregunta_bloque in preguntas_bloque:
                            flowables.append(KeepTogether(pregunta_bloque))

                if usar_distribucion_fija:
                    total_hojas = (total_textos + 1) // 2
                    for hoja in range(total_hojas):
                        textos_disponibles = total_textos - (hoja * 2)
                        esperadas = {0, 1} if textos_disponibles >= 2 else {0}
                        ocupadas = columnas_por_hoja.get(hoja, set())
                        if ocupadas != esperadas:
                            raise ValueError(
                                "Distribución inválida de textos en PDF: "
                                f"hoja {hoja + 1}, columnas esperadas={sorted(esperadas)}, "
                                f"columnas ocupadas={sorted(ocupadas)}"
                            )
            else:
                flowables.append(Spacer(1, intro_top_spacing))
                flowables.append(Paragraph(instruccion_lectura, style_instruction))
                for _, row in grupos[0][1].iterrows():
                    if cantidad is not None and pregunta_num > cantidad:
                        break
                    flowables.append(
                        KeepTogether(_build_question_block(row, pregunta_num))
                    )
                    pregunta_num += 1

        if integrar_hoja_ultima_pagina:
            flowables.append(AnswerSheetAnchorFlowable())

        doc.build(flowables, canvasmaker=NumberedCanvas)
        return exam_id

    def maestro_modificar_nota(self):
        sel = self._maestro_get_selected_row()
        if not sel:
            return
        _, vals, documento = sel
        area_sel = vals[6] if len(vals) > 6 else None

        nuevo = simpledialog.askfloat(
            "Modificar nota", "Nueva nota:", parent=self.win, minvalue=0.0
        )
        if nuevo is None:
            return

        if area_sel:
            self.cur.execute(
                "UPDATE resultados SET nota=? WHERE id = (SELECT id FROM resultados WHERE documento = ? AND area = ? ORDER BY id DESC LIMIT 1)",
                (nuevo, documento, area_sel),
            )
        else:
            self.cur.execute(
                "UPDATE resultados SET nota=? WHERE id = (SELECT id FROM resultados WHERE documento = ? ORDER BY id DESC LIMIT 1)",
                (nuevo, documento),
            )
        self.conn.commit()
        self.maestro_cargar_resultados()

    def maestro_intervenir_intento(self):
        sel = self._maestro_get_selected_row()
        if not sel:
            return
        _, vals, documento = sel
        area_sel = vals[6] if len(vals) > 6 else None

        nuevo = simpledialog.askinteger(
            "Intervenir intento",
            "Nuevo número de intento:",
            parent=self.win,
            minvalue=0,
        )
        if nuevo is None:
            return

        if area_sel:
            self.cur.execute(
                "UPDATE resultados SET intento=? WHERE id = (SELECT id FROM resultados WHERE documento = ? AND area = ? ORDER BY id DESC LIMIT 1)",
                (nuevo, documento, area_sel),
            )
        else:
            self.cur.execute(
                "UPDATE resultados SET intento=? WHERE id = (SELECT id FROM resultados WHERE documento = ? ORDER BY id DESC LIMIT 1)",
                (nuevo, documento),
            )
        self.conn.commit()
        self.maestro_cargar_resultados()

    # ---------- Configuración de Plantel ----------
    def _ensure_configuracion_plantel(self):
        """Asegura que exista la tabla de configuración de plantel."""
        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS configuracion_plantel (
                   clave TEXT PRIMARY KEY,
                   valor TEXT
               )"""
        )
        self.conn.commit()
        # tabla para escala de valoración (si no existe)
        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS escala_valoracion (
                   id INTEGER PRIMARY KEY AUTOINCREMENT,
                   desde REAL NOT NULL,
                   hasta REAL NOT NULL,
                   letra TEXT,
                   concepto TEXT,
                   desempeno TEXT,
                   recomendacion TEXT
               )"""
        )
        self.conn.commit()

    def _get_config_plantel(self, clave):
        """Obtiene un valor de configuración del plantel."""
        self._ensure_configuracion_plantel()
        self.cur.execute(
            "SELECT valor FROM configuracion_plantel WHERE clave=?", (clave,)
        )
        row = self.cur.fetchone()
        return row[0] if row else ""

    def _set_config_plantel(self, clave, valor):
        """Establece un valor de configuración del plantel."""
        self._ensure_configuracion_plantel()
        self.cur.execute(
            "REPLACE INTO configuracion_plantel(clave,valor) VALUES(?,?)",
            (clave, valor),
        )
        self.conn.commit()

    def _config_plantel_cargar_sedes(self):
        """Carga la configuración de sedes educativas desde la configuración del plantel."""
        sedes_guardadas = []
        raw = self._get_config_plantel("sedes_educativas") or ""
        if raw:
            try:
                data = json.loads(raw)
                if isinstance(data, list):
                    for item in data:
                        if not isinstance(item, dict):
                            continue
                        jornadas = item.get("jornadas") or {}
                        sedes_guardadas.append(
                            {
                                "nombre": str(item.get("nombre", "")).strip(),
                                "manana": bool(jornadas.get("manana", False)),
                                "tarde": bool(jornadas.get("tarde", False)),
                                "noche": bool(jornadas.get("noche", False)),
                            }
                        )
            except Exception:
                sedes_guardadas = []

        if not sedes_guardadas:
            jornadas_prev = self._get_config_plantel("jornadas") or ""
            jparts = [p.strip().lower() for p in jornadas_prev.split(",") if p.strip()]
            if jornadas_prev.strip():
                sedes_guardadas.append(
                    {
                        "nombre": "Sede Principal",
                        "manana": "mañana" in jparts or "manana" in jparts,
                        "tarde": "tarde" in jparts,
                        "noche": "nocturna" in jparts or "noche" in jparts,
                    }
                )

        return sedes_guardadas or [
            {"nombre": "", "manana": False, "tarde": False, "noche": False}
        ]

    def _config_plantel_render_sedes(self):
        """Renderiza la tabla editable de sedes educativas."""
        if not hasattr(self, "sedes_body_frame"):
            return

        for widget in self.sedes_body_frame.winfo_children():
            widget.destroy()

        headers = ["No.", "Nombre de la sede", "Mañana", "Tarde", "Noche", "Acción"]
        for column, header in enumerate(headers):
            ttk.Label(
                self.sedes_body_frame,
                text=header,
                font=("Arial", 9, "bold"),
            ).grid(row=0, column=column, padx=4, pady=2, sticky="w")
            self.sedes_body_frame.grid_columnconfigure(
                column, weight=1 if column == 1 else 0
            )

        if hasattr(self, "sedes_summary_label"):
            total = len(getattr(self, "sedes_data", []) or [])
            texto = (
                "1 sede configurada" if total == 1 else f"{total} sedes configuradas"
            )
            self.sedes_summary_label.config(text=texto)

        self.sedes_rows = []
        for index, sede in enumerate(getattr(self, "sedes_data", []), start=1):
            numero_label = ttk.Label(self.sedes_body_frame, text=str(index))
            numero_label.grid(row=index, column=0, padx=4, pady=2, sticky="w")

            nombre_var = tk.StringVar(value=str(sede.get("nombre", "")).strip())
            nombre_entry = ttk.Entry(
                self.sedes_body_frame, textvariable=nombre_var, width=44
            )
            nombre_entry.grid(row=index, column=1, padx=4, pady=2, sticky="ew")

            manana_var = tk.BooleanVar(value=bool(sede.get("manana", False)))
            tarde_var = tk.BooleanVar(value=bool(sede.get("tarde", False)))
            noche_var = tk.BooleanVar(value=bool(sede.get("noche", False)))

            ttk.Checkbutton(self.sedes_body_frame, variable=manana_var).grid(
                row=index, column=2, padx=4, pady=2
            )
            ttk.Checkbutton(self.sedes_body_frame, variable=tarde_var).grid(
                row=index, column=3, padx=4, pady=2
            )
            ttk.Checkbutton(self.sedes_body_frame, variable=noche_var).grid(
                row=index, column=4, padx=4, pady=2
            )
            ttk.Button(
                self.sedes_body_frame,
                text="Quitar",
                command=lambda row_index=index - 1: self._config_plantel_eliminar_sede(
                    row_index
                ),
            ).grid(row=index, column=5, padx=4, pady=2, sticky="e")

            self.sedes_rows.append(
                {
                    "nombre": nombre_var,
                    "manana": manana_var,
                    "tarde": tarde_var,
                    "noche": noche_var,
                }
            )

    def _config_plantel_agregar_sede(self):
        """Agrega una nueva fila de sede educativa."""
        if not hasattr(self, "sedes_data"):
            self.sedes_data = []
        self.sedes_data.append(
            {"nombre": "", "manana": False, "tarde": False, "noche": False}
        )
        self._config_plantel_render_sedes()

    def _config_plantel_eliminar_sede(self, index):
        """Elimina una sede educativa de la tabla editable."""
        if not hasattr(self, "sedes_data"):
            self.sedes_data = []
        if index < 0 or index >= len(self.sedes_data):
            return
        if len(self.sedes_data) <= 1:
            self.sedes_data = [
                {"nombre": "", "manana": False, "tarde": False, "noche": False}
            ]
        else:
            del self.sedes_data[index]
        self._config_plantel_render_sedes()

    def _config_plantel_obtener_sedes(self):
        """Recolecta las sedes configuradas actualmente en la tabla."""
        sedes = []
        for row in getattr(self, "sedes_rows", []):
            nombre = row["nombre"].get().strip()
            manana = bool(row["manana"].get())
            tarde = bool(row["tarde"].get())
            noche = bool(row["noche"].get())
            if nombre or manana or tarde or noche:
                sedes.append(
                    {
                        "nombre": nombre,
                        "jornadas": {
                            "manana": manana,
                            "tarde": tarde,
                            "noche": noche,
                        },
                    }
                )
        return sedes

    # ---------- Contenido Institucional ----------
    def _ensure_contenido_institucional(self):
        """Asegura la tabla para el carrusel del login."""
        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS contenido_institucional (
                   id INTEGER PRIMARY KEY AUTOINCREMENT,
                   titulo TEXT,
                   mensaje TEXT NOT NULL,
                   ruta_imagen TEXT,
                   estado TEXT NOT NULL DEFAULT 'Activo',
                   orden INTEGER NOT NULL DEFAULT 1,
                   fecha_creacion TEXT NOT NULL
               )"""
        )
        self.conn.commit()

    def _listar_contenido_institucional(self, solo_activos=False):
        self._ensure_contenido_institucional()
        sql = (
            "SELECT id, titulo, mensaje, ruta_imagen, estado, orden, fecha_creacion "
            "FROM contenido_institucional"
        )
        params = []
        if solo_activos:
            sql += " WHERE LOWER(COALESCE(estado, 'Activo')) = ?"
            params.append("activo")
        sql += " ORDER BY COALESCE(orden, 0) ASC, id ASC"
        rows = self.cur.execute(sql, params).fetchall()
        return [
            {
                "id": row[0],
                "titulo": row[1] or "",
                "mensaje": row[2] or "",
                "ruta_imagen": row[3] or "",
                "estado": row[4] or "Activo",
                "orden": row[5] if row[5] is not None else 0,
                "fecha_creacion": row[6] or "",
            }
            for row in rows
        ]

    def _obtener_contenido_institucional(self, contenido_id):
        self._ensure_contenido_institucional()
        row = self.cur.execute(
            "SELECT id, titulo, mensaje, ruta_imagen, estado, orden, fecha_creacion "
            "FROM contenido_institucional WHERE id=?",
            (contenido_id,),
        ).fetchone()
        if not row:
            return None
        return {
            "id": row[0],
            "titulo": row[1] or "",
            "mensaje": row[2] or "",
            "ruta_imagen": row[3] or "",
            "estado": row[4] or "Activo",
            "orden": row[5] if row[5] is not None else 0,
            "fecha_creacion": row[6] or "",
        }

    def _guardar_contenido_institucional(self, payload, contenido_id=None):
        self._ensure_contenido_institucional()
        titulo = str(payload.get("titulo", "") or "").strip()
        mensaje = str(payload.get("mensaje", "") or "").strip()
        ruta_imagen = str(payload.get("ruta_imagen", "") or "").strip()
        estado = (
            "Activo"
            if str(payload.get("estado", "Activo")).strip().lower() == "activo"
            else "Inactivo"
        )
        try:
            orden = int(str(payload.get("orden", 1) or 1).strip())
        except Exception:
            orden = 1
        orden = max(0, orden)
        fecha_creacion = datetime.now().isoformat(timespec="seconds")

        if contenido_id:
            self.cur.execute(
                "UPDATE contenido_institucional SET titulo=?, mensaje=?, ruta_imagen=?, estado=?, orden=? WHERE id=?",
                (titulo, mensaje, ruta_imagen, estado, orden, contenido_id),
            )
            self.conn.commit()
            return contenido_id

        self.cur.execute(
            "INSERT INTO contenido_institucional (titulo, mensaje, ruta_imagen, estado, orden, fecha_creacion) VALUES (?, ?, ?, ?, ?, ?)",
            (titulo, mensaje, ruta_imagen, estado, orden, fecha_creacion),
        )
        self.conn.commit()
        return self.cur.lastrowid

    def _eliminar_contenido_institucional(self, contenido_id):
        self._ensure_contenido_institucional()
        self.cur.execute(
            "DELETE FROM contenido_institucional WHERE id=?", (contenido_id,)
        )
        self.conn.commit()

    def _contenido_institucional_siguiente_orden(self):
        self._ensure_contenido_institucional()
        row = self.cur.execute(
            "SELECT COALESCE(MAX(orden), 0) + 1 FROM contenido_institucional"
        ).fetchone()
        return int(row[0] or 1)

    def _contenido_institucional_resolver_ruta(self, ruta_ref):
        txt = str(ruta_ref or "").strip()
        if not txt:
            return ""
        if os.path.isabs(txt):
            return txt
        return os.path.join(self.base_dir, txt)

    def _contenido_institucional_actualizar_preview(self):
        if not hasattr(self, "lbl_contenido_preview"):
            return
        ruta_ref = self.var_contenido_ruta_imagen.get().strip()
        self.lbl_contenido_preview.configure(image="", text="Sin imagen seleccionada")
        self.lbl_contenido_preview.image = None
        if not ruta_ref:
            return
        ruta_abs = self._contenido_institucional_resolver_ruta(ruta_ref)
        if _HAS_PIL and os.path.isfile(ruta_abs):
            try:
                img = Image.open(ruta_abs)
                img.thumbnail((160, 90))
                tkimg = ImageTk.PhotoImage(img)
                self.lbl_contenido_preview.configure(image=tkimg, text="")
                self.lbl_contenido_preview.image = tkimg
                return
            except Exception:
                pass
        self.lbl_contenido_preview.configure(text=os.path.basename(ruta_ref))

    def _contenido_institucional_seleccionar_imagen(self):
        archivo = filedialog.askopenfilename(
            parent=self.win,
            title="Seleccionar imagen para contenido institucional",
            filetypes=[
                ("Imágenes", "*.png *.jpg *.jpeg *.bmp *.gif *.webp"),
                ("Todos", "*.*"),
            ],
        )
        if not archivo:
            return
        try:
            extension = os.path.splitext(archivo)[1].lower() or ".png"
            nombre = f"contenido_institucional_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}{extension}"
            destino = os.path.join(self.imagenes_dir, nombre)
            shutil.copy2(archivo, destino)
            self.var_contenido_ruta_imagen.set(os.path.relpath(destino, self.base_dir))
            self._contenido_institucional_actualizar_preview()
        except Exception as e:
            messagebox.showerror(
                "Contenido Institucional",
                f"No se pudo copiar la imagen seleccionada.\n{e}",
                parent=self.win,
            )

    def _contenido_institucional_limpiar_form(self):
        self.contenido_institucional_id_actual = None
        self.var_contenido_titulo.set("")
        self.txt_contenido_mensaje.delete("1.0", "end")
        self.var_contenido_ruta_imagen.set("")
        self.var_contenido_estado.set("Activo")
        self.var_contenido_orden.set(
            str(self._contenido_institucional_siguiente_orden())
        )
        self.lbl_contenido_form_estado.configure(text="Nuevo contenido")
        if (
            hasattr(self, "btn_contenido_eliminar")
            and self.btn_contenido_eliminar is not None
        ):
            self.btn_contenido_eliminar.state(["disabled"])
        if hasattr(self, "tree_contenido_institucional"):
            self.tree_contenido_institucional.selection_remove(
                self.tree_contenido_institucional.selection()
            )
        self._contenido_institucional_actualizar_preview()

    def _contenido_institucional_cargar_form(self, data):
        self.contenido_institucional_id_actual = data["id"]
        self.var_contenido_titulo.set(data.get("titulo", ""))
        self.txt_contenido_mensaje.delete("1.0", "end")
        self.txt_contenido_mensaje.insert("1.0", data.get("mensaje", ""))
        self.var_contenido_ruta_imagen.set(data.get("ruta_imagen", ""))
        self.var_contenido_estado.set(data.get("estado", "Activo") or "Activo")
        self.var_contenido_orden.set(str(data.get("orden", 0) or 0))
        self.lbl_contenido_form_estado.configure(
            text=f"Editando registro #{data['id']}"
        )
        if (
            hasattr(self, "btn_contenido_eliminar")
            and self.btn_contenido_eliminar is not None
        ):
            self.btn_contenido_eliminar.state(["!disabled"])
        self._contenido_institucional_actualizar_preview()

    def _contenido_institucional_recargar_tree(self, seleccionar_id=None):
        if not hasattr(self, "tree_contenido_institucional"):
            return
        tree = self.tree_contenido_institucional
        for item in tree.get_children():
            tree.delete(item)

        registros = self._listar_contenido_institucional(solo_activos=False)
        for registro in registros:
            titulo = registro["titulo"] or "(Sin título)"
            mensaje = " ".join(registro["mensaje"].split())
            if len(mensaje) > 70:
                mensaje = mensaje[:67].rstrip() + "..."
            tree.insert(
                "",
                "end",
                iid=str(registro["id"]),
                values=(
                    registro["orden"],
                    registro["estado"],
                    titulo,
                    "Sí" if registro["ruta_imagen"] else "No",
                    mensaje,
                ),
            )

        self.lbl_contenido_total.configure(text=f"Registros: {len(registros)}")
        if seleccionar_id is not None and tree.exists(str(seleccionar_id)):
            tree.selection_set(str(seleccionar_id))
            tree.focus(str(seleccionar_id))
            tree.see(str(seleccionar_id))

    def _contenido_institucional_on_select(self, _event=None):
        seleccion = getattr(self, "tree_contenido_institucional", None)
        if seleccion is None:
            return
        ids = seleccion.selection()
        if not ids:
            return
        data = self._obtener_contenido_institucional(int(ids[0]))
        if data:
            self._contenido_institucional_cargar_form(data)

    def _contenido_institucional_guardar_desde_form(self):
        if not self._requiere_permiso(
            "desktop.superadmin.contenido_institucional.guardar"
        ):
            return
        mensaje = self.txt_contenido_mensaje.get("1.0", "end").strip()
        if not mensaje:
            messagebox.showerror(
                "Contenido Institucional",
                "El mensaje es obligatorio.",
                parent=self.win,
            )
            return
        try:
            orden = int(str(self.var_contenido_orden.get() or "0").strip())
        except Exception:
            messagebox.showerror(
                "Contenido Institucional",
                "El orden debe ser un número entero.",
                parent=self.win,
            )
            return

        payload = {
            "titulo": self.var_contenido_titulo.get().strip(),
            "mensaje": mensaje,
            "ruta_imagen": self.var_contenido_ruta_imagen.get().strip(),
            "estado": self.var_contenido_estado.get().strip() or "Activo",
            "orden": orden,
        }
        contenido_id = self._guardar_contenido_institucional(
            payload,
            contenido_id=self.contenido_institucional_id_actual,
        )
        self._contenido_institucional_recargar_tree(seleccionar_id=contenido_id)
        self._contenido_institucional_cargar_form(
            self._obtener_contenido_institucional(contenido_id)
        )
        messagebox.showinfo(
            "Contenido Institucional",
            "Contenido guardado correctamente.",
            parent=self.win,
        )

    def _contenido_institucional_eliminar_desde_form(self):
        if not self._requiere_permiso(
            "desktop.superadmin.contenido_institucional.eliminar"
        ):
            return
        if not self.contenido_institucional_id_actual:
            messagebox.showwarning(
                "Contenido Institucional",
                "Seleccione un registro para eliminar.",
                parent=self.win,
            )
            return
        confirmar = messagebox.askyesno(
            "Contenido Institucional",
            "¿Desea eliminar el contenido seleccionado?",
            parent=self.win,
        )
        if not confirmar:
            return
        self._eliminar_contenido_institucional(self.contenido_institucional_id_actual)
        self._contenido_institucional_recargar_tree()
        self._contenido_institucional_limpiar_form()
        messagebox.showinfo(
            "Contenido Institucional",
            "Contenido eliminado correctamente.",
            parent=self.win,
        )

    def _build_contenido_institucional_tab(self, parent=None):
        self._ensure_contenido_institucional()
        frame = parent if parent is not None else self.tab_contenido_institucional
        for widget in frame.winfo_children():
            widget.destroy()

        self.contenido_institucional_id_actual = None
        self.var_contenido_titulo = tk.StringVar()
        self.var_contenido_ruta_imagen = tk.StringVar()
        self.var_contenido_estado = tk.StringVar(value="Activo")
        self.var_contenido_orden = tk.StringVar(
            value=str(self._contenido_institucional_siguiente_orden())
        )

        shell = ttk.Frame(frame)
        shell.pack(fill="both", expand=True)

        hero = tk.Frame(shell, bg="#0f5db8", padx=24, pady=18)
        hero.pack(fill="x", padx=18, pady=(18, 10))
        tk.Label(
            hero,
            text="Contenido Institucional",
            font=("Segoe UI", 17, "bold"),
            bg="#0f5db8",
            fg="white",
        ).pack(anchor="w")
        tk.Label(
            hero,
            text="Administra los mensajes y piezas gráficas que se muestran en el acceso principal. Usa el orden para controlar la rotación del carrusel.",
            font=("Segoe UI", 10),
            bg="#0f5db8",
            fg="#dbeafe",
            wraplength=980,
            justify="left",
        ).pack(anchor="w", pady=(6, 0))

        contenedor = ttk.Frame(shell, padding=14)
        contenedor.pack(fill="both", expand=True, padx=4, pady=(0, 8))
        contenedor.grid_columnconfigure(0, weight=1, uniform="contenido")
        contenedor.grid_columnconfigure(1, weight=1, uniform="contenido")
        contenedor.grid_rowconfigure(0, weight=1)

        editor = ttk.LabelFrame(
            contenedor,
            text="Editor de Contenido Institucional",
            padding=14,
            style="Card.TLabelframe",
        )
        editor.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        editor.grid_columnconfigure(1, weight=1)
        editor.grid_rowconfigure(2, weight=1)

        listado = ttk.LabelFrame(
            contenedor,
            text="Contenido Registrado",
            padding=14,
            style="Card.TLabelframe",
        )
        listado.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        listado.grid_columnconfigure(0, weight=1)
        listado.grid_rowconfigure(1, weight=1)

        ttk.Label(
            editor,
            text="Crea anuncios visuales para el login. Si agregas imagen, el acceso mostrará el título y la imagen priorizando el espacio disponible.",
            style="CardHint.TLabel",
        ).grid(row=0, column=0, columnspan=3, sticky="w", padx=4, pady=(0, 8))

        self.lbl_contenido_form_estado = ttk.Label(
            editor, text="Nuevo contenido", font=("Arial", 10, "bold")
        )
        self.lbl_contenido_form_estado.grid(
            row=1, column=0, columnspan=3, sticky="w", pady=(0, 10)
        )

        ttk.Label(editor, text="Título:").grid(
            row=2, column=0, sticky="w", padx=4, pady=4
        )
        ttk.Entry(editor, textvariable=self.var_contenido_titulo).grid(
            row=2, column=1, columnspan=2, sticky="ew", padx=4, pady=4
        )

        ttk.Label(editor, text="Mensaje:").grid(
            row=3, column=0, sticky="nw", padx=4, pady=4
        )
        self.txt_contenido_mensaje = tk.Text(
            editor, height=6, wrap="word", font=("Segoe UI", 10)
        )
        self.txt_contenido_mensaje.grid(
            row=3, column=1, columnspan=2, sticky="nsew", padx=4, pady=4
        )

        ttk.Label(editor, text="Imagen:").grid(
            row=4, column=0, sticky="w", padx=4, pady=4
        )
        ttk.Entry(editor, textvariable=self.var_contenido_ruta_imagen).grid(
            row=4, column=1, sticky="ew", padx=4, pady=4
        )
        acciones_imagen = ttk.Frame(editor)
        acciones_imagen.grid(row=4, column=2, sticky="e", padx=4, pady=4)
        ttk.Button(
            acciones_imagen,
            text="Seleccionar",
            command=self._contenido_institucional_seleccionar_imagen,
        ).pack(side="left", padx=(0, 4))
        ttk.Button(
            acciones_imagen,
            text="Quitar",
            command=lambda: (
                self.var_contenido_ruta_imagen.set(""),
                self._contenido_institucional_actualizar_preview(),
            ),
        ).pack(side="left")

        ttk.Label(editor, text="Estado:").grid(
            row=5, column=0, sticky="w", padx=4, pady=4
        )
        ttk.Combobox(
            editor,
            textvariable=self.var_contenido_estado,
            values=["Activo", "Inactivo"],
            state="readonly",
            width=18,
        ).grid(row=5, column=1, sticky="w", padx=4, pady=4)

        ttk.Label(editor, text="Orden:").grid(
            row=5, column=2, sticky="w", padx=4, pady=4
        )
        ttk.Entry(editor, textvariable=self.var_contenido_orden, width=8).grid(
            row=5, column=2, sticky="e", padx=4, pady=4
        )

        preview_frame = ttk.LabelFrame(
            editor, text="Vista previa", padding=8, style="Card.TLabelframe"
        )
        preview_frame.grid(
            row=6, column=0, columnspan=3, sticky="ew", padx=4, pady=(6, 6)
        )
        preview_frame.grid_columnconfigure(0, weight=1)
        preview_body = tk.Frame(preview_frame, bg="#ffffff", height=105)
        preview_body.grid(row=0, column=0, sticky="ew")
        preview_body.grid_propagate(False)
        self.lbl_contenido_preview = ttk.Label(
            preview_body, text="Sin imagen seleccionada", anchor="center"
        )
        self.lbl_contenido_preview.pack(fill="both", expand=True)

        acciones = ttk.Frame(editor)
        acciones.grid(row=7, column=0, columnspan=3, sticky="ew", pady=(6, 0))
        ttk.Button(
            acciones, text="Nuevo", command=self._contenido_institucional_limpiar_form
        ).pack(side="left")
        self._crear_boton_si_permiso(
            ttk.Button,
            acciones,
            "desktop.superadmin.contenido_institucional.guardar",
            text="Guardar",
            command=self._contenido_institucional_guardar_desde_form,
            layout_kwargs={"side": "right"},
        )
        self.btn_contenido_eliminar = self._crear_boton_si_permiso(
            ttk.Button,
            acciones,
            "desktop.superadmin.contenido_institucional.eliminar",
            text="Eliminar",
            command=self._contenido_institucional_eliminar_desde_form,
            layout_kwargs={"side": "right", "padx": (0, 6)},
        )
        if self.btn_contenido_eliminar is not None:
            self.btn_contenido_eliminar.state(["disabled"])

        ttk.Label(
            listado,
            text="Consulta el contenido ya creado, selecciona un registro y ajusta estado, orden o material gráfico sin salir del panel.",
            style="CardHint.TLabel",
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))
        self.lbl_contenido_total = ttk.Label(
            listado, text="Registros: 0", font=("Arial", 9, "bold")
        )
        self.lbl_contenido_total.grid(row=0, column=0, sticky="e", pady=(0, 8))

        columnas = ("orden", "estado", "titulo", "imagen", "mensaje")
        self.tree_contenido_institucional = ttk.Treeview(
            listado,
            columns=columnas,
            show="headings",
            height=16,
        )
        encabezados = {
            "orden": "Orden",
            "estado": "Estado",
            "titulo": "Título",
            "imagen": "Imagen",
            "mensaje": "Mensaje",
        }
        anchos = {
            "orden": 70,
            "estado": 90,
            "titulo": 170,
            "imagen": 70,
            "mensaje": 260,
        }
        for columna in columnas:
            self.tree_contenido_institucional.heading(
                columna, text=encabezados[columna]
            )
            self.tree_contenido_institucional.column(
                columna,
                width=anchos[columna],
                anchor="w",
                stretch=(columna in {"titulo", "mensaje"}),
            )
        self.tree_contenido_institucional.grid(row=1, column=0, sticky="nsew")
        scrollbar_tree = ttk.Scrollbar(
            listado,
            orient="vertical",
            command=self.tree_contenido_institucional.yview,
        )
        scrollbar_tree.grid(row=1, column=1, sticky="ns")
        self.tree_contenido_institucional.configure(yscrollcommand=scrollbar_tree.set)
        self.tree_contenido_institucional.bind(
            "<<TreeviewSelect>>", self._contenido_institucional_on_select
        )

        self._contenido_institucional_limpiar_form()
        self._contenido_institucional_recargar_tree()

    def _build_configuracion_plantel_tab(self, parent=None):
        """Construye la interfaz de Configuración de Plantel."""
        frame = parent if parent is not None else self.tab_configuracion_plantel

        # Contenedor centrado para el canvas
        container = ttk.Frame(frame)
        container.pack(fill="both", expand=True)

        # Canvas con scrollbar para formularios largos
        canvas = tk.Canvas(container, bg="#eef4fb", highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Centrar el contenido del formulario
        scrollable_frame = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=scrollable_frame, anchor="n", width=1100)

        def resize_canvas(event):
            # Centrar el canvas y ajustar ancho
            canvas_width = min(max(event.width, 700), 1200)
            canvas.itemconfig("all", width=canvas_width)
            canvas.config(width=canvas_width)

        scrollable_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        container.bind("<Configure>", resize_canvas)
        canvas.configure(yscrollcommand=scrollbar.set)

        nombre_actual = (
            self._get_config_plantel("nombre_institucion")
            or "Institución no configurada"
        )
        caracter_actual = (
            self._get_config_plantel("caracter_institucion") or "Sin definir"
        )
        anio_actual = self.obtener_anio_lectivo_activo() or "Sin definir"
        total_sedes = len(self._config_plantel_cargar_sedes())

        hero = tk.Frame(scrollable_frame, bg="#0f5db8", padx=24, pady=20)
        hero.pack(fill="x", padx=22, pady=(20, 12))

        hero_top = tk.Frame(hero, bg="#0f5db8")
        hero_top.pack(fill="x")
        hero_info = tk.Frame(hero_top, bg="#0f5db8")
        hero_info.pack(side="left", fill="x", expand=True)
        tk.Label(
            hero_info,
            text="Configuración Institución Educativa",
            font=("Segoe UI", 18, "bold"),
            bg="#0f5db8",
            fg="white",
        ).pack(anchor="w")
        tk.Label(
            hero_info,
            text="Administra la identidad institucional, la estructura académica y la información oficial del plantel desde un solo panel.",
            font=("Segoe UI", 10),
            bg="#0f5db8",
            fg="#dbeafe",
            wraplength=760,
            justify="left",
        ).pack(anchor="w", pady=(6, 0))

        hero_actions = tk.Frame(hero_top, bg="#0f5db8")
        hero_actions.pack(side="right", anchor="ne")
        self._crear_boton_si_permiso(
            tk.Button,
            hero_actions,
            "desktop.superadmin.configuracion_plantel.guardar",
            text="Guardar configuración",
            font=("Segoe UI", 10, "bold"),
            bg="#ffffff",
            fg="#0f5db8",
            relief="flat",
            padx=14,
            pady=8,
            cursor="hand2",
            command=self.configuracion_plantel_guardar,
            layout_kwargs={"side": "left", "padx": (0, 8)},
        )
        tk.Button(
            hero_actions,
            text="Limpiar",
            font=("Segoe UI", 10, "bold"),
            bg="#0b4a92",
            fg="white",
            relief="flat",
            padx=14,
            pady=8,
            cursor="hand2",
            command=self._limpiar_configuracion_plantel,
        ).pack(side="left")

        hero_stats = tk.Frame(hero, bg="#0f5db8")
        hero_stats.pack(fill="x", pady=(16, 0))

        def _stat_card(parent_stat, titulo, valor):
            card = tk.Frame(
                parent_stat,
                bg="#1559a6",
                padx=14,
                pady=10,
                highlightthickness=1,
                highlightbackground="#2d73c4",
            )
            card.pack(side="left", padx=(0, 10), fill="x", expand=True)
            tk.Label(
                card,
                text=titulo,
                font=("Segoe UI", 9, "bold"),
                bg="#1559a6",
                fg="#bfdbfe",
            ).pack(anchor="w")
            tk.Label(
                card,
                text=str(valor),
                font=("Segoe UI", 13, "bold"),
                bg="#1559a6",
                fg="white",
                wraplength=220,
                justify="left",
            ).pack(anchor="w", pady=(4, 0))

        _stat_card(hero_stats, "Institución", nombre_actual)
        _stat_card(hero_stats, "Carácter", caracter_actual)
        _stat_card(hero_stats, "Año lectivo", anio_actual)
        _stat_card(hero_stats, "Sedes", total_sedes)

        intro = ttk.Frame(scrollable_frame)
        intro.pack(fill="x", padx=22, pady=(0, 8))
        ttk.Label(
            intro,
            text="Panel de configuración",
            style="SectionTitle.TLabel",
        ).pack(anchor="w")
        ttk.Label(
            intro,
            text="Organiza la información por secciones y actualiza solo los datos necesarios. Los cambios se conservan con el botón de guardar.",
            style="SectionSub.TLabel",
        ).pack(anchor="w", pady=(4, 0))

        # Frame para campos del formulario
        form_frame = ttk.Frame(scrollable_frame)
        form_frame.pack(fill="both", expand=True, padx=22, pady=(8, 24))
        form_frame.grid_columnconfigure(0, weight=1, uniform="col")
        form_frame.grid_columnconfigure(1, weight=1, uniform="col")

        # ---------------- SECCIÓN 1: Información Legal e Identificación ----------------
        sec1 = ttk.LabelFrame(
            form_frame,
            text="1. Información Legal e Identificación",
            padding=20,
            style="Card.TLabelframe",
        )
        sec1.grid(row=0, column=0, columnspan=2, sticky="nsew", pady=10)
        sec1.grid_columnconfigure(0, weight=0)
        sec1.grid_columnconfigure(1, weight=1)

        ttk.Label(sec1, text="Nombre de la institución:").grid(
            row=0, column=0, sticky="w", padx=4, pady=4
        )

        self.entry_nombre_institucion = ttk.Entry(sec1, width=80)
        self.entry_nombre_institucion.grid(row=0, column=1, sticky="ew", padx=4, pady=4)
        self.entry_nombre_institucion.insert(
            0, self._get_config_plantel("nombre_institucion")
        )

        # Campo ID de la Institución (nuevo, editable, corto, alineado a la derecha de Nombre de la institución)
        ttk.Label(sec1, text="ID de la Institución:").grid(
            row=0, column=2, sticky="w", padx=(8, 2), pady=4
        )
        self.entry_id_institucion = ttk.Entry(sec1, width=14)
        self.entry_id_institucion.grid(row=0, column=3, sticky="w", padx=(0, 4), pady=4)
        # No se inicializa con ningún valor ni se conecta a backend, solo UI

        # Fila: Código DANE | NIT | Carácter de la Institución
        # --- NUEVA DISTRIBUCIÓN ---
        # Código DANE
        # Código DANE (ancho fijo para 12 caracteres)
        # Código DANE
        ttk.Label(sec1, text="Código DANE:", width=13, anchor="w").grid(
            row=1, column=0, sticky="w", padx=(4, 0), pady=4
        )
        self.entry_codigo_dane = ttk.Entry(sec1, width=13)
        # Sin padding a la derecha en Código DANE
        self.entry_codigo_dane.grid(row=1, column=1, sticky="w", padx=(0, 0), pady=4)
        self.entry_codigo_dane.insert(0, self._get_config_plantel("codigo_dane"))
        # NIT (label y entry juntos, después de Código DANE), sin separación
        frame_nit = ttk.Frame(sec1)
        # Padding izquierdo de 1 cm (20 px) para desplazar NIT respecto a Código DANE
        frame_nit.grid(row=1, column=1, sticky="w", padx=(120, 0), pady=4)
        ttk.Label(frame_nit, text="NIT:", width=4, anchor="w").pack(
            side="left", padx=(15, 5)
        )
        self.entry_nit = ttk.Entry(frame_nit, width=12)
        self.entry_nit.pack(side="left", padx=(0, 0))
        self.entry_nit.insert(0, self._get_config_plantel("nit"))
        # Carácter de la Institución (label y combobox juntos, después de NIT)
        ttk.Label(sec1, text="Carácter de la Institución:", anchor="w").grid(
            row=1, column=2, sticky="w", padx=(10, 2), pady=4
        )
        self.var_caracter_institucion = tk.StringVar()
        self.combo_caracter_institucion = ttk.Combobox(
            sec1,
            textvariable=self.var_caracter_institucion,
            values=["Público", "Privado"],
            state="readonly",
            width=28,
        )
        self.combo_caracter_institucion.grid(
            row=1, column=3, sticky="w", padx=(0, 4), pady=4
        )
        valor_caracter = self._get_config_plantel("caracter_institucion") or ""
        if valor_caracter in ("Público", "Privado"):
            self.combo_caracter_institucion.set(valor_caracter)
        else:
            self.combo_caracter_institucion.set("")
        # Configurar columnas para alineación compacta y expansión del carácter
        sec1.grid_columnconfigure(0, minsize=80, weight=0)
        sec1.grid_columnconfigure(1, minsize=110, weight=0)
        sec1.grid_columnconfigure(2, minsize=5, weight=0)
        sec1.grid_columnconfigure(3, minsize=90, weight=0)
        sec1.grid_columnconfigure(4, minsize=120, weight=0)
        sec1.grid_columnconfigure(5, minsize=200, weight=1)
        sec1.grid_columnconfigure(3, minsize=80, weight=0)
        sec1.grid_columnconfigure(4, minsize=120, weight=0)
        sec1.grid_columnconfigure(5, minsize=200, weight=1)  # Carácter ocupa el resto

        ttk.Label(sec1, text="Decreto de funcionamiento:").grid(
            row=2, column=0, sticky="w", padx=4, pady=4
        )
        self.entry_decreto = ttk.Entry(sec1, width=80)
        self.entry_decreto.grid(
            row=2, column=1, columnspan=3, sticky="ew", padx=4, pady=4
        )
        self.entry_decreto.insert(0, self._get_config_plantel("decreto_funcionamiento"))

        ttk.Label(sec1, text="Resolución de aprobación:").grid(
            row=3, column=0, sticky="w", padx=4, pady=4
        )
        self.entry_resolucion = ttk.Entry(sec1, width=80)
        self.entry_resolucion.grid(
            row=3, column=1, columnspan=3, sticky="ew", padx=4, pady=4
        )
        self.entry_resolucion.insert(
            0, self._get_config_plantel("resolucion_aprobacion")
        )

        ttk.Label(sec1, text="Año lectivo activo:").grid(
            row=5, column=0, sticky="w", padx=4, pady=4
        )
        self.entry_anio_lectivo = ttk.Entry(sec1, width=20)
        self.entry_anio_lectivo.grid(row=5, column=1, sticky="ew", padx=4, pady=4)
        self.entry_anio_lectivo.insert(0, self.obtener_anio_lectivo_activo())
        # ---------------- SECCIÓN 7: Configuración Académica ----------------
        sec_acad = ttk.LabelFrame(
            form_frame,
            text="7. Configuración Académica",
            padding=20,
            style="Card.TLabelframe",
        )
        sec_acad.grid(row=7, column=0, columnspan=2, sticky="nsew", pady=10)
        sec_acad.grid_columnconfigure(0, weight=0)
        sec_acad.grid_columnconfigure(1, weight=1)

        # --- 7.1 Cantidad de períodos académicos ---
        frame_periodos = ttk.LabelFrame(
            sec_acad,
            text="7.1 Cantidad de períodos académicos",
            padding=12,
            style="Card.TLabelframe",
        )
        frame_periodos.grid(row=0, column=0, columnspan=4, sticky="ew", pady=(4, 4))
        ttk.Label(frame_periodos, text="Cantidad de períodos académicos:").grid(
            row=0, column=0, sticky="w", padx=4, pady=4
        )
        self.var_cantidad_periodos = tk.IntVar(
            value=int(self._get_config_plantel("cantidad_periodos") or 1)
        )
        self.spin_cantidad_periodos = ttk.Spinbox(
            frame_periodos,
            from_=1,
            to=10,
            textvariable=self.var_cantidad_periodos,
            state="readonly",
            width=5,
        )
        self.spin_cantidad_periodos.grid(row=0, column=1, sticky="ew", padx=4, pady=4)
        ttk.Label(frame_periodos, text="Activar distribución porcentual:").grid(
            row=0, column=2, sticky="w", padx=4, pady=4
        )
        self.var_usa_porcentajes = tk.StringVar()
        self.combo_usa_porcentajes = ttk.Combobox(
            frame_periodos,
            textvariable=self.var_usa_porcentajes,
            values=["No", "Sí"],
            state="readonly",
            width=5,
        )
        self.combo_usa_porcentajes.grid(row=0, column=3, sticky="ew", padx=4, pady=4)
        valor_usa = self._get_config_plantel("usa_porcentajes")
        self.combo_usa_porcentajes.set("Sí" if valor_usa == "Sí" else "No")
        # --- Frame de porcentajes dinámicos (row=1) ---
        self.frame_porcentajes = None

        def actualizar_porcentajes(*_):
            if self.frame_porcentajes:
                self.frame_porcentajes.destroy()
                self.frame_porcentajes = None
            if self.combo_usa_porcentajes.get() == "Sí":
                self.frame_porcentajes = ttk.LabelFrame(
                    frame_periodos, text="Porcentajes por período", padding=8
                )
                self.frame_porcentajes.grid(
                    row=1, column=0, columnspan=4, sticky="ew", pady=(8, 4)
                )
                self.vars_porcentajes = []
                cantidad = self.var_cantidad_periodos.get()
                # Leer valores guardados
                try:
                    valores_guardados = json.loads(
                        self._get_config_plantel("porcentajes_periodos") or "[]"
                    )
                except Exception:
                    valores_guardados = []
                for i in range(1, cantidad + 1):
                    ttk.Label(self.frame_porcentajes, text=f"Período {i} (%)").grid(
                        row=0, column=i - 1, padx=4, pady=2
                    )
                    valor = (
                        valores_guardados[i - 1]
                        if i - 1 < len(valores_guardados)
                        else 0
                    )
                    var = tk.IntVar(value=valor)
                    entry = ttk.Entry(self.frame_porcentajes, textvariable=var, width=6)
                    entry.grid(row=1, column=i - 1, padx=4, pady=2)
                    self.vars_porcentajes.append(var)

        self.combo_usa_porcentajes.bind("<<ComboboxSelected>>", actualizar_porcentajes)
        self.var_cantidad_periodos.trace_add(
            "write", lambda *_: actualizar_porcentajes()
        )
        actualizar_porcentajes()

        # --- 7.1 Sistema de Evaluación (row=2) ---
        # --- 7.2 Sistema de Evaluación ---
        sep_eval = ttk.LabelFrame(
            sec_acad,
            text="7.2 Sistema de Evaluación",
            padding=12,
            style="Card.TLabelframe",
        )
        sep_eval.grid(row=1, column=0, columnspan=4, sticky="ew", pady=(12, 4))

        cfg_componentes_eval = self._obtener_componentes_evaluacion()
        self.var_eval_cognitivo = tk.IntVar(
            value=cfg_componentes_eval["cognitivo"]["porcentaje"]
        )
        self.var_eval_examen = tk.IntVar(
            value=cfg_componentes_eval["examen"]["porcentaje"]
        )
        self.var_eval_auto = tk.IntVar(
            value=cfg_componentes_eval["autoevaluacion"]["porcentaje"]
        )
        self.var_eval_notas_cognitivo = tk.IntVar(
            value=cfg_componentes_eval["cognitivo"]["cantidad"]
        )
        self.var_eval_notas_examen = tk.IntVar(
            value=cfg_componentes_eval["examen"]["cantidad"]
        )
        self.var_eval_nombre_cognitivo = tk.StringVar(
            value=cfg_componentes_eval["cognitivo"]["nombre"]
        )
        self.var_eval_nombre_examen = tk.StringVar(
            value=cfg_componentes_eval["examen"]["nombre"]
        )
        self.var_eval_nombre_auto = tk.StringVar(
            value=cfg_componentes_eval["autoevaluacion"]["nombre"]
        )
        self.var_eval_notas_auto = tk.IntVar(
            value=cfg_componentes_eval["autoevaluacion"]["cantidad"]
        )

        ttk.Label(
            sep_eval,
            text="Define nombre, porcentaje y cantidad de notas por componente. Si un componente no aplica, puede quedar en 0.",
            style="CardHint.TLabel",
        ).grid(row=0, column=0, columnspan=7, sticky="w", padx=4, pady=(0, 8))

        filas_componentes = [
            (
                self.var_eval_nombre_cognitivo,
                self.var_eval_cognitivo,
                self.var_eval_notas_cognitivo,
            ),
            (
                self.var_eval_nombre_examen,
                self.var_eval_examen,
                self.var_eval_notas_examen,
            ),
            (
                self.var_eval_nombre_auto,
                self.var_eval_auto,
                self.var_eval_notas_auto,
            ),
        ]
        for indice, (var_nombre, var_porcentaje, var_notas) in enumerate(
            filas_componentes, start=1
        ):
            ttk.Label(sep_eval, text="Componente:").grid(
                row=indice, column=0, sticky="w", padx=4, pady=4
            )
            ttk.Entry(sep_eval, textvariable=var_nombre, width=18).grid(
                row=indice, column=1, sticky="w", padx=4, pady=4
            )
            ttk.Label(sep_eval, text="Porcentaje (%):").grid(
                row=indice, column=2, sticky="w", padx=(18, 4), pady=4
            )
            ttk.Entry(sep_eval, textvariable=var_porcentaje, width=8).grid(
                row=indice, column=3, sticky="w", padx=4, pady=4
            )
            ttk.Label(sep_eval, text="Cantidad de notas:").grid(
                row=indice, column=4, sticky="w", padx=(18, 4), pady=4
            )
            ttk.Spinbox(
                sep_eval,
                from_=0,
                to=20,
                textvariable=var_notas,
                width=5,
                state="readonly",
            ).grid(row=indice, column=5, sticky="w", padx=4, pady=4)

        self.lbl_eval_sum = ttk.Label(sep_eval, text="Suma: 0%", foreground="red")
        self.lbl_eval_sum.grid(row=1, column=6, padx=8, pady=2, sticky="w")
        self.lbl_eval_total_notas = ttk.Label(
            sep_eval, text="Total notas: 0", style="CardHint.TLabel"
        )
        self.lbl_eval_total_notas.grid(row=2, column=6, padx=8, pady=2, sticky="w")

        def actualizar_suma_eval(*_):
            total = (
                self.var_eval_cognitivo.get()
                + self.var_eval_examen.get()
                + self.var_eval_auto.get()
            )
            self.lbl_eval_sum.config(
                text=f"Suma: {total}%", foreground=("green" if total == 100 else "red")
            )
            total_notas = (
                self.var_eval_notas_cognitivo.get()
                + self.var_eval_notas_examen.get()
                + self.var_eval_notas_auto.get()
            )
            self.lbl_eval_total_notas.config(text=f"Total notas: {total_notas}")

        for v in (
            self.var_eval_cognitivo,
            self.var_eval_examen,
            self.var_eval_auto,
            self.var_eval_notas_cognitivo,
            self.var_eval_notas_examen,
            self.var_eval_notas_auto,
        ):
            v.trace_add("write", actualizar_suma_eval)
        actualizar_suma_eval()

        # Frame para porcentajes dinámicos (solo si corresponde)

        # --- 7.1 Sistema de Evaluación ---

        def guardar_config_evaluacion():
            # Validación suma 100%
            total = (
                self.var_eval_cognitivo.get()
                + self.var_eval_examen.get()
                + self.var_eval_auto.get()
            )
            if total != 100:
                messagebox.showerror(
                    "Error de validación", "La suma de los porcentajes debe ser 100%."
                )
                return False
            cantidades = {
                "notas_cognitivo": self.var_eval_notas_cognitivo.get(),
                "notas_examen": self.var_eval_notas_examen.get(),
                "notas_autoevaluacion": self.var_eval_notas_auto.get(),
            }
            if any(valor < 0 for valor in cantidades.values()):
                messagebox.showerror(
                    "Error de validación",
                    "Cada cantidad de notas debe ser mayor o igual a 0.",
                )
                return False
            nombres_componentes = {
                "nombre_cognitivo": self.var_eval_nombre_cognitivo.get().strip(),
                "nombre_examen": self.var_eval_nombre_examen.get().strip(),
                "nombre_autoevaluacion": self.var_eval_nombre_auto.get().strip(),
            }
            if any(not valor for valor in nombres_componentes.values()):
                messagebox.showerror(
                    "Error de validación",
                    "Cada componente debe tener un nombre.",
                )
                return False
            # Guardar en tabla config_evaluacion (debe implementarse la lógica de guardado real)
            config = {
                "cognitivo": self.var_eval_cognitivo.get(),
                "examen": self.var_eval_examen.get(),
                "autoevaluacion": self.var_eval_auto.get(),
                "cantidad_notas": sum(cantidades.values()),
                "nota_min": self._coerce_float(
                    self._get_config_plantel("nota_min") or 1.0, 1.0
                ),
                "nota_max": self._coerce_float(
                    self._get_config_plantel("nota_max") or 5.0, 5.0
                ),
                "decimales": self._coerce_int(
                    self._get_config_plantel("decimales") or 1, 1
                ),
                "metodo": self._get_config_plantel("metodo_evaluacion") or "promedio",
                "notas_cognitivo": cantidades["notas_cognitivo"],
                "notas_examen": cantidades["notas_examen"],
                "notas_autoevaluacion": cantidades["notas_autoevaluacion"],
            }
            # Guardar en base de datos real (tabla config_evaluacion)
            self._guardar_config_evaluacion(config)
            for clave, valor in nombres_componentes.items():
                self._set_config_plantel(clave, valor)
            messagebox.showinfo(
                "Guardado", "Configuración de evaluación guardada correctamente."
            )
            return True

        self.guardar_config_evaluacion = guardar_config_evaluacion
        for i in range(4):
            sec_acad.grid_columnconfigure(i, weight=1)

        # Frame para porcentajes dinámicos

        # ---------------- SECCIÓN 2: Ubicación Geográfica ----------------
        sec2 = ttk.LabelFrame(
            form_frame,
            text="2. Ubicación Geográfica",
            padding=20,
            style="Card.TLabelframe",
        )
        sec2.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=10)
        sec2.grid_columnconfigure(0, weight=0)
        sec2.grid_columnconfigure(1, weight=1)
        sec2.grid_columnconfigure(2, weight=0)
        sec2.grid_columnconfigure(3, weight=1)

        # Departamento y Municipio en la misma fila
        ttk.Label(sec2, text="Departamento:").grid(
            row=0, column=0, sticky="w", padx=4, pady=4
        )
        self.entry_departamento = ttk.Entry(sec2, width=30)
        self.entry_departamento.grid(row=0, column=1, sticky="ew", padx=4, pady=4)
        self.entry_departamento.insert(0, self._get_config_plantel("departamento"))

        ttk.Label(sec2, text="Municipio:").grid(
            row=0, column=2, sticky="w", padx=4, pady=4
        )
        self.entry_municipio = ttk.Entry(sec2, width=30)
        self.entry_municipio.grid(row=0, column=3, sticky="ew", padx=4, pady=4)
        self.entry_municipio.insert(0, self._get_config_plantel("municipio"))

        # Corregimiento y Dirección en la misma fila
        ttk.Label(sec2, text="Corregimiento / Localidad:").grid(
            row=1, column=0, sticky="w", padx=4, pady=4
        )
        self.entry_corregimiento = ttk.Entry(sec2, width=40)
        self.entry_corregimiento.grid(row=1, column=1, sticky="ew", padx=4, pady=4)
        self.entry_corregimiento.insert(
            0, self._get_config_plantel("corregimiento_localidad")
        )

        ttk.Label(sec2, text="Dirección:").grid(
            row=1, column=2, sticky="w", padx=4, pady=4
        )
        self.entry_direccion = ttk.Entry(sec2, width=40)
        self.entry_direccion.grid(row=1, column=3, sticky="ew", padx=4, pady=4)
        self.entry_direccion.insert(0, self._get_config_plantel("direccion"))

        # ---------------- SECCIÓN 3: Sedes Educativas ----------------
        sec3 = ttk.LabelFrame(
            form_frame,
            text="3. Sedes Educativas",
            padding=12,
            style="Card.TLabelframe",
        )
        sec3.grid(row=2, column=0, columnspan=2, sticky="we", pady=6)

        top_sec3 = ttk.Frame(sec3)
        top_sec3.grid(row=0, column=0, sticky="ew", padx=4, pady=(0, 8))
        top_sec3.grid_columnconfigure(0, weight=1)
        info_sec3 = ttk.Frame(top_sec3)
        info_sec3.grid(row=0, column=0, sticky="w")
        ttk.Label(
            info_sec3,
            text="Registra las sedes habilitadas y las jornadas activas en cada una.",
            style="CardHint.TLabel",
        ).pack(anchor="w")
        self.sedes_summary_label = ttk.Label(
            info_sec3,
            text="0 sedes configuradas",
            style="CardHint.TLabel",
        )
        self.sedes_summary_label.pack(anchor="w", pady=(2, 0))
        ttk.Button(
            top_sec3,
            text="Agregar Sede",
            command=self._config_plantel_agregar_sede,
        ).grid(row=0, column=1, sticky="e")

        self.sedes_data = self._config_plantel_cargar_sedes()
        self.sedes_body_frame = ttk.Frame(sec3)
        self.sedes_body_frame.grid(row=1, column=0, sticky="ew", padx=4, pady=2)
        sec3.grid_columnconfigure(0, weight=1)
        self._config_plantel_render_sedes()

        # ---------------- SECCIÓN 4: Información Directiva ----------------
        sec4 = ttk.LabelFrame(
            form_frame,
            text="4. Información Directiva",
            padding=12,
            style="Card.TLabelframe",
        )
        sec4.grid(row=3, column=0, columnspan=2, sticky="we", pady=6)
        sec4.grid_columnconfigure(0, weight=0)
        sec4.grid_columnconfigure(1, weight=1)
        ttk.Label(
            sec4,
            text="Asocia el rector a la planta docente y conserva un bloque separado para datos administrativos de secretaría.",
            style="CardHint.TLabel",
        ).grid(row=0, column=0, columnspan=2, sticky="w", padx=4, pady=(0, 8))

        # --- NUEVO: ComboBox para seleccionar Rector desde Planta Docente ---
        ttk.Label(sec4, text="Rector (seleccione docente):").grid(
            row=1, column=0, sticky="w", padx=4, pady=2
        )
        self.var_rector_docente_id = tk.StringVar()
        self.combo_rector_docente = ttk.Combobox(
            sec4, textvariable=self.var_rector_docente_id, state="readonly", width=60
        )
        self.combo_rector_docente.grid(row=1, column=1, sticky="ew", padx=4, pady=2)

        # Cargar docentes activos (nombre completo)
        from core import docentes as core_docentes

        docentes = []
        try:
            data = core_docentes.listar_docentes(limit=10000, offset=0)
            for d in data.get("docentes", []):
                if d.get("estado", "Activo") == "Activo":
                    docentes.append((d.get("documento", ""), d.get("nombre", "")))
        except Exception:
            docentes = []
        self._rector_docente_map = {
            f"{nombre} [{doc}]": doc for doc, nombre in docentes if doc and nombre
        }
        self._rector_docente_invmap = {
            doc: f"{nombre} [{doc}]" for doc, nombre in docentes if doc and nombre
        }
        self.combo_rector_docente["values"] = list(self._rector_docente_map.keys())

        # Cargar valor guardado si existe
        rector_docente_id = self._get_config_plantel("rector_docente_id")
        if rector_docente_id and rector_docente_id in self._rector_docente_invmap:
            self.combo_rector_docente.set(
                self._rector_docente_invmap[rector_docente_id]
            )
        elif rector_docente_id and rector_docente_id in self._rector_docente_map:
            # Compatibilidad con valores guardados erróneamente como etiqueta visible.
            self.combo_rector_docente.set(rector_docente_id)
        else:
            self.combo_rector_docente.set("")

        # Campos de solo lectura para mostrar nombre e identificación
        ttk.Label(sec4, text="Nombre completo:").grid(
            row=2, column=0, sticky="w", padx=4, pady=2
        )
        self.entry_rector_nombre = ttk.Entry(sec4, width=60, state="readonly")
        self.entry_rector_nombre.grid(row=2, column=1, sticky="ew", padx=4, pady=2)

        ttk.Label(sec4, text="Identificación:").grid(
            row=3, column=0, sticky="w", padx=4, pady=2
        )
        self.entry_rector_id = ttk.Entry(sec4, width=30, state="readonly")
        self.entry_rector_id.grid(row=3, column=1, sticky="ew", padx=4, pady=2)

        ttk.Label(sec4, text="Cargo:").grid(row=4, column=0, sticky="w", padx=4, pady=2)
        self.entry_rector_cargo = ttk.Entry(sec4, width=40, state="readonly")
        self.entry_rector_cargo.grid(row=4, column=1, sticky="ew", padx=4, pady=2)
        self.entry_rector_cargo.insert(0, "Rector")

        # Mostrar datos si hay selección previa
        def actualizar_campos_rector(event=None):
            sel = self.combo_rector_docente.get()
            doc_id = self._rector_docente_map.get(sel, "")
            nombre = ""
            identificacion = ""
            if doc_id:
                # Buscar nombre e identificación
                for d in docentes:
                    if d[0] == doc_id:
                        nombre = d[1]
                        identificacion = d[0]
                        break
            self.entry_rector_nombre.config(state="normal")
            self.entry_rector_nombre.delete(0, "end")
            self.entry_rector_nombre.insert(0, nombre)
            self.entry_rector_nombre.config(state="readonly")
            self.entry_rector_id.config(state="normal")
            self.entry_rector_id.delete(0, "end")
            self.entry_rector_id.insert(0, identificacion)
            self.entry_rector_id.config(state="readonly")
            self.entry_rector_cargo.config(state="normal")
            self.entry_rector_cargo.delete(0, "end")
            self.entry_rector_cargo.insert(0, "Rector")
            self.entry_rector_cargo.config(state="readonly")

        self.combo_rector_docente.bind("<<ComboboxSelected>>", actualizar_campos_rector)
        # Inicializar campos si hay valor
        actualizar_campos_rector()

        # --- RESPALDO: Mostrar campos antiguos si existen (solo lectura, no eliminar) ---
        rector_nombre_ant = self._get_config_plantel("rector_nombre")
        rector_id_ant = self._get_config_plantel("rector_identificacion")
        rector_cargo_ant = self._get_config_plantel("rector_cargo")
        secretaria_row_base = 5
        if rector_nombre_ant or rector_id_ant or rector_cargo_ant:
            ttk.Label(sec4, text="(Respaldo) Rector - Nombre:").grid(
                row=5, column=0, sticky="w", padx=4, pady=2
            )
            entry_rector_nombre_ant = ttk.Entry(sec4, width=60, state="readonly")
            entry_rector_nombre_ant.grid(row=5, column=1, sticky="ew", padx=4, pady=2)
            entry_rector_nombre_ant.insert(0, rector_nombre_ant)
            ttk.Label(sec4, text="(Respaldo) Rector - Identificación:").grid(
                row=6, column=0, sticky="w", padx=4, pady=2
            )
            entry_rector_id_ant = ttk.Entry(sec4, width=30, state="readonly")
            entry_rector_id_ant.grid(row=6, column=1, sticky="ew", padx=4, pady=2)
            entry_rector_id_ant.insert(0, rector_id_ant)
            ttk.Label(sec4, text="(Respaldo) Rector - Cargo:").grid(
                row=7, column=0, sticky="w", padx=4, pady=2
            )
            entry_rector_cargo_ant = ttk.Entry(sec4, width=40, state="readonly")
            entry_rector_cargo_ant.grid(row=7, column=1, sticky="ew", padx=4, pady=2)
            entry_rector_cargo_ant.insert(0, rector_cargo_ant)
            secretaria_row_base = 8

        # Secretaría
        ttk.Label(sec4, text="Secretaría - Nombre:").grid(
            row=secretaria_row_base,
            column=0,
            sticky="w",
            padx=4,
            pady=(10, 2),
        )
        self.entry_sec_nombre = ttk.Entry(sec4, width=60)
        self.entry_sec_nombre.grid(
            row=secretaria_row_base, column=1, sticky="ew", padx=4, pady=(10, 2)
        )
        self.entry_sec_nombre.insert(0, self._get_config_plantel("secretaria_nombre"))

        ttk.Label(sec4, text="Secretaría - Identificación:").grid(
            row=secretaria_row_base + 1, column=0, sticky="w", padx=4, pady=2
        )
        self.entry_sec_id = ttk.Entry(sec4, width=30)
        self.entry_sec_id.grid(
            row=secretaria_row_base + 1, column=1, sticky="ew", padx=4, pady=2
        )
        self.entry_sec_id.insert(
            0, self._get_config_plantel("secretaria_identificacion")
        )

        ttk.Label(sec4, text="Secretaría - Cargo:").grid(
            row=secretaria_row_base + 2, column=0, sticky="w", padx=4, pady=2
        )
        self.entry_sec_cargo = ttk.Entry(sec4, width=40)
        self.entry_sec_cargo.grid(
            row=secretaria_row_base + 2, column=1, sticky="ew", padx=4, pady=2
        )
        self.entry_sec_cargo.insert(0, self._get_config_plantel("secretaria_cargo"))

        # ---------------- SECCIÓN 5: Información de Contacto ----------------
        sec5 = ttk.LabelFrame(
            form_frame,
            text="5. Información de Contacto",
            padding=12,
            style="Card.TLabelframe",
        )
        sec5.grid(row=4, column=0, columnspan=2, sticky="we", pady=6)
        sec5.grid_columnconfigure(0, weight=0)
        sec5.grid_columnconfigure(1, weight=1)
        sec5.grid_columnconfigure(2, weight=0)
        sec5.grid_columnconfigure(3, weight=1)
        ttk.Label(
            sec5,
            text="Centraliza los canales institucionales que se mostrarán en reportes, documentos y comunicación oficial.",
            style="CardHint.TLabel",
        ).grid(row=0, column=0, columnspan=4, sticky="w", padx=4, pady=(0, 8))

        # Teléfono y Celular en la misma fila
        ttk.Label(sec5, text="Teléfono:").grid(
            row=1, column=0, sticky="w", padx=4, pady=4
        )
        self.entry_telefono = ttk.Entry(sec5, width=20)
        self.entry_telefono.grid(row=1, column=1, sticky="ew", padx=4, pady=4)
        self.entry_telefono.insert(0, self._get_config_plantel("telefono"))
        ttk.Label(sec5, text="Celular:").grid(
            row=1, column=2, sticky="w", padx=4, pady=4
        )
        self.entry_celular = ttk.Entry(sec5, width=20)
        self.entry_celular.grid(row=1, column=3, sticky="ew", padx=4, pady=4)
        self.entry_celular.insert(0, self._get_config_plantel("celular"))

        # Correo y Dominio en la misma fila
        ttk.Label(sec5, text="Correo institucional:").grid(
            row=2, column=0, sticky="w", padx=4, pady=4
        )
        self.entry_correo = ttk.Entry(sec5, width=30)
        self.entry_correo.grid(row=2, column=1, sticky="ew", padx=4, pady=4)
        self.entry_correo.insert(0, self._get_config_plantel("correo_institucional"))
        ttk.Label(sec5, text="Dominio Web:").grid(
            row=2, column=2, sticky="w", padx=4, pady=4
        )
        self.entry_dominio = ttk.Entry(sec5, width=30)
        self.entry_dominio.grid(row=2, column=3, sticky="ew", padx=4, pady=4)
        self.entry_dominio.insert(0, self._get_config_plantel("dominio_web"))

        # ---------------- SECCIÓN 6: Escala de Valoración ----------------
        sec6 = ttk.LabelFrame(
            form_frame,
            text="6. Escala de Valoración (orden descendente)",
            padding=12,
            style="Card.TLabelframe",
        )
        sec6.grid(row=5, column=0, columnspan=2, sticky="nsew", pady=6)
        for i in range(6):
            sec6.grid_columnconfigure(i, weight=1)

        ttk.Label(
            sec6,
            text="Define los rangos finales en orden descendente. Cada tramo debe ser exclusivo y sin cruces entre valores.",
            style="CardHint.TLabel",
        ).grid(row=0, column=0, columnspan=5, sticky="w", padx=4, pady=(0, 8))

        headings = ["Desde", "Hasta", "Letra", "Desempeño", "Recomendación"]
        for i, h in enumerate(headings):
            ttk.Label(sec6, text=h, font=("Arial", 9, "bold")).grid(
                row=1, column=i, padx=4, pady=2, sticky="ew"
            )
        # Ajustar columnas: las primeras 4 fijas, la última (Recomendación) expandible
        for i in range(4):
            sec6.grid_columnconfigure(i, weight=0, minsize=0)
        sec6.grid_columnconfigure(4, weight=1, minsize=120)

        # crear 4 filas editables (fijas) para la escala
        self.escala_entries = []
        escala_db = self.cur.execute(
            "SELECT desde, hasta, letra, desempeno, recomendacion FROM escala_valoracion ORDER BY desde DESC"
        ).fetchall()
        default_rows = [
            (4.6, 5.0, "S", "Superior", ""),
            (4.0, 4.5, "A", "Alto", ""),
            (3.0, 3.9, "B", "Básico", ""),
            (1.0, 2.9, "I", "Bajo", ""),
        ]
        rows_to_use = escala_db if escala_db else default_rows
        for r_idx in range(4):
            vals = (
                rows_to_use[r_idx] if r_idx < len(rows_to_use) else default_rows[r_idx]
            )
            row_widgets = {}
            # Desde, Hasta, Letra, Desempeño, Recomendación
            widths = [6, 6, 4, 10, 10]
            for c_idx, h in enumerate(headings):
                if h == "Recomendación":
                    e = ttk.Entry(sec6)
                    e.grid(
                        row=r_idx + 2,
                        column=c_idx,
                        columnspan=5 - c_idx,
                        padx=4,
                        pady=2,
                        sticky="ew",
                    )
                else:
                    e = ttk.Entry(sec6, width=widths[c_idx])
                    e.grid(row=r_idx + 2, column=c_idx, padx=4, pady=2, sticky="w")
                e.insert(0, str(vals[c_idx] if vals[c_idx] is not None else ""))
                row_widgets[h] = e
            self.escala_entries.append(row_widgets)

        # ---------------- SECCIÓN 8: Logo Institucional ----------------
        sec8 = ttk.LabelFrame(
            form_frame,
            text="8. Logo Institucional",
            padding=12,
            style="Card.TLabelframe",
        )
        sec8.grid(row=8, column=0, columnspan=2, sticky="nsew", pady=6)
        sec8.grid_columnconfigure(1, weight=1)
        ttk.Label(
            sec8,
            text="Carga el emblema oficial para usarlo en reportes, certificados y la portada de acceso.",
            style="CardHint.TLabel",
        ).grid(row=0, column=0, columnspan=4, sticky="w", padx=4, pady=(0, 8))

        ttk.Button(
            sec8,
            text="Seleccionar / Reemplazar Logo",
            command=self.configuracion_plantel_seleccionar_logo,
        ).grid(row=1, column=0, padx=6, pady=4, sticky="w")
        self.config_logo_label = ttk.Label(
            sec8, text="No hay logo seleccionado", foreground="#666666"
        )
        self.config_logo_label.grid(row=1, column=1, padx=8, pady=4, sticky="w")
        logo_path = self._get_config_plantel("logo_path")
        if logo_path and os.path.exists(logo_path):
            self.config_logo_label.config(text=os.path.basename(logo_path))

        footer_actions = ttk.Frame(sec8)
        footer_actions.grid(
            row=2, column=0, columnspan=4, sticky="e", padx=4, pady=(10, 0)
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            footer_actions,
            "desktop.superadmin.configuracion_plantel.guardar",
            text="Guardar Configuración",
            command=self.configuracion_plantel_guardar,
            layout_kwargs={"side": "left", "padx": (0, 8)},
        )
        ttk.Button(
            footer_actions,
            text="Limpiar formulario",
            command=self._limpiar_configuracion_plantel,
        ).pack(side="left")

        if not frame.winfo_children():
            # Forzar reconstrucción solo una vez
            self._reconstruir_configuracion_plantel()

    def _reconstruir_configuracion_plantel(self):
        """Reconstruye el contenido de la pestaña Configuración Plantel si está vacía."""
        frame = self.tab_configuracion_plantel
        for w in frame.winfo_children():
            w.destroy()
        try:
            # Canvas con scrollbar para formularios largos
            canvas = tk.Canvas(frame, bg="#f6f8fb", highlightthickness=0)
            scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all")),
            )
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            ttk.Label(
                scrollable_frame,
                text="Configuración Institución Educativa",
                font=("Arial", 14, "bold"),
            ).pack(pady=10, padx=10)
            form_frame = ttk.Frame(scrollable_frame)
            form_frame.pack(fill="both", expand=True, padx=10, pady=10)
            # ... (resto del código igual que antes, omitido aquí por brevedad)
            # Copia aquí todo el bloque de construcción si necesitas depurar más campos
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
        except Exception as e:
            import traceback

            print("[ERROR reconstrucción Configuración Plantel]:", e)
            traceback.print_exc()

    def configuracion_plantel_seleccionar_logo(self):
        """Permite seleccionar el logo institucional."""
        archivo = filedialog.askopenfilename(
            parent=self.win,
            title="Seleccionar Logo Institucional",
            filetypes=[("Imágenes", "*.png *.jpg *.jpeg *.bmp"), ("Todos", "*.*")],
        )
        if archivo:
            # Forzar PNG y nombre estandarizado: logo_institucion.png
            try:
                nombre_logo = "logo_institucion.png"
                logo_destino = os.path.join(self.imagenes_dir, nombre_logo)
                if _HAS_PIL:
                    try:
                        img = Image.open(archivo).convert("RGBA")
                        img.save(logo_destino, format="PNG")
                    except Exception:
                        shutil.copy2(archivo, logo_destino)
                else:
                    shutil.copy2(archivo, logo_destino)
                # guardar ruta absoluta en la configuración
                self._set_config_plantel("logo_path", logo_destino)
                self.config_logo_label.config(text=nombre_logo)
                messagebox.showinfo(
                    "Logo", "Logo institucional cargado correctamente.", parent=self.win
                )
            except Exception as e:
                messagebox.showerror(
                    "Error", f"Error al procesar logo: {e}", parent=self.win
                )

    def configuracion_plantel_guardar(self):
        if not self._requiere_permiso(
            "desktop.superadmin.configuracion_plantel.guardar"
        ):
            return
        # Guardar carácter de la institución
        self._set_config_plantel(
            "caracter_institucion", self.var_caracter_institucion.get()
        )

        # Guardar configuración académica
        self._set_config_plantel(
            "cantidad_periodos", str(self.var_cantidad_periodos.get())
        )
        self._set_config_plantel("usa_porcentajes", self.var_usa_porcentajes.get())
        porcentajes_eval = {
            "porcentaje_cognitivo": self.var_eval_cognitivo.get(),
            "porcentaje_examen": self.var_eval_examen.get(),
            "porcentaje_autoevaluacion": self.var_eval_auto.get(),
        }
        nombres_eval = {
            "nombre_cognitivo": self.var_eval_nombre_cognitivo.get().strip(),
            "nombre_examen": self.var_eval_nombre_examen.get().strip(),
            "nombre_autoevaluacion": self.var_eval_nombre_auto.get().strip(),
        }
        cantidades_eval = {
            "notas_cognitivo": self.var_eval_notas_cognitivo.get(),
            "notas_examen": self.var_eval_notas_examen.get(),
            "notas_autoevaluacion": self.var_eval_notas_auto.get(),
        }
        if sum(porcentajes_eval.values()) != 100:
            messagebox.showerror(
                "Error",
                "La suma de los porcentajes de Cognitivo, Examen y Autoevaluación debe ser 100%.",
                parent=self.win,
            )
            return
        if any(valor < 0 for valor in cantidades_eval.values()):
            messagebox.showerror(
                "Error",
                "Cada cantidad de notas debe ser mayor o igual a 0.",
                parent=self.win,
            )
            return
        if any(not valor for valor in nombres_eval.values()):
            messagebox.showerror(
                "Error",
                "Cada componente del sistema de evaluación debe tener un nombre.",
                parent=self.win,
            )
            return
        for clave, valor in nombres_eval.items():
            self._set_config_plantel(clave, valor)
        for clave, valor in porcentajes_eval.items():
            self._set_config_plantel(clave, str(valor))
        for clave, valor in cantidades_eval.items():
            self._set_config_plantel(clave, str(valor))
        # Compatibilidad con configuraciones antiguas y componentes que aún leen total global
        self._set_config_plantel("cantidad_notas", str(sum(cantidades_eval.values())))
        if self.var_usa_porcentajes.get() == "Sí":
            porcentajes = [v.get() for v in getattr(self, "vars_porcentajes", [])]
            suma = sum(porcentajes)
            if len(porcentajes) != self.var_cantidad_periodos.get():
                messagebox.showerror(
                    "Error",
                    "Debe ingresar el porcentaje para cada período.",
                    parent=self.win,
                )
                return
            if abs(suma - 100) > 0.01:
                messagebox.showerror(
                    "Error",
                    "La suma de los porcentajes debe ser igual a 100%.",
                    parent=self.win,
                )
                return
            self._set_config_plantel("porcentajes_periodos", json.dumps(porcentajes))
        else:
            self._set_config_plantel("porcentajes_periodos", "")

        try:
            # Validaciones básicas
            nombre = self.entry_nombre_institucion.get().strip()
            if not nombre:
                messagebox.showerror(
                    "Error",
                    "El nombre de la institución es obligatorio.",
                    parent=self.win,
                )
                return

            dane = self.entry_codigo_dane.get().strip()
            if dane and not re.fullmatch(r"\d{12}", dane):
                messagebox.showerror(
                    "Error",
                    "Código DANE debe tener exactamente 12 dígitos numéricos.",
                    parent=self.win,
                )
                return

            nit = self.entry_nit.get().strip()
            if nit and not re.fullmatch(r"\d+-?\d*|\d+", nit):
                messagebox.showerror(
                    "Error",
                    "NIT debe ser numérico (se permite guion).",
                    parent=self.win,
                )
                return

            correo = self.entry_correo.get().strip()
            if correo and not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", correo):
                messagebox.showerror(
                    "Error",
                    "Correo institucional con formato inválido.",
                    parent=self.win,
                )
                return

            anio_lectivo = (
                self.entry_anio_lectivo.get().strip()
                if hasattr(self, "entry_anio_lectivo")
                else ""
            )
            if not re.fullmatch(r"\d{4}", anio_lectivo):
                messagebox.showerror(
                    "Error",
                    "El Año lectivo activo debe tener 4 dígitos (ej. 2026).",
                    parent=self.win,
                )
                return
            # Guardar el año lectivo en configuracion_plantel
            self._set_config_plantel("anio_lectivo", anio_lectivo)

            # Validar y recolectar escala
            escala_list = []
            for row in self.escala_entries:
                desde_s = row["Desde"].get().strip()
                hasta_s = row["Hasta"].get().strip()
                letra = row["Letra"].get().strip()
                desempeno = row["Desempeño"].get().strip()
                recomend = row["Recomendación"].get().strip()
                if desde_s == "" or hasta_s == "":
                    messagebox.showerror(
                        "Error",
                        "Los valores 'Desde' y 'Hasta' son obligatorios en la escala.",
                        parent=self.win,
                    )
                    return
                try:
                    desde = float(desde_s)
                    hasta = float(hasta_s)
                except Exception:
                    messagebox.showerror(
                        "Error",
                        "Los campos 'Desde' y 'Hasta' deben ser numéricos.",
                        parent=self.win,
                    )
                    return
                if desde > hasta:
                    messagebox.showerror(
                        "Error",
                        "En la escala 'Desde' debe ser menor o igual a 'Hasta'.",
                        parent=self.win,
                    )
                    return
                escala_list.append((desde, hasta, letra, desempeno, recomend))

            # comprobar solapamientos
            escala_sorted = sorted(escala_list, key=lambda x: x[0])
            for i in range(len(escala_sorted) - 1):
                a = escala_sorted[i]
                b = escala_sorted[i + 1]
                if a[1] >= b[0]:
                    messagebox.showerror(
                        "Error",
                        f"Los rangos se cruzan o están adyacentes: {a[0]}-{a[1]} y {b[0]}-{b[1]}",
                        parent=self.win,
                    )
                    return

            # Guardar configuraciones
            self._set_config_plantel("nombre_institucion", nombre)
            self._set_config_plantel("codigo_dane", dane)
            self._set_config_plantel("nit", nit)
            self._set_config_plantel(
                "decreto_funcionamiento", self.entry_decreto.get().strip()
            )
            self._set_config_plantel(
                "resolucion_aprobacion", self.entry_resolucion.get().strip()
            )

            self._set_config_plantel(
                "departamento", self.entry_departamento.get().strip()
            )
            self._set_config_plantel("municipio", self.entry_municipio.get().strip())
            self._set_config_plantel(
                "corregimiento_localidad", self.entry_corregimiento.get().strip()
            )
            self._set_config_plantel("direccion", self.entry_direccion.get().strip())

            # sedes educativas y resumen de jornadas para compatibilidad
            sedes = self._config_plantel_obtener_sedes()
            self._set_config_plantel(
                "sedes_educativas",
                json.dumps(sedes, ensure_ascii=False),
            )

            jornadas = []
            if any(sede.get("jornadas", {}).get("manana") for sede in sedes):
                jornadas.append("Mañana")
            if any(sede.get("jornadas", {}).get("tarde") for sede in sedes):
                jornadas.append("Tarde")
            if any(sede.get("jornadas", {}).get("noche") for sede in sedes):
                jornadas.append("Nocturna")
            self._set_config_plantel("jornadas", ",".join(jornadas))

            # directiva
            rector_label = self.combo_rector_docente.get().strip()
            rector_docente_id = self._rector_docente_map.get(rector_label, "")
            self._set_config_plantel("rector_docente_id", rector_docente_id)
            # Mantener compatibilidad con reportes o flujos que aún consumen los campos legacy.
            self._set_config_plantel(
                "rector_nombre", self.entry_rector_nombre.get().strip()
            )
            self._set_config_plantel(
                "rector_identificacion", self.entry_rector_id.get().strip()
            )
            self._set_config_plantel(
                "rector_cargo", self.entry_rector_cargo.get().strip()
            )
            self._set_config_plantel(
                "secretaria_nombre", self.entry_sec_nombre.get().strip()
            )
            self._set_config_plantel(
                "secretaria_identificacion", self.entry_sec_id.get().strip()
            )
            self._set_config_plantel(
                "secretaria_cargo", self.entry_sec_cargo.get().strip()
            )

            # contacto
            self._set_config_plantel("telefono", self.entry_telefono.get().strip())
            self._set_config_plantel("correo_institucional", correo)
            self._set_config_plantel("dominio_web", self.entry_dominio.get().strip())
            # self.establecer_anio_lectivo_activo(anio_lectivo)  # Ya no se usa, ahora se guarda en configuracion_plantel

            # Escala: reemplazar registros existentes
            self.cur.execute("DELETE FROM escala_valoracion")
            for item in escala_list:
                self.cur.execute(
                    "INSERT INTO escala_valoracion (desde, hasta, letra, desempeno, recomendacion) VALUES (?, ?, ?, ?, ?)",
                    item,
                )
            self._guardar_config_evaluacion(
                {
                    "cognitivo": porcentajes_eval["porcentaje_cognitivo"],
                    "examen": porcentajes_eval["porcentaje_examen"],
                    "autoevaluacion": porcentajes_eval["porcentaje_autoevaluacion"],
                    "cantidad_notas": sum(cantidades_eval.values()),
                    "nota_min": self._coerce_float(
                        self._get_config_plantel("nota_min") or 1.0, 1.0
                    ),
                    "nota_max": self._coerce_float(
                        self._get_config_plantel("nota_max") or 5.0, 5.0
                    ),
                    "decimales": self._coerce_int(
                        self._get_config_plantel("decimales") or 1, 1
                    ),
                    "metodo": self._get_config_plantel("metodo_evaluacion")
                    or "promedio",
                    "notas_cognitivo": cantidades_eval["notas_cognitivo"],
                    "notas_examen": cantidades_eval["notas_examen"],
                    "notas_autoevaluacion": cantidades_eval["notas_autoevaluacion"],
                }
            )
            self.conn.commit()

            messagebox.showinfo(
                "Éxito",
                "Configuración del plantel guardada correctamente.",
                parent=self.win,
            )
        except Exception as e:
            messagebox.showerror(
                "Error", f"Error al guardar configuración: {e}", parent=self.win
            )

    def _limpiar_configuracion_plantel(self):
        """Limpia los campos del formulario de configuracion de plantel."""
        # Limpiar secciones principales
        try:
            for w in (
                self.entry_nombre_institucion,
                self.entry_codigo_dane,
                self.entry_nit,
                self.entry_decreto,
                self.entry_resolucion,
                self.entry_departamento,
                self.entry_municipio,
                self.entry_corregimiento,
                self.entry_direccion,
                self.entry_rector_nombre,
                self.entry_rector_id,
                self.entry_rector_cargo,
                self.entry_sec_nombre,
                self.entry_sec_id,
                self.entry_sec_cargo,
                self.entry_telefono,
                self.entry_correo,
                self.entry_dominio,
            ):
                try:
                    w.delete(0, tk.END)
                except Exception:
                    pass
        except Exception:
            pass

        try:
            if hasattr(self, "entry_anio_lectivo"):
                self.entry_anio_lectivo.delete(0, tk.END)
                self.entry_anio_lectivo.insert(0, self.obtener_anio_lectivo_activo())
        except Exception:
            pass

        # sedes educativas
        try:
            self.sedes_data = [
                {"nombre": "", "manana": False, "tarde": False, "noche": False}
            ]
            self._config_plantel_render_sedes()
        except Exception:
            pass

        # escala
        try:
            for row in getattr(self, "escala_entries", []):
                for e in row.values():
                    try:
                        e.delete(0, tk.END)
                    except Exception:
                        pass
        except Exception:
            pass

    # ---------- Plan de Estudios ----------
    def _build_plan_estudio_tab(self):
        frame = self.tab_plan_estudio

        # --- Barra de herramientas ---
        toolbar = ttk.Frame(frame)
        toolbar.pack(fill="x", padx=8, pady=(8, 4))

        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.plan_estudio.importar",
            text="📂 Importar (Excel/CSV)",
            command=self.importar_plan_estudio,
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.plan_estudio.agregar",
            text="➕ Agregar",
            command=self._plan_agregar,
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.plan_estudio.editar",
            text="✏️ Editar",
            command=self._plan_editar,
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.plan_estudio.copiar",
            text="📋 Copiar Plan",
            command=self._plan_copiar_plan,
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.plan_estudio.catalogo_areas",
            text="Catálogo de Áreas",
            command=self._plan_abrir_catalogo_areas,
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.plan_estudio.eliminar",
            text="🗑️ Eliminar seleccionado",
            command=self._plan_eliminar_seleccionado,
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.plan_estudio.vaciar",
            text="🗑️ Vaciar todo",
            command=self._plan_vaciar_todo,
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.plan_estudio.actualizar",
            text="🔄 Actualizar",
            command=self._plan_load_treeview,
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )

        # --- Filtros ---
        filtros = ttk.Frame(frame)
        filtros.pack(fill="x", padx=8, pady=(0, 4))

        ttk.Label(filtros, text="Nivel:").pack(side="left")
        self.plan_filter_nivel = ttk.Combobox(filtros, state="readonly", width=14)
        self.plan_filter_nivel.pack(side="left", padx=(4, 10))

        ttk.Label(filtros, text="Grado:").pack(side="left")
        self.plan_filter_grado = ttk.Combobox(filtros, state="readonly", width=10)
        self.plan_filter_grado.pack(side="left", padx=(4, 10))

        ttk.Label(filtros, text="Curso:").pack(side="left")
        self.plan_filter_curso = ttk.Combobox(filtros, state="readonly", width=10)
        self.plan_filter_curso.pack(side="left", padx=(4, 10))

        ttk.Label(filtros, text="Área:").pack(side="left")
        self.plan_filter_area = ttk.Combobox(filtros, state="readonly", width=22)
        self.plan_filter_area.pack(side="left", padx=(4, 10))

        ttk.Button(
            filtros, text="Limpiar filtros", command=self._plan_limpiar_filtros
        ).pack(side="left", padx=(6, 0))

        for cb in (
            self.plan_filter_nivel,
            self.plan_filter_grado,
            self.plan_filter_curso,
            self.plan_filter_area,
        ):
            cb.bind("<<ComboboxSelected>>", lambda e: self._plan_load_treeview())

        # --- Treeview ---
        cols = ("id", "nivel", "grado", "curso", "area", "horas", "estado")
        self.tree_plan = ttk.Treeview(frame, columns=cols, show="headings")

        headers = {
            "id": "ID",
            "nivel": "Nivel",
            "grado": "Grado",
            "curso": "Curso",
            "area": "Área",
            "horas": "Horas",
            "estado": "Estado",
        }
        widths = {
            "id": 50,
            "nivel": 110,
            "grado": 80,
            "curso": 80,
            "area": 220,
            "horas": 70,
            "estado": 80,
        }
        for c in cols:
            self.tree_plan.heading(c, text=headers[c])
            self.tree_plan.column(
                c,
                width=widths[c],
                anchor=(
                    "center"
                    if c in ("id", "grado", "curso", "horas", "estado")
                    else "w"
                ),
                stretch=(c == "area"),
            )

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree_plan.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.tree_plan.xview)
        self.tree_plan.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree_plan.pack(fill="both", expand=True, padx=8, pady=(0, 2))

        # Contador de registros
        self.plan_lbl_total = ttk.Label(frame, text="")
        self.plan_lbl_total.pack(anchor="e", padx=10, pady=(2, 4))

        self._plan_refresh_filtros()
        self._plan_load_treeview()

    def _plan_refresh_filtros(self):
        """Recarga los valores disponibles en los combos de filtro."""
        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()

            cur.execute(
                "SELECT DISTINCT nivel FROM plan_estudio "
                "WHERE nivel IS NOT NULL ORDER BY nivel"
            )
            niveles = [""] + [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT grado FROM plan_estudio "
                "WHERE grado IS NOT NULL ORDER BY grado"
            )
            grados = [""] + [str(r[0]) for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT curso FROM plan_estudio "
                "WHERE curso IS NOT NULL ORDER BY curso"
            )
            cursos = [""] + [str(r[0]) for r in cur.fetchall()]

            cur.execute(
                """
                SELECT DISTINCT COALESCE(a.nombre, p.area)
                FROM plan_estudio p
                LEFT JOIN areas a ON a.id = p.IdArea
                WHERE COALESCE(a.nombre, p.area) IS NOT NULL
                  AND TRIM(COALESCE(a.nombre, p.area)) <> ''
                ORDER BY COALESCE(a.nombre, p.area)
                """
            )
            areas = [""] + [r[0] for r in cur.fetchall()]

            conn.close()
        except Exception:
            niveles = grados = cursos = areas = [""]

        self.plan_filter_nivel["values"] = niveles
        self.plan_filter_grado["values"] = grados
        self.plan_filter_curso["values"] = cursos
        self.plan_filter_area["values"] = areas

    def _plan_load_treeview(self):
        """Carga/recarga el Treeview aplicando los filtros activos."""
        for item in self.tree_plan.get_children():
            self.tree_plan.delete(item)

        nivel = self.plan_filter_nivel.get().strip()
        grado = self.plan_filter_grado.get().strip()
        curso = self.plan_filter_curso.get().strip()
        area = self.plan_filter_area.get().strip()

        conditions = []
        params = []
        if nivel:
            conditions.append("p.nivel=?")
            params.append(nivel)
        if grado:
            conditions.append("TRIM(CAST(p.grado AS TEXT))=?")
            params.append(grado)
        if curso:
            conditions.append("TRIM(CAST(p.curso AS TEXT))=?")
            params.append(curso)
        if area:
            conditions.append("COALESCE(a.nombre, p.area)=?")
            params.append(area)

        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        sql = (
            f"SELECT p.id, p.nivel, p.grado, p.curso, COALESCE(a.nombre, p.area) AS area, "
            f"p.horas, p.estado "
            f"FROM plan_estudio p "
            f"LEFT JOIN areas a ON a.id = p.IdArea "
            f"{where} "
            f"ORDER BY p.grado, p.curso, COALESCE(a.nombre, p.area)"
        )

        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute(sql, params)
            rows = cur.fetchall()
            conn.close()
        except Exception as exc:
            messagebox.showerror(
                "Plan de Estudios", f"Error al cargar datos:\n{exc}", parent=self.win
            )
            return

        for row in rows:
            id_, nivel_, grado_, curso_, area_, horas_, estado_ = row
            estado_txt = "Activo" if estado_ else "Inactivo"
            self.tree_plan.insert(
                "",
                "end",
                values=(
                    id_,
                    nivel_ or "",
                    grado_ or "",
                    curso_ or "",
                    area_ or "",
                    horas_ or 0,
                    estado_txt,
                ),
            )

        total = len(rows)
        self.plan_lbl_total.config(text=f"Total: {total} registro(s)")
        self._plan_refresh_filtros()

    def _plan_limpiar_filtros(self):
        for cb in (
            self.plan_filter_nivel,
            self.plan_filter_grado,
            self.plan_filter_curso,
            self.plan_filter_area,
        ):
            cb.set("")
        self._plan_load_treeview()

    def _plan_agregar(self):
        """Abre formulario para agregar varias áreas a un curso."""
        if not self._requiere_permiso("desktop.superadmin.plan_estudio.agregar"):
            return
        self._plan_abrir_formulario_agregar_masivo()

    def _plan_niveles_grados(self):
        return {
            "Preescolar": ["0", "JA", "PREJ"],
            "Básica Primaria (1° a 5°)": ["1", "2", "3", "4", "5"],
            "Básica Secundaria (6° a 9°)": ["6", "7", "8", "9"],
            "Educación Media (10° y 11°)": ["10", "11"],
            "Educación por Ciclos (CLEI)": ["C1", "C2", "C3", "C4", "C5", "C6"],
        }

    def _plan_normalizar_grado(self, valor):
        raw = str(valor or "").strip().upper()
        if not raw:
            raise ValueError("Grado vacío")
        if raw in {"JA", "JARDIN", "JARDÍN"}:
            return "JA"
        if raw in {"PREJ", "PREJARDIN", "PREJARDÍN"}:
            return "PREJ"
        if raw in {"0", "TRANSICION", "TRANSICIÓN", "TRANSICION PREESCOLAR"}:
            return "0"
        if re.fullmatch(r"C[1-6]", raw):
            return raw
        if re.fullmatch(r"\d+", raw):
            return str(int(raw))
        raise ValueError("Grado inválido")

    def _plan_normalizar_curso(self, valor):
        raw = str(valor or "").strip().upper()
        if not raw:
            raise ValueError("Curso vacío")
        if re.fullmatch(r"C[1-6]", raw):
            return raw
        if re.fullmatch(r"\d+", raw):
            return str(int(raw))
        raise ValueError("Curso inválido")

    def _plan_parse_curso_compuesto(self, curso_raw):
        """Convierte 'GG-CC' en (grado, curso), soportando CLEI."""
        raw = str(curso_raw or "").strip().upper()
        if not raw:
            raise ValueError("Curso vacío")
        if "-" in raw:
            g_raw, c_raw = raw.split("-", 1)
            return self._plan_normalizar_grado(g_raw), self._plan_normalizar_curso(
                c_raw
            )

        grado = self._plan_normalizar_grado(raw)
        # Compatibilidad: cuando solo viene el grado, curso por defecto = 1.
        if re.fullmatch(r"C[1-6]|JA", grado):
            return grado, grado if re.fullmatch(r"C[1-6]", grado) else "1"
        return grado, "1"

    def _plan_abrir_catalogo_areas(self):
        """Abre una ventana dedicada para gestionar el catálogo de áreas."""
        if not self._requiere_permiso("desktop.superadmin.plan_estudio.catalogo_areas"):
            return
        win_existente = getattr(self, "win_catalogo_areas", None)
        if win_existente is not None:
            try:
                if win_existente.winfo_exists():
                    win_existente.deiconify()
                    win_existente.lift()
                    win_existente.focus_force()
                    self._areas_load_treeview()
                    return
            except Exception:
                pass

        win = tk.Toplevel(self.win)
        win.title("Catálogo de Áreas")
        win.geometry("620x460")
        win.transient(self.win)

        self.win_catalogo_areas = win

        toolbar = ttk.Frame(win)
        toolbar.pack(fill="x", padx=10, pady=(10, 6))

        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.areas.agregar",
            text="Agregar Área",
            command=self._area_agregar,
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.areas.editar",
            text="Editar Área",
            command=self._area_editar,
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.areas.eliminar",
            text="Eliminar Área",
            command=self._area_eliminar,
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.areas.importar",
            text="Importar CSV",
            command=self._area_importar_csv,
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )
        self._crear_boton_si_permiso(
            ttk.Button,
            toolbar,
            "desktop.superadmin.areas.actualizar",
            text="Actualizar",
            command=self._areas_load_treeview,
            layout_kwargs={"side": "left", "padx": (0, 6)},
        )

        tree_wrap = ttk.Frame(win)
        tree_wrap.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        cols = ("id", "nombre")
        self.tree_areas_catalog = ttk.Treeview(tree_wrap, columns=cols, show="headings")
        self.tree_areas_catalog.heading("id", text="ID")
        self.tree_areas_catalog.heading("nombre", text="Área")
        self.tree_areas_catalog.column("id", width=70, anchor="center", stretch=False)
        self.tree_areas_catalog.column("nombre", width=420, anchor="w", stretch=True)

        vsb = ttk.Scrollbar(
            tree_wrap, orient="vertical", command=self.tree_areas_catalog.yview
        )
        self.tree_areas_catalog.configure(yscrollcommand=vsb.set)

        self.tree_areas_catalog.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        def _on_close_catalogo():
            try:
                win.destroy()
            finally:
                self.win_catalogo_areas = None
                self.tree_areas_catalog = None

        win.protocol("WM_DELETE_WINDOW", _on_close_catalogo)
        self._areas_load_treeview()

    def cargar_areas(self):
        """Devuelve áreas activas del catálogo general de áreas."""
        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute(
                """
                SELECT nombre
                FROM areas
                WHERE estado=1
                ORDER BY nombre
                """
            )
            rows = [str(r[0]).strip() for r in cur.fetchall() if r and r[0]]
            conn.close()
            return rows
        except Exception:
            return []

    def _area_parent(self):
        win = getattr(self, "win_catalogo_areas", None)
        if win is not None:
            try:
                if win.winfo_exists():
                    return win
            except Exception:
                pass
        return self.win

    def _areas_get_tree(self):
        tree = getattr(self, "tree_areas_catalog", None)
        if tree is None:
            return None
        try:
            if tree.winfo_exists():
                return tree
        except Exception:
            pass
        return None

    def _areas_load_treeview(self):
        tree = self._areas_get_tree()
        if tree is None:
            return
        for item in tree.get_children():
            tree.delete(item)
        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute("SELECT id, nombre FROM areas ORDER BY id")
            rows = cur.fetchall()
            conn.close()
        except Exception:
            rows = []

        for area_id, nombre in rows:
            tree.insert("", "end", values=(area_id, nombre or ""))

    def _area_agregar(self):
        if not self._requiere_permiso("desktop.superadmin.areas.agregar"):
            return
        self._area_abrir_formulario(modo="agregar")

    def _area_editar(self):
        if not self._requiere_permiso("desktop.superadmin.areas.editar"):
            return
        tree = self._areas_get_tree()
        if tree is None:
            return
        sel = tree.selection()
        if not sel:
            messagebox.showwarning(
                "Áreas",
                "Seleccione un área para editar.",
                parent=self._area_parent(),
            )
            return
        valores = tree.item(sel[0], "values")
        self._area_abrir_formulario(modo="editar", valores=valores)

    def _area_abrir_formulario(self, modo="agregar", valores=None):
        titulo = "Agregar Área" if modo == "agregar" else "Editar Área"
        dlg = tk.Toplevel(self._area_parent())
        dlg.title(titulo)
        dlg.transient(self._area_parent())
        dlg.grab_set()
        dlg.resizable(False, False)

        frm = ttk.Frame(dlg)
        frm.pack(fill="both", expand=True, padx=12, pady=10)

        ttk.Label(frm, text="Nombre del área:").grid(
            row=0, column=0, sticky="e", padx=(0, 6), pady=4
        )
        var_nombre = tk.StringVar()
        ttk.Entry(frm, textvariable=var_nombre, width=32).grid(
            row=0, column=1, sticky="w", pady=4
        )

        area_id = None
        if modo == "editar" and valores:
            area_id = int(valores[0])
            var_nombre.set(str(valores[1] or "").strip())

        def _guardar():
            nombre = var_nombre.get().strip()
            if not nombre:
                messagebox.showwarning(
                    "Validación",
                    "El nombre del área es obligatorio.",
                    parent=dlg,
                )
                return
            try:
                conn = sqlite3.connect(self.db_path)
                cur = conn.cursor()
                if modo == "agregar":
                    cur.execute("INSERT INTO areas (nombre) VALUES (?)", (nombre,))
                else:
                    cur.execute(
                        "UPDATE areas SET nombre=? WHERE id=?",
                        (nombre, area_id),
                    )
                conn.commit()
                conn.close()
            except sqlite3.IntegrityError:
                messagebox.showwarning(
                    "Áreas",
                    "Ya existe un área con ese nombre.",
                    parent=dlg,
                )
                return
            except Exception as exc:
                messagebox.showerror(
                    "Áreas",
                    f"Error al guardar área:\n{exc}",
                    parent=dlg,
                )
                return

            dlg.destroy()
            self._areas_load_treeview()

        btns = ttk.Frame(frm)
        btns.grid(row=1, column=0, columnspan=2, sticky="w", pady=(8, 0))
        ttk.Button(btns, text="Guardar", command=_guardar).pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(btns, text="Cancelar", command=dlg.destroy).pack(side="left")

        dlg.wait_window()

    def _area_eliminar(self):
        if not self._requiere_permiso("desktop.superadmin.areas.eliminar"):
            return
        tree = self._areas_get_tree()
        if tree is None:
            return
        sel = tree.selection()
        if not sel:
            messagebox.showwarning(
                "Áreas",
                "Seleccione un área para eliminar.",
                parent=self._area_parent(),
            )
            return

        valores = tree.item(sel[0], "values")
        if not valores:
            return
        area_id = int(valores[0])
        nombre = str(valores[1] or "").strip()

        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute(
                """
                SELECT id
                FROM plan_estudio
                WHERE CAST(COALESCE(IdArea, '') AS TEXT) = CAST(? AS TEXT)
                   OR (
                        (IdArea IS NULL OR TRIM(CAST(IdArea AS TEXT)) = '')
                        AND LOWER(TRIM(COALESCE(area, ''))) = LOWER(TRIM(?))
                   )
                LIMIT 1
                """,
                (area_id, nombre),
            )
            en_uso = cur.fetchone() is not None
            if en_uso:
                conn.close()
                messagebox.showwarning(
                    "Áreas",
                    "Esta área está registrada en el plan de estudio.",
                    parent=self._area_parent(),
                )
                return

            if not messagebox.askyesno(
                "Confirmar",
                f"¿Eliminar el área '{nombre}'?",
                parent=self._area_parent(),
            ):
                conn.close()
                return

            cur.execute("DELETE FROM areas WHERE id=?", (area_id,))
            conn.commit()
            conn.close()
        except Exception as exc:
            messagebox.showerror(
                "Áreas",
                f"Error al eliminar área:\n{exc}",
                parent=self._area_parent(),
            )
            return

        self._areas_load_treeview()

    def _area_importar_csv(self):
        if not self._requiere_permiso("desktop.superadmin.areas.importar"):
            return
        parent = self._area_parent()
        path = filedialog.askopenfilename(
            title="Importar catálogo de áreas (CSV)",
            filetypes=[("CSV", "*.csv"), ("Todos", "*.*")],
            parent=parent,
        )
        if not path:
            return

        import csv

        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                muestra = f.read(4096)
            sep = ","
            for candidato in ("|", ";", "\t", ","):
                if candidato in muestra:
                    sep = candidato
                    break

            nombres = []
            with open(path, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f, delimiter=sep)
                if reader.fieldnames:
                    for row in reader:
                        fila = {
                            str(k).strip().lower().replace(" ", "_"): v
                            for k, v in row.items()
                            if k is not None
                        }
                        nombre = ""
                        for key in ("nombre", "area", "área"):
                            if key in fila and fila[key] is not None:
                                nombre = str(fila[key]).strip()
                                if nombre:
                                    break
                        if not nombre:
                            for val in fila.values():
                                if val is not None and str(val).strip():
                                    nombre = str(val).strip()
                                    break
                        if nombre:
                            nombres.append(nombre)
                else:
                    f.seek(0)
                    raw_reader = csv.reader(f, delimiter=sep)
                    for row in raw_reader:
                        if not row:
                            continue
                        nombre = str(row[0]).strip()
                        if nombre:
                            nombres.append(nombre)
        except Exception as exc:
            messagebox.showerror(
                "Áreas",
                f"Error al leer CSV:\n{exc}",
                parent=parent,
            )
            return

        if not nombres:
            messagebox.showwarning(
                "Áreas",
                "No se encontraron nombres de áreas válidos en el archivo.",
                parent=parent,
            )
            return

        insertados = 0
        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            for nombre in nombres:
                cur.execute(
                    "INSERT OR IGNORE INTO areas (nombre) VALUES (?)", (nombre,)
                )
                if cur.rowcount > 0:
                    insertados += 1
            conn.commit()
            conn.close()
        except Exception as exc:
            messagebox.showerror(
                "Áreas",
                f"Error al importar áreas:\n{exc}",
                parent=parent,
            )
            return

        self._areas_load_treeview()
        omitidos = len(nombres) - insertados
        messagebox.showinfo(
            "Áreas",
            f"Importación completada.\nInsertadas: {insertados}\nOmitidas: {omitidos}",
            parent=parent,
        )

    def _plan_catalogo_areas_disponibles(self):
        """Devuelve catálogo activo como (id_area, nombre, horas_defecto)."""
        base = [
            (7, "Matemáticas", 5),
            (14, "Español", 5),
            (8, "Ciencias Sociales", 4),
            (16, "Ciencias Naturales", 4),
            (9, "Idioma Extranjero", 3),
            (11, "Tecnología e Informática", 2),
            (15, "Educación Física", 2),
            (12, "Educación Artística", 2),
        ]

        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute(
                """
                SELECT a.id, a.nombre
                FROM areas a
                WHERE a.estado = 1
                ORDER BY a.nombre
                """
            )
            catalogo_rows = cur.fetchall()

            cur.execute(
                """
                SELECT IdArea, COALESCE(MAX(horas), 0)
                FROM plan_estudio
                WHERE IdArea IS NOT NULL
                GROUP BY IdArea
                """
            )
            horas_por_id = {
                int(r[0]): int(r[1] or 0) for r in cur.fetchall() if r[0] is not None
            }

            conn.close()

            if catalogo_rows:
                return [
                    (int(area_id), str(nombre), int(horas_por_id.get(int(area_id), 0)))
                    for area_id, nombre in catalogo_rows
                ]
        except Exception:
            pass
        return base

    def _plan_abrir_formulario_agregar_masivo(self):
        """Formulario para registrar múltiples áreas en un mismo curso."""
        if not self._requiere_permiso("desktop.superadmin.plan_estudio.agregar"):
            return
        dlg = tk.Toplevel(self.win)
        dlg.title("Agregar Plan de Estudio")
        dlg.transient(self.win)
        dlg.grab_set()
        dlg.geometry("560x520")

        top = ttk.Frame(dlg)
        top.pack(fill="x", padx=10, pady=(10, 6))

        niveles_grados = self._plan_niveles_grados()
        lista_niveles = list(niveles_grados.keys())

        ttk.Label(top, text="Nivel:").grid(
            row=0, column=0, sticky="e", padx=(0, 6), pady=4
        )
        var_nivel = tk.StringVar()
        cb_nivel = ttk.Combobox(
            top,
            textvariable=var_nivel,
            state="readonly",
            values=lista_niveles,
            width=28,
        )
        cb_nivel.grid(row=0, column=1, sticky="w", pady=4)

        ttk.Label(top, text="Grado:").grid(
            row=0, column=2, sticky="e", padx=(12, 6), pady=4
        )
        var_grado = tk.StringVar()
        cb_grado = ttk.Combobox(top, textvariable=var_grado, state="readonly", width=10)
        cb_grado.grid(row=0, column=3, sticky="w", pady=4)

        ttk.Label(top, text="Curso:").grid(
            row=0, column=4, sticky="e", padx=(12, 6), pady=4
        )
        var_curso = tk.StringVar()
        cb_curso = ttk.Combobox(top, textvariable=var_curso, state="normal", width=10)
        cb_curso.grid(row=0, column=5, sticky="w", pady=4)

        def _actualizar_grados(*_):
            nivel_sel = var_nivel.get().strip()
            grados = niveles_grados.get(nivel_sel, [])
            cb_grado["values"] = grados
            var_grado.set("")
            var_curso.set("")
            cb_curso["values"] = []

        def _actualizar_cursos(*_):
            grado_sel = var_grado.get().strip().upper()
            if re.fullmatch(r"C[1-6]", grado_sel):
                cb_curso["values"] = ["C1", "C2", "C3", "C4", "C5", "C6"]
            else:
                cb_curso["values"] = [str(i) for i in range(1, 11)]

        cb_nivel.bind("<<ComboboxSelected>>", _actualizar_grados)
        cb_grado.bind("<<ComboboxSelected>>", _actualizar_cursos)

        ttk.Separator(dlg, orient="horizontal").pack(fill="x", padx=10, pady=(0, 6))
        ttk.Label(dlg, text="Seleccione áreas y defina horas:").pack(
            anchor="w", padx=12
        )

        holder = ttk.Frame(dlg)
        holder.pack(fill="both", expand=True, padx=10, pady=(6, 8))

        canvas = tk.Canvas(holder, highlightthickness=0)
        scrollbar = ttk.Scrollbar(holder, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas)
        inner.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")),
        )
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        ttk.Label(inner, text="Seleccionar", width=12).grid(
            row=0, column=0, sticky="w", padx=6
        )
        ttk.Label(inner, text="Área", width=28).grid(
            row=0, column=1, sticky="w", padx=6
        )
        ttk.Label(inner, text="Horas", width=10).grid(
            row=0, column=2, sticky="w", padx=6
        )

        areas_config = []
        for idx, (id_area, area_nombre, horas_def) in enumerate(
            self._plan_catalogo_areas_disponibles(), start=1
        ):
            var_sel = tk.BooleanVar(value=False)
            var_horas = tk.StringVar(value=str(int(horas_def or 0)))
            ttk.Checkbutton(inner, variable=var_sel).grid(
                row=idx, column=0, sticky="w", padx=6, pady=2
            )
            ttk.Label(inner, text=area_nombre).grid(
                row=idx, column=1, sticky="w", padx=6, pady=2
            )
            ttk.Entry(inner, textvariable=var_horas, width=8).grid(
                row=idx, column=2, sticky="w", padx=6, pady=2
            )
            areas_config.append((int(id_area), area_nombre, var_sel, var_horas))

        def _guardar():
            nivel = var_nivel.get().strip()
            grado_raw = var_grado.get().strip()
            curso_raw = var_curso.get().strip()

            if not nivel:
                messagebox.showwarning("Validación", "Seleccione un nivel.", parent=dlg)
                return

            if not grado_raw:
                messagebox.showwarning(
                    "Validación", "El campo Grado es obligatorio.", parent=dlg
                )
                return
            if not curso_raw:
                messagebox.showwarning(
                    "Validación", "El campo Curso es obligatorio.", parent=dlg
                )
                return

            try:
                grado_val = self._plan_normalizar_grado(grado_raw)
            except ValueError:
                messagebox.showwarning(
                    "Validación",
                    "Grado inválido para el nivel seleccionado.",
                    parent=dlg,
                )
                return
            try:
                curso_val = self._plan_normalizar_curso(curso_raw)
            except ValueError:
                messagebox.showwarning("Validación", "Curso inválido.", parent=dlg)
                return

            seleccionadas = []
            for id_area, area_nombre, var_sel, var_horas in areas_config:
                if not var_sel.get():
                    continue
                horas_raw = var_horas.get().strip()
                try:
                    horas_val = int(float(horas_raw)) if horas_raw else 0
                except ValueError:
                    messagebox.showwarning(
                        "Validación",
                        f"Horas inválidas para el área '{area_nombre}'.",
                        parent=dlg,
                    )
                    return
                seleccionadas.append((id_area, area_nombre, horas_val))

            if not seleccionadas:
                messagebox.showwarning(
                    "Plan de Estudios",
                    "Seleccione al menos un área para guardar.",
                    parent=dlg,
                )
                return

            insertados = 0
            duplicados = []
            try:
                conn = sqlite3.connect(self.db_path)
                cur = conn.cursor()
                for id_area, area_nombre, horas_val in seleccionadas:
                    cur.execute(
                        """
                        SELECT id
                        FROM plan_estudio
                        WHERE grado=? AND curso=?
                          AND (
                                CAST(COALESCE(IdArea,'') AS TEXT)=CAST(? AS TEXT)
                                OR (
                                    (IdArea IS NULL OR TRIM(CAST(IdArea AS TEXT))='')
                                    AND LOWER(TRIM(COALESCE(area,'')))=LOWER(TRIM(?))
                                )
                          )
                        """,
                        (grado_val, curso_val, id_area, area_nombre),
                    )
                    if cur.fetchone():
                        duplicados.append(area_nombre)
                        continue

                    cur.execute(
                        "INSERT INTO plan_estudio (nivel, grado, curso, IdArea, area, horas) "
                        "VALUES (?, ?, ?, ?, ?, ?)",
                        (
                            nivel,
                            grado_val,
                            curso_val,
                            int(id_area),
                            area_nombre,
                            horas_val,
                        ),
                    )
                    insertados += 1
                conn.commit()
                conn.close()
            except Exception as exc:
                messagebox.showerror(
                    "Plan de Estudios",
                    f"Error al guardar registros:\n{exc}",
                    parent=dlg,
                )
                return

            dlg.destroy()
            self._plan_load_treeview()

            if insertados == 0 and duplicados:
                messagebox.showwarning(
                    "Plan de Estudios",
                    "El área ya existe en este grado y curso.",
                    parent=self.win,
                )
                return

            msg = f"Áreas insertadas: {insertados}."
            if duplicados:
                msg += "\nOmitidas por duplicado: " + ", ".join(duplicados)
            messagebox.showinfo("Plan de Estudios", msg, parent=self.win)

        bot = ttk.Frame(dlg)
        bot.pack(fill="x", padx=10, pady=(2, 10))
        ttk.Button(bot, text="Guardar", command=_guardar).pack(side="left", padx=(0, 6))
        ttk.Button(bot, text="Cancelar", command=dlg.destroy).pack(side="left")

        dlg.wait_window()

    def _plan_copiar_plan(self):
        """Copia todas las áreas/horas desde un curso origen a uno destino."""
        if not self._requiere_permiso("desktop.superadmin.plan_estudio.copiar"):
            return
        dlg = tk.Toplevel(self.win)
        dlg.title("Copiar Plan de Estudios")
        dlg.transient(self.win)
        dlg.grab_set()
        dlg.resizable(False, False)

        frm = ttk.Frame(dlg)
        frm.pack(fill="both", expand=True, padx=12, pady=10)

        ttk.Label(frm, text="Curso origen", font=("Segoe UI", 10, "bold")).grid(
            row=0, column=0, columnspan=4, sticky="w", pady=(0, 6)
        )
        ttk.Label(frm, text="Grado:").grid(
            row=1, column=0, sticky="e", padx=(0, 6), pady=3
        )
        var_g_origen = tk.StringVar()
        ttk.Entry(frm, textvariable=var_g_origen, width=10).grid(
            row=1, column=1, sticky="w", pady=3
        )
        ttk.Label(frm, text="Curso:").grid(
            row=1, column=2, sticky="e", padx=(12, 6), pady=3
        )
        var_c_origen = tk.StringVar()
        ttk.Entry(frm, textvariable=var_c_origen, width=10).grid(
            row=1, column=3, sticky="w", pady=3
        )

        ttk.Separator(frm, orient="horizontal").grid(
            row=2, column=0, columnspan=4, sticky="ew", pady=8
        )

        ttk.Label(frm, text="Curso destino", font=("Segoe UI", 10, "bold")).grid(
            row=3, column=0, columnspan=4, sticky="w", pady=(0, 6)
        )
        ttk.Label(frm, text="Grado:").grid(
            row=4, column=0, sticky="e", padx=(0, 6), pady=3
        )
        var_g_destino = tk.StringVar()
        ttk.Entry(frm, textvariable=var_g_destino, width=10).grid(
            row=4, column=1, sticky="w", pady=3
        )
        ttk.Label(frm, text="Curso:").grid(
            row=4, column=2, sticky="e", padx=(12, 6), pady=3
        )
        var_c_destino = tk.StringVar()
        ttk.Entry(frm, textvariable=var_c_destino, width=10).grid(
            row=4, column=3, sticky="w", pady=3
        )

        def _confirmar_reemplazo(cantidad):
            ask = tk.Toplevel(dlg)
            ask.title("Curso destino con plan existente")
            ask.transient(dlg)
            ask.grab_set()
            ask.resizable(False, False)

            ttk.Label(
                ask,
                text=(
                    "El curso destino ya tiene un plan de estudios registrado.\n"
                    f"Registros actuales: {cantidad}."
                ),
                justify="left",
            ).pack(anchor="w", padx=12, pady=(12, 8))

            decision = {"reemplazar": False}

            btns = ttk.Frame(ask)
            btns.pack(fill="x", padx=12, pady=(0, 12))

            def _si():
                decision["reemplazar"] = True
                ask.destroy()

            def _no():
                ask.destroy()

            ttk.Button(btns, text="Reemplazar", command=_si).pack(
                side="left", padx=(0, 6)
            )
            ttk.Button(btns, text="Cancelar", command=_no).pack(side="left")

            ask.wait_window()
            return decision["reemplazar"]

        def _copiar():
            try:
                g_origen = self._plan_normalizar_grado(var_g_origen.get().strip())
                c_origen = self._plan_normalizar_curso(var_c_origen.get().strip())
                g_destino = self._plan_normalizar_grado(var_g_destino.get().strip())
                c_destino = self._plan_normalizar_curso(var_c_destino.get().strip())
            except ValueError:
                messagebox.showwarning(
                    "Validación",
                    "Grado y Curso inválidos en origen o destino.",
                    parent=dlg,
                )
                return

            try:
                conn = sqlite3.connect(self.db_path)
                cur = conn.cursor()

                cur.execute(
                    "SELECT nivel, area, IdArea, horas FROM plan_estudio "
                    "WHERE grado=? AND curso=? ORDER BY area",
                    (g_origen, c_origen),
                )
                origen_rows = cur.fetchall()

                if not origen_rows:
                    conn.close()
                    messagebox.showwarning(
                        "Plan de Estudios",
                        "El curso origen no tiene áreas registradas.",
                        parent=dlg,
                    )
                    return

                cur.execute(
                    "SELECT COUNT(*) FROM plan_estudio WHERE grado=? AND curso=?",
                    (g_destino, c_destino),
                )
                destino_count = int(cur.fetchone()[0] or 0)

                if destino_count > 0:
                    if not _confirmar_reemplazo(destino_count):
                        conn.close()
                        return
                    cur.execute(
                        "DELETE FROM plan_estudio WHERE grado=? AND curso=?",
                        (g_destino, c_destino),
                    )

                for nivel, area, id_area, horas in origen_rows:
                    id_area_val = id_area
                    if id_area_val is None and str(area or "").strip():
                        cur.execute(
                            "SELECT id FROM areas WHERE LOWER(TRIM(nombre))=LOWER(TRIM(?)) LIMIT 1",
                            (str(area).strip(),),
                        )
                        r_area = cur.fetchone()
                        if r_area:
                            id_area_val = int(r_area[0])

                    cur.execute(
                        "INSERT INTO plan_estudio (nivel, grado, curso, IdArea, area, horas) "
                        "VALUES (?, ?, ?, ?, ?, ?)",
                        (
                            nivel,
                            g_destino,
                            c_destino,
                            id_area_val,
                            area,
                            int(horas or 0),
                        ),
                    )

                conn.commit()
                conn.close()
            except Exception as exc:
                messagebox.showerror(
                    "Plan de Estudios",
                    f"Error al copiar plan:\n{exc}",
                    parent=dlg,
                )
                return

            dlg.destroy()
            self._plan_load_treeview()
            messagebox.showinfo(
                "Plan de Estudios",
                "Plan copiado correctamente.",
                parent=self.win,
            )

        btns = ttk.Frame(frm)
        btns.grid(row=5, column=0, columnspan=4, sticky="w", pady=(10, 0))
        ttk.Button(btns, text="Copiar", command=_copiar).pack(side="left", padx=(0, 6))
        ttk.Button(btns, text="Cancelar", command=dlg.destroy).pack(side="left")

        dlg.wait_window()

    def _plan_editar(self):
        """Abre el formulario cargando los datos del registro seleccionado."""
        if not self._requiere_permiso("desktop.superadmin.plan_estudio.editar"):
            return
        sel = self.tree_plan.selection()
        if not sel:
            messagebox.showwarning(
                "Plan de Estudios",
                "Seleccione un registro para editar.",
                parent=self.win,
            )
            return
        valores = self.tree_plan.item(sel[0], "values")
        # valores: (id, nivel, grado, curso, area, horas, estado)
        self._plan_abrir_formulario(modo="editar", valores=valores)

    def _plan_abrir_formulario(self, modo="agregar", valores=None):
        """Formulario modal para agregar o editar un registro de plan de estudio.

        Args:
            modo: 'agregar' o 'editar'.
            valores: tupla (id, nivel, grado, curso, area, horas, estado) cuando
                     modo='editar'.
        """
        permiso = (
            "desktop.superadmin.plan_estudio.agregar"
            if modo == "agregar"
            else "desktop.superadmin.plan_estudio.editar"
        )
        if not self._requiere_permiso(permiso):
            return
        titulo = (
            "Agregar Plan de Estudio" if modo == "agregar" else "Editar Plan de Estudio"
        )
        dlg = tk.Toplevel(self.win)
        dlg.title(titulo)
        dlg.resizable(False, False)
        dlg.transient(self.win)
        dlg.grab_set()

        pad = {"padx": 10, "pady": 5}
        campos = [
            ("Nivel:", "nivel"),
            ("Grado:", "grado"),
            ("Curso:", "curso"),
            ("Área:", "area"),
            ("Horas:", "horas"),
        ]
        entradas = {}
        for i, (label, key) in enumerate(campos):
            ttk.Label(dlg, text=label, anchor="e", width=8).grid(
                row=i, column=0, sticky="e", **pad
            )
            var = tk.StringVar()
            if key == "area":
                entry = ttk.Combobox(
                    dlg,
                    textvariable=var,
                    values=self.cargar_areas(),
                    state="readonly",
                    width=26,
                )
            else:
                entry = ttk.Entry(dlg, textvariable=var, width=28)
            entry.grid(row=i, column=1, sticky="ew", **pad)
            entradas[key] = var

        # Pre-cargar valores si es modo editar
        id_registro = None
        if modo == "editar" and valores:
            id_registro = valores[0]
            entradas["nivel"].set(valores[1] if valores[1] else "")
            entradas["grado"].set(valores[2] if valores[2] else "")
            entradas["curso"].set(valores[3] if valores[3] else "")
            entradas["area"].set(valores[4] if valores[4] else "")
            entradas["horas"].set(valores[5] if valores[5] else "")

        def _guardar():
            nivel = entradas["nivel"].get().strip()
            grado_raw = entradas["grado"].get().strip()
            curso_raw = entradas["curso"].get().strip()
            area = entradas["area"].get().strip()
            horas_raw = entradas["horas"].get().strip()

            # Validaciones básicas
            if not grado_raw:
                messagebox.showwarning(
                    "Validación", "El campo Grado es obligatorio.", parent=dlg
                )
                return
            if not curso_raw:
                messagebox.showwarning(
                    "Validación", "El campo Curso es obligatorio.", parent=dlg
                )
                return
            if not area:
                messagebox.showwarning(
                    "Validación", "El campo Área es obligatorio.", parent=dlg
                )
                return

            try:
                grado_val = self._plan_normalizar_grado(grado_raw)
            except ValueError:
                messagebox.showwarning("Validación", "Grado inválido.", parent=dlg)
                return
            try:
                curso_val = self._plan_normalizar_curso(curso_raw)
            except ValueError:
                messagebox.showwarning("Validación", "Curso inválido.", parent=dlg)
                return
            try:
                horas_val = int(float(horas_raw)) if horas_raw else 0
            except ValueError:
                messagebox.showwarning(
                    "Validación", "Horas debe ser un número entero.", parent=dlg
                )
                return

            try:
                conn = sqlite3.connect(self.db_path)
                cur = conn.cursor()

                cur.execute("SELECT id FROM areas WHERE nombre=? LIMIT 1", (area,))
                row_area = cur.fetchone()
                if not row_area:
                    conn.close()
                    messagebox.showwarning(
                        "Validación",
                        "El área seleccionada no existe en el catálogo.",
                        parent=dlg,
                    )
                    return
                id_area = int(row_area[0])

                # Verificar duplicado (excluir el propio registro en modo editar)
                if modo == "agregar":
                    cur.execute(
                        "SELECT id FROM plan_estudio WHERE grado=? AND curso=? AND CAST(COALESCE(IdArea,'') AS TEXT)=CAST(? AS TEXT)",
                        (grado_val, curso_val, id_area),
                    )
                else:
                    cur.execute(
                        "SELECT id FROM plan_estudio "
                        "WHERE grado=? AND curso=? AND CAST(COALESCE(IdArea,'') AS TEXT)=CAST(? AS TEXT) AND id<>?",
                        (grado_val, curso_val, id_area, int(id_registro)),
                    )

                if cur.fetchone():
                    conn.close()
                    messagebox.showwarning(
                        "Duplicado",
                        "El área ya existe en este grado y curso.",
                        parent=dlg,
                    )
                    return

                if modo == "agregar":
                    cur.execute(
                        "INSERT INTO plan_estudio (nivel, grado, curso, IdArea, area, horas) "
                        "VALUES (?, ?, ?, ?, ?, ?)",
                        (nivel, grado_val, curso_val, id_area, area, horas_val),
                    )
                else:
                    cur.execute(
                        "UPDATE plan_estudio "
                        "SET nivel=?, grado=?, curso=?, IdArea=?, area=?, horas=? "
                        "WHERE id=?",
                        (
                            nivel,
                            grado_val,
                            curso_val,
                            id_area,
                            area,
                            horas_val,
                            int(id_registro),
                        ),
                    )

                conn.commit()
                conn.close()
            except Exception as exc:
                messagebox.showerror(
                    "Plan de Estudios", f"Error al guardar:\n{exc}", parent=dlg
                )
                return

            dlg.destroy()
            self._plan_load_treeview()

        btn_frame = ttk.Frame(dlg)
        btn_frame.grid(row=len(campos), column=0, columnspan=2, pady=(6, 10))
        ttk.Button(btn_frame, text="Guardar", command=_guardar).pack(
            side="left", padx=6
        )
        ttk.Button(btn_frame, text="Cancelar", command=dlg.destroy).pack(
            side="left", padx=6
        )

        dlg.columnconfigure(1, weight=1)
        dlg.wait_window()

    def importar_plan_estudio(self):
        """Importa plan de estudios desde un archivo Excel o CSV.

        El archivo debe tener las columnas: nivel, curso, IdArea, horas
        El campo 'curso' tiene el formato 'GG-CC' (ej. '04-01') donde
        GG = grado y CC = número de curso.
        """
        if not self._requiere_permiso("desktop.superadmin.plan_estudio.importar"):
            return
        path = filedialog.askopenfilename(
            title="Seleccionar archivo Plan de Estudios",
            filetypes=[
                ("Excel / CSV", "*.xlsx;*.xls;*.csv"),
                ("Excel", "*.xlsx;*.xls"),
                ("CSV", "*.csv"),
                ("Todos", "*.*"),
            ],
            parent=self.win,
        )
        if not path:
            return

        # ---- Leer archivo ----
        registros = self._plan_leer_archivo(path)
        if registros is None:
            return
        if not registros:
            messagebox.showerror(
                "Plan de Estudios",
                "El archivo está vacío o no tiene datos válidos.",
                parent=self.win,
            )
            return

        # ---- Procesar e insertar ----
        insertados = 0
        omitidos = 0
        errores = []

        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()

            for i, fila in enumerate(registros, start=2):  # fila 2 = primera de datos
                try:
                    nivel = str(fila.get("nivel") or "").strip()
                    curso_raw = str(fila.get("curso") or "").strip()
                    id_area_raw = fila.get("idarea")
                    horas_raw = fila.get("horas") or 0

                    if not curso_raw or id_area_raw in (None, ""):
                        omitidos += 1
                        continue

                    grado_val, curso_val = self._plan_parse_curso_compuesto(curso_raw)

                    try:
                        id_area_val = int(str(id_area_raw).strip())
                    except Exception:
                        errores.append(f"Fila {i}: IdArea inválido '{id_area_raw}'")
                        continue

                    cur.execute(
                        "SELECT nombre FROM areas WHERE id=? AND estado=1",
                        (id_area_val,),
                    )
                    row_area = cur.fetchone()
                    if not row_area:
                        errores.append(
                            f"Fila {i}: IdArea no existe o inactivo ({id_area_val})"
                        )
                        continue
                    area_nombre = str(row_area[0] or "").strip()

                    try:
                        horas_val = int(float(str(horas_raw)))
                    except (ValueError, TypeError):
                        horas_val = 0

                    # Verificar duplicado (grado + curso + IdArea)
                    cur.execute(
                        "SELECT id FROM plan_estudio "
                        "WHERE grado=? AND curso=? AND ("
                        "CAST(COALESCE(IdArea,'') AS TEXT)=CAST(? AS TEXT) "
                        "OR ((IdArea IS NULL OR TRIM(CAST(IdArea AS TEXT))='') "
                        "AND LOWER(TRIM(COALESCE(area,'')))=LOWER(TRIM(?))))",
                        (grado_val, curso_val, id_area_val, area_nombre),
                    )
                    if cur.fetchone():
                        omitidos += 1
                        continue

                    cur.execute(
                        "INSERT INTO plan_estudio (nivel, grado, curso, IdArea, area, horas) "
                        "VALUES (?, ?, ?, ?, ?, ?)",
                        (
                            nivel,
                            grado_val,
                            curso_val,
                            id_area_val,
                            area_nombre,
                            horas_val,
                        ),
                    )
                    insertados += 1

                except Exception as exc_fila:
                    errores.append(f"Fila {i}: {exc_fila}")

            conn.commit()
            conn.close()
        except Exception as exc:
            messagebox.showerror(
                "Plan de Estudios",
                f"Error al acceder a la base de datos:\n{exc}",
                parent=self.win,
            )
            return

        self._plan_load_treeview()

        msg = f"Importación completada.\n\nInsertados: {insertados}\nOmitidos (duplicados o vacíos): {omitidos}"
        if errores:
            msg += f"\nFilas con error: {len(errores)}\n" + "\n".join(errores[:10])
        messagebox.showinfo("Plan de Estudios", msg, parent=self.win)

    def _plan_leer_archivo(self, path):
        """Lee un Excel o CSV y devuelve lista de dicts con llaves normalizadas."""
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == ".csv":
                registros = self._plan_leer_csv(path)
            else:
                registros = self._plan_leer_excel(path)
        except Exception as exc:
            messagebox.showerror(
                "Plan de Estudios",
                f"No se pudo leer el archivo:\n{exc}",
                parent=self.win,
            )
            return None

        if registros is None:
            messagebox.showerror(
                "Plan de Estudios",
                "El archivo no pudo ser leído. Verifique el formato.",
                parent=self.win,
            )
            return None

        # Normalizar nombres de columnas
        alias = {
            "nivel": ["nivel", "level"],
            "curso": ["curso", "course", "grado_curso"],
            "idarea": ["idarea", "id_area", "idárea", "id-área"],
            "horas": ["horas", "hours", "intensidad", "intensidad_horaria"],
        }
        resultado = []
        for fila in registros:
            fila_norm = {}
            fila_lower = {
                k.strip().lower().replace(" ", "_"): v for k, v in fila.items()
            }
            for campo, posibles in alias.items():
                for p in posibles:
                    if p in fila_lower:
                        fila_norm[campo] = fila_lower[p]
                        break
                else:
                    fila_norm[campo] = None
            resultado.append(fila_norm)
        return resultado

    def _plan_leer_csv(self, path):
        """Lee un CSV con separador auto-detectado y devuelve lista de dicts."""
        import csv

        # Detectar separador (coma, punto y coma, tabulador o pipe)
        with open(path, newline="", encoding="utf-8-sig") as f:
            muestra = f.read(4096)
        sep = ","
        for candidato in ("|", ";", "\t", ","):
            if candidato in muestra:
                sep = candidato
                break
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f, delimiter=sep)
            return [dict(row) for row in reader]

    def _plan_leer_excel(self, path):
        """Lee un Excel y devuelve lista de dicts."""
        if _HAS_PANDAS:
            df = pd.read_excel(path)
            return df.to_dict(orient="records")
        # Fallback openpyxl
        wb = load_workbook(path)
        ws = wb.active
        rows = list(ws.values)
        if not rows:
            return []
        headers = [str(h) if h is not None else "" for h in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:]]

    def _plan_eliminar_seleccionado(self):
        if not self._requiere_permiso("desktop.superadmin.plan_estudio.eliminar"):
            return
        sel = self.tree_plan.selection()
        if not sel:
            messagebox.showwarning(
                "Plan de Estudios",
                "Seleccione un registro para eliminar.",
                parent=self.win,
            )
            return
        valores = self.tree_plan.item(sel[0], "values")
        id_reg = valores[0] if valores else None
        if not id_reg:
            return
        if not messagebox.askyesno(
            "Confirmar",
            f"¿Eliminar el registro ID {id_reg}?",
            parent=self.win,
        ):
            return
        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute("DELETE FROM plan_estudio WHERE id=?", (int(id_reg),))
            conn.commit()
            conn.close()
        except Exception as exc:
            messagebox.showerror(
                "Plan de Estudios", f"Error al eliminar:\n{exc}", parent=self.win
            )
            return
        self._plan_load_treeview()

    def _plan_vaciar_todo(self):
        if not self._requiere_permiso("desktop.superadmin.plan_estudio.vaciar"):
            return
        total = len(self.tree_plan.get_children())
        if total == 0:
            messagebox.showinfo(
                "Plan de Estudios", "No hay registros para eliminar.", parent=self.win
            )
            return
        if not messagebox.askyesno(
            "Confirmar",
            f"¿Eliminar TODOS los {total} registro(s) del plan de estudios?\n"
            "Esta acción no se puede deshacer.",
            parent=self.win,
        ):
            return
        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute("DELETE FROM plan_estudio")
            conn.commit()
            conn.close()
        except Exception as exc:
            messagebox.showerror(
                "Plan de Estudios", f"Error al vaciar:\n{exc}", parent=self.win
            )
            return
        self._plan_load_treeview()
        messagebox.showinfo(
            "Plan de Estudios", "Plan de estudios vaciado.", parent=self.win
        )

    def cargar_areas_plan_estudio(self, grado, curso):
        """Devuelve la lista de áreas activas para el grado y curso indicados.

        Uso futuro: banco de preguntas, configuración de examen, carga académica.

        Args:
            grado: número o texto del grado (ej. 4 o '4').
            curso: número o texto del curso (ej. 1 o '1').

        Returns:
            list[str]: áreas ordenadas alfabéticamente.
        """
        try:
            grado_val = self._plan_normalizar_grado(grado)
            curso_val = self._plan_normalizar_curso(curso)
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute(
                """
                SELECT COALESCE(a.nombre, p.area)
                FROM plan_estudio p
                LEFT JOIN areas a ON a.id = p.IdArea
                WHERE p.grado=? AND p.curso=? AND p.estado=1
                  AND TRIM(COALESCE(a.nombre, p.area, '')) <> ''
                ORDER BY COALESCE(a.nombre, p.area)
                """,
                (grado_val, curso_val),
            )
            areas = [r[0] for r in cur.fetchall()]
            conn.close()
            return areas
        except Exception:
            return []

    # ---------- Seguridad ----------

    def _validar_accion_critica(self, titulo, mensaje_confirmacion):
        if not self._tiene_permiso("desktop.superadmin.seguridad"):
            messagebox.showerror("Error", "Acceso no autorizado", parent=self.win)
            return False

        confirmar = messagebox.askyesno(titulo, mensaje_confirmacion, parent=self.win)
        if not confirmar:
            return False

        clave = simpledialog.askstring(
            "Seguridad",
            "Ingrese la clave maestra:",
            show="*",
            parent=self.win,
        )
        if clave is None:
            return False
        if clave != self.get_master_key():
            messagebox.showerror("Error", "Clave incorrecta", parent=self.win)
            return False
        return True

    def vaciar_matricula(self):
        if not self._requiere_permiso("desktop.superadmin.matricula.vaciar"):
            return
        if not self._validar_accion_critica(
            "Confirmación",
            "Esta acción eliminará toda la matrícula y no se puede deshacer.",
        ):
            return
        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute("DELETE FROM estudiantes;")
            cur.execute("DELETE FROM sqlite_sequence WHERE name='estudiantes';")
            conn.commit()
            conn.close()
            messagebox.showinfo(
                "Éxito", "Matrícula eliminada correctamente.", parent=self.win
            )
        except Exception as e:
            messagebox.showerror(
                "Error", f"No se pudo eliminar matrícula: {e}", parent=self.win
            )

    def vaciar_calificaciones(self):
        if not self._validar_accion_critica(
            "Confirmación",
            "Esta acción eliminará todas las calificaciones registradas y no se puede deshacer.",
        ):
            return
        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute("DELETE FROM calificaciones;")
            cur.execute("DELETE FROM sqlite_sequence WHERE name='calificaciones';")
            conn.commit()
            conn.close()
            messagebox.showinfo(
                "Éxito",
                "Calificaciones eliminadas correctamente.",
                parent=self.win,
            )
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"No se pudo eliminar calificaciones: {e}",
                parent=self.win,
            )

    def _roles_permisos_editables(self):
        return {
            "Docente": core_usuarios.ROL_DOCENTE,
            "Rector": core_usuarios.ROL_RECTOR,
            "Secretaría": core_usuarios.ROL_SECRETARIA,
            "Coordinador": core_usuarios.ROL_COORDINADOR,
            "Estudiante": core_usuarios.ROL_ESTUDIANTE,
        }

    def _etiqueta_rol_permiso(self, rol):
        roles = self._roles_permisos_editables()
        for etiqueta, valor in roles.items():
            if valor == rol:
                return etiqueta
        return str(rol or "").strip().title()

    def _crear_selector_rol_permisos(self, parent, row, column):
        roles = self._roles_permisos_editables()
        cb = ttk.Combobox(
            parent,
            textvariable=self.var_permiso_rol,
            state="readonly",
            values=list(roles.keys()),
            width=18,
        )
        cb.grid(row=row, column=column, sticky="w", padx=6, pady=4)
        return cb

    def _permisos_resolver_rol_actual(self):
        return self._roles_permisos_editables().get(
            str(self.var_permiso_rol.get() or "").strip(),
            core_usuarios.ROL_DOCENTE,
        )

    def _permisos_modo_actual(self):
        valor = str(self.var_permiso_modo.get() or "Por rol").strip().lower()
        return "usuario" if "usuario" in valor else "rol"

    def _permisos_refrescar_usuarios(self, *_args, preservar_actual=True):
        rol = self._permisos_resolver_rol_actual()
        usuarios = core_usuarios.listar_usuarios_para_permisos(rol)
        self._permisos_usuario_label_to_doc = {}
        valores = []
        actual = str(self.var_permiso_usuario.get() or "").strip()
        for usuario in usuarios:
            documento = str(usuario.get("documento") or "").strip()
            nombre = str(usuario.get("nombre") or "").strip() or documento
            if not documento:
                continue
            etiqueta = f"{nombre} ({documento})"
            self._permisos_usuario_label_to_doc[etiqueta] = documento
            valores.append(etiqueta)
        self.combo_permiso_usuario["values"] = valores
        if preservar_actual and actual in valores:
            self.var_permiso_usuario.set(actual)
        elif valores:
            self.var_permiso_usuario.set(valores[0])
        else:
            self.var_permiso_usuario.set("")

    def _permisos_marcar_catalogo(self, permisos_activos):
        activos = set(permisos_activos or [])
        for permiso, var in getattr(self, "_permiso_vars", {}).items():
            var.set(permiso in activos)

    def _permisos_estado_actual(self):
        modo = self._permisos_modo_actual()
        rol = self._permisos_resolver_rol_actual()
        if modo == "usuario":
            etiqueta = str(self.var_permiso_usuario.get() or "").strip()
            documento = self._permisos_usuario_label_to_doc.get(etiqueta, "")
            return core_usuarios.obtener_configuracion_permisos_usuario(documento, rol)
        return core_usuarios.obtener_configuracion_permisos_rol(rol)

    def _permisos_actualizar_resumen(self, cfg=None):
        if not isinstance(cfg, dict):
            cfg = self._permisos_estado_actual()
        rol = self._etiqueta_rol_permiso(cfg.get("rol"))
        modo = self._permisos_modo_actual()
        if modo == "usuario":
            etiqueta = str(self.var_permiso_usuario.get() or "").strip()
            if not etiqueta:
                self.lbl_permiso_estado_var.set(
                    "Seleccione un usuario para revisar o asignar permisos."
                )
                return
            if cfg.get("personalizado"):
                self.lbl_permiso_estado_var.set(
                    f"Usuario con permisos personalizados sobre el rol {rol}."
                )
            else:
                self.lbl_permiso_estado_var.set(
                    f"Usuario heredando permisos del rol {rol}."
                )
            return

        if cfg.get("personalizado"):
            self.lbl_permiso_estado_var.set(
                f"Rol {rol} con perfil de permisos personalizado."
            )
        else:
            self.lbl_permiso_estado_var.set(
                f"Rol {rol} usando permisos predeterminados del sistema."
            )

    def _permisos_cargar_actual(self, *_):
        self._permisos_refrescar_usuarios(preservar_actual=True)
        modo = self._permisos_modo_actual()
        mostrar_usuario = modo == "usuario"
        estado_usuario = "readonly" if mostrar_usuario else "disabled"
        try:
            self.combo_permiso_usuario.configure(state=estado_usuario)
        except Exception:
            pass

        if mostrar_usuario:
            try:
                self.lbl_permiso_usuario.grid()
                self.combo_permiso_usuario.grid()
            except Exception:
                pass
        else:
            self.var_permiso_usuario.set("")
            try:
                self.lbl_permiso_usuario.grid_remove()
                self.combo_permiso_usuario.grid_remove()
            except Exception:
                pass

        if modo == "usuario" and not str(self.var_permiso_usuario.get() or "").strip():
            self._permisos_marcar_catalogo(set())
            self._permisos_actualizar_resumen(
                {
                    "rol": self._permisos_resolver_rol_actual(),
                    "personalizado": False,
                }
            )
            return

        cfg = self._permisos_estado_actual()
        self._permisos_marcar_catalogo(cfg.get("permisos", set()))
        self._permisos_actualizar_resumen(cfg)

    def _permisos_seleccionar_todos(self):
        for var in getattr(self, "_permiso_vars", {}).values():
            var.set(True)

    def _permisos_limpiar_todos(self):
        for var in getattr(self, "_permiso_vars", {}).values():
            var.set(False)

    def _permisos_guardar_actual(self):
        rol = self._permisos_resolver_rol_actual()
        permisos = [
            permiso
            for permiso, var in getattr(self, "_permiso_vars", {}).items()
            if bool(var.get())
        ]

        try:
            if self._permisos_modo_actual() == "usuario":
                etiqueta = str(self.var_permiso_usuario.get() or "").strip()
                documento = self._permisos_usuario_label_to_doc.get(etiqueta, "")
                if not documento:
                    messagebox.showwarning(
                        "Permisos",
                        "Seleccione un usuario para guardar permisos personalizados.",
                        parent=self.win,
                    )
                    return
                core_usuarios.guardar_permisos_usuario(documento, rol, permisos)
                messagebox.showinfo(
                    "Permisos",
                    "Permisos del usuario guardados y aplicados correctamente.",
                    parent=self.win,
                )
            else:
                core_usuarios.guardar_permisos_rol(rol, permisos)
                messagebox.showinfo(
                    "Permisos",
                    "Permisos del rol guardados y aplicados correctamente.",
                    parent=self.win,
                )
        except Exception as exc:
            messagebox.showerror(
                "Permisos",
                f"No se pudieron guardar los permisos.\n{exc}",
                parent=self.win,
            )
            return

        self._permisos_cargar_actual()

    def _permisos_restablecer_actual(self):
        rol = self._permisos_resolver_rol_actual()
        try:
            if self._permisos_modo_actual() == "usuario":
                etiqueta = str(self.var_permiso_usuario.get() or "").strip()
                documento = self._permisos_usuario_label_to_doc.get(etiqueta, "")
                if not documento:
                    messagebox.showwarning(
                        "Permisos",
                        "Seleccione un usuario para restablecer permisos.",
                        parent=self.win,
                    )
                    return
                core_usuarios.restablecer_permisos_usuario(documento, rol)
                messagebox.showinfo(
                    "Permisos",
                    "El usuario volvió a heredar los permisos del rol y el cambio ya fue aplicado.",
                    parent=self.win,
                )
            else:
                core_usuarios.restablecer_permisos_rol(rol)
                messagebox.showinfo(
                    "Permisos",
                    "El rol volvió a su configuración predeterminada y el cambio ya fue aplicado.",
                    parent=self.win,
                )
        except Exception as exc:
            messagebox.showerror(
                "Permisos",
                f"No se pudieron restablecer los permisos.\n{exc}",
                parent=self.win,
            )
            return

        self._permisos_cargar_actual()

    def _build_permisos_tab(self, parent):
        frame = ttk.LabelFrame(
            parent,
            text="Roles y permisos",
            padding=16,
            style="Card.TLabelframe",
        )
        frame.pack(fill="both", expand=True, padx=22, pady=(0, 14))
        frame.grid_columnconfigure(0, weight=1)

        ttk.Label(
            frame,
            text="Administra por funcionalidad lo que cada usuario puede ver y usar. Lo no seleccionado queda oculto y sin acceso.",
            style="CardHint.TLabel",
        ).grid(row=0, column=0, sticky="w", pady=(0, 6))

        controles = ttk.Frame(frame)
        controles.grid(row=1, column=0, sticky="ew", pady=(0, 6))
        controles.grid_columnconfigure(7, weight=1)

        self.var_permiso_modo = tk.StringVar(value="Por rol")
        self.var_permiso_rol = tk.StringVar(value="Docente")
        self.var_permiso_usuario = tk.StringVar()
        self.lbl_permiso_estado_var = tk.StringVar(
            value="Seleccione un perfil para revisar permisos."
        )
        self.lbl_permiso_ayuda_var = tk.StringVar(
            value="Modo Rol: define la base del perfil. Modo Usuario: habilita o restringe funcionalidades específicas sobre ese perfil."
        )
        self._permiso_vars = {}
        self._permisos_usuario_label_to_doc = {}

        ttk.Label(controles, text="Administrar:").grid(
            row=0, column=0, sticky="e", padx=6, pady=2
        )
        self.combo_permiso_modo = ttk.Combobox(
            controles,
            textvariable=self.var_permiso_modo,
            state="readonly",
            values=["Por rol", "Por usuario"],
            width=14,
        )
        self.combo_permiso_modo.grid(row=0, column=1, sticky="w", padx=6, pady=2)

        ttk.Label(controles, text="Rol:").grid(
            row=0, column=2, sticky="e", padx=6, pady=2
        )
        self.combo_permiso_rol = self._crear_selector_rol_permisos(controles, 0, 3)

        self.lbl_permiso_usuario = ttk.Label(controles, text="Usuario:")
        self.lbl_permiso_usuario.grid(row=0, column=4, sticky="e", padx=6, pady=2)
        self.combo_permiso_usuario = ttk.Combobox(
            controles,
            textvariable=self.var_permiso_usuario,
            state="disabled",
            width=42,
        )
        self.combo_permiso_usuario.grid(
            row=0, column=5, columnspan=3, sticky="ew", padx=6, pady=2
        )

        acciones = ttk.Frame(controles)
        acciones.grid(row=1, column=0, columnspan=8, sticky="w", pady=(4, 0))
        ttk.Button(
            acciones, text="Seleccionar todo", command=self._permisos_seleccionar_todos
        ).pack(side="left", padx=(0, 6))
        ttk.Button(acciones, text="Limpiar", command=self._permisos_limpiar_todos).pack(
            side="left", padx=(0, 12)
        )
        ttk.Button(
            acciones, text="Guardar y aplicar", command=self._permisos_guardar_actual
        ).pack(side="left", padx=(0, 6))
        ttk.Button(
            acciones, text="Restablecer", command=self._permisos_restablecer_actual
        ).pack(side="left", padx=(0, 6))
        ttk.Button(
            acciones, text="Recargar", command=self._permisos_cargar_actual
        ).pack(side="left")

        ttk.Label(frame, textvariable=self.lbl_permiso_estado_var).grid(
            row=2, column=0, sticky="w", pady=(0, 4)
        )
        ttk.Label(
            frame, textvariable=self.lbl_permiso_ayuda_var, style="CardHint.TLabel"
        ).grid(row=3, column=0, sticky="w", pady=(0, 6))

        editor = ttk.Frame(frame)
        editor.grid(row=4, column=0, sticky="nsew")
        frame.grid_rowconfigure(4, weight=1)
        editor.grid_columnconfigure(0, weight=1)
        editor.grid_rowconfigure(0, weight=1)

        inner = ttk.Frame(editor)
        inner.grid_columnconfigure(0, weight=1)
        inner.grid(row=0, column=0, sticky="nsew")

        for idx, (grupo, permisos) in enumerate(
            core_usuarios.listar_catalogo_permisos_asignables().items()
        ):
            card = ttk.LabelFrame(
                inner,
                text=grupo,
                padding=6,
                style="Card.TLabelframe",
            )
            card.grid(row=idx, column=0, sticky="ew", padx=4, pady=4)
            total_permisos = len(permisos)
            num_columnas = 2 if total_permisos >= 6 else 1
            for col in range(num_columnas):
                card.grid_columnconfigure(col, weight=1)
            for jdx, permiso in enumerate(permisos):
                var = tk.BooleanVar(value=False)
                self._permiso_vars[permiso] = var
                descripcion = core_usuarios.describir_permiso(permiso)
                texto = descripcion if descripcion else permiso
                fila = jdx // num_columnas
                columna = jdx % num_columnas
                ttk.Checkbutton(
                    card,
                    text=texto,
                    variable=var,
                ).grid(
                    row=fila,
                    column=columna,
                    sticky="w",
                    padx=(0, 12),
                    pady=1,
                )

        self.combo_permiso_modo.bind(
            "<<ComboboxSelected>>", self._permisos_cargar_actual
        )
        self.combo_permiso_rol.bind(
            "<<ComboboxSelected>>", self._permisos_cargar_actual
        )
        self.combo_permiso_rol.bind(
            "<<ComboboxSelected>>", self._permisos_refrescar_usuarios, add="+"
        )
        self.combo_permiso_usuario.bind(
            "<<ComboboxSelected>>", self._permisos_cargar_actual
        )
        self.combo_permiso_modo.bind(
            "<<ComboboxSelected>>", self._permisos_refrescar_usuarios, add="+"
        )
        self.combo_permiso_usuario.bind(
            "<<ComboboxSelected>>", self._permisos_actualizar_resumen, add="+"
        )

        self._permisos_refrescar_usuarios(preservar_actual=False)
        self._permisos_cargar_actual()

        return frame

    def _build_seguridad_tab(self, parent=None):
        frame = parent if parent is not None else self.tab_seguridad
        for widget in frame.winfo_children():
            widget.destroy()

        container = ttk.Frame(frame)
        container.pack(fill="both", expand=True)
        container.grid_columnconfigure(0, weight=1)
        container.grid_rowconfigure(0, weight=1)

        canvas = tk.Canvas(container, bg="#eef4fb", highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        shell = ttk.Frame(canvas)
        shell.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        shell_window = canvas.create_window((0, 0), window=shell, anchor="nw")

        def _ajustar_ancho_seguridad(event):
            try:
                canvas.itemconfigure(shell_window, width=max(200, event.width - 2))
            except Exception:
                pass

        canvas.bind("<Configure>", _ajustar_ancho_seguridad)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        hero = tk.Frame(shell, bg="#16324f", padx=24, pady=18)
        hero.pack(fill="x", padx=18, pady=(18, 10))
        tk.Label(
            hero,
            text="Seguridad",
            font=("Segoe UI", 17, "bold"),
            bg="#16324f",
            fg="white",
        ).pack(anchor="w")
        tk.Label(
            hero,
            text="Gestiona la clave maestra y administra los permisos de acceso desde un entorno controlado.",
            font=("Segoe UI", 10),
            bg="#16324f",
            fg="#d1e3f4",
            wraplength=980,
            justify="left",
        ).pack(anchor="w", pady=(6, 0))

        acceso = ttk.LabelFrame(
            shell,
            text="Acceso administrativo",
            padding=16,
            style="Card.TLabelframe",
        )
        acceso.pack(fill="x", padx=22, pady=(0, 14))
        ttk.Label(
            acceso,
            text="Actualiza la clave maestra usada para operaciones sensibles del sistema.",
            style="CardHint.TLabel",
        ).pack(anchor="w", pady=(0, 10))
        ttk.Button(
            acceso,
            text="Cambiar clave maestra",
            command=self.seguridad_cambiar_clave,
        ).pack(anchor="w")

        perfil = ttk.LabelFrame(
            shell,
            text="Perfil de SuperAdmin",
            padding=16,
            style="Card.TLabelframe",
        )
        perfil.pack(fill="x", padx=22, pady=(0, 14))
        perfil.grid_columnconfigure(1, weight=1)
        perfil.grid_columnconfigure(3, weight=1)
        ttk.Label(
            perfil,
            text="Define los datos de identificación del rol SuperAdmin sin cambiar el método actual de ingreso por admin y clave maestra.",
            style="CardHint.TLabel",
            wraplength=980,
            justify="left",
        ).grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 10))

        perfil_actual = self.get_superadmin_profile()
        self.var_superadmin_nombres = tk.StringVar(value=perfil_actual["nombres"])
        self.var_superadmin_apellidos = tk.StringVar(value=perfil_actual["apellidos"])
        self.var_superadmin_documento = tk.StringVar(value=perfil_actual["documento"])
        self.var_superadmin_estado = tk.StringVar(
            value="Actualice nombres, apellidos y documento del rol SuperAdmin."
        )

        ttk.Label(perfil, text="Nombres:").grid(
            row=1, column=0, sticky="e", padx=6, pady=4
        )
        ttk.Entry(perfil, textvariable=self.var_superadmin_nombres, width=28).grid(
            row=1, column=1, sticky="ew", padx=6, pady=4
        )
        ttk.Label(perfil, text="Apellidos:").grid(
            row=1, column=2, sticky="e", padx=6, pady=4
        )
        ttk.Entry(perfil, textvariable=self.var_superadmin_apellidos, width=28).grid(
            row=1, column=3, sticky="ew", padx=6, pady=4
        )
        ttk.Label(perfil, text="Documento:").grid(
            row=2, column=0, sticky="e", padx=6, pady=4
        )
        ttk.Entry(perfil, textvariable=self.var_superadmin_documento, width=28).grid(
            row=2, column=1, sticky="ew", padx=6, pady=4
        )
        ttk.Button(
            perfil,
            text="Guardar perfil",
            command=self.seguridad_guardar_perfil_superadmin,
        ).grid(row=2, column=3, sticky="e", padx=6, pady=4)
        ttk.Label(
            perfil,
            textvariable=self.var_superadmin_estado,
            style="CardHint.TLabel",
            wraplength=980,
            justify="left",
        ).grid(row=3, column=0, columnspan=4, sticky="w", pady=(8, 0))

        self._build_permisos_tab(shell)

        def _on_mousewheel(event):
            try:
                delta = -1 * int(event.delta / 120)
                canvas.yview_scroll(delta, "units")
            except Exception:
                pass

        try:
            canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
        except Exception:
            pass

    def configuracion_configurar_horas_docente(self):
        if not self._requiere_permiso("desktop.superadmin.carga_academica.horas"):
            return
        try:
            core_docentes.asegurar_esquema_carga_academica()
        except Exception as e:
            messagebox.showerror(
                "Configuración",
                f"No se pudo preparar la tabla de configuración de horas.\n{e}",
                parent=self.win,
            )
            return

        d = tk.Toplevel(self.win)
        d.transient(self.win)
        d.grab_set()
        d.title("Configurar Horas Docente")
        d.geometry("560x280")
        d.minsize(520, 250)

        frame = ttk.Frame(d, padding=12)
        frame.pack(fill="both", expand=True)
        frame.grid_columnconfigure(1, weight=1)

        docentes_activos = self._ca_todos_docentes(solo_activos=True)
        docente_label_to_doc = {}
        labels_doc = []
        for doc, nom in docentes_activos:
            doc_txt = str(doc or "").strip()
            nom_txt = str(nom or "").strip()
            if not doc_txt:
                continue
            label = f"{nom_txt} ({doc_txt})" if nom_txt else doc_txt
            labels_doc.append(label)
            docente_label_to_doc[label] = doc_txt

        ttk.Label(frame, text="Docente:").grid(
            row=0, column=0, sticky="e", padx=6, pady=8
        )
        cb_docente = ttk.Combobox(frame, state="readonly", width=42)
        cb_docente["values"] = labels_doc
        cb_docente.grid(row=0, column=1, sticky="we", padx=6, pady=8)

        ttk.Label(frame, text="Horas máximas normales:").grid(
            row=1, column=0, sticky="e", padx=6, pady=8
        )
        en_normales = ttk.Entry(frame, width=14)
        en_normales.grid(row=1, column=1, sticky="w", padx=6, pady=8)

        ttk.Label(frame, text="Horas extras máximas:").grid(
            row=2, column=0, sticky="e", padx=6, pady=8
        )
        en_extras = ttk.Entry(frame, width=14)
        en_extras.grid(row=2, column=1, sticky="w", padx=6, pady=8)

        info_var = tk.StringVar(
            value="Si no existe configuración, se usarán valores por defecto: 22 normales y 0 extras."
        )
        ttk.Label(frame, textvariable=info_var).grid(
            row=3, column=0, columnspan=2, sticky="w", padx=6, pady=(4, 10)
        )

        def _cargar_config_docente(*_):
            label = cb_docente.get().strip()
            doc = docente_label_to_doc.get(label, "")
            if not doc:
                return
            normales, extras, configurado = self._ca_obtener_limites_docente(doc)
            en_normales.delete(0, "end")
            en_normales.insert(0, str(normales))
            en_extras.delete(0, "end")
            en_extras.insert(0, str(extras))
            if configurado:
                info_var.set("Docente con configuración personalizada.")
            else:
                info_var.set(
                    "Docente sin configuración personalizada. Al guardar se creará el registro."
                )

        def _guardar_config_docente():
            label = cb_docente.get().strip()
            docente_documento = docente_label_to_doc.get(label, "")
            if not docente_documento:
                messagebox.showerror(
                    "Configuración",
                    "Seleccione un docente válido.",
                    parent=d,
                )
                return

            try:
                horas_normales = int(str(en_normales.get()).strip())
                horas_extras = int(str(en_extras.get()).strip())
            except Exception:
                messagebox.showerror(
                    "Configuración",
                    "Las horas deben ser números enteros.",
                    parent=d,
                )
                return

            if horas_normales < 0 or horas_extras < 0:
                messagebox.showerror(
                    "Configuración",
                    "Las horas no pueden ser negativas.",
                    parent=d,
                )
                return

            try:
                core_docentes.guardar_limites_docente(
                    docente_documento,
                    horas_normales,
                    horas_extras,
                )
            except Exception as e:
                messagebox.showerror(
                    "Configuración",
                    f"No se pudo guardar la configuración del docente.\n{e}",
                    parent=d,
                )
                return

            info_var.set("Configuración guardada correctamente.")
            messagebox.showinfo(
                "Configuración",
                "Configuración de horas guardada correctamente.",
                parent=d,
            )

        cb_docente.bind("<<ComboboxSelected>>", _cargar_config_docente)
        if labels_doc:
            cb_docente.set(labels_doc[0])
            _cargar_config_docente()

        frame_btn = ttk.Frame(frame)
        frame_btn.grid(row=4, column=0, columnspan=2, sticky="w", padx=6, pady=(0, 0))
        ttk.Button(frame_btn, text="Guardar", command=_guardar_config_docente).pack(
            side="left"
        )
        ttk.Button(frame_btn, text="Cerrar", command=d.destroy).pack(
            side="left", padx=(8, 0)
        )

        d.wait_window()

    def seguridad_cambiar_clave(self):
        if not self._requiere_permiso("desktop.superadmin.seguridad"):
            return
        actual = simpledialog.askstring(
            "Clave actual", "Ingrese clave actual:", show="*", parent=self.win
        )
        if actual is None:
            return
        if actual != self.get_master_key():
            messagebox.showerror("Error", "Clave actual incorrecta.", parent=self.win)
            return
        nueva = simpledialog.askstring(
            "Nueva clave", "Ingrese nueva clave:", show="*", parent=self.win
        )
        if not nueva:
            return
        confirm = simpledialog.askstring(
            "Confirmar", "Confirme nueva clave:", show="*", parent=self.win
        )
        if nueva != confirm:
            messagebox.showerror("Error", "Las claves no coinciden.", parent=self.win)
            return
        self.set_master_key(nueva)
        messagebox.showinfo("OK", "Clave maestra actualizada.", parent=self.win)

    def seguridad_guardar_perfil_superadmin(self):
        if not self._requiere_permiso("desktop.superadmin.seguridad"):
            return

        nombres = str(self.var_superadmin_nombres.get() or "").strip()
        apellidos = str(self.var_superadmin_apellidos.get() or "").strip()
        documento = str(self.var_superadmin_documento.get() or "").strip()

        if not nombres:
            messagebox.showwarning(
                "Perfil SuperAdmin",
                "Ingrese los nombres del rol SuperAdmin.",
                parent=self.win,
            )
            return
        if not apellidos:
            messagebox.showwarning(
                "Perfil SuperAdmin",
                "Ingrese los apellidos del rol SuperAdmin.",
                parent=self.win,
            )
            return
        if not documento:
            messagebox.showwarning(
                "Perfil SuperAdmin",
                "Ingrese el número de documento del rol SuperAdmin.",
                parent=self.win,
            )
            return

        self.set_superadmin_profile(nombres, apellidos, documento)
        self._actualizar_resumen_superadmin()
        self.var_superadmin_estado.set(
            f"Perfil actualizado: {nombres} {apellidos} | Documento: {documento}. El ingreso sigue siendo por admin y clave maestra."
        )
        messagebox.showinfo(
            "Perfil SuperAdmin",
            "Los datos del rol SuperAdmin fueron guardados correctamente.",
            parent=self.win,
        )

    # ---------- Cierre ----------
    def close(self):
        try:
            self.conn.close()
        except Exception:
            pass


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    msa = ModuloSuperAdmin(root, db_path=str(DB_FILE), base_dir=str(BASE_DIR))
    if msa.authenticate():
        msa.open_interface()
        root.mainloop()
